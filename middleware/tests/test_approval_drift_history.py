import os
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from approval_drift_history import (
    accept_reviewed_drift,
    build_history_rows,
    load_current_participants,
    parse_drift_events,
    run,
)


def _write_status_workbook(path: Path, extra_rows: list[list[object]] | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts-Status"
    ws.append(
        [
            "Church Team",
            "ChMeetings ID",
            "First Name",
            "Last Name",
            "Participant ID (WP)",
            "Approval_Status (WP)",
            "Is_Member_ChM",
            "Sports Registered",
            "First_Open_ERROR_Desc (WP)",
            "Update_on_ChM",
        ]
    )
    ws.append(
        [
            "GAC",
            "3744979",
            "Joshua",
            "Dang",
            "446",
            "reapproval_required",
            True,
            "Soccer - Coed Exhibition Mixed Team, Table Tennis Men Singles",
            "",
            "2026-07-13 03:01:02",
        ]
    )
    ws.append(
        [
            "RPC",
            "123",
            "Approved",
            "Player",
            "10",
            "approved",
            True,
            "Basketball Men Team",
            "",
            "2026-07-13 03:01:02",
        ]
    )
    ws.append(
        [
            "MWC",
            "999",
            "No",
            "Log",
            "20",
            "reapproval_required",
            False,
            "Volleyball Women Team",
            "Needs attention",
            "",
        ]
    )
    for row in extra_rows or []:
        ws.append(row)
    wb.save(path)


def _write_log(logs_dir: Path, lines: list[str]) -> Path:
    logs_dir.mkdir()
    log_file = logs_dir / "sportsfest_20260713.log"
    log_file.write_text("\n".join(lines), encoding="utf-8")
    return log_file


def test_parse_drift_events_splits_hard_drift_fields(tmp_path):
    log_file = tmp_path / "sportsfest_20260713.log"
    log_file.write_text(
        "\n".join(
            [
                "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
                "2026-07-13 03:33:41 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=4363690 (WP participant_id=117): Secondary sport: 'Volleyball - Men Team' -> 'Basketball - Men Team'. Prior 'approved' invalidated -> 'reapproval_required'.",
            ]
        ),
        encoding="utf-8",
    )

    events = parse_drift_events([log_file])

    assert len(events) == 2
    assert events[0].chmeetings_id == "3744979"
    assert events[0].field == "Other events"
    assert events[0].old_value == ""
    assert events[0].new_value == "Soccer - Coed Exhibition"
    assert events[0].prior_status == "approved"
    assert events[1].field == "Secondary sport"


def test_build_history_rows_includes_current_status_and_missing_log_marker(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(workbook)
    current = load_current_participants(workbook)
    log_file = tmp_path / "sportsfest_20260713.log"
    log_file.write_text(
        "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.\n",
        encoding="utf-8",
    )

    rows = build_history_rows(current, parse_drift_events([log_file]))

    assert [row["Name"] for row in rows] == ["Joshua Dang", "No Log"]
    assert rows[0]["Changed Field"] == "Other events"
    assert rows[0]["Is Member"] == "Yes"
    assert rows[0]["Current Approval Status"] == "reapproval_required"
    assert rows[0]["Latest ChMeetings Update"] == "2026-07-13 03:01:02"
    assert rows[0]["First Detected At"] == "2026-07-13 03:32:48"
    assert rows[0]["Last Seen At"] == "2026-07-13 03:32:48"
    assert rows[0]["Times Seen"] == "1"
    assert rows[1]["Event Type"] == "no_local_drift_log_found"
    assert rows[1]["Is Member"] == "No"
    assert rows[1]["First Open Error"] == "Needs attention"


def test_run_writes_history_workbook(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(workbook)
    logs_dir = tmp_path / "logs"
    logs_dir.mkdir()
    (logs_dir / "sportsfest_20260713.log").write_text(
        "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.\n",
        encoding="utf-8",
    )
    output = tmp_path / "approval_drift_history.xlsx"

    summary = run(
        workbook_path=workbook,
        logs_dir=logs_dir,
        output_path=output,
        since="2026-07-11",
    )

    assert summary["participants"] == 2
    assert summary["participants_with_history"] == 1
    assert summary["rows"] == 2
    assert output.exists()
    wb = openpyxl.load_workbook(output, data_only=True)
    assert "Approval-Drift-History" in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_repeated_identical_changes_are_collapsed(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(workbook)
    current = load_current_participants(workbook)
    log_file = tmp_path / "sportsfest_20260713.log"
    log_file.write_text(
        "\n".join(
            [
                "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
                "2026-07-13 06:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
            ]
        ),
        encoding="utf-8",
    )

    rows = build_history_rows(current, parse_drift_events([log_file]))

    joshua_rows = [row for row in rows if row["Name"] == "Joshua Dang"]
    assert len(joshua_rows) == 1
    assert joshua_rows[0]["Times Seen"] == "2"
    assert joshua_rows[0]["First Detected At"] == "2026-07-13 03:32:48"
    assert joshua_rows[0]["Last Seen At"] == "2026-07-13 06:32:48"


class _FakeWordPress:
    def __init__(self):
        self.participants = [
            {
                "participant_id": 446,
                "chmeetings_id": "3744979",
                "church_code": "GAC",
                "approval_status": "reapproval_required",
            },
            {
                "participant_id": 20,
                "chmeetings_id": "999",
                "church_code": "MWC",
                "approval_status": "reapproval_required",
            },
            {
                "participant_id": 73,
                "chmeetings_id": "3739994",
                "church_code": "MWC",
                "approval_status": "reapproval_required",
            },
            {
                "participant_id": 74,
                "chmeetings_id": "3636103",
                "church_code": "MWC",
                "approval_status": "reapproval_required",
            },
        ]
        self.approvals = {
            446: [
                {
                    "approval_id": 91,
                    "participant_id": 446,
                    "approval_status": "pending",
                    "approval_notes": "Pastor approved before drift.",
                }
            ],
            20: [
                {
                    "approval_id": 92,
                    "participant_id": 20,
                    "approval_status": "pending",
                    "approval_notes": "",
                }
            ],
            73: [
                {
                    "approval_id": 93,
                    "participant_id": 73,
                    "approval_status": "pending",
                    "approval_notes": "",
                }
            ],
            74: [
                {
                    "approval_id": 94,
                    "participant_id": 74,
                    "approval_status": "pending",
                    "approval_notes": "",
                }
            ],
        }
        self.issues = {
            446: [
                {
                    "issue_id": 55,
                    "participant_id": 446,
                    "issue_type": "approval_identity_drift",
                    "status": "open",
                }
            ]
        }
        self.updated_participants = []
        self.updated_approvals = []
        self.updated_issues = []

    def get_participants(self, params=None):
        chmeetings_id = str((params or {}).get("chmeetings_id") or "")
        return [
            participant
            for participant in self.participants
            if str(participant["chmeetings_id"]) == chmeetings_id
        ]

    def update_participant(self, participant_id, participant_data):
        self.updated_participants.append((participant_id, participant_data))
        for participant in self.participants:
            if participant["participant_id"] == participant_id:
                participant.update(participant_data)
                return participant
        return None

    def get_approvals(self, params=None):
        participant_id = int((params or {}).get("participant_id") or 0)
        return list(self.approvals.get(participant_id, []))

    def update_approval(self, approval_id, approval_data):
        self.updated_approvals.append((approval_id, approval_data))
        for approvals in self.approvals.values():
            for approval in approvals:
                if approval["approval_id"] == approval_id:
                    approval.update(approval_data)
                    return {"approval_id": approval_id, "updated": True}
        return None

    def get_validation_issues(self, params=None):
        participant_id = int((params or {}).get("participant_id") or 0)
        issue_type = (params or {}).get("issue_type")
        status = (params or {}).get("status")
        return [
            issue
            for issue in self.issues.get(participant_id, [])
            if issue.get("issue_type") == issue_type and issue.get("status") == status
        ]

    def update_validation_issue(self, issue_id, issue_data):
        self.updated_issues.append((issue_id, issue_data))
        for issues in self.issues.values():
            for issue in issues:
                if issue["issue_id"] == issue_id:
                    issue.update(issue_data)
                    return {"issue_id": issue_id, **issue_data}
        return None


def test_accept_reviewed_drift_dry_run_does_not_update_wordpress(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(workbook)
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
        ],
    )
    output = tmp_path / "approval_drift_acceptance.xlsx"
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=output,
        logs_dir=logs_dir,
        church_code="GAC",
        reason="GAC confirmed final-week sport changes.",
        execute=False,
    )

    assert summary["targets"] == 1
    assert summary["would_accept"] == 1
    assert summary["accepted"] == 0
    assert wp.updated_participants == []
    assert wp.updated_approvals == []
    assert wp.updated_issues == []
    assert output.exists()


def test_accept_reviewed_drift_execute_restores_participant_and_approval(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(workbook)
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 03:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3744979 (WP participant_id=446): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
        ],
    )
    output = tmp_path / "approval_drift_acceptance.xlsx"
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=output,
        logs_dir=logs_dir,
        chm_id="3744979",
        reason="GAC confirmed final-week sport changes.",
        execute=True,
    )

    assert summary["targets"] == 1
    assert summary["accepted"] == 1
    assert wp.updated_participants == [(446, {"approval_status": "approved"})]
    assert wp.updated_approvals[0][0] == 91
    assert wp.updated_approvals[0][1]["approval_status"] == "approved"
    assert wp.updated_approvals[0][1]["synced_to_chmeetings"] is False
    assert "Pastor approved before drift." in wp.updated_approvals[0][1]["approval_notes"]
    assert "GAC confirmed final-week sport changes." in wp.updated_approvals[0][1]["approval_notes"]
    assert wp.updated_issues == [(55, {"status": "resolved"})]


def test_accept_reviewed_drift_restores_single_pending_prior_status(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(
        workbook,
        extra_rows=[
            [
                "MWC",
                "3739994",
                "Benjamin",
                "Schaner",
                "73",
                "reapproval_required",
                True,
                "Volleyball Men Team",
                "",
                "2026-07-13 04:01:02",
            ],
        ],
    )
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 04:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3739994 (WP participant_id=73): Main sport: 'Volleyball - Men Team' -> 'Basketball - Men Team'. Prior 'pending' invalidated -> 'reapproval_required'.",
        ],
    )
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=tmp_path / "approval_drift_acceptance.xlsx",
        logs_dir=logs_dir,
        chm_id="3739994",
        reason="MWC confirmed this should return to prior state.",
        execute=True,
    )

    assert summary["accepted"] == 1
    assert wp.updated_participants == [(73, {"approval_status": "pending"})]
    assert wp.updated_approvals[0][0] == 93
    assert wp.updated_approvals[0][1]["approval_status"] == "pending"
    assert "Approval restored to 'pending'" in wp.updated_approvals[0][1]["approval_notes"]


def test_accept_reviewed_drift_blocks_ambiguous_prior_statuses(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(
        workbook,
        extra_rows=[
            [
                "MWC",
                "3636103",
                "Joanna",
                "Nguyen",
                "74",
                "reapproval_required",
                True,
                "Basketball Women Team",
                "",
                "2026-07-13 04:01:02",
            ],
        ],
    )
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 04:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Main sport: 'Basketball - Women Team' -> 'Volleyball - Women Team'. Prior 'pending' invalidated -> 'reapproval_required'.",
            "2026-07-13 05:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
        ],
    )
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=tmp_path / "approval_drift_acceptance.xlsx",
        logs_dir=logs_dir,
        chm_id="3636103",
        reason="MWC confirmed sport changes.",
        execute=True,
    )

    assert summary["accepted"] == 0
    assert summary["skipped"] == 1
    assert summary["actions"] == {"blocked_ambiguous_prior_status": 1}
    assert wp.updated_participants == []
    assert wp.updated_approvals == []
    assert wp.updated_issues == []


def test_accept_reviewed_drift_ignores_reapproval_required_as_prior_status(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(
        workbook,
        extra_rows=[
            [
                "MWC",
                "3636103",
                "Joanna",
                "Nguyen",
                "74",
                "reapproval_required",
                True,
                "Basketball Women Team",
                "",
                "2026-07-13 04:01:02",
            ],
        ],
    )
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 04:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Main sport: 'Basketball - Women Team' -> 'Volleyball - Women Team'. Prior 'pending' invalidated -> 'reapproval_required'.",
            "2026-07-13 05:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'reapproval_required' invalidated -> 'reapproval_required'.",
        ],
    )
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=tmp_path / "approval_drift_acceptance.xlsx",
        logs_dir=logs_dir,
        chm_id="3636103",
        reason="MWC confirmed sport changes.",
        execute=True,
    )

    assert summary["accepted"] == 1
    assert wp.updated_participants == [(74, {"approval_status": "pending"})]
    assert wp.updated_approvals[0][0] == 94
    assert wp.updated_approvals[0][1]["approval_status"] == "pending"


def test_accept_reviewed_drift_force_approved_overrides_ambiguous_prior_statuses(tmp_path):
    workbook = tmp_path / "Church_Team_Status_ALL_2026-07-14.xlsx"
    _write_status_workbook(
        workbook,
        extra_rows=[
            [
                "MWC",
                "3636103",
                "Joanna",
                "Nguyen",
                "74",
                "reapproval_required",
                True,
                "Basketball Women Team",
                "",
                "2026-07-13 04:01:02",
            ],
        ],
    )
    logs_dir = tmp_path / "logs"
    _write_log(
        logs_dir,
        [
            "2026-07-13 04:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Main sport: 'Basketball - Women Team' -> 'Volleyball - Women Team'. Prior 'pending' invalidated -> 'reapproval_required'.",
            "2026-07-13 05:32:48 | WARNING | [VAY SM] APPROVAL IDENTITY DRIFT for chm_id=3636103 (WP participant_id=74): Other events: '' -> 'Soccer - Coed Exhibition'. Prior 'approved' invalidated -> 'reapproval_required'.",
        ],
    )
    wp = _FakeWordPress()

    summary = accept_reviewed_drift(
        wordpress_connector=wp,
        workbook_path=workbook,
        output_path=tmp_path / "approval_drift_acceptance.xlsx",
        logs_dir=logs_dir,
        chm_id="3636103",
        reason="MWC explicitly approved the reviewed changes.",
        execute=True,
        force_approved=True,
    )

    assert summary["accepted"] == 1
    assert wp.updated_participants == [(74, {"approval_status": "approved"})]
    assert wp.updated_approvals[0][0] == 94
    assert wp.updated_approvals[0][1]["approval_status"] == "approved"
