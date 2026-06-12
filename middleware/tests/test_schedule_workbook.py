import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from church_teams_export import ChurchTeamsExporter
from config import (
    POD_RESOURCE_TYPE_PICKLEBALL,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
    TEAM_RESOURCE_TYPE_SOCCER,
)
from schedule_workbook import ScheduleWorkbookBuilder


def _find_row_by_first_cell(ws, value: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == value:
            return row
    raise AssertionError(f"Could not find row starting with {value!r}")


def _header_row(ws) -> int:
    for row_idx in range(1, min(ws.max_row, 3) + 1):
        if ws.cell(row=row_idx, column=1).value == "Event":
            return row_idx
    raise AssertionError("Could not find header row starting with 'Event'")


def _sheet_rows(ws) -> list[dict]:
    header_row = _header_row(ws)
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row = {
            headers[col_idx - 1]: ws.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, ws.max_column + 1)
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def _status_banner_values(ws) -> set[str]:
    return {
        str(cell.value)
        for row in ws.iter_rows(min_row=1, max_row=2)
        for cell in row
        if cell.value is not None
    }


def _make_gym_roster(n_churches: int = 8) -> list[dict]:
    """Return minimal Basketball-Men roster rows for n_churches churches."""
    codes = ["RPC", "ANH", "FVC", "GAC", "NSD", "TLC", "GLA", "ORN"][:n_churches]
    return _make_gym_roster_from_codes(codes)


def _make_gym_roster_from_codes(codes: list[str]) -> list[dict]:
    """Return minimal Basketball-Men roster rows for explicit church codes."""
    rows = []
    for code in codes:
        for _ in range(5):  # 5 members per church -> meets min team size
            rows.append(
                {
                    "Church Team": code,
                    "sport_type": SPORT_TYPE["BASKETBALL"],
                    "sport_gender": "Men",
                    "sport_format": "Team",
                    "Participant ID (WP)": 1,
                }
            )
    return rows


def _make_schedule_pair() -> tuple[dict, dict]:
    schedule_input = {
        "games": [
            {
                "game_id": "BBM-01",
                "event": "Basketball - Men Team",
                "stage": "Pool",
                "pool_id": "P1",
                "round": 1,
                "team_a_id": "BBM-P1-T1",
                "team_b_id": "BBM-P1-T2",
                "team_a_label": "OCB",
                "team_b_label": "RPC",
                "duration_minutes": 60,
                "resource_type": "Gym Court",
                "earliest_slot": None,
                "latest_slot": None,
            }
        ],
        "resources": [
            {
                "resource_id": "GYM-Sat-1-1",
                "resource_type": "Gym Court",
                "label": "Court-1",
                "day": "Sat-1",
                "open_time": "08:00",
                "close_time": "09:00",
                "slot_minutes": 60,
                "exclusive_group": "",
            }
        ],
        "playoff_slots": [],
    }
    schedule_output = {
        "solved_at": "2026-05-01T10:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {"game_id": "BBM-01", "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-08:00"}
        ],
        "unscheduled": [],
        "conflict_audit_summary": {
            "total_edges": 1,
            "separated_edges": 1,
            "overlapping_edges": 0,
            "incomplete_edges": 0,
            "remaining_primary_overlap_penalty": 0,
            "remaining_secondary_overlap_penalty": 0,
        },
        "conflict_audit": [
            {
                "team_a_label": "OCB",
                "event_a": "Basketball - Men Team",
                "team_b_label": "OCB",
                "event_b": "Volleyball - Men Team",
                "shared_count": 2,
                "primary_overlap_count": 2,
                "secondary_only_count": 0,
                "status": "SeparatedInSchedule",
                "overlap_count": 0,
                "scheduled_team_a_games": 1,
                "scheduled_team_b_games": 1,
                "shared_participant_names": "An, Binh",
                "overlap_game_pairs": "",
            }
        ],
    }
    return schedule_output, schedule_input


def test_church_teams_exporter_schedule_methods_delegate_to_builder():
    """ChurchTeamsExporter should reuse the extracted scheduling implementation."""
    method_names = (
        "_build_schedule_input",
        "_write_schedule_input_tab",
        "_build_schedule_output_flat_rows",
        "_write_schedule_output_report",
        "_build_gym_resources_from_allocator",
        "_partner_validation_key",
        "_reconcile_pod_validation",
        "write_schedule_input_json",
        "write_schedule_workbook",
        "write_schedule_output_workbook",
    )
    for method_name in method_names:
        assert ChurchTeamsExporter.__dict__[method_name] is ScheduleWorkbookBuilder.__dict__[method_name]


def test_church_teams_exporter_reconciliation_has_delegated_key_helper():
    """The live export path must carry reconciliation's helper dependencies."""
    unprotected = [{
        "participant_id": "99",
        "participant_name": "Mai Tran",
        "church_code": "TLC",
        "division_id": "BAD-Women-Doubles",
        "sport_type": SPORT_TYPE["BADMINTON"],
        "sport_format": "Doubles",
        "sport_gender": "Women",
        "reason": "MissingPartner",
    }]
    validation_rows = [{
        "Participant ID (WP)": "99",
        "sport_type": SPORT_TYPE["BADMINTON"],
        "sport_format": "Women Double",
        "Issue Type": "missing_doubles_partner",
        "Status": "open",
    }]

    reconciliation = ChurchTeamsExporter._reconcile_pod_validation(
        unprotected,
        {},
        validation_rows,
    )

    assert reconciliation["is_clean"] is True
    assert reconciliation["matched_count"] == 1
    assert unprotected[0]["validation_issue_status"] == "Matched"


def test_write_schedule_input_json_writes_file(tmp_path):
    """The extracted builder should own the schedule_input.json write path."""
    builder = ScheduleWorkbookBuilder()
    json_path = tmp_path / "schedule_input.json"

    schedule_input = builder.write_schedule_input_json(
        _make_gym_roster(),
        [],
        tmp_path / "missing.xlsx",
        json_path,
    )

    assert json_path.exists()
    written = json.loads(json_path.read_text(encoding="utf-8"))
    assert written["game_count"] == schedule_input["game_count"]
    assert written["resource_count"] == schedule_input["resource_count"]
    assert written["games"]
    assert written["resources"]


def test_write_schedule_workbook_creates_planning_tabs(tmp_path):
    """The planning workbook entry point should always create the operator + planning tabs."""
    builder = ScheduleWorkbookBuilder()
    roster_rows = _make_gym_roster()
    schedule_input = builder.write_schedule_input_json(
        roster_rows,
        [],
        tmp_path / "missing.xlsx",
        tmp_path / "schedule_input.json",
    )
    schedule_input["gym_modes"] = {
        "Main Gym": {
            "Basketball Court": 2,
            "Volleyball Court": 1,
        }
    }
    schedule_input["gym_allocation"] = {
        "source": "allocator",
        "switch_count": 1,
        "decisions": [
            {
                "gym_name": "Main Gym",
                "day": "Sat-1",
                "open_time": "08:00",
                "close_time": "12:00",
                "mode": "Basketball Court",
                "courts": 2,
                "slot_minutes": 60,
            }
        ],
        "mode_demand": {"Basketball Court": 4},
        "mode_supply": {"Basketball Court": 4},
        "mode_shortfall": {"Basketball Court": 0},
    }
    workbook_path = tmp_path / "schedule_workbook.xlsx"

    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        schedule_input,
        tmp_path / "missing.xlsx",
    )

    wb = load_workbook(workbook_path)
    assert wb.sheetnames == [
        "Summary",
        "Venue-Estimator",
        "Pool-Assignment",
        "Pod-Divisions",
        "Pod-Entries-Review",
        "Court-Schedule-Sketch",
        "Pod-Resource-Estimate",
        "Schedule-Input",
        "Gym-Allocation",
    ]
    summary_ws = wb["Summary"]
    assert summary_ws["A1"].value == "VAY Sports Fest — Schedule Workbook Guide"
    assert "Layer 1 planning aid" in str(summary_ws["B4"].value)
    summary_text = "".join(
        str(summary_ws.cell(row=row, column=2).value or "")
        for row in range(1, summary_ws.max_row + 1)
    )
    assert "Church_Team_Status_ALL" in summary_text
    assert "run-schedule.bat" in summary_text
    assert "assign-pools" in summary_text
    assert "BB/VBM/VBW/BC/SOC" in summary_text
    assert "STATUS: READ-ONLY OUTPUT" in _status_banner_values(summary_ws)
    assert "STATUS: TEMPORARY EDIT SURFACE" in _status_banner_values(wb["Pool-Assignment"])
    assert "STATUS: GENERATED LOOKUP / MACHINE CONTRACT VIEW" in _status_banner_values(
        wb["Schedule-Input"]
    )
    assert "STATUS: READ-ONLY ALLOCATION AUDIT" in _status_banner_values(wb["Gym-Allocation"])
    venue_ws = wb["Venue-Estimator"]
    assert venue_ws["A1"].comment is not None
    assert "Canonical event name" in venue_ws["A1"].comment.text
    pool_ws = wb["Pool-Assignment"]
    assert ScheduleWorkbookBuilder._POOL_ASSIGNMENT_NOTE in str(pool_ws["A1"].value or "")
    assert pool_ws["A2"].comment is not None
    assert "Canonical team-sport event" in pool_ws["A2"].comment.text
    assert "SOC" in pool_ws["A2"].comment.text
    assert pool_ws["I2"].comment is not None
    assert "Leave blank or enter 0" in pool_ws["I2"].comment.text
    pod_ws = wb["Pod-Divisions"]
    assert pod_ws["A1"].comment is not None
    assert "Canonical division label" in pod_ws["A1"].comment.text
    assert pod_ws["K1"].comment is not None
    assert "Ready = clean to plan" in pod_ws["K1"].comment.text
    entry_ws = wb["Pod-Entries-Review"]
    assert entry_ws["A1"].comment is not None
    assert "Unique row ID" in entry_ws["A1"].comment.text
    sketch_ws = wb["Court-Schedule-Sketch"]
    assert sketch_ws["A4"].comment is not None
    assert "Start time for the slot" in sketch_ws["A4"].comment.text
    resource_ws = wb["Pod-Resource-Estimate"]
    assert resource_ws["A1"].comment is not None
    assert "Racquet event" in resource_ws["A1"].comment.text
    schedule_input_ws = wb["Schedule-Input"]
    games_header_row = _find_row_by_first_cell(schedule_input_ws, "GAMES") + 1
    resources_header_row = _find_row_by_first_cell(schedule_input_ws, "RESOURCES") + 1
    playoff_header_row = _find_row_by_first_cell(schedule_input_ws, "PLAYOFF-SLOTS") + 1
    assert schedule_input_ws.cell(row=games_header_row, column=1).comment is not None
    assert (
        "Unique placeholder game ID"
        in schedule_input_ws.cell(row=games_header_row, column=1).comment.text
    )
    assert schedule_input_ws.cell(row=resources_header_row, column=1).comment is not None
    assert (
        "Exact solver resource ID"
        in schedule_input_ws.cell(row=resources_header_row, column=1).comment.text
    )
    assert schedule_input_ws.cell(row=playoff_header_row, column=1).comment is not None
    assert (
        "Exact playoff game ID"
        in schedule_input_ws.cell(row=playoff_header_row, column=1).comment.text
    )
    gym_ws = wb["Gym-Allocation"]
    assert gym_ws["A4"].comment is not None
    assert "Exclusive Venue Group / gym block name" in gym_ws["A4"].comment.text
    assert gym_ws["A8"].comment is not None
    assert "Sport mode or resource mode" in gym_ws["A8"].comment.text


def test_write_schedule_workbook_loads_pool_assignment_sidecar(tmp_path):
    builder = ScheduleWorkbookBuilder()
    roster_rows = _make_gym_roster()
    schedule_input = builder.write_schedule_input_json(
        roster_rows,
        [],
        tmp_path / "missing.xlsx",
        tmp_path / "schedule_input.json",
    )
    sidecar_path = tmp_path / "pool_assignments.json"
    sidecar_path.write_text(
        json.dumps(
            {
                "version": 1,
                "rows": [
                    {
                        "event": SPORT_TYPE["BASKETBALL"],
                        "team_id": "RPC",
                        "seed": 1,
                        "random_draw_order": 4,
                        "notes": "Returning champion",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    workbook_path = tmp_path / "schedule_workbook.xlsx"

    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        schedule_input,
        tmp_path / "missing.xlsx",
        pool_assignment_path=sidecar_path,
    )

    ws = load_workbook(workbook_path)["Pool-Assignment"]
    rpc_row = next(
        row for row in _sheet_rows(ws)
        if row["Event"] == SPORT_TYPE["BASKETBALL"] and row["Team ID"] == "RPC"
    )
    assert rpc_row["Seed"] == 1
    assert rpc_row["Notes"] == "Returning champion"
    assert rpc_row["Pool ID"] is not None
    assert rpc_row["Pool Slot"] is not None


def test_refresh_pool_assignments_persists_seed_edits_and_recomputes_draw(tmp_path):
    builder = ScheduleWorkbookBuilder()
    roster_rows = _make_gym_roster(n_churches=4)
    schedule_input = builder.write_schedule_input_json(
        roster_rows,
        [],
        tmp_path / "missing.xlsx",
        tmp_path / "schedule_input.json",
    )
    workbook_path = tmp_path / "schedule_workbook.xlsx"
    sidecar_path = tmp_path / "pool_assignments.json"

    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        schedule_input,
        tmp_path / "missing.xlsx",
        pool_assignment_path=sidecar_path,
    )

    wb = load_workbook(workbook_path)
    ws = wb["Pool-Assignment"]
    header_row = _header_row(ws)
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    seed_col = headers.index("Seed") + 1
    notes_col = headers.index("Notes") + 1
    team_id_col = headers.index("Team ID") + 1
    for row_idx in range(header_row + 1, ws.max_row + 1):
        team_id = ws.cell(row=row_idx, column=team_id_col).value
        if team_id == "RPC":
            ws.cell(row=row_idx, column=seed_col, value=1)
            ws.cell(row=row_idx, column=notes_col, value="Top seed")
        elif team_id == "ANH":
            ws.cell(row=row_idx, column=seed_col, value=2)
    wb.save(workbook_path)

    builder.refresh_pool_assignments(
        workbook_path,
        sidecar_path=sidecar_path,
    )

    payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
    saved_rpc = next(
        row for row in payload["rows"]
        if row["event"] == SPORT_TYPE["BASKETBALL"] and row["team_id"] == "RPC"
    )
    assert saved_rpc["seed"] == 1
    assert saved_rpc["notes"] == "Top seed"

    refreshed_ws = load_workbook(workbook_path)["Pool-Assignment"]
    refreshed_rows = _sheet_rows(refreshed_ws)
    rpc_row = next(
        row for row in refreshed_rows
        if row["Event"] == SPORT_TYPE["BASKETBALL"] and row["Team ID"] == "RPC"
    )
    anh_row = next(
        row for row in refreshed_rows
        if row["Event"] == SPORT_TYPE["BASKETBALL"] and row["Team ID"] == "ANH"
    )
    assert rpc_row["Draw Position"] == 1
    assert anh_row["Draw Position"] == 2


def test_refresh_pool_assignments_flags_duplicate_seeds(tmp_path):
    builder = ScheduleWorkbookBuilder()
    roster_rows = _make_gym_roster(n_churches=4)
    schedule_input = builder.write_schedule_input_json(
        roster_rows,
        [],
        tmp_path / "missing.xlsx",
        tmp_path / "schedule_input.json",
    )
    workbook_path = tmp_path / "schedule_workbook.xlsx"
    sidecar_path = tmp_path / "pool_assignments.json"

    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        schedule_input,
        tmp_path / "missing.xlsx",
        pool_assignment_path=sidecar_path,
    )

    wb = load_workbook(workbook_path)
    ws = wb["Pool-Assignment"]
    header_row = _header_row(ws)
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    seed_col = headers.index("Seed") + 1
    team_id_col = headers.index("Team ID") + 1
    for row_idx in range(header_row + 1, ws.max_row + 1):
        team_id = ws.cell(row=row_idx, column=team_id_col).value
        if team_id in {"RPC", "ANH"}:
            ws.cell(row=row_idx, column=seed_col, value=1)
    wb.save(workbook_path)

    builder.refresh_pool_assignments(
        workbook_path,
        sidecar_path=sidecar_path,
    )

    refreshed_rows = _sheet_rows(load_workbook(workbook_path)["Pool-Assignment"])
    duplicate_rows = [
        row for row in refreshed_rows
        if row["Event"] == SPORT_TYPE["BASKETBALL"] and row["Team ID"] in {"RPC", "ANH"}
    ]
    assert len(duplicate_rows) == 2
    assert {row["Assignment Basis"] for row in duplicate_rows} == {"SeededDuplicate"}


def test_write_schedule_output_workbook_creates_schedule_tabs(tmp_path):
    """The output-workbook entry point should produce the renderer and audit tabs."""
    schedule_output, schedule_input = _make_schedule_pair()
    workbook_path = tmp_path / "schedule_output.xlsx"

    ScheduleWorkbookBuilder.write_schedule_output_workbook(
        workbook_path,
        schedule_output,
        schedule_input,
    )

    wb = load_workbook(workbook_path)
    assert wb.sheetnames == [
        "Schedule-by-Time",
        "Schedule-by-Sport",
        "Conflict-Audit",
        "Schedule-Diagnostics",
        "Master-Schedule",
    ]
    assert "STATUS: FINAL OUTPUT" in _status_banner_values(wb["Schedule-by-Time"])
    assert "STATUS: AUDIT OUTPUT" in _status_banner_values(wb["Conflict-Audit"])
    assert "STATUS: DIAGNOSTIC OUTPUT" in _status_banner_values(wb["Schedule-Diagnostics"])


def test_read_roster_validation_rows_missing_path_degrades():
    """A missing ALL workbook should yield empty lists, not raise."""
    roster_rows, validation_rows = ScheduleWorkbookBuilder.read_roster_validation_rows(None)
    assert roster_rows == []
    assert validation_rows == []


def test_read_roster_validation_rows_parses_tabs(tmp_path):
    """Roster and Validation-Issues tabs round-trip into builder-shaped dicts."""
    xlsx_path = tmp_path / "Church_Team_Status_ALL.xlsx"
    roster_df = pd.DataFrame(
        [
            {
                "Church Team": "RPC",
                "sport_type": SPORT_TYPE["BASKETBALL"],
                "sport_gender": "Men",
                "sport_format": "Team",
                "Participant ID (WP)": 1,
                "First Name": "An",
                "Last Name": "Nguyen",
                "partner_name": None,
            }
        ]
    )
    validation_df = pd.DataFrame(
        [
            {
                "Church Team": "RPC",
                "Severity": "ERROR",
                "Status": "open",
                "Participant ID (WP)": 1,
                "sport_type": SPORT_TYPE["BASKETBALL"],
            }
        ]
    )
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        roster_df.to_excel(writer, sheet_name="Roster", index=False)
        validation_df.to_excel(writer, sheet_name="Validation-Issues", index=False)

    roster_rows, validation_rows = ScheduleWorkbookBuilder.read_roster_validation_rows(
        xlsx_path
    )

    assert len(roster_rows) == 1
    assert roster_rows[0]["Church Team"] == "RPC"
    # NaN cells normalize to None so `str(v or "")` collapses blanks cleanly.
    assert roster_rows[0]["partner_name"] is None
    assert len(validation_rows) == 1
    assert validation_rows[0]["Severity"] == "ERROR"


def test_read_roster_validation_rows_missing_tab_degrades(tmp_path):
    """A workbook without the expected tabs degrades to empty lists."""
    xlsx_path = tmp_path / "no_tabs.xlsx"
    pd.DataFrame([{"x": 1}]).to_excel(xlsx_path, sheet_name="Summary", index=False)

    roster_rows, validation_rows = ScheduleWorkbookBuilder.read_roster_validation_rows(
        xlsx_path
    )
    assert roster_rows == []
    assert validation_rows == []


def test_load_available_slots_from_schedule_input_counts_slots():
    schedule_input = {
        "resources": [
            {
                "resource_id": "PCK-1",
                "resource_type": "Pickleball Court",
                "label": "Court-1",
                "day": "Day-1",
                "open_time": "09:00",
                "close_time": "10:00",
                "slot_minutes": 30,
            },
            {
                "resource_id": "PCK-2",
                "resource_type": "Pickleball Court",
                "label": "Court-2",
                "day": "Day-1",
                "open_time": "09:00",
                "close_time": "09:30",
                "slot_minutes": 30,
            },
        ]
    }

    totals = ScheduleWorkbookBuilder._load_available_slots_from_schedule_input(
        schedule_input
    )

    assert totals["Pickleball Court"] == 3


def test_write_schedule_workbook_uses_schedule_input_resources_offline(tmp_path):
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {
            "Church Team": "RPC",
            "sport_type": SPORT_TYPE["PICKLEBALL"],
            "sport_gender": "Mixed",
            "sport_format": "Singles",
            "Participant ID (WP)": 1,
            "First Name": "Alex",
            "Last Name": "Tran",
        }
    ]
    schedule_input = {
        "generated_at": "2026-05-16T00:00:00",
        "gym_court_scenario": 4,
        "game_count": 0,
        "resource_count": 1,
        "games": [],
        "resources": [
            {
                "resource_id": "PCK-1",
                "resource_type": "Pickleball Court",
                "label": "Court-1",
                "day": "Day-1",
                "open_time": "09:00",
                "close_time": "10:00",
                "slot_minutes": 30,
                "exclusive_group": "",
            }
        ],
        "playoff_slots": [],
        "gym_modes": {},
    }
    workbook_path = tmp_path / "offline_schedule_workbook.xlsx"

    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        schedule_input,
        venue_input_path=None,
    )

    wb = load_workbook(workbook_path)
    ws = wb["Pod-Resource-Estimate"]

    found_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == SPORT_TYPE["PICKLEBALL"]:
            found_row = row
            break

    assert found_row is not None
    assert ws.cell(row=found_row, column=5).value == 2
    all_values = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value is not None
    ]
    assert any(
        "schedule_input.json resources" in str(value) for value in all_values
    )


# ===========================================================================
# Pure scheduling-method tests migrated from test_church_teams_export.py
# (Issue #98 Step 4 — ScheduleWorkbookBuilder extraction)
# ===========================================================================


def _write_venue_input(path, headers, data_rows, gym_modes_rows=None, playoff_rows=None):
    """Write a venue_input.xlsx with a Venue-Input tab (and optional Gym-Modes /
    Playoff-Slots tabs; the first row of each optional tab is its header row)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, data in enumerate(data_rows, start=2):
        for c, val in enumerate(data, start=1):
            ws.cell(row=r, column=c, value=val)
    if gym_modes_rows is not None:
        gm = wb.create_sheet("Gym-Modes")
        for r, data in enumerate(gym_modes_rows, start=1):
            for c, val in enumerate(data, start=1):
                gm.cell(row=r, column=c, value=val)
    if playoff_rows is not None:
        ps = wb.create_sheet("Playoff-Slots")
        for r, data in enumerate(playoff_rows, start=1):
            for c, val in enumerate(data, start=1):
                ps.cell(row=r, column=c, value=val)
    wb.save(path)


def _make_render_schedule_pair():
    """Return (schedule_output, schedule_input) test fixtures."""
    schedule_input = {
        "games": [
            {
                "game_id": "BBM-01", "event": "Basketball - Men Team",
                "stage": "Pool", "pool_id": "P1", "round": 1,
                "team_a_id": "BBM-P1-T1", "team_b_id": "BBM-P1-T2",
                "duration_minutes": 60, "resource_type": "Gym Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "BBM-Final", "event": "Basketball - Men Team",
                "stage": "Final", "pool_id": "", "round": 1,
                "team_a_id": "WIN-BBM-Semi-1", "team_b_id": "WIN-BBM-Semi-2",
                "duration_minutes": 60, "resource_type": "Gym Court",
                "earliest_slot": None, "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "GYM-Sat-1-1", "resource_type": "Gym Court",
                "label": "Court-1", "day": "Sat-1",
                "open_time": "08:00", "close_time": "12:00", "slot_minutes": 60,
            }
        ],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-01T10:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {"game_id": "BBM-01",    "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-08:00"},
            {"game_id": "BBM-Final", "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-10:00"},
        ],
        "unscheduled": [],
    }
    return schedule_output, schedule_input


def test_venue_capacity_court_slot_math():
    """Pool/playoff/total slot math (Issue #83)."""
    builder = ScheduleWorkbookBuilder()

    # 0 teams -> all zeros
    s0 = builder._compute_court_slots(0)
    assert s0["pool_slots"] == 0
    assert s0["playoff_teams"] == 0
    assert s0["playoff_slots"] == 0
    assert s0["total_slots"] == 0
    assert s0["court_hours"] == 0.0

    # 6 teams, 2 pool games each -> ceil(6*2/2) = 6 pool, 4-team playoff = 3 playoff games
    s6 = builder._compute_court_slots(6)
    assert s6["pool_slots"] == 6
    assert s6["playoff_teams"] == 4
    assert s6["playoff_slots"] == 3
    assert s6["third_place_slots"] == 1
    assert s6["total_slots"] == 10
    assert s6["court_hours"] == 10.0  # 60 min/game

    # 8 teams -> ceil(8*2/2)=8 pool, 8-team playoff = 7 playoff games
    s8 = builder._compute_court_slots(8)
    assert s8["pool_slots"] == 8
    assert s8["playoff_teams"] == 8
    assert s8["playoff_slots"] == 7
    assert s8["third_place_slots"] == 1
    assert s8["total_slots"] == 16

    # 3 teams -> only pool play, no playoff
    s3 = builder._compute_court_slots(3)
    assert s3["pool_slots"] == 3
    assert s3["playoff_teams"] == 0
    assert s3["playoff_slots"] == 0
    assert s3["total_slots"] == 3


def test_count_estimating_teams_uses_min_team_size():
    """A church only counts when its roster meets the min team size (Issue #83)."""
    builder = ScheduleWorkbookBuilder()

    # RPC has 5 basketball players (meets min=5), TLC has 4 (potential only)
    roster_rows = [
        {"Church Team": "RPC", "sport_type": "Basketball", "sport_gender": "Men"} for _ in range(5)
    ] + [
        {"Church Team": "TLC", "sport_type": "Basketball", "sport_gender": "Men"} for _ in range(4)
    ]

    result = builder._count_estimating_teams(roster_rows, "Basketball - Men Team", min_team_size=5)
    assert result["n_estimating"] == 1       # only RPC qualifies
    assert result["n_potential"] == 2        # RPC (estimating) + TLC (partial) = all with >= 1
    assert result["team_codes"] == "RPC"     # sorted, comma-separated


def test_count_estimating_teams_separates_volleyball_men_and_women():
    """Volleyball Men and Women are distinct events; team_codes is sorted (Issue #83)."""
    builder = ScheduleWorkbookBuilder()

    roster_rows = (
        [{"Church Team": "RPC", "sport_type": "Volleyball", "sport_gender": "Men"} for _ in range(6)]
        + [{"Church Team": "RPC", "sport_type": "Volleyball", "sport_gender": "Women"} for _ in range(6)]
        + [{"Church Team": "TLC", "sport_type": "Volleyball", "sport_gender": "Women"} for _ in range(6)]
    )

    men = builder._count_estimating_teams(roster_rows, "Volleyball - Men Team", 6)
    assert men["n_estimating"] == 1
    assert men["team_codes"] == "RPC"

    women = builder._count_estimating_teams(roster_rows, "Volleyball - Women Team", 6)
    assert women["n_estimating"] == 2
    assert women["team_codes"] == "RPC, TLC"  # alphabetically sorted


def test_count_estimating_teams_soccer_full_label():
    """Soccer sport_type is stored as the full Other-Events label, not just 'Soccer'."""
    builder = ScheduleWorkbookBuilder()

    # Other-events registrations store the full SPORT_TYPE constant value verbatim
    roster_rows = [
        {"Church Team": "RPC", "sport_type": SPORT_TYPE["SOCCER"], "sport_gender": "Mixed"}
        for _ in range(5)
    ] + [
        {"Church Team": "TLC", "sport_type": SPORT_TYPE["SOCCER"], "sport_gender": "Mixed"}
        for _ in range(3)
    ]

    result = builder._count_estimating_teams(
        roster_rows, SPORT_TYPE["SOCCER"], min_team_size=4
    )
    assert result["n_estimating"] == 1      # only RPC has >= 4
    assert result["n_potential"] == 2       # RPC + TLC both have >= 1
    assert result["team_codes"] == "RPC"


def test_count_estimating_teams_respects_explicit_team_order():
    """Explicit A/B team splits should count as separate estimating teams."""
    builder = ScheduleWorkbookBuilder()

    roster_rows = (
        [
            {
                "Church Team": "RPC",
                "team_order": "A",
                "sport_type": "Basketball",
                "sport_gender": "Men",
            }
            for _ in range(5)
        ]
        + [
            {
                "Church Team": "RPC",
                "team_order": "B",
                "sport_type": "Basketball",
                "sport_gender": "Men",
            }
            for _ in range(5)
        ]
        + [
            {
                "Church Team": "OCB",
                "team_order": "A",
                "sport_type": "Basketball",
                "sport_gender": "Men",
            }
            for _ in range(5)
        ]
    )

    result = builder._count_estimating_teams(
        roster_rows, "Basketball - Men Team", min_team_size=5
    )

    assert result["n_estimating"] == 3
    assert result["n_potential"] == 3
    assert result["team_codes"] == "OCB-A, RPC-A, RPC-B"


def test_count_racquet_entries():
    """Racquet potential should respect church-level 2026 caps while counting complete pairs normally."""
    builder = ScheduleWorkbookBuilder()

    roster_rows = [
        # RPC: 5 mixed doubles registrations -> ceil(5/2)=3 teams, capped to 1
        {
            "Church Team": "RPC",
            "sport_type": "Badminton",
            "sport_gender": "Mixed",
            "sport_format": "Mixed Doubles",
        }
        for _ in range(5)
    ] + [
        # RPC: 2 men doubles registrations -> 1 team, within cap 2
        {
            "Church Team": "RPC",
            "sport_type": "Badminton",
            "sport_gender": "Men",
            "sport_format": "Men Double",
        }
        for _ in range(2)
    ] + [
        # ANH: singles are not allowed in badminton, so they should not raise the potential ceiling
        {
            "Church Team": "ANH",
            "sport_type": "Badminton",
            "sport_gender": "Men",
            "sport_format": "Men Singles",
        },
        {
            "Church Team": "ANH",
            "sport_type": "Badminton",
            "sport_gender": "Women",
            "sport_format": "Women Singles",
        },
        # Another sport should not bleed into Badminton count
        {
            "Church Team": "RPC",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Mixed Doubles",
        },
    ]

    result = builder._count_racquet_entries(roster_rows, "Badminton")
    assert result["n_estimating"] == 5  # floor(7/2)=3 doubles teams + 2 singles
    assert result["n_potential"] == 2   # RPC mixed=1 cap, RPC men=1 cap, ANH singles cap to 0
    assert result["team_codes"] == ""


def test_count_racquet_entries_applies_format_total_caps():
    """Format-total rules should cap the combined doubles ceiling per church."""
    builder = ScheduleWorkbookBuilder()

    roster_rows = (
        [
            {
                "Church Team": "RPC",
                "sport_type": "Pickleball",
                "sport_gender": "Men",
                "sport_format": "Men Doubles",
            }
            for _ in range(4)
        ]
        + [
            {
                "Church Team": "RPC",
                "sport_type": "Pickleball",
                "sport_gender": "Women",
                "sport_format": "Women Doubles",
            }
            for _ in range(4)
        ]
        + [
            {
                "Church Team": "RPC",
                "sport_type": "Pickleball",
                "sport_gender": "Mixed",
                "sport_format": "Mixed Doubles",
            }
            for _ in range(6)
        ]
    )

    result = builder._count_racquet_entries(roster_rows, "Pickleball")
    assert result["n_estimating"] == 7  # floor(14/2)
    assert result["n_potential"] == 3   # church total capped by Pickleball doubles total rule


def test_pod_format_class():
    builder = ScheduleWorkbookBuilder()
    assert builder._pod_format_class("Men Single") == "singles"
    assert builder._pod_format_class("Singles") == "singles"
    assert builder._pod_format_class("Women Singles") == "singles"
    assert builder._pod_format_class("Men Double") == "doubles"
    assert builder._pod_format_class("Doubles") == "doubles"
    assert builder._pod_format_class("Mixed Double") == "doubles"
    assert builder._pod_format_class("Team") == "anomaly"
    assert builder._pod_format_class("") == "anomaly"
    assert builder._pod_format_class(None) == "anomaly"


def test_make_division_id():
    builder = ScheduleWorkbookBuilder()
    assert builder._make_division_id("Badminton", "Men", "singles") == "BAD-Men-Singles"
    assert builder._make_division_id("Table Tennis", "Women", "doubles") == "TT-Women-Doubles"
    assert builder._make_division_id("Pickleball 35+", "Mixed", "doubles") == "PCK35-Mixed-Doubles"
    assert builder._make_division_id("Tennis", "Men", "anomaly") == "TEN-Men-Anomaly"
    assert builder._make_division_id("Table Tennis 35+", "Men", "singles") == "TT35-Men-Singles"


def test_build_pod_error_lookup():
    builder = ScheduleWorkbookBuilder()
    validation_rows = [
        {
            "Participant ID (WP)": "42",
            "sport_type": "Badminton",
            "Severity": "ERROR",
            "Status": "open",
        },
        {
            "Participant ID (WP)": "42",
            "sport_type": "Pickleball",
            "Severity": "WARNING",  # warnings excluded
            "Status": "open",
        },
        {
            "Participant ID (WP)": "99",
            "sport_type": "Table Tennis",
            "Severity": "ERROR",
            "Status": "resolved",  # resolved excluded
        },
        {
            "Participant ID (WP)": "55",
            "sport_type": "Tennis",
            "Severity": "ERROR",
            "Status": "open",
        },
    ]
    lookup = builder._build_pod_error_lookup(validation_rows)
    assert lookup == {"42": {"Badminton"}, "55": {"Tennis"}}


def test_build_pod_divisions_rows_singles():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "1", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "2", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "3", "Church Team": "TLC"},
    ]
    # Participant 2 has an error
    validation_rows = [
        {"Participant ID (WP)": "2", "sport_type": "Badminton", "Severity": "ERROR", "Status": "open"},
    ]
    rows = builder._build_pod_divisions_rows(roster_rows, validation_rows)

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "BAD-Men-Singles"
    assert div["sport_type"] == "Badminton"
    assert div["resource_type"] == "Badminton Court"
    assert div["planning_entries"] == 3
    assert div["confirmed_entries"] == 2  # participant 2 has error
    assert div["provisional_entries"] == 1
    assert div["anomaly_count"] == 0
    assert div["division_status"] == "Partial"


def test_build_pod_divisions_rows_doubles():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "10", "Church Team": "RPC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "11", "Church Team": "RPC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "12", "Church Team": "TLC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "13", "Church Team": "TLC"},
    ]
    rows = builder._build_pod_divisions_rows(roster_rows, [])

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "TT-Men-Doubles"
    assert div["resource_type"] == "Table Tennis Table"
    assert div["planning_entries"] == 2   # floor(4/2)
    assert div["confirmed_entries"] == 2  # no errors
    assert div["provisional_entries"] == 0
    assert div["division_status"] == "Ready"


def test_build_pod_divisions_rows_anomaly():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Pickleball", "sport_gender": "Men", "sport_format": "Team",
         "Participant ID (WP)": "20", "Church Team": "RPC"},
    ]
    rows = builder._build_pod_divisions_rows(roster_rows, [])

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "PCK-Men-Anomaly"
    assert div["planning_entries"] == 0
    assert div["confirmed_entries"] == 0
    assert div["anomaly_count"] == 1
    assert div["division_status"] == "AnomalyOnly"


def test_build_pod_entries_review_singles():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Tennis", "sport_gender": "Women", "sport_format": "Women Single",
         "Participant ID (WP)": "30", "First Name": "Lan", "Last Name": "Tran", "Church Team": "RPC"},
        {"sport_type": "Tennis", "sport_gender": "Women", "sport_format": "Women Single",
         "Participant ID (WP)": "31", "First Name": "Hoa", "Last Name": "Le", "Church Team": "RPC"},
    ]
    validation_rows = [
        {"Participant ID (WP)": "31", "sport_type": "Tennis", "Severity": "ERROR", "Status": "open"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, validation_rows)

    assert len(rows) == 2
    singles = [r for r in rows if r["entry_type"] == "Singles"]
    assert len(singles) == 2

    lan = next(r for r in singles if r["participant_1_name"] == "Lan Tran")
    assert lan["review_status"] == "OK"
    assert lan["partner_status"] == "N/A"
    assert lan["division_id"] == "TEN-Women-Singles"

    hoa = next(r for r in singles if r["participant_1_name"] == "Hoa Le")
    assert hoa["review_status"] == "NeedsReview"


def test_build_pod_entries_review_doubles_reciprocal():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "40", "First Name": "Anh", "Last Name": "Nguyen",
         "partner_name": "Binh Tran", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "41", "First Name": "Binh", "Last Name": "Tran",
         "partner_name": "Anh Nguyen", "Church Team": "RPC"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    pair = rows[0]
    assert pair["entry_type"] == "DoublesPair"
    assert pair["partner_status"] == "Confirmed"
    assert pair["review_status"] == "OK"
    assert "Anh Nguyen" in pair["participant_1_name"] or "Binh Tran" in pair["participant_1_name"]
    assert "Anh Nguyen" in pair["participant_2_name"] or "Binh Tran" in pair["participant_2_name"]
    assert pair["church_team"] == "RPC"


def test_build_pod_entries_review_doubles_cross_church_unresolved():
    """Cross-church partner declarations must not confirm; both stay unresolved."""
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "40", "First Name": "Anh", "Last Name": "Nguyen",
         "partner_name": "Binh Tran", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "41", "First Name": "Binh", "Last Name": "Tran",
         "partner_name": "Anh Nguyen", "Church Team": "TLC"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 2
    types = {r["entry_type"] for r in rows}
    assert types == {"UnresolvedDoubles"}
    assert all(r["partner_status"] != "Confirmed" for r in rows)


def test_build_pod_entries_review_doubles_missing_partner():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Pickleball", "sport_gender": "Women", "sport_format": "Women Double",
         "Participant ID (WP)": "50", "First Name": "Cam", "Last Name": "Ho",
         "partner_name": "", "Church Team": "RPC"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    entry = rows[0]
    assert entry["entry_type"] == "UnresolvedDoubles"
    assert entry["partner_status"] == "MissingPartner"
    assert entry["review_status"] == "NeedsReview"


def test_build_pod_entries_review_doubles_non_reciprocal():
    builder = ScheduleWorkbookBuilder()
    # A claims B, B claims someone else (non-reciprocal)
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Mixed", "sport_format": "Mixed Double",
         "Participant ID (WP)": "60", "First Name": "Dan", "Last Name": "Vo",
         "partner_name": "Linh Pham", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Mixed", "sport_format": "Mixed Double",
         "Participant ID (WP)": "61", "First Name": "Linh", "Last Name": "Pham",
         "partner_name": "Khoa Bui", "Church Team": "RPC"},  # claims Khoa, not Dan
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    unresolved = [r for r in rows if r["entry_type"] == "UnresolvedDoubles"]
    assert len(unresolved) >= 1
    reasons = {r["partner_status"] for r in unresolved}
    assert "NonReciprocal" in reasons


def test_build_pod_entries_review_doubles_self_paired():
    """A participant who lists their own name as partner is flagged as SelfPaired (Issue #160)."""
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "62", "First Name": "Anh", "Last Name": "Nguyen",
         "partner_name": "Anh Nguyen", "Church Team": "RPC"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    assert rows[0]["entry_type"] == "UnresolvedDoubles"
    assert rows[0]["partner_status"] == "SelfPaired"


def test_build_pod_unprotected_entries_enriched(tmp_path):
    """UnresolvedDoubles in pod_unprotected_entries carry participant_id, sport_type,
    sport_format, and church_code (Issue #160)."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Double", "Participant ID (WP)": "99",
         "First Name": "Mai", "Last Name": "Tran",
         "partner_name": "", "Church Team": "TLC",
         "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    si = builder._build_schedule_input(roster, [], tmp_path / "no_venue.xlsx")
    unprotected = si["pod_unprotected_entries"]
    assert len(unprotected) == 1
    rec = unprotected[0]
    assert rec["participant_id"] == "99"
    assert rec["sport_type"] == SPORT_TYPE["BADMINTON"]
    assert rec["church_code"] == "TLC"
    assert rec["reason"] == "MissingPartner"


def test_pod_validation_reconciliation_subset_holds(tmp_path):
    """Every scheduler-unprotected doubles registration with a matching open
    partner validation issue reconciles cleanly (Issue #160 acceptance)."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Double", "Participant ID (WP)": "99",
         "First Name": "Mai", "Last Name": "Tran",
         "partner_name": "", "Church Team": "TLC",
         "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    validation_rows = [
        {"Participant ID (WP)": "99", "sport_type": SPORT_TYPE["BADMINTON"],
         "sport_format": "Women Double",
         "Issue Type": "missing_doubles_partner", "Severity": "ERROR",
         "Status": "open"},
    ]
    si = builder._build_schedule_input(roster, validation_rows, tmp_path / "no_venue.xlsx")

    recon = si["pod_validation_reconciliation"]
    assert recon["matched_count"] == 1
    assert recon["missing_validation_issues"] == []
    assert recon["mismatched_issue_types"] == []
    assert recon["contradictory_open_issues"] == []
    assert si["pod_unprotected_entries"][0]["validation_issue_status"] == "Matched"


def test_pod_validation_reconciliation_detects_missing_issue(tmp_path):
    """An unprotected doubles registration without an open partner validation
    issue must be reported — the Church Rep workbook would otherwise hide it."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["TABLE_TENNIS"], "sport_gender": "Men",
         "sport_format": "Men Double", "Participant ID (WP)": "214",
         "First Name": "Duc", "Last Name": "Pham",
         "partner_name": "Khong Ai Ca", "Church Team": "TLC",
         "participant_primary_sport": SPORT_TYPE["TABLE_TENNIS"]},
    ]
    si = builder._build_schedule_input(roster, [], tmp_path / "no_venue.xlsx")

    recon = si["pod_validation_reconciliation"]
    assert recon["matched_count"] == 0
    assert len(recon["missing_validation_issues"]) == 1
    rec = recon["missing_validation_issues"][0]
    assert rec["participant_id"] == "214"
    assert rec["sport_type"] == SPORT_TYPE["TABLE_TENNIS"]
    assert rec["reason"] == "PartnerNotFound"
    assert rec["expected_issue_type"] == "doubles_partner_unmatched"
    assert (
        si["pod_unprotected_entries"][0]["validation_issue_status"]
        == "MissingValidationIssue"
    )


def test_pod_validation_reconciliation_isolated_by_format_and_gender():
    """One format's issue must not satisfy another division in the same sport."""
    builder = ScheduleWorkbookBuilder()
    unprotected = [
        {
            "participant_id": "7",
            "participant_name": "Player Seven",
            "church_code": "TLC",
            "division_id": "TT-Men-Doubles",
            "sport_type": SPORT_TYPE["TABLE_TENNIS"],
            "sport_format": "Doubles",
            "sport_gender": "Men",
            "reason": "MissingPartner",
        },
        {
            "participant_id": "7",
            "participant_name": "Player Seven",
            "church_code": "TLC",
            "division_id": "TT-Mixed-Doubles",
            "sport_type": SPORT_TYPE["TABLE_TENNIS"],
            "sport_format": "Doubles",
            "sport_gender": "Mixed",
            "reason": "MissingPartner",
        },
    ]
    validation_rows = [
        {
            "Participant ID (WP)": "7",
            "sport_type": SPORT_TYPE["TABLE_TENNIS"],
            "sport_format": "Men Double",
            "Issue Type": "missing_doubles_partner",
            "Status": "open",
        },
    ]

    recon = builder._reconcile_pod_validation(
        unprotected, {}, validation_rows
    )

    assert recon["matched_count"] == 1
    assert len(recon["missing_validation_issues"]) == 1
    assert recon["missing_validation_issues"][0]["division_id"] == "TT-Mixed-Doubles"
    assert [entry["validation_issue_status"] for entry in unprotected] == [
        "Matched",
        "MissingValidationIssue",
    ]


def test_pod_validation_reconciliation_detects_contradiction(tmp_path):
    """A confirmed reciprocal pair member with a still-open partner validation
    issue is contradictory: scheduler and validation disagree (Issue #160)."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Double", "Participant ID (WP)": "446",
         "First Name": "Anh", "Last Name": "Nguyen",
         "partner_name": "Binh Tran", "Church Team": "GAC",
         "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Double", "Participant ID (WP)": "447",
         "First Name": "Binh", "Last Name": "Tran",
         "partner_name": "Anh Nguyen", "Church Team": "GAC",
         "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    validation_rows = [
        {"Participant ID (WP)": "446", "sport_type": SPORT_TYPE["BADMINTON"],
         "sport_format": "Men Double",
         "Issue Type": "doubles_partner_unmatched", "Severity": "WARNING",
         "Status": "open"},
    ]
    si = builder._build_schedule_input(roster, validation_rows, tmp_path / "no_venue.xlsx")

    recon = si["pod_validation_reconciliation"]
    assert recon["missing_validation_issues"] == []
    assert len(recon["contradictory_open_issues"]) == 1
    rec = recon["contradictory_open_issues"][0]
    assert rec["participant_id"] == "446"
    assert rec["open_issue_types"] == ["doubles_partner_unmatched"]


def test_pod_validation_reconciliation_reports_validation_only(tmp_path):
    """An open partner issue whose participant is absent from the exported
    roster is reported as validation-only instead of silently ignored."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Double", "Participant ID (WP)": "271",
         "First Name": "Lan", "Last Name": "Vu",
         "partner_name": "", "Church Team": "WAG",
         "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    validation_rows = [
        {"Participant ID (WP)": "271", "sport_type": SPORT_TYPE["BADMINTON"],
         "sport_format": "Women Double",
         "Issue Type": "missing_doubles_partner", "Severity": "ERROR",
         "Status": "open"},
        # Participant 273 has an open issue but no roster row at all.
        {"Participant ID (WP)": "273", "sport_type": SPORT_TYPE["BADMINTON"],
         "sport_format": "Women Double",
         "Issue Type": "doubles_partner_unmatched", "Severity": "WARNING",
         "Status": "open"},
    ]
    si = builder._build_schedule_input(roster, validation_rows, tmp_path / "no_venue.xlsx")

    recon = si["pod_validation_reconciliation"]
    assert recon["matched_count"] == 1
    assert recon["missing_validation_issues"] == []
    assert len(recon["validation_only_issues"]) == 1
    assert recon["validation_only_issues"][0]["participant_id"] == "273"
    assert recon["is_clean"] is False
    assert recon["problem_count"] == 1


def test_pod_validation_reconciliation_resolved_issues_ignored(tmp_path):
    """Resolved partner issues do not satisfy the subset criterion — the
    matching issue must still be OPEN."""
    from config import SPORT_TYPE
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["PICKLEBALL"], "sport_gender": "Mixed",
         "sport_format": "Mixed Double", "Participant ID (WP)": "479",
         "First Name": "Hoa", "Last Name": "Dang",
         "partner_name": "", "Church Team": "NHC",
         "participant_primary_sport": SPORT_TYPE["PICKLEBALL"]},
    ]
    validation_rows = [
        {"Participant ID (WP)": "479", "sport_type": SPORT_TYPE["PICKLEBALL"],
         "sport_format": "Mixed Double",
         "Issue Type": "missing_doubles_partner", "Severity": "ERROR",
         "Status": "resolved"},
    ]
    si = builder._build_schedule_input(roster, validation_rows, tmp_path / "no_venue.xlsx")

    recon = si["pod_validation_reconciliation"]
    assert recon["matched_count"] == 0
    assert len(recon["missing_validation_issues"]) == 1


def test_build_pod_entries_review_anomaly():
    builder = ScheduleWorkbookBuilder()
    roster_rows = [
        {"sport_type": "Table Tennis 35+", "sport_gender": "Men", "sport_format": "Team",
         "Participant ID (WP)": "70", "First Name": "Tri", "Last Name": "Nguyen",
         "Church Team": "RPC"},
    ]
    rows = builder._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    assert rows[0]["entry_type"] == "Anomaly"
    assert rows[0]["review_status"] == "NeedsReview"
    assert "Team" in rows[0]["notes"]


def test_build_scenario_schedule_pool_before_playoffs():
    """Pool before early playoffs, early playoffs on sat2, finals pinned to sun2."""
    pool_queues = [
        [f"BBM-{i:02d}" for i in range(1, 5)],
        [f"VBM-{i:02d}" for i in range(1, 5)],
        [f"VBW-{i:02d}" for i in range(1, 5)],
    ]
    early_playoff_queues = [
        ["BBM-Semi-1", "BBM-Semi-2"],
        ["VBM-Semi-1", "VBM-Semi-2"],
        ["VBW-Semi-1", "VBW-Semi-2"],
    ]
    final_queues = [
        ["BBM-Final"],
        ["VBM-Final"],
        ["VBW-Final"],
    ]

    n_sat, n_sun = 13, 8

    for n_courts in [3, 4, 5]:
        grids = ScheduleWorkbookBuilder._build_scenario_schedule(
            n_courts, pool_queues, early_playoff_queues, final_queues, n_sat, n_sun
        )
        sat1_cells = [cell for row in grids[0] for cell in row if cell]
        sun1_cells = [cell for row in grids[1] for cell in row if cell]
        sat2_cells = [cell for row in grids[2] for cell in row if cell]
        sun2_cells = [cell for row in grids[3] for cell in row if cell]

        all_pool  = {g for q in pool_queues for g in q}
        all_early = {g for q in early_playoff_queues for g in q}
        all_final = {g for q in final_queues for g in q}

        # All pool games appear in sat1/sun1/sat2
        assert set(sat1_cells + sun1_cells + sat2_cells) & all_pool == all_pool, \
            f"n_courts={n_courts}: missing pool games"

        # No playoff/final games in sat1 or sun1
        assert not (set(sat1_cells) & (all_early | all_final)), \
            f"n_courts={n_courts}: playoff/final in sat1"
        assert not (set(sun1_cells) & (all_early | all_final)), \
            f"n_courts={n_courts}: playoff/final in sun1"

        # Early playoffs land on sat2, not sun2
        assert set(sat2_cells) & all_early == all_early, \
            f"n_courts={n_courts}: early playoffs missing from sat2"
        assert not (set(sun2_cells) & all_early), \
            f"n_courts={n_courts}: early playoffs leaked into sun2"

        # Finals land on sun2, not sat2
        assert set(sun2_cells) & all_final == all_final, \
            f"n_courts={n_courts}: finals missing from sun2"
        assert not (set(sat2_cells) & all_final), \
            f"n_courts={n_courts}: finals leaked into sat2"

        # Pool games never appear in sun2
        assert not (set(sun2_cells) & all_pool), \
            f"n_courts={n_courts}: pool games leaked into sun2"

        # Playoffs/finals stay on their primary court blocks
        n_sports = len(pool_queues)
        base = n_courts // n_sports
        extras = n_courts % n_sports
        cur = 0
        for sport_idx, (early_q, final_q) in enumerate(zip(early_playoff_queues, final_queues)):
            k = base + (1 if sport_idx < extras else 0)
            sport_courts = set(range(cur, cur + k))
            cur += k
            playoff_ids = set(early_q) | set(final_q)
            for sess_idx in [2, 3]:  # sat2, sun2 only
                for t, row in enumerate(grids[sess_idx]):
                    for c_idx, game_id in enumerate(row):
                        if game_id in playoff_ids:
                            assert c_idx in sport_courts, (
                                f"n_courts={n_courts} sport={sport_idx}: "
                                f"{game_id} on court {c_idx}, expected {sport_courts}"
                            )


def test_pod_resource_estimate_no_venue_input(tmp_path):
    """When no venue_input.xlsx exists, tab still renders with notice row."""
    builder = ScheduleWorkbookBuilder()
    # Provide zero registrations — all entries will be 0.
    available = {}  # empty — no venue file
    pod_rows = builder._build_pod_resource_rows([], available)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    builder._write_pod_resource_estimate(ws, pod_rows, available)

    # All Fit Status cells say "No venue data"
    fit_values = {ws.cell(row=r, column=7).value for r in range(2, 2 + len(pod_rows))}
    assert fit_values == {"No venue data"}, f"Unexpected fit values: {fit_values}"


def test_pod_resource_estimate_fit_status_rules():
    """Green/Yellow/Red thresholds from POD_FIT_YELLOW_MAX (= 3)."""
    builder = ScheduleWorkbookBuilder()
    from config import POD_RESOURCE_TYPE_BADMINTON, POD_FIT_YELLOW_MAX

    # Build roster rows: 10 Badminton singles → Required = 9
    roster_rows = [
        {"sport_type": "Badminton", "sport_format": "Men Singles"} for _ in range(10)
    ]

    # Green: available >= required (9)
    green = builder._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9}
    )
    badminton_green = next(r for r in green if "Badminton" in r["Event"])
    assert badminton_green["Fit Status"] == "Green"
    assert badminton_green["Surplus / Shortage"] == 0

    # Yellow: short by 1 to POD_FIT_YELLOW_MAX (3)
    yellow = builder._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9 - POD_FIT_YELLOW_MAX}
    )
    badminton_yellow = next(r for r in yellow if "Badminton" in r["Event"])
    assert badminton_yellow["Fit Status"] == "Yellow"

    # Red: short by more than POD_FIT_YELLOW_MAX
    red = builder._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9 - POD_FIT_YELLOW_MAX - 1}
    )
    badminton_red = next(r for r in red if "Badminton" in r["Event"])
    assert badminton_red["Fit Status"] == "Red"


def test_load_venue_input_aggregates_by_resource_type(tmp_path):
    """Available Slots are summed across multiple rows of the same resource type."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_PICKLEBALL

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # Two Pickleball Court rows: 24 + 18 = 42 total
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=2, column=9, value=24)
    ws.cell(row=3, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=3, column=9, value=18)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ScheduleWorkbookBuilder._load_venue_input(path)
    assert result[POD_RESOURCE_TYPE_PICKLEBALL] == 42


def test_load_venue_input_fallback_formula(tmp_path):
    """When Available Slots is zero/missing, compute from Quantity/times/Slot Minutes."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # 4 courts, 13:00–18:00, 60-min slots → (18-13)*60/60 + 1 = 6 starts → 4*6 = 24
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=4)    # Quantity
    ws.cell(row=2, column=6, value=13)   # Start Time (decimal hour)
    ws.cell(row=2, column=7, value=18)   # Last Start Time
    ws.cell(row=2, column=8, value=60)   # Slot Minutes
    ws.cell(row=2, column=9, value=0)    # Available Slots = 0 → triggers fallback

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ScheduleWorkbookBuilder._load_venue_input(path)
    assert result[POD_RESOURCE_TYPE_TENNIS] == 24


def test_load_venue_input_ignores_blank_resource_rows(tmp_path):
    """Blank resource rows should be ignored instead of creating a literal 'nan' bucket."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_PICKLEBALL

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=2, column=9, value=24)
    ws.cell(row=3, column=3, value=None)
    ws.cell(row=3, column=9, value=None)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ScheduleWorkbookBuilder._load_venue_input(path)
    assert result == {POD_RESOURCE_TYPE_PICKLEBALL: 24}


def test_build_gym_game_objects_structure():
    """Each game object has all required OR-Tools schema fields."""
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster())
    assert games, "Expected at least one game"
    required_fields = {
        "game_id", "event", "stage", "pool_id", "round",
        "team_a_id", "team_b_id", "duration_minutes",
        "resource_type", "earliest_slot", "latest_slot",
    }
    for g in games:
        assert required_fields <= g.keys(), f"Missing fields in {g}"
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    valid_types = {GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL}
    assert all(g["resource_type"] in valid_types for g in games), \
        f"Unexpected resource_type(s): {[g['resource_type'] for g in games]}"
    # team_a_id and team_b_id must be non-null strings for all games (pool + playoff)
    assert all(
        isinstance(g["team_a_id"], str) and isinstance(g["team_b_id"], str)
        for g in games
    ), "All games must have non-null team_a_id and team_b_id"


def test_build_gym_game_objects_stages():
    """With 8 BBM teams, only Pool stage is present (playoffs go in Playoff-Slots tab)."""
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster(8))
    bbm_stages = {g["stage"] for g in games if g["event"] == SPORT_TYPE["BASKETBALL"]}
    assert bbm_stages == {"Pool"}, f"Expected only Pool stage; got {bbm_stages}"


def test_build_gym_game_objects_prefix_format():
    """Pool game IDs follow the BBM-01 format."""
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster())
    pool_ids = [g["game_id"] for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    assert pool_ids, "Expected Basketball pool games"
    import re
    assert all(re.match(r"BBM-\d{2}$", gid) for gid in pool_ids)


def test_build_gym_game_objects_stable_team_ids():
    """The same placeholder team ID is reused across multiple pool games for that team."""
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster(8))
    bbm_pool = [g for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    assert bbm_pool, "Expected Basketball pool games"

    # Collect all team IDs and count appearances
    from collections import Counter
    appearances: Counter = Counter()
    for g in bbm_pool:
        appearances[g["team_a_id"]] += 1
        appearances[g["team_b_id"]] += 1

    # The live Basketball default is now 3 games/team, so 8 teams split into
    # two 4-team round-robin pools and every placeholder appears 3 times.
    assert appearances, "Expected stable placeholder team IDs to be reused"
    assert all(count == 3 for count in appearances.values()), appearances


def test_build_gym_game_objects_pool_id_nonempty():
    """Pool games carry a non-empty pool_id; playoff/final games have empty pool_id."""
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster(8))
    pool_games = [g for g in games if g["stage"] == "Pool"]
    playoff_games = [g for g in games if g["stage"] not in ("Pool",)]

    assert pool_games, "Expected pool games"
    assert all(g["pool_id"] != "" for g in pool_games), "Pool games must have non-empty pool_id"
    assert all(g["pool_id"] == "" for g in playoff_games), "Playoff/final games must have empty pool_id"


def test_build_gym_game_objects_team_id_format():
    """Pool game team IDs follow the stable placeholder format PREFIX-Px-Ty."""
    import re as _re
    builder = ScheduleWorkbookBuilder()
    games = builder._build_gym_game_objects(_make_gym_roster(8))
    bbm_pool = [g for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    for g in bbm_pool:
        assert _re.match(r"BBM-P\d+-T\d+$", g["team_a_id"]), f"Unexpected team_a_id: {g['team_a_id']}"
        assert _re.match(r"BBM-P\d+-T\d+$", g["team_b_id"]), f"Unexpected team_b_id: {g['team_b_id']}"


def test_build_schedule_input_gym_court_scenario(tmp_path):
    """Fallback path: gym resources cover all four sessions split by basketball/volleyball."""
    from config import SCHEDULE_SOLVER_GYM_COURTS, GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], tmp_path / "missing.xlsx")
    assert si["gym_court_scenario"] == SCHEDULE_SOLVER_GYM_COURTS
    gym_resources = [r for r in si["resources"] if r["resource_type"] in (GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL)]
    n_sessions = 4  # Sat-1, Sun-1, Sat-2, Sun-2
    assert len(gym_resources) == SCHEDULE_SOLVER_GYM_COURTS * n_sessions
    # Allocation source should be "fallback" when venue_input.xlsx is absent
    assert si.get("gym_allocation", {}).get("source") == "fallback"


def test_build_schedule_input_legacy_venue_rows_do_not_add_fallback_gyms(tmp_path):
    """Legacy Venue-Input rows should be used directly instead of adding fallback gym courts."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL,
         1, "2026-07-18", 8, 10, 60, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    assert si["gym_allocation"]["source"] == "direct_venue_input"
    assert si["resource_count"] == 1
    assert {g["resource_type"] for g in si["games"]} == {GYM_RESOURCE_TYPE_BASKETBALL}
    assert [r["resource_type"] for r in si["resources"]] == [GYM_RESOURCE_TYPE_BASKETBALL]
    assert si["resources"][0]["day"] == "Sat-1"
    assert si["resources"][0]["venue_name"] == "Church Main Gym"


def test_build_schedule_input_grouped_rows_without_gym_modes_use_direct_resources(tmp_path):
    """Grouped venue rows should not be dropped when Gym-Modes is missing."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sat-1",
         "2026-07-18", 8, 12, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 12, 60, None, "Main Gym", None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    assert si["gym_allocation"]["source"] == "direct_venue_input"
    assert si["gym_allocation"]["reason"] == "grouped_rows_without_gym_modes"
    assert {r["resource_type"] for r in si["resources"]} == {
        GYM_RESOURCE_TYPE_BASKETBALL,
        GYM_RESOURCE_TYPE_VOLLEYBALL,
    }
    assert {g["resource_type"] for g in si["games"]} == {GYM_RESOURCE_TYPE_BASKETBALL}


def test_build_schedule_input_allocator_omits_zero_team_gym_sports(tmp_path):
    """Allocator-backed schedule input should omit placeholder sports with no estimating teams."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sat-1",
         "2026-07-18", 8, 12, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 12, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 1, 2],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows, gym_modes_rows=gym_modes)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    assert si["gym_allocation"]["source"] == "allocator"
    assert {g["resource_type"] for g in si["games"]} == {GYM_RESOURCE_TYPE_BASKETBALL}
    assert {r["resource_type"] for r in si["resources"]} == {GYM_RESOURCE_TYPE_BASKETBALL}


def test_build_schedule_input_uncovered_exclusive_group_included_as_direct_resource(tmp_path):
    """Venue rows whose exclusive_group has no Gym-Modes entry must not be silently
    dropped when the overall gym_resource_strategy is 'allocator'.

    Real-world trigger: EHS Tennis Court rows (Tennis Court + Pickleball Court) share
    an Exclusive Venue Group but are absent from Gym-Modes.  Without this fix they were
    excluded by the allocator AND by the direct_resources filter, leaving 0 Pickleball /
    Tennis resources and making those pools INFEASIBLE."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        # Gym courts — covered by Gym-Modes → go through allocator
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        # Tennis courts — exclusive group NOT in Gym-Modes → must become direct resources
        ["Tennis Pod", "EHS Tennis Court", "Tennis Court", 6, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "EHS Tennis Court", None, None, None],
        ["Tennis Pod", "EHS Tennis Court", "Pickleball Court", 12, "Sat-1",
         "2026-07-18", 8, 17, 30, None, "EHS Tennis Court", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 2, 2],
        # NOTE: "EHS Tennis Court" is intentionally absent from Gym-Modes
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows, gym_modes_rows=gym_modes)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    resource_types = {r["resource_type"] for r in si["resources"]}
    assert "Tennis Court" in resource_types, "Tennis Court must not be dropped"
    assert "Pickleball Court" in resource_types, "Pickleball Court must not be dropped"
    # Allocator-backed gym sports must also still be present
    assert GYM_RESOURCE_TYPE_BASKETBALL in resource_types


def test_build_schedule_input_playoff_pinned_resource_promoted_as_single_slot(tmp_path):
    """A playoff slot referencing an allocator-unvisited gym block should be promoted
    as a one-slot synthetic resource that is excluded from the Gym Core solver pool."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    from openpyxl import Workbook

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    # Sat-1 block has enough capacity to satisfy all basketball demand; Sun-2 block
    # will be left unallocated by the greedy allocator.
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 4, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        # Sun-2 block — deliberately left unallocated after Sat-1 satisfies demand.
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 4, 2],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "resource_id", "slot"],
        ["BBM-Final", "Basketball - Men Team", "Final", "GYM-Sun-2-1", "Sun-2-14:00"],
    ]

    path = tmp_path / "venue_input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    gm = wb.create_sheet("Gym-Modes")
    for r, row in enumerate(gym_modes, start=1):
        for c, val in enumerate(row, start=1):
            gm.cell(row=r, column=c, value=val)
    ps = wb.create_sheet("Playoff-Slots")
    for r, row in enumerate(playoff_rows, start=1):
        for c, val in enumerate(row, start=1):
            ps.cell(row=r, column=c, value=val)
    wb.save(path)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    # The promoted synthetic resource must appear in the resource list.
    promoted = [r for r in si["resources"] if r["resource_id"] == "GYM-Sun-2-1"]
    assert len(promoted) == 1, "GYM-Sun-2-1 should be promoted as a synthetic resource"
    p = promoted[0]

    # Single-slot: open_time == slot time.
    assert p["open_time"] == "14:00"
    # close_time = 14:00 + 60 min = 15:00.
    assert p["close_time"] == "15:00"
    # Must be flagged so capacity diagnostics skip it.
    assert p.get("playoff_pinned") is True
    # Must join the Gym Core pool: a pinned game takes its pool from its
    # pinned resource, and QF→Semi→Final precedence is only enforceable when
    # the pinned game shares a pool with its pool-play siblings.  Pool play
    # still cannot use the resource — its only slot is reserved.
    assert p.get("solver_pool") == ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL


def test_build_schedule_input_playoff_pinned_resource_merges_multiple_slots(tmp_path):
    """Reusing the same playoff-pinned GYM resource_id across multiple slots should
    expand one synthetic resource instead of leaving later slots invalid."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    from openpyxl import Workbook
    from scheduler import validate_playoff_slots

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 4, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 4, 2],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "resource_id", "slot"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi", "GYM-Sun-2-1", "Sun-2-14:00"],
        ["BBM-Final", "Basketball - Men Team", "Final", "GYM-Sun-2-1", "Sun-2-16:00"],
    ]

    path = tmp_path / "venue_input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    gm = wb.create_sheet("Gym-Modes")
    for r, row in enumerate(gym_modes, start=1):
        for c, val in enumerate(row, start=1):
            gm.cell(row=r, column=c, value=val)
    ps = wb.create_sheet("Playoff-Slots")
    for r, row in enumerate(playoff_rows, start=1):
        for c, val in enumerate(row, start=1):
            ps.cell(row=r, column=c, value=val)
    wb.save(path)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    promoted = [r for r in si["resources"] if r["resource_id"] == "GYM-Sun-2-1"]
    assert len(promoted) == 1
    resource = promoted[0]
    assert resource["open_time"] == "14:00"
    assert resource["close_time"] == "17:00"

    validated, _blocked = validate_playoff_slots(si["playoff_slots"], si["resources"])
    assert len(validated) == 2


def test_build_schedule_input_playoff_pinned_resource_rejects_implausible_ordinal(tmp_path):
    """An unknown GYM-{day}-{n} playoff resource should not be promoted when n exceeds
    the physically plausible court ordinal range for that day/block."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    from openpyxl import Workbook

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Practice Gym", "Church Practice Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sun-2",
         "2026-07-26", 8, 10, 60, None, "Practice Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 1, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Practice Gym", 0, 2],
        ["Main Gym", 2, 1],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "resource_id", "slot"],
        ["BBM-Final", "Basketball - Men Team", "Final", "GYM-Sun-2-99", "Sun-2-14:00"],
    ]

    path = tmp_path / "venue_input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    gm = wb.create_sheet("Gym-Modes")
    for r, row in enumerate(gym_modes, start=1):
        for c, val in enumerate(row, start=1):
            gm.cell(row=r, column=c, value=val)
    ps = wb.create_sheet("Playoff-Slots")
    for r, row in enumerate(playoff_rows, start=1):
        for c, val in enumerate(row, start=1):
            ps.cell(row=r, column=c, value=val)
    wb.save(path)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    assert not any(r["resource_id"] == "GYM-Sun-2-99" for r in si["resources"])


_VENUE_CENTRIC_HEADERS = [
    "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
    "Date", "Start Time", "Last Start Time", "Slot Minutes",
    "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
]


def _venue_centric_allocator_fixture(tmp_path, playoff_rows):
    """venue_input.xlsx with an allocator-managed gym (Sat-1 + Sun-2 blocks)."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 4, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 4, 2],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows,
                       gym_modes_rows=gym_modes, playoff_rows=playoff_rows)
    return path


def test_build_schedule_input_venue_centric_playoff_pin_creates_pinned_resource(tmp_path):
    """A Playoff-Slots row with gym_name + date + start_time (no resource_id)
    should resolve to a synthetic playoff-pinned single-slot resource (#127)."""
    from scheduler import validate_playoff_slots

    path = _venue_centric_allocator_fixture(tmp_path, [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Church Main Gym", "2026-07-26", "14:00"],
    ])

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    pinned = [r for r in si["resources"] if r.get("playoff_pinned")]
    assert len(pinned) == 1
    p = pinned[0]
    assert p["resource_id"] == "BB-Sun-2-PF1"
    assert p["resource_type"] == "Basketball Court"
    assert p["open_time"] == "14:00"
    assert p["close_time"] == "15:00"
    assert p["venue_name"] == "Church Main Gym"
    # Joins Gym Core so precedence rules involving the pinned game stay
    # same-pool and enforceable; all its slots are reserved from pool play.
    assert p.get("solver_pool") == ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL

    # The playoff entry must be rewritten to the resolved resource_id + slot
    # and still pass solver-side validation.
    assert len(si["playoff_slots"]) == 1
    entry = si["playoff_slots"][0]
    assert entry["resource_id"] == "BB-Sun-2-PF1"
    assert entry["slot"] == "Sun-2-14:00"
    assert entry["gym_name"] == "Church Main Gym"
    validated, _blocked = validate_playoff_slots(si["playoff_slots"], si["resources"])
    assert len(validated) == 1


def test_build_schedule_input_venue_centric_pin_reserves_allocator_window(tmp_path):
    """The pinned window must be carved out of the allocator inventory: no
    Stage-A decision may overlap it, even on a demand-claimed day (#127/#133)."""
    path = _venue_centric_allocator_fixture(tmp_path, [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Main Gym", "2026-07-26", "14:00"],
    ])

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    assert si["gym_allocation"]["source"] == "allocator"
    for decision in si["gym_allocation"]["decisions"]:
        if decision["day"] != "Sun-2":
            continue
        # Decision windows must not overlap the reserved 14:00–15:00 window.
        assert (
            decision["close_time"] <= "14:00" or decision["open_time"] >= "15:00"
        ), f"allocator claimed reserved playoff window: {decision}"
    # The same exclusion applies to allocator-emitted pool-play resources.
    for resource in si["resources"]:
        if resource.get("playoff_pinned") or resource["day"] != "Sun-2":
            continue
        assert resource["close_time"] <= "14:00" or resource["open_time"] >= "15:00"


def test_build_schedule_input_venue_centric_contiguous_pins_merge_one_track(tmp_path):
    """Contiguous venue-centric pins on the same gym/sport share one synthetic
    court, mirroring the legacy same-resource_id merge behavior."""
    from scheduler import validate_playoff_slots

    path = _venue_centric_allocator_fixture(tmp_path, [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi",
         "Main Gym", "2026-07-26", "14:00"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Main Gym", "2026-07-26", "15:00"],
    ])

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    pinned = [r for r in si["resources"] if r.get("playoff_pinned")]
    assert len(pinned) == 1, "contiguous pins should merge onto one court track"
    assert pinned[0]["open_time"] == "14:00"
    assert pinned[0]["close_time"] == "16:00"

    entries = si["playoff_slots"]
    assert [e["resource_id"] for e in entries] == ["BB-Sun-2-PF1", "BB-Sun-2-PF1"]
    assert [e["slot"] for e in entries] == ["Sun-2-14:00", "Sun-2-15:00"]
    validated, _blocked = validate_playoff_slots(entries, si["resources"])
    assert len(validated) == 2


def test_build_schedule_input_venue_centric_pin_resolves_direct_resource(tmp_path):
    """Venue-centric rows for standalone (non-allocator) venues resolve to the
    existing expanded resource IDs; concurrent pins land on distinct courts."""
    from config import SPORT_TYPE
    from scheduler import validate_playoff_slots

    rows = [
        ["TT Pod", "Orange Chapel", "Table Tennis Table", 2, "Sun-2",
         "2026-07-26", 9, 16, 20, None, None, None, None, None],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["TT-Semi-1", SPORT_TYPE["TABLE_TENNIS"], "Semi",
         "Orange Chapel", "2026-07-26", "15:00"],
        ["TT-Semi-2", SPORT_TYPE["TABLE_TENNIS"], "Semi",
         "Orange Chapel", "2026-07-26", "15:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows, playoff_rows=playoff_rows)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    # No synthetic resources for standalone venues — existing IDs are reused.
    assert not any(r.get("playoff_pinned") for r in si["resources"])
    entries = si["playoff_slots"]
    assert {e["resource_id"] for e in entries} == {"TT-Sun-2-1", "TT-Sun-2-2"}
    assert all(e["slot"] == "Sun-2-15:00" for e in entries)
    validated, _blocked = validate_playoff_slots(entries, si["resources"])
    assert len(validated) == 2


def test_build_schedule_input_venue_centric_direct_overlap_fails(tmp_path):
    """Multi-slot pins may not overlap on the same standalone resource."""
    from config import SPORT_TYPE

    rows = [
        ["TT Pod", "Orange Chapel", "Table Tennis Table", 1, "Sun-2",
         "2026-07-26", 9, 17, 20, None, None, None, None, None],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time",
         "duration_minutes"],
        ["TT-Semi", SPORT_TYPE["TABLE_TENNIS"], "Semi",
         "Orange Chapel", "2026-07-26", "14:00", 120],
        ["TT-Final", SPORT_TYPE["TABLE_TENNIS"], "Final",
         "Orange Chapel", "2026-07-26", "15:00", 120],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows, playoff_rows=playoff_rows)

    with pytest.raises(ValueError, match="already occupied during"):
        ScheduleWorkbookBuilder()._build_schedule_input(
            _make_gym_roster(), [], path
        )


def test_build_schedule_input_venue_centric_managed_capacity_fails(tmp_path):
    """Two concurrent pins cannot be placed on a one-court gym mode."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL

    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 1, 0],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi",
         "Church Main Gym", "2026-07-26", "14:00"],
        ["BBM-Semi-2", "Basketball - Men Team", "Semi",
         "Church Main Gym", "2026-07-26", "14:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(
        path, _VENUE_CENTRIC_HEADERS, rows,
        gym_modes_rows=gym_modes, playoff_rows=playoff_rows,
    )

    with pytest.raises(ValueError, match="Gym-Modes provides 1"):
        ScheduleWorkbookBuilder()._build_schedule_input(
            _make_gym_roster(), [], path
        )


def test_build_schedule_input_venue_centric_cross_mode_overlap_fails(tmp_path):
    """Mutually-exclusive gym modes cannot be pinned at the same time."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 2, 2],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Church Main Gym", "2026-07-26", "14:00"],
        ["VBM-Final", "Volleyball - Men Team", "Final",
         "Church Main Gym", "2026-07-26", "14:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(
        path, _VENUE_CENTRIC_HEADERS, rows,
        gym_modes_rows=gym_modes, playoff_rows=playoff_rows,
    )

    with pytest.raises(ValueError, match="mutually-exclusive gym"):
        ScheduleWorkbookBuilder()._build_schedule_input(
            _make_gym_roster(), [], path
        )


def test_build_schedule_input_venue_pin_uses_generated_game_duration(tmp_path):
    """A 60-minute generated game on a 30-minute grid reserves two slots."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL
    from scheduler import validate_playoff_slots

    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 12, 17, 30, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 2, 0],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Church Main Gym", "2026-07-26", "14:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(
        path, _VENUE_CENTRIC_HEADERS, rows,
        gym_modes_rows=gym_modes, playoff_rows=playoff_rows,
    )

    si = ScheduleWorkbookBuilder()._build_schedule_input(
        _make_gym_roster(), [], path
    )

    entry = si["playoff_slots"][0]
    pinned = next(r for r in si["resources"] if r.get("playoff_pinned"))
    assert entry["duration_minutes"] == 60
    assert pinned["slot_minutes"] == 30
    assert pinned["open_time"] == "14:00"
    assert pinned["close_time"] == "15:00"
    _validated, blocked = validate_playoff_slots(
        si["playoff_slots"], si["resources"]
    )
    assert blocked["Basketball Court"][pinned["resource_id"]] == {
        "Sun-2-14:00", "Sun-2-14:30",
    }


def test_build_schedule_input_venue_centric_pin_unknown_gym_fails_build(tmp_path):
    """An unknown venue fails generation instead of silently omitting the pin."""
    from loguru import logger as _loguru_logger

    path = _venue_centric_allocator_fixture(tmp_path, [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final",
         "Nonexistent Gym", "2026-07-26", "14:00"],
    ])

    messages = []
    sink_id = _loguru_logger.add(messages.append, level="ERROR")
    try:
        builder = ScheduleWorkbookBuilder()
        with pytest.raises(
            ValueError, match="Invalid venue-centric Playoff-Slots configuration"
        ):
            builder._build_schedule_input(_make_gym_roster(), [], path)
    finally:
        _loguru_logger.remove(sink_id)

    assert any(
        "no Venue-Input row found for gym 'Nonexistent Gym'" in str(m)
        for m in messages
    )


def test_build_schedule_input_venue_centric_and_explicit_forms_coexist(tmp_path):
    """Both Playoff-Slots forms work in the same file; an explicit
    resource_id + slot row passes through unchanged."""
    from scheduler import validate_playoff_slots

    path = _venue_centric_allocator_fixture(tmp_path, [
        ["game_id", "event", "stage", "resource_id", "slot",
         "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final", None, None,
         "Main Gym", "2026-07-26", "14:00"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi", "GYM-Sat-1-1", "Sat-1-16:00",
         None, None, None],
    ])

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], path)

    by_game = {e["game_id"]: e for e in si["playoff_slots"]}
    assert by_game["BBM-Final"]["resource_id"] == "BB-Sun-2-PF1"
    assert by_game["BBM-Semi-1"]["resource_id"] == "GYM-Sat-1-1"
    assert by_game["BBM-Semi-1"]["slot"] == "Sat-1-16:00"
    validated, _blocked = validate_playoff_slots(si["playoff_slots"], si["resources"])
    assert len(validated) == 2


def test_venue_centric_pins_solve_end_to_end_with_precedence(tmp_path):
    """Full #127 scenario: all BBM late rounds pinned by venue name on an
    allocator-unvisited Sun-2 block.  The build must pass the schedule
    contract, the QFs must solve onto the day before the pinned Semis
    (precedence is enforceable because the pinned games share the Gym Core
    pool), and no pool game may touch the pinned court."""
    import scheduler
    from schedule_contracts import validate_schedule_input
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL

    rows = [
        ["Main Gym", "EHS Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 4, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "EHS Main Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Practice Gym", "EHS Practice Gym", GYM_RESOURCE_TYPE_VOLLEYBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Practice Gym", None, None, None],
        ["Practice Gym", "EHS Practice Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Practice Gym", None, None, None],
        ["Main Gym", "EHS Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-2",
         "2026-07-26", 9, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 4, 2],
        ["Practice Gym", 1, 2],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi", "EHS Main Gym", "2026-07-26", "12:00"],
        ["BBM-Semi-2", "Basketball - Men Team", "Semi", "EHS Main Gym", "2026-07-26", "13:00"],
        ["BBM-3rd",    "Basketball - Men Team", "3rd",  "EHS Main Gym", "2026-07-26", "14:00"],
        ["BBM-Final",  "Basketball - Men Team", "Final", "EHS Main Gym", "2026-07-26", "15:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows,
                       gym_modes_rows=gym_modes, playoff_rows=playoff_rows)

    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(8), [], path)

    assert validate_schedule_input(si) == []
    pinned = [r for r in si["resources"] if r.get("playoff_pinned")]
    assert [(r["resource_id"], r["open_time"], r["close_time"]) for r in pinned] == [
        ("BB-Sun-2-PF1", "12:00", "16:00"),
    ], "the four contiguous pins must merge onto one playoff court track"

    out = scheduler.solve(si)
    assert out["status"] == "OPTIMAL"
    assert out["unscheduled"] == []
    by_game = {a["game_id"]: a for a in out["assignments"]}
    assert by_game["BBM-Final"]["slot"] == "Sun-2-15:00"
    # QF precedence vs the pinned Semis must be enforced: QFs stay on Sat-1.
    for n in range(1, 5):
        day, _time = by_game[f"BBM-QF-{n}"]["slot"].rsplit("-", 1)
        assert day == "Sat-1"
    # Pool play never touches the pinned playoff court.
    pinned_users = {
        a["game_id"] for a in out["assignments"]
        if a["resource_id"] == "BB-Sun-2-PF1"
    }
    assert pinned_users == {"BBM-Semi-1", "BBM-Semi-2", "BBM-3rd", "BBM-Final"}


def test_load_playoff_slots_accepts_venue_centric_columns(tmp_path):
    """The loader accepts gym_name/date/start_time rows, normalizes date and
    start_time, and skips rows with neither complete placement form."""
    rows = [
        ["X", "Some Gym", "Basketball Court", 1, "Sat-1",
         "2026-07-18", 8, 10, 60, None, None, None, None, None],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["BBM-Final", "Basketball - Men Team", "Final", "Some Gym", "2026-07-26", "14:00"],
        ["BBM-Semi-1", "Basketball - Men Team", "Semi", "Some Gym", None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows, playoff_rows=playoff_rows)

    slots = ScheduleWorkbookBuilder._load_playoff_slots(path)
    assert len(slots) == 1, "incomplete venue-centric row must be skipped"
    entry = slots[0]
    assert entry["gym_name"] == "Some Gym"
    assert entry["date"] == "2026-07-26"
    assert entry["start_time"] == "14:00"
    assert entry["resource_id"] == ""


def test_build_schedule_input_qf_games_get_latest_slot_day_before_finals(tmp_path):
    """BBM-QF-* games should receive a latest_slot pinning them to the day
    before the pinned Final.  When BBM-Final is at Sun-2-14:00 and Sat-2 has
    BB capacity with Last Start Time = 17 (→ close_time 18:00, 60-min slots),
    each QF game's latest_slot must be 'Sat-2-17:00' (the last valid start)."""
    from config import GYM_RESOURCE_TYPE_BASKETBALL
    from openpyxl import Workbook

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    # All BB rows so the allocator has only one mode to consider — the
    # spreading-pass round-robin between modes is not what's under test here.
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-1",
         "2026-07-19", 8, 17, 60, None, "Main Gym", None, None, None],
        # Sat-2 BB capacity for QFs — Last Start Time=17 → close 18:00, last start 17:00.
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sat-2",
         "2026-07-25", 8, 17, 60, None, "Main Gym", None, None, None],
        # Sun-2 marker so day_order includes the Finals day.
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sun-2",
         "2026-07-26", 13, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 2, 0],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "resource_id", "slot"],
        ["BBM-Final", "Basketball - Men Team", "Final", "GYM-Sun-2-1", "Sun-2-14:00"],
    ]

    path = tmp_path / "venue_input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    gm = wb.create_sheet("Gym-Modes")
    for r, row in enumerate(gym_modes, start=1):
        for c, val in enumerate(row, start=1):
            gm.cell(row=r, column=c, value=val)
    ps = wb.create_sheet("Playoff-Slots")
    for r, row in enumerate(playoff_rows, start=1):
        for c, val in enumerate(row, start=1):
            ps.cell(row=r, column=c, value=val)
    wb.save(path)

    builder = ScheduleWorkbookBuilder()
    # 16 churches → enough BB demand to fill Sat-1 + Sun-1 in the demand pass.
    si = builder._build_schedule_input(_make_gym_roster(n_churches=8), [], path)

    # QF and Semi games should both be constrained to the day before Finals.
    constrained = [g for g in si["games"]
                   if g["stage"] in ("QF", "Semi") and g["event"] == "Basketball - Men Team"]
    assert constrained, "Expected BBM-QF-* and BBM-Semi-* games in schedule input"
    for g in constrained:
        assert g["latest_slot"] == "Sat-2-17:00", (
            f"BBM {g['stage']} game {g['game_id']} expected latest_slot 'Sat-2-17:00', "
            f"got {g['latest_slot']!r}"
        )

    # Final and 3rd-place games must NOT carry latest_slot.
    unconstrained = [g for g in si["games"]
                     if g["stage"] in ("Final", "3rd")
                     and g["event"] == "Basketball - Men Team"]
    for g in unconstrained:
        assert g["latest_slot"] is None, (
            f"{g['game_id']} (stage={g['stage']}) should not have latest_slot set"
        )


def test_last_slot_label_on_day_helper():
    """The helper should return the last valid slot start time (close_time - slot_min)."""
    resources = [
        {"resource_type": "Basketball Court", "day": "Sat-2",
         "close_time": "17:00", "slot_minutes": 60},
        {"resource_type": "Basketball Court", "day": "Sat-2",
         "close_time": "16:30", "slot_minutes": 30},  # earlier close
        {"resource_type": "Volleyball Court", "day": "Sat-2",
         "close_time": "21:00", "slot_minutes": 60},  # different type
        {"resource_type": "Basketball Court", "day": "Sun-2",
         "close_time": "18:00", "slot_minutes": 60},  # different day
    ]
    # Last BB start on Sat-2 = max(17:00-60, 16:30-30) = max(16:00, 16:00) = 16:00
    label = ScheduleWorkbookBuilder._last_slot_label_on_day(
        resources, "Basketball Court", "Sat-2"
    )
    assert label == "Sat-2-16:00"

    # No matches → None
    assert ScheduleWorkbookBuilder._last_slot_label_on_day(
        resources, "Tennis Court", "Sat-2"
    ) is None
    assert ScheduleWorkbookBuilder._last_slot_label_on_day(
        resources, "Basketball Court", "Fri-1"
    ) is None


def test_build_schedule_input_via_church_teams_exporter_does_not_raise(tmp_path):
    """Regression: _build_schedule_input is copied onto ChurchTeamsExporter via
    setattr at the bottom of church_teams_export.py, but _last_slot_label_on_day
    is NOT in that copy list.  The QF latest_slot patching code must therefore
    invoke the helper via the class name (ScheduleWorkbookBuilder), not via
    self, otherwise running export-church-teams raises
    `AttributeError: 'ChurchTeamsExporter' object has no attribute
    '_last_slot_label_on_day'` and the schedule_input.json is never written —
    leaving the solver to consume a stale file from a previous run."""
    from church_teams_export import ChurchTeamsExporter
    from config import GYM_RESOURCE_TYPE_BASKETBALL
    from openpyxl import Workbook

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity", "Day",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Exclusive Venue Group", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sat-1",
         "2026-07-18", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sun-1",
         "2026-07-19", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 2, "Sat-2",
         "2026-07-25", 8, 17, 60, None, "Main Gym", None, None, None],
        ["Main Gym", "Church Main Gym", GYM_RESOURCE_TYPE_BASKETBALL, 1, "Sun-2",
         "2026-07-26", 13, 17, 60, None, "Main Gym", None, None, None],
    ]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts"],
        ["Main Gym", 2, 0],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "resource_id", "slot"],
        ["BBM-Final", "Basketball - Men Team", "Final", "GYM-Sun-2-1", "Sun-2-14:00"],
    ]

    path = tmp_path / "venue_input.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    gm = wb.create_sheet("Gym-Modes")
    for r, row in enumerate(gym_modes, start=1):
        for c, val in enumerate(row, start=1):
            gm.cell(row=r, column=c, value=val)
    ps = wb.create_sheet("Playoff-Slots")
    for r, row in enumerate(playoff_rows, start=1):
        for c, val in enumerate(row, start=1):
            ps.cell(row=r, column=c, value=val)
    wb.save(path)

    # The exporter does NOT inherit ScheduleWorkbookBuilder — its methods are
    # injected via setattr.  This call would raise AttributeError before the fix.
    exporter = ChurchTeamsExporter()
    si = exporter._build_schedule_input(_make_gym_roster(n_churches=8), [], path)

    # Both QF and Semi games should be constrained to the day before Finals.
    constrained = [g for g in si["games"]
                   if g["stage"] in ("QF", "Semi") and g["event"] == "Basketball - Men Team"]
    assert constrained, "Expected BBM-QF-* and BBM-Semi-* games to be generated"
    for g in constrained:
        assert g["latest_slot"] == "Sat-2-17:00", (
            f"BBM {g['stage']} game {g['game_id']} expected latest_slot 'Sat-2-17:00', "
            f"got {g['latest_slot']!r} — the latest_slot patch must work when "
            f"_build_schedule_input is invoked through a ChurchTeamsExporter instance"
        )


def test_build_pod_game_objects_single_elimination():
    """With 3 entries, 2 game placeholders with stage-aware IDs are generated.

    N=3, P=4: round 1 = Semi (1 game), round 2 = Final (1 game).
    """
    from config import POD_RESOURCE_EVENT_TYPE

    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Single"}
        for i in range(1, 4)  # 3 entries
    ]
    builder = ScheduleWorkbookBuilder()
    games, prec = builder._build_pod_game_objects(roster_rows, [])
    assert len(games) == 2, f"Expected 2 games (3-1=2), got {len(games)}"
    assert all(g["game_id"].startswith("BAD-Women-Singles-") for g in games)
    assert all(g["resource_type"] == POD_RESOURCE_EVENT_TYPE[SPORT_TYPE["BADMINTON"]] for g in games)
    game_ids = [g["game_id"] for g in games]
    assert game_ids == ["BAD-Women-Singles-Semi-1", "BAD-Women-Singles-Final"]
    stages = [g["stage"] for g in games]
    assert stages == ["Semi", "Final"]
    # Precedence: Semi must complete before Final
    assert len(prec) == 1
    assert prec[0]["before_game_id"] == "BAD-Women-Singles-Semi-1"
    assert prec[0]["after_game_id"] == "BAD-Women-Singles-Final"


def test_build_pod_game_objects_skips_empty_divisions():
    """Divisions with fewer than 2 entries produce no game objects."""
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": 1,
         "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Single"},
    ]
    builder = ScheduleWorkbookBuilder()
    games, prec = builder._build_pod_game_objects(roster_rows, [])
    assert games == [], "Single-entry division should produce no games"
    assert prec == []


def test_build_gym_resource_objects_count():
    """4 sessions × (n_basketball + n_volleyball) resources are returned."""
    resources = ScheduleWorkbookBuilder._build_gym_resource_objects(n_basketball=2, n_volleyball=2)
    assert len(resources) == 16, f"Expected 4 sessions × 4 courts = 16, got {len(resources)}"
    from config import GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL
    valid_types = {GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL}
    assert all(r["resource_type"] in valid_types for r in resources)
    days = {r["day"] for r in resources}
    assert days == {"Sat-1", "Sun-1", "Sat-2", "Sun-2"}
    bb = [r for r in resources if r["resource_type"] == GYM_RESOURCE_TYPE_BASKETBALL]
    vb = [r for r in resources if r["resource_type"] == GYM_RESOURCE_TYPE_VOLLEYBALL]
    assert len(bb) == 8, f"Expected 4 sessions × 2 basketball courts = 8, got {len(bb)}"
    assert len(vb) == 8, f"Expected 4 sessions × 2 volleyball courts = 8, got {len(vb)}"


def test_build_gym_resource_objects_labels():
    """Court labels and resource_ids are formatted correctly."""
    resources = ScheduleWorkbookBuilder._build_gym_resource_objects(n_basketball=2, n_volleyball=1)
    labels = {r["label"] for r in resources}
    assert labels == {"Court-1", "Court-2", "Court-3"}
    ids = {r["resource_id"] for r in resources}
    assert "GYM-Sat-1-1" in ids
    assert "GYM-Sun-2-3" in ids


def test_build_gym_resource_objects_include_blank_exclusive_group():
    """Gym resources carry the same exclusive_group field as venue-loaded resources."""
    resources = ScheduleWorkbookBuilder._build_gym_resource_objects(n_basketball=1, n_volleyball=1)
    assert resources
    assert all("exclusive_group" in r for r in resources)
    assert all(r["exclusive_group"] == "" for r in resources)


def test_load_venue_input_rows_missing_file(tmp_path):
    """Returns empty list (and empty day_order) when venue_input.xlsx does not exist."""
    result, day_order = ScheduleWorkbookBuilder._load_venue_input_rows(tmp_path / "missing.xlsx")
    assert result == []
    assert day_order == []


def test_load_venue_input_rows_expands_quantity(tmp_path):
    """Quantity=2 for a Tennis Court row yields 2 resource objects."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=2)    # Quantity = 2
    ws.cell(row=2, column=6, value=9)    # Start Time
    ws.cell(row=2, column=7, value=17)   # Last Start Time
    ws.cell(row=2, column=8, value=30)   # Slot Minutes

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result, day_order = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert len(result) == 2
    assert result[0]["resource_type"] == POD_RESOURCE_TYPE_TENNIS
    assert result[0]["label"] == "Court-1"
    assert result[1]["label"] == "Court-2"
    assert result[0]["open_time"] == "09:00"
    assert result[0]["day"] == "Day-1"


def test_load_venue_input_rows_table_label(tmp_path):
    """Table Tennis Table rows get Table-N labels instead of Court-N."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TABLE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TABLE_TENNIS)
    ws.cell(row=2, column=4, value=3)
    ws.cell(row=2, column=6, value=9)
    ws.cell(row=2, column=7, value=17)
    ws.cell(row=2, column=8, value=20)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result, _ = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert len(result) == 3
    assert all(r["label"].startswith("Table-") for r in result)


def test_load_venue_input_rows_skips_blank_resource_rows(tmp_path):
    """Blank/NaN venue rows should be ignored instead of crashing Schedule-Input generation."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=2)
    ws.cell(row=2, column=6, value=9)
    ws.cell(row=2, column=7, value=17)
    ws.cell(row=2, column=8, value=30)
    ws.cell(row=3, column=3, value=None)
    ws.cell(row=3, column=4, value=None)
    ws.cell(row=3, column=6, value=None)
    ws.cell(row=3, column=7, value=None)
    ws.cell(row=3, column=8, value=None)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result, _ = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert len(result) == 2
    assert all(r["resource_type"] == POD_RESOURCE_TYPE_TENNIS for r in result)


def test_load_venue_input_rows_reads_exclusive_group(tmp_path):
    """Exclusive Venue Group column is attached to each emitted resource object."""
    from config import POD_RESOURCE_TYPE_TENNIS

    headers = [
        "Pod Name", "Venue Name", "Exclusive Venue Group", "Resource Type",
        "Quantity", "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["BB Pod", "Midsize Gym", "Midsize Gym", POD_RESOURCE_TYPE_TENNIS,
         2, None, 9, 17, 30, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result, _ = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert len(result) == 2
    assert all(r["exclusive_group"] == "Midsize Gym" for r in result)
    assert all(r["venue_name"] == "Midsize Gym" for r in result)


def test_load_venue_input_rows_blank_exclusive_group(tmp_path):
    """A row with no Exclusive Venue Group yields an empty-string group."""
    from config import POD_RESOURCE_TYPE_TENNIS

    headers = [
        "Pod Name", "Venue Name", "Exclusive Venue Group", "Resource Type",
        "Quantity", "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Tennis Pod", "Chapman", None, POD_RESOURCE_TYPE_TENNIS,
         1, None, 9, 17, 60, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result, _ = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert result[0]["exclusive_group"] == ""
    assert result[0]["venue_name"] == "Chapman"


def test_load_venue_input_rows_derives_day_labels_from_date_column(tmp_path):
    """Date-only venue rows should map to logical weekday labels like Fri-*/Sat-*."""
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Basketball Pod", "HS Small Gym", "Basketball Court", 1,
         "2026-07-17", 17, 21, 60, None, None, None, None],
        ["Basketball Pod", "HS Small Gym", "Basketball Court", 1,
         "2026-07-18", 8, 16, 60, None, None, None, None],
        ["Basketball Pod", "HS Small Gym", "Basketball Court", 1,
         "2026-07-19", 12, 20, 60, None, None, None, None],
        ["Basketball Pod", "HS Small Gym", "Basketball Court", 1,
         "2026-07-25", 8, 17, 60, None, None, None, None],
        ["Basketball Pod", "HS Small Gym", "Basketball Court", 1,
         "2026-07-26", 12, 17, 60, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result, day_order = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert [r["day"] for r in result] == ["Fri-1", "Sat-1", "Sun-1", "Sat-2", "Sun-2"]
    # day_order reflects actual calendar date order: Fri-1 is the first unique date
    # (2026-07-17), then Sat-1 (07-18), Sun-1 (07-19), Sat-2 (07-25), Sun-2 (07-26).
    assert day_order == ["Fri-1", "Sat-1", "Sun-1", "Sat-2", "Sun-2"]


def test_load_venue_input_rows_normalizes_resource_types_and_prefixes(tmp_path):
    """Human-friendly resource names should normalize to canonical scheduler IDs."""
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Bible Challenge", "Library", "Jeopardy stage", 1,
         "2026-07-18", 11, 20, 60, None, None, None, None],
        ["Soccer", "Field", "Soccer field", 1,
         "2026-07-18", 11, 17, 60, None, None, None, None],
        ["Table Tennis", "Chapel", "Table Tennis station", 1,
         "2026-07-18", 17, 21, 20, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result, _ = ScheduleWorkbookBuilder._load_venue_input_rows(path)
    assert [r["resource_type"] for r in result] == [
        "BC Station",
        "Soccer Field",
        "Table Tennis Table",
    ]
    assert [r["resource_id"] for r in result] == [
        "BC-Sat-1-1",
        "SOC-Sat-1-1",
        "TT-Sat-1-1",
    ]


def test_load_gym_modes_missing_file(tmp_path):
    """Returns empty dict when venue_input.xlsx does not exist."""
    assert ScheduleWorkbookBuilder._load_gym_modes(tmp_path / "missing.xlsx") == {}


def test_load_gym_modes_missing_tab(tmp_path):
    """File present but no Gym-Modes tab → empty dict (warning, no crash)."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]])
    assert ScheduleWorkbookBuilder._load_gym_modes(path) == {}


def test_load_gym_modes_reads_capacities(tmp_path):
    """Gym-Modes tab is parsed into {gym: {resource_type: courts_per_block}}."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts",
         "Badminton Courts", "Pickleball Courts", "Soccer Fields", "Notes"],
        ["Midsize Gym", 1, 2, 6, 8, 1, "either-or"],
        ["Big Gym", 2, 3, 12, 0, 0, "larger"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ScheduleWorkbookBuilder._load_gym_modes(path)
    assert result["Midsize Gym"] == {
        "Basketball Court": 1, "Volleyball Court": 2,
        "Badminton Court": 6, "Pickleball Court": 8, "Soccer Field": 1,
    }
    assert result["Big Gym"]["Volleyball Court"] == 3
    assert result["Big Gym"]["Pickleball Court"] == 0


def test_load_gym_modes_trims_header_whitespace(tmp_path):
    """Operator-edited headers with trailing spaces are normalized before row access."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name ", "Basketball Courts ", "Volleyball Courts ", "Notes "],
        ["Midsize Gym", 1, 2, "either-or"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ScheduleWorkbookBuilder._load_gym_modes(path)
    assert result["Midsize Gym"]["Basketball Court"] == 1
    assert result["Midsize Gym"]["Volleyball Court"] == 2


def test_load_gym_modes_skips_note_row(tmp_path):
    """A footer row with text in Gym Name but no capacities is ignored."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts",
         "Badminton Courts", "Pickleball Courts", "Soccer Fields", "Notes"],
        ["Midsize Gym", 1, 2, 6, 8, 1, "either-or"],
        ["Capacity-per-mode coefficients for the LP estimator.",
         None, None, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ScheduleWorkbookBuilder._load_gym_modes(path)
    assert list(result.keys()) == ["Midsize Gym"]


def test_build_schedule_input_keys(tmp_path):
    """_build_schedule_input returns dict with all required top-level keys."""
    builder = ScheduleWorkbookBuilder()
    si = builder._build_schedule_input(_make_gym_roster(), [], tmp_path / "missing.xlsx")
    assert set(si.keys()) == {
        "generated_at", "gym_court_scenario", "game_count", "resource_count",
        "games", "resources", "playoff_slots", "gym_modes", "gym_allocation",
        "team_conflicts", "pod_unprotected_entries",
        "pod_validation_reconciliation", "precedence", "day_order",
    }
    assert si["game_count"] == len(si["games"])
    assert si["resource_count"] == len(si["resources"])
    assert si["game_count"] > 0
    assert si["resource_count"] > 0  # at least gym resources


def test_build_schedule_input_uses_pool_assignments_for_core_gym_games(tmp_path):
    """Layer 2 schedule input should replace gym placeholders with assigned real teams."""
    builder = ScheduleWorkbookBuilder()
    roster_rows = []
    for participant_id in range(1, 11):
        roster_rows.append({
            "Church Team": "OCB",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": participant_id,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        })
        roster_rows.append({
            "Church Team": "RPC",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": participant_id + 100,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        })
        roster_rows.append({
            "Church Team": "OCB",
            "sport_type": SPORT_TYPE["VOLLEYBALL_MEN"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": participant_id,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        })
        roster_rows.append({
            "Church Team": "ANH",
            "sport_type": SPORT_TYPE["VOLLEYBALL_MEN"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": participant_id + 200,
            "participant_primary_sport": SPORT_TYPE["VOLLEYBALL_MEN"],
        })

    si = builder._build_schedule_input(roster_rows, [], tmp_path / "missing.xlsx")

    bb_game = next(game for game in si["games"] if game["event"] == SPORT_TYPE["BASKETBALL"])
    assert bb_game["team_a_id"].startswith("BBM::") or bb_game["team_b_id"].startswith("BBM::")
    assert bb_game["solver_pool"] == "Gym Core"
    assert any(
        {"BBM::", "VBM::"} == {
            edge["team_a_id"].split("::")[0] + "::",
            edge["team_b_id"].split("::")[0] + "::",
        }
        for edge in si["team_conflicts"]
    )


def test_build_assigned_gym_game_objects_fallback_counts_team_order_units():
    roster_rows = []
    for church_code in ("RPC", "OCB"):
        for team_order in ("A", "B"):
            for participant_id in range(5):
                roster_rows.append(
                    {
                        "Church Team": church_code,
                        "team_order": team_order,
                        "sport_type": SPORT_TYPE["BASKETBALL"],
                        "sport_gender": "Men",
                        "sport_format": "Team",
                        "Participant ID (WP)": f"{church_code}-{team_order}-{participant_id}",
                    }
                )

    games, _precedence = ScheduleWorkbookBuilder._build_assigned_gym_game_objects(
        roster_rows,
        [],
        allow_placeholder_fallback=False,
    )

    basketball_pool_games = [
        game for game in games
        if game["event"] == SPORT_TYPE["BASKETBALL"] and game.get("stage") == "Pool"
    ]
    assert len(basketball_pool_games) == 6


def test_build_assigned_gym_game_objects_auto_generates_playoffs_for_8_teams():
    """8-team fallback should auto-generate QF/Semi/Final/3rd plus precedence."""
    roster_rows = []
    for church_code in ("RPC", "OCB", "ANH", "GLA", "TLC", "FVC", "PHX", "WAG"):
        for participant_id in range(5):
            roster_rows.append({
                "Church Team": church_code,
                "team_order": "A",
                "sport_type": SPORT_TYPE["BASKETBALL"],
                "sport_gender": "Men",
                "sport_format": "Team",
                "Participant ID (WP)": f"{church_code}-A-{participant_id}",
            })

    games, precedence = ScheduleWorkbookBuilder._build_assigned_gym_game_objects(
        roster_rows,
        [],
        allow_placeholder_fallback=False,
    )

    bbm_games = [g for g in games if g["event"] == SPORT_TYPE["BASKETBALL"]]
    stages = {g["game_id"]: g.get("stage") for g in bbm_games}

    qf_ids = {gid for gid, st in stages.items() if st == "QF"}
    semi_ids = {gid for gid, st in stages.items() if st == "Semi"}
    final_ids = {gid for gid, st in stages.items() if st == "Final"}
    third_ids = {gid for gid, st in stages.items() if st == "3rd"}

    assert qf_ids == {"BBM-QF-1", "BBM-QF-2", "BBM-QF-3", "BBM-QF-4"}
    assert semi_ids == {"BBM-Semi-1", "BBM-Semi-2"}
    assert final_ids == {"BBM-Final"}
    assert third_ids == {"BBM-3rd"}

    # Final pulls winners of semis
    final_game = next(g for g in bbm_games if g["game_id"] == "BBM-Final")
    assert final_game["team_a_id"] == "WIN-BBM-Semi-1"
    assert final_game["team_b_id"] == "WIN-BBM-Semi-2"

    # Precedence covers Pool→QF, QF→Semi, Semi→Final, Semi→3rd
    bbm_prec = {
        (str(r["before_game_id"]), str(r["after_game_id"]))
        for r in precedence
        if str(r.get("before_game_id") or "").startswith("BBM-")
    }
    pool_ids = {gid for gid, st in stages.items() if st == "Pool"}
    assert all((p, q) in bbm_prec for p in pool_ids for q in qf_ids)
    assert ("BBM-QF-1", "BBM-Semi-1") in bbm_prec
    assert ("BBM-QF-2", "BBM-Semi-1") in bbm_prec
    assert ("BBM-QF-3", "BBM-Semi-2") in bbm_prec
    assert ("BBM-QF-4", "BBM-Semi-2") in bbm_prec
    assert ("BBM-QF-1", "BBM-Semi-2") not in bbm_prec
    assert ("BBM-QF-2", "BBM-Semi-2") not in bbm_prec
    assert ("BBM-QF-3", "BBM-Semi-1") not in bbm_prec
    assert ("BBM-QF-4", "BBM-Semi-1") not in bbm_prec
    assert all((s, "BBM-Final") in bbm_prec for s in semi_ids)
    assert all((s, "BBM-3rd") in bbm_prec for s in semi_ids)

    # QF→Semi, Semi→Final, Semi→3rd all require a 1-hour rest buffer (min_gap_slots=2)
    qf_semi_rules = [
        r for r in precedence
        if str(r.get("before_game_id") or "").startswith("BBM-QF-")
        and str(r.get("after_game_id") or "").startswith("BBM-Semi-")
    ]
    semi_final_rules = [
        r for r in precedence
        if str(r.get("before_game_id") or "").startswith("BBM-Semi-")
        and str(r.get("after_game_id") or "") in ("BBM-Final", "BBM-3rd")
    ]
    assert all(r["min_gap_slots"] == 2 for r in qf_semi_rules), "QF→Semi must require 2-slot gap"
    assert all(r["min_gap_slots"] == 2 for r in semi_final_rules), "Semi→Final/3rd must require 2-slot gap"


def test_build_assigned_gym_game_objects_4_team_bracket_skips_qf():
    """4-team bracket should generate Semi/Final/3rd directly, no QFs."""
    roster_rows = []
    for church_code in ("RPC", "OCB", "ANH", "GLA"):
        for participant_id in range(12):
            roster_rows.append({
                "Church Team": church_code,
                "team_order": "A",
                "sport_type": SPORT_TYPE["VOLLEYBALL_MEN"],
                "sport_gender": "Men",
                "sport_format": "Team",
                "Participant ID (WP)": f"{church_code}-A-{participant_id}",
            })

    games, _precedence = ScheduleWorkbookBuilder._build_assigned_gym_game_objects(
        roster_rows,
        [],
        allow_placeholder_fallback=False,
    )

    vbm_stages = {g["game_id"]: g.get("stage") for g in games if g["event"] == SPORT_TYPE["VOLLEYBALL_MEN"]}
    assert not any(st == "QF" for st in vbm_stages.values())
    assert {gid for gid, st in vbm_stages.items() if st == "Semi"} == {"VBM-Semi-1", "VBM-Semi-2"}
    assert {gid for gid, st in vbm_stages.items() if st == "Final"} == {"VBM-Final"}


def test_classmethod_schedule_helpers_do_not_instantiate_exporter_subclass():
    class CountingBuilder(ScheduleWorkbookBuilder):
        init_calls = 0

        def __init__(self):
            type(self).init_calls += 1
            super().__init__()

    CountingBuilder._build_core_gym_team_lookup([])
    assert CountingBuilder.init_calls == 0


def test_build_schedule_output_flat_rows_count():
    """Returns one row per assignment."""
    so, si = _make_render_schedule_pair()
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    assert len(rows) == 2


def test_build_schedule_output_flat_rows_fields():
    """Each row contains expected keys with non-empty event."""
    so, si = _make_render_schedule_pair()
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    required = {"game_id", "event", "stage", "round", "team_a_id", "team_b_id",
                "resource_label", "day", "slot", "duration_minutes"}
    for row in rows:
        assert required.issubset(row.keys())
        assert row["event"] == "Basketball - Men Team"


def test_build_schedule_output_flat_rows_sorted():
    """Rows are sorted Pool before Final (stage order)."""
    so, si = _make_render_schedule_pair()
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    stages = [r["stage"] for r in rows]
    assert stages == ["Pool", "Final"]


def test_build_schedule_output_flat_rows_time_part():
    """The slot field extracts the HH:MM part from the full slot label."""
    so, si = _make_render_schedule_pair()
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    pool_row = next(r for r in rows if r["game_id"] == "BBM-01")
    assert pool_row["slot"] == "08:00"


def test_build_schedule_output_flat_rows_day_display():
    """Sat-1 is translated to '1st Sat'."""
    so, si = _make_render_schedule_pair()
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    assert all(r["day"] == "1st Sat" for r in rows)


def test_build_schedule_output_flat_rows_empty():
    """Empty assignments list returns empty rows."""
    so = {"assignments": [], "unscheduled": []}
    si = {"games": [], "resources": [], "precedence": []}
    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(so, si)
    assert rows == []


def test_compute_court_slots_matches_normalized_policy_8teams_gpg2():
    """8 teams / gpg=2 now stays at 8 pool games under the normalized policy."""

    n_teams, gpg = 8, 2
    actual = len(ScheduleWorkbookBuilder._make_pool_game_pairs("_", n_teams, gpg))
    assert actual == 8  # pools [4,4] with 4-match matrices -> 4 + 4

    builder = ScheduleWorkbookBuilder()
    s_formula = builder._compute_court_slots(n_teams, pool_games_per_team=gpg)
    s_actual  = builder._compute_court_slots(n_teams, pool_games_per_team=gpg,
                                              actual_pool_games=actual)
    assert s_formula["pool_slots"] == 8
    assert s_actual["pool_slots"]  == 8


def test_make_pool_game_pairs_exact_two_games_per_team():
    """Normalized B4 policy keeps every team at the same exact pool-game count."""
    from collections import Counter

    cases = [
        (3, 2, 2, 3),
        (4, 2, 2, 4),
        (5, 2, 2, 5),
        (7, 2, 2, 7),
        (8, 2, 2, 8),
        (12, 2, 2, 12),
        (20, 2, 2, 20),
    ]
    for n_teams, gpg, expected_games_per_team, expected_total_games in cases:
        pairs = ScheduleWorkbookBuilder._make_pool_game_pairs("T", n_teams, gpg)
        games_per_team: Counter = Counter()
        for a, b, _ in pairs:
            games_per_team[a] += 1
            games_per_team[b] += 1
        assert len(pairs) == expected_total_games, (
            f"n_teams={n_teams}: total pool games {len(pairs)} != {expected_total_games}"
        )
        assert games_per_team, f"Expected generated games for n_teams={n_teams}"
        assert all(count == expected_games_per_team for count in games_per_team.values()), (
            f"n_teams={n_teams}, gpg={gpg}: {dict(games_per_team)}"
        )


def test_make_pool_game_pairs_direct_match_for_two_teams():
    """Two teams stay as a single direct match rather than an over-built pool."""

    pairs = ScheduleWorkbookBuilder._make_pool_game_pairs("T", 2, 2)
    assert pairs == [("T-P1-T1", "T-P1-T2", "P1")]


def test_make_pool_game_pairs_exact_three_games_per_team_even_counts():
    """3-game mode keeps even-count pools at exactly 3 games/team."""
    from collections import Counter

    cases = [
        (4, 6),
        (6, 9),
        (8, 12),
        (10, 15),
        (12, 18),
    ]
    for n_teams, expected_total_games in cases:
        pairs = ScheduleWorkbookBuilder._make_pool_game_pairs("T", n_teams, 3)
        games_per_team: Counter = Counter()
        for a, b, _ in pairs:
            games_per_team[a] += 1
            games_per_team[b] += 1
        assert len(pairs) == expected_total_games
        assert games_per_team
        assert all(count == 3 for count in games_per_team.values()), dict(games_per_team)


def test_make_pool_game_pairs_three_game_policy_odd_counts_use_one_extra_slot():
    """Odd-size 3-game pools keep everyone at 3 except the highest odd-pool slot."""
    from collections import Counter

    cases = [
        (5, [3, 3, 3, 3, 4]),
        (7, [3, 3, 3, 3, 3, 3, 4]),
        (9, [3, 3, 3, 3, 3, 3, 3, 3, 4]),
        (13, [3] * 12 + [4]),
    ]
    for n_teams, expected_counts in cases:
        pairs = ScheduleWorkbookBuilder._make_pool_game_pairs("T", n_teams, 3)
        games_per_team: Counter = Counter()
        for a, b, _ in pairs:
            games_per_team[a] += 1
            games_per_team[b] += 1
        assert sorted(games_per_team.values()) == expected_counts, (
            f"n_teams={n_teams}: {dict(games_per_team)}"
        )


def test_make_pool_game_pairs_reject_unsupported_target():
    """Only 2 and 3 games/team are supported for live team-sport planning."""
    import pytest

    with pytest.raises(ValueError, match="Only 2 and 3 games/team are supported"):
        ScheduleWorkbookBuilder._make_pool_game_pairs("T", 8, 4)


def test_summarize_pool_policy_three_game_note_for_odd_pool():
    """3-game mode should explain which odd-pool slot gets the extra game."""
    summary = ScheduleWorkbookBuilder._summarize_pool_policy(13, 3)

    assert summary["target_pool_games_per_team"] == 3
    assert summary["actual_pool_games_per_team"] == 3.08
    assert "5-team pools give T5 the extra 4th game" in summary["pool_composition"]
    assert summary["actual_pool_games"] == 20


def test_pool_sizes_for_assignment_follow_three_game_policy(monkeypatch):
    """Pool-Assignment should switch to 4/4/5 sizing when Basketball uses 3 games/team."""
    builder = ScheduleWorkbookBuilder()
    codes = [
        "RPC", "ANH", "FVC", "GAC", "NSD", "TLC", "GLA",
        "ORN", "WSD", "PCC", "PHX", "WAG", "NHC",
    ]
    monkeypatch.setitem(COURT_ESTIMATE_POOL_GAMES_PER_TEAM, SPORT_TYPE["BASKETBALL"], 3)

    assert builder._pool_sizes_for_assignment(
        SPORT_TYPE["BASKETBALL"],
        len(codes),
    ) == [4, 4, 5]


def test_build_gym_game_objects_follow_three_game_policy(monkeypatch):
    """Layer-2 bridge game generation should honor the explicit 3-game policy."""
    from collections import Counter

    builder = ScheduleWorkbookBuilder()
    codes = [
        "RPC", "ANH", "FVC", "GAC", "NSD", "TLC", "GLA",
        "ORN", "WSD", "PCC", "PHX", "WAG", "NHC",
    ]
    monkeypatch.setitem(COURT_ESTIMATE_POOL_GAMES_PER_TEAM, SPORT_TYPE["BASKETBALL"], 3)
    games = builder._build_gym_game_objects(_make_gym_roster_from_codes(codes))
    bbm_pool = [g for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]

    appearances: Counter = Counter()
    for g in bbm_pool:
        appearances[g["team_a_id"]] += 1
        appearances[g["team_b_id"]] += 1

    assert len(bbm_pool) == 20
    assert sorted(appearances.values()) == ([3] * 12) + [4]


def test_compute_court_slots_consistent_with_make_pool_game_pairs():
    """pool_slots from _compute_court_slots equals len(_make_pool_game_pairs) for
    several (n_teams, gpg) combinations."""

    builder = ScheduleWorkbookBuilder()
    cases = [(2, 2), (5, 2), (8, 2), (12, 2), (9, 3), (13, 3)]
    for n_teams, gpg in cases:
        actual = len(ScheduleWorkbookBuilder._make_pool_game_pairs("_", n_teams, gpg))
        s = builder._compute_court_slots(n_teams, pool_games_per_team=gpg,
                                          actual_pool_games=actual)
        assert s["pool_slots"] == actual, (
            f"n_teams={n_teams}, gpg={gpg}: pool_slots={s['pool_slots']} != actual={actual}"
        )


def test_warn_if_schedules_mismatched_clean():
    """Returns True and no warning when all assignment IDs are in schedule_input."""

    so = {"assignments": [{"game_id": "G1", "resource_id": "R1", "slot": "Sat-1-08:00"}]}
    si = {"games": [{"game_id": "G1"}], "playoff_slots": []}
    assert ScheduleWorkbookBuilder._warn_if_schedules_mismatched(so, si) is True


def test_warn_if_schedules_mismatched_playoff_ok():
    """Returns True when assignment ID matches a playoff_slot, not a game."""

    so = {"assignments": [{"game_id": "BBM-Final", "resource_id": "R1", "slot": "Sat-2-14:00"}]}
    si = {"games": [], "playoff_slots": [{"game_id": "BBM-Final"}]}
    assert ScheduleWorkbookBuilder._warn_if_schedules_mismatched(so, si) is True


def test_warn_if_schedules_mismatched_detects_orphan():
    """Returns False and logs a warning when an assignment game_id is unknown."""
    from loguru import logger

    messages = []
    sink_id = logger.add(lambda msg: messages.append(msg), level="WARNING")
    try:
        so = {"assignments": [{"game_id": "STALE-99", "resource_id": "R1", "slot": "Sat-1-08:00"}]}
        si = {"games": [{"game_id": "G1"}], "playoff_slots": []}
        result = ScheduleWorkbookBuilder._warn_if_schedules_mismatched(so, si)
    finally:
        logger.remove(sink_id)
    assert result is False
    assert any("STALE-99" in m for m in messages)
    assert any("different runs" in m for m in messages)


def test_warn_if_resource_slot_minutes_differ_from_config_is_advisory():
    """Mismatched venue slot sizes should warn without blocking Layer 2 generation."""
    from loguru import logger

    messages = []
    sink_id = logger.add(lambda msg: messages.append(msg), level="WARNING")
    try:
        ScheduleWorkbookBuilder._warn_if_resource_slot_minutes_differ_from_config(
            all_games=[
                {
                    "game_id": "PCK-01",
                    "event": SPORT_TYPE["PICKLEBALL"],
                    "resource_type": POD_RESOURCE_TYPE_PICKLEBALL,
                    "duration_minutes": 30,
                }
            ],
            all_resources=[
                {
                    "resource_id": "PCK-Sat-1-1",
                    "resource_type": POD_RESOURCE_TYPE_PICKLEBALL,
                    "slot_minutes": 20,
                }
            ],
        )
    finally:
        logger.remove(sink_id)

    assert any("Layer 2 duration mismatch" in m for m in messages)
    assert any("Pickleball Court" in m for m in messages)
    assert any("config.py game duration is 30m" in m for m in messages)
    assert any("ignore this warning" in m for m in messages)


def test_write_schedule_output_report_creates_file(tmp_path):
    """_write_schedule_output_report writes an xlsx with all expected tabs."""
    import openpyxl
    so, si = _make_render_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "Schedule-by-Time" in wb.sheetnames
    assert "Schedule-by-Sport" in wb.sheetnames
    assert "Conflict-Audit" in wb.sheetnames


def test_write_schedule_output_report_tab1_has_data(tmp_path):
    """Schedule-by-Time tab has a title in row 1 and game text in the grid."""
    import openpyxl
    so, si = _make_render_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]
    title = ws.cell(row=1, column=1).value
    assert title and "Sports Fest" in title
    # At least one game should appear somewhere in the grid
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    assert any("BBM" in str(v) for v in all_values if v)


def test_write_schedule_output_report_tab2_flat_list(tmp_path):
    """Schedule-by-Sport tab has a header row and one data row per assignment."""
    import openpyxl
    so, si = _make_render_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Sport"]
    assert ws.cell(row=1, column=1).value == "game_id"
    assert ws.cell(row=2, column=1).value == "BBM-01"    # Pool comes first
    assert ws.cell(row=3, column=1).value == "BBM-Final"


def test_write_schedule_output_report_renders_three_team_bc_cell(tmp_path):
    """Schedule-by-Time should render BC Jeopardy games with all three team labels."""
    import openpyxl

    schedule_input = {
        "games": [
            {
                "game_id": "BC-P1-RR-1",
                "event": SPORT_TYPE["BIBLE_CHALLENGE"],
                "stage": "Pool",
                "pool_id": "P1",
                "round": 1,
                "team_a_id": "BC::RPC",
                "team_b_id": "BC::ANH",
                "team_c_id": "BC::OCB",
                "team_a_label": "RPC",
                "team_b_label": "ANH",
                "team_c_label": "OCB",
                "duration_minutes": 60,
                "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
                "earliest_slot": None,
                "latest_slot": None,
            }
        ],
        "resources": [
            {
                "resource_id": "BC-Sat-1-1",
                "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
                "label": "BC-1",
                "day": "Sat-1",
                "open_time": "08:00",
                "close_time": "10:00",
                "slot_minutes": 60,
                "exclusive_group": "",
            }
        ],
        "playoff_slots": [],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-01T10:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {"game_id": "BC-P1-RR-1", "resource_id": "BC-Sat-1-1", "slot": "Sat-1-08:00"}
        ],
        "unscheduled": [],
        "conflict_audit_summary": {},
        "conflict_audit": [],
    }

    out = tmp_path / "bc_sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    assert any("RPC / ANH / OCB" in str(value) for value in all_values if value)


def test_write_schedule_output_report_tab3_conflict_audit(tmp_path):
    """Conflict-Audit tab should render audit rows and summary content."""
    import openpyxl
    so, si = _make_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Conflict-Audit"]
    assert ws.cell(row=1, column=1).value == "Cross-Sport Conflict Audit"
    assert "Separated" in str(ws.cell(row=3, column=2).value)
    assert ws.cell(row=6, column=1).value == "team_a_label"
    assert ws.cell(row=7, column=1).value == "OCB"
    assert ws.cell(row=7, column=8).value == "SeparatedInSchedule"


def test_write_schedule_output_report_tab3_planning_only_conflict_audit(tmp_path):
    """Conflict-Audit should render planning-only rows distinctly."""
    import openpyxl

    so, si = _make_schedule_pair()
    so["conflict_audit_summary"]["planning_only_edges"] = 1
    so["conflict_audit_summary"]["separated_edges"] = 0
    so["conflict_audit"][0]["event_b"] = SPORT_TYPE["BIBLE_CHALLENGE"]
    so["conflict_audit"][0]["status"] = "PlanningOnly"
    so["conflict_audit"][0]["scheduled_team_b_games"] = 0

    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Conflict-Audit"]
    assert "Planning-only" in str(ws.cell(row=3, column=2).value)
    assert ws.cell(row=7, column=8).value == "PlanningOnly"


def test_write_schedule_output_report_adds_diagnostics_tab(tmp_path):
    """Schedule-Diagnostics should surface diagnose-schedule next actions."""
    import openpyxl

    so, si = _make_schedule_pair()
    si["gym_allocation"] = {
        "source": "allocator",
        "mode_shortfall": {"Badminton Court": 13.0},
    }

    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-Diagnostics"]
    all_values = [
        str(ws.cell(row=row, column=col).value or "")
        for row in range(1, ws.max_row + 1)
        for col in range(1, ws.max_column + 1)
    ]
    assert "VAY Sports Fest - Schedule Diagnostics" in all_values
    assert "DIAGNOSTIC OUTPUT" in "".join(all_values)
    assert "Resource Contract" in all_values
    assert "clean" in all_values
    assert "capacity note" in all_values
    assert any("but all games were scheduled" in value for value in all_values)


def test_write_schedule_output_report_diagnostics_tab_flags_physical_overlaps(tmp_path):
    """Physical venue overlaps should be visible in the generated workbook."""
    import openpyxl

    so, si = _make_schedule_pair()
    si["resources"].extend(
        [
            {
                "resource_id": "EHS-Main-Gym-BAD-1",
                "resource_type": "Badminton Court",
                "label": "Court-1",
                "day": "Sat-2",
                "open_time": "13:00",
                "close_time": "17:00",
                "slot_minutes": 60,
                "exclusive_group": "EHS Main Gym",
            },
            {
                "resource_id": "EHS-Main-Gym-BBM-1",
                "resource_type": "Basketball Court",
                "label": "Court-1",
                "day": "Sat-2",
                "open_time": "08:00",
                "close_time": "17:00",
                "slot_minutes": 60,
                "exclusive_group": "EHS Main Gym",
            },
        ]
    )

    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-Diagnostics"]
    all_values = [
        str(ws.cell(row=row, column=col).value or "")
        for row in range(1, ws.max_row + 1)
        for col in range(1, ws.max_column + 1)
    ]
    assert "Physical Venue Overlaps" in all_values
    assert "Resource Contract" in all_values
    assert "physical_mode_overlap" in all_values
    assert "EHS Main Gym" in all_values
    assert any("one physical gym is in only one mode" in value for value in all_values)


def test_write_schedule_output_report_unscheduled_section(tmp_path):
    """Unscheduled section appears in Schedule-by-Sport when games are unscheduled."""
    import openpyxl
    so, si = _make_render_schedule_pair()
    so["unscheduled"] = ["BBM-QF-1"]
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Sport"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("Unscheduled" in str(v) for v in all_values if v)
    assert any("BBM-QF-1" in str(v) for v in all_values if v)


def test_write_schedule_output_report_groups_mixed_pod_windows(tmp_path):
    """Pod resources with different slot windows render as separate time-grid sections."""
    import openpyxl

    schedule_input = {
        "games": [
            {
                "game_id": "PCK-01", "event": "Pickleball",
                "stage": "R1", "pool_id": "", "round": 1,
                "team_a_id": None, "team_b_id": None,
                "duration_minutes": 20, "resource_type": "Pickleball Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "TT-01", "event": "Table Tennis",
                "stage": "R1", "pool_id": "", "round": 1,
                "team_a_id": None, "team_b_id": None,
                "duration_minutes": 30, "resource_type": "Table Tennis Table",
                "earliest_slot": None, "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "PCK-1", "resource_type": "Pickleball Court",
                "label": "Court-1", "day": "Day-1",
                "open_time": "13:00", "close_time": "13:40", "slot_minutes": 20,
            },
            {
                "resource_id": "TT-1", "resource_type": "Table Tennis Table",
                "label": "Table-1", "day": "Day-1",
                "open_time": "18:00", "close_time": "19:00", "slot_minutes": 30,
            },
        ],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-15T07:07:48",
        "status": "PARTIAL",
        "solver_wall_seconds": 0.2,
        "assignments": [
            {"game_id": "PCK-01", "resource_id": "PCK-1", "slot": "Day-1-13:00"},
            {"game_id": "TT-01", "resource_id": "TT-1", "slot": "Day-1-18:00"},
        ],
        "unscheduled": [],
        "pool_results": [],
    }
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]

    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    string_values = [str(v) for v in all_values if v is not None]

    assert any("Pickleball Court" in v for v in string_values)
    assert any("Table Tennis Table" in v for v in string_values)
    assert "13:00" in string_values
    assert "18:00" in string_values
    assert any("PCK-01" in v for v in string_values)
    assert any("TT-01" in v for v in string_values)


def test_write_schedule_output_report_merges_gym_core_day_sections(tmp_path):
    """Gym Core resources for one day/resource_type render as one continuous section."""
    import openpyxl

    schedule_input = {
        "games": [
            {
                "game_id": "VBM-01", "event": "Volleyball - Men Team",
                "stage": "Pool", "pool_id": "P1", "round": 1,
                "team_a_id": "VBM::RPC", "team_b_id": "VBM::SDC",
                "duration_minutes": 60, "resource_type": "Volleyball Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "VBM-02", "event": "Volleyball - Men Team",
                "stage": "Pool", "pool_id": "P1", "round": 2,
                "team_a_id": "VBM::FVC", "team_b_id": "VBM::ORN",
                "duration_minutes": 60, "resource_type": "Volleyball Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "VBW-01", "event": "Volleyball - Women Team",
                "stage": "Pool", "pool_id": "P1", "round": 1,
                "team_a_id": "VBW::RPC", "team_b_id": "VBW::PCC",
                "duration_minutes": 60, "resource_type": "Volleyball Court",
                "earliest_slot": None, "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "GYM-Day-1-1", "resource_type": "Volleyball Court",
                "label": "Court-1", "day": "Day-1", "open_time": "08:00",
                "close_time": "18:00", "slot_minutes": 60,
                "solver_pool": ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL,
                "exclusive_group": "HS Big Gym",
            },
            {
                "resource_id": "GYM-Day-1-4", "resource_type": "Volleyball Court",
                "label": "Court-1", "day": "Day-1", "open_time": "12:00",
                "close_time": "21:00", "slot_minutes": 60,
                "solver_pool": ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL,
                "exclusive_group": "HS Big Gym",
            },
            {
                "resource_id": "GYM-Day-1-6", "resource_type": "Volleyball Court",
                "label": "Court-3", "day": "Day-1", "open_time": "12:00",
                "close_time": "21:00", "slot_minutes": 60,
                "solver_pool": ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL,
                "exclusive_group": "HS Big Gym",
            },
            {
                "resource_id": "VB-Day-1-1", "resource_type": "Volleyball Court",
                "label": "Court-2", "day": "Day-1", "open_time": "10:00",
                "close_time": "13:00", "slot_minutes": 60,
                "solver_pool": ScheduleWorkbookBuilder._GYM_CORE_SOLVER_POOL,
                "venue_name": "Orange Gym",
                "exclusive_group": "",
            },
        ],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-20T14:07:19",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {"game_id": "VBM-01", "resource_id": "GYM-Day-1-1", "slot": "Day-1-08:00"},
            {"game_id": "VBW-01", "resource_id": "VB-Day-1-1", "slot": "Day-1-10:00"},
            {"game_id": "VBM-02", "resource_id": "GYM-Day-1-4", "slot": "Day-1-14:00"},
        ],
        "unscheduled": [],
        "pool_results": [],
    }
    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]

    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    string_values = [str(v) for v in all_values if v is not None]

    volleyball_sections = [v for v in string_values if "Volleyball Court" in v]
    assert len(volleyball_sections) == 1
    assert "08:00-21:00" in volleyball_sections[0]
    assert "08:00" in string_values
    assert "20:00" in string_values
    assert any("HS Big Gym Court-1 [08:00-18:00]" in v for v in string_values)
    assert any("HS Big Gym Court-1 [12:00-21:00]" in v for v in string_values)
    assert any("HS Big Gym Court-3" in v for v in string_values)
    assert any("Orange Gym Court-2" in v for v in string_values)
    assert any("VBM-01" in v for v in string_values)
    assert any("VBM-02" in v for v in string_values)
    assert any("VBW-01" in v for v in string_values)


def test_write_schedule_output_report_labels_allocator_badminton_by_gym(tmp_path):
    """Allocator-created non-Gym-Core resources should still show their gym name."""
    import openpyxl
    from types import SimpleNamespace

    resources = ScheduleWorkbookBuilder._build_gym_resources_from_allocator(
        [
            SimpleNamespace(
                gym_name="EHS Main Gym",
                day="Sat-2",
                open_time="13:00",
                close_time="17:00",
                mode="Badminton Court",
                courts=2,
                slot_minutes=60,
            )
        ]
    )
    assert all(resource["venue_name"] == "EHS Main Gym" for resource in resources)
    schedule_input = {
        "games": [
            {
                "game_id": "BAD-Men-Doubles-01",
                "event": "Badminton - Men Doubles",
                "stage": "Pool",
                "pool_id": "P1",
                "round": 1,
                "team_a_id": "BAD::RPC",
                "team_b_id": "BAD::TLC",
                "duration_minutes": 60,
                "resource_type": "Badminton Court",
            }
        ],
        "resources": resources,
        "day_order": ["Sat-2"],
    }
    schedule_output = {
        "solved_at": "2026-05-25T09:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {
                "game_id": "BAD-Men-Doubles-01",
                "resource_id": "GYM-Sat-2-1",
                "slot": "Sat-2-13:00",
            }
        ],
        "unscheduled": [],
        "pool_results": [],
    }

    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Master-Schedule"]
    header_values = [
        ws.cell(row=2, column=col).value
        for col in range(1, ws.max_column + 1)
    ]

    assert "EHS Main Gym" in header_values
    assert "Other" not in header_values
    assert ws.cell(row=3, column=3).value == "Court-1"


def test_master_schedule_first_write_wins_on_exclusive_group_double_booking(tmp_path):
    """When two resources share the same physical column (exclusive_group double-booking),
    the first assignment (Basketball) must stay visible; the cell turns red but is not
    overwritten by the second resource (Badminton). Reproduces the BBM-Semi-1 hidden
    bug where Badminton overwrote Basketball on the same physical court."""
    import openpyxl
    import zipfile

    schedule_input = {
        "games": [
            {
                "game_id": "BBM-Semi-1",
                "event": "Basketball - Men Team",
                "stage": "Semi",
                "pool_id": "",
                "round": 1,
                "team_a_id": "WIN-BBM-QF-1",
                "team_b_id": "WIN-BBM-QF-2",
                "team_a_label": "Winner QF-1",
                "team_b_label": "Winner QF-2",
                "duration_minutes": 60,
                "resource_type": "Basketball Court",
                "solver_pool": "Gym Core",
                "earliest_slot": None,
                "latest_slot": None,
            },
            {
                "game_id": "BAD-Men-Doubles-01",
                "event": "Badminton",
                "stage": "R1",
                "pool_id": "",
                "round": 1,
                "team_a_id": None,
                "team_b_id": None,
                "duration_minutes": 60,
                "resource_type": "Badminton Court",
                "solver_pool": None,
                "earliest_slot": None,
                "latest_slot": None,
            },
        ],
        "resources": [
            # Basketball Court-1: open all day, Gym Core pool
            {
                "resource_id": "GYM-Sat-2-1",
                "resource_type": "Basketball Court",
                "label": "Court-1",
                "day": "Sat-2",
                "open_time": "08:00",
                "close_time": "17:00",
                "slot_minutes": 60,
                "solver_pool": "Gym Core",
                "venue_name": "EHS Main Gym",
                "exclusive_group": "EHS Main Gym",
            },
            # Badminton Court-1: same physical court, afternoon only, no pool
            {
                "resource_id": "GYM-Sat-2-6",
                "resource_type": "Badminton Court",
                "label": "Court-1",
                "day": "Sat-2",
                "open_time": "13:00",
                "close_time": "17:00",
                "slot_minutes": 60,
                "solver_pool": None,
                "venue_name": "EHS Main Gym",
                "exclusive_group": "EHS Main Gym",
            },
        ],
        "day_order": ["Sat-2"],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-25T09:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            # Basketball placed at 13:00 (same slot as Badminton — double-booking)
            {"game_id": "BBM-Semi-1",        "resource_id": "GYM-Sat-2-1", "slot": "Sat-2-13:00"},
            {"game_id": "BAD-Men-Doubles-01", "resource_id": "GYM-Sat-2-6", "slot": "Sat-2-13:00"},
        ],
        "unscheduled": [],
        "pool_results": [],
    }

    out = tmp_path / "sched.xlsx"
    ScheduleWorkbookBuilder._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Master-Schedule"]

    # Collect all Master-Schedule cell values
    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    string_values = [str(v) for v in all_values if v is not None]

    # BBM-Semi-1 MUST appear (first write wins)
    assert any("BBM-Semi-1" in v for v in string_values), "BBM-Semi-1 should appear in Master-Schedule"
    # BAD-Men-Doubles-01 must NOT overwrite the basketball cell
    bad_texts = [v for v in string_values if "BAD-Men-Doubles-01" in v]
    assert not bad_texts, "Badminton game must not overwrite Basketball in shared column"

    # The conflicted cell must be marked red with yellow text and carry a Note.
    conflict_cell = next(
        (
            ws.cell(row=r, column=c)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if "BBM-Semi-1" in str(ws.cell(row=r, column=c).value or "")
        ),
        None,
    )
    assert conflict_cell is not None
    assert conflict_cell.fill.fgColor.rgb.endswith("FFCC00"), "Conflict cell must have yellow fill"
    assert conflict_cell.font.color.rgb.endswith("FF2400"), "Conflict cell must have red text"
    assert conflict_cell.comment is not None, "Conflict cell must carry a Note"
    assert "SCHEDULING CONFLICT" in conflict_cell.comment.text

    # Excel stores comment visibility in VML. openpyxl can read the workbook
    # after our post-save patch, and the VML should request an open note.
    with zipfile.ZipFile(out) as zf:
        vml_names = [name for name in zf.namelist() if name.endswith(".vml")]
        assert vml_names
        vml_text = "\n".join(zf.read(name).decode("utf-8") for name in vml_names)
    assert "visibility:visible" in vml_text
    assert ":Visible" in vml_text
    assert 'xmlns="urn:schemas-microsoft-com:office:excel"' not in vml_text


# ---------------------------------------------------------------------------
# Bible Challenge Venue-Estimator tests (Issue #118)
# ---------------------------------------------------------------------------

def _bc_roster(church_codes, n_per_church=3):
    """Build minimal BC roster rows — enough to meet min_team_size=3."""
    rows = []
    for code in church_codes:
        for _ in range(n_per_church):
            rows.append({"Church Team": code, "sport_type": SPORT_TYPE["BIBLE_CHALLENGE"], "sport_gender": "Mixed"})
    return rows


def test_bc_venue_estimator_rr_game_count_12_teams():
    """12 BC teams × 2 games/team ÷ 3 teams/game = 8 RR games."""
    from config import (
        COURT_ESTIMATE_BC_TEAMS_PER_GAME,
        COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM,
        COURT_ESTIMATE_BC_PLAYOFF_GAMES,
        COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF,
        COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
    )
    builder = ScheduleWorkbookBuilder()
    churches = [f"C{i:02d}" for i in range(1, 13)]  # 12 teams
    rows = _bc_roster(churches)
    capacity_rows = builder._build_venue_capacity_rows(rows)
    bc_row = next(r for r in capacity_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"])

    assert bc_row["Estimating Teams/Entries"] == 12
    assert bc_row["Pool Slots"] == 12                       # ceil(12*3/3)
    assert bc_row["Actual Pool Games/Team"] == 3
    assert bc_row["Playoff Teams"] == COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF
    assert bc_row["Playoff Slots"] == COURT_ESTIMATE_BC_PLAYOFF_GAMES  # 4
    assert bc_row["Total Court Slots"] == 16                # 12 RR + 4 playoff
    assert bc_row["Minutes Per Game"] == COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE  # 60
    assert bc_row["Estimated Court Hours"] == 16.0          # 16 games × 60 min / 60
    assert bc_row["Third Place?"] == "No"
    assert "Sequential" in bc_row["Pool Composition"]
    assert "1 classroom" in bc_row["Pool Composition"]


def test_bc_venue_estimator_no_playoff_when_fewer_than_9_teams():
    """8 BC teams → no playoff (need ≥ 9). Pool Slots = ceil(8*2/3) = 6."""
    from config import COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF
    builder = ScheduleWorkbookBuilder()
    churches = [f"C{i:02d}" for i in range(1, 9)]  # 8 teams
    rows = _bc_roster(churches)
    capacity_rows = builder._build_venue_capacity_rows(rows)
    bc_row = next(r for r in capacity_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"])

    assert bc_row["Estimating Teams/Entries"] == 8
    assert bc_row["Pool Slots"] == 8                        # ceil(8*3/3) = 8
    assert bc_row["Actual Pool Games/Team"] == 3
    assert bc_row["Playoff Teams"] == 0
    assert bc_row["Playoff Slots"] == 0
    assert bc_row["Total Court Slots"] == 8
    assert f"< {COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF}" in bc_row["Pool Composition"]


def test_bc_venue_estimator_zero_teams():
    """0 BC teams → all zeros, no crash."""
    builder = ScheduleWorkbookBuilder()
    capacity_rows = builder._build_venue_capacity_rows([])
    bc_row = next(r for r in capacity_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"])

    assert bc_row["Estimating Teams/Entries"] == 0
    assert bc_row["Actual Pool Games/Team"] == 0
    assert bc_row["Pool Slots"] == 0
    assert bc_row["Playoff Slots"] == 0
    assert bc_row["Total Court Slots"] == 0
    assert bc_row["Estimated Court Hours"] == 0.0


def test_bc_venue_estimator_not_in_court_hours_model():
    """BC must NOT appear via the standard concurrent court-hours COURT_ESTIMATE_EVENTS path."""
    from config import COURT_ESTIMATE_EVENTS
    assert SPORT_TYPE["BIBLE_CHALLENGE"] not in COURT_ESTIMATE_EVENTS


def test_bc_venue_estimator_waits_for_first_three_team_game():
    """With only 2 BC teams, the sequential queue should remain in waiting mode."""
    builder = ScheduleWorkbookBuilder()
    rows = _bc_roster(["RPC", "OCB"])
    capacity_rows = builder._build_venue_capacity_rows(rows)
    bc_row = next(r for r in capacity_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"])

    assert bc_row["Estimating Teams/Entries"] == 2
    assert bc_row["Actual Pool Games/Team"] == 0
    assert bc_row["Pool Slots"] == 0
    assert bc_row["Total Court Slots"] == 0
    assert "waiting for at least 3 teams" in bc_row["Pool Composition"]


def test_bc_minutes_per_game_is_60():
    """Confirm COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE was updated to 60."""
    from config import COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE
    assert COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE == 60


# ---------------------------------------------------------------------------
# Bible Challenge Pool-Assignment & conflict-edge tests (Issue #118)
# ---------------------------------------------------------------------------

def test_bc_appears_in_pool_assignment_base_rows():
    """BC teams meeting min_team_size=3 should produce Pool-Assignment rows."""
    builder = ScheduleWorkbookBuilder()
    churches = ["RPC", "OCB", "ANH"]
    rows = _bc_roster(churches, n_per_church=3)
    base_rows = builder._build_pool_assignment_base_rows(rows)
    bc_rows = [r for r in base_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"]]
    assert {r["Team ID"] for r in bc_rows} == set(churches)
    for r in bc_rows:
        assert r["Min Team Size"] == 3
        assert r["Roster Count"] == 3


def test_bc_pool_assignment_below_min_team_size_excluded():
    """BC team with fewer than 3 roster members is excluded from Pool-Assignment."""
    builder = ScheduleWorkbookBuilder()
    rows = (
        _bc_roster(["RPC"], n_per_church=3)
        + _bc_roster(["OCB"], n_per_church=2)  # below min
    )
    base_rows = builder._build_pool_assignment_base_rows(rows)
    bc_rows = [r for r in base_rows if r["Event"] == SPORT_TYPE["BIBLE_CHALLENGE"]]
    assert {r["Team ID"] for r in bc_rows} == {"RPC"}


def test_bc_cross_sport_conflict_edge_with_basketball(tmp_path):
    """BC and BB teams sharing an athlete must produce a team_conflicts edge."""
    builder = ScheduleWorkbookBuilder()
    shared_id = 42
    # Need ≥ 2 BB teams and ≥ 2 BC teams so the pool-assignment logic
    # doesn't fall into the "WaitingForMoreTeams" branch.
    bb_rpc = [
        {
            "Church Team": "RPC",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        }
        for pid in range(40, 45)  # includes shared_id=42
    ]
    bb_anh = [
        {
            "Church Team": "ANH",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        }
        for pid in range(50, 55)
    ]
    bc_ocb = [
        {
            "Church Team": "OCB",
            "sport_type": SPORT_TYPE["BIBLE_CHALLENGE"],
            "sport_gender": "Mixed",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],  # primary is BB
        }
        for pid in (shared_id, 90, 91)  # #42 is shared with RPC's BB team
    ]
    bc_tlc = [
        {
            "Church Team": "TLC",
            "sport_type": SPORT_TYPE["BIBLE_CHALLENGE"],
            "sport_gender": "Mixed",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BIBLE_CHALLENGE"],
        }
        for pid in (95, 96, 97)
    ]
    roster_rows = bb_rpc + bb_anh + bc_ocb + bc_tlc

    si = builder._build_schedule_input(roster_rows, [], tmp_path / "no_venue.xlsx")

    bc_bb_edges = [
        edge for edge in si["team_conflicts"]
        if {edge.get("event_a"), edge.get("event_b")}
           == {SPORT_TYPE["BASKETBALL"], SPORT_TYPE["BIBLE_CHALLENGE"]}
    ]
    assert len(bc_bb_edges) == 1
    edge = bc_bb_edges[0]
    assert edge["shared_count"] == 1
    assert edge["primary_overlap_count"] == 1  # athlete's primary is BB
    assert edge["secondary_only_count"] == 0


def test_bc_pool_assignment_creates_bc_station_games(tmp_path):
    """BC pool games: all teams treated as one global pool.

    For 9 teams: 9 pool games (n * 3 / 3 = n); each team appears exactly 3
    times; no pair of teams meets in more than one game."""
    builder = ScheduleWorkbookBuilder()

    bc_rows_nine = _bc_roster(
        ["RPC", "OCB", "ANH", "GLA", "TLC", "FVC", "MWC", "NSD", "WCC"], n_per_church=3
    )
    si_nine = builder._build_schedule_input(bc_rows_nine, [], tmp_path / "nine.xlsx")
    bc_nine = [g for g in si_nine.get("games", []) if g.get("event") == SPORT_TYPE["BIBLE_CHALLENGE"]
               and g.get("stage") == "Pool"]
    assert len(bc_nine) == 9

    # All pool game IDs follow the flat BC-RR-N scheme.
    assert all(g["game_id"].startswith("BC-RR-") for g in bc_nine)

    # Each of the 9 teams appears in exactly 3 pool games.
    appearances: dict = {}
    for g in bc_nine:
        for key in ("team_a_id", "team_b_id", "team_c_id"):
            appearances[g[key]] = appearances.get(g[key], 0) + 1
    assert all(v == 3 for v in appearances.values()), appearances

    # No pair of teams meets in more than one game.
    seen_pairs: set = set()
    for g in bc_nine:
        ids = [g["team_a_id"], g["team_b_id"], g["team_c_id"]]
        for i in range(3):
            for j in range(i + 1, 3):
                pair = (min(ids[i], ids[j]), max(ids[i], ids[j]))
                assert pair not in seen_pairs, \
                    f"Pair {pair} appears in more than one BC game"
                seen_pairs.add(pair)


def test_bc_no_repeat_triplets_nine_teams():
    """_bc_no_repeat_triplets produces 9 valid games for 9 teams: each team
    plays 3 times and no pair meets twice."""
    rows = [
        {"Pool ID": "P1", "Pool Slot": f"T{i+1}", "Team ID": chr(65 + i)}
        for i in range(9)
    ]
    triplets = ScheduleWorkbookBuilder._bc_no_repeat_triplets(rows)
    assert len(triplets) == 9

    appearances: dict = {}
    seen_pairs: set = set()
    for trio in triplets:
        ids = [r["Team ID"] for r in trio]
        for team_id in ids:
            appearances[team_id] = appearances.get(team_id, 0) + 1
        for i in range(3):
            for j in range(i + 1, 3):
                pair = (min(ids[i], ids[j]), max(ids[i], ids[j]))
                assert pair not in seen_pairs, f"Pair {pair} repeats"
                seen_pairs.add(pair)

    assert appearances == {chr(65 + i): 3 for i in range(9)}


def test_bc_no_repeat_triplets_seeded_teams_never_share_game():
    """Seeded teams must never appear in the same triplet; they should only
    meet in playoffs.  Verified with 9 teams where A, B, C are seeded."""
    rows = []
    for i in range(9):
        team_id = chr(65 + i)
        seed = str(i + 1) if i < 3 else ""  # A=seed1, B=seed2, C=seed3
        rows.append({"Pool ID": "P1", "Pool Slot": f"T{i+1}", "Team ID": team_id, "Seed": seed})

    triplets = ScheduleWorkbookBuilder._bc_no_repeat_triplets(rows)
    assert len(triplets) == 9

    seeded_ids = {"A", "B", "C"}
    for trio in triplets:
        ids = {r["Team ID"] for r in trio}
        seeded_in_game = ids & seeded_ids
        assert len(seeded_in_game) <= 1, \
            f"Two seeded teams appear in the same game: {ids}"


def test_bc_schedule_input_adds_playoff_precedence(tmp_path):
    """Nine BC teams should keep all BC prelims ahead of semis, then semis ahead of the final."""
    builder = ScheduleWorkbookBuilder()
    bc_rows = _bc_roster(
        ["RPC", "OCB", "ANH", "GLA", "TLC", "FVC", "MWC", "NSD", "WCC"],
        n_per_church=3,
    )

    si = builder._build_schedule_input(bc_rows, [], tmp_path / "no_venue.xlsx")
    bc_games = [g for g in si["games"] if g.get("event") == SPORT_TYPE["BIBLE_CHALLENGE"]]
    pool_game_ids = {
        str(g["game_id"])
        for g in bc_games
        if g.get("stage") == "Pool"
    }
    semi_ids = {
        str(g["game_id"])
        for g in bc_games
        if g.get("stage") == "Semi"
    }
    final_ids = {
        str(g["game_id"])
        for g in bc_games
        if g.get("stage") == "Final"
    }
    assert semi_ids == {"BC-Semi-1", "BC-Semi-2", "BC-Semi-3"}
    assert final_ids == {"BC-Final"}

    precedence_pairs = {
        (str(rule["before_game_id"]), str(rule["after_game_id"]))
        for rule in si["precedence"]
        if str(rule.get("before_game_id") or "").startswith("BC-")
           or str(rule.get("after_game_id") or "").startswith("BC-")
    }
    expected_pairs = {
        (pool_game_id, semi_id)
        for pool_game_id in pool_game_ids
        for semi_id in semi_ids
    } | {
        (semi_id, "BC-Final")
        for semi_id in semi_ids
    }
    assert precedence_pairs == expected_pairs
    bc_rules = [
        r for r in si["precedence"]
        if str(r.get("before_game_id") or "").startswith("BC-")
        or str(r.get("after_game_id") or "").startswith("BC-")
    ]
    assert all(int(rule.get("min_gap_slots") or 0) == 1 for rule in bc_rules)


def test_bc_event_in_pool_assignment_defs():
    """BC must be the 4th entry in _POOL_ASSIGNMENT_EVENT_DEFS with prefix 'BC'."""
    defs = ScheduleWorkbookBuilder._POOL_ASSIGNMENT_EVENT_DEFS
    assert (SPORT_TYPE["BIBLE_CHALLENGE"], "BC") in defs
    # Sort index makes BC the last (4th) event
    idx = ScheduleWorkbookBuilder._event_sort_index(SPORT_TYPE["BIBLE_CHALLENGE"])
    assert idx == 3


# ---------------------------------------------------------------------------
# Soccer (optional / config-driven) tests — Issue #118
# ---------------------------------------------------------------------------

def _soccer_roster(church_codes, n_per_church=4):
    """Build minimal Soccer roster rows — enough to meet min_team_size=4."""
    rows = []
    for code in church_codes:
        for _ in range(n_per_church):
            rows.append({"Church Team": code, "sport_type": SPORT_TYPE["SOCCER"], "sport_gender": "Mixed"})
    return rows


def test_soccer_in_pool_assignment_defs_when_enabled():
    """When SOCCER_ENABLED is True (default), Soccer is included in pool defs as 5th entry."""
    from config import SOCCER_ENABLED
    assert SOCCER_ENABLED is True
    defs = ScheduleWorkbookBuilder._POOL_ASSIGNMENT_EVENT_DEFS
    assert (SPORT_TYPE["SOCCER"], "SOC") in defs
    idx = ScheduleWorkbookBuilder._event_sort_index(SPORT_TYPE["SOCCER"])
    assert idx == 4  # after BB(0), VBM(1), VBW(2), BC(3)


def test_soccer_in_court_estimate_events_when_enabled():
    """Soccer should appear in the standard court-hours estimator when enabled."""
    from config import COURT_ESTIMATE_EVENTS, SOCCER_ENABLED
    assert SOCCER_ENABLED is True
    assert SPORT_TYPE["SOCCER"] in COURT_ESTIMATE_EVENTS


def test_soccer_appears_in_pool_assignment_base_rows():
    """Soccer teams meeting min_team_size=4 should produce Pool-Assignment rows."""
    builder = ScheduleWorkbookBuilder()
    churches = ["RPC", "OCB", "ANH"]
    rows = _soccer_roster(churches, n_per_church=4)
    base_rows = builder._build_pool_assignment_base_rows(rows)
    soccer_rows = [r for r in base_rows if r["Event"] == SPORT_TYPE["SOCCER"]]
    assert {r["Team ID"] for r in soccer_rows} == set(churches)
    for r in soccer_rows:
        assert r["Min Team Size"] == 4


def test_soccer_pool_assignment_below_min_team_size_excluded():
    """Soccer team with fewer than 4 roster members is excluded from Pool-Assignment."""
    builder = ScheduleWorkbookBuilder()
    rows = (
        _soccer_roster(["RPC"], n_per_church=4)
        + _soccer_roster(["OCB"], n_per_church=3)  # below min
    )
    base_rows = builder._build_pool_assignment_base_rows(rows)
    soccer_rows = [r for r in base_rows if r["Event"] == SPORT_TYPE["SOCCER"]]
    assert {r["Team ID"] for r in soccer_rows} == {"RPC"}


def test_soccer_cross_sport_conflict_edge_with_basketball(tmp_path):
    """Soccer and BB teams sharing an athlete must produce a team_conflicts edge."""
    builder = ScheduleWorkbookBuilder()
    shared_id = 42
    bb_rpc = [
        {
            "Church Team": "RPC",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        }
        for pid in range(40, 45)  # includes shared_id=42
    ]
    bb_anh = [
        {
            "Church Team": "ANH",
            "sport_type": SPORT_TYPE["BASKETBALL"],
            "sport_gender": "Men",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["BASKETBALL"],
        }
        for pid in range(50, 55)
    ]
    soccer_ocb = [
        {
            "Church Team": "OCB",
            "sport_type": SPORT_TYPE["SOCCER"],
            "sport_gender": "Mixed",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["SOCCER"],  # primary is Soccer
        }
        for pid in (shared_id, 90, 91, 92)
    ]
    soccer_tlc = [
        {
            "Church Team": "TLC",
            "sport_type": SPORT_TYPE["SOCCER"],
            "sport_gender": "Mixed",
            "sport_format": "Team",
            "Participant ID (WP)": pid,
            "participant_primary_sport": SPORT_TYPE["SOCCER"],
        }
        for pid in (95, 96, 97, 98)
    ]
    roster_rows = bb_rpc + bb_anh + soccer_ocb + soccer_tlc

    si = builder._build_schedule_input(roster_rows, [], tmp_path / "no_venue.xlsx")

    bb_soccer_edges = [
        edge for edge in si["team_conflicts"]
        if {edge.get("event_a"), edge.get("event_b")}
           == {SPORT_TYPE["BASKETBALL"], SPORT_TYPE["SOCCER"]}
    ]
    assert len(bb_soccer_edges) == 1
    edge = bb_soccer_edges[0]
    assert edge["shared_count"] == 1
    # Athlete's primary is Soccer, which is one of the events on this edge.
    assert edge["primary_overlap_count"] == 1


# ── Issue #158: racquet (pod) cross-sport conflict modeling ──────────────────

def _badminton_doubles_pair(genders, pairs):
    """Return roster rows for confirmed Badminton doubles pairs.

    pairs: list of (pid_a, name_a, pid_b, name_b, primary_a, primary_b).
    """
    rows = []
    for (pid_a, name_a, pid_b, name_b, prim_a, prim_b) in pairs:
        rows.append({
            "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": genders,
            "sport_format": "Men Doubles", "Participant ID (WP)": pid_a,
            "First Name": name_a, "Last Name": "X", "partner_name": f"{name_b} X",
            "Church Team": "RPC", "participant_primary_sport": prim_a,
        })
        rows.append({
            "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": genders,
            "sport_format": "Men Doubles", "Participant ID (WP)": pid_b,
            "First Name": name_b, "Last Name": "X", "partner_name": f"{name_a} X",
            "Church Team": "RPC", "participant_primary_sport": prim_b,
        })
    return rows


def test_resolve_pod_doubles_assigns_stable_reproducible_ids():
    """Confirmed doubles pairs get stable {division}-E{nn} IDs across re-runs."""
    builder = ScheduleWorkbookBuilder()
    roster = _badminton_doubles_pair("Men", [
        ("11", "Anh", "12", "Binh", "Badminton", "Badminton"),
        ("13", "Cuong", "14", "Dung", "Badminton", "Badminton"),
    ])

    confirmed, unprotected = builder._resolve_pod_doubles(roster, [])
    assert unprotected == []
    entries = confirmed["BAD-Men-Doubles"]
    assert [e["entry_id"] for e in entries] == ["BAD-Men-Doubles-E01", "BAD-Men-Doubles-E02"]
    assert all(len(e["participant_ids"]) == 2 for e in entries)
    assert entries[0]["sport_type"] == SPORT_TYPE["BADMINTON"]

    # Re-running on a reshuffled roster yields identical IDs (deterministic).
    confirmed2, _ = builder._resolve_pod_doubles(list(reversed(roster)), [])
    assert [e["entry_id"] for e in confirmed2["BAD-Men-Doubles"]] == \
        ["BAD-Men-Doubles-E01", "BAD-Men-Doubles-E02"]
    assert {tuple(sorted(e["participant_ids"])) for e in entries} == \
        {tuple(sorted(e["participant_ids"])) for e in confirmed2["BAD-Men-Doubles"]}


def test_pod_game_objects_assigns_r1_team_ids_for_confirmed_doubles():
    """Round-1 games for confirmed doubles carry real entry IDs; later rounds stay None.

    N=4, P=4: round 1 = Semi (2 games with team IDs), round 2 = Final (None).
    """
    builder = ScheduleWorkbookBuilder()
    # 4 confirmed pairs → 4 planning entries → 3 games; R1 matchups = floor(4/2)=2.
    roster = _badminton_doubles_pair("Men", [
        ("11", "Anh", "12", "Binh", "Badminton", "Badminton"),
        ("13", "Cuong", "14", "Dung", "Badminton", "Badminton"),
        ("15", "Em", "16", "Phuc", "Badminton", "Badminton"),
        ("17", "Giang", "18", "Hai", "Badminton", "Badminton"),
    ])
    games, prec = builder._build_pod_game_objects(roster, [])
    assert len(games) == 4  # 4 - 1 elimination games + third place
    # N=4, P=4: Semi-1 and Semi-2 get team IDs; Final has None.
    semi_games = [g for g in games if g["stage"] == "Semi"]
    assert len(semi_games) == 2
    for g in semi_games:
        assert g["team_a_id"].startswith("BAD-Men-Doubles-E")
        assert g["team_b_id"].startswith("BAD-Men-Doubles-E")
        assert g["team_a_id"] != g["team_b_id"]
    final_game = next(g for g in games if g["stage"] == "Final")
    assert final_game["team_a_id"] is None and final_game["team_b_id"] is None
    third_game = next(g for g in games if g["stage"] == "3rd")
    assert third_game["game_id"] == "BAD-Men-Doubles-3rd"
    assert third_game["team_a_id"] is None and third_game["team_b_id"] is None
    # Precedence: Semi-1, Semi-2 → Final and 3rd
    assert len(prec) == 4
    after_ids = {p["after_game_id"] for p in prec}
    assert after_ids == {"BAD-Men-Doubles-Final", "BAD-Men-Doubles-3rd"}


def test_pod_game_objects_r1_matchups_are_bye_aware():
    """Bye-aware bracket math: N=5 yields 1 first-round match (QF-1), not 2.

    N=5, P=8, byes=3: round 1 = QF (1 game), round 2 = Semi (2 games), round 3 = Final.
    The two lowest-seeded entries (E04, E05) play in QF-1; E01-E03 get byes to Semi.
    """
    builder = ScheduleWorkbookBuilder()
    roster = _badminton_doubles_pair("Men", [
        ("11", "Anh", "12", "Binh", "Badminton", "Badminton"),
        ("13", "Cuong", "14", "Dung", "Badminton", "Badminton"),
        ("15", "Em", "16", "Phuc", "Badminton", "Badminton"),
        ("17", "Giang", "18", "Hai", "Badminton", "Badminton"),
        ("19", "Kiet", "20", "Long", "Badminton", "Badminton"),
    ])
    games, prec = builder._build_pod_game_objects(roster, [])
    assert len(games) == 5  # 5 - 1 elimination games + third place

    stages = [g["stage"] for g in games]
    assert stages == ["QF", "Semi", "Semi", "Final", "3rd"]
    game_ids = [g["game_id"] for g in games]
    assert game_ids == [
        "BAD-Men-Doubles-QF-1",
        "BAD-Men-Doubles-Semi-1",
        "BAD-Men-Doubles-Semi-2",
        "BAD-Men-Doubles-Final",
        "BAD-Men-Doubles-3rd",
    ]

    # Only QF-1 has known team IDs (the 2 non-bye players)
    qf_game = games[0]
    matched_ids = {qf_game["team_a_id"], qf_game["team_b_id"]}
    assert matched_ids == {"BAD-Men-Doubles-E04", "BAD-Men-Doubles-E05"}
    # Semi and Final games have no known participants yet
    assert all(g["team_a_id"] is None for g in games[1:])

    # Precedence: QF-1 → Semi-1, QF-1 → Semi-2, Semi-1 → Final, Semi-2 → Final
    assert len(prec) == 6
    prec_pairs = {(p["before_game_id"], p["after_game_id"]) for p in prec}
    assert prec_pairs == {
        ("BAD-Men-Doubles-QF-1", "BAD-Men-Doubles-Semi-1"),
        ("BAD-Men-Doubles-QF-1", "BAD-Men-Doubles-Semi-2"),
        ("BAD-Men-Doubles-Semi-1", "BAD-Men-Doubles-Final"),
        ("BAD-Men-Doubles-Semi-2", "BAD-Men-Doubles-Final"),
        ("BAD-Men-Doubles-Semi-1", "BAD-Men-Doubles-3rd"),
        ("BAD-Men-Doubles-Semi-2", "BAD-Men-Doubles-3rd"),
    }


def test_pod_game_objects_byes_include_unresolved_planned_entries():
    """Unresolved entries affect bye math without receiving false team IDs."""
    builder = ScheduleWorkbookBuilder()
    roster = _badminton_doubles_pair("Men", [
        ("11", "Anh", "12", "Binh", "Badminton", "Badminton"),
        ("13", "Cuong", "14", "Dung", "Badminton", "Badminton"),
        ("15", "Em", "16", "Phuc", "Badminton", "Badminton"),
        ("17", "Giang", "18", "Hai", "Badminton", "Badminton"),
    ])
    roster.extend([
        {
            "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
            "sport_format": "Men Doubles", "Participant ID (WP)": "90",
            "First Name": "Unknown", "Last Name": "One",
            "partner_name": "Missing One", "Church Team": "RPC",
            "participant_primary_sport": SPORT_TYPE["BADMINTON"],
        },
        {
            "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
            "sport_format": "Men Doubles", "Participant ID (WP)": "91",
            "First Name": "Unknown", "Last Name": "Two",
            "partner_name": "Missing Two", "Church Team": "RPC",
            "participant_primary_sport": SPORT_TYPE["BADMINTON"],
        },
    ])

    games, _prec = builder._build_pod_game_objects(roster, [])

    assert len(games) == 5  # four elimination games + third place
    # First game (QF-1) has one known player; the other slot is None (unresolved entry)
    identified_r1 = [g for g in games if g["team_a_id"] or g["team_b_id"]]
    assert len(identified_r1) == 1
    assert {
        identified_r1[0]["team_a_id"],
        identified_r1[0]["team_b_id"],
    } == {"BAD-Men-Doubles-E04", None}
    assert all(g["division_entry_count"] == 5 for g in games)


# ── Issue #130: stage-aware racquet late-round game IDs ──────────────────────


def test_build_pod_game_objects_tt_singles_bracket_p8():
    """TT Men Singles with N=6 uses a P=8 bracket: QF-1..2, Semi-1..2, Final.

    N=6, P=8, byes=2: round 1 (QF) has 2 games, round 2 (Semi) 2 games, round 3 (Final) 1 game.
    """
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["TABLE_TENNIS"], "sport_gender": "Men",
         "sport_format": "Men Single"}
        for i in range(1, 7)  # 6 entries
    ]
    games, prec = ScheduleWorkbookBuilder()._build_pod_game_objects(roster_rows, [])

    assert len(games) == 6  # 6 - 1 elimination games + third place
    game_ids = [g["game_id"] for g in games]
    assert game_ids == [
        "TT-Men-Singles-QF-1",
        "TT-Men-Singles-QF-2",
        "TT-Men-Singles-Semi-1",
        "TT-Men-Singles-Semi-2",
        "TT-Men-Singles-Final",
        "TT-Men-Singles-3rd",
    ]
    stages = [g["stage"] for g in games]
    assert stages == ["QF", "QF", "Semi", "Semi", "Final", "3rd"]
    assert all(g["event"] == SPORT_TYPE["TABLE_TENNIS"] for g in games)
    assert all(g["resource_type"] == "Table Tennis Table" for g in games)
    assert all(g["duration_minutes"] == 20 for g in games)

    # 4 QF→Semi + 2 Semi→Final + 2 Semi→3rd edges = 8 total
    assert len(prec) == 8
    qf_to_semi = [p for p in prec if "QF" in p["before_game_id"]]
    assert len(qf_to_semi) == 4
    semi_to_final = [p for p in prec if "Semi" in p["before_game_id"]]
    assert len(semi_to_final) == 4
    assert {
        p["after_game_id"] for p in semi_to_final
    } == {"TT-Men-Singles-Final", "TT-Men-Singles-3rd"}
    assert all(p["min_gap_slots"] == 1 for p in prec)


def test_build_pod_game_objects_tt_full_bracket_p8():
    """TT Women Singles with N=8 (no byes): QF-1..4, Semi-1..2, Final — 7 games."""
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["TABLE_TENNIS"], "sport_gender": "Women",
         "sport_format": "Women Single"}
        for i in range(1, 9)  # 8 entries — full P=8 bracket, no byes
    ]
    games, prec = ScheduleWorkbookBuilder()._build_pod_game_objects(roster_rows, [])

    assert len(games) == 8  # 8 - 1 elimination games + third place
    game_ids = [g["game_id"] for g in games]
    assert game_ids == [
        "TT-Women-Singles-QF-1",
        "TT-Women-Singles-QF-2",
        "TT-Women-Singles-QF-3",
        "TT-Women-Singles-QF-4",
        "TT-Women-Singles-Semi-1",
        "TT-Women-Singles-Semi-2",
        "TT-Women-Singles-Final",
        "TT-Women-Singles-3rd",
    ]
    # 8 QF→Semi + 2 Semi→Final + 2 Semi→3rd edges = 12 total
    assert len(prec) == 12


def test_build_pod_game_objects_pickleball_singles_bracket_p4():
    """Pickleball Women Singles with N=4 (P=4): Semi-1, Semi-2, Final — 3 games."""
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["PICKLEBALL"], "sport_gender": "Women",
         "sport_format": "Women Single"}
        for i in range(1, 5)  # 4 entries — P=4 bracket
    ]
    games, prec = ScheduleWorkbookBuilder()._build_pod_game_objects(roster_rows, [])

    assert len(games) == 4  # 4 - 1 elimination games + third place
    game_ids = [g["game_id"] for g in games]
    assert game_ids == [
        "PCK-Women-Singles-Semi-1",
        "PCK-Women-Singles-Semi-2",
        "PCK-Women-Singles-Final",
        "PCK-Women-Singles-3rd",
    ]
    stages = [g["stage"] for g in games]
    assert stages == ["Semi", "Semi", "Final", "3rd"]
    assert all(g["event"] == SPORT_TYPE["PICKLEBALL"] for g in games)
    assert all(g["resource_type"] == "Pickleball Court" for g in games)

    # 2 Semi→Final + 2 Semi→3rd edges
    assert len(prec) == 4
    assert {
        p["after_game_id"] for p in prec
    } == {"PCK-Women-Singles-Final", "PCK-Women-Singles-3rd"}
    assert all(p["min_gap_slots"] == 1 for p in prec)


def test_pod_precedence_included_in_schedule_input(tmp_path):
    """Racquet sport precedence edges appear in the schedule_input.json output."""
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["TABLE_TENNIS"], "sport_gender": "Men",
         "sport_format": "Men Single"}
        for i in range(1, 5)  # 4 TT Men: P=4, Semi-1, Semi-2, Final
    ]
    si = ScheduleWorkbookBuilder()._build_schedule_input(
        roster_rows, [], tmp_path / "no_venue.xlsx"
    )

    # TT-Men-Singles-Final must be in the games list
    game_ids = {g["game_id"] for g in si["games"]}
    assert "TT-Men-Singles-Final" in game_ids
    assert "TT-Men-Singles-Semi-1" in game_ids
    assert "TT-Men-Singles-Semi-2" in game_ids
    assert "TT-Men-Singles-3rd" in game_ids

    # Precedence edges from the TT bracket must be included
    tt_prec = [p for p in si["precedence"] if "TT-Men-Singles" in p["before_game_id"]]
    assert len(tt_prec) == 4  # both Semis → Final and 3rd
    after_ids = {p["after_game_id"] for p in tt_prec}
    assert after_ids == {"TT-Men-Singles-Final", "TT-Men-Singles-3rd"}


def test_schedule_output_orders_all_early_racquet_rounds_before_qf():
    """R2+ stages sort chronologically before QF/Semi/Final in the report."""
    resource = {
        "resource_id": "TT-Sat-1-1",
        "resource_type": "Table Tennis Table",
        "label": "Table-1",
        "day": "Sat-1",
        "open_time": "08:00",
        "close_time": "12:00",
        "slot_minutes": 20,
    }
    games = [
        {
            "game_id": f"TT-Men-Singles-{suffix}",
            "event": SPORT_TYPE["TABLE_TENNIS"],
            "stage": stage,
            "round": round_num,
            "duration_minutes": 20,
            "resource_type": "Table Tennis Table",
        }
        for suffix, stage, round_num in (
            ("Final", "Final", 5),
            ("QF-1", "QF", 3),
            ("02", "R2", 2),
            ("01", "R1", 1),
            ("Semi-1", "Semi", 4),
        )
    ]
    assignments = [
        {
            "game_id": game["game_id"],
            "resource_id": resource["resource_id"],
            "slot": f"Sat-1-{8 + index:02d}:00",
        }
        for index, game in enumerate(games)
    ]

    rows = ScheduleWorkbookBuilder._build_schedule_output_flat_rows(
        {"assignments": assignments},
        {"games": games, "resources": [resource]},
    )

    assert [row["stage"] for row in rows] == [
        "R1", "R2", "QF", "Semi", "Final",
    ]


def test_tt_final_pinned_via_playoff_slots(tmp_path):
    """A TT-Men-Singles-Final game is generated and can be pinned via Playoff-Slots."""
    from config import POD_RESOURCE_TYPE_TABLE_TENNIS

    # 6 TT Men Singles entries → TT-Men-Singles-Final is generated (P=8 bracket)
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["TABLE_TENNIS"], "sport_gender": "Men",
         "sport_format": "Men Single"}
        for i in range(1, 7)
    ]
    # Venue-Input: standalone TT table at "Church Hall" on Sun-2
    rows = [
        ["TT Pod", "Church Hall", POD_RESOURCE_TYPE_TABLE_TENNIS, 2, "Sun-2",
         "2026-07-26", 8, 17, 20, None, None, None, None, None],
    ]
    playoff_rows = [
        ["game_id", "event", "stage", "gym_name", "date", "start_time"],
        ["TT-Men-Singles-Final", SPORT_TYPE["TABLE_TENNIS"], "Final",
         "Church Hall", "2026-07-26", "14:00"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, _VENUE_CENTRIC_HEADERS, rows, playoff_rows=playoff_rows)

    si = ScheduleWorkbookBuilder()._build_schedule_input(roster_rows, [], path)

    # The Final game must exist in the GAMES list
    final_game = next(
        (g for g in si["games"] if g["game_id"] == "TT-Men-Singles-Final"),
        None,
    )
    assert final_game is not None, "TT-Men-Singles-Final game not generated"
    assert final_game["stage"] == "Final"
    assert final_game["event"] == SPORT_TYPE["TABLE_TENNIS"]

    # The playoff slot for the Final must be resolved (resource_id assigned)
    pinned = [ps for ps in si["playoff_slots"]
              if ps.get("game_id") == "TT-Men-Singles-Final"]
    assert len(pinned) == 1, "Expected 1 pinned TT Final slot"
    assert pinned[0].get("resource_id"), "TT Final slot should have a resource_id"
    assert pinned[0].get("slot"), "TT Final slot should have a slot label"

    # Precedence chain: Semi-1 and Semi-2 must precede the Final
    tt_prec = [p for p in si["precedence"]
               if p.get("after_game_id") == "TT-Men-Singles-Final"]
    assert len(tt_prec) == 2
    before_ids = {p["before_game_id"] for p in tt_prec}
    assert before_ids == {"TT-Men-Singles-Semi-1", "TT-Men-Singles-Semi-2"}


def test_pod_doubles_cross_sport_conflict_edge_with_basketball(tmp_path):
    """A player on a BB team and in a Badminton doubles pair → team↔racquet edge."""
    builder = ScheduleWorkbookBuilder()
    shared_id = "42"
    # Two BB teams so pool assignment does not wait for more teams.
    bb_rpc = [
        {"Church Team": "RPC", "sport_type": SPORT_TYPE["BASKETBALL"],
         "sport_gender": "Men", "sport_format": "Team", "Participant ID (WP)": pid,
         "First Name": f"P{pid}", "Last Name": "X",
         "participant_primary_sport": SPORT_TYPE["BASKETBALL"]}
        for pid in ["40", "41", shared_id, "43", "44"]
    ]
    bb_anh = [
        {"Church Team": "ANH", "sport_type": SPORT_TYPE["BASKETBALL"],
         "sport_gender": "Men", "sport_format": "Team", "Participant ID (WP)": pid,
         "First Name": f"P{pid}", "Last Name": "X",
         "participant_primary_sport": SPORT_TYPE["BASKETBALL"]}
        for pid in ["50", "51", "52", "53", "54"]
    ]
    # Two Badminton Men Doubles confirmed pairs (so the division has an R1 game);
    # one member of the first pair is the shared basketball player.
    bad = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": shared_id,
         "First Name": "P42", "Last Name": "X", "partner_name": "Buddy X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BASKETBALL"]},
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": "99",
         "First Name": "Buddy", "Last Name": "X", "partner_name": "P42 X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": "95",
         "First Name": "Tam", "Last Name": "X", "partner_name": "Quan X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": "96",
         "First Name": "Quan", "Last Name": "X", "partner_name": "Tam X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    si = builder._build_schedule_input(bb_rpc + bb_anh + bad, [], tmp_path / "no_venue.xlsx")

    edges = [
        e for e in si["team_conflicts"]
        if {e.get("event_a"), e.get("event_b")}
        == {SPORT_TYPE["BASKETBALL"], SPORT_TYPE["BADMINTON"]}
    ]
    assert len(edges) == 1
    edge = edges[0]
    assert edge["shared_count"] == 1
    # Shared athlete's primary is Basketball (one of the two events) → primary.
    assert edge["primary_overlap_count"] == 1
    # The racquet side references a stable entry id, also present as a game team.
    racquet_team = edge["team_a_id"] if "BAD-" in edge["team_a_id"] else edge["team_b_id"]
    assert racquet_team.startswith("BAD-Men-Doubles-E")
    assert any(
        racquet_team in (g.get("team_a_id"), g.get("team_b_id"))
        for g in si["games"]
    )


def test_pod_doubles_racquet_to_racquet_conflict_edge(tmp_path):
    """A player in Badminton doubles and Pickleball doubles → racquet↔racquet edge."""
    builder = ScheduleWorkbookBuilder()
    shared_id = "70"
    bad = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": shared_id,
         "First Name": "Sang", "Last Name": "X", "partner_name": "Tho X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": "71",
         "First Name": "Tho", "Last Name": "X", "partner_name": "Sang X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    pck = [
        {"sport_type": SPORT_TYPE["PICKLEBALL"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": shared_id,
         "First Name": "Sang", "Last Name": "X", "partner_name": "Vinh X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
        {"sport_type": SPORT_TYPE["PICKLEBALL"], "sport_gender": "Men",
         "sport_format": "Men Doubles", "Participant ID (WP)": "72",
         "First Name": "Vinh", "Last Name": "X", "partner_name": "Sang X",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["PICKLEBALL"]},
    ]
    si = builder._build_schedule_input(bad + pck, [], tmp_path / "no_venue.xlsx")

    edges = [
        e for e in si["team_conflicts"]
        if {e.get("event_a"), e.get("event_b")}
        == {SPORT_TYPE["BADMINTON"], SPORT_TYPE["PICKLEBALL"]}
    ]
    assert len(edges) == 1
    assert edges[0]["shared_count"] == 1


def test_pod_unresolved_doubles_reported_as_unprotected(tmp_path):
    """UnresolvedDoubles surface in pod_unprotected_entries (not silently dropped)."""
    builder = ScheduleWorkbookBuilder()
    roster = [
        {"sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Doubles", "Participant ID (WP)": "80",
         "First Name": "Mai", "Last Name": "X", "partner_name": "",
         "Church Team": "RPC", "participant_primary_sport": SPORT_TYPE["BADMINTON"]},
    ]
    si = builder._build_schedule_input(roster, [], tmp_path / "no_venue.xlsx")
    unprotected = si["pod_unprotected_entries"]
    assert len(unprotected) == 1
    assert unprotected[0]["division_id"] == "BAD-Women-Doubles"
    assert unprotected[0]["reason"] == "MissingPartner"
    # An unprotected entry produces no cross-sport conflict edge.
    assert not [
        e for e in si["team_conflicts"]
        if SPORT_TYPE["BADMINTON"] in {e.get("event_a"), e.get("event_b")}
    ]


def test_soccer_schedule_input_creates_soccer_field_games(tmp_path):
    """Soccer pool-assignment rows should create real Soccer Field games, not Gym Core games."""
    builder = ScheduleWorkbookBuilder()
    rows = _soccer_roster(["RPC", "OCB", "ANH"], n_per_church=4)
    si = builder._build_schedule_input(rows, [], tmp_path / "no_venue.xlsx")
    soccer_games = [g for g in si.get("games", []) if g.get("event") == SPORT_TYPE["SOCCER"]]
    assert len(soccer_games) == 3
    assert all(g["resource_type"] == TEAM_RESOURCE_TYPE_SOCCER for g in soccer_games)
    assert all(g["stage"] == "Pool" for g in soccer_games)
    assert all(g.get("team_c_id") in (None, "") for g in soccer_games)


def test_soccer_schedule_input_adds_playoff_precedence(tmp_path):
    """Six Soccer teams should add semis/final/3rd plus pool-before-semi precedence."""
    builder = ScheduleWorkbookBuilder()
    rows = _soccer_roster(["RPC", "OCB", "ANH", "GLA", "TLC", "FVC"], n_per_church=4)

    si = builder._build_schedule_input(rows, [], tmp_path / "no_venue.xlsx")
    soccer_games = [g for g in si["games"] if g.get("event") == SPORT_TYPE["SOCCER"]]
    pool_game_ids = {
        str(g["game_id"])
        for g in soccer_games
        if g.get("stage") == "Pool"
    }
    semi_ids = {
        str(g["game_id"])
        for g in soccer_games
        if g.get("stage") == "Semi"
    }
    final_ids = {
        str(g["game_id"])
        for g in soccer_games
        if g.get("stage") == "Final"
    }
    third_ids = {
        str(g["game_id"])
        for g in soccer_games
        if g.get("stage") == "3rd"
    }

    assert len(pool_game_ids) == 6
    assert semi_ids == {"SOC-Semi-1", "SOC-Semi-2"}
    assert final_ids == {"SOC-Final"}
    assert third_ids == {"SOC-3rd"}

    precedence_pairs = {
        (str(rule["before_game_id"]), str(rule["after_game_id"]))
        for rule in si["precedence"]
        if str(rule.get("before_game_id") or "").startswith("SOC-")
           or str(rule.get("after_game_id") or "").startswith("SOC-")
    }
    expected_pairs = {
        (pool_game_id, semi_id)
        for pool_game_id in pool_game_ids
        for semi_id in semi_ids
    } | {
        (semi_id, "SOC-Final")
        for semi_id in semi_ids
    } | {
        (semi_id, "SOC-3rd")
        for semi_id in semi_ids
    }
    assert precedence_pairs == expected_pairs


def test_soccer_disabled_removes_from_pool_assignment(monkeypatch):
    """When SOCCER_ENABLED=False, Soccer is excluded from Pool-Assignment outputs."""
    # Simulate disabled mode by patching the class attribute to the without-Soccer list.
    defs_without_soccer = [
        (event, prefix)
        for event, prefix in ScheduleWorkbookBuilder._POOL_ASSIGNMENT_EVENT_DEFS
        if event != SPORT_TYPE["SOCCER"]
    ]
    monkeypatch.setattr(
        ScheduleWorkbookBuilder,
        "_POOL_ASSIGNMENT_EVENT_DEFS",
        defs_without_soccer,
    )

    builder = ScheduleWorkbookBuilder()
    rows = _soccer_roster(["RPC", "OCB"], n_per_church=4)
    base_rows = builder._build_pool_assignment_base_rows(rows)
    soccer_rows = [r for r in base_rows if r["Event"] == SPORT_TYPE["SOCCER"]]
    assert soccer_rows == []
