import json

from openpyxl import Workbook

from config import SPORT_TYPE
from schedule_workbook import ScheduleWorkbookBuilder
from scheduling import manual_matchups


def _manual_roster_rows(event: str, teams: list[str]) -> list[dict]:
    if event == SPORT_TYPE["BASKETBALL"]:
        sport_type, gender = "Basketball", "Men"
    elif event == SPORT_TYPE["VOLLEYBALL_MEN"]:
        sport_type, gender = "Volleyball", "Men"
    elif event == SPORT_TYPE["VOLLEYBALL_WOMEN"]:
        sport_type, gender = "Volleyball", "Women"
    elif event == SPORT_TYPE["SOCCER"]:
        sport_type, gender = "Soccer - Coed Exhibition", "Mixed"
    else:
        sport_type, gender = "Bible Challenge", "Mixed"
    return [
        {
            "Church Team": team,
            "sport_type": sport_type,
            "sport_gender": gender,
            "sport_format": "Team",
            "Participant ID (WP)": f"{event}-{team}",
        }
        for team in teams
    ]


def _write_manual_sheet(
    ws,
    pools: list[tuple[str, list[tuple[int, str]]]],
    matches,
    *,
    team_count: int = 2,
):
    """Create the workbook shape the importer expects: pool table + >> rows."""
    if team_count == 2:
        for col_idx, value in enumerate(
            ["Slot A", "v", "Slot B", ">>", "Team A", "v", "Team B"],
            start=14,
        ):
            ws.cell(row=1, column=col_idx, value=value)
    else:
        for col_idx, value in enumerate(
            ["Slot A", "v", "Slot B", "v", "Slot C", ">>", "Team A", "v", "Team B", "v", "Team C"],
            start=14,
        ):
            ws.cell(row=1, column=col_idx, value=value)

    pool_row = 1
    for pool_label, members in pools:
        ws.cell(row=pool_row, column=5, value=f"POOL {pool_label}")
        for offset, (slot, team) in enumerate(members, start=1):
            ws.cell(row=pool_row + offset, column=5, value=team)
            ws.cell(row=pool_row + offset, column=6, value=slot)
        pool_row += len(members) + 2

    for row_idx, match in enumerate(matches, start=2):
        if team_count == 2:
            slot_a, slot_b, team_a, team_b = match
            ws.cell(row=row_idx, column=14, value=slot_a)
            ws.cell(row=row_idx, column=15, value="v")
            ws.cell(row=row_idx, column=16, value=slot_b)
            ws.cell(row=row_idx, column=17, value=">>")
            ws.cell(row=row_idx, column=18, value=team_a)
            ws.cell(row=row_idx, column=19, value="v")
            ws.cell(row=row_idx, column=20, value=team_b)
        else:
            slot_a, slot_b, slot_c, team_a, team_b, team_c = match
            ws.cell(row=row_idx, column=14, value=slot_a)
            ws.cell(row=row_idx, column=15, value="v")
            ws.cell(row=row_idx, column=16, value=slot_b)
            ws.cell(row=row_idx, column=17, value="v")
            ws.cell(row=row_idx, column=18, value=slot_c)
            ws.cell(row=row_idx, column=19, value=">>")
            ws.cell(row=row_idx, column=20, value=team_a)
            ws.cell(row=row_idx, column=21, value="v")
            ws.cell(row=row_idx, column=22, value=team_b)
            ws.cell(row=row_idx, column=23, value="v")
            ws.cell(row=row_idx, column=24, value=team_c)


def _create_2026_manual_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    bb = wb.create_sheet("BB_round2")
    _write_manual_sheet(
        bb,
        [
            ("A", [(1, "MWC"), (2, "RPC"), (3, "LBC"), (4, "GLA")]),
            ("B", [(5, "NHC"), (6, "ORN"), (7, "TLC"), (8, "GAC")]),
            ("C", [(9, "FVC"), (10, "WAG"), (11, "ANH"), (12, "OCB"), (13, "bye"), (14, "GEC")]),
        ],
        [
            (1, 4, "MWC", "GLA"),
            (2, 1, "RPC", "MWC"),
            (3, 2, "LBC", "RPC"),
            (4, 3, "GLA", "LBC"),
            (1, 3, "MWC", "LBC"),
            (2, 4, "RPC", "GLA"),
            (5, 8, "NHC", "GAC"),
            (6, 5, "ORN", "NHC"),
            (7, 6, "TLC", "ORN"),
            (8, 7, "GAC", "TLC"),
            (5, 7, "NHC", "TLC"),
            (6, 8, "ORN", "GAC"),
            (9, 14, "FVC", "GEC"),
            (10, 9, "WAG", "FVC"),
            (11, 10, "ANH", "WAG"),
            (12, 11, "OCB", "ANH"),
            (14, 12, "GEC", "OCB"),
            (9, 12, "FVC", "OCB"),
            (10, 13, "WAG", "bye"),
            (11, 14, "ANH", "GEC"),
        ],
    )

    mvb = wb.create_sheet("MVB")
    _write_manual_sheet(
        mvb,
        [
            ("A", [(1, "FVC"), (2, "PCC"), (3, "ORN"), (4, "RPC")]),
            ("B", [(5, "GLA"), (6, "SFV"), (7, "NSD"), (8, "GEC")]),
            ("C", [(9, "GAC"), (10, "SDC"), (11, "WSD"), (12, "ANH"), (13, "MWC"), (14, "TLC")]),
        ],
        [
            (1, 4, "FVC", "RPC"),
            (2, 1, "PCC", "FVC"),
            (3, 2, "ORN", "PCC"),
            (4, 3, "RPC", "ORN"),
            (1, 3, "FVC", "ORN"),
            (2, 4, "PCC", "RPC"),
            (5, 8, "GLA", "GEC"),
            (6, 5, "SFV", "GLA"),
            (7, 6, "NSD", "SFV"),
            (8, 7, "GEC", "NSD"),
            (5, 7, "GLA", "NSD"),
            (6, 8, "SFV", "GEC"),
            (9, 14, "GAC", "TLC"),
            (10, 9, "SDC", "GAC"),
            (11, 10, "WSD", "SDC"),
            (12, 11, "ANH", "WSD"),
            (13, 12, "MWC", "ANH"),
            (14, 13, "TLC", "MWC"),
            (9, 12, "GAC", "ANH"),
            (10, 13, "SDC", "MWC"),
            (11, 14, "WSD", "TLC"),
        ],
    )

    wvb = wb.create_sheet("WVB")
    _write_manual_sheet(
        wvb,
        [
            ("A", [(1, "RPC"), (2, "PCC"), (3, "GLA"), (4, "GAC"), (5, "SDC"), (6, "FVC"), (7, "NSD"), (8, "ORN")]),
            ("B", [(9, "ANH"), (10, "WAG"), (11, "MWC"), (12, "NHC"), (13, "WCC")]),
        ],
        [
            (1, 8, "RPC", "ORN"),
            (2, 1, "PCC", "RPC"),
            (3, 2, "GLA", "PCC"),
            (4, 3, "GAC", "GLA"),
            (5, 4, "SDC", "GAC"),
            (6, 5, "FVC", "SDC"),
            (7, 6, "NSD", "FVC"),
            (8, 7, "ORN", "NSD"),
            (1, 7, "RPC", "NSD"),
            (2, 8, "PCC", "ORN"),
            (3, 1, "GLA", "RPC"),
            (4, 2, "GAC", "PCC"),
            (5, 3, "SDC", "GLA"),
            (6, 4, "FVC", "GAC"),
            (7, 5, "NSD", "SDC"),
            (8, 6, "ORN", "FVC"),
            (9, 13, "ANH", "WCC"),
            (10, 9, "WAG", "ANH"),
            (11, 10, "MWC", "WAG"),
            (12, 11, "NHC", "MWC"),
            (13, 12, "WCC", "NHC"),
            (9, 12, "ANH", "NHC"),
            (10, 13, "WAG", "WCC"),
            (11, 9, "MWC", "ANH"),
            (12, 10, "NHC", "WAG"),
            (13, 11, "WCC", "MWC"),
        ],
    )

    soc = wb.create_sheet("SOC")
    _write_manual_sheet(
        soc,
        [
            ("A", [(1, "RPC"), (2, "FVC"), (3, "ANH"), (4, "TLC")]),
        ],
        [
            (1, 2, "RPC", "FVC"),
            (3, 4, "ANH", "TLC"),
            (1, 3, "RPC", "ANH"),
            (2, 4, "FVC", "TLC"),
        ],
    )

    bc = wb.create_sheet("BC")
    _write_manual_sheet(
        bc,
        [
            (
                "A",
                [
                    (1, "RPC"),
                    (2, "FVC"),
                    (3, "ANH"),
                    (4, "TLC"),
                    (5, "MWC"),
                    (6, "GAC"),
                    (7, "NHC"),
                    (8, "ORN"),
                    (9, "WAG"),
                ],
            ),
        ],
        [
            (1, 2, 3, "RPC", "FVC", "ANH"),
            (4, 5, 6, "TLC", "MWC", "GAC"),
            (7, 8, 9, "NHC", "ORN", "WAG"),
            (1, 4, 7, "RPC", "TLC", "NHC"),
            (2, 5, 8, "FVC", "MWC", "ORN"),
            (3, 6, 9, "ANH", "GAC", "WAG"),
            (1, 5, 9, "RPC", "MWC", "WAG"),
            (2, 6, 7, "FVC", "GAC", "NHC"),
            (3, 4, 8, "ANH", "TLC", "ORN"),
        ],
        team_count=3,
    )

    wb.create_sheet("BB (2)")
    wb.save(path)


def test_manual_matchup_import_accepts_2026_shapes(tmp_path):
    workbook = tmp_path / manual_matchups.MANUAL_TEAM_MATCHUP_WORKBOOK_FILENAME
    _create_2026_manual_workbook(workbook)
    roster_rows = []
    roster_rows += _manual_roster_rows(
        SPORT_TYPE["BASKETBALL"],
        ["MWC", "RPC", "LBC", "GLA", "NHC", "ORN", "TLC", "GAC", "FVC", "WAG", "ANH", "OCB", "GEC"],
    )
    roster_rows += _manual_roster_rows(
        SPORT_TYPE["VOLLEYBALL_MEN"],
        ["FVC", "PCC", "ORN", "RPC", "GLA", "SFV", "NSD", "GEC", "GAC", "SDC", "WSD", "ANH", "MWC", "TLC"],
    )
    roster_rows += _manual_roster_rows(
        SPORT_TYPE["VOLLEYBALL_WOMEN"],
        ["RPC", "PCC", "GLA", "GAC", "SDC", "FVC", "NSD", "ORN", "ANH", "WAG", "MWC", "NHC", "WCC"],
    )
    roster_rows += _manual_roster_rows(
        SPORT_TYPE["SOCCER"],
        ["RPC", "FVC", "ANH", "TLC"],
    )
    roster_rows += _manual_roster_rows(
        SPORT_TYPE["BIBLE_CHALLENGE"],
        ["RPC", "FVC", "ANH", "TLC", "MWC", "GAC", "NHC", "ORN", "WAG"],
    )

    payload = manual_matchups.build_manual_matchup_payload(
        workbook,
        roster_rows=roster_rows,
    )

    assert payload["validation"]["errors"] == []
    assert payload["validation"]["game_count"] == 79
    assert payload["validation"]["bye_count"] == 1
    by_event = {event["event"]: event for event in payload["events"]}
    assert len(by_event[SPORT_TYPE["BASKETBALL"]]["games"]) == 19
    assert len(by_event[SPORT_TYPE["VOLLEYBALL_MEN"]]["games"]) == 21
    assert len(by_event[SPORT_TYPE["VOLLEYBALL_WOMEN"]]["games"]) == 26
    assert len(by_event[SPORT_TYPE["SOCCER"]]["games"]) == 4
    assert len(by_event[SPORT_TYPE["BIBLE_CHALLENGE"]]["games"]) == 9
    assert by_event[SPORT_TYPE["VOLLEYBALL_WOMEN"]]["expected_games_per_team"] == 4
    assert by_event[SPORT_TYPE["VOLLEYBALL_WOMEN"]]["team_game_counts"]["RPC"] == 4
    assert by_event[SPORT_TYPE["SOCCER"]]["team_game_counts"]["RPC"] == 2
    assert by_event[SPORT_TYPE["BIBLE_CHALLENGE"]]["team_game_counts"]["RPC"] == 3
    assert by_event[SPORT_TYPE["BIBLE_CHALLENGE"]]["games"][0]["team_c_label"] == "ANH"
    assert by_event[SPORT_TYPE["BIBLE_CHALLENGE"]]["games"][0]["game_id"] == "BC-RR-1"
    assert by_event[SPORT_TYPE["BASKETBALL"]]["team_game_counts"]["WAG"] == 2
    assert by_event[SPORT_TYPE["BASKETBALL"]]["count_warnings"]


def test_schedule_input_uses_imported_manual_matchups(tmp_path):
    workbook = tmp_path / manual_matchups.MANUAL_TEAM_MATCHUP_WORKBOOK_FILENAME
    _create_2026_manual_workbook(workbook)
    payload = manual_matchups.build_manual_matchup_payload(workbook, roster_rows=[])
    sidecar = tmp_path / manual_matchups.MANUAL_TEAM_MATCHUP_SIDECAR_FILENAME
    manual_matchups.write_manual_matchup_sidecar(payload, sidecar)
    roster_rows = []
    for event, teams, participant_count in [
        (SPORT_TYPE["SOCCER"], ["RPC", "FVC", "ANH", "TLC"], 4),
        (
            SPORT_TYPE["BIBLE_CHALLENGE"],
            ["RPC", "FVC", "ANH", "TLC", "MWC", "GAC", "NHC", "ORN", "WAG"],
            3,
        ),
    ]:
        base_rows = _manual_roster_rows(event, teams)
        for row in base_rows:
            for idx in range(participant_count):
                participant_row = dict(row)
                participant_row["Participant ID (WP)"] = (
                    f"{row['Participant ID (WP)']}-{idx + 1}"
                )
                roster_rows.append(participant_row)

    schedule_input = ScheduleWorkbookBuilder()._build_schedule_input(
        roster_rows=roster_rows,
        validation_rows=[],
        venue_input_path=tmp_path / "missing_venue.xlsx",
        manual_matchup_path=sidecar,
    )

    basketball_pool_games = [
        game for game in schedule_input["games"]
        if game["event"] == SPORT_TYPE["BASKETBALL"] and game["stage"] == "Pool"
    ]
    wvb_pool_games = [
        game for game in schedule_input["games"]
        if game["event"] == SPORT_TYPE["VOLLEYBALL_WOMEN"] and game["stage"] == "Pool"
    ]
    soccer_pool_games = [
        game for game in schedule_input["games"]
        if game["event"] == SPORT_TYPE["SOCCER"] and game["stage"] == "Pool"
    ]
    bc_pool_games = [
        game for game in schedule_input["games"]
        if game["event"] == SPORT_TYPE["BIBLE_CHALLENGE"] and game["stage"] == "Pool"
    ]

    assert [game["team_a_label"] for game in basketball_pool_games[:3]] == ["MWC", "RPC", "LBC"]
    assert [game["team_b_label"] for game in basketball_pool_games[:3]] == ["GLA", "MWC", "RPC"]
    assert len(basketball_pool_games) == 19
    assert len(wvb_pool_games) == 26
    assert len(soccer_pool_games) == 4
    assert len(bc_pool_games) == 9
    assert bc_pool_games[0]["team_c_label"] == "ANH"
    assert all("bye" not in json.dumps(game).casefold() for game in basketball_pool_games)
    assert any(
        game["game_id"] == "BC-Final" and game["event"] == SPORT_TYPE["BIBLE_CHALLENGE"]
        for game in schedule_input["games"]
    )
    assert any(
        game["game_id"] == "SOC-Final" and game["event"] == SPORT_TYPE["SOCCER"]
        for game in schedule_input["games"]
    )
    assert schedule_input["manual_matchups"]["validation"]["game_count"] == 79


def test_manual_matchups_do_not_require_pool_map_when_match_list_is_complete(tmp_path):
    workbook = tmp_path / manual_matchups.MANUAL_TEAM_MATCHUP_WORKBOOK_FILENAME
    wb = Workbook()
    ws = wb.active
    ws.title = "SOC"
    _write_manual_sheet(
        ws,
        [],
        [
            (1, 2, "ANH", "GAC"),
        ],
    )
    bc = wb.create_sheet("BC")
    _write_manual_sheet(
        bc,
        [],
        [
            (1, 2, 3, "ANH", "GAC", "GLA"),
        ],
        team_count=3,
    )
    wb.save(workbook)

    payload = manual_matchups.build_manual_matchup_payload(
        workbook,
        active_sheets=["SOC", "BC"],
    )

    assert payload["validation"]["errors"] == []
    assert payload["validation"]["game_count"] == 2
    warning_text = "\n".join(payload["validation"]["warnings"])
    assert "no pool could be inferred" not in warning_text
    assert [event["games"][0]["pool_id"] for event in payload["events"]] == ["", ""]
