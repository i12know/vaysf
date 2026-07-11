from datetime import time

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from config import (
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
)
from scheduling import master_schedule


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _create_master_schedule_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "2026_Draft (4)"

    fills = {
        "Badminton": _fill("CCFFFF"),
        "Basketball": _fill("F4B084"),
        "MVB": _fill("9DC3E6"),
        "WVB": _fill("A9D18E"),
        "BC": _fill("FFD966"),
        "Table Tennis": _fill("BFBFBF"),
    }
    for coord, sport in {
        "B55": "Badminton",
        "F55": "Basketball",
        "J55": "MVB",
        "N55": "WVB",
        "R55": "BC",
        "V55": "Table Tennis",
    }.items():
        ws[coord] = sport
        ws[coord].fill = fills[sport]

    ws["A1"] = "SAT 7/18"
    ws["B1"] = "Main Gym BB1"
    ws["F1"] = "Main Gym BB2"
    ws["J1"] = "Prac. Gym VB1"
    ws["N1"] = "Prac. Gym VB2"
    ws["R1"] = "Prac. Gym VB3"
    ws["V1"] = "BC [LIBRARY]"

    ws["A5"] = time(12, 0)
    ws["K5"] = 1
    ws["K5"].fill = fills["MVB"]
    ws["O5"] = 1
    ws["O5"].fill = fills["WVB"]
    for coord, value in {"V5": "WSD", "W5": "ANH", "X5": "MWC"}.items():
        ws[coord] = value
        ws[coord].fill = fills["BC"]

    ws["A7"] = "13:59:59.998"
    ws["C7"] = 1
    ws["C7"].fill = fills["Basketball"]
    ws["G7"] = 20
    ws["G7"].fill = fills["Basketball"]

    ws["A9"] = "15:00"
    ws["C9"] = "BB QF"
    ws["C9"].fill = fills["Basketball"]
    ws["O9"] = "s"
    ws["O9"].fill = fills["WVB"]

    wb.save(path)


def _games():
    return [
        {
            "game_id": "BBM-01",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "duration_minutes": 45,
        },
        {
            "game_id": "VBM-01",
            "event": SPORT_TYPE["VOLLEYBALL_MEN"],
            "stage": "Pool",
            "duration_minutes": 45,
        },
        {
            "game_id": "VBW-01",
            "event": SPORT_TYPE["VOLLEYBALL_WOMEN"],
            "stage": "Pool",
            "duration_minutes": 45,
        },
        {
            "game_id": "BC-RR-7",
            "event": SPORT_TYPE["BIBLE_CHALLENGE"],
            "stage": "Pool",
            "team_a_label": "WSD",
            "team_b_label": "ANH",
            "team_c_label": "MWC",
            "duration_minutes": 60,
        },
        {
            "game_id": "BBM-QF-1",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "QF",
            "duration_minutes": 45,
        },
    ]


def _resources():
    resources = [
        {
            "resource_id": "BB-1",
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-1",
            "day": "Sat-1",
            "open_time": "14:00",
            "close_time": "16:00",
            "slot_minutes": 60,
        },
        {
            "resource_id": "BB-2",
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-2",
            "day": "Sat-1",
            "open_time": "14:00",
            "close_time": "16:00",
            "slot_minutes": 60,
        },
        {
            "resource_id": "BC-1",
            "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
            "label": "Station-1",
            "day": "Sat-1",
            "open_time": "12:00",
            "close_time": "13:00",
            "slot_minutes": 60,
        },
    ]
    for idx in range(1, 4):
        resources.append(
            {
                "resource_id": f"VB-{idx}",
                "resource_type": GYM_RESOURCE_TYPE_VOLLEYBALL,
                "label": f"Court-{idx}",
                "day": "Sat-1",
                "open_time": "12:00",
                "close_time": "13:00",
                "slot_minutes": 60,
            }
        )
    return resources


def test_master_schedule_parser_reads_visual_workbook(tmp_path):
    workbook = tmp_path / master_schedule.MASTER_SCHEDULE_WORKBOOK_FILENAME
    _create_master_schedule_workbook(workbook)

    payload = master_schedule.parse_master_schedule_workbook(workbook)
    rows_by_cell = {
        row.get("source_cell"): row
        for row in payload["rows"]
        if isinstance(row, dict)
    }

    assert rows_by_cell["C7"]["game_id"] == "BBM-01"
    assert rows_by_cell["C7"]["slot"] == "Sat-1-14:00"
    assert rows_by_cell["K5"]["game_id"] == "VBM-01"
    assert rows_by_cell["O5"]["game_id"] == "VBW-01"
    assert rows_by_cell["V5:X5"]["teams"] == ["WSD", "ANH", "MWC"]
    assert rows_by_cell["C9"]["game_id"] == "BBM-QF-1"
    assert rows_by_cell["O9"]["kind"] == "block"
    assert payload["diagnostics"]["block_count"] == 1


def test_master_schedule_resolves_confident_rows_and_reports_unknown_game(tmp_path):
    workbook = tmp_path / master_schedule.MASTER_SCHEDULE_WORKBOOK_FILENAME
    _create_master_schedule_workbook(workbook)
    payload = master_schedule.parse_master_schedule_workbook(workbook)

    fixed, diagnostics = master_schedule.resolve_master_schedule_payload(
        payload,
        _games(),
        _resources(),
    )
    fixed_by_game = {row["game_id"]: row for row in fixed}

    assert fixed_by_game["BBM-01"]["resource_id"] == "BB-1"
    assert fixed_by_game["BBM-01"]["slot"] == "Sat-1-14:00"
    assert fixed_by_game["VBM-01"]["resource_id"] == "VB-1"
    assert fixed_by_game["BC-RR-7"]["resource_id"] == "BC-1"
    assert fixed_by_game["BBM-QF-1"]["slot"] == "Sat-1-15:00"
    assert "BBM-20" not in fixed_by_game
    assert any("BBM-20" in warning for warning in diagnostics["warnings"])
    assert diagnostics["errors"] == []


def test_master_schedule_merge_overrides_existing_fixed_slot(tmp_path):
    workbook = tmp_path / master_schedule.MASTER_SCHEDULE_WORKBOOK_FILENAME
    _create_master_schedule_workbook(workbook)
    payload = master_schedule.parse_master_schedule_workbook(workbook)

    merged, summary = master_schedule.merge_master_schedule_into_playoff_slots(
        [
            {
                "game_id": "BBM-01",
                "event": SPORT_TYPE["BASKETBALL"],
                "stage": "Pool",
                "resource_id": "OLD",
                "slot": "Sat-1-09:00",
            }
        ],
        payload,
        _games(),
        _resources(),
    )
    by_game = {row["game_id"]: row for row in merged}

    assert by_game["BBM-01"]["resource_id"] == "BB-1"
    assert by_game["BBM-01"]["x_master_schedule_cell"] == "C7"
    assert summary["fixed_count"] == 5
    assert summary["unresolved_count"] == 1
