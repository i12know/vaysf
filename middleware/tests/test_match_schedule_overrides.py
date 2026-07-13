from datetime import time

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill

from config import (
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
    SPORT_FORMAT,
    SPORT_TYPE,
)
from scheduling import match_schedule_overrides as mso

# Real fill colors observed in the actual "2026 Main Schedule draft 11.xlsx"
# export -- deliberately NOT the schedule_styles.py canonical swatches, to
# exercise the nearest-color match rather than an exact-match shortcut.
_BB_FILL = PatternFill("solid", fgColor="F8CBAD")
_MVB_FILL = PatternFill("solid", fgColor="9BC2E6")
_WVB_FILL = PatternFill("solid", fgColor="FFCCFF")
_BC_FILL = PatternFill("solid", fgColor="92D050")
_BB_THEME_FILL = PatternFill("solid", fgColor=Color(theme=5, tint=0.5999938962981048))
_MVB_THEME_FILL = PatternFill("solid", fgColor=Color(theme=8, tint=0.3999755851924192))


def _venue_row(ws, row: int, day_label: str, venues: list):
    ws.cell(row=row, column=1, value=day_label)
    for start_col, name in venues:
        ws.cell(row=row, column=start_col, value=name)
        ws.merge_cells(
            start_row=row, start_column=start_col, end_row=row, end_column=start_col + 2
        )


def _game_row(ws, row: int, time_label: str, cells: dict):
    ws.cell(row=row, column=1, value=time_label)
    for col, (value, fill) in cells.items():
        cell = ws.cell(row=row, column=col, value=value)
        if fill is not None:
            cell.fill = fill


def _build_diagnostics_workbook(path):
    """One header block, then rows exercising unknown-code / duplicate /
    event-mismatch / bye scenarios. Column layout mirrors the real file:
    B-D=BB1, E-G=BB2, H-J=VB1, K-M=VB2, N-P=BC.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _venue_row(
        ws, 1, "SAT 7/18",
        [(2, "Main Gym BB1"), (5, "Main Gym BB2"), (8, "Prac. Gym VB1"),
         (11, "Prac. Gym VB2"), (14, "BC [LIBRARY]")],
    )

    _game_row(ws, 2, "12:00 PM", {
        2: ("WAG", _BB_FILL), 3: ("v", _BB_FILL), 4: ("FVC", _BB_FILL),
        5: ("SDC", _BB_FILL), 6: ("v", _BB_FILL), 7: ("MWC", _BB_FILL),
        8: ("GLA", _MVB_FILL), 9: ("v", _MVB_FILL), 10: ("PCC", _MVB_FILL),
        11: ("NSD", _WVB_FILL), 12: ("v", _WVB_FILL), 13: ("FVC", _WVB_FILL),
        14: ("AAA", _BC_FILL), 15: ("BBB", _BC_FILL), 16: ("CCC", _BC_FILL),
    })
    _game_row(ws, 3, "1:00 PM", {
        8: ("GLA", _MVB_FILL), 9: ("v", _MVB_FILL), 10: ("ZZZ", _MVB_FILL),
    })
    _game_row(ws, 4, "2:00 PM", {
        8: ("GLA", _MVB_FILL), 9: ("v", _MVB_FILL), 10: ("PCC", _MVB_FILL),
        11: ("NSD", _MVB_FILL), 12: ("v", _MVB_FILL), 13: ("GLA", _MVB_FILL),
    })
    _game_row(ws, 5, "3:00 PM", {
        11: ("GLA", _WVB_FILL), 12: ("v", _WVB_FILL), 13: ("FVC", _WVB_FILL),
    })
    _game_row(ws, 6, "4:00 PM", {
        2: ("WAG", _BB_FILL), 3: ("v", _BB_FILL), 4: ("bye", _BB_FILL),
    })

    wb.save(path)


def _build_clean_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    _venue_row(ws, 1, "SAT 7/18", [(2, "Main Gym BB1"), (5, "Prac. Gym VB1")])
    _game_row(ws, 2, "12:00 PM", {
        2: ("WAG", _BB_FILL), 3: ("v", _BB_FILL), 4: ("FVC", _BB_FILL),
        5: ("GLA", _MVB_FILL), 6: ("v", _MVB_FILL), 7: ("PCC", _MVB_FILL),
    })
    wb.save(path)


def _roster_rows():
    def rows(event, gender, codes):
        return [
            {
                "sport_type": event, "sport_gender": gender,
                "sport_format": SPORT_FORMAT["TEAM"], "Church Team": code,
            }
            for code in codes
        ]

    return (
        rows(SPORT_TYPE["BASKETBALL"], "Men", ["WAG", "FVC", "SDC", "MWC", "GLA", "PCC", "NSD"])
        + rows(SPORT_TYPE["VOLLEYBALL_MEN"], "Men", ["GLA", "PCC", "NSD", "FVC"])
        + rows(SPORT_TYPE["VOLLEYBALL_WOMEN"], "Women", ["NSD", "FVC", "MWC"])
    )


def test_parser_classifies_real_world_fill_colors(tmp_path):
    workbook = tmp_path / "diag.xlsx"
    _build_diagnostics_workbook(workbook)

    payload = mso.parse_match_schedule_overrides_workbook(workbook)
    two_team = {row["source_cell"]: row for row in payload["rows"] if row["kind"] == "two_team_game"}
    bc = [row for row in payload["rows"] if row["kind"] == "three_team_game"]
    bye = [row for row in payload["rows"] if row["kind"] == "bye"]

    assert two_team["B2:D2"]["sport"] == "Basketball"
    assert two_team["H2:J2"]["sport"] == "MVB"
    assert two_team["K2:M2"]["sport"] == "WVB"
    assert two_team["B2:D2"]["team_a"] == "WAG"
    assert two_team["B2:D2"]["team_b"] == "FVC"
    assert bc[0]["teams"] == ["AAA", "BBB", "CCC"]
    assert bye[0]["team_a"] == "WAG" and bye[0]["team_b"] is None


def test_parser_classifies_theme_fill_colors_from_draft_12_layout(tmp_path, monkeypatch):
    monkeypatch.setattr(mso, "_theme_rgb_by_index", lambda workbook: {5: "ED7D31", 8: "5B9BD5"})

    workbook = tmp_path / "draft12-theme.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "2026_Draft (12)"
    _venue_row(
        ws, 1, "SAT 7/18",
        [(2, "Main Gym BB1"), (10, "Prac. Gym VB1")],
    )
    ws.cell(row=2, column=1, value=time(13, 59, 59, 999000))
    for col, value in ((2, "WAG"), (3, "v"), (4, "FVC")):
        cell = ws.cell(row=2, column=col, value=value)
        cell.fill = _BB_THEME_FILL
    for col, value in ((10, "SDC"), (11, "v"), (12, "MWC")):
        cell = ws.cell(row=2, column=col, value=value)
        cell.fill = _MVB_THEME_FILL
    wb.save(workbook)

    payload = mso.parse_match_schedule_overrides_workbook(workbook)
    two_team = {row["source_cell"]: row for row in payload["rows"] if row["kind"] == "two_team_game"}

    assert payload["diagnostics"]["errors"] == []
    assert two_team["B2:D2"]["sport"] == "Basketball"
    assert two_team["J2:L2"]["sport"] == "MVB"
    assert two_team["B2:D2"]["slot"] == "Sat-1-14:00"
    assert two_team["J2:L2"]["slot"] == "Sat-1-14:00"


def test_parser_ignores_repeated_wide_control_blocks(tmp_path):
    workbook = tmp_path / "wide-controls.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "2026_Draft (12)"
    ws.cell(row=1, column=1, value="SAT 7/25")
    ws.cell(row=1, column=2, value="Main Gym BB1 / BB2")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=8)
    ws.cell(row=2, column=1, value="9:00 AM")
    ws.cell(row=2, column=2, value="BB QF")
    ws.cell(row=2, column=6, value="BB QF")
    ws.cell(row=3, column=1, value="2:00 PM")
    ws.cell(row=3, column=2, value="BADMINTON\nPLAYOFF / 3RD / FINAL\n[3 COURTS]")
    wb.save(workbook)

    payload = mso.parse_match_schedule_overrides_workbook(workbook)

    assert payload["rows"] == []
    assert payload["diagnostics"]["block_count"] == 0
    assert payload["diagnostics"]["warnings"] == []
    assert payload["diagnostics"]["unmapped_cells"] == []


def test_unknown_team_code_is_reported_with_cell_and_event(tmp_path):
    workbook = tmp_path / "diag.xlsx"
    _build_diagnostics_workbook(workbook)
    payload = mso.build_match_schedule_overrides_payload(
        workbook, roster_rows=_roster_rows(), games=[], resources=[],
    )
    errors = payload["validation"]["errors"]
    assert any("ZZZ" in error and "H3:J3" in error for error in errors)


def test_duplicate_team_in_same_slot_is_reported(tmp_path):
    workbook = tmp_path / "diag.xlsx"
    _build_diagnostics_workbook(workbook)
    payload = mso.build_match_schedule_overrides_payload(
        workbook, roster_rows=_roster_rows(), games=[], resources=[],
    )
    errors = payload["validation"]["errors"]
    assert any("GLA" in error and "double-booked" in error for error in errors)


def test_event_mismatch_names_the_actual_event(tmp_path):
    workbook = tmp_path / "diag.xlsx"
    _build_diagnostics_workbook(workbook)
    payload = mso.build_match_schedule_overrides_payload(
        workbook, roster_rows=_roster_rows(), games=[], resources=[],
    )
    errors = payload["validation"]["errors"]
    assert any(
        "GLA" in error and "K5:M5" in error and "Volleyball - Men Team instead" in error
        for error in errors
    )


def test_valid_import_pins_existing_game_and_creates_missing_one(tmp_path):
    workbook = tmp_path / "clean.xlsx"
    _build_clean_workbook(workbook)

    games = [{
        "game_id": "VBM-01",
        "event": SPORT_TYPE["VOLLEYBALL_MEN"],
        "stage": "Pool",
        "team_a_label": "GLA",
        "team_b_label": "PCC",
        "duration_minutes": 45,
    }]
    resources = [
        {
            "resource_id": "BB-1", "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-1", "day": "Sat-1", "open_time": "12:00",
            "close_time": "13:00", "slot_minutes": 60,
        },
        {
            "resource_id": "VB-1", "resource_type": GYM_RESOURCE_TYPE_VOLLEYBALL,
            "label": "Court-1", "day": "Sat-1", "open_time": "12:00",
            "close_time": "13:00", "slot_minutes": 60,
        },
    ]

    payload = mso.build_match_schedule_overrides_payload(
        workbook, roster_rows=_roster_rows(), games=games, resources=resources,
    )
    validation = payload["validation"]
    assert validation["errors"] == []
    assert validation["fixed_count"] == 2
    assert validation["created_game_count"] == 1

    slots_by_game = {row["game_id"]: row for row in validation["resolved_slots"]}
    assert slots_by_game["VBM-01"]["resource_id"] == "VB-1"
    assert any(
        game["team_a_label"] == "WAG" and game["team_b_label"] == "FVC"
        for game in validation["resolved_games"]
    )


def test_merge_pins_existing_game_and_appends_new_game(tmp_path):
    workbook = tmp_path / "clean.xlsx"
    _build_clean_workbook(workbook)
    payload = mso.build_match_schedule_overrides_payload(
        workbook, roster_rows=_roster_rows(), games=[], resources=[],
    )
    payload["events"] = list(mso.DEFAULT_EVENT_CODES)

    existing_games = [{
        "game_id": "VBM-01",
        "event": SPORT_TYPE["VOLLEYBALL_MEN"],
        "stage": "Pool",
        "team_a_label": "GLA",
        "team_b_label": "PCC",
        "duration_minutes": 45,
    }]
    resources = [
        {
            "resource_id": "BB-1", "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-1", "day": "Sat-1", "open_time": "12:00",
            "close_time": "13:00", "slot_minutes": 60,
        },
        {
            "resource_id": "VB-1", "resource_type": GYM_RESOURCE_TYPE_VOLLEYBALL,
            "label": "Court-1", "day": "Sat-1", "open_time": "12:00",
            "close_time": "13:00", "slot_minutes": 60,
        },
    ]

    merged_games, merged_slots, summary = mso.merge_match_schedule_overrides_into_schedule_input(
        existing_games, [], payload, resources,
    )

    assert len(merged_games) == 2  # VBM-01 preserved, one new BB game appended
    assert any(game["game_id"] == "VBM-01" for game in merged_games)
    new_game = next(game for game in merged_games if game["game_id"] != "VBM-01")
    assert new_game["team_a_label"] == "WAG" and new_game["team_b_label"] == "FVC"

    slots_by_game = {row["game_id"]: row for row in merged_slots}
    assert slots_by_game["VBM-01"]["resource_id"] == "VB-1"
    assert slots_by_game[new_game["game_id"]]["resource_id"] == "BB-1"
    assert summary["fixed_count"] == 2


def test_merge_reports_existing_fixed_slot_conflict():
    payload = {
        "events": ["BB"],
        "source_workbook": "unit.xlsx",
        "rows": [{
            "kind": "two_team_game",
            "sport": "Basketball",
            "event": SPORT_TYPE["BASKETBALL"],
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "team_a": "WAG",
            "team_b": "FVC",
            "day": "Sat-1",
            "slot": "Sat-1-12:00",
            "visual_venue": "Main Gym BB1",
            "source_cell": "B2:D2",
            "source_sheet": "Sheet1",
            "raw_value": "WAG / v / FVC",
        }],
    }
    games = [
        {
            "game_id": "BBM-01",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "team_a_label": "WAG",
            "team_b_label": "FVC",
            "duration_minutes": 60,
        },
        {
            "game_id": "BBM-02",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "team_a_label": "SDC",
            "team_b_label": "MWC",
            "duration_minutes": 60,
        },
    ]
    existing_slots = [{
        "game_id": "BBM-02",
        "event": SPORT_TYPE["BASKETBALL"],
        "resource_id": "BB-1",
        "slot": "Sat-1-12:00",
    }]
    resources = [{
        "resource_id": "BB-1",
        "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
        "label": "Court-1",
        "day": "Sat-1",
        "open_time": "12:00",
        "close_time": "13:00",
        "slot_minutes": 60,
    }]

    _merged_games, merged_slots, summary = mso.merge_match_schedule_overrides_into_schedule_input(
        games, existing_slots, payload, resources,
    )

    assert any("conflicts with existing fixed assignment BBM-02" in error for error in summary["errors"])
    assert {slot["game_id"] for slot in merged_slots} == {"BBM-02"}


def test_merge_supersedes_numbered_master_pool_pin_for_team_code_schedule():
    payload = {
        "events": ["BB"],
        "source_workbook": "unit.xlsx",
        "rows": [{
            "kind": "two_team_game",
            "sport": "Basketball",
            "event": SPORT_TYPE["BASKETBALL"],
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "team_a": "WAG",
            "team_b": "FVC",
            "day": "Sat-1",
            "slot": "Sat-1-12:00",
            "visual_venue": "Main Gym BB1",
            "source_cell": "B2:D2",
            "source_sheet": "Sheet1",
            "raw_value": "WAG / v / FVC",
        }],
    }
    games = [
        {
            "game_id": "BBM-01",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "team_a_label": "WAG",
            "team_b_label": "FVC",
            "duration_minutes": 60,
        },
        {
            "game_id": "BBM-02",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "team_a_label": "SDC",
            "team_b_label": "MWC",
            "duration_minutes": 60,
        },
    ]
    existing_slots = [
        {
            "game_id": "BBM-01",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "resource_id": "BB-2",
            "slot": "Sat-1-13:00",
            "x_master_schedule_cell": "E3",
        },
        {
            "game_id": "BBM-02",
            "event": SPORT_TYPE["BASKETBALL"],
            "stage": "Pool",
            "resource_id": "BB-1",
            "slot": "Sat-1-12:00",
            "x_master_schedule_cell": "B2",
        },
    ]
    resources = [
        {
            "resource_id": "BB-1",
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-1",
            "day": "Sat-1",
            "open_time": "12:00",
            "close_time": "13:00",
            "slot_minutes": 60,
        },
        {
            "resource_id": "BB-2",
            "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
            "label": "Court-2",
            "day": "Sat-1",
            "open_time": "13:00",
            "close_time": "14:00",
            "slot_minutes": 60,
        },
    ]

    _merged_games, merged_slots, summary = mso.merge_match_schedule_overrides_into_schedule_input(
        games, existing_slots, payload, resources,
    )

    assert summary["errors"] == []
    assert any("superseded numbered master-schedule pool pin BBM-02" in warning for warning in summary["warnings"])
    assert {slot["game_id"] for slot in merged_slots} == {"BBM-01"}
    assert merged_slots[0]["resource_id"] == "BB-1"
    assert merged_slots[0]["slot"] == "Sat-1-12:00"
