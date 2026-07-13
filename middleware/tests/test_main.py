import argparse
import datetime as dt
import json

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import main
from schedule_workbook import ScheduleWorkbookBuilder


def _run_main_expect_exit(expected_code: int) -> None:
    with pytest.raises(SystemExit) as exc:
        main.main()
    assert exc.value.code == expected_code


def _status_banner_values(ws) -> set[str]:
    return {
        str(cell.value)
        for row in ws.iter_rows(min_row=1, max_row=2)
        for cell in row
        if cell.value is not None
    }


def _sheet_texts(ws) -> list[str]:
    return [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _minimal_schedule_input() -> dict:
    return {
        "generated_at": "2026-05-15T00:00:00",
        "gym_court_scenario": 4,
        "game_count": 2,
        "resource_count": 1,
        "games": [
            {
                "game_id": "BBM-01",
                "event": "Basketball - Men Team",
                "stage": "Pool",
                "pool_id": "P1",
                "round": 1,
                "team_a_id": "BBM-P1-T1",
                "team_b_id": "BBM-P1-T2",
                "duration_minutes": 60,
                "resource_type": "Basketball Court",
                "earliest_slot": None,
                "latest_slot": None,
            },
            {
                "game_id": "BBM-02",
                "event": "Basketball - Men Team",
                "stage": "Pool",
                "pool_id": "P1",
                "round": 2,
                "team_a_id": "BBM-P1-T3",
                "team_b_id": "BBM-P1-T4",
                "duration_minutes": 60,
                "resource_type": "Basketball Court",
                "earliest_slot": None,
                "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "GYM-Sat-1-1",
                "resource_type": "Basketball Court",
                "label": "Court-1",
                "day": "Sat-1",
                "open_time": "08:00",
                "close_time": "10:00",
                "slot_minutes": 60,
                "exclusive_group": "",
            }
        ],
        "playoff_slots": [],
        "gym_modes": {},
    }


def test_parse_args_solve_schedule_defaults(monkeypatch):
    monkeypatch.setattr(main.sys, "argv", ["main.py", "solve-schedule"])
    args = main.parse_args()
    assert args.command == "solve-schedule"
    assert args.input is None
    assert args.output is None


def test_parse_args_diagnose_schedule_defaults(monkeypatch):
    monkeypatch.setattr(main.sys, "argv", ["main.py", "diagnose-schedule"])
    args = main.parse_args()
    assert args.command == "diagnose-schedule"
    assert args.input is None
    assert args.schedule_output is None
    assert args.output is None


def test_parse_args_produce_schedule_aliases(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "produce-schedule",
            "--input",
            "out.json",
            "--constraint",
            "in.json",
            "--output",
            "schedule.xlsx",
        ],
    )
    args = main.parse_args()
    assert args.command == "produce-schedule"
    assert args.schedule_output == "out.json"
    assert args.schedule_input == "in.json"
    assert args.output == "schedule.xlsx"


def test_parse_args_build_schedule_workbook_defaults(monkeypatch):
    monkeypatch.setattr(main.sys, "argv", ["main.py", "build-schedule-workbook"])
    args = main.parse_args()
    assert args.command == "build-schedule-workbook"
    assert args.input_json is None
    assert args.input_xlsx is None
    assert args.output is None
    assert args.pool_assignments is None


def test_parse_args_assign_pools_defaults(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        ["main.py", "assign-pools", "--workbook", "Schedule_Workbook.xlsx"],
    )
    args = main.parse_args()
    assert args.command == "assign-pools"
    assert args.workbook == "Schedule_Workbook.xlsx"
    assert args.output is None
    assert args.pool_assignments is None


def test_parse_args_import_team_matchups_file_alias(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        ["main.py", "import-team-matchups", "--file", "Manual_Matchups.xlsx"],
    )
    args = main.parse_args()
    assert args.command == "import-team-matchups"
    assert args.workbook == "Manual_Matchups.xlsx"


def test_parse_args_import_master_schedule_file_alias(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "import-master-schedule",
            "--file",
            "VAY2026_Main_Schedule_draft_4.xlsx",
            "--schedule-input",
            "schedule_input.json",
        ],
    )
    args = main.parse_args()
    assert args.command == "import-master-schedule"
    assert args.workbook == "VAY2026_Main_Schedule_draft_4.xlsx"
    assert args.schedule_input == "schedule_input.json"
    assert args.output is None


def test_parse_args_import_match_schedule_overrides_dry_run(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "import-match-schedule-overrides",
            "--file",
            "2026 Main Schedule draft 11.xlsx",
            "--events",
            "BB,MVB,WVB",
            "--dry-run",
        ],
    )
    args = main.parse_args()
    assert args.command == "import-match-schedule-overrides"
    assert args.workbook == "2026 Main Schedule draft 11.xlsx"
    assert args.events == "BB,MVB,WVB"
    assert args.dry_run is True
    assert args.execute is False


def test_parse_args_import_match_schedule_overrides_requires_mode(monkeypatch, capsys):
    monkeypatch.setattr(
        main.sys, "argv", ["main.py", "import-match-schedule-overrides"]
    )
    with pytest.raises(SystemExit) as exc:
        main.parse_args()
    assert exc.value.code == 2


def test_parse_args_import_approved_games_execute(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "import-approved-games",
            "--main-schedule",
            "main.xlsx",
            "--badminton",
            "badminton.xlsx",
            "--soccer",
            "soccer.xlsx",
            "--table-tennis",
            "tt.xlsx",
            "--schedule-input",
            "schedule_input.json",
            "--input-xlsx",
            "Church_Team_Status_ALL.xlsx",
            "--execute",
        ],
    )
    args = main.parse_args()
    assert args.command == "import-approved-games"
    assert args.main_schedule == "main.xlsx"
    assert args.badminton == "badminton.xlsx"
    assert args.soccer == "soccer.xlsx"
    assert args.table_tennis == "tt.xlsx"
    assert args.schedule_input == "schedule_input.json"
    assert args.input_xlsx == "Church_Team_Status_ALL.xlsx"
    assert args.execute is True
    assert args.dry_run is False


def test_parse_args_upload_person_photo_execute(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "upload-person-photo",
            "--chm-id",
            "999001",
            "--photo-file",
            "athlete.png",
            "--execute",
        ],
    )
    args = main.parse_args()
    assert args.command == "upload-person-photo"
    assert args.chm_id == "999001"
    assert args.photo_file == "athlete.png"
    assert args.execute is True
    assert args.dry_run is False


def test_parse_args_upload_person_photo_url_dry_run(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "upload-person-photo",
            "--chm-id",
            "999001",
            "--photo-url",
            "https://cdne-chmeetings-content.azureedge.net/images/photo.jpg",
            "--dry-run",
        ],
    )
    args = main.parse_args()
    assert args.command == "upload-person-photo"
    assert args.chm_id == "999001"
    assert args.photo_url.startswith("https://")
    assert args.photo_file is None
    assert args.dry_run is True


def test_parse_args_publish_schedule_requires_mode(monkeypatch, capsys):
    monkeypatch.setattr(main.sys, "argv", ["main.py", "publish-schedule"])
    with pytest.raises(SystemExit) as exc:
        main.parse_args()
    assert exc.value.code == 2


def test_parse_args_publish_schedule_rejects_both_modes(monkeypatch, capsys):
    monkeypatch.setattr(
        main.sys, "argv", ["main.py", "publish-schedule", "--dry-run", "--execute"]
    )
    with pytest.raises(SystemExit) as exc:
        main.parse_args()
    assert exc.value.code == 2


def test_parse_args_publish_schedule_dry_run_defaults(monkeypatch):
    monkeypatch.setattr(main.sys, "argv", ["main.py", "publish-schedule", "--dry-run"])
    args = main.parse_args()
    assert args.command == "publish-schedule"
    assert args.dry_run is True
    assert args.execute is False
    assert args.force_cancel is False
    assert args.input is None
    assert args.schedule_output is None
    assert args.output is None


def test_parse_args_publish_schedule_execute_force_cancel(monkeypatch):
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "publish-schedule",
            "--execute",
            "--force-cancel",
            "--input",
            "schedule_input.json",
            "--schedule-output",
            "schedule_output.json",
            "--output",
            "audit.json",
        ],
    )
    args = main.parse_args()
    assert args.execute is True
    assert args.force_cancel is True
    assert args.input == "schedule_input.json"
    assert args.schedule_output == "schedule_output.json"
    assert args.output == "audit.json"


def test_main_publish_schedule_force_cancel_without_execute_exits_1(monkeypatch):
    # --force-cancel is independent of the --dry-run/--execute mutex, so
    # "--force-cancel --dry-run" parses fine at the argparse level; main()
    # itself enforces "force-cancel requires --execute" before touching
    # WordPress at all (no connector needed for this test).
    monkeypatch.setattr(
        main.sys,
        "argv",
        ["main.py", "publish-schedule", "--force-cancel", "--dry-run"],
    )
    _run_main_expect_exit(1)


class _FakeWordPressConnectorForPublish:
    """Stand-in for WordPressConnector used by publish-schedule dispatch tests."""

    def __init__(self):
        self.get_schedules_calls = 0
        self.upsert_calls = []

    def get_schedules(self, params=None):
        self.get_schedules_calls += 1
        return []

    def upsert_schedules(self, games, schedule_version, force_cancel=False):
        self.upsert_calls.append(
            {"games": games, "schedule_version": schedule_version, "force_cancel": force_cancel}
        )
        return {
            "success": True,
            "schedule_version": schedule_version,
            "created_count": len(games),
            "updated_count": 0,
            "skipped_count": 0,
            "results": [],
        }

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        return False


def _write_publish_schedule_fixtures(tmp_path):
    schedule_input = _minimal_schedule_input()
    schedule_input_path = tmp_path / "schedule_input.json"
    schedule_input_path.write_text(json.dumps(schedule_input), encoding="utf-8")

    schedule_output = {
        "solved_at": "2026-07-11T00:00:00",
        "status": "OPTIMAL",
        "assignments": [
            {"game_id": "BBM-01", "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-08:00"},
            {"game_id": "BBM-02", "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-09:00"},
        ],
        "unscheduled": [],
    }
    schedule_output_path = tmp_path / "schedule_output.json"
    schedule_output_path.write_text(json.dumps(schedule_output), encoding="utf-8")
    return schedule_input_path, schedule_output_path


def test_main_publish_schedule_dry_run_never_upserts(monkeypatch, tmp_path):
    schedule_input_path, schedule_output_path = _write_publish_schedule_fixtures(tmp_path)
    fake_connector = _FakeWordPressConnectorForPublish()
    monkeypatch.setattr(main, "WordPressConnector", lambda: fake_connector)
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "publish-schedule",
            "--dry-run",
            "--input",
            str(schedule_input_path),
            "--schedule-output",
            str(schedule_output_path),
        ],
    )
    _run_main_expect_exit(0)
    assert fake_connector.get_schedules_calls == 1
    assert fake_connector.upsert_calls == []


def test_main_publish_schedule_execute_upserts_once(monkeypatch, tmp_path):
    schedule_input_path, schedule_output_path = _write_publish_schedule_fixtures(tmp_path)
    fake_connector = _FakeWordPressConnectorForPublish()
    monkeypatch.setattr(main, "WordPressConnector", lambda: fake_connector)
    monkeypatch.setattr(
        main.sys,
        "argv",
        [
            "main.py",
            "publish-schedule",
            "--execute",
            "--input",
            str(schedule_input_path),
            "--schedule-output",
            str(schedule_output_path),
        ],
    )
    _run_main_expect_exit(0)
    assert fake_connector.get_schedules_calls == 1
    assert len(fake_connector.upsert_calls) == 1
    upserted_keys = {game["game_key"] for game in fake_connector.upsert_calls[0]["games"]}
    assert upserted_keys == {"BBM-01", "BBM-02"}


def test_main_build_schedule_workbook_writes_xlsx(monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "export"
    data_dir.mkdir()
    export_dir.mkdir()
    monkeypatch.setattr(main, "DATA_DIR", data_dir)
    monkeypatch.setattr(main, "EXPORT_DIR", export_dir)

    schedule_input_path = export_dir / "schedule_input.json"
    schedule_input_path.write_text(
        json.dumps(_minimal_schedule_input()), encoding="utf-8"
    )

    class FakeDate(dt.date):
        @classmethod
        def today(cls):
            return cls(2026, 5, 15)

    monkeypatch.setattr(main.datetime, "date", FakeDate)
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="build-schedule-workbook",
            input_json=None,
            input_xlsx=None,
            output=None,
            pool_assignments=None,
        ),
    )

    _run_main_expect_exit(0)

    out_path = export_dir / "Schedule_Workbook_2026-05-15.xlsx"
    assert out_path.exists()
    wb = load_workbook(out_path)
    assert "Schedule-Input" in wb.sheetnames


def test_main_build_schedule_workbook_auto_detects_latest_all_workbook(monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "export"
    data_dir.mkdir()
    export_dir.mkdir()
    monkeypatch.setattr(main, "DATA_DIR", data_dir)
    monkeypatch.setattr(main, "EXPORT_DIR", export_dir)

    schedule_input_path = export_dir / "schedule_input.json"
    schedule_input_path.write_text(
        json.dumps(_minimal_schedule_input()), encoding="utf-8"
    )

    workbook_path = export_dir / "Church_Team_Status_ALL_2026-05-15.xlsx"
    wb = Workbook()
    roster_ws = wb.active
    roster_ws.title = "Roster"
    roster_headers = [
        "Church Team", "sport_type", "sport_gender", "sport_format",
        "Participant ID (WP)", "First Name", "Last Name", "partner_name",
    ]
    roster_ws.append(roster_headers)
    for idx in range(1, 6):
        roster_ws.append(
            ["RPC", "Basketball", "Men", "Team", idx, f"P{idx}", "Nguyen", None]
        )
    validation_ws = wb.create_sheet("Validation-Issues")
    validation_ws.append(["Status", "Severity", "Participant ID (WP)", "sport_type"])
    wb.save(workbook_path)

    class FakeDate(dt.date):
        @classmethod
        def today(cls):
            return cls(2026, 5, 15)

    monkeypatch.setattr(main.datetime, "date", FakeDate)
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="build-schedule-workbook",
            input_json=None,
            input_xlsx=None,
            output=None,
            pool_assignments=None,
        ),
    )

    _run_main_expect_exit(0)

    out_path = export_dir / "Schedule_Workbook_2026-05-15.xlsx"
    assert out_path.exists()
    venue_ws = load_workbook(out_path)["Venue-Estimator"]
    assert venue_ws["B2"].value == 1
    assert venue_ws["C2"].value == 1
    assert venue_ws["D2"].value == "RPC"


def test_main_build_schedule_workbook_prefers_input_xlsx_sibling_json(monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "export"
    custom_dir = tmp_path / "custom"
    data_dir.mkdir()
    export_dir.mkdir()
    custom_dir.mkdir()
    monkeypatch.setattr(main, "DATA_DIR", data_dir)
    monkeypatch.setattr(main, "EXPORT_DIR", export_dir)

    workbook_path = custom_dir / "Church_Team_Status_ALL_2026-05-15.xlsx"
    wb = Workbook()
    wb.save(workbook_path)

    sibling_json = custom_dir / "schedule_input.json"
    sibling_json.write_text(json.dumps(_minimal_schedule_input()), encoding="utf-8")

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="build-schedule-workbook",
            input_json=None,
            input_xlsx=str(workbook_path),
            output=str(custom_dir / "Schedule_Workbook.xlsx"),
            pool_assignments=None,
        ),
    )

    _run_main_expect_exit(0)

    assert (custom_dir / "Schedule_Workbook.xlsx").exists()


def test_main_build_schedule_workbook_missing_input_fails(monkeypatch, tmp_path):
    monkeypatch.setattr(main, "DATA_DIR", tmp_path)
    monkeypatch.setattr(main, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="build-schedule-workbook",
            input_json=str(tmp_path / "does_not_exist.json"),
            input_xlsx=None,
            output=None,
            pool_assignments=None,
        ),
    )

    _run_main_expect_exit(1)


def test_main_assign_pools_writes_sidecar(monkeypatch, tmp_path):
    workbook_path = tmp_path / "Schedule_Workbook.xlsx"
    schedule_input_path = tmp_path / "schedule_input.json"
    builder = ScheduleWorkbookBuilder()
    roster_rows = []
    for code in ("RPC", "ANH", "TLC", "FVC"):
        for idx in range(5):
            roster_rows.append(
                {
                    "Church Team": code,
                    "sport_type": "Basketball",
                    "sport_gender": "Men",
                    "sport_format": "Team",
                    "Participant ID (WP)": idx + 1,
                }
            )
    schedule_input_path.write_text(
        json.dumps(_minimal_schedule_input()),
        encoding="utf-8",
    )
    builder.write_schedule_workbook(
        workbook_path,
        roster_rows,
        [],
        json.loads(schedule_input_path.read_text(encoding="utf-8")),
        tmp_path / "missing.xlsx",
        pool_assignment_path=tmp_path / "pool_assignments.json",
    )

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="assign-pools",
            workbook=str(workbook_path),
            output=None,
            pool_assignments=None,
        ),
    )

    _run_main_expect_exit(0)

    assert (tmp_path / "pool_assignments.json").exists()


def test_generate_venue_template_example_dates_match_day_labels(tmp_path):
    out_path = tmp_path / "venue_template.xlsx"

    assert main.generate_venue_template(out_path) is True

    wb = load_workbook(out_path)
    ws = wb["Venue-Input"]
    assert ws["E2"].value == "Sat-1"
    assert ws["F2"].value == "2026-07-18"
    assert ws["E4"].value == "Sun-1"
    assert ws["F4"].value == "2026-07-19"
    assert "STATUS: EDITABLE INPUT" in _status_banner_values(wb["Venue-Input"])
    assert "STATUS: EDITABLE INPUT" in _status_banner_values(wb["Gym-Modes"])
    assert "STATUS: EDITABLE OVERRIDE INPUT" in _status_banner_values(wb["Playoff-Slots"])


def test_main_solve_schedule_uses_default_paths(mocker, monkeypatch, tmp_path):
    monkeypatch.setattr(main, "DATA_DIR", tmp_path)
    mock_run = mocker.patch("scheduler.run_solve_schedule", return_value=7)
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(command="solve-schedule", input=None, output=None),
    )

    _run_main_expect_exit(7)

    mock_run.assert_called_once_with(
        tmp_path / "schedule_input.json",
        tmp_path / "schedule_output.json",
    )


def test_main_diagnose_schedule_uses_default_paths(mocker, monkeypatch, tmp_path):
    monkeypatch.setattr(main, "EXPORT_DIR", tmp_path)
    (tmp_path / "schedule_output.json").write_text("{}", encoding="utf-8")
    mock_run = mocker.patch("schedule_diagnostics.run_diagnose_schedule", return_value=0)
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="diagnose-schedule",
            input=None,
            schedule_output=None,
            output=None,
        ),
    )

    _run_main_expect_exit(0)

    mock_run.assert_called_once_with(
        tmp_path / "schedule_input.json",
        schedule_output_path=tmp_path / "schedule_output.json",
        output_path=None,
    )


def test_main_produce_schedule_uses_default_paths(mocker, monkeypatch, tmp_path):
    monkeypatch.setattr(main, "DATA_DIR", tmp_path)
    monkeypatch.setattr(main, "EXPORT_DIR", tmp_path)

    schedule_output_path = tmp_path / "schedule_output.json"
    schedule_input_path = tmp_path / "schedule_input.json"
    schedule_output_path.write_text(
        json.dumps(
            {
                "solved_at": "2026-05-15T12:00:00",
                "status": "OPTIMAL",
                "solver_wall_seconds": 0.1,
                "assignments": [],
                "unscheduled": [],
            }
        ),
        encoding="utf-8",
    )
    schedule_input_path.write_text(
        json.dumps({"games": [], "resources": [], "playoff_slots": []}),
        encoding="utf-8",
    )

    class FakeDate(dt.date):
        @classmethod
        def today(cls):
            return cls(2026, 5, 15)

    monkeypatch.setattr(main.datetime, "date", FakeDate)
    mock_write = mocker.patch.object(
        ScheduleWorkbookBuilder,
        "write_schedule_output_workbook",
    )
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="produce-schedule",
            schedule_output=None,
            schedule_input=None,
            output=None,
        ),
    )

    _run_main_expect_exit(0)

    mock_write.assert_called_once()
    out_path, so_data, si_data = mock_write.call_args.args
    assert out_path == tmp_path / "VAYSF_Schedule_2026-05-15.xlsx"
    assert so_data["status"] == "OPTIMAL"
    assert si_data["games"] == []


def test_main_produce_schedule_malformed_json_fails_controlled(
    mocker, monkeypatch, tmp_path
):
    """A damaged event-week JSON file must exit 1 via the contract-error
    path, not raise an uncaught JSONDecodeError (#161 review finding 3)."""
    monkeypatch.setattr(main, "DATA_DIR", tmp_path)
    monkeypatch.setattr(main, "EXPORT_DIR", tmp_path)

    (tmp_path / "schedule_output.json").write_text(
        '{"status": "OPTIMAL", "assignments": [',  # truncated mid-write
        encoding="utf-8",
    )
    (tmp_path / "schedule_input.json").write_text(
        json.dumps({"games": [], "resources": []}), encoding="utf-8"
    )

    mock_write = mocker.patch.object(
        ScheduleWorkbookBuilder, "write_schedule_output_workbook"
    )
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="produce-schedule",
            schedule_output=None,
            schedule_input=None,
            output=None,
        ),
    )

    _run_main_expect_exit(1)
    mock_write.assert_not_called()


def test_main_produce_schedule_preserves_vietnamese_diacritics(
    monkeypatch, tmp_path
):
    """UTF-8 JSON names survive produce-schedule in every operator-facing tab."""
    label_a = "Hội Thánh"
    label_b = "Tin Lành"
    participant_names = "Nguyễn Văn Đức; Trần Thị Mỹ Linh"
    input_path = tmp_path / "schedule_input.json"
    output_path = tmp_path / "schedule_output.json"
    workbook_path = tmp_path / "schedule.xlsx"

    schedule_input = {
        "games": [{
            "game_id": "BBM-01",
            "event": "Basketball - Men Team",
            "stage": "Pool",
            "pool_id": "P1",
            "round": 1,
            "team_a_id": "BBM-P1-T1",
            "team_b_id": "BBM-P1-T2",
            "team_a_label": label_a,
            "team_b_label": label_b,
            "duration_minutes": 60,
            "resource_type": "Gym Court",
            "earliest_slot": None,
            "latest_slot": None,
        }],
        "resources": [{
            "resource_id": "GYM-Sat-1-1",
            "resource_type": "Gym Court",
            "label": "Court-1",
            "day": "Sat-1",
            "open_time": "08:00",
            "close_time": "09:00",
            "slot_minutes": 60,
            "exclusive_group": "",
        }],
        "playoff_slots": [],
    }
    schedule_output = {
        "solved_at": "2026-05-01T10:00:00+00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [{
            "game_id": "BBM-01",
            "resource_id": "GYM-Sat-1-1",
            "slot": "Sat-1-08:00",
        }],
        "unscheduled": [],
        "pool_results": [],
        "conflict_audit_summary": {
            "total_edges": 1,
            "separated_edges": 1,
            "overlapping_edges": 0,
            "incomplete_edges": 0,
            "remaining_primary_overlap_penalty": 0,
            "remaining_secondary_overlap_penalty": 0,
        },
        "conflict_audit": [{
            "team_a_label": label_a,
            "event_a": "Basketball - Men Team",
            "team_b_label": label_b,
            "event_b": "Volleyball - Men Team",
            "shared_count": 2,
            "primary_overlap_count": 2,
            "secondary_only_count": 0,
            "status": "SeparatedInSchedule",
            "overlap_count": 0,
            "scheduled_team_a_games": 1,
            "scheduled_team_b_games": 1,
            "shared_participant_names": participant_names,
            "overlap_game_pairs": "",
        }],
    }
    input_path.write_text(
        json.dumps(schedule_input, ensure_ascii=False), encoding="utf-8"
    )
    output_path.write_text(
        json.dumps(schedule_output, ensure_ascii=False), encoding="utf-8"
    )
    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="produce-schedule",
            schedule_output=str(output_path),
            schedule_input=str(input_path),
            output=str(workbook_path),
        ),
    )

    _run_main_expect_exit(0)

    wb = load_workbook(workbook_path)
    by_time = _sheet_texts(wb["Schedule-by-Time"])
    by_sport = _sheet_texts(wb["Schedule-by-Sport"])
    conflict_audit = _sheet_texts(wb["Conflict-Audit"])
    assert any(f"{label_a} vs {label_b}" in text for text in by_time)
    assert label_a in by_sport
    assert label_b in by_sport
    assert label_a in conflict_audit
    assert label_b in conflict_audit
    assert participant_names in conflict_audit


def test_schedule_pipeline_export_solve_produce_local(mocker, monkeypatch, tmp_path):
    pytest.importorskip("ortools")

    export_dir = tmp_path / "export"
    schedule_input_path = export_dir / "schedule_input.json"
    schedule_output_path = export_dir / "schedule_output.json"
    workbook_path = export_dir / "VAYSF_Schedule_test.xlsx"
    real_exporter_cls = main.ChurchTeamsExporter

    class FakeExporter:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def generate_reports(
            self,
            target_church_code,
            output_dir,
            force_resend_pending,
            force_resend_validated1,
            force_resend_validated2,
            dry_run,
            target_resend_chm_id,
        ):
            output_dir.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws["A1"] = "Fake export for CLI pipeline test"
            wb.save(output_dir / "Church_Team_Status_ALL_2026-05-15.xlsx")
            data = _minimal_schedule_input()
            schedule_input = output_dir / "schedule_input.json"
            schedule_input.write_text(json.dumps(data), encoding="utf-8")
            return True

    monkeypatch.setattr(main, "ChurchTeamsExporter", FakeExporter)

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="export-church-teams",
            church_code=None,
            output=str(export_dir),
            force_resend_pending=False,
            force_resend_validated1=False,
            force_resend_validated2=False,
            dry_run=False,
            chm_id=None,
        ),
    )
    _run_main_expect_exit(0)
    assert schedule_input_path.exists()
    monkeypatch.setattr(main, "ChurchTeamsExporter", real_exporter_cls)

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="solve-schedule",
            input=str(schedule_input_path),
            output=str(schedule_output_path),
        ),
    )
    _run_main_expect_exit(0)
    assert schedule_output_path.exists()

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="produce-schedule",
            schedule_output=str(schedule_output_path),
            schedule_input=str(schedule_input_path),
            output=str(workbook_path),
        ),
    )
    _run_main_expect_exit(0)

    wb = load_workbook(workbook_path)
    assert "Schedule-by-Time" in wb.sheetnames
    assert "Schedule-by-Sport" in wb.sheetnames
    assert "Conflict-Audit" in wb.sheetnames

    schedule_output = json.loads(schedule_output_path.read_text(encoding="utf-8"))
    assert schedule_output["status"] == "OPTIMAL"
    assert len(schedule_output["assignments"]) == 2


def _write_match_schedule_overrides_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    bb_fill = PatternFill("solid", fgColor="F8CBAD")
    mvb_fill = PatternFill("solid", fgColor="9BC2E6")

    ws.cell(row=1, column=1, value="SAT 7/18")
    ws.cell(row=1, column=2, value="Main Gym BB1")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
    ws.cell(row=1, column=5, value="Prac. Gym VB1")
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="12:00 PM")
    for col, (value, fill) in {
        2: ("WAG", bb_fill), 3: ("v", bb_fill), 4: ("FVC", bb_fill),
        5: ("GLA", mvb_fill), 6: ("v", mvb_fill), 7: ("PCC", mvb_fill),
    }.items():
        cell = ws.cell(row=2, column=col, value=value)
        cell.fill = fill

    wb.save(path)


def test_main_import_match_schedule_overrides_dry_run_writes_audit(monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "export"
    data_dir.mkdir()
    export_dir.mkdir()
    monkeypatch.setattr(main, "DATA_DIR", data_dir)
    monkeypatch.setattr(main, "EXPORT_DIR", export_dir)

    workbook_path = data_dir / "2026 Main Schedule draft 11.xlsx"
    _write_match_schedule_overrides_workbook(workbook_path)

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="import-match-schedule-overrides",
            workbook=str(workbook_path),
            events=None,
            input_xlsx=None,
            schedule_input=None,
            output=None,
            dry_run=True,
            execute=False,
        ),
    )
    _run_main_expect_exit(0)

    audit_path = export_dir / "match_schedule_overrides.audit.json"
    assert audit_path.exists()
    payload = json.loads(audit_path.read_text(encoding="utf-8"))
    assert payload["validation"]["errors"] == []
    assert payload["validation"]["created_game_count"] == 2
    assert not (export_dir / "match_schedule_overrides.json").exists()


def test_main_import_match_schedule_overrides_execute_writes_sidecar(monkeypatch, tmp_path):
    data_dir = tmp_path / "data"
    export_dir = tmp_path / "export"
    data_dir.mkdir()
    export_dir.mkdir()
    monkeypatch.setattr(main, "DATA_DIR", data_dir)
    monkeypatch.setattr(main, "EXPORT_DIR", export_dir)

    workbook_path = data_dir / "2026 Main Schedule draft 11.xlsx"
    _write_match_schedule_overrides_workbook(workbook_path)

    monkeypatch.setattr(
        main,
        "parse_args",
        lambda: argparse.Namespace(
            command="import-match-schedule-overrides",
            workbook=str(workbook_path),
            events=None,
            input_xlsx=None,
            schedule_input=None,
            output=None,
            dry_run=False,
            execute=True,
        ),
    )
    _run_main_expect_exit(0)

    sidecar_path = export_dir / "match_schedule_overrides.json"
    assert sidecar_path.exists()
    payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
    assert payload["validation"]["created_game_count"] == 2
