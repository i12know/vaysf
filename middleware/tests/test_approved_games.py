from openpyxl import Workbook
from openpyxl.styles import PatternFill

from config import (
    POD_RESOURCE_TYPE_BADMINTON,
    POD_RESOURCE_TYPE_TABLE_TENNIS,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_SOCCER,
)
from schedule_contracts import (
    validate_output_against_input,
    validate_schedule_input,
    validate_schedule_output,
)
from schedule_styles import SPORT_STYLES
from scheduling import approved_games


def test_default_main_schedule_path_uses_current_official_draft_12(tmp_path):
    assert (
        approved_games.default_main_schedule_path(tmp_path)
        == tmp_path / "VAY2026_Main_Schedule_draft_12.xlsx"
    )


def _solid(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _write_main_schedule(path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "SAT 7/18"
    ws.merge_cells("B1:D1")
    ws["B1"] = "Main Gym BB1"
    ws["A2"] = "1:00 PM"
    for coord, value in (("B2", "ANH"), ("C2", "v"), ("D2", "GAC")):
        ws[coord] = value
        ws[coord].fill = _solid(SPORT_STYLES[SPORT_TYPE["BASKETBALL"]].fill_color)
    wb.save(path)


def _write_badminton(path, team_a="GLA", team_b="ANH"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["2026 VAY Badminton Preliminary Schedule"])
    ws.append(["Time", "#", "Court 1", None, None])
    ws.append(["5:00 PM", "", "Opening and Check-Ins"])
    ws.append(["5:20 PM", 1, team_a, "v", team_b])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["", "BADMINTON - MEN'S DOUBLES", None, None])
    ws2.append(["No.", "Full Name", "Gender", "Team"])
    ws2.append([1, "Player One", "Male", "GLA"])
    ws2.append([2, "Player Two", "Male", "ANH"])
    wb.save(path)


def _write_soccer(path, team_a="ANH", team_b="GAC"):
    wb = Workbook()
    ws = wb.active
    for _ in range(10):
        ws.append([])
    ws.append(["SAT 7/18", "1:00 PM", "G1", team_a, "v", team_b, "RPC"])
    wb.save(path)


def _write_table_tennis(
    path,
    include_sbc=True,
    matches=None,
    roster_entries=None,
    date_header="Fri 7/24",
):
    wb = Workbook()
    ws = wb.active
    ws.append(["", "VAY SPORTS FEST 2026 - TABLE TENNIS SCHEDULE"])
    ws.append([date_header, "Table 1", "Table 2", "Table 3", "Table 4"])
    matches = matches or ["Nhan Micah (ORN) - Phan Dora (WSD)"]
    for index, match in enumerate(matches):
        minutes = index * 15
        ws.append([f"{17 + minutes // 60:02d}:{minutes % 60:02d}", match])
    ws.append([])
    ws.append(["TEAM", "ATHLETES"])
    for team, athletes in roster_entries or []:
        ws.append([team, athletes])
    if include_sbc:
        ws.append(["(U35) SBC", "Player A & Player B"])
    wb.save(path)


def _tt_roster_row(
    church,
    first,
    last,
    *,
    event=None,
    gender="Women",
    sport_format="Singles",
    team_order=None,
):
    return {
        "Church Team": church,
        "sport_type": event or SPORT_TYPE["TABLE_TENNIS"],
        "sport_gender": gender,
        "sport_format": sport_format,
        "team_order": team_order,
        "Participant ID (WP)": f"{church}-{first}-{last}",
        "First Name": first,
        "Last Name": last,
    }


def _schedule_input():
    return {
        "games": [
            {
                "game_id": "BBM-01",
                "event": SPORT_TYPE["BASKETBALL"],
                "stage": "Pool",
                "pool_id": "A",
                "round": 1,
                "team_a_id": "BBM::ANH",
                "team_b_id": "BBM::GAC",
                "team_a_label": "ANH",
                "team_b_label": "GAC",
                "duration_minutes": 60,
                "resource_type": "Basketball Court",
            }
        ],
        "resources": [
            {
                "resource_id": "BB-Sat-1-1",
                "resource_type": "Basketball Court",
                "label": "Court-1",
                "day": "Sat-1",
                "open_time": "13:00",
                "close_time": "14:00",
                "slot_minutes": 60,
            },
            {
                "resource_id": "BAD-Fri-1-1",
                "resource_type": POD_RESOURCE_TYPE_BADMINTON,
                "label": "Court-1",
                "day": "Fri-1",
                "open_time": "17:00",
                "close_time": "18:00",
                "slot_minutes": 60,
            },
            {
                "resource_id": "SOC-Sat-1-1",
                "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
                "label": "Court-1",
                "day": "Sat-1",
                "open_time": "13:00",
                "close_time": "14:00",
                "slot_minutes": 60,
            },
            {
                "resource_id": "TT-Fri-1-1",
                "resource_type": POD_RESOURCE_TYPE_TABLE_TENNIS,
                "label": "Table-1",
                "day": "Fri-1",
                "open_time": "17:00",
                "close_time": "18:00",
                "slot_minutes": 20,
                "venue_name": "Orange Chapel",
            },
        ],
    }


def test_approved_games_payload_builds_publishable_artifacts(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=False)

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
    )

    assert payload["validation"]["errors"] == []
    keys = {game["game_key"] for game in payload["games"]}
    assert {"BBM-01", "BAD-MD-01", "SOC-G1", "TT-W-S-01"} <= keys

    publish_input = payload["publish_artifacts"]["schedule_input"]
    publish_output = payload["publish_artifacts"]["schedule_output"]
    assert validate_schedule_input(publish_input) == []
    assert validate_schedule_output(publish_output) == []
    assert validate_output_against_input(publish_output, publish_input) == []


def test_soccer_approved_schedule_can_create_override_resource(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=False)
    schedule_input = _schedule_input()
    schedule_input["resources"] = [
        resource for resource in schedule_input["resources"]
        if resource["resource_type"] != TEAM_RESOURCE_TYPE_SOCCER
    ]

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=schedule_input,
    )

    assert payload["validation"]["errors"] == []
    assert any("approved Soccer workbook override" in warning for warning in payload["validation"]["warnings"])
    publish_resources = payload["publish_artifacts"]["schedule_input"]["resources"]
    assert any(resource["resource_id"] == "SOC-APPROVED-Sat-1-1300" for resource in publish_resources)


def test_table_tennis_sbc_discrepancy_blocks_execute_without_waiver(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=True)

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
    )

    assert any("SBC" in error for error in payload["validation"]["errors"])

    waived = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
        waive_table_tennis_discrepancy=True,
    )
    assert waived["validation"]["errors"] == []
    assert any("waived" in warning for warning in waived["validation"]["warnings"])


def test_bye_rows_are_not_imported_as_real_games(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton, team_b="BYE")
    _write_soccer(soccer, team_b="BYE")
    _write_table_tennis(
        table_tennis,
        include_sbc=False,
        matches=["Bye - Phan Dora (WSD)", "Nhan Micah (ORN) - Bye"],
    )

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
    )

    assert payload["validation"]["errors"] == []
    labels = {
        label
        for game in payload["games"]
        for label in (game.get("team_a_label"), game.get("team_b_label"))
    }
    assert "BYE" not in labels
    keys = {game["game_key"] for game in payload["games"]}
    assert not any(key.startswith(("BAD-", "SOC-", "TT-")) for key in keys)

    # Every sport must record its skipped bye row for audit, not just soccer
    # (issue: badminton/table-tennis byes used to vanish with no trace).
    bye_placeholders = [
        placeholder for placeholder in payload["placeholders"]
        if placeholder.get("classification") == "bye"
    ]
    bye_workbooks = {placeholder["source_workbook"] for placeholder in bye_placeholders}
    assert str(badminton) in bye_workbooks
    assert str(soccer) in bye_workbooks
    assert str(table_tennis) in bye_workbooks
    # Two bye rows were written for table tennis (left-side and right-side bye).
    assert sum(1 for wb in bye_placeholders if wb["source_workbook"] == str(table_tennis)) == 2


def test_table_tennis_source_validation_flags_unregistered_team(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=True)

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
        roster_rows=[
            _tt_roster_row("ORN", "Micah", "Nhan"),
            _tt_roster_row("WSD", "Dora", "Phan"),
        ],
    )

    errors = payload["validation"]["errors"]
    assert any("source team SBC" in error for error in errors)


def test_table_tennis_source_validation_flags_unregistered_athlete(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(
        table_tennis,
        include_sbc=False,
        matches=["(U35) GAC-1 - BYE"],
        roster_entries=[("(U35) GAC-1", "Joshua Dang & Noah Vo")],
    )

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
        roster_rows=[
            _tt_roster_row(
                "GAC",
                "Joshua",
                "Dang",
                gender="Mixed",
                sport_format="Mixed Double",
                team_order=1,
            ),
        ],
    )

    errors = payload["validation"]["errors"]
    assert any("Noah Vo" in error and "not registered" in error for error in errors)


def test_table_tennis_source_validation_warns_for_unique_name_typo(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(
        table_tennis,
        include_sbc=False,
        matches=["(U35) GAC-2 - BYE"],
        roster_entries=[("(U35) GAC-2", "Phillip Tran & Justin Pham")],
    )

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
        roster_rows=[
            _tt_roster_row(
                "GAC",
                "Phillip",
                "Tran",
                gender="Mixed",
                sport_format="Mixed Double",
            ),
            _tt_roster_row(
                "GAC",
                "Justtin",
                "Pham",
                gender="Mixed",
                sport_format="Mixed Double",
            ),
        ],
    )

    assert payload["validation"]["errors"] == []
    warnings = payload["validation"]["warnings"]
    assert any("Justin Pham" in warning and "Justtin Pham" in warning for warning in warnings)


def test_table_tennis_prelim_balance_flags_low_and_high_appearances(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(
        table_tennis,
        include_sbc=False,
        matches=[
            "Nhan Micah (ORN) - Phan Dora (WSD)",
            "Phan Dora (WSD) - To Jacklyn (WSD)",
            "To Jacklyn (WSD) - Ly Jennifer (RPC)",
            "Ly Jennifer (RPC) - Nguyen Christina (WSD)",
            "Nguyen Christina (WSD) - Nhan Micah (ORN)",
            "Nhan Micah (ORN) - To Jacklyn (WSD)",
            "Phan Dora (WSD) - Ly Jennifer (RPC)",
            "To Jacklyn (WSD) - Bye",
        ],
    )
    schedule_input = _schedule_input()
    schedule_input["resources"][-1]["close_time"] = "20:00"

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=schedule_input,
    )

    errors = payload["validation"]["errors"]
    balance_error = next(
        error for error in errors
        if "Women's Singles preliminary row counts are unbalanced" in error
    )
    assert "Nguyen Christina=2" in balance_error
    assert "To Jacklyn=4" in balance_error


def test_table_tennis_side_parts_accepts_lowercase_team_code():
    # A hand-typed workbook cell can write the church code in any case; the
    # extracted team code should still normalize to uppercase either way.
    for text in ("Nhan Micah (ORN)", "Nhan Micah (orn)", "Nhan Micah (Orn)"):
        label, team_code = approved_games._table_tennis_side_parts(text)
        assert label == "Nhan Micah"
        assert team_code == "ORN"


def test_table_tennis_source_validation_matches_lowercase_church_code(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(
        table_tennis,
        include_sbc=False,
        matches=["Nhan Micah (orn) - Phan Dora (WSD)"],
    )

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
        roster_rows=[
            _tt_roster_row("ORN", "Micah", "Nhan"),
            _tt_roster_row("WSD", "Dora", "Phan"),
        ],
    )

    assert payload["validation"]["errors"] == []


def test_table_tennis_requires_friday_724_header(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=False, date_header="Sat 7/24")

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
    )

    assert any("Friday 7/24" in error for error in payload["validation"]["errors"])


def test_table_tennis_requires_orange_resource(tmp_path):
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=False)
    schedule_input = _schedule_input()
    schedule_input["resources"][-1]["venue_name"] = "Esperanza Chapel"

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=schedule_input,
    )

    assert any("expected Orange" in error for error in payload["validation"]["errors"])


def test_table_tennis_wrong_day_workbook_still_blocked_by_header_check(tmp_path):
    """The header scan alone must catch a workbook that isn't Friday 7/24 at
    all (not just a Friday-dated header with the wrong weekday label) — this
    is the case the removed, unreachable per-game day check used to look
    like it was covering. _parse_table_tennis always builds scheduled_slot
    with the hardcoded required day, so a per-record comparison against that
    same hardcoded value could never disagree with it regardless of what the
    workbook's header actually says; the header scan is what has to do the
    real work.
    """
    main = tmp_path / "main.xlsx"
    badminton = tmp_path / "badminton.xlsx"
    soccer = tmp_path / "soccer.xlsx"
    table_tennis = tmp_path / "tt.xlsx"
    _write_main_schedule(main)
    _write_badminton(badminton)
    _write_soccer(soccer)
    _write_table_tennis(table_tennis, include_sbc=False, date_header="Sat 7/25")

    payload = approved_games.build_approved_games_payload(
        main_schedule_path=main,
        badminton_path=badminton,
        soccer_path=soccer,
        table_tennis_path=table_tennis,
        schedule_input=_schedule_input(),
    )

    assert any("Friday 7/24" in error for error in payload["validation"]["errors"])
