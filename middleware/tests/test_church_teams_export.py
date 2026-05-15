import os
import sys
from unittest.mock import MagicMock

import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from church_teams_export import ChurchTeamsExporter, CHM_FIELDS, MEMBERSHIP_QUESTION
from config import SPORT_TYPE


@pytest.fixture()
def mock_connectors(mocker):
    chm_connector = MagicMock()
    chm_connector.get_groups.return_value = []
    chm_connector.get_group_people.return_value = []
    chm_connector.last_get_person_status = "ok"

    wp_connector = MagicMock()

    mocker.patch("church_teams_export.ChMeetingsConnector", return_value=chm_connector)
    mocker.patch("church_teams_export.WordPressConnector", return_value=wp_connector)

    return chm_connector, wp_connector


def test_fetch_chm_church_team_data_skips_orphaned_memberships(mock_connectors):
    chm_connector, _ = mock_connectors

    chm_connector.get_groups.return_value = [{"id": "870578", "name": "Team RPC"}]
    chm_connector.get_group_people.return_value = [
        {"person_id": "999999"},
        {"person_id": "101"},
    ]

    valid_person = {
        "id": "101",
        "first_name": "Alice",
        "last_name": "Nguyen",
        "gender": "Female",
        "birth_date": "2000-01-02",
        "mobile": "555-0101",
        "email": "alice@test.com",
        "updated_on": "2026-05-07T22:17:30+00:00",
        "additional_fields": [
            {"field_name": MEMBERSHIP_QUESTION, "value": "Yes"},
            {"field_name": CHM_FIELDS["ROLES"], "value": "Athlete"},
            {"field_name": CHM_FIELDS["COMPLETION_CHECKLIST"], "value": ""},
        ],
    }

    def fake_get_person(person_id):
        if person_id == "999999":
            chm_connector.last_get_person_status = "not_found"
            return None
        chm_connector.last_get_person_status = "ok"
        return valid_person

    chm_connector.get_person.side_effect = fake_get_person

    exporter = ChurchTeamsExporter()
    data = exporter._fetch_chm_church_team_data()

    assert "RPC" in data
    assert len(data["RPC"]) == 1
    assert data["RPC"][0]["ChMeetings ID"] == "101"
    assert exporter.last_orphaned_memberships_by_church == {"RPC": 1}


def test_handle_force_resend_filters_to_one_chm_id(mock_connectors, mocker):
    _, wp_connector = mock_connectors

    fake_sync_manager = MagicMock()
    fake_sync_manager.wordpress_connector.get_churches.return_value = [
        {"church_code": "RPC", "pastor_email": "pastor@rpc.org", "church_rep_email": "rep@rpc.org"}
    ]
    fake_sync_manager.__enter__.return_value = fake_sync_manager
    fake_sync_manager.__exit__.return_value = None

    mocker.patch("sync.manager.SyncManager", return_value=fake_sync_manager)

    exporter = ChurchTeamsExporter()
    contacts = [
        {
            "ChMeetings ID": "3318927",
            "First Name": "Sam",
            "Last Name": "Le",
            "Church Team": "RPC",
            "Email": "sam@example.com",
            "Approval_Status (WP)": "pending_approval",
        },
        {
            "ChMeetings ID": "4363698",
            "First Name": "Thomas",
            "Last Name": "Phan",
            "Church Team": "RPC",
            "Email": "thomas@example.com",
            "Approval_Status (WP)": "pending_approval",
        },
    ]

    resend_count = exporter._handle_force_resend(
        contacts,
        force_pending=True,
        force_validated1=False,
        force_validated2=False,
        dry_run=True,
        target_resend_chm_id="3318927",
    )

    assert resend_count == 1


def test_generate_reports_surfaces_open_validation_issues(mock_connectors, mocker, tmp_path):
    chm_connector, wp_connector = mock_connectors
    chm_connector.authenticate.return_value = True

    exporter = ChurchTeamsExporter()
    exporter.latest_chm_update_by_church = {"RPC": "2026-05-08 10:00:00"}
    mocker.patch.object(
        exporter,
        "_fetch_chm_church_team_data",
        return_value={
            "RPC": [
                {
                    "Church Team": "RPC",
                    "ChMeetings ID": "101",
                    "First Name": "Alice",
                    "Last Name": "Nguyen",
                    "Gender": "Female",
                    "Birthdate": "2000-01-02",
                    "Mobile Phone": "555-0101",
                    "Email": "alice@test.com",
                    "Is_Member_ChM": True,
                    "ChM_Roles": "Athlete",
                    "ChM_Completion_Checklist": "",
                    "Update_on_ChM": "2026-05-08 10:00:00",
                }
            ]
        },
    )

    wp_connector.get_church_by_code.return_value = {"church_id": 1, "church_code": "RPC"}
    wp_connector.get_participants.return_value = [
        {
            "participant_id": 42,
            "approval_status": "pending",
            "photo_url": "https://example.com/photo.jpg",
        }
    ]
    wp_connector.get_validation_issues.return_value = [
        {
            "issue_id": 1,
            "participant_id": 42,
            "issue_type": "missing_photo",
            "issue_description": "No photo uploaded",
            "rule_code": "PHOTO_REQUIRED",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": None,
            "sport_format": None,
        },
        {
            "issue_id": 2,
            "participant_id": None,
            "issue_type": "team_non_member_limit",
            "issue_description": "Basketball - Men Team has 3 non-members, exceeding limit of 2",
            "rule_code": "MAX_NON_MEMBERS_TEAM",
            "rule_level": "TEAM",
            "severity": "ERROR",
            "status": "open",
            "sport_type": "Basketball - Men Team",
            "sport_format": None,
        },
        {
            "issue_id": 3,
            "participant_id": None,
            "issue_type": "doubles_non_member_limit",
            "issue_description": "Badminton Men Double pair Alice / Guest has 2 non-members, exceeding limit of 1",
            "rule_code": "MAX_NON_MEMBERS_DOUBLES",
            "rule_level": "TEAM",
            "severity": "ERROR",
            "status": "open",
            "sport_type": "Badminton",
            "sport_format": "Men Double",
        },
        {
            "issue_id": 4,
            "participant_id": 42,
            "issue_type": "missing_consent",
            "issue_description": "Missing consent form",
            "rule_code": "CONSENT_REQUIRED",
            "rule_level": "INDIVIDUAL",
            "severity": "WARNING",
            "status": "open",
            "sport_type": None,
            "sport_format": None,
        },
    ]
    wp_connector.get_rosters.return_value = [
        {
            "sport_type": "Basketball",
            "sport_gender": "Men",
            "sport_format": "Team",
            "team_order": None,
            "partner_name": None,
        },
        {
            "sport_type": "Badminton",
            "sport_gender": "Men",
            "sport_format": "Doubles",
            "team_order": None,
            "partner_name": "Guest",
        },
    ]

    write_report = mocker.patch.object(exporter, "_write_excel_report")

    result = exporter.generate_reports("RPC", tmp_path)

    assert result is True
    write_report.assert_called_once()
    _, summary_rows, contacts_rows, roster_rows, validation_rows = write_report.call_args.args

    assert summary_rows[0]["Total Participants w/ Open ERRORs (WP)"] == 1
    assert summary_rows[0]["Total Open Individual ERRORs (WP)"] == 1
    assert summary_rows[0]["Total Open TEAM ERRORs (WP)"] == 2
    assert summary_rows[0]["Total Open WARNINGs (WP)"] == 1
    assert summary_rows[0]["Total Sports w/ Open TEAM Issues (WP)"] == 2

    assert contacts_rows[0]["Total_Open_ERRORs (WP)"] == 1
    assert contacts_rows[0]["First_Open_ERROR_Desc (WP)"] == "No photo uploaded"

    basketball_row = next(row for row in roster_rows if row["sport_type"] == "Basketball")
    badminton_row = next(row for row in roster_rows if row["sport_type"] == "Badminton")
    assert basketball_row["Open_TEAM_Issue_Count (WP)"] == 1
    assert "Basketball - Men Team has 3 non-members" in basketball_row["Open_TEAM_Issue_Desc (WP)"]
    assert badminton_row["Open_TEAM_Issue_Count (WP)"] == 1
    assert "Badminton Men Double pair Alice / Guest" in badminton_row["Open_TEAM_Issue_Desc (WP)"]

    assert len(validation_rows) == 4
    participant_issue = next(row for row in validation_rows if row["Issue Type"] == "missing_photo")
    assert participant_issue["Participant Name"] == "Alice Nguyen"
    team_issue = next(row for row in validation_rows if row["Issue Type"] == "team_non_member_limit")
    assert team_issue["Rule Level"] == "TEAM"
    assert team_issue["Participant Name"] == ""


def test_generate_reports_filters_stale_individual_validation_issues(mock_connectors, mocker, tmp_path):
    chm_connector, wp_connector = mock_connectors
    chm_connector.authenticate.return_value = True

    exporter = ChurchTeamsExporter()
    exporter.latest_chm_update_by_church = {"RPC": "2026-05-08 10:00:00"}
    mocker.patch.object(
        exporter,
        "_fetch_chm_church_team_data",
        return_value={
            "RPC": [
                {
                    "Church Team": "RPC",
                    "ChMeetings ID": "101",
                    "First Name": "Alice",
                    "Last Name": "Nguyen",
                    "Gender": "Female",
                    "Birthdate": "2000-01-02",
                    "Mobile Phone": "555-0101",
                    "Email": "alice@test.com",
                    "Is_Member_ChM": True,
                    "ChM_Roles": "Athlete",
                    "ChM_Completion_Checklist": "",
                    "Update_on_ChM": "2026-05-08 10:00:00",
                }
            ]
        },
    )

    wp_connector.get_church_by_code.return_value = {"church_id": 1, "church_code": "RPC"}
    wp_connector.get_participants.return_value = [
        {
            "participant_id": 42,
            "approval_status": "pending",
            "photo_url": "https://example.com/photo.jpg",
        }
    ]
    wp_connector.get_validation_issues.return_value = [
        {
            "issue_id": 1,
            "participant_id": 42,
            "issue_type": "missing_photo",
            "issue_description": "No photo uploaded",
            "rule_code": "PHOTO_REQUIRED",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": None,
            "sport_format": None,
        },
        {
            "issue_id": 2,
            "participant_id": 99,
            "issue_type": "missing_consent",
            "issue_description": "Consent form status unknown or not provided",
            "rule_code": "CONSENT_REQUIRED",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": None,
            "sport_format": None,
        },
    ]
    wp_connector.get_rosters.return_value = []

    write_report = mocker.patch.object(exporter, "_write_excel_report")

    result = exporter.generate_reports("RPC", tmp_path)

    assert result is True
    _, summary_rows, contacts_rows, _, validation_rows = write_report.call_args.args

    assert summary_rows[0]["Total Participants w/ Open ERRORs (WP)"] == 1
    assert summary_rows[0]["Total Open Individual ERRORs (WP)"] == 1
    assert contacts_rows[0]["Total_Open_ERRORs (WP)"] == 1
    assert len(validation_rows) == 1
    assert validation_rows[0]["Participant ID (WP)"] == 42
    assert validation_rows[0]["Issue Type"] == "missing_photo"


def test_write_excel_report_adds_validation_issues_tab(mock_connectors, tmp_path):
    exporter = ChurchTeamsExporter()
    filepath = tmp_path / "church-report.xlsx"

    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 1,
        "Total Participants (in WP)": 1,
        "Total Approved (WP)": 0,
        "Total Pending Approval (WP)": 1,
        "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 1,
        "Total Open Individual ERRORs (WP)": 1,
        "Total Open TEAM ERRORs (WP)": 1,
        "Total Open WARNINGs (WP)": 0,
        "Total Sports w/ Open TEAM Issues (WP)": 1,
        "Latest ChM Record Update for Team": "2026-05-08 10:00:00",
    }]
    contacts_rows = [{
        "Church Team": "RPC",
        "ChMeetings ID": "101",
        "First Name": "Alice",
        "Last Name": "Nguyen",
        "Is_Participant": "Yes",
        "Is_Member_ChM": "Yes",
        "Participant ID (WP)": 42,
        "Approval_Status (WP)": "pending",
        "Total_Open_ERRORs (WP)": 1,
        "Gender": "Female",
        "Birthdate": "2000-01-02",
        "Age (at Event)": 26,
        "Mobile Phone": "555-0101",
        "Email": "alice@test.com",
        "First_Open_ERROR_Desc (WP)": "No photo uploaded",
        "Box 1": "",
        "Box 2": "",
        "Box 3": "",
        "Box 4": "",
        "Box 5": "",
        "Box 6": "",
        "Photo URL (WP)": "N/A",
        "Update_on_ChM": "2026-05-08 10:00:00",
    }]
    roster_rows = [{
        "Church Team": "RPC",
        "ChMeetings ID": "101",
        "Participant ID (WP)": 42,
        "Approval_Status (WP)": "pending",
        "Is_Member_ChM": True,
        "Photo": "",
        "First Name": "Alice",
        "Last Name": "Nguyen",
        "Gender": "Female",
        "Age (at Event)": 26,
        "Mobile Phone": "555-0101",
        "Email": "alice@test.com",
        "sport_type": "Badminton",
        "sport_gender": "Men",
        "sport_format": "Doubles",
        "team_order": None,
        "partner_name": "Guest",
        "Open_TEAM_Issue_Count (WP)": 1,
        "Open_TEAM_Issue_Desc (WP)": "Badminton Men Double pair Alice / Guest has 2 non-members, exceeding limit of 1",
    }]
    validation_rows = [{
        "Church Team": "RPC",
        "Rule Level": "TEAM",
        "Severity": "ERROR",
        "Status": "open",
        "Issue Type": "doubles_non_member_limit",
        "Rule Code": "MAX_NON_MEMBERS_DOUBLES",
        "Participant ID (WP)": "",
        "ChMeetings ID": "",
        "Participant Name": "",
        "Approval_Status (WP)": "",
        "sport_type": "Badminton",
        "sport_format": "Men Double",
        "Issue Description": "Badminton Men Double pair Alice / Guest has 2 non-members, exceeding limit of 1",
    }]

    exporter._write_excel_report(filepath, summary_rows, contacts_rows, roster_rows, validation_rows)

    workbook = pd.ExcelFile(filepath)
    assert "Validation-Issues" in workbook.sheet_names

    validation_df = pd.read_excel(filepath, sheet_name="Validation-Issues")
    roster_df = pd.read_excel(filepath, sheet_name="Roster")
    assert "Rule Level" in validation_df.columns
    assert validation_df.loc[0, "Issue Type"] == "doubles_non_member_limit"
    assert "Open_TEAM_Issue_Count (WP)" in roster_df.columns
    assert int(roster_df.loc[0, "Open_TEAM_Issue_Count (WP)"]) == 1


def test_validation_issue_rows_add_reverse_partner_suggestion(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {
            "Church Team": "RPC",
            "Participant ID (WP)": 72,
            "First Name": "Dean",
            "Last Name": "Nguyen",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "Janice",
        },
        {
            "Church Team": "RPC",
            "Participant ID (WP)": 75,
            "First Name": "Janice",
            "Last Name": "Vu",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "",
        },
    ]
    reverse_lookup = exporter._build_reverse_partner_suggestion_lookup(roster_rows)

    issue_rows = exporter._build_validation_issue_rows(
        "RPC",
        [{
            "participant_id": 75,
            "issue_type": "missing_doubles_partner",
            "issue_description": "Partner name required for Pickleball (Mixed Double)",
            "rule_code": "PARTNER_REQUIRED_DOUBLES",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": "Pickleball",
            "sport_format": "Mixed Double",
        }],
        {
            "75": {
                "ChMeetings ID": "43636",
                "First Name": "Janice",
                "Last Name": "Vu",
                "Approval_Status (WP)": "pending",
            }
        },
        reverse_lookup,
    )

    assert len(issue_rows) == 1
    assert "perhaps Dean Nguyen listed you as partner." in issue_rows[0]["Issue Description"]


def test_reverse_partner_suggestion_lookup_skips_ambiguous_partial_match(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {
            "Church Team": "RPC",
            "Participant ID (WP)": 72,
            "First Name": "Dean",
            "Last Name": "Nguyen",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "Janice",
        },
        {
            "Church Team": "RPC",
            "Participant ID (WP)": 75,
            "First Name": "Janice",
            "Last Name": "Vu",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "",
        },
        {
            "Church Team": "RPC",
            "Participant ID (WP)": 76,
            "First Name": "Janice",
            "Last Name": "Nguyen",
            "sport_type": "Pickleball",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "",
        },
    ]

    reverse_lookup = exporter._build_reverse_partner_suggestion_lookup(roster_rows)

    janice_vu_key = exporter._reverse_partner_suggestion_key(75, "Pickleball", "Mixed", "Doubles")
    janice_nguyen_key = exporter._reverse_partner_suggestion_key(76, "Pickleball", "Mixed", "Doubles")
    assert janice_vu_key not in reverse_lookup
    assert janice_nguyen_key not in reverse_lookup


def test_validation_issue_rows_do_not_cross_gendered_doubles_formats(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {
            "Church Team": "TLC",
            "Participant ID (WP)": 26,
            "First Name": "Hyewon",
            "Last Name": "Yun",
            "sport_type": "Badminton",
            "sport_gender": "Men",
            "sport_format": "Doubles",
            "partner_name": "Shawn Le",
        },
        {
            "Church Team": "TLC",
            "Participant ID (WP)": 31,
            "First Name": "Shawn",
            "Last Name": "Le",
            "sport_type": "Badminton",
            "sport_gender": "Mixed",
            "sport_format": "Doubles",
            "partner_name": "",
        },
    ]

    reverse_lookup = exporter._build_reverse_partner_suggestion_lookup(roster_rows)

    issue_rows = exporter._build_validation_issue_rows(
        "TLC",
        [{
            "participant_id": 31,
            "issue_type": "missing_doubles_partner",
            "issue_description": "Partner name required for Badminton (Mixed Double)",
            "rule_code": "PARTNER_REQUIRED_DOUBLES",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": "Badminton",
            "sport_format": "Mixed Double",
        }],
        {
            "31": {
                "ChMeetings ID": "35628",
                "First Name": "Shawn",
                "Last Name": "Le",
                "Approval_Status (WP)": "pending",
            }
        },
        reverse_lookup,
    )

    assert len(issue_rows) == 1
    assert "perhaps Hyewon Yun listed you as partner." not in issue_rows[0]["Issue Description"]


def test_issue_based_reverse_partner_suggestion_handles_incomplete_roster_data(mock_connectors):
    exporter = ChurchTeamsExporter()
    issues = [
        {
            "participant_id": 149,
            "issue_type": "missing_doubles_partner",
            "issue_description": "Partner name required for Table Tennis 35+ (Men Double)",
            "rule_code": "PARTNER_REQUIRED_DOUBLES",
            "rule_level": "INDIVIDUAL",
            "severity": "ERROR",
            "status": "open",
            "sport_type": "Table Tennis 35+",
            "sport_format": "Men Double",
        },
        {
            "participant_id": 156,
            "issue_type": "doubles_partner_unmatched",
            "issue_description": (
                "Long Chung listed Andrew Nguyen as their partner for Table Tennis 35+ "
                "(Men Double), but Andrew Nguyen did not reciprocally list Long Chung."
            ),
            "rule_code": "PARTNER_RECIPROCAL_DOUBLES",
            "rule_level": "TEAM",
            "severity": "WARNING",
            "status": "open",
            "sport_type": "Table Tennis 35+",
            "sport_format": "Men Double",
        },
    ]
    participants_by_wp_id = {
        "149": {
            "ChMeetings ID": "43644",
            "First Name": "Andrew",
            "Last Name": "Nguyen",
            "Approval_Status (WP)": "pending",
        },
        "156": {
            "ChMeetings ID": "43644",
            "First Name": "Long",
            "Last Name": "Chung",
            "Approval_Status (WP)": "pending",
        },
    }

    reverse_lookup = exporter._build_issue_based_reverse_partner_suggestion_lookup(
        issues,
        participants_by_wp_id,
    )

    issue_rows = exporter._build_validation_issue_rows(
        "TLC",
        [issues[0]],
        participants_by_wp_id,
        reverse_lookup,
    )

    assert len(issue_rows) == 1
    assert "perhaps Long Chung listed you as partner." in issue_rows[0]["Issue Description"]


def test_contacts_status_tab_includes_sports_registered_column(mock_connectors, tmp_path):
    exporter = ChurchTeamsExporter()
    filepath = tmp_path / "church-report.xlsx"

    contacts_rows = [
        {
            "Church Team": "RPC",
            "ChMeetings ID": "101",
            "First Name": "Alice",
            "Last Name": "Nguyen",
            "Is_Participant": "Yes",
            "Is_Member_ChM": "Yes",
            "Participant ID (WP)": 42,
            "Approval_Status (WP)": "pending",
            "Total_Open_ERRORs (WP)": 0,
            "Gender": "Female",
            "Birthdate": "2000-01-02",
            "Age (at Event)": 26,
            "Mobile Phone": "555-0101",
            "Email": "alice@test.com",
            "Registration Date (WP)": "2026-03-01",
            "Athlete Fee": 30,
            "First_Open_ERROR_Desc (WP)": "",
            "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
            "Photo URL (WP)": "N/A",
            "Update_on_ChM": "2026-05-08",
        },
        {
            "Church Team": "RPC",
            "ChMeetings ID": "102",
            "First Name": "Bob",
            "Last Name": "Tran",
            "Is_Participant": "Yes",
            "Is_Member_ChM": "Yes",
            "Participant ID (WP)": 99,
            "Approval_Status (WP)": "approved",
            "Total_Open_ERRORs (WP)": 0,
            "Gender": "Male",
            "Birthdate": "1998-05-10",
            "Age (at Event)": 28,
            "Mobile Phone": "555-0202",
            "Email": "bob@test.com",
            "Registration Date (WP)": "2026-03-02",
            "Athlete Fee": 30,
            "First_Open_ERROR_Desc (WP)": "",
            "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
            "Photo URL (WP)": "N/A",
            "Update_on_ChM": "2026-05-08",
        },
        # No matching roster entry
        {
            "Church Team": "RPC",
            "ChMeetings ID": "103",
            "First Name": "Carol",
            "Last Name": "Pham",
            "Is_Participant": "No",
            "Is_Member_ChM": "Yes",
            "Participant ID (WP)": None,
            "Approval_Status (WP)": "",
            "Total_Open_ERRORs (WP)": 0,
            "Gender": "Female",
            "Birthdate": "2003-09-15",
            "Age (at Event)": 22,
            "Mobile Phone": "555-0303",
            "Email": "carol@test.com",
            "Registration Date (WP)": "",
            "Athlete Fee": "",
            "First_Open_ERROR_Desc (WP)": "",
            "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
            "Photo URL (WP)": "N/A",
            "Update_on_ChM": "2026-05-08",
        },
    ]
    roster_rows = [
        # Alice plays two sports
        {
            "Church Team": "RPC", "ChMeetings ID": "101", "Participant ID (WP)": 42,
            "sport_type": "Badminton", "sport_gender": "Women", "sport_format": "Doubles",
        },
        {
            "Church Team": "RPC", "ChMeetings ID": "101", "Participant ID (WP)": 42,
            "sport_type": "Basketball", "sport_gender": "", "sport_format": "",
        },
        # Bob plays one sport; duplicate roster row should not duplicate the label
        {
            "Church Team": "RPC", "ChMeetings ID": "102", "Participant ID (WP)": 99,
            "sport_type": "Volleyball", "sport_gender": "Men", "sport_format": "",
        },
        {
            "Church Team": "RPC", "ChMeetings ID": "102", "Participant ID (WP)": 99,
            "sport_type": "Volleyball", "sport_gender": "Men", "sport_format": "",
        },
    ]
    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 3,
        "Total Participants (in WP)": 2,
        "Total Approved (WP)": 1,
        "Total Pending Approval (WP)": 1,
        "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0,
        "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0,
        "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-08",
    }]

    exporter._write_excel_report(filepath, summary_rows, contacts_rows, roster_rows, [])

    contacts_df = pd.read_excel(filepath, sheet_name="Contacts-Status")
    assert "Sports Registered" in contacts_df.columns

    # Column must appear immediately before "Athlete Fee"
    cols = list(contacts_df.columns)
    assert cols.index("Sports Registered") == cols.index("Athlete Fee") - 1

    alice_row = contacts_df[contacts_df["First Name"] == "Alice"].iloc[0]
    sports = [s.strip() for s in alice_row["Sports Registered"].split(",")]
    assert sorted(sports) == sorted(["Badminton Women Doubles", "Basketball"])

    bob_row = contacts_df[contacts_df["First Name"] == "Bob"].iloc[0]
    assert bob_row["Sports Registered"] == "Volleyball Men"

    carol_row = contacts_df[contacts_df["First Name"] == "Carol"].iloc[0]
    assert carol_row["Sports Registered"] == "" or pd.isna(carol_row["Sports Registered"])


def test_venue_capacity_court_slot_math(mock_connectors):
    """Pool/playoff/total slot math (Issue #83)."""
    exporter = ChurchTeamsExporter()

    # 0 teams -> all zeros
    s0 = exporter._compute_court_slots(0)
    assert s0["pool_slots"] == 0
    assert s0["playoff_teams"] == 0
    assert s0["playoff_slots"] == 0
    assert s0["total_slots"] == 0
    assert s0["court_hours"] == 0.0

    # 6 teams, 2 pool games each -> ceil(6*2/2) = 6 pool, 4-team playoff = 3 playoff games
    s6 = exporter._compute_court_slots(6)
    assert s6["pool_slots"] == 6
    assert s6["playoff_teams"] == 4
    assert s6["playoff_slots"] == 3
    assert s6["third_place_slots"] == 0  # default off
    assert s6["total_slots"] == 9
    assert s6["court_hours"] == 9.0  # 60 min/game

    # 8 teams -> ceil(8*2/2)=8 pool, 8-team playoff = 7 playoff games
    s8 = exporter._compute_court_slots(8)
    assert s8["pool_slots"] == 8
    assert s8["playoff_teams"] == 8
    assert s8["playoff_slots"] == 7
    assert s8["total_slots"] == 15

    # 3 teams -> only pool play, no playoff
    s3 = exporter._compute_court_slots(3)
    assert s3["pool_slots"] == 3
    assert s3["playoff_teams"] == 0
    assert s3["playoff_slots"] == 0
    assert s3["total_slots"] == 3


def test_count_estimating_teams_uses_min_team_size(mock_connectors):
    """A church only counts when its roster meets the min team size (Issue #83)."""
    exporter = ChurchTeamsExporter()

    # RPC has 5 basketball players (meets min=5), TLC has 4 (potential only)
    roster_rows = [
        {"Church Team": "RPC", "sport_type": "Basketball", "sport_gender": "Men"} for _ in range(5)
    ] + [
        {"Church Team": "TLC", "sport_type": "Basketball", "sport_gender": "Men"} for _ in range(4)
    ]

    result = exporter._count_estimating_teams(roster_rows, "Basketball - Men Team", min_team_size=5)
    assert result["n_estimating"] == 1       # only RPC qualifies
    assert result["n_potential"] == 2        # RPC (estimating) + TLC (partial) = all with >= 1
    assert result["team_codes"] == "RPC"     # sorted, comma-separated


def test_count_estimating_teams_separates_volleyball_men_and_women(mock_connectors):
    """Volleyball Men and Women are distinct events; team_codes is sorted (Issue #83)."""
    exporter = ChurchTeamsExporter()

    roster_rows = (
        [{"Church Team": "RPC", "sport_type": "Volleyball", "sport_gender": "Men"} for _ in range(6)]
        + [{"Church Team": "RPC", "sport_type": "Volleyball", "sport_gender": "Women"} for _ in range(6)]
        + [{"Church Team": "TLC", "sport_type": "Volleyball", "sport_gender": "Women"} for _ in range(6)]
    )

    men = exporter._count_estimating_teams(roster_rows, "Volleyball - Men Team", 6)
    assert men["n_estimating"] == 1
    assert men["team_codes"] == "RPC"

    women = exporter._count_estimating_teams(roster_rows, "Volleyball - Women Team", 6)
    assert women["n_estimating"] == 2
    assert women["team_codes"] == "RPC, TLC"  # alphabetically sorted


def test_count_estimating_teams_soccer_full_label(mock_connectors):
    """Soccer sport_type is stored as the full Other-Events label, not just 'Soccer'."""
    exporter = ChurchTeamsExporter()

    # Other-events registrations store the full SPORT_TYPE constant value verbatim
    roster_rows = [
        {"Church Team": "RPC", "sport_type": SPORT_TYPE["SOCCER"], "sport_gender": "Mixed"}
        for _ in range(5)
    ] + [
        {"Church Team": "TLC", "sport_type": SPORT_TYPE["SOCCER"], "sport_gender": "Mixed"}
        for _ in range(3)
    ]

    result = exporter._count_estimating_teams(
        roster_rows, SPORT_TYPE["SOCCER"], min_team_size=4
    )
    assert result["n_estimating"] == 1      # only RPC has >= 4
    assert result["n_potential"] == 2       # RPC + TLC both have >= 1
    assert result["team_codes"] == "RPC"


def test_count_racquet_entries(mock_connectors):
    """Racquet entries: complete pairs counted as 1, singles as 1; potential = all regs."""
    exporter = ChurchTeamsExporter()

    roster_rows = [
        # 5 Badminton doubles registrations → 2 complete pairs + 1 waiting
        {"sport_type": "Badminton", "sport_format": "Mixed Doubles"} for _ in range(5)
    ] + [
        # 2 Badminton singles
        {"sport_type": "Badminton", "sport_format": "Men Singles"},
        {"sport_type": "Badminton", "sport_format": "Women Singles"},
    ] + [
        # Pickleball should not bleed into Badminton count
        {"sport_type": "Pickleball", "sport_format": "Mixed Doubles"},
    ]

    result = exporter._count_racquet_entries(roster_rows, "Badminton")
    assert result["n_estimating"] == 2 + 2   # floor(5/2)=2 pairs + 2 singles
    assert result["n_potential"] == 5 + 2    # 5 doubles + 2 singles = 7 registrations
    assert result["team_codes"] == ""


def test_venue_capacity_tab_only_in_consolidated_export(mock_connectors, tmp_path):
    """Venue-Estimator tab appears only when include_venue_capacity=True (Issue #83)."""
    exporter = ChurchTeamsExporter()

    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 6, "Total Participants (in WP)": 6,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 6, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-13",
    }]
    contacts_rows = [{
        "Church Team": "RPC", "ChMeetings ID": str(100 + i), "First Name": f"P{i}",
        "Last Name": "X", "Is_Participant": "Yes", "Is_Member_ChM": "Yes",
        "Participant ID (WP)": i, "Approval_Status (WP)": "pending",
        "Total_Open_ERRORs (WP)": 0, "Gender": "Male", "Birthdate": "2000-01-01",
        "Age (at Event)": 26, "Mobile Phone": "", "Email": "",
        "Registration Date (WP)": "2026-03-01", "Athlete Fee": 30,
        "First_Open_ERROR_Desc (WP)": "",
        "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
        "Photo URL (WP)": "N/A", "Update_on_ChM": "",
    } for i in range(6)]
    roster_rows = [{
        "Church Team": "RPC", "ChMeetings ID": str(100 + i), "Participant ID (WP)": i,
        "sport_type": "Basketball", "sport_gender": "Men", "sport_format": "Team",
    } for i in range(6)]

    # Single-church export: no Venue-Estimator tab
    single_path = tmp_path / "single.xlsx"
    exporter._write_excel_report(single_path, summary_rows, contacts_rows, roster_rows, [])
    assert "Venue-Estimator" not in pd.ExcelFile(single_path).sheet_names

    # Consolidated ALL export: tab present, snapshot note appended after data
    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(all_path, summary_rows, contacts_rows, roster_rows, [],
                                 include_venue_capacity=True)
    sheets = pd.ExcelFile(all_path).sheet_names
    assert "Venue-Estimator" in sheets

    venue_df = pd.read_excel(all_path, sheet_name="Venue-Estimator", header=0)
    assert list(venue_df.columns)[0] == "Event"
    assert "Potential Teams/Entries" in venue_df.columns
    assert "Estimating Teams/Entries" in venue_df.columns
    assert "Teams" in venue_df.columns
    assert "Target Pool Games/Team" in venue_df.columns
    assert "Actual Pool Games/Team" in venue_df.columns
    assert "Pool Composition" in venue_df.columns
    assert "BYE Slots" in venue_df.columns
    assert "Estimated Court Hours" in venue_df.columns
    # 5 team sports + 6 racquet sports
    assert len(venue_df[venue_df["Minutes Per Game"].notna()]) == 11  # 5 team + 6 racquet sports

    # Column order: Potential before Estimating before Teams
    cols = list(venue_df.columns)
    assert cols.index("Potential Teams/Entries") < cols.index("Estimating Teams/Entries") < cols.index("Teams")

    bball = venue_df[venue_df["Event"] == "Basketball - Men Team"].iloc[0]
    assert int(bball["Estimating Teams/Entries"]) == 1   # RPC's 6 basketball players qualify
    assert int(bball["Potential Teams/Entries"]) == 1    # RPC (estimating) counts in potential too
    assert str(bball["Teams"]) == "RPC"
    assert int(bball["Pool Slots"]) == 0         # 1 team can't play pool games (B3: actual=0)
    assert int(bball["Playoff Teams"]) == 0      # 1 team → no playoff
    assert int(bball["Total Court Slots"]) == 0

    vb_men = venue_df[venue_df["Event"] == "Volleyball - Men Team"].iloc[0]
    assert int(vb_men["Estimating Teams/Entries"]) == 0  # no volleyball rosters
    assert str(vb_men["Teams"]) in ("", "nan")

    # Snapshot disclaimer appears after the data (header + 11 rows + blank = row 13)
    raw = pd.read_excel(all_path, sheet_name="Venue-Estimator", header=None)
    note_row_idx = 13  # 0-based: row 14 in Excel (1 header + 11 data + 1 blank + note)
    assert "Roster snapshot as of" in str(raw.iloc[note_row_idx, 0])


# ── Pod-Divisions / Pod-Entries-Review tests (Issue #88) ────────────────────

def test_pod_format_class(mock_connectors):
    exporter = ChurchTeamsExporter()
    assert exporter._pod_format_class("Men Single") == "singles"
    assert exporter._pod_format_class("Singles") == "singles"
    assert exporter._pod_format_class("Women Singles") == "singles"
    assert exporter._pod_format_class("Men Double") == "doubles"
    assert exporter._pod_format_class("Doubles") == "doubles"
    assert exporter._pod_format_class("Mixed Double") == "doubles"
    assert exporter._pod_format_class("Team") == "anomaly"
    assert exporter._pod_format_class("") == "anomaly"
    assert exporter._pod_format_class(None) == "anomaly"


def test_make_division_id(mock_connectors):
    exporter = ChurchTeamsExporter()
    assert exporter._make_division_id("Badminton", "Men", "singles") == "BAD-Men-Singles"
    assert exporter._make_division_id("Table Tennis", "Women", "doubles") == "TT-Women-Doubles"
    assert exporter._make_division_id("Pickleball 35+", "Mixed", "doubles") == "PCK35-Mixed-Doubles"
    assert exporter._make_division_id("Tennis", "Men", "anomaly") == "TEN-Men-Anomaly"
    assert exporter._make_division_id("Table Tennis 35+", "Men", "singles") == "TT35-Men-Singles"


def test_build_pod_error_lookup(mock_connectors):
    exporter = ChurchTeamsExporter()
    validation_rows = [
        {
            "Participant ID (WP)": "42",
            "sport_type": "Badminton",
            "Severity": "ERROR",
            "Status": "open",
        },
        {
            "Participant ID (WP)": "42",
            "sport_type": "Pickleball",
            "Severity": "WARNING",  # warnings excluded
            "Status": "open",
        },
        {
            "Participant ID (WP)": "99",
            "sport_type": "Table Tennis",
            "Severity": "ERROR",
            "Status": "resolved",  # resolved excluded
        },
        {
            "Participant ID (WP)": "55",
            "sport_type": "Tennis",
            "Severity": "ERROR",
            "Status": "open",
        },
    ]
    lookup = exporter._build_pod_error_lookup(validation_rows)
    assert lookup == {"42": {"Badminton"}, "55": {"Tennis"}}


def test_build_pod_divisions_rows_singles(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "1", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "2", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Single",
         "Participant ID (WP)": "3", "Church Team": "TLC"},
    ]
    # Participant 2 has an error
    validation_rows = [
        {"Participant ID (WP)": "2", "sport_type": "Badminton", "Severity": "ERROR", "Status": "open"},
    ]
    rows = exporter._build_pod_divisions_rows(roster_rows, validation_rows)

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "BAD-Men-Singles"
    assert div["sport_type"] == "Badminton"
    assert div["resource_type"] == "Badminton Court"
    assert div["planning_entries"] == 3
    assert div["confirmed_entries"] == 2  # participant 2 has error
    assert div["provisional_entries"] == 1
    assert div["anomaly_count"] == 0
    assert div["division_status"] == "Partial"


def test_build_pod_divisions_rows_doubles(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "10", "Church Team": "RPC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "11", "Church Team": "RPC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "12", "Church Team": "TLC"},
        {"sport_type": "Table Tennis", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "13", "Church Team": "TLC"},
    ]
    rows = exporter._build_pod_divisions_rows(roster_rows, [])

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "TT-Men-Doubles"
    assert div["resource_type"] == "Table Tennis Table"
    assert div["planning_entries"] == 2   # floor(4/2)
    assert div["confirmed_entries"] == 2  # no errors
    assert div["provisional_entries"] == 0
    assert div["division_status"] == "Ready"


def test_build_pod_divisions_rows_anomaly(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Pickleball", "sport_gender": "Men", "sport_format": "Team",
         "Participant ID (WP)": "20", "Church Team": "RPC"},
    ]
    rows = exporter._build_pod_divisions_rows(roster_rows, [])

    assert len(rows) == 1
    div = rows[0]
    assert div["division_id"] == "PCK-Men-Anomaly"
    assert div["planning_entries"] == 0
    assert div["confirmed_entries"] == 0
    assert div["anomaly_count"] == 1
    assert div["division_status"] == "AnomalyOnly"


def test_build_pod_entries_review_singles(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Tennis", "sport_gender": "Women", "sport_format": "Women Single",
         "Participant ID (WP)": "30", "First Name": "Lan", "Last Name": "Tran", "Church Team": "RPC"},
        {"sport_type": "Tennis", "sport_gender": "Women", "sport_format": "Women Single",
         "Participant ID (WP)": "31", "First Name": "Hoa", "Last Name": "Le", "Church Team": "RPC"},
    ]
    validation_rows = [
        {"Participant ID (WP)": "31", "sport_type": "Tennis", "Severity": "ERROR", "Status": "open"},
    ]
    rows = exporter._build_pod_entries_review_rows(roster_rows, validation_rows)

    assert len(rows) == 2
    singles = [r for r in rows if r["entry_type"] == "Singles"]
    assert len(singles) == 2

    lan = next(r for r in singles if r["participant_1_name"] == "Lan Tran")
    assert lan["review_status"] == "OK"
    assert lan["partner_status"] == "N/A"
    assert lan["division_id"] == "TEN-Women-Singles"

    hoa = next(r for r in singles if r["participant_1_name"] == "Hoa Le")
    assert hoa["review_status"] == "NeedsReview"


def test_build_pod_entries_review_doubles_reciprocal(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "40", "First Name": "Anh", "Last Name": "Nguyen",
         "partner_name": "Binh Tran", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Men", "sport_format": "Men Double",
         "Participant ID (WP)": "41", "First Name": "Binh", "Last Name": "Tran",
         "partner_name": "Anh Nguyen", "Church Team": "TLC"},
    ]
    rows = exporter._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    pair = rows[0]
    assert pair["entry_type"] == "DoublesPair"
    assert pair["partner_status"] == "Confirmed"
    assert pair["review_status"] == "OK"
    assert "Anh Nguyen" in pair["participant_1_name"] or "Binh Tran" in pair["participant_1_name"]
    assert "Anh Nguyen" in pair["participant_2_name"] or "Binh Tran" in pair["participant_2_name"]
    # cross-church pair shows both church codes
    assert "RPC" in pair["church_team"] and "TLC" in pair["church_team"]


def test_build_pod_entries_review_doubles_missing_partner(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Pickleball", "sport_gender": "Women", "sport_format": "Women Double",
         "Participant ID (WP)": "50", "First Name": "Cam", "Last Name": "Ho",
         "partner_name": "", "Church Team": "RPC"},
    ]
    rows = exporter._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    entry = rows[0]
    assert entry["entry_type"] == "UnresolvedDoubles"
    assert entry["partner_status"] == "MissingPartner"
    assert entry["review_status"] == "NeedsReview"


def test_build_pod_entries_review_doubles_non_reciprocal(mock_connectors):
    exporter = ChurchTeamsExporter()
    # A claims B, B claims someone else (non-reciprocal)
    roster_rows = [
        {"sport_type": "Badminton", "sport_gender": "Mixed", "sport_format": "Mixed Double",
         "Participant ID (WP)": "60", "First Name": "Dan", "Last Name": "Vo",
         "partner_name": "Linh Pham", "Church Team": "RPC"},
        {"sport_type": "Badminton", "sport_gender": "Mixed", "sport_format": "Mixed Double",
         "Participant ID (WP)": "61", "First Name": "Linh", "Last Name": "Pham",
         "partner_name": "Khoa Bui", "Church Team": "RPC"},  # claims Khoa, not Dan
    ]
    rows = exporter._build_pod_entries_review_rows(roster_rows, [])

    unresolved = [r for r in rows if r["entry_type"] == "UnresolvedDoubles"]
    assert len(unresolved) >= 1
    reasons = {r["partner_status"] for r in unresolved}
    assert "NonReciprocal" in reasons


def test_build_pod_entries_review_anomaly(mock_connectors):
    exporter = ChurchTeamsExporter()
    roster_rows = [
        {"sport_type": "Table Tennis 35+", "sport_gender": "Men", "sport_format": "Team",
         "Participant ID (WP)": "70", "First Name": "Tri", "Last Name": "Nguyen",
         "Church Team": "RPC"},
    ]
    rows = exporter._build_pod_entries_review_rows(roster_rows, [])

    assert len(rows) == 1
    assert rows[0]["entry_type"] == "Anomaly"
    assert rows[0]["review_status"] == "NeedsReview"
    assert "Team" in rows[0]["notes"]


def test_pod_tabs_present_in_consolidated_export(mock_connectors, tmp_path):
    """Pod-Divisions and Pod-Entries-Review tabs appear only in the ALL export."""
    exporter = ChurchTeamsExporter()

    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 2, "Total Participants (in WP)": 2,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 2, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-14",
    }]
    contacts_rows = [{
        "Church Team": "RPC", "ChMeetings ID": "101", "First Name": "Alice", "Last Name": "Nguyen",
        "Is_Participant": "Yes", "Is_Member_ChM": "Yes", "Participant ID (WP)": 1,
        "Approval_Status (WP)": "pending", "Total_Open_ERRORs (WP)": 0,
        "Gender": "Female", "Birthdate": "2000-01-02", "Age (at Event)": 26,
        "Mobile Phone": "", "Email": "", "Registration Date (WP)": "2026-03-01",
        "Athlete Fee": 30, "First_Open_ERROR_Desc (WP)": "",
        "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
        "Photo URL (WP)": "N/A", "Update_on_ChM": "",
    }]
    # Two badminton singles participants — one confirmed, one with error
    roster_rows = [
        {"Church Team": "RPC", "ChMeetings ID": "101", "Participant ID (WP)": 1,
         "First Name": "Alice", "Last Name": "Nguyen", "Gender": "Female", "Age (at Event)": 26,
         "Mobile Phone": "", "Email": "", "Is_Member_ChM": True, "Photo": "",
         "Approval_Status (WP)": "pending",
         "sport_type": "Badminton", "sport_gender": "Women", "sport_format": "Women Single",
         "team_order": None, "partner_name": None,
         "Open_TEAM_Issue_Count (WP)": 0, "Open_TEAM_Issue_Desc (WP)": ""},
        {"Church Team": "RPC", "ChMeetings ID": "102", "Participant ID (WP)": 2,
         "First Name": "Binh", "Last Name": "Le", "Gender": "Female", "Age (at Event)": 28,
         "Mobile Phone": "", "Email": "", "Is_Member_ChM": True, "Photo": "",
         "Approval_Status (WP)": "pending",
         "sport_type": "Badminton", "sport_gender": "Women", "sport_format": "Women Single",
         "team_order": None, "partner_name": None,
         "Open_TEAM_Issue_Count (WP)": 0, "Open_TEAM_Issue_Desc (WP)": ""},
    ]
    validation_rows = [
        {"Church Team": "RPC", "Rule Level": "INDIVIDUAL", "Severity": "ERROR",
         "Status": "open", "Issue Type": "missing_photo", "Rule Code": "PHOTO_REQUIRED",
         "Participant ID (WP)": 2, "ChMeetings ID": "102", "Participant Name": "Binh Le",
         "Approval_Status (WP)": "pending", "sport_type": "Badminton", "sport_format": None,
         "Issue Description": "No photo uploaded"},
    ]

    # Single-church: no pod tabs
    single_path = tmp_path / "single.xlsx"
    exporter._write_excel_report(single_path, summary_rows, contacts_rows, roster_rows, validation_rows)
    single_sheets = pd.ExcelFile(single_path).sheet_names
    assert "Pod-Divisions" not in single_sheets
    assert "Pod-Entries-Review" not in single_sheets

    # ALL export: pod tabs present
    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(all_path, summary_rows, contacts_rows, roster_rows, validation_rows,
                                 include_venue_capacity=True)
    all_sheets = pd.ExcelFile(all_path).sheet_names
    assert "Pod-Divisions" in all_sheets
    assert "Pod-Entries-Review" in all_sheets

    pod_div_df = pd.read_excel(all_path, sheet_name="Pod-Divisions")
    assert list(pod_div_df.columns)[:3] == ["division_id", "sport_type", "sport_gender"]
    assert len(pod_div_df) == 1
    row = pod_div_df.iloc[0]
    assert row["division_id"] == "BAD-Women-Singles"
    assert int(row["planning_entries"]) == 2
    assert int(row["confirmed_entries"]) == 1   # participant 2 has ERROR
    assert int(row["provisional_entries"]) == 1
    assert int(row["anomaly_count"]) == 0

    pod_entry_df = pd.read_excel(all_path, sheet_name="Pod-Entries-Review")
    assert "entry_type" in pod_entry_df.columns
    assert len(pod_entry_df) == 2
    assert set(pod_entry_df["entry_type"]) == {"Singles"}
    ok_rows = pod_entry_df[pod_entry_df["review_status"] == "OK"]
    needs_review_rows = pod_entry_df[pod_entry_df["review_status"] == "NeedsReview"]
    assert len(ok_rows) == 1
    assert len(needs_review_rows) == 1


def test_court_schedule_sketch_tab_present(mock_connectors, tmp_path):
    """Court-Schedule-Sketch tab appears only in consolidated ALL export."""
    exporter = ChurchTeamsExporter()
    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 6, "Total Participants (in WP)": 6,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 6, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-13",
    }]
    roster_rows = [
        {"Church Team": "RPC", "ChMeetings ID": str(100 + i), "Participant ID (WP)": i,
         "sport_type": "Basketball", "sport_gender": "Men", "sport_format": "Team"}
        for i in range(6)
    ]

    single_path = tmp_path / "single.xlsx"
    exporter._write_excel_report(single_path, summary_rows, [], roster_rows, [])
    assert "Court-Schedule-Sketch" not in pd.ExcelFile(single_path).sheet_names

    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(all_path, summary_rows, [], roster_rows, [],
                                 include_venue_capacity=True)
    sheets = pd.ExcelFile(all_path).sheet_names
    assert "Court-Schedule-Sketch" in sheets


def test_court_schedule_sketch_structure(mock_connectors, tmp_path):
    """Court-Schedule-Sketch tab has three scenario blocks with correct structure."""
    from config import SCHEDULE_SKETCH_N_COURTS
    exporter = ChurchTeamsExporter()
    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 6, "Total Participants (in WP)": 6,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 6, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-13",
    }]
    # Six BBM, six VBM, six VBW players → one team each, falls back to 8 teams for planning
    sports = [
        ("Basketball", "Men"),
        ("Volleyball", "Men"),
        ("Volleyball", "Women"),
    ]
    roster_rows = []
    for sport_type, sport_gender in sports:
        for i in range(6):
            roster_rows.append({
                "Church Team": "RPC", "ChMeetings ID": str(200 + len(roster_rows)),
                "Participant ID (WP)": len(roster_rows),
                "sport_type": sport_type, "sport_gender": sport_gender, "sport_format": "Team",
            })

    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(all_path, summary_rows, [], roster_rows, [],
                                 include_venue_capacity=True)

    from openpyxl import load_workbook
    wb = load_workbook(all_path)
    ws = wb["Court-Schedule-Sketch"]

    # Row 3 should contain scenario headers for each court count
    row3_values = [ws.cell(row=3, column=c).value for c in range(1, 20)]
    scenario_headers = [v for v in row3_values if v and "Scenario" in str(v)]
    assert len(scenario_headers) == len(SCHEDULE_SKETCH_N_COURTS)
    for n_courts, hdr in zip(SCHEDULE_SKETCH_N_COURTS, scenario_headers):
        assert str(n_courts) in str(hdr)

    # Row 4 should have "Time" and "Court N" sub-headers in each block
    row4_values = [ws.cell(row=4, column=c).value for c in range(1, 20)]
    time_headers = [v for v in row4_values if v == "Time"]
    court_headers = [v for v in row4_values if v and str(v).startswith("Court")]
    assert len(time_headers) == len(SCHEDULE_SKETCH_N_COURTS)
    # Total court columns = 3 + 4 + 5 = 12
    assert len(court_headers) == sum(SCHEDULE_SKETCH_N_COURTS)

    # Section labels ("1st Saturday" etc.) must appear somewhere in the sheet
    all_values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                all_values.append(str(cell.value))
    assert any("1st Saturday" in v for v in all_values)
    assert any("1st Sunday" in v for v in all_values)
    assert any("2nd Saturday" in v for v in all_values)
    assert any("2nd Sunday" in v for v in all_values)


def test_build_scenario_schedule_pool_before_playoffs(mock_connectors):
    """Pool before early playoffs, early playoffs on sat2, finals pinned to sun2."""
    pool_queues = [
        [f"BBM-{i:02d}" for i in range(1, 5)],
        [f"VBM-{i:02d}" for i in range(1, 5)],
        [f"VBW-{i:02d}" for i in range(1, 5)],
    ]
    early_playoff_queues = [
        ["BBM-Semi-1", "BBM-Semi-2"],
        ["VBM-Semi-1", "VBM-Semi-2"],
        ["VBW-Semi-1", "VBW-Semi-2"],
    ]
    final_queues = [
        ["BBM-Final"],
        ["VBM-Final"],
        ["VBW-Final"],
    ]

    n_sat, n_sun = 13, 8

    for n_courts in [3, 4, 5]:
        grids = ChurchTeamsExporter._build_scenario_schedule(
            n_courts, pool_queues, early_playoff_queues, final_queues, n_sat, n_sun
        )
        sat1_cells = [cell for row in grids[0] for cell in row if cell]
        sun1_cells = [cell for row in grids[1] for cell in row if cell]
        sat2_cells = [cell for row in grids[2] for cell in row if cell]
        sun2_cells = [cell for row in grids[3] for cell in row if cell]

        all_pool  = {g for q in pool_queues for g in q}
        all_early = {g for q in early_playoff_queues for g in q}
        all_final = {g for q in final_queues for g in q}

        # All pool games appear in sat1/sun1/sat2
        assert set(sat1_cells + sun1_cells + sat2_cells) & all_pool == all_pool, \
            f"n_courts={n_courts}: missing pool games"

        # No playoff/final games in sat1 or sun1
        assert not (set(sat1_cells) & (all_early | all_final)), \
            f"n_courts={n_courts}: playoff/final in sat1"
        assert not (set(sun1_cells) & (all_early | all_final)), \
            f"n_courts={n_courts}: playoff/final in sun1"

        # Early playoffs land on sat2, not sun2
        assert set(sat2_cells) & all_early == all_early, \
            f"n_courts={n_courts}: early playoffs missing from sat2"
        assert not (set(sun2_cells) & all_early), \
            f"n_courts={n_courts}: early playoffs leaked into sun2"

        # Finals land on sun2, not sat2
        assert set(sun2_cells) & all_final == all_final, \
            f"n_courts={n_courts}: finals missing from sun2"
        assert not (set(sat2_cells) & all_final), \
            f"n_courts={n_courts}: finals leaked into sat2"

        # Pool games never appear in sun2
        assert not (set(sun2_cells) & all_pool), \
            f"n_courts={n_courts}: pool games leaked into sun2"

        # Playoffs/finals stay on their primary court blocks
        n_sports = len(pool_queues)
        base = n_courts // n_sports
        extras = n_courts % n_sports
        cur = 0
        for sport_idx, (early_q, final_q) in enumerate(zip(early_playoff_queues, final_queues)):
            k = base + (1 if sport_idx < extras else 0)
            sport_courts = set(range(cur, cur + k))
            cur += k
            playoff_ids = set(early_q) | set(final_q)
            for sess_idx in [2, 3]:  # sat2, sun2 only
                for t, row in enumerate(grids[sess_idx]):
                    for c_idx, game_id in enumerate(row):
                        if game_id in playoff_ids:
                            assert c_idx in sport_courts, (
                                f"n_courts={n_courts} sport={sport_idx}: "
                                f"{game_id} on court {c_idx}, expected {sport_courts}"
                            )


def test_court_schedule_sketch_game_id_prefixes(mock_connectors, tmp_path):
    """Game IDs use BBM, VBM, VBW prefixes with two-digit sequential numbering."""
    exporter = ChurchTeamsExporter()

    # Provide enough players for 2 teams per sport (min 6 for VB)
    roster_rows = []
    sports = [
        ("Basketball", "Men"),
        ("Volleyball", "Men"),
        ("Volleyball", "Women"),
    ]
    for sport_type, sport_gender in sports:
        for i in range(12):
            roster_rows.append({
                "Church Team": f"CH{i}",
                "ChMeetings ID": str(300 + len(roster_rows)),
                "Participant ID (WP)": len(roster_rows),
                "sport_type": sport_type,
                "sport_gender": sport_gender,
                "sport_format": "Team",
            })

    all_path = tmp_path / "all.xlsx"
    summary_rows = [{
        "Church Code": "CH0",
        "Total Members (ChM Team Group)": 6, "Total Participants (in WP)": 6,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 6, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-13",
    }]
    exporter._write_excel_report(all_path, summary_rows, [], roster_rows, [],
                                 include_venue_capacity=True)

    from openpyxl import load_workbook
    wb = load_workbook(all_path)
    ws = wb["Court-Schedule-Sketch"]

    cell_values = set()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "-" in v and v.split("-")[0] in ("BBM", "VBM", "VBW"):
                cell_values.add(v)

    bbm_ids = {v for v in cell_values if v.startswith("BBM-")}
    vbm_ids = {v for v in cell_values if v.startswith("VBM-")}
    vbw_ids = {v for v in cell_values if v.startswith("VBW-")}

    assert bbm_ids, "Expected BBM game IDs in Court-Schedule-Sketch"
    assert vbm_ids, "Expected VBM game IDs in Court-Schedule-Sketch"
    assert vbw_ids, "Expected VBW game IDs in Court-Schedule-Sketch"

    # Pool IDs: BBM-01, BBM-02, … — two-digit numeric suffix after the dash
    pool_ids = {v for v in cell_values if v[4:].isdigit() and len(v[4:]) == 2}
    assert pool_ids, "Expected pool game IDs with two-digit suffix (e.g. BBM-01)"

    # Playoff IDs: named labels (Final, Semi-N, QF-N, 3rd)
    playoff_labels = {"Final", "Semi-1", "Semi-2", "QF-1", "QF-2", "QF-3", "QF-4", "3rd"}
    found_playoff_labels = {v.split("-", 1)[1] for v in cell_values if not v[4:].isdigit()}
    assert found_playoff_labels & playoff_labels, (
        f"Expected named playoff labels in Court-Schedule-Sketch, got: {found_playoff_labels}"
    )
    assert "Final" in found_playoff_labels, "Expected BBM/VBM/VBW-Final in sketch"


# ── Pod-Resource-Estimate tests (Issue #86) ─────────────────────────────────


def test_pod_resource_estimate_tab_present(mock_connectors, tmp_path):
    """Pod-Resource-Estimate tab appears only in the consolidated ALL export."""
    exporter = ChurchTeamsExporter()
    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 6, "Total Participants (in WP)": 6,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 6, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-13",
    }]

    single_path = tmp_path / "single.xlsx"
    exporter._write_excel_report(single_path, summary_rows, [], [], [])
    assert "Pod-Resource-Estimate" not in pd.ExcelFile(single_path).sheet_names

    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(all_path, summary_rows, [], [], [], include_venue_capacity=True)
    assert "Pod-Resource-Estimate" in pd.ExcelFile(all_path).sheet_names


def test_pod_resource_estimate_no_venue_input(mock_connectors, tmp_path):
    """When no venue_input.xlsx exists, tab still renders with notice row."""
    exporter = ChurchTeamsExporter()
    # Provide zero registrations — all entries will be 0.
    available = {}  # empty — no venue file
    pod_rows = exporter._build_pod_resource_rows([], available)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    exporter._write_pod_resource_estimate(ws, pod_rows, available)

    # All Fit Status cells say "No venue data"
    fit_values = {ws.cell(row=r, column=7).value for r in range(2, 2 + len(pod_rows))}
    assert fit_values == {"No venue data"}, f"Unexpected fit values: {fit_values}"


def test_pod_resource_estimate_fit_status_rules(mock_connectors):
    """Green/Yellow/Red thresholds from POD_FIT_YELLOW_MAX (= 3)."""
    exporter = ChurchTeamsExporter()
    from config import POD_RESOURCE_TYPE_BADMINTON, POD_FIT_YELLOW_MAX

    # Build roster rows: 10 Badminton singles → Required = 9
    roster_rows = [
        {"sport_type": "Badminton", "sport_format": "Men Singles"} for _ in range(10)
    ]

    # Green: available >= required (9)
    green = exporter._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9}
    )
    badminton_green = next(r for r in green if "Badminton" in r["Event"])
    assert badminton_green["Fit Status"] == "Green"
    assert badminton_green["Surplus / Shortage"] == 0

    # Yellow: short by 1 to POD_FIT_YELLOW_MAX (3)
    yellow = exporter._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9 - POD_FIT_YELLOW_MAX}
    )
    badminton_yellow = next(r for r in yellow if "Badminton" in r["Event"])
    assert badminton_yellow["Fit Status"] == "Yellow"

    # Red: short by more than POD_FIT_YELLOW_MAX
    red = exporter._build_pod_resource_rows(
        roster_rows, {POD_RESOURCE_TYPE_BADMINTON: 9 - POD_FIT_YELLOW_MAX - 1}
    )
    badminton_red = next(r for r in red if "Badminton" in r["Event"])
    assert badminton_red["Fit Status"] == "Red"


def test_load_venue_input_aggregates_by_resource_type(mock_connectors, tmp_path):
    """Available Slots are summed across multiple rows of the same resource type."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_PICKLEBALL

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # Two Pickleball Court rows: 24 + 18 = 42 total
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=2, column=9, value=24)
    ws.cell(row=3, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=3, column=9, value=18)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input(path)
    assert result[POD_RESOURCE_TYPE_PICKLEBALL] == 42


def test_load_venue_input_fallback_formula(mock_connectors, tmp_path):
    """When Available Slots is zero/missing, compute from Quantity/times/Slot Minutes."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # 4 courts, 13:00–18:00, 60-min slots → (18-13)*60/60 + 1 = 6 starts → 4*6 = 24
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=4)    # Quantity
    ws.cell(row=2, column=6, value=13)   # Start Time (decimal hour)
    ws.cell(row=2, column=7, value=18)   # Last Start Time
    ws.cell(row=2, column=8, value=60)   # Slot Minutes
    ws.cell(row=2, column=9, value=0)    # Available Slots = 0 → triggers fallback

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input(path)
    assert result[POD_RESOURCE_TYPE_TENNIS] == 24


def test_load_venue_input_ignores_blank_resource_rows(mock_connectors, tmp_path):
    """Blank resource rows should be ignored instead of creating a literal 'nan' bucket."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_PICKLEBALL

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_PICKLEBALL)
    ws.cell(row=2, column=9, value=24)
    ws.cell(row=3, column=3, value=None)
    ws.cell(row=3, column=9, value=None)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input(path)
    assert result == {POD_RESOURCE_TYPE_PICKLEBALL: 24}


# ── Schedule-Input helpers tests (Issue #87) ────────────────────────────────


def _make_gym_roster(n_churches: int = 8) -> list:
    """Return minimal Basketball-Men roster rows for n_churches churches."""
    codes = ["RPC", "ANH", "FVC", "GAC", "NSD", "TLC", "GLA", "ORN"][:n_churches]
    rows = []
    for code in codes:
        for _ in range(5):  # 5 members per church → meets min team size
            rows.append({
                "Church Team": code,
                "sport_type": SPORT_TYPE["BASKETBALL"],
                "sport_gender": "Men",
                "sport_format": "Team",
                "Participant ID (WP)": 1,
            })
    return rows


def test_build_gym_game_objects_structure(mock_connectors):
    """Each game object has all required OR-Tools schema fields."""
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster())
    assert games, "Expected at least one game"
    required_fields = {
        "game_id", "event", "stage", "pool_id", "round",
        "team_a_id", "team_b_id", "duration_minutes",
        "resource_type", "earliest_slot", "latest_slot",
    }
    for g in games:
        assert required_fields <= g.keys(), f"Missing fields in {g}"
    from config import GYM_RESOURCE_TYPE
    assert all(g["resource_type"] == GYM_RESOURCE_TYPE for g in games)
    # team_a_id and team_b_id must be non-null strings for all games (pool + playoff)
    assert all(
        isinstance(g["team_a_id"], str) and isinstance(g["team_b_id"], str)
        for g in games
    ), "All games must have non-null team_a_id and team_b_id"


def test_build_gym_game_objects_stages(mock_connectors):
    """With 8 BBM teams, only Pool stage is present (playoffs go in Playoff-Slots tab)."""
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster(8))
    bbm_stages = {g["stage"] for g in games if g["event"] == SPORT_TYPE["BASKETBALL"]}
    assert bbm_stages == {"Pool"}, f"Expected only Pool stage; got {bbm_stages}"


def test_build_gym_game_objects_prefix_format(mock_connectors):
    """Pool game IDs follow the BBM-01 format."""
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster())
    pool_ids = [g["game_id"] for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    assert pool_ids, "Expected Basketball pool games"
    import re
    assert all(re.match(r"BBM-\d{2}$", gid) for gid in pool_ids)


def test_build_gym_game_objects_stable_team_ids(mock_connectors):
    """The same placeholder team ID is reused across multiple pool games for that team."""
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster(8))
    bbm_pool = [g for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    assert bbm_pool, "Expected Basketball pool games"

    # Collect all team IDs and count appearances
    from collections import Counter
    appearances: Counter = Counter()
    for g in bbm_pool:
        appearances[g["team_a_id"]] += 1
        appearances[g["team_b_id"]] += 1

    # The normalized 2-game policy keeps every team at exactly 2 pool games.
    assert appearances, "Expected stable placeholder team IDs to be reused"
    assert all(count == 2 for count in appearances.values()), appearances


def test_build_gym_game_objects_pool_id_nonempty(mock_connectors):
    """Pool games carry a non-empty pool_id; playoff/final games have empty pool_id."""
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster(8))
    pool_games = [g for g in games if g["stage"] == "Pool"]
    playoff_games = [g for g in games if g["stage"] not in ("Pool",)]

    assert pool_games, "Expected pool games"
    assert all(g["pool_id"] != "" for g in pool_games), "Pool games must have non-empty pool_id"
    assert all(g["pool_id"] == "" for g in playoff_games), "Playoff/final games must have empty pool_id"


def test_build_gym_game_objects_team_id_format(mock_connectors):
    """Pool game team IDs follow the stable placeholder format PREFIX-Px-Ty."""
    import re as _re
    exporter = ChurchTeamsExporter()
    games = exporter._build_gym_game_objects(_make_gym_roster(8))
    bbm_pool = [g for g in games if g["stage"] == "Pool" and g["event"] == SPORT_TYPE["BASKETBALL"]]
    for g in bbm_pool:
        assert _re.match(r"BBM-P\d+-T\d+$", g["team_a_id"]), f"Unexpected team_a_id: {g['team_a_id']}"
        assert _re.match(r"BBM-P\d+-T\d+$", g["team_b_id"]), f"Unexpected team_b_id: {g['team_b_id']}"


def test_build_schedule_input_gym_court_scenario(mock_connectors, tmp_path):
    """gym_court_scenario in schedule_input matches SCHEDULE_SOLVER_GYM_COURTS."""
    from config import SCHEDULE_SOLVER_GYM_COURTS
    exporter = ChurchTeamsExporter()
    si = exporter._build_schedule_input(_make_gym_roster(), [], tmp_path / "missing.xlsx")
    assert si["gym_court_scenario"] == SCHEDULE_SOLVER_GYM_COURTS
    gym_resources = [r for r in si["resources"] if r["resource_type"] == "Gym Court"]
    n_sessions = 4  # Sat-1, Sun-1, Sat-2, Sun-2
    assert len(gym_resources) == SCHEDULE_SOLVER_GYM_COURTS * n_sessions


def test_build_pod_game_objects_single_elimination(mock_connectors):
    """With 3 entries in a division, 2 game placeholders are generated."""
    from config import POD_RESOURCE_EVENT_TYPE

    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": i,
         "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Women",
         "sport_format": "Women Single"}
        for i in range(1, 4)  # 3 entries
    ]
    exporter = ChurchTeamsExporter()
    games = exporter._build_pod_game_objects(roster_rows, [])
    assert len(games) == 2, f"Expected 2 games (3-1=2), got {len(games)}"
    assert all(g["game_id"].startswith("BAD-Women-Singles-") for g in games)
    assert all(g["stage"] == "R1" for g in games)
    assert all(g["resource_type"] == POD_RESOURCE_EVENT_TYPE[SPORT_TYPE["BADMINTON"]] for g in games)
    assert games[0]["game_id"] == "BAD-Women-Singles-01"


def test_build_pod_game_objects_skips_empty_divisions(mock_connectors):
    """Divisions with fewer than 2 entries produce no game objects."""
    roster_rows = [
        {"Church Team": "RPC", "Participant ID (WP)": 1,
         "sport_type": SPORT_TYPE["BADMINTON"], "sport_gender": "Men",
         "sport_format": "Men Single"},
    ]
    exporter = ChurchTeamsExporter()
    games = exporter._build_pod_game_objects(roster_rows, [])
    assert games == [], "Single-entry division should produce no games"


def test_build_gym_resource_objects_count(mock_connectors):
    """4 sessions × n_courts resources are returned."""
    resources = ChurchTeamsExporter._build_gym_resource_objects(n_courts=4)
    assert len(resources) == 16, f"Expected 4 sessions × 4 courts = 16, got {len(resources)}"
    from config import GYM_RESOURCE_TYPE
    assert all(r["resource_type"] == GYM_RESOURCE_TYPE for r in resources)
    days = {r["day"] for r in resources}
    assert days == {"Sat-1", "Sun-1", "Sat-2", "Sun-2"}


def test_build_gym_resource_objects_labels(mock_connectors):
    """Court labels and resource_ids are formatted correctly."""
    resources = ChurchTeamsExporter._build_gym_resource_objects(n_courts=3)
    labels = {r["label"] for r in resources}
    assert labels == {"Court-1", "Court-2", "Court-3"}
    ids = {r["resource_id"] for r in resources}
    assert "GYM-Sat-1-1" in ids
    assert "GYM-Sun-2-3" in ids


def test_build_gym_resource_objects_include_blank_exclusive_group(mock_connectors):
    """Gym resources carry the same exclusive_group field as venue-loaded resources."""
    resources = ChurchTeamsExporter._build_gym_resource_objects(n_courts=2)
    assert resources
    assert all("exclusive_group" in r for r in resources)
    assert all(r["exclusive_group"] == "" for r in resources)


def test_load_venue_input_rows_missing_file(mock_connectors, tmp_path):
    """Returns empty list when venue_input.xlsx does not exist."""
    result = ChurchTeamsExporter._load_venue_input_rows(tmp_path / "missing.xlsx")
    assert result == []


def test_load_venue_input_rows_expands_quantity(mock_connectors, tmp_path):
    """Quantity=2 for a Tennis Court row yields 2 resource objects."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=2)    # Quantity = 2
    ws.cell(row=2, column=6, value=9)    # Start Time
    ws.cell(row=2, column=7, value=17)   # Last Start Time
    ws.cell(row=2, column=8, value=30)   # Slot Minutes

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input_rows(path)
    assert len(result) == 2
    assert result[0]["resource_type"] == POD_RESOURCE_TYPE_TENNIS
    assert result[0]["label"] == "Court-1"
    assert result[1]["label"] == "Court-2"
    assert result[0]["open_time"] == "09:00"


def test_load_venue_input_rows_table_label(mock_connectors, tmp_path):
    """Table Tennis Table rows get Table-N labels instead of Court-N."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TABLE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TABLE_TENNIS)
    ws.cell(row=2, column=4, value=3)
    ws.cell(row=2, column=6, value=9)
    ws.cell(row=2, column=7, value=17)
    ws.cell(row=2, column=8, value=20)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input_rows(path)
    assert len(result) == 3
    assert all(r["label"].startswith("Table-") for r in result)


def test_load_venue_input_rows_skips_blank_resource_rows(mock_connectors, tmp_path):
    """Blank/NaN venue rows should be ignored instead of crashing Schedule-Input generation."""
    from openpyxl import Workbook
    from config import POD_RESOURCE_TYPE_TENNIS

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=3, value=POD_RESOURCE_TYPE_TENNIS)
    ws.cell(row=2, column=4, value=2)
    ws.cell(row=2, column=6, value=9)
    ws.cell(row=2, column=7, value=17)
    ws.cell(row=2, column=8, value=30)
    ws.cell(row=3, column=3, value=None)
    ws.cell(row=3, column=4, value=None)
    ws.cell(row=3, column=6, value=None)
    ws.cell(row=3, column=7, value=None)
    ws.cell(row=3, column=8, value=None)

    path = tmp_path / "venue_input.xlsx"
    wb.save(path)

    result = ChurchTeamsExporter._load_venue_input_rows(path)
    assert len(result) == 2
    assert all(r["resource_type"] == POD_RESOURCE_TYPE_TENNIS for r in result)


def _write_venue_input(path, headers, data_rows, gym_modes_rows=None):
    """Write a venue_input.xlsx with a Venue-Input tab (and optional Gym-Modes)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, data in enumerate(data_rows, start=2):
        for c, val in enumerate(data, start=1):
            ws.cell(row=r, column=c, value=val)
    if gym_modes_rows is not None:
        gm = wb.create_sheet("Gym-Modes")
        for r, data in enumerate(gym_modes_rows, start=1):
            for c, val in enumerate(data, start=1):
                gm.cell(row=r, column=c, value=val)
    wb.save(path)


def test_load_venue_input_rows_reads_exclusive_group(mock_connectors, tmp_path):
    """Exclusive Venue Group column is attached to each emitted resource object."""
    from config import POD_RESOURCE_TYPE_TENNIS

    headers = [
        "Pod Name", "Venue Name", "Exclusive Venue Group", "Resource Type",
        "Quantity", "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["BB Pod", "Midsize Gym", "Midsize Gym", POD_RESOURCE_TYPE_TENNIS,
         2, None, 9, 17, 30, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result = ChurchTeamsExporter._load_venue_input_rows(path)
    assert len(result) == 2
    assert all(r["exclusive_group"] == "Midsize Gym" for r in result)


def test_load_venue_input_rows_blank_exclusive_group(mock_connectors, tmp_path):
    """A row with no Exclusive Venue Group yields an empty-string group."""
    from config import POD_RESOURCE_TYPE_TENNIS

    headers = [
        "Pod Name", "Venue Name", "Exclusive Venue Group", "Resource Type",
        "Quantity", "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]
    rows = [
        ["Tennis Pod", "Chapman", None, POD_RESOURCE_TYPE_TENNIS,
         1, None, 9, 17, 60, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, rows)

    result = ChurchTeamsExporter._load_venue_input_rows(path)
    assert result[0]["exclusive_group"] == ""


def test_load_gym_modes_missing_file(mock_connectors, tmp_path):
    """Returns empty dict when venue_input.xlsx does not exist."""
    assert ChurchTeamsExporter._load_gym_modes(tmp_path / "missing.xlsx") == {}


def test_load_gym_modes_missing_tab(mock_connectors, tmp_path):
    """File present but no Gym-Modes tab → empty dict (warning, no crash)."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]])
    assert ChurchTeamsExporter._load_gym_modes(path) == {}


def test_load_gym_modes_reads_capacities(mock_connectors, tmp_path):
    """Gym-Modes tab is parsed into {gym: {resource_type: courts_per_block}}."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts",
         "Badminton Courts", "Pickleball Courts", "Soccer Fields", "Notes"],
        ["Midsize Gym", 1, 2, 6, 8, 1, "either-or"],
        ["Big Gym", 2, 3, 12, 0, 0, "larger"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ChurchTeamsExporter._load_gym_modes(path)
    assert result["Midsize Gym"] == {
        "Basketball Court": 1, "Volleyball Court": 2,
        "Badminton Court": 6, "Pickleball Court": 8, "Soccer Field": 1,
    }
    assert result["Big Gym"]["Volleyball Court"] == 3
    assert result["Big Gym"]["Pickleball Court"] == 0


def test_load_gym_modes_trims_header_whitespace(mock_connectors, tmp_path):
    """Operator-edited headers with trailing spaces are normalized before row access."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name ", "Basketball Courts ", "Volleyball Courts ", "Notes "],
        ["Midsize Gym", 1, 2, "either-or"],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ChurchTeamsExporter._load_gym_modes(path)
    assert result["Midsize Gym"]["Basketball Court"] == 1
    assert result["Midsize Gym"]["Volleyball Court"] == 2


def test_load_gym_modes_skips_note_row(mock_connectors, tmp_path):
    """A footer row with text in Gym Name but no capacities is ignored."""
    headers = ["Pod Name", "Resource Type", "Quantity"]
    gym_modes = [
        ["Gym Name", "Basketball Courts", "Volleyball Courts",
         "Badminton Courts", "Pickleball Courts", "Soccer Fields", "Notes"],
        ["Midsize Gym", 1, 2, 6, 8, 1, "either-or"],
        ["Capacity-per-mode coefficients for the LP estimator.",
         None, None, None, None, None, None],
    ]
    path = tmp_path / "venue_input.xlsx"
    _write_venue_input(path, headers, [["P", "Tennis Court", 1]], gym_modes)

    result = ChurchTeamsExporter._load_gym_modes(path)
    assert list(result.keys()) == ["Midsize Gym"]


def test_build_schedule_input_keys(mock_connectors, tmp_path):
    """_build_schedule_input returns dict with all required top-level keys."""
    exporter = ChurchTeamsExporter()
    si = exporter._build_schedule_input(_make_gym_roster(), [], tmp_path / "missing.xlsx")
    assert set(si.keys()) == {
        "generated_at", "gym_court_scenario", "game_count", "resource_count",
        "games", "resources", "playoff_slots", "gym_modes",
    }
    assert si["game_count"] == len(si["games"])
    assert si["resource_count"] == len(si["resources"])
    assert si["game_count"] > 0
    assert si["resource_count"] > 0  # at least gym resources


def test_schedule_input_tab_in_consolidated_export(mock_connectors, tmp_path):
    """Schedule-Input tab and schedule_input.json are written in the ALL export."""
    exporter = ChurchTeamsExporter()

    summary_rows = [{
        "Church Code": "RPC",
        "Total Members (ChM Team Group)": 5, "Total Participants (in WP)": 5,
        "Total Approved (WP)": 0, "Total Pending Approval (WP)": 5, "Total Denied (WP)": 0,
        "Total Participants w/ Open ERRORs (WP)": 0,
        "Total Open Individual ERRORs (WP)": 0, "Total Open TEAM ERRORs (WP)": 0,
        "Total Open WARNINGs (WP)": 0, "Total Sports w/ Open TEAM Issues (WP)": 0,
        "Latest ChM Record Update for Team": "2026-05-14",
        "Total Athlete Fees": 150,
    }]
    contacts_rows = [{
        "Church Team": "RPC", "ChMeetings ID": "101", "First Name": "Alice", "Last Name": "Ng",
        "Is_Participant": "Yes", "Is_Member_ChM": "Yes", "Participant ID (WP)": 1,
        "Approval_Status (WP)": "pending", "Total_Open_ERRORs (WP)": 0,
        "Gender": "Female", "Birthdate": "2000-01-02", "Age (at Event)": 26,
        "Mobile Phone": "", "Email": "", "Registration Date (WP)": "2026-03-01",
        "Athlete Fee": 30, "First_Open_ERROR_Desc (WP)": "",
        "Box 1": "", "Box 2": "", "Box 3": "", "Box 4": "", "Box 5": "", "Box 6": "",
        "Photo URL (WP)": "N/A", "Update_on_ChM": "",
    }]

    all_path = tmp_path / "all.xlsx"
    exporter._write_excel_report(
        all_path, summary_rows, contacts_rows,
        _make_gym_roster(4), [],
        include_venue_capacity=True,
    )

    sheets = pd.ExcelFile(all_path).sheet_names
    assert "Schedule-Input" in sheets, f"Schedule-Input tab missing; sheets: {sheets}"

    json_path = tmp_path / "schedule_input.json"
    assert json_path.exists(), "schedule_input.json was not written"

    import json as _json
    data = _json.loads(json_path.read_text(encoding="utf-8"))
    assert "games" in data and "resources" in data and "playoff_slots" in data
    assert "gym_court_scenario" in data
    assert data["game_count"] > 0
    assert data["resource_count"] == len(data["resources"])
    assert data["resource_count"] > 0


def test_schedule_input_tab_absent_in_single_church_export(mock_connectors, tmp_path):
    """Schedule-Input tab does NOT appear in single-church exports."""
    exporter = ChurchTeamsExporter()
    path = tmp_path / "single.xlsx"
    exporter._write_excel_report(path, [], [], _make_gym_roster(2), [])
    sheets = pd.ExcelFile(path).sheet_names
    assert "Schedule-Input" not in sheets


# ---------------------------------------------------------------------------
# Issue #94 — _build_schedule_output_flat_rows
# ---------------------------------------------------------------------------

def _make_schedule_pair():
    """Return (schedule_output, schedule_input) test fixtures."""
    schedule_input = {
        "games": [
            {
                "game_id": "BBM-01", "event": "Basketball - Men Team",
                "stage": "Pool", "pool_id": "P1", "round": 1,
                "team_a_id": "BBM-P1-T1", "team_b_id": "BBM-P1-T2",
                "duration_minutes": 60, "resource_type": "Gym Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "BBM-Final", "event": "Basketball - Men Team",
                "stage": "Final", "pool_id": "", "round": 1,
                "team_a_id": "WIN-BBM-Semi-1", "team_b_id": "WIN-BBM-Semi-2",
                "duration_minutes": 60, "resource_type": "Gym Court",
                "earliest_slot": None, "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "GYM-Sat-1-1", "resource_type": "Gym Court",
                "label": "Court-1", "day": "Sat-1",
                "open_time": "08:00", "close_time": "12:00", "slot_minutes": 60,
            }
        ],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-01T10:00:00",
        "status": "OPTIMAL",
        "solver_wall_seconds": 0.1,
        "assignments": [
            {"game_id": "BBM-01",    "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-08:00"},
            {"game_id": "BBM-Final", "resource_id": "GYM-Sat-1-1", "slot": "Sat-1-10:00"},
        ],
        "unscheduled": [],
    }
    return schedule_output, schedule_input


def test_build_schedule_output_flat_rows_count():
    """Returns one row per assignment."""
    from church_teams_export import ChurchTeamsExporter
    so, si = _make_schedule_pair()
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    assert len(rows) == 2


def test_build_schedule_output_flat_rows_fields():
    """Each row contains expected keys with non-empty event."""
    from church_teams_export import ChurchTeamsExporter
    so, si = _make_schedule_pair()
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    required = {"game_id", "event", "stage", "round", "team_a_id", "team_b_id",
                "resource_label", "day", "slot", "duration_minutes"}
    for row in rows:
        assert required.issubset(row.keys())
        assert row["event"] == "Basketball - Men Team"


def test_build_schedule_output_flat_rows_sorted():
    """Rows are sorted Pool before Final (stage order)."""
    from church_teams_export import ChurchTeamsExporter
    so, si = _make_schedule_pair()
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    stages = [r["stage"] for r in rows]
    assert stages == ["Pool", "Final"]


def test_build_schedule_output_flat_rows_time_part():
    """The slot field extracts the HH:MM part from the full slot label."""
    from church_teams_export import ChurchTeamsExporter
    so, si = _make_schedule_pair()
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    pool_row = next(r for r in rows if r["game_id"] == "BBM-01")
    assert pool_row["slot"] == "08:00"


def test_build_schedule_output_flat_rows_day_display():
    """Sat-1 is translated to '1st Sat'."""
    from church_teams_export import ChurchTeamsExporter
    so, si = _make_schedule_pair()
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    assert all(r["day"] == "1st Sat" for r in rows)


def test_build_schedule_output_flat_rows_empty():
    """Empty assignments list returns empty rows."""
    from church_teams_export import ChurchTeamsExporter
    so = {"assignments": [], "unscheduled": []}
    si = {"games": [], "resources": [], "precedence": []}
    rows = ChurchTeamsExporter._build_schedule_output_flat_rows(so, si)
    assert rows == []


# ---------------------------------------------------------------------------
# B3 — _compute_court_slots matches _make_pool_game_pairs actual count
# ---------------------------------------------------------------------------

def test_compute_court_slots_matches_normalized_policy_8teams_gpg2(mock_connectors):
    """8 teams / gpg=2 now stays at 8 pool games under the normalized policy."""
    from church_teams_export import ChurchTeamsExporter

    n_teams, gpg = 8, 2
    actual = len(ChurchTeamsExporter._make_pool_game_pairs("_", n_teams, gpg))
    assert actual == 8  # pools [4,4] with 4-match matrices -> 4 + 4

    exporter = ChurchTeamsExporter()
    s_formula = exporter._compute_court_slots(n_teams, pool_games_per_team=gpg)
    s_actual  = exporter._compute_court_slots(n_teams, pool_games_per_team=gpg,
                                              actual_pool_games=actual)
    assert s_formula["pool_slots"] == 8
    assert s_actual["pool_slots"]  == 8


def test_make_pool_game_pairs_exact_two_games_per_team(mock_connectors):
    """Normalized B4 policy keeps every team at the same exact pool-game count."""
    from church_teams_export import ChurchTeamsExporter
    from collections import Counter

    cases = [
        (3, 2, 2, 3),
        (4, 2, 2, 4),
        (5, 2, 2, 5),
        (7, 2, 2, 7),
        (8, 2, 2, 8),
        (12, 2, 2, 12),
        (20, 2, 2, 20),
    ]
    for n_teams, gpg, expected_games_per_team, expected_total_games in cases:
        pairs = ChurchTeamsExporter._make_pool_game_pairs("T", n_teams, gpg)
        games_per_team: Counter = Counter()
        for a, b, _ in pairs:
            games_per_team[a] += 1
            games_per_team[b] += 1
        assert len(pairs) == expected_total_games, (
            f"n_teams={n_teams}: total pool games {len(pairs)} != {expected_total_games}"
        )
        assert games_per_team, f"Expected generated games for n_teams={n_teams}"
        assert all(count == expected_games_per_team for count in games_per_team.values()), (
            f"n_teams={n_teams}, gpg={gpg}: {dict(games_per_team)}"
        )


def test_make_pool_game_pairs_direct_match_for_two_teams(mock_connectors):
    """Two teams stay as a single direct match rather than an over-built pool."""
    from church_teams_export import ChurchTeamsExporter

    pairs = ChurchTeamsExporter._make_pool_game_pairs("T", 2, 2)
    assert pairs == [("T-P1-T1", "T-P1-T2", "P1")]


def test_make_pool_game_pairs_legacy_fallback_for_nondefault_target(mock_connectors):
    """Non-default targets keep the legacy balanced round-robin fallback."""
    from church_teams_export import ChurchTeamsExporter
    from collections import Counter

    pairs = ChurchTeamsExporter._make_pool_game_pairs("T", 9, 3)
    games_per_team: Counter = Counter()
    for a, b, _ in pairs:
        games_per_team[a] += 1
        games_per_team[b] += 1
    assert min(games_per_team.values()) >= 3


def test_compute_court_slots_consistent_with_make_pool_game_pairs(mock_connectors):
    """pool_slots from _compute_court_slots equals len(_make_pool_game_pairs) for
    several (n_teams, gpg) combinations."""
    from church_teams_export import ChurchTeamsExporter

    exporter = ChurchTeamsExporter()
    cases = [(2, 2), (5, 2), (8, 2), (12, 2), (9, 3)]
    for n_teams, gpg in cases:
        actual = len(ChurchTeamsExporter._make_pool_game_pairs("_", n_teams, gpg))
        s = exporter._compute_court_slots(n_teams, pool_games_per_team=gpg,
                                          actual_pool_games=actual)
        assert s["pool_slots"] == actual, (
            f"n_teams={n_teams}, gpg={gpg}: pool_slots={s['pool_slots']} != actual={actual}"
        )


# ---------------------------------------------------------------------------
# B5 — _warn_if_schedules_mismatched
# ---------------------------------------------------------------------------

def test_warn_if_schedules_mismatched_clean(mock_connectors):
    """Returns True and no warning when all assignment IDs are in schedule_input."""
    from church_teams_export import ChurchTeamsExporter

    so = {"assignments": [{"game_id": "G1", "resource_id": "R1", "slot": "Sat-1-08:00"}]}
    si = {"games": [{"game_id": "G1"}], "playoff_slots": []}
    assert ChurchTeamsExporter._warn_if_schedules_mismatched(so, si) is True


def test_warn_if_schedules_mismatched_playoff_ok(mock_connectors):
    """Returns True when assignment ID matches a playoff_slot, not a game."""
    from church_teams_export import ChurchTeamsExporter

    so = {"assignments": [{"game_id": "BBM-Final", "resource_id": "R1", "slot": "Sat-2-14:00"}]}
    si = {"games": [], "playoff_slots": [{"game_id": "BBM-Final"}]}
    assert ChurchTeamsExporter._warn_if_schedules_mismatched(so, si) is True


def test_warn_if_schedules_mismatched_detects_orphan(mock_connectors):
    """Returns False and logs a warning when an assignment game_id is unknown."""
    from loguru import logger
    from church_teams_export import ChurchTeamsExporter

    messages = []
    sink_id = logger.add(lambda msg: messages.append(msg), level="WARNING")
    try:
        so = {"assignments": [{"game_id": "STALE-99", "resource_id": "R1", "slot": "Sat-1-08:00"}]}
        si = {"games": [{"game_id": "G1"}], "playoff_slots": []}
        result = ChurchTeamsExporter._warn_if_schedules_mismatched(so, si)
    finally:
        logger.remove(sink_id)
    assert result is False
    assert any("STALE-99" in m for m in messages)
    assert any("different runs" in m for m in messages)


# ---------------------------------------------------------------------------
# Issue #94 — _write_schedule_output_report
# ---------------------------------------------------------------------------

def test_write_schedule_output_report_creates_file(tmp_path):
    """_write_schedule_output_report writes an xlsx with both expected tabs."""
    from church_teams_export import ChurchTeamsExporter
    import openpyxl
    so, si = _make_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ChurchTeamsExporter._write_schedule_output_report(out, so, si)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "Schedule-by-Time" in wb.sheetnames
    assert "Schedule-by-Sport" in wb.sheetnames


def test_write_schedule_output_report_tab1_has_data(tmp_path):
    """Schedule-by-Time tab has a title in row 1 and game text in the grid."""
    from church_teams_export import ChurchTeamsExporter
    import openpyxl
    so, si = _make_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ChurchTeamsExporter._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]
    title = ws.cell(row=1, column=1).value
    assert title and "Sports Fest" in title
    # At least one game should appear somewhere in the grid
    all_values = [ws.cell(row=r, column=c).value
                  for r in range(1, ws.max_row + 1)
                  for c in range(1, ws.max_column + 1)]
    assert any("BBM" in str(v) for v in all_values if v)


def test_write_schedule_output_report_tab2_flat_list(tmp_path):
    """Schedule-by-Sport tab has a header row and one data row per assignment."""
    from church_teams_export import ChurchTeamsExporter
    import openpyxl
    so, si = _make_schedule_pair()
    out = tmp_path / "sched.xlsx"
    ChurchTeamsExporter._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Sport"]
    assert ws.cell(row=1, column=1).value == "game_id"
    assert ws.cell(row=2, column=1).value == "BBM-01"    # Pool comes first
    assert ws.cell(row=3, column=1).value == "BBM-Final"


def test_write_schedule_output_report_unscheduled_section(tmp_path):
    """Unscheduled section appears in Schedule-by-Sport when games are unscheduled."""
    from church_teams_export import ChurchTeamsExporter
    import openpyxl
    so, si = _make_schedule_pair()
    so["unscheduled"] = ["BBM-QF-1"]
    out = tmp_path / "sched.xlsx"
    ChurchTeamsExporter._write_schedule_output_report(out, so, si)
    ws = openpyxl.load_workbook(out)["Schedule-by-Sport"]
    all_values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any("Unscheduled" in str(v) for v in all_values if v)
    assert any("BBM-QF-1" in str(v) for v in all_values if v)


def test_write_schedule_output_report_groups_mixed_pod_windows(tmp_path):
    """Pod resources with different slot windows render as separate time-grid sections."""
    from church_teams_export import ChurchTeamsExporter
    import openpyxl

    schedule_input = {
        "games": [
            {
                "game_id": "PCK-01", "event": "Pickleball",
                "stage": "R1", "pool_id": "", "round": 1,
                "team_a_id": None, "team_b_id": None,
                "duration_minutes": 20, "resource_type": "Pickleball Court",
                "earliest_slot": None, "latest_slot": None,
            },
            {
                "game_id": "TT-01", "event": "Table Tennis",
                "stage": "R1", "pool_id": "", "round": 1,
                "team_a_id": None, "team_b_id": None,
                "duration_minutes": 30, "resource_type": "Table Tennis Table",
                "earliest_slot": None, "latest_slot": None,
            },
        ],
        "resources": [
            {
                "resource_id": "PCK-1", "resource_type": "Pickleball Court",
                "label": "Court-1", "day": "Day-1",
                "open_time": "13:00", "close_time": "13:40", "slot_minutes": 20,
            },
            {
                "resource_id": "TT-1", "resource_type": "Table Tennis Table",
                "label": "Table-1", "day": "Day-1",
                "open_time": "18:00", "close_time": "19:00", "slot_minutes": 30,
            },
        ],
        "precedence": [],
    }
    schedule_output = {
        "solved_at": "2026-05-15T07:07:48",
        "status": "PARTIAL",
        "solver_wall_seconds": 0.2,
        "assignments": [
            {"game_id": "PCK-01", "resource_id": "PCK-1", "slot": "Day-1-13:00"},
            {"game_id": "TT-01", "resource_id": "TT-1", "slot": "Day-1-18:00"},
        ],
        "unscheduled": [],
        "pool_results": [],
    }
    out = tmp_path / "sched.xlsx"
    ChurchTeamsExporter._write_schedule_output_report(out, schedule_output, schedule_input)
    ws = openpyxl.load_workbook(out)["Schedule-by-Time"]

    all_values = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    string_values = [str(v) for v in all_values if v is not None]

    assert any("Pickleball Court" in v for v in string_values)
    assert any("Table Tennis Table" in v for v in string_values)
    assert "13:00" in string_values
    assert "18:00" in string_values
    assert any("PCK-01" in v for v in string_values)
    assert any("TT-01" in v for v in string_values)
