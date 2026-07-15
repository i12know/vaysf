"""Generate print-ready paper score sheets for event-day games.

Issue #211 starts with Basketball because those sheets need player rosters for
foul tracking. Issue #250 extends the same print pipeline to Volleyball. Issue
#254 adds Soccer, and #255 adds Bible Challenge. The renderer intentionally
produces an image-backed PDF using Pillow so it can run with the existing
middleware dependencies.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import math
import re
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urlencode

import qrcode
from loguru import logger
from PIL import Image, ImageDraw, ImageFont

from config import Config
from schedule_publisher import merge_schedule

BASKETBALL_EVENT = "Basketball - Men Team"
BIBLE_CHALLENGE_EVENT = "Bible Challenge - Mixed Team"
SOCCER_EVENT = "Soccer - Coed Exhibition"
VOLLEYBALL_MEN_EVENT = "Volleyball - Men Team"
VOLLEYBALL_WOMEN_EVENT = "Volleyball - Women Team"
VOLLEYBALL_EVENTS = {VOLLEYBALL_MEN_EVENT, VOLLEYBALL_WOMEN_EVENT}
PAGE_W, PAGE_H = 1275, 1650  # Letter page at 150 DPI.
MARGIN = 70
LOGO_BOX = (72, 58, 182, 168)
QR_SIZE = 150
QR_BOX = (PAGE_W - MARGIN - QR_SIZE, 52)
MAX_ROSTER_ROWS = 15
MAX_SOCCER_ROSTER_ROWS_PER_TEAM = 12
MAX_VOLLEYBALL_ROSTER_ROWS_PER_COLUMN = 9
MAX_VOLLEYBALL_ROSTER_ROWS = MAX_VOLLEYBALL_ROSTER_ROWS_PER_COLUMN * 2

COL_BLACK = (32, 38, 46)
COL_BLUE = (20, 72, 118)
COL_BORDER = (88, 99, 115)
COL_LIGHT = (242, 246, 250)
COL_HEADER = (232, 240, 248)
COL_MUTED = (94, 105, 116)
COL_STRIKE = (170, 31, 45)

DAY_LABELS = {
    "Fri-1": "Fri 7/24",
    "Sat-1": "Sat 7/18",
    "Sun-1": "Sun 7/19",
    "Sat-2": "Sat 7/25",
    "Sun-2": "Sun 7/26",
}

_FONT_CANDIDATES = {
    "regular": [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ],
    "bold": [
        Path("C:/Windows/Fonts/arialbd.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ],
    "italic": [
        Path("C:/Windows/Fonts/ariali.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Italic.ttf"),
    ],
}


class ScoreSheetError(RuntimeError):
    """Raised for controlled score-sheet generation failures."""


def default_logo_path() -> Path:
    """Return the committed plugin logo used on printed score sheets."""

    return Path(__file__).resolve().parents[1] / "plugins" / "vaysf" / "assets" / "logo.png"


def default_score_entry_base_url() -> str:
    """Return the coordinator score-entry page URL from the configured WP_URL."""

    if not Config.WP_URL:
        raise ScoreSheetError("WP_URL is required to generate score-entry QR codes.")
    return f"{Config.WP_URL.rstrip('/')}/coordinator-score-entry/"


def score_entry_url_for_game(game_key: str, score_entry_base_url: Optional[str] = None) -> str:
    """Build a stable score-entry URL for the published schedule game_key."""

    base_url = (score_entry_base_url or default_score_entry_base_url()).strip()
    if not base_url:
        raise ScoreSheetError("Score-entry base URL is required to generate score-entry QR codes.")
    separator = "&" if "?" in base_url else "?"
    return f"{base_url}{separator}{urlencode({'action': 'score', 'game_key': game_key})}"


def _font(kind: str, size: int) -> ImageFont.ImageFont:
    for path in _FONT_CANDIDATES.get(kind, []):
        if path.exists():
            return ImageFont.truetype(str(path), size)
    return ImageFont.load_default()


def _text_size(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> tuple[int, int]:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def _center_text(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: str,
    font: ImageFont.ImageFont,
    fill: tuple[int, int, int] = COL_BLACK,
) -> None:
    w, h = _text_size(draw, text, font)
    x = box[0] + (box[2] - box[0] - w) // 2
    y = box[1] + (box[3] - box[1] - h) // 2
    draw.text((x, y), text, font=font, fill=fill)


def _draw_wrapped(
    draw: ImageDraw.ImageDraw,
    text: str,
    xy: tuple[int, int],
    font: ImageFont.ImageFont,
    max_width: int,
    fill: tuple[int, int, int] = COL_BLACK,
    line_gap: int = 6,
) -> int:
    words = str(text or "").split()
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if not current or _text_size(draw, candidate, font)[0] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    if current:
        lines.append(current)
    if not lines:
        lines = [""]

    x, y = xy
    line_h = _text_size(draw, "Ag", font)[1] + line_gap
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_h
    return y


def _load_logo(logo_path: Path) -> Optional[Image.Image]:
    if not logo_path.exists():
        logger.warning(f"Score-sheet logo not found at {logo_path}; rendering without logo")
        return None
    try:
        return Image.open(logo_path).convert("RGBA")
    except Exception as exc:  # pragma: no cover - defensive around corrupt local assets
        logger.warning(f"Could not load score-sheet logo at {logo_path}: {exc}")
        return None


def _paste_contained(canvas: Image.Image, source: Image.Image, box: tuple[int, int, int, int]) -> None:
    max_w = box[2] - box[0]
    max_h = box[3] - box[1]
    image = source.copy()
    image.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
    x = box[0] + (max_w - image.width) // 2
    y = box[1] + (max_h - image.height) // 2
    canvas.paste(image, (x, y), image if image.mode == "RGBA" else None)


class PhotoCache:
    """Small per-run cache for roster photo thumbnails."""

    def __init__(self) -> None:
        self._cache: dict[str, Optional[Image.Image]] = {}

    def load(self, photo_ref: Any) -> Optional[Image.Image]:
        url_or_path = _extract_photo_ref(photo_ref)
        if not url_or_path:
            return None
        if url_or_path in self._cache:
            cached = self._cache[url_or_path]
            return cached.copy() if cached is not None else None

        image = _load_photo_image(url_or_path)
        self._cache[url_or_path] = image
        return image.copy() if image is not None else None


def _extract_photo_ref(photo_ref: Any) -> str:
    if _is_blank_photo_ref(photo_ref):
        return ""
    text = str(photo_ref).strip()
    formula_match = re.search(r'IMAGE\("([^"]+)"', text, flags=re.IGNORECASE)
    if formula_match:
        return formula_match.group(1).strip()
    return text


def _is_blank_photo_ref(photo_ref: Any) -> bool:
    if photo_ref is None:
        return True
    if isinstance(photo_ref, float) and math.isnan(photo_ref):
        return True
    text = str(photo_ref).strip()
    return not text or text.upper() in {"N/A", "NAN", "NONE", "NULL"}


def _load_photo_image(url_or_path: str) -> Optional[Image.Image]:
    try:
        if url_or_path.startswith(("http://", "https://")):
            with urllib.request.urlopen(url_or_path, timeout=8) as response:
                data = response.read(2_000_000)
            return Image.open(io.BytesIO(data)).convert("RGB")

        path = Path(url_or_path)
        if path.exists():
            return Image.open(path).convert("RGB")
    except (OSError, urllib.error.URLError, ValueError) as exc:
        logger.debug(f"Could not load roster photo {url_or_path!r}: {exc}")
    return None


def enrich_roster_photos_from_workbook(
    roster_rows: Iterable[dict[str, Any]],
    workbook_path: Path,
) -> list[dict[str, Any]]:
    """Fill blank Photo values with formula text from the source roster workbook.

    pandas/openpyxl reads Excel IMAGE() cells as their cached display value, which
    is often blank. The score sheets need the formula URL, so this helper reopens
    only the Roster tab with formulas preserved and patches copied row dicts.
    """

    rows = [dict(row) for row in roster_rows]
    if not rows:
        return rows

    try:
        from openpyxl import load_workbook

        wb = load_workbook(workbook_path, data_only=False, read_only=True)
        ws = wb["Roster"]
    except Exception as exc:
        logger.debug(f"Could not read roster photo formulas from {workbook_path}: {exc}")
        return rows

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        photo_idx = headers.index("Photo")
    except ValueError:
        wb.close()
        return rows

    for row_dict, cells in zip(rows, ws.iter_rows(min_row=2), strict=False):
        if not _is_blank_photo_ref(row_dict.get("Photo")):
            continue
        if photo_idx < len(cells):
            formula_or_value = cells[photo_idx].value
            if not _is_blank_photo_ref(formula_or_value):
                row_dict["Photo"] = formula_or_value
    wb.close()
    return rows


def _friendly_slot(slot: Any) -> str:
    raw = str(slot or "").strip()
    if not raw:
        return "Time TBD"
    match = re.match(r"^([A-Za-z]+-\d+)-(\d{1,2}):(\d{2})$", raw)
    if not match:
        return raw
    day, hour, minute = match.groups()
    clock = dt.time(int(hour), int(minute)).strftime("%I:%M %p").lstrip("0")
    return f"{DAY_LABELS.get(day, day)}, {clock}"


def _friendly_location(game: dict[str, Any]) -> str:
    location = str(game.get("scheduled_location") or "").strip()
    if location:
        return location

    resource_id = str(game.get("resource_id") or "").strip()
    gym_match = re.match(r"^(?:GYM|BB)-[A-Za-z]+-\d+-(\d+)$", resource_id)
    if gym_match:
        return f"EHS Main Gym - Court {gym_match.group(1)}"
    return resource_id or "TBD"


def _team_code(value: Any) -> str:
    text = str(value or "").strip()
    if "::" in text:
        text = text.rsplit("::", 1)[-1]
    tokens = re.findall(r"[A-Z0-9]{2,5}", text.upper())
    return tokens[0] if tokens else text.upper()


def _team_label(game: dict[str, Any], side: str) -> str:
    return str(game.get(f"team_{side}_label") or game.get(f"team_{side}_key") or "TBD").strip()


def _team_labels(game: dict[str, Any], sides: tuple[str, ...]) -> list[str]:
    return [_team_label(game, side) for side in sides if _team_label(game, side) != "TBD"]


def _roster_name(row: dict[str, Any]) -> str:
    first = str(row.get("First Name") or "").strip()
    last = str(row.get("Last Name") or "").strip()
    full = f"{last} {first}".strip()
    return full or str(row.get("Full Name") or "").strip() or "Player"


def _event_matches_roster_row(row: dict[str, Any], event: str) -> bool:
    sport_type = str(row.get("sport_type") or "").strip()
    sport_gender = str(row.get("sport_gender") or "").strip()
    sport_format = str(row.get("sport_format") or "").strip()
    if sport_type == event:
        return True
    if event == BASKETBALL_EVENT:
        return sport_type == "Basketball" and sport_gender == "Men" and sport_format == "Team"
    if event == VOLLEYBALL_MEN_EVENT:
        return sport_type == "Volleyball" and sport_gender == "Men" and sport_format == "Team"
    if event == VOLLEYBALL_WOMEN_EVENT:
        return sport_type == "Volleyball" and sport_gender == "Women" and sport_format == "Team"
    if event == SOCCER_EVENT:
        return sport_type == "Soccer" and sport_gender == "Coed" and sport_format == "Exhibition"
    return False


def _approval_status(row: dict[str, Any]) -> str:
    for key in ("approval_status", "Approval Status", "Approval_Status", "approval status"):
        status = str(row.get(key) or "").strip()
        if status:
            return status
    return ""


def _is_approved_roster_row(row: dict[str, Any]) -> bool:
    return _approval_status(row).casefold() == "approved"


def _draw_unapproved_roster_mark(
    draw: ImageDraw.ImageDraw,
    row_box: tuple[int, int, int, int],
    name_box: tuple[int, int, int, int],
    status: str,
) -> None:
    x0, y0, x1, y1 = row_box
    draw.line((x0 + 4, (y0 + y1) // 2, x1 - 4, (y0 + y1) // 2), fill=COL_STRIKE, width=3)
    status_text = status or "not approved"
    draw.text(
        (name_box[0], max(y0 + 2, name_box[3] - 13)),
        status_text[:24],
        font=_font("regular", 10),
        fill=COL_STRIKE,
    )


def _sort_roster_index(indexed: dict[str, list[dict[str, Any]]]) -> None:
    for rows in indexed.values():
        rows.sort(
            key=lambda item: (
                str(item.get("Last Name") or "").casefold(),
                str(item.get("First Name") or "").casefold(),
            )
        )


def build_roster_index(roster_rows: Iterable[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Index Basketball roster rows by church/team code."""

    indexed: dict[str, list[dict[str, Any]]] = {}
    for row in roster_rows:
        if not _event_matches_roster_row(row, BASKETBALL_EVENT):
            continue
        code = _team_code(row.get("Church Team"))
        if not code:
            continue
        indexed.setdefault(code, []).append(row)
    _sort_roster_index(indexed)
    return indexed


def build_volleyball_roster_index(roster_rows: Iterable[dict[str, Any]]) -> dict[str, dict[str, list[dict[str, Any]]]]:
    """Index Volleyball roster rows by event and church/team code."""

    indexed: dict[str, dict[str, list[dict[str, Any]]]] = {
        VOLLEYBALL_MEN_EVENT: {},
        VOLLEYBALL_WOMEN_EVENT: {},
    }
    for row in roster_rows:
        for event in VOLLEYBALL_EVENTS:
            if not _event_matches_roster_row(row, event):
                continue
            code = _team_code(row.get("Church Team"))
            if not code:
                continue
            indexed[event].setdefault(code, []).append(row)

    for event_index in indexed.values():
        _sort_roster_index(event_index)
    return indexed


def build_soccer_roster_index(roster_rows: Iterable[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Index Soccer roster rows by church/team code."""

    indexed: dict[str, list[dict[str, Any]]] = {}
    for row in roster_rows:
        if not _event_matches_roster_row(row, SOCCER_EVENT):
            continue
        code = _team_code(row.get("Church Team"))
        if not code:
            continue
        indexed.setdefault(code, []).append(row)
    _sort_roster_index(indexed)
    return indexed


def _warn_if_approval_status_missing(roster_rows: Iterable[dict[str, Any]], sport_label: str) -> None:
    rows = list(roster_rows)
    if not rows:
        return
    if any(_approval_status(row) for row in rows):
        return
    logger.warning(
        f"{sport_label} score sheets: roster rows have no approval_status/Approval Status values; "
        "all printed roster rows will be marked not approved."
    )


def _draw_logo(canvas: Image.Image, logo: Optional[Image.Image]) -> None:
    draw = ImageDraw.Draw(canvas)
    if logo is not None:
        _paste_contained(canvas, logo, LOGO_BOX)
    else:
        draw.rectangle(LOGO_BOX, outline=COL_BORDER, width=2)
        _center_text(draw, LOGO_BOX, "VAY", _font("bold", 28), COL_BLUE)


def _draw_qr(canvas: Image.Image, game_key: str, score_entry_base_url: Optional[str]) -> None:
    qr_payload = score_entry_url_for_game(game_key, score_entry_base_url)
    qr = qrcode.QRCode(border=1, box_size=10)
    qr.add_data(qr_payload)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    qr_image = qr_image.resize((QR_SIZE, QR_SIZE), Image.Resampling.NEAREST)
    canvas.paste(qr_image, QR_BOX)


def _draw_header(draw: ImageDraw.ImageDraw, game: dict[str, Any]) -> None:
    team_a = _team_label(game, "a")
    team_b = _team_label(game, "b")
    game_key = str(game.get("game_key") or "")
    location = _friendly_location(game)
    draw.text((220, 58), "Vietnamese Alliance Youth Sports Festival", font=_font("regular", 25), fill=COL_BLACK)
    draw.text((220, 96), f"Basketball SCORESHEET: {team_a} vs. {team_b}", font=_font("bold", 31), fill=COL_BLACK)
    draw.text((220, 140), f"Game ID: {game_key}", font=_font("bold", 23), fill=COL_BLUE)
    draw.text(
        (220, 174),
        f"{_friendly_slot(game.get('scheduled_slot'))}     Location: {location}",
        font=_font("regular", 22),
        fill=COL_BLACK,
    )
    draw.text((PAGE_W - MARGIN - QR_SIZE, 210), "Scan to enter score", font=_font("regular", 15), fill=COL_MUTED)


def _draw_generic_header(draw: ImageDraw.ImageDraw, game: dict[str, Any], title: str) -> None:
    team_a = _team_label(game, "a")
    team_b = _team_label(game, "b")
    game_key = str(game.get("game_key") or "")
    location = _friendly_location(game)
    draw.text((220, 58), "Vietnamese Alliance Youth Sports Festival", font=_font("regular", 25), fill=COL_BLACK)
    draw.text((220, 96), f"{title}: {team_a} vs. {team_b}", font=_font("bold", 31), fill=COL_BLACK)
    draw.text((220, 140), f"Game ID: {game_key}", font=_font("bold", 23), fill=COL_BLUE)
    draw.text(
        (220, 174),
        f"{_friendly_slot(game.get('scheduled_slot'))}     Location: {location}",
        font=_font("regular", 22),
        fill=COL_BLACK,
    )
    draw.text((PAGE_W - MARGIN - QR_SIZE, 210), "Scan to enter score", font=_font("regular", 15), fill=COL_MUTED)


def _draw_multi_team_header(
    draw: ImageDraw.ImageDraw,
    game: dict[str, Any],
    title: str,
    sides: tuple[str, ...],
    line_break_teams: bool = False,
) -> None:
    teams = _team_labels(game, sides)
    game_key = str(game.get("game_key") or "")
    location = _friendly_location(game)
    if str(game.get("event") or "").strip() == BIBLE_CHALLENGE_EVENT:
        location = re.sub(r"\s+-\s+Station\s+\d+\b", "", location, flags=re.IGNORECASE)
    draw.text((220, 58), "Vietnamese Alliance Youth Sports Festival", font=_font("regular", 25), fill=COL_BLACK)
    if line_break_teams:
        draw.text((220, 96), f"{title}:", font=_font("bold", 30), fill=COL_BLACK)
        draw.text((220, 132), " vs. ".join(teams), font=_font("bold", 26), fill=COL_BLACK)
        game_id_y = 164
        detail_y = 198
    else:
        draw.text((220, 96), f"{title}: {' vs. '.join(teams)}", font=_font("bold", 31), fill=COL_BLACK)
        game_id_y = 140
        detail_y = 174
    draw.text((220, game_id_y), f"Game ID: {game_key}", font=_font("bold", 23), fill=COL_BLUE)
    draw.text(
        (220, detail_y),
        f"{_friendly_slot(game.get('scheduled_slot'))}     Location: {location}",
        font=_font("regular", 22),
        fill=COL_BLACK,
    )
    draw.text((PAGE_W - MARGIN - QR_SIZE, 210), "Scan to enter score", font=_font("regular", 15), fill=COL_MUTED)


def _draw_score_boxes(draw: ImageDraw.ImageDraw, game: dict[str, Any]) -> None:
    top = 250
    left = MARGIN
    right = PAGE_W - MARGIN
    draw.rounded_rectangle((left, top, right, top + 92), radius=12, outline=COL_BORDER, width=2, fill=COL_LIGHT)
    draw.text((left + 24, top + 16), "FINAL SCORE", font=_font("bold", 22), fill=COL_BLUE)
    team_w = (right - left - 260) // 2
    for idx, side in enumerate(("a", "b")):
        x = left + 190 + idx * (team_w + 28)
        draw.text((x, top + 12), _team_label(game, side), font=_font("bold", 21), fill=COL_BLACK)
        draw.rectangle((x, top + 42, x + team_w, top + 78), outline=COL_BORDER, width=2, fill=(255, 255, 255))


def _draw_three_team_score_boxes(draw: ImageDraw.ImageDraw, game: dict[str, Any]) -> None:
    top = 250
    left = MARGIN
    right = PAGE_W - MARGIN
    draw.rounded_rectangle((left, top, right, top + 104), radius=12, outline=COL_BORDER, width=2, fill=COL_LIGHT)
    draw.text((left + 24, top + 18), "FINAL SCORE", font=_font("bold", 22), fill=COL_BLUE)
    team_w = (right - left - 252) // 3
    for idx, side in enumerate(("a", "b", "c")):
        x = left + 190 + idx * (team_w + 22)
        draw.text((x, top + 12), _team_label(game, side), font=_font("bold", 21), fill=COL_BLACK)
        draw.rectangle((x, top + 48, x + team_w, top + 86), outline=COL_BORDER, width=2, fill=(255, 255, 255))


def _draw_referee_section(draw: ImageDraw.ImageDraw, y: int = 370) -> None:
    label_font = _font("bold", 18)
    text_font = _font("regular", 18)
    for label in ("PRIMARY REFEREE:", "SECONDARY REFEREE:"):
        draw.text((MARGIN, y), label, font=label_font, fill=COL_BLACK)
        draw.line((270, y + 22, 560, y + 22), fill=COL_BORDER, width=2)
        draw.text((590, y), "CHURCH:", font=text_font, fill=COL_BLACK)
        draw.line((682, y + 22, 780, y + 22), fill=COL_BORDER, width=2)
        y += 42
    draw.text((MARGIN, y + 12), "Opening Prayer Verse:", font=_font("bold", 16), fill=COL_BLACK)
    draw.text(
        (270, y + 12),
        "You, dear children, are from God and have overcome them, because the one who is in you is greater. (1 John 4:4)",
        font=_font("italic", 15),
        fill=COL_BLACK,
    )


def _draw_bible_challenge_official_section(draw: ImageDraw.ImageDraw, y: int = 384) -> None:
    label_font = _font("bold", 18)
    text_font = _font("regular", 18)
    left_y = y
    for label in ("MODERATOR:", "SCOREKEEPER:"):
        draw.text((MARGIN, left_y), label, font=label_font, fill=COL_BLACK)
        draw.line((230, left_y + 22, 520, left_y + 22), fill=COL_BORDER, width=2)
        left_y += 34

    ref_y = y
    for label in ("REF 1:", "REF 2:", "REF 3:"):
        draw.text((610, ref_y), label, font=label_font, fill=COL_BLACK)
        draw.line((704, ref_y + 22, 870, ref_y + 22), fill=COL_BORDER, width=2)
        draw.text((900, ref_y), "CHURCH:", font=text_font, fill=COL_BLACK)
        draw.line((994, ref_y + 22, PAGE_W - MARGIN, ref_y + 22), fill=COL_BORDER, width=2)
        ref_y += 34

    y = max(left_y, ref_y, 486)
    draw.text((MARGIN, y + 10), "Opening Prayer Verse:", font=_font("bold", 16), fill=COL_BLACK)
    draw.text(
        (270, y + 10),
        "Your word is a lamp for my feet, a light on my path. (Psalm 119:105)",
        font=_font("italic", 15),
        fill=COL_BLACK,
    )


def _draw_bible_challenge_score_grid(draw: ImageDraw.ImageDraw, game: dict[str, Any], y: int = 520) -> int:
    teams = [_team_label(game, side) for side in ("a", "b", "c")]
    x = MARGIN
    width = PAGE_W - MARGIN * 2
    label_w = 180
    team_w = (width - label_w) // 3
    team_header_h = 48
    row_h = 56
    rows = ["Round 1", "Round 2"]
    height = 40 + team_header_h + row_h * len(rows)
    draw.rectangle((x, y, x + width, y + height), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 40), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 40), "BIBLE CHALLENGE SCORE TRACKER", _font("bold", 20), COL_BLACK)

    for idx, team in enumerate(teams):
        left = x + label_w + idx * team_w
        right = left + team_w
        draw.line((left, y + 40, left, y + height), fill=COL_BORDER, width=1)
        _center_text(draw, (left, y + 40, right, y + 40 + team_header_h), team, _font("bold", 19), COL_BLACK)
    draw.line((x + width, y + 40, x + width, y + height), fill=COL_BORDER, width=1)
    draw.line((x, y + 40 + team_header_h, x + width, y + 40 + team_header_h), fill=COL_BORDER, width=1)

    for row_idx, label in enumerate(rows):
        row_y = y + 40 + team_header_h + row_idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        draw.text((x + 16, row_y + 17), label, font=_font("bold", 17), fill=COL_BLACK)
        for idx in range(3):
            left = x + label_w + idx * team_w
            right = left + team_w
            box_w = 150
            box_h = 36
            bx = left + (right - left - box_w) // 2
            by = row_y + (row_h - box_h) // 2
            draw.rectangle((bx, by, bx + box_w, by + box_h), outline=COL_BORDER, width=2, fill=(255, 255, 255))
    return y + height


def _draw_bible_challenge_notes_grid(draw: ImageDraw.ImageDraw, y: int) -> int:
    x = MARGIN
    width = PAGE_W - MARGIN * 2
    header_h = 40
    row_h = 44
    rows = 8
    height = header_h + row_h * rows
    draw.rectangle((x, y, x + width, y + height), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + header_h), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + header_h), "QUESTION / APPEAL NOTES", _font("bold", 18), COL_BLACK)

    col_round = x + 110
    col_team = col_round + 155
    col_ref = col_team + 250
    for line_x in (col_round, col_team, col_ref):
        draw.line((line_x, y + header_h, line_x, y + height), fill=COL_BORDER, width=1)
    draw.text((x + 18, y + header_h + 11), "ROUND", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_round + 18, y + header_h + 11), "TEAM", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_team + 18, y + header_h + 11), "QUESTION / REF", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_ref + 18, y + header_h + 11), "NOTES", font=_font("bold", 15), fill=COL_BLACK)

    for idx in range(rows):
        row_y = y + header_h + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        if idx == 0:
            continue
        draw.line((x + 18, row_y + 24, col_round - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_round + 18, row_y + 24, col_team - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_team + 18, row_y + 24, col_ref - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_ref + 18, row_y + 24, x + width - 18, row_y + 24), fill=(190, 196, 204), width=1)
    return y + height


def _draw_bible_challenge_footer(draw: ImageDraw.ImageDraw) -> None:
    y = 1392
    draw.text((MARGIN, y), "CERTIFICATION / COMMENTS", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((MARGIN, y + 50, PAGE_W - MARGIN, y + 50), fill=COL_BORDER, width=1)
    draw.line((MARGIN, y + 96, PAGE_W - MARGIN, y + 96), fill=COL_BORDER, width=1)
    draw.text((MARGIN, y + 148), "MODERATOR SIGNATURE:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((340, y + 170, 650, y + 170), fill=COL_BORDER, width=2)
    draw.text((690, y + 148), "SCOREKEEPER:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((858, y + 170, PAGE_W - MARGIN, y + 170), fill=COL_BORDER, width=2)


def _draw_volleyball_score_grid(draw: ImageDraw.ImageDraw, game: dict[str, Any], y: int = 500) -> None:
    team_a = _team_label(game, "a")
    team_b = _team_label(game, "b")
    x = MARGIN
    width = PAGE_W - MARGIN * 2
    label_w = 140
    team_w = 112
    row_h = 44
    rows = [
        ("Set 1", team_a, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 CAP"),
        ("", team_b, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 CAP"),
        ("Set 2", team_a, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 CAP"),
        ("", team_b, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 CAP"),
        ("Tiebreaker", team_a, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 CAP"),
        ("", team_b, "1 2 3 4 5 6 7 8 9 10 11 12 13 14 15 CAP"),
    ]
    height = row_h * len(rows) + 34
    draw.rectangle((x, y, x + width, y + height), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 34), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 34), "VOLLEYBALL SET SCORE TRACKER", _font("bold", 18), COL_BLACK)
    draw.line((x + label_w, y + 34, x + label_w, y + height), fill=COL_BORDER, width=1)
    draw.line((x + label_w + team_w, y + 34, x + label_w + team_w, y + height), fill=COL_BORDER, width=1)

    for idx, (label, team, score_line) in enumerate(rows):
        row_y = y + 34 + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        if label:
            draw.text((x + 12, row_y + 10), label, font=_font("bold", 18), fill=COL_BLACK)
        draw.text((x + label_w + 16, row_y + 10), team, font=_font("bold", 18), fill=COL_BLACK)
        draw.text((x + label_w + team_w + 16, row_y + 11), score_line, font=_font("regular", 20), fill=COL_BLACK)


def _draw_volleyball_roster_table(
    canvas: Image.Image,
    draw: ImageDraw.ImageDraw,
    origin: tuple[int, int],
    width: int,
    title: str,
    roster_rows: list[dict[str, Any]],
    photo_cache: PhotoCache,
) -> None:
    x, y = origin
    header_h = 48
    row_h = 57
    table_h = header_h + row_h * MAX_VOLLEYBALL_ROSTER_ROWS_PER_COLUMN
    draw.rectangle((x, y, x + width, y + table_h), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 29), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 29), title, _font("bold", 16), COL_BLACK)

    gap = 10
    column_w = (width - gap) // 2
    columns = [x, x + column_w + gap]
    draw.line((columns[1] - gap // 2, y + 29, columns[1] - gap // 2, y + table_h), fill=COL_BORDER, width=2)

    col_no = 28
    col_photo = 54
    col_age = 34
    for col_x in columns:
        photo_x = col_x + col_no
        name_x = photo_x + col_photo + 7
        age_x = col_x + column_w - col_age
        for line_x in (photo_x, name_x - 7, age_x):
            draw.line((line_x, y + 29, line_x, y + table_h), fill=COL_BORDER, width=1)
        draw.text((col_x + 9, y + 33), "#", font=_font("bold", 10), fill=COL_BLACK)
        draw.text((name_x, y + 33), "ATHLETE", font=_font("bold", 10), fill=COL_BLACK)
        draw.text((age_x + 8, y + 33), "AGE", font=_font("bold", 10), fill=COL_BLACK)

    rows = roster_rows[:MAX_VOLLEYBALL_ROSTER_ROWS]
    for col_idx, col_x in enumerate(columns):
        start_idx = col_idx * MAX_VOLLEYBALL_ROSTER_ROWS_PER_COLUMN
        photo_x = col_x + col_no
        name_x = photo_x + col_photo + 7
        age_x = col_x + column_w - col_age
        for row_idx in range(MAX_VOLLEYBALL_ROSTER_ROWS_PER_COLUMN):
            idx = start_idx + row_idx
            row_y = y + header_h + row_idx * row_h
            draw.line((col_x, row_y, col_x + column_w, row_y), fill=COL_BORDER, width=1)
            draw.text((col_x + 8, row_y + 21), str(idx + 1), font=_font("regular", 10), fill=COL_MUTED)
            photo_box = (photo_x + 4, row_y + 5, photo_x + 48, row_y + 51)
            draw.rectangle(photo_box, outline=(185, 193, 203), width=1)
            if idx >= len(rows):
                draw.line((name_x, row_y + 30, age_x - 8, row_y + 30), fill=(190, 196, 204), width=1)
                continue
            row = rows[idx]
            name = _roster_name(row)
            age = row.get("Age (at Event)")
            age_text = f"{int(age)}" if isinstance(age, (int, float)) else str(age or "").strip()
            photo = photo_cache.load(row.get("Photo"))
            if photo is not None:
                _paste_contained(canvas, photo, photo_box)
            _draw_wrapped(draw, name, (name_x, row_y + 10), _font("regular", 14), age_x - name_x - 6, COL_BLACK, line_gap=1)
            if age_text:
                _center_text(draw, (age_x, row_y, col_x + column_w, row_y + row_h), age_text, _font("regular", 12), COL_BLACK)
            if not _is_approved_roster_row(row):
                _draw_unapproved_roster_mark(
                    draw,
                    (col_x, row_y, col_x + column_w, row_y + row_h),
                    (name_x, row_y, age_x - 4, row_y + row_h),
                    _approval_status(row),
                )

    if len(roster_rows) > MAX_VOLLEYBALL_ROSTER_ROWS:
        draw.text(
            (x + 12, y + table_h + 5),
            f"+{len(roster_rows) - MAX_VOLLEYBALL_ROSTER_ROWS} more",
            font=_font("regular", 12),
            fill=COL_MUTED,
        )


def _draw_roster_table(
    canvas: Image.Image,
    draw: ImageDraw.ImageDraw,
    origin: tuple[int, int],
    width: int,
    title: str,
    roster_rows: list[dict[str, Any]],
    photo_cache: PhotoCache,
) -> None:
    x, y = origin
    header_h = 58
    row_h = 56
    table_h = header_h + row_h * MAX_ROSTER_ROWS
    draw.rectangle((x, y, x + width, y + table_h), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 30), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 30), title, _font("bold", 17), COL_BLACK)

    col_no = 44
    col_photo = 62
    col_age = 34
    col_foul = 82
    no_x = x + col_no
    photo_x = no_x + col_photo
    foul_x = x + width - col_foul
    age_x = foul_x - col_age
    name_x = photo_x + 7
    label_y = y + 34

    for line_x in (no_x, photo_x, age_x, foul_x):
        draw.line((line_x, y + 30, line_x, y + table_h), fill=COL_BORDER, width=1)
    draw.text((x + 11, label_y), "NO", font=_font("bold", 10), fill=COL_BLACK)
    draw.text((photo_x + 8, label_y), "ATHLETE", font=_font("bold", 10), fill=COL_BLACK)
    draw.text((age_x + 8, label_y), "AGE", font=_font("bold", 10), fill=COL_BLACK)
    draw.text((foul_x + 22, label_y), "FOUL", font=_font("bold", 10), fill=COL_BLACK)
    draw.line((foul_x, y + header_h, foul_x, y + table_h), fill=COL_BORDER, width=2)

    rows = roster_rows[:MAX_ROSTER_ROWS]
    for idx in range(MAX_ROSTER_ROWS):
        row_y = y + header_h + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        photo_box = (no_x + 3, row_y + 4, no_x + 53, row_y + 54)
        draw.rectangle(photo_box, outline=(185, 193, 203), width=1)
        if idx < len(rows):
            row = rows[idx]
            name = _roster_name(row)
            age = row.get("Age (at Event)")
            age_text = f"{int(age)}" if isinstance(age, (int, float)) else str(age or "").strip()
            photo = photo_cache.load(row.get("Photo"))
            if photo is not None:
                _paste_contained(canvas, photo, photo_box)
            draw.line((x + 8, row_y + 30, no_x - 8, row_y + 30), fill=(170, 178, 188), width=1)
            _draw_wrapped(draw, name, (name_x, row_y + 9), _font("regular", 16), age_x - name_x - 5, COL_BLACK, line_gap=2)
            if age_text:
                _center_text(draw, (age_x, row_y, foul_x, row_y + row_h), age_text, _font("regular", 13), COL_BLACK)
            if not _is_approved_roster_row(row):
                _draw_unapproved_roster_mark(
                    draw,
                    (x, row_y, foul_x, row_y + row_h),
                    (name_x, row_y, age_x - 4, row_y + row_h),
                    _approval_status(row),
                )
        else:
            draw.line((x + 8, row_y + 30, no_x - 8, row_y + 30), fill=(190, 196, 204), width=1)
            draw.line((name_x, row_y + 30, age_x - 8, row_y + 30), fill=(190, 196, 204), width=1)

        foul_start = foul_x + 8
        for foul_idx in range(5):
            cx = foul_start + foul_idx * 13
            cy = row_y + 28
            draw.ellipse((cx - 5, cy - 5, cx + 5, cy + 5), outline=COL_BORDER, width=1)

    if len(roster_rows) > MAX_ROSTER_ROWS:
        draw.text(
            (x + 12, y + table_h + 8),
            f"+{len(roster_rows) - MAX_ROSTER_ROWS} additional roster row(s) not printed",
            font=_font("regular", 13),
            fill=COL_MUTED,
        )


def _draw_footer(draw: ImageDraw.ImageDraw) -> None:
    y = 1430
    draw.text((MARGIN, y + 8), "REFEREE COMMENTS", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((MARGIN, y + 58, PAGE_W - MARGIN, y + 58), fill=COL_BORDER, width=1)
    draw.line((MARGIN, y + 104, PAGE_W - MARGIN, y + 104), fill=COL_BORDER, width=1)
    draw.text((MARGIN, y + 150), "REFEREE SIGNATURES:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((300, y + 172, 570, y + 172), fill=COL_BORDER, width=2)
    draw.text((596, y + 150), "AND", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((650, y + 172, 920, y + 172), fill=COL_BORDER, width=2)


def _draw_volleyball_footer(draw: ImageDraw.ImageDraw) -> None:
    y = 1390
    draw.text((MARGIN, y + 8), "REFEREE COMMENTS", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((MARGIN, y + 58, PAGE_W - MARGIN, y + 58), fill=COL_BORDER, width=1)
    draw.line((MARGIN, y + 104, PAGE_W - MARGIN, y + 104), fill=COL_BORDER, width=1)
    draw.text((MARGIN, y + 150), "REFEREE SIGNATURES:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((300, y + 172, 570, y + 172), fill=COL_BORDER, width=2)
    draw.text((596, y + 150), "AND", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((650, y + 172, 920, y + 172), fill=COL_BORDER, width=2)


def _draw_soccer_score_grid(draw: ImageDraw.ImageDraw, game: dict[str, Any], y: int = 360) -> int:
    team_a = _team_label(game, "a")
    team_b = _team_label(game, "b")
    x = MARGIN
    width = PAGE_W - MARGIN * 2
    label_w = 190
    team_w = (width - label_w) // 2
    team_header_h = 46
    row_h = 54
    rows = [
        ("1st Half", team_a, team_b),
        ("2nd Half", team_a, team_b),
        ("Final", team_a, team_b),
        ("Shootout / OT", team_a, team_b),
    ]
    height = 40 + team_header_h + row_h * len(rows)
    draw.rectangle((x, y, x + width, y + height), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 40), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 40), "SOCCER SCORE TRACKER", _font("bold", 20), COL_BLACK)

    col_a_x = x + label_w
    col_b_x = col_a_x + team_w
    draw.line((col_a_x, y + 40, col_a_x, y + height), fill=COL_BORDER, width=1)
    draw.line((col_b_x, y + 40, col_b_x, y + height), fill=COL_BORDER, width=1)
    _center_text(draw, (col_a_x, y + 40, col_b_x, y + 40 + team_header_h), team_a, _font("bold", 19), COL_BLACK)
    _center_text(draw, (col_b_x, y + 40, x + width, y + 40 + team_header_h), team_b, _font("bold", 19), COL_BLACK)
    draw.line((x, y + 40 + team_header_h, x + width, y + 40 + team_header_h), fill=COL_BORDER, width=1)

    for idx, (label, _team_a, _team_b) in enumerate(rows):
        row_y = y + 40 + team_header_h + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        draw.text((x + 16, row_y + 16), label, font=_font("bold", 18), fill=COL_BLACK)
        for left, right in ((col_a_x, col_b_x), (col_b_x, x + width)):
            box_w = 118
            box_h = 34
            bx = left + (right - left - box_w) // 2
            by = row_y + (row_h - box_h) // 2
            draw.rectangle((bx, by, bx + box_w, by + box_h), outline=COL_BORDER, width=2, fill=(255, 255, 255))
    return y + height


def _draw_soccer_roster_table(
    canvas: Image.Image,
    draw: ImageDraw.ImageDraw,
    origin: tuple[int, int],
    width: int,
    title: str,
    roster_rows: list[dict[str, Any]],
    photo_cache: PhotoCache,
) -> None:
    x, y = origin
    header_h = 48
    row_h = 46
    table_h = header_h + row_h * MAX_SOCCER_ROSTER_ROWS_PER_TEAM
    draw.rectangle((x, y, x + width, y + table_h), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + 29), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + 29), title, _font("bold", 16), COL_BLACK)

    col_no = 32
    col_photo = 48
    col_age = 36
    no_x = x + col_no
    photo_x = no_x + col_photo
    age_x = x + width - col_age
    name_x = photo_x + 7
    for line_x in (no_x, photo_x, age_x):
        draw.line((line_x, y + 29, line_x, y + table_h), fill=COL_BORDER, width=1)
    draw.text((x + 9, y + 33), "#", font=_font("bold", 10), fill=COL_BLACK)
    draw.text((name_x, y + 33), "ATHLETE", font=_font("bold", 10), fill=COL_BLACK)
    draw.text((age_x + 8, y + 33), "AGE", font=_font("bold", 10), fill=COL_BLACK)

    rows = roster_rows[:MAX_SOCCER_ROSTER_ROWS_PER_TEAM]
    for idx in range(MAX_SOCCER_ROSTER_ROWS_PER_TEAM):
        row_y = y + header_h + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        draw.text((x + 9, row_y + 17), str(idx + 1), font=_font("regular", 10), fill=COL_MUTED)
        photo_box = (no_x + 4, row_y + 4, no_x + 40, row_y + 40)
        draw.rectangle(photo_box, outline=(185, 193, 203), width=1)
        if idx >= len(rows):
            draw.line((name_x, row_y + 24, age_x - 8, row_y + 24), fill=(190, 196, 204), width=1)
            continue

        row = rows[idx]
        name = _roster_name(row)
        age = row.get("Age (at Event)")
        age_text = f"{int(age)}" if isinstance(age, (int, float)) else str(age or "").strip()
        photo = photo_cache.load(row.get("Photo"))
        if photo is not None:
            _paste_contained(canvas, photo, photo_box)
        _draw_wrapped(draw, name, (name_x, row_y + 6), _font("regular", 13), age_x - name_x - 6, COL_BLACK, line_gap=1)
        if age_text:
            _center_text(draw, (age_x, row_y, x + width, row_y + row_h), age_text, _font("regular", 11), COL_BLACK)
        if not _is_approved_roster_row(row):
            _draw_unapproved_roster_mark(
                draw,
                (x, row_y, x + width, row_y + row_h),
                (name_x, row_y, age_x - 4, row_y + row_h),
                _approval_status(row),
            )

    if len(roster_rows) > MAX_SOCCER_ROSTER_ROWS_PER_TEAM:
        draw.text(
            (x + 12, y + table_h + 4),
            f"+{len(roster_rows) - MAX_SOCCER_ROSTER_ROWS_PER_TEAM} more",
            font=_font("regular", 11),
            fill=COL_MUTED,
        )


def _draw_soccer_event_log(draw: ImageDraw.ImageDraw, y: int) -> int:
    x = MARGIN
    width = PAGE_W - MARGIN * 2
    header_h = 40
    row_h = 44
    rows = 8
    height = header_h + row_h * rows
    draw.rectangle((x, y, x + width, y + height), outline=COL_BORDER, width=2)
    draw.rectangle((x, y, x + width, y + header_h), fill=COL_HEADER, outline=COL_BORDER, width=2)
    _center_text(draw, (x, y, x + width, y + header_h), "GOALS / CARDS / MATCH NOTES", _font("bold", 18), COL_BLACK)

    col_minute = x + 95
    col_team = col_minute + 155
    col_player = col_team + 365
    for line_x in (col_minute, col_team, col_player):
        draw.line((line_x, y + header_h, line_x, y + height), fill=COL_BORDER, width=1)

    draw.text((x + 18, y + header_h + 11), "MIN", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_minute + 18, y + header_h + 11), "TEAM", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_team + 18, y + header_h + 11), "PLAYER / EVENT", font=_font("bold", 15), fill=COL_BLACK)
    draw.text((col_player + 18, y + header_h + 11), "NOTES", font=_font("bold", 15), fill=COL_BLACK)

    for idx in range(rows):
        row_y = y + header_h + idx * row_h
        draw.line((x, row_y, x + width, row_y), fill=COL_BORDER, width=1)
        if idx == 0:
            continue
        draw.line((x + 18, row_y + 24, col_minute - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_minute + 18, row_y + 24, col_team - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_team + 18, row_y + 24, col_player - 18, row_y + 24), fill=(190, 196, 204), width=1)
        draw.line((col_player + 18, row_y + 24, x + width - 18, row_y + 24), fill=(190, 196, 204), width=1)
    return y + height


def _draw_soccer_footer(draw: ImageDraw.ImageDraw, y: int = 1288) -> None:
    draw.text((MARGIN, y), "REFEREE COMMENTS", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((MARGIN, y + 44, PAGE_W - MARGIN, y + 44), fill=COL_BORDER, width=1)
    draw.line((MARGIN, y + 82, PAGE_W - MARGIN, y + 82), fill=COL_BORDER, width=1)
    draw.text((MARGIN, y + 122), "REFEREE SIGNATURE:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((312, y + 144, 690, y + 144), fill=COL_BORDER, width=2)
    draw.text((730, y + 122), "SCOREKEEPER:", font=_font("bold", 18), fill=COL_BLACK)
    draw.line((898, y + 144, PAGE_W - MARGIN, y + 144), fill=COL_BORDER, width=2)


def render_basketball_scoresheet_page(
    game: dict[str, Any],
    roster_index: Optional[dict[str, list[dict[str, Any]]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    photo_cache: Optional[PhotoCache] = None,
) -> Image.Image:
    """Render one basketball game score sheet as an RGB image."""

    roster_index = roster_index or {}
    photo_cache = photo_cache or PhotoCache()
    logo = _load_logo(logo_path or default_logo_path())
    page = Image.new("RGB", (PAGE_W, PAGE_H), "white")
    draw = ImageDraw.Draw(page)

    _draw_logo(page, logo)
    _draw_qr(page, str(game.get("game_key") or ""), score_entry_base_url)
    _draw_header(draw, game)
    _draw_score_boxes(draw, game)
    _draw_referee_section(draw)

    team_a_code = _team_code(_team_label(game, "a"))
    team_b_code = _team_code(_team_label(game, "b"))
    table_y = 500
    table_w = (PAGE_W - MARGIN * 2 - 28) // 2
    _draw_roster_table(
        page,
        draw,
        (MARGIN, table_y),
        table_w,
        _team_label(game, "a"),
        roster_index.get(team_a_code, []),
        photo_cache,
    )
    _draw_roster_table(
        page,
        draw,
        (MARGIN + table_w + 28, table_y),
        table_w,
        _team_label(game, "b"),
        roster_index.get(team_b_code, []),
        photo_cache,
    )
    _draw_footer(draw)
    return page


def render_bible_challenge_scoresheet_page(
    game: dict[str, Any],
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
) -> Image.Image:
    """Render one three-team Bible Challenge score sheet as an RGB image."""

    logo = _load_logo(logo_path or default_logo_path())
    page = Image.new("RGB", (PAGE_W, PAGE_H), "white")
    draw = ImageDraw.Draw(page)

    _draw_logo(page, logo)
    _draw_qr(page, str(game.get("game_key") or ""), score_entry_base_url)
    _draw_multi_team_header(draw, game, "BIBLE CHALLENGE SCORESHEET", ("a", "b", "c"), line_break_teams=True)
    _draw_three_team_score_boxes(draw, game)
    _draw_bible_challenge_official_section(draw, y=386)
    next_y = _draw_bible_challenge_score_grid(draw, game, y=548)
    _draw_bible_challenge_notes_grid(draw, next_y + 24)
    _draw_bible_challenge_footer(draw)
    return page


def render_soccer_scoresheet_page(
    game: dict[str, Any],
    roster_index: Optional[dict[str, list[dict[str, Any]]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    photo_cache: Optional[PhotoCache] = None,
) -> Image.Image:
    """Render one soccer game score sheet as an RGB image."""

    roster_index = roster_index or {}
    photo_cache = photo_cache or PhotoCache()
    logo = _load_logo(logo_path or default_logo_path())
    page = Image.new("RGB", (PAGE_W, PAGE_H), "white")
    draw = ImageDraw.Draw(page)

    _draw_logo(page, logo)
    _draw_qr(page, str(game.get("game_key") or ""), score_entry_base_url)
    _draw_generic_header(draw, game, "SOCCER SCORESHEET")
    _draw_score_boxes(draw, game)
    _draw_referee_section(draw, y=372)
    next_y = _draw_soccer_score_grid(draw, game, y=520)
    team_a_code = _team_code(_team_label(game, "a"))
    team_b_code = _team_code(_team_label(game, "b"))
    table_y = next_y + 28
    table_w = (PAGE_W - MARGIN * 2 - 28) // 2
    _draw_soccer_roster_table(
        page,
        draw,
        (MARGIN, table_y),
        table_w,
        _team_label(game, "a"),
        roster_index.get(team_a_code, []),
        photo_cache,
    )
    _draw_soccer_roster_table(
        page,
        draw,
        (MARGIN + table_w + 28, table_y),
        table_w,
        _team_label(game, "b"),
        roster_index.get(team_b_code, []),
        photo_cache,
    )
    _draw_soccer_footer(draw, y=1450)
    return page


def render_volleyball_scoresheet_page(
    game: dict[str, Any],
    roster_index: Optional[dict[str, dict[str, list[dict[str, Any]]]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    photo_cache: Optional[PhotoCache] = None,
) -> Image.Image:
    """Render one volleyball game score sheet as an RGB image."""

    roster_index = roster_index or {}
    photo_cache = photo_cache or PhotoCache()
    logo = _load_logo(logo_path or default_logo_path())
    page = Image.new("RGB", (PAGE_W, PAGE_H), "white")
    draw = ImageDraw.Draw(page)

    _draw_logo(page, logo)
    _draw_qr(page, str(game.get("game_key") or ""), score_entry_base_url)
    title = "MVB SCORESHEET" if game.get("event") == VOLLEYBALL_MEN_EVENT else "WVB SCORESHEET"
    _draw_generic_header(draw, game, title)
    _draw_referee_section(draw, y=288)
    _draw_volleyball_score_grid(draw, game, y=420)

    event = str(game.get("event") or "").strip()
    event_index = roster_index.get(event, {})
    team_a_code = _team_code(_team_label(game, "a"))
    team_b_code = _team_code(_team_label(game, "b"))
    table_y = 740
    table_w = (PAGE_W - MARGIN * 2 - 28) // 2
    _draw_volleyball_roster_table(
        page,
        draw,
        (MARGIN, table_y),
        table_w,
        _team_label(game, "a"),
        event_index.get(team_a_code, []),
        photo_cache,
    )
    _draw_volleyball_roster_table(
        page,
        draw,
        (MARGIN + table_w + 28, table_y),
        table_w,
        _team_label(game, "b"),
        event_index.get(team_b_code, []),
        photo_cache,
    )
    _draw_volleyball_footer(draw)
    return page


def _load_json(path: Path) -> dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ScoreSheetError(f"JSON file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise ScoreSheetError(f"JSON file is not valid at {path}: {exc}") from exc


def _basketball_games(schedule_input: dict[str, Any], schedule_output: dict[str, Any]) -> list[dict[str, Any]]:
    games = [
        game
        for game in merge_schedule(schedule_input, schedule_output)
        if str(game.get("event") or "").strip() == BASKETBALL_EVENT
        and str(game.get("game_key") or "").startswith("BBM-")
    ]
    games.sort(key=lambda game: (str(game.get("scheduled_slot") or ""), str(game.get("resource_id") or ""), str(game.get("game_key") or "")))
    return games


def _bible_challenge_games(schedule_input: dict[str, Any], schedule_output: dict[str, Any]) -> list[dict[str, Any]]:
    games = [
        game
        for game in merge_schedule(schedule_input, schedule_output)
        if str(game.get("event") or "").strip() == BIBLE_CHALLENGE_EVENT
        and str(game.get("game_key") or "").startswith("BC-")
    ]
    games.sort(key=lambda game: (str(game.get("scheduled_slot") or ""), str(game.get("resource_id") or ""), str(game.get("game_key") or "")))
    return games


def _volleyball_games(schedule_input: dict[str, Any], schedule_output: dict[str, Any]) -> list[dict[str, Any]]:
    games = [
        game
        for game in merge_schedule(schedule_input, schedule_output)
        if str(game.get("event") or "").strip() in VOLLEYBALL_EVENTS
        and str(game.get("game_key") or "").startswith(("VBM-", "VBW-"))
    ]
    games.sort(key=lambda game: (str(game.get("scheduled_slot") or ""), str(game.get("resource_id") or ""), str(game.get("game_key") or "")))
    return games


def _soccer_games(schedule_input: dict[str, Any], schedule_output: dict[str, Any]) -> list[dict[str, Any]]:
    games = [
        game
        for game in merge_schedule(schedule_input, schedule_output)
        if str(game.get("event") or "").strip() == SOCCER_EVENT
        and str(game.get("game_key") or "").startswith("SOC-")
    ]
    games.sort(key=lambda game: (str(game.get("scheduled_slot") or ""), str(game.get("resource_id") or ""), str(game.get("game_key") or "")))
    return games


def write_basketball_scoresheets_pdf(
    schedule_input_path: Path,
    schedule_output_path: Path,
    output_dir: Path,
    roster_rows: Optional[list[dict[str, Any]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    output_filename: Optional[str] = None,
) -> tuple[Path, int]:
    """Write one combined basketball score-sheet PDF and return (path, pages)."""

    schedule_input = _load_json(schedule_input_path)
    schedule_output = _load_json(schedule_output_path)
    games = _basketball_games(schedule_input, schedule_output)
    if not games:
        raise ScoreSheetError("No scheduled basketball games found in the supplied schedule artifacts.")

    roster_rows = list(roster_rows or [])
    _warn_if_approval_status_missing(roster_rows, "Basketball")
    roster_index = build_roster_index(roster_rows)
    photo_cache = PhotoCache()
    pages = [
        render_basketball_scoresheet_page(
            game,
            roster_index=roster_index,
            logo_path=logo_path,
            score_entry_base_url=score_entry_base_url,
            photo_cache=photo_cache,
        )
        for game in games
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"Basketball_Scoresheets_{dt.date.today().isoformat()}.pdf"
    output_path = output_dir / filename
    pages[0].save(output_path, "PDF", save_all=True, append_images=pages[1:], resolution=150.0)
    return output_path, len(pages)


def write_bible_challenge_scoresheets_pdf(
    schedule_input_path: Path,
    schedule_output_path: Path,
    output_dir: Path,
    roster_rows: Optional[list[dict[str, Any]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    output_filename: Optional[str] = None,
) -> tuple[Path, int]:
    """Write one combined Bible Challenge score-sheet PDF and return (path, pages)."""

    del roster_rows  # Bible Challenge sheets are match-score sheets, not roster sheets.
    schedule_input = _load_json(schedule_input_path)
    schedule_output = _load_json(schedule_output_path)
    games = _bible_challenge_games(schedule_input, schedule_output)
    if not games:
        raise ScoreSheetError("No scheduled Bible Challenge games found in the supplied schedule artifacts.")

    pages = [
        render_bible_challenge_scoresheet_page(
            game,
            logo_path=logo_path,
            score_entry_base_url=score_entry_base_url,
        )
        for game in games
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"Bible_Challenge_Scoresheets_{dt.date.today().isoformat()}.pdf"
    output_path = output_dir / filename
    pages[0].save(output_path, "PDF", save_all=True, append_images=pages[1:], resolution=150.0)
    return output_path, len(pages)


def write_soccer_scoresheets_pdf(
    schedule_input_path: Path,
    schedule_output_path: Path,
    output_dir: Path,
    roster_rows: Optional[list[dict[str, Any]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    output_filename: Optional[str] = None,
) -> tuple[Path, int]:
    """Write one combined soccer score-sheet PDF and return (path, pages)."""

    schedule_input = _load_json(schedule_input_path)
    schedule_output = _load_json(schedule_output_path)
    games = _soccer_games(schedule_input, schedule_output)
    if not games:
        raise ScoreSheetError("No scheduled soccer games found in the supplied schedule artifacts.")

    roster_rows = list(roster_rows or [])
    _warn_if_approval_status_missing(roster_rows, "Soccer")
    roster_index = build_soccer_roster_index(roster_rows)
    photo_cache = PhotoCache()
    pages = [
        render_soccer_scoresheet_page(
            game,
            roster_index=roster_index,
            logo_path=logo_path,
            score_entry_base_url=score_entry_base_url,
            photo_cache=photo_cache,
        )
        for game in games
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"Soccer_Scoresheets_{dt.date.today().isoformat()}.pdf"
    output_path = output_dir / filename
    pages[0].save(output_path, "PDF", save_all=True, append_images=pages[1:], resolution=150.0)
    return output_path, len(pages)


def write_volleyball_scoresheets_pdf(
    schedule_input_path: Path,
    schedule_output_path: Path,
    output_dir: Path,
    roster_rows: Optional[list[dict[str, Any]]] = None,
    logo_path: Optional[Path] = None,
    score_entry_base_url: Optional[str] = None,
    output_filename: Optional[str] = None,
) -> tuple[Path, int]:
    """Write one combined volleyball score-sheet PDF and return (path, pages)."""

    schedule_input = _load_json(schedule_input_path)
    schedule_output = _load_json(schedule_output_path)
    games = _volleyball_games(schedule_input, schedule_output)
    if not games:
        raise ScoreSheetError("No scheduled volleyball games found in the supplied schedule artifacts.")

    roster_rows = list(roster_rows or [])
    _warn_if_approval_status_missing(roster_rows, "Volleyball")
    roster_index = build_volleyball_roster_index(roster_rows)
    photo_cache = PhotoCache()
    pages = [
        render_volleyball_scoresheet_page(
            game,
            roster_index=roster_index,
            logo_path=logo_path,
            score_entry_base_url=score_entry_base_url,
            photo_cache=photo_cache,
        )
        for game in games
    ]

    output_dir.mkdir(parents=True, exist_ok=True)
    filename = output_filename or f"Volleyball_Scoresheets_{dt.date.today().isoformat()}.pdf"
    output_path = output_dir / filename
    pages[0].save(output_path, "PDF", save_all=True, append_images=pages[1:], resolution=150.0)
    return output_path, len(pages)
