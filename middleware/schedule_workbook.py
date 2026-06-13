# schedule_workbook.py
# Version 1.0.0
# ScheduleWorkbookBuilder — extracted from church_teams_export.py (Issue #98 Step 1)
#
# Contains all scheduling-related methods that previously lived in
# ChurchTeamsExporter.  One-way dependency: church_teams_export.py may import
# from here; this module must NOT import from church_teams_export.
import json
import pandas as pd
from pathlib import Path
from loguru import logger
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from collections import defaultdict
import random
import re
from math import ceil

from config import (
    SPORT_TYPE,
    SPORT_FORMAT,
    FORMAT_MAPPINGS,
    SOCCER_ENABLED,
    RACQUET_SPORTS,
    is_racquet_sport,
    COURT_ESTIMATE_EVENTS,
    COURT_ESTIMATE_RACQUET_EVENTS,
    COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
    COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME,
    COURT_ESTIMATE_MIN_TEAM_SIZE,
    COURT_ESTIMATE_MINUTES_PER_GAME,
    COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
    COURT_ESTIMATE_BC_TEAMS_PER_GAME,
    COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM,
    COURT_ESTIMATE_BC_PLAYOFF_GAMES,
    COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF,
    COURT_ESTIMATE_PLAYOFF_RULES,
    POD_SPORT_ABBREV,
    SCHEDULE_SKETCH_SATURDAY_START,
    SCHEDULE_SKETCH_SATURDAY_LAST_GAME,
    SCHEDULE_SKETCH_SUNDAY_START,
    SCHEDULE_SKETCH_SUNDAY_LAST_GAME,
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
    TEAM_RESOURCE_TYPE_SOCCER,
    SCHEDULE_SOLVER_GYM_COURTS,
    VENUE_INPUT_FILENAME,
    POD_RESOURCE_EVENT_TYPE,
)
from validation.doubles_resolver import Selection as _DblSelection, resolve_doubles as _resolve_dbl
from validation.name_matcher import normalized_name as _norm_name
from validation.models import RulesManager
from schedule_styles import (
    category_style,
    category_prefix,
)
from scheduling import xlsx_utils, venue_loader, output_report
from scheduling import planning_tabs


class ScheduleWorkbookBuilder:
    """Builds schedule-planning workbooks and schedule_input.json from roster data.

    This class contains no API connections and no WordPress connector.
    It is instantiated without arguments and used by export-church-teams
    (via ChurchTeamsExporter) and the standalone build-schedule-workbook command.

    The class-level _rules_manager_cache attribute is set lazily by
    _get_min_team_size; __init__ intentionally does not touch it.
    """

    _rules_manager_cache: Optional[RulesManager] = None
    _rules_manager_cache_failed: bool = False
    _racquet_entry_limit_cache: Dict[str, Tuple[Dict[str, int], Optional[int]]] = {}

    def __init__(self) -> None:
        pass

    _GYM_CORE_SOLVER_POOL = "Gym Core"
    _SPORT_EXPORT_TAB_STATUS: Tuple[str, str, str] = (
        "READ-ONLY OUTPUT",
        "Generated sport-specific roster view. Do not edit; rerun export-church-teams.",
        "D9EAF7",
    )

    _POOL_ASSIGNMENT_COLUMNS: List[str] = [
        "Event",
        "Church Team",
        "Team Order",
        "Team ID",
        "Team Label",
        "Team Source",
        "Roster Count",
        "Min Team Size",
        "Seed",
        "Random Draw Order",
        "Draw Position",
        "Pool ID",
        "Pool Slot",
        "Assignment Basis",
        "Notes",
    ]
    _POOL_ASSIGNMENT_NOTE = (
        "If you add in a seed number for the top teams for fair pool-play, remember to rerun "
        "assign-pools --workbook command (see HOW-TO for details)"
    )
    _POOL_ASSIGNMENT_EVENT_DEFS: List[Tuple[str, str]] = [
        (SPORT_TYPE["BASKETBALL"], "BBM"),
        (SPORT_TYPE["VOLLEYBALL_MEN"], "VBM"),
        (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW"),
        (SPORT_TYPE["BIBLE_CHALLENGE"], "BC"),
    ] + (
        [(SPORT_TYPE["SOCCER"], "SOC")] if SOCCER_ENABLED else []
    )
    _POOL_ASSIGNMENT_HEADER_NOTES: Dict[str, str] = {
        "Event": (
            "Canonical team-sport event this row belongs to. Current Phase 1 "
            "pool-assignment rows cover BB / VBM / VBW / BC"
            + (" / SOC." if SOCCER_ENABLED else ".")
        ),
        "Church Team": (
            "Church code contributing this team row, such as RPC or TLC."
        ),
        "Team Order": (
            "Explicit team order when roster rows already carry one, such as A or B. "
            "Blank means the row currently represents the church-level team grouping."
        ),
        "Team ID": (
            "Stable team identifier used by the planning workbook. Usually the church code, "
            "or church code plus team order when team_order is available."
        ),
        "Team Label": (
            "Human-readable team label for operators. Usually matches Team ID in this first version."
        ),
        "Team Source": (
            "How this row was derived. ChurchLevel means one team row per church. "
            "ExplicitTeamOrder means the roster already carries a team_order value."
        ),
        "Roster Count": (
            "Number of participant roster rows currently contributing to this team row."
        ),
        "Min Team Size": (
            "Minimum roster size from the validation rules used to decide whether this team is eligible."
        ),
        "Seed": (
            "Operator-editable. Leave blank or enter 0 for an unseeded random draw. "
            "Enter 1, 2, 3... for ranked teams placed first in ascending order."
        ),
        "Random Draw Order": (
            "Stable random-draw order used for unseeded teams. Normally not edited by hand."
        ),
        "Draw Position": (
            "Computed overall placement order after seeded teams are sorted first and unseeded teams follow."
        ),
        "Pool ID": (
            "Computed assigned pool, such as P1 or P2."
        ),
        "Pool Slot": (
            "Computed slot within the assigned pool, such as T1 or T3. Under the 3-game "
            "policy, the highest slot in an odd pool (T5 in a 5-team pool, T7 in a 7-team "
            "pool) receives the extra 4th game."
        ),
        "Assignment Basis": (
            "Computed explanation for the row placement: Seeded, SeededDuplicate, "
            "RandomDraw, or WaitingForMoreTeams."
        ),
        "Notes": (
            "Operator notes. Safe to edit."
        ),
    }

    _POD_DIVISION_HEADER_NOTES: Dict[str, str] = {
        "division_id": (
            "Canonical division label used across the planning tabs. "
            "Built from sport + gender + format, for example BAD-Men-Doubles."
        ),
        "sport_type": (
            "Base sport name for this division, such as Badminton, Pickleball, "
            "Table Tennis, or Tennis."
        ),
        "sport_gender": (
            "Gender bucket for the division: Men, Women, or Mixed."
        ),
        "sport_format": (
            "Division format used for planning: Singles, Doubles, or an anomaly placeholder "
            "when the roster data does not fit a normal pod division."
        ),
        "resource_type": (
            "Exact court or table resource type this division needs in the venue plan and solver. "
            "This must line up with the resource names used in venue_input.xlsx and schedule_input.json."
        ),
        "minutes_per_game": (
            "Planning duration per game in minutes for this division."
        ),
        "planning_entries": (
            "Total entries the planner can currently schedule for this division. "
            "For doubles, this counts pairs, not individual people."
        ),
        "confirmed_entries": (
            "Entries currently considered ready for planning after open ERROR-level validation issues "
            "are excluded. For doubles, this also counts pairs, not people."
        ),
        "provisional_entries": (
            "Entries that appear to exist mathematically but still need cleanup before they are fully "
            "confirmed, such as unresolved validation issues."
        ),
        "anomaly_count": (
            "Roster rows that did not fit a normal Singles or Doubles planning entry and therefore "
            "need manual review."
        ),
        "division_status": (
            "Planner readiness flag. Ready = clean to plan, Partial = some entries still need review, "
            "AnomalyOnly = only anomaly rows exist, Empty = no entries found."
        ),
        "notes": (
            "Operator notes column for manual comments or follow-up reminders."
        ),
    }
    _VENUE_ESTIMATOR_HEADER_NOTES: Dict[str, str] = {
        "Event": (
            "Canonical event name used for the venue estimate."
        ),
        "Potential Teams/Entries": (
            "Rule-aware ceiling for this event based on current registrations. Team sports count "
            "all current team units with at least one roster entry; racquet sports cap each "
            "church using the 2026 entry-limit rules while still allowing incomplete doubles "
            "pairs to mature into valid entries."
        ),
        "Estimating Teams/Entries": (
            "Entries currently counted for court estimation. Team sports count only churches "
            "that meet minimum team size; racquet sports count singles plus complete doubles pairs."
        ),
        "Teams": (
            "Comma-separated qualifying team identifiers. Usually these are church codes, but "
            "explicit A/B team splits appear as church code plus team order, such as RPC-A."
        ),
        "Target Pool Games/Team": (
            "Configured planning target for pool games per team. This column is generated "
            "from middleware/config.py and is not edited in Excel. Supported core team-sport "
            "targets are 2 and 3."
        ),
        "Actual Pool Games/Team": (
            "Actual pool-game planning assumption for this event. For the standard team-sport "
            "model this is the average implied by the normalized pool layout; for 3-game "
            "team-sport mode, odd team counts may show a value slightly above 3.0 when one "
            "odd-size pool gives its highest slot an extra game. For Bible "
            "Challenge it reflects the organizer-facing 2-games-per-team target once enough "
            "teams exist to run the Jeopardy format."
        ),
        "Pool Composition": (
            "Pool sizes used by the current planning policy, such as 4 + 3 + 3. Under the "
            "3-game policy, a 5-team or 7-team pool note means the highest slot in that odd "
            "pool takes the extra 4th game."
        ),
        "BYE Slots": (
            "Implicit bye slots created by the current pool layout. Most relevant when 5-team pools "
            "are used under the 2-game planning policy."
        ),
        "Minutes Per Game": (
            "Planning duration in minutes for one game in this event."
        ),
        "Pool Slots": (
            "Total pool-stage game slots needed for this event."
        ),
        "Playoff Teams": (
            "Number of teams assumed to advance from pools into the playoff bracket."
        ),
        "Playoff Slots": (
            "Bracket game slots required, excluding any third-place game."
        ),
        "Third Place?": (
            "Whether the estimate assumes a third-place game for this event."
        ),
        "Third Place Slots": (
            "Additional slots reserved for a third-place game when enabled."
        ),
        "Total Court Slots": (
            "Total planned slots for this event: pool + playoff + third-place."
        ),
        "Estimated Court Hours": (
            "Approximate court-hours required for this event based on total slots and minutes per game."
        ),
    }
    _POD_ENTRY_HEADER_NOTES: Dict[str, str] = {
        "entry_id": (
            "Unique row ID for this review entry within the workbook."
        ),
        "division_id": (
            "Canonical pod division label that this entry belongs to, such as BAD-Men-Doubles."
        ),
        "entry_type": (
            "Planner classification for the entry: Singles, DoublesPair, UnresolvedDoubles, or Anomaly."
        ),
        "participant_1_name": (
            "Primary participant name for the entry."
        ),
        "participant_2_name": (
            "Second participant name for a confirmed doubles pair. Blank for singles, unresolved, or anomaly rows."
        ),
        "source_participant_ids": (
            "Underlying participant IDs that produced this planning entry. Usually WordPress participant IDs, "
            "with ChMeetings IDs as fallback if needed."
        ),
        "church_team": (
            "Church code or codes contributing to this entry."
        ),
        "partner_status": (
            "Partner-matching result. Examples: N/A, Confirmed, MissingPartner, PartnerNotFound, or NonReciprocal."
        ),
        "review_status": (
            "Operator readiness flag. OK means clean enough for planning; NeedsReview means the entry still needs manual attention."
        ),
        "notes": (
            "Reason the entry needs review, plus room for operator notes."
        ),
    }
    _POD_RESOURCE_HEADER_NOTES: Dict[str, str] = {
        "Event": (
            "Racquet event being evaluated for pod-resource fit."
        ),
        "Resource Type": (
            "Exact court or table type that this event requires."
        ),
        "Entries / Teams": (
            "Planning entries counted for this event. Doubles count as complete pairs, not individual people."
        ),
        "Required Slots": (
            "Approximate match slots required under the current single-elimination assumption: entries minus one."
        ),
        "Available Slots": (
            "Slots currently available for this resource type from venue_input.xlsx or the derived schedule-input resources."
        ),
        "Surplus / Shortage": (
            "Available slots minus required slots. Negative values mean the venue is short."
        ),
        "Fit Status": (
            "Traffic-light summary. Green = enough slots, Yellow = small shortage, Red = larger shortage, "
            "No venue data = availability could not be loaded."
        ),
    }
    _SCHEDULE_INPUT_GAME_HEADER_NOTES: Dict[str, str] = {
        "game_id": (
            "Unique placeholder game ID used by the solver and by playoff-slot pinning."
        ),
        "event": (
            "Canonical event name for the game."
        ),
        "stage": (
            "Tournament stage, such as Pool, QF, Semi, Final, or 3rd."
        ),
        "pool_id": (
            "Pool label for pool-stage games. Blank for non-pool games."
        ),
        "round": (
            "Round number within the stage when applicable."
        ),
        "team_a_id": (
            "Placeholder team slot or team ID for side A."
        ),
        "team_b_id": (
            "Placeholder team slot or team ID for side B."
        ),
        "team_c_id": (
            "Optional third team ID for multi-team games such as Bible Challenge."
        ),
        "duration_minutes": (
            "Game duration in minutes used by the solver."
        ),
        "resource_type": (
            "Exact resource type this game must be assigned to."
        ),
        "earliest_slot": (
            "Optional earliest allowed slot constraint for this game."
        ),
        "latest_slot": (
            "Optional latest allowed slot constraint for this game."
        ),
    }
    _SCHEDULE_INPUT_PRECEDENCE_HEADER_NOTES: Dict[str, str] = {
        "before_game_id": (
            "Game that must start before the paired after_game_id."
        ),
        "after_game_id": (
            "Game that must start after before_game_id."
        ),
        "min_gap_slots": (
            "Minimum number of solver slot starts that must separate the two games."
        ),
    }
    _SCHEDULE_INPUT_RESOURCE_HEADER_NOTES: Dict[str, str] = {
        "resource_id": (
            "Exact solver resource ID. This is the value to copy into Playoff-Slots when pinning games."
        ),
        "resource_type": (
            "Resource category offered by this row, such as Basketball Court or Badminton Court."
        ),
        "label": (
            "Human-readable label for the resource, usually what operators will recognize on-site."
        ),
        "day": (
            "Schedule day key, such as Fri-1, Sat-1, Sun-1, Fri-2, or Sun-2."
        ),
        "open_time": (
            "Start of the available scheduling window for this resource."
        ),
        "close_time": (
            "End of the available scheduling window for this resource."
        ),
        "slot_minutes": (
            "Length of each solver slot for this resource."
        ),
        "exclusive_group": (
            "Mutually-exclusive venue group, if any. Resources in the same group cannot all be used at once."
        ),
    }
    _SCHEDULE_INPUT_PLAYOFF_HEADER_NOTES: Dict[str, str] = {
        "game_id": (
            "Exact playoff game ID to pin to a specific slot."
        ),
        "event": (
            "Canonical event name for the pinned playoff game."
        ),
        "stage": (
            "Playoff stage being pinned, such as QF, Semi, Final, or 3rd."
        ),
        "resource_id": (
            "Exact resource ID the playoff game must use."
        ),
        "slot": (
            "Exact solver slot label, typically Day-HH:MM, that the playoff game must occupy."
        ),
    }
    _GYM_ALLOCATION_DECISION_HEADER_NOTES: Dict[str, str] = {
        "gym_name": (
            "Exclusive Venue Group / gym block name being allocated."
        ),
        "day": (
            "Schedule day for this gym block."
        ),
        "open_time": (
            "Start time of the allocated gym block."
        ),
        "close_time": (
            "End time of the allocated gym block."
        ),
        "mode": (
            "Chosen sport mode or resource mode for this block."
        ),
        "courts": (
            "Number of courts assigned to the chosen mode in this block."
        ),
        "slot_minutes": (
            "Slot length in minutes for this block."
        ),
    }
    _GYM_ALLOCATION_SUPPLY_HEADER_NOTES: Dict[str, str] = {
        "mode": (
            "Sport mode or resource mode being compared."
        ),
        "demand": (
            "Total slot demand from games needing this mode."
        ),
        "supply": (
            "Total slot supply produced by the allocator for this mode."
        ),
        "shortfall": (
            "Unmet slots after allocation. Zero means the current allocation covers demand."
        ),
    }

    # ── Shared helpers ───────────────────────────────────────────────────────

    @staticmethod
    def _decompose_event_name(event_name: str) -> Tuple[str, str, str]:
        """Split a canonical event name like 'Basketball - Men Team' into
        (sport_type, sport_gender, sport_format) using the same casing the
        roster rows store after sync."""
        parts = event_name.split(" - ", 1)
        sport_type = parts[0].strip()
        if len(parts) < 2:
            return sport_type, "", "Team"
        suffix = parts[1].upper()
        if "WOMEN" in suffix:
            gender = "Women"
        elif "MEN" in suffix:
            gender = "Men"
        elif "MIXED" in suffix or "COED" in suffix:
            gender = "Mixed"
        else:
            gender = ""
        sport_format = "Singles" if "SINGLES" in suffix else "Team"
        return sport_type, gender, sport_format

    @classmethod
    def _get_rules_manager(cls) -> Optional[RulesManager]:
        """Return the cached SUMMER_2026 RulesManager when available."""
        rules_manager_cache = getattr(cls, "_rules_manager_cache", None)
        rules_manager_cache_failed = bool(getattr(cls, "_rules_manager_cache_failed", False))
        if rules_manager_cache is None and not rules_manager_cache_failed:
            try:
                setattr(cls, "_rules_manager_cache", RulesManager(collection="SUMMER_2026"))
            except Exception as e:
                logger.warning(f"Could not load validation rules for venue estimate: {e}")
                setattr(cls, "_rules_manager_cache_failed", True)
        return getattr(cls, "_rules_manager_cache", None)

    @classmethod
    def _get_min_team_size(cls, event_name: str) -> int:
        """Look up minimum team size from the validation ruleset; fall back
        to COURT_ESTIMATE_MIN_TEAM_SIZE if the JSON rule is absent."""
        rules_manager = cls._get_rules_manager()
        if rules_manager is not None:
            for rule in rules_manager.get_rules_for_sport(event_name):
                if rule.get("rule_type") == "team_size" and rule.get("category") == "min":
                    try:
                        return int(rule.get("value"))
                    except (TypeError, ValueError):
                        pass
        return int(COURT_ESTIMATE_MIN_TEAM_SIZE.get(event_name, 0))

    @classmethod
    def _count_estimating_teams(cls, roster_rows: List[Dict[str, Any]],
                                 event_name: str, min_team_size: int) -> Dict[str, Any]:
        """Return estimating/potential team counts and the qualifying team ids.

        Approval-agnostic — every roster entry counts.

        Returns a dict with:
          n_estimating  – teams with >= min_team_size entries (ready to compete)
          n_potential   – teams with >= 1 but < min_team_size entries (still forming)
          team_codes    – sorted, comma-separated list of estimating team ids
        """
        if min_team_size <= 0:
            return {"n_estimating": 0, "n_potential": 0, "team_codes": ""}
        target_type, target_gender, _ = cls._decompose_event_name(event_name)
        counts_by_team: Dict[Tuple[str, str], int] = {}
        for r in roster_rows:
            r_type = str(r.get("sport_type") or "").strip()
            r_gender = str(r.get("sport_gender") or "").strip()
            # Primary/secondary sports are stored as the base name (e.g. "Basketball");
            # Other-Events sports are stored as the full SPORT_TYPE value verbatim.
            # Accept either so both paths match.
            if (r_type.casefold() != target_type.casefold() and
                    r_type.casefold() != event_name.casefold()):
                continue
            if target_gender and r_gender.casefold() != target_gender.casefold():
                continue
            church = str(r.get("Church Team") or "").strip().upper()
            if not church:
                continue
            team_order = str(r.get("team_order") or "").strip().upper()
            team_key = (church, team_order)
            counts_by_team[team_key] = counts_by_team.get(team_key, 0) + 1
        estimating = sorted(
            church if not team_order else f"{church}-{team_order}"
            for (church, team_order), n in counts_by_team.items()
            if n >= min_team_size
        )
        partial = [
            (church, team_order)
            for (church, team_order), n in counts_by_team.items()
            if 0 < n < min_team_size
        ]
        return {
            "n_estimating": len(estimating),
            "n_potential": len(estimating) + len(partial),  # all team units with >= 1 entry
            "team_codes": ", ".join(estimating),
        }

    @staticmethod
    def _get_playoff_teams(n_teams: int) -> int:
        for rule in COURT_ESTIMATE_PLAYOFF_RULES:
            if rule["min_teams"] <= n_teams <= rule["max_teams"]:
                return int(rule["playoff_teams"])
        return 0

    @staticmethod
    def _get_playoff_teams_for_event(event_name: str, n_teams: int) -> int:
        """Return playoff-team count for one event under the live planning policy."""
        if event_name == SPORT_TYPE["SOCCER"]:
            return 4 if n_teams >= 4 else 0
        return ScheduleWorkbookBuilder._get_playoff_teams(n_teams)

    def _compute_court_slots(self, n_teams: int,
                              minutes_per_game: int = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                              pool_games_per_team: int = COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM,
                              actual_pool_games: Optional[int] = None,
                              event_name: str = "") -> Dict[str, Any]:
        include_third = COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME

        # When the caller already knows the exact game count (e.g. from
        # _make_pool_game_pairs), use it directly so the Venue-Estimator and
        # schedule_input.json stay aligned with the current pool-generation
        # policy instead of falling back to a rough ceil(n*gpg/2) estimate.
        if actual_pool_games is not None:
            pool_slots = actual_pool_games
        else:
            pool_slots = ceil((n_teams * pool_games_per_team) / 2) if n_teams > 0 else 0
        playoff_teams = self._get_playoff_teams_for_event(event_name, n_teams)
        playoff_slots = max(playoff_teams - 1, 0)
        third_place_slots = 1 if include_third and playoff_teams >= 4 else 0
        total_slots = pool_slots + playoff_slots + third_place_slots
        return {
            "pool_games_per_team": pool_games_per_team,
            "minutes_per_game": minutes_per_game,
            "pool_slots": pool_slots,
            "playoff_teams": playoff_teams,
            "playoff_slots": playoff_slots,
            "include_third_place": include_third,
            "third_place_slots": third_place_slots,
            "total_slots": total_slots,
            "court_hours": round(total_slots * minutes_per_game / 60, 2),
        }

    @staticmethod
    def _normalize_racquet_gender(raw_gender: str, raw_format: str = "") -> str:
        """Normalize a racquet roster row's gender to Men/Women/Mixed when possible."""
        tokens = f"{raw_gender} {raw_format}".casefold()
        if "women" in tokens:
            return "Women"
        if "mixed" in tokens or "coed" in tokens:
            return "Mixed"
        if "men" in tokens:
            return "Men"
        return ""

    @classmethod
    def _get_racquet_entry_limits(cls, sport_name: str) -> Tuple[Dict[str, int], Optional[int]]:
        """Return ({format parameter -> max entries}, doubles_total_limit) for one racquet sport."""
        cache = getattr(cls, "_racquet_entry_limit_cache", None)
        if cache is None:
            cache = {}
            setattr(cls, "_racquet_entry_limit_cache", cache)
        cached = cache.get(sport_name)
        if cached is not None:
            return cached

        format_limits: Dict[str, int] = {}
        doubles_total_limit: Optional[int] = None
        rules_manager = cls._get_rules_manager()
        if rules_manager is not None:
            for rule in rules_manager.get_rules_for_sport(sport_name):
                if rule.get("rule_type") != "entry_limit":
                    continue
                parameter = str(rule.get("parameter") or "").strip()
                category = str(rule.get("category") or "").strip()
                try:
                    max_entries = int(rule.get("value"))
                except (TypeError, ValueError):
                    continue
                if category == "format" and parameter:
                    format_limits[parameter] = max_entries
                elif (
                    category == "format_total"
                    and parameter.casefold() == SPORT_FORMAT["DOUBLES"].casefold()
                ):
                    doubles_total_limit = max_entries

        cache[sport_name] = (format_limits, doubles_total_limit)
        return format_limits, doubles_total_limit

    @classmethod
    def _racquet_rule_parameter_for_row(cls, row: Dict[str, Any]) -> Optional[str]:
        """Map one racquet roster row to a church-level entry-limit parameter."""
        raw_format = str(row.get("sport_format") or "").strip()
        format_cf = raw_format.casefold()
        gender = cls._normalize_racquet_gender(
            str(row.get("sport_gender") or "").strip(),
            raw_format,
        )

        if "single" in format_cf:
            if gender in {"Men", "Women"}:
                return f"{gender} Single"
            return None
        if "double" in format_cf or format_cf == SPORT_FORMAT["TEAM"].casefold() or not format_cf:
            if gender in {"Men", "Women", "Mixed"}:
                return f"{gender} Double"
            return None
        return None

    def _count_racquet_entries(self, roster_rows: List[Dict[str, Any]],
                               sport_name: str) -> Dict[str, Any]:
        """Count racquet sport entries for the venue estimator.

        Estimating Entries = complete pairs floor(n_doubles / 2) + n_singles.
        Potential Entries  = rule-aware per-church ceiling using the SUMMER_2026
                             entry-limit rules, while still allowing incomplete
                             doubles registrations to mature into valid pairs.
        """
        n_singles = 0
        n_doubles = 0
        singles_by_church: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        doubles_regs_by_church: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        fallback_potential_by_church: Dict[str, int] = defaultdict(int)
        for r in roster_rows:
            if str(r.get("sport_type") or "").strip().casefold() != sport_name.casefold():
                continue
            fmt = str(r.get("sport_format") or "").strip().casefold()
            church = str(r.get("Church Team") or "").strip().upper()
            parameter = self._racquet_rule_parameter_for_row(r)
            if "single" in fmt:
                n_singles += 1
                if church and parameter:
                    singles_by_church[church][parameter] += 1
                elif church:
                    fallback_potential_by_church[church] += 1
            else:
                n_doubles += 1
                if church and parameter:
                    doubles_regs_by_church[church][parameter] += 1
                elif church:
                    fallback_potential_by_church[church] += 1
        n_estimating = n_singles + (n_doubles // 2)
        format_limits, doubles_total_limit = self._get_racquet_entry_limits(sport_name)
        if format_limits or doubles_total_limit is not None:
            churches = (
                set(singles_by_church.keys())
                | set(doubles_regs_by_church.keys())
                | set(fallback_potential_by_church.keys())
            )
            n_potential = 0
            for church in churches:
                church_total = fallback_potential_by_church.get(church, 0)
                for parameter, count in singles_by_church.get(church, {}).items():
                    limit = format_limits.get(parameter)
                    church_total += min(count, limit) if limit is not None else count
                doubles_total = 0
                for parameter, count in doubles_regs_by_church.get(church, {}).items():
                    team_count = ceil(count / 2)
                    limit = format_limits.get(parameter)
                    doubles_total += min(team_count, limit) if limit is not None else team_count
                if doubles_total_limit is not None:
                    doubles_total = min(doubles_total, doubles_total_limit)
                church_total += doubles_total
                n_potential += church_total
        else:
            n_potential = n_singles + n_doubles
        return {
            "n_estimating": n_estimating,
            "n_potential": n_potential,
            "team_codes": "",
        }

    # ── Pod helpers ──────────────────────────────────────────────────────────

    @staticmethod
    def _pod_format_class(sport_format: str) -> str:
        """Classify a sport_format string into 'singles', 'doubles', or 'anomaly'."""
        fmt = (sport_format or "").strip().casefold()
        if "single" in fmt:
            return "singles"
        if "double" in fmt:
            return "doubles"
        return "anomaly"

    @staticmethod
    def _make_division_id(sport_type: str, sport_gender: str, format_class: str) -> str:
        """Return a canonical division_id string, e.g. 'BAD-Men-Singles'."""
        abbrev = POD_SPORT_ABBREV.get(sport_type, sport_type[:3].upper())
        gender_part = sport_gender or "Unknown"
        format_part = format_class.title() if format_class != "anomaly" else "Anomaly"
        return f"{abbrev}-{gender_part}-{format_part}"

    @staticmethod
    def _build_pod_error_lookup(
        validation_rows: List[Dict[str, Any]],
    ) -> Dict[str, set]:
        """Return {str(participant_id) → set(sport_types)} for open ERRORs."""
        lookup: Dict[str, set] = {}
        for v in validation_rows:
            if str(v.get("Severity", "")).upper() != "ERROR":
                continue
            if str(v.get("Status", "")).lower() != "open":
                continue
            pid = str(v.get("Participant ID (WP)") or "").strip()
            if not pid or pid in ("0",):
                continue
            sport = str(v.get("sport_type") or "").strip()
            lookup.setdefault(pid, set()).add(sport)
        return lookup

    def _build_pod_divisions_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Build one summary row per pod division for the Pod-Divisions tab."""
        error_lookup = self._build_pod_error_lookup(validation_rows)

        # Accumulate per-division counts.
        # key: (sport_type, sport_gender, format_class)
        divs: Dict[tuple, Dict[str, Any]] = {}
        seen_singles_ids: Dict[tuple, set[str]] = defaultdict(set)

        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            sport_gender = str(r.get("sport_gender") or "").strip()
            sport_format = str(r.get("sport_format") or "").strip()
            fmt_class = self._pod_format_class(sport_format)

            key = (sport_type, sport_gender, fmt_class)
            if key not in divs:
                divs[key] = {
                    "sport_type": sport_type,
                    "sport_gender": sport_gender,
                    "sport_format": sport_format,
                    "n_total": 0,
                    "n_confirmed": 0,
                    "n_anomaly": 0,
                }

            pid = str(r.get("Participant ID (WP)") or "").strip()
            if fmt_class == "singles" and pid:
                if pid in seen_singles_ids[key]:
                    continue
                seen_singles_ids[key].add(pid)
            has_error = bool(pid and pid not in ("0",) and sport_type in error_lookup.get(pid, set()))

            if fmt_class == "anomaly":
                divs[key]["n_anomaly"] += 1
            else:
                divs[key]["n_total"] += 1
                if not has_error:
                    divs[key]["n_confirmed"] += 1

        rows: List[Dict[str, Any]] = []
        for (sport_type, sport_gender, fmt_class), div in sorted(divs.items()):
            division_id = self._make_division_id(sport_type, sport_gender, fmt_class)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(sport_type, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)

            if fmt_class == "doubles":
                planning = div["n_total"] // 2
                confirmed = div["n_confirmed"] // 2
            else:
                planning = div["n_total"]
                confirmed = div["n_confirmed"]

            provisional = planning - confirmed
            anomaly_count = div["n_anomaly"]

            if planning == 0 and anomaly_count == 0:
                status = "Empty"
            elif planning == 0:
                status = "AnomalyOnly"
            elif anomaly_count > 0 or provisional > 0:
                status = "Partial"
            else:
                status = "Ready"

            rows.append({
                "division_id": division_id,
                "sport_type": sport_type,
                "sport_gender": sport_gender,
                "sport_format": div["sport_format"],
                # Keep pod division rows aligned with venue_input.xlsx and the solver's
                # exact C4 resource-type matching. Generic "Court" breaks all pod games.
                "resource_type": POD_RESOURCE_EVENT_TYPE.get(sport_type, "Court"),
                "minutes_per_game": mpg,
                "planning_entries": planning,
                "confirmed_entries": confirmed,
                "provisional_entries": provisional,
                "anomaly_count": anomaly_count,
                "division_status": status,
                "notes": "",
            })

        return rows

    def _build_pod_entries_review_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Build one review row per singles player, doubles pair, or anomaly.

        Doubles are matched by reciprocal partner_name declarations using
        normalized name comparison. Unmatched doubles surface as UnresolvedDoubles.
        Confirmed/provisional status is determined by open ERROR presence.
        """
        error_lookup = self._build_pod_error_lookup(validation_rows)

        # Separate racquet entries by format class.
        singles_rows: List[Dict[str, Any]] = []
        doubles_rows: List[Dict[str, Any]] = []
        anomaly_rows: List[Dict[str, Any]] = []

        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            fmt_class = self._pod_format_class(str(r.get("sport_format") or ""))
            if fmt_class == "singles":
                singles_rows.append(r)
            elif fmt_class == "doubles":
                doubles_rows.append(r)
            else:
                anomaly_rows.append(r)

        entry_rows: List[Dict[str, Any]] = []
        entry_counter = 0

        def _full_name(r: Dict[str, Any]) -> str:
            return f"{r.get('First Name', '')} {r.get('Last Name', '')}".strip()

        def _pid(r: Dict[str, Any]) -> str:
            return str(r.get("Participant ID (WP)") or r.get("ChMeetings ID") or "").strip()

        def _has_error(r: Dict[str, Any]) -> bool:
            pid = _pid(r)
            sport = str(r.get("sport_type") or "").strip()
            return bool(pid and pid not in ("0",) and sport in error_lookup.get(pid, set()))

        # Singles — one entry per participant.
        for r in singles_rows:
            entry_counter += 1
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            division_id = self._make_division_id(sport_type, sport_gender, "singles")
            entry_rows.append({
                "entry_id": entry_counter,
                "division_id": division_id,
                "entry_type": "Singles",
                "participant_1_name": _full_name(r),
                "participant_2_name": "",
                "source_participant_ids": _pid(r),
                "church_team": str(r.get("Church Team") or ""),
                "partner_status": "N/A",
                "review_status": "NeedsReview" if _has_error(r) else "OK",
                "notes": "",
            })

        # Doubles — use canonical resolver grouped by division (sport_type + sport_gender).
        doubles_by_div: Dict[tuple, List[Dict[str, Any]]] = {}
        for r in doubles_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            doubles_by_div.setdefault((sport_type, sport_gender), []).append(r)

        for (sport_type, sport_gender), div_rows in sorted(doubles_by_div.items()):
            division_id = self._make_division_id(sport_type, sport_gender, "doubles")

            # Build Selection objects; use index-scoped synthetic IDs when PID is absent.
            row_by_sel_id: Dict[str, Dict[str, Any]] = {}
            sel_objs: List[_DblSelection] = []
            for i, r in enumerate(div_rows):
                pid = _pid(r)
                sel_id = pid if pid else f"_div_{division_id}_idx_{i}"
                name = _full_name(r)
                partner_decl = str(r.get("partner_name") or "").strip()
                sel_objs.append(_DblSelection(
                    participant_id=sel_id,
                    name=name,
                    norm_name=_norm_name(name),
                    partner_name=partner_decl,
                    partner_norm_name=_norm_name(partner_decl),
                    sport_type=sport_type,
                    sport_format=str(r.get("sport_format") or "").strip(),
                    church_code=str(r.get("Church Team") or "").strip(),
                    group_key=division_id,
                ))
                row_by_sel_id[sel_id] = r

            confirmed_pairs, unresolved_recs = _resolve_dbl(sel_objs)

            for pair in confirmed_pairs:
                pid_a, pid_b = pair.participant_ids
                r_a = row_by_sel_id.get(pid_a)
                r_b = row_by_sel_id.get(pid_b)
                both_ok = bool(
                    r_a and not _has_error(r_a)
                    and r_b and not _has_error(r_b)
                )
                churches = ", ".join(c for c in sorted({
                    str((r_a or {}).get("Church Team") or ""),
                    str((r_b or {}).get("Church Team") or ""),
                }) if c)
                entry_counter += 1
                entry_rows.append({
                    "entry_id": entry_counter,
                    "division_id": division_id,
                    "entry_type": "DoublesPair",
                    "participant_1_name": pair.participant_names.get(pid_a, ""),
                    "participant_2_name": pair.participant_names.get(pid_b, ""),
                    "source_participant_ids": ", ".join(filter(None, [pid_a, pid_b])),
                    "church_team": churches,
                    "partner_status": "Confirmed",
                    "sport_format": pair.sport_format,
                    "review_status": "OK" if both_ok else "NeedsReview",
                    "notes": "",
                })

            for rec in unresolved_recs:
                if rec.reason == "SelfPaired":
                    note = "Participant listed themselves as their own partner"
                elif rec.reason == "NonReciprocal":
                    note = f"Partner '{rec.partner_name}' found but did not reciprocate"
                elif rec.reason == "AmbiguousPartner":
                    note = f"Partner '{rec.partner_name}' matches multiple participants"
                elif rec.reason == "MissingPartner":
                    note = "No partner declared"
                else:
                    note = f"Partner '{rec.partner_name}' not in same-division roster"
                # Recover the real PID (may differ from sel_id when synthetic).
                real_pid = _pid(row_by_sel_id.get(rec.participant_id) or {})
                entry_counter += 1
                entry_rows.append({
                    "entry_id": entry_counter,
                    "division_id": division_id,
                    "entry_type": "UnresolvedDoubles",
                    "participant_1_name": rec.name,
                    "participant_2_name": "",
                    "source_participant_ids": real_pid,
                    "church_team": rec.church_code,
                    "partner_status": rec.reason,
                    "sport_format": rec.sport_format,
                    "review_status": "NeedsReview",
                    "notes": note,
                })

        # Anomalies — non-standard format rows for racquet sports.
        for r in anomaly_rows:
            entry_counter += 1
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            division_id = self._make_division_id(sport_type, sport_gender, "anomaly")
            entry_rows.append({
                "entry_id": entry_counter,
                "division_id": division_id,
                "entry_type": "Anomaly",
                "participant_1_name": _full_name(r),
                "participant_2_name": "",
                "source_participant_ids": _pid(r),
                "church_team": str(r.get("Church Team") or ""),
                "partner_status": "N/A",
                "review_status": "NeedsReview",
                "notes": f"Unexpected format '{r.get('sport_format', '')}' for racquet sport",
            })

        return entry_rows

    def _resolve_pod_doubles(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> Tuple[Dict[str, List[Dict[str, Any]]], List[Dict[str, Any]]]:
        """Resolve racquet DOUBLES entries for conflict modeling (Issue #158).

        Reuses the reciprocal-partner pairing in `_build_pod_entries_review_rows`
        so confirmed pairs are resolved exactly once and never diverge.  Returns
        a tuple:

          confirmed_by_division: {division_id: [entry, ...]} where each entry is
            a confirmed doubles pair carrying a stable ``entry_id`` of the form
            ``{division_id}-E{nn}`` (e.g. ``BAD-Men-Doubles-E01``), the pair's
            participant IDs/names, and each member's normalized primary sport.
            Entries are sorted by participant IDs so the IDs are reproducible
            across re-runs of the same roster.
          unprotected: [{division_id, participant_name, reason, notes}] for
            UnresolvedDoubles entries — flagged but not conflict-protected
            because their membership is unknown.

        Singles are intentionally ignored (Decision 3: doubles first).
        """
        # division_id -> (sport_type, sport_gender) for doubles divisions only.
        div_meta: Dict[str, Tuple[str, str]] = {}
        # participant_id -> {name, primary_sport} across all racquet entries.
        pid_info: Dict[str, Dict[str, str]] = {}
        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            sport_gender = str(r.get("sport_gender") or "").strip()
            pid = str(r.get("Participant ID (WP)") or r.get("ChMeetings ID") or "").strip()
            if pid:
                full_name = (
                    f"{str(r.get('First Name') or '').strip()} "
                    f"{str(r.get('Last Name') or '').strip()}"
                ).strip()
                pid_info[pid] = {
                    "name": full_name or pid,
                    "primary_sport": self._normalize_primary_sport_name(
                        r.get("participant_primary_sport")
                    ),
                }
            if self._pod_format_class(str(r.get("sport_format") or "")) == "doubles":
                division_id = self._make_division_id(sport_type, sport_gender, "doubles")
                div_meta[division_id] = (sport_type, sport_gender)

        confirmed_by_division: Dict[str, List[Dict[str, Any]]] = {}
        unprotected: List[Dict[str, Any]] = []
        for row in self._build_pod_entries_review_rows(roster_rows, validation_rows):
            division_id = str(row.get("division_id") or "").strip()
            if row.get("entry_type") == "DoublesPair":
                pids = [
                    p.strip()
                    for p in str(row.get("source_participant_ids") or "").split(",")
                    if p.strip()
                ]
                pair_names = [
                    str(row.get("participant_1_name") or "").strip(),
                    str(row.get("participant_2_name") or "").strip(),
                ]
                sport_type, sport_gender = div_meta.get(division_id, ("", ""))
                confirmed_by_division.setdefault(division_id, []).append({
                    "division_id": division_id,
                    "sport_type": sport_type,
                    "sport_gender": sport_gender,
                    "sport_format": SPORT_FORMAT["DOUBLES"],
                    "participant_ids": pids,
                    "participant_names": {
                        pid: pid_info.get(pid, {}).get("name", pid) for pid in pids
                    },
                    "primary_sports": {
                        pid: pid_info.get(pid, {}).get("primary_sport", "") for pid in pids
                    },
                    "pair_names": pair_names,
                })
            elif row.get("entry_type") == "UnresolvedDoubles":
                sport_type_u, sport_gender_u = div_meta.get(division_id, ("", ""))
                raw_pid = str(row.get("source_participant_ids") or "").strip()
                unprotected.append({
                    "division_id": division_id,
                    "participant_id": raw_pid or None,
                    "participant_name": str(row.get("participant_1_name") or "").strip(),
                    "sport_type": sport_type_u,
                    "sport_format": str(row.get("sport_format") or "").strip(),
                    "sport_gender": sport_gender_u,
                    "church_code": str(row.get("church_team") or "").strip(),
                    "reason": str(row.get("partner_status") or "Unresolved").strip(),
                    "notes": str(row.get("notes") or "").strip(),
                })

        # Assign deterministic, reproducible entry IDs per division.
        for division_id, entries in confirmed_by_division.items():
            entries.sort(key=lambda e: (tuple(sorted(e["participant_ids"])), tuple(e["pair_names"])))
            for idx, entry in enumerate(entries, start=1):
                entry["entry_id"] = f"{division_id}-E{idx:02d}"
                pair_label = " / ".join(n for n in entry["pair_names"] if n)
                entry["label"] = (
                    f"{entry['entry_id']} ({pair_label})" if pair_label else entry["entry_id"]
                )

        return confirmed_by_division, unprotected

    def _resolve_pod_singles(
        self,
        roster_rows: List[Dict[str, Any]],
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Resolve racquet SINGLES entries for conflict modeling (Issue #164).

        Singles membership is always known — one participant per entry, no
        partner declaration to fail — so every singles roster row becomes an
        entry.  Returns ``{division_id: [entry, ...]}`` where each entry
        carries a stable ``entry_id`` of the form ``{division_id}-S{nn}``
        (e.g. ``BAD-Men-Singles-S01``), parallel to the doubles ``-E{nn}``
        model from Issue #158.  Entries are sorted by participant ID so the
        IDs are reproducible across re-runs of the same roster.

        Only the bracket's Round-1 games carry these entry IDs (assigned in
        `_build_pod_game_objects`); bye entries and post-R1 rounds remain
        unprotected because their participation depends on match results.
        """
        by_division: Dict[str, List[Dict[str, Any]]] = {}
        seen_in_div: Dict[str, set] = {}
        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            if self._pod_format_class(str(r.get("sport_format") or "")) != "singles":
                continue
            sport_gender = str(r.get("sport_gender") or "").strip()
            division_id = self._make_division_id(sport_type, sport_gender, "singles")
            pid = str(r.get("Participant ID (WP)") or r.get("ChMeetings ID") or "").strip()
            if pid:
                if pid in seen_in_div.setdefault(division_id, set()):
                    continue  # duplicate roster row for the same player
                seen_in_div[division_id].add(pid)
            full_name = (
                f"{str(r.get('First Name') or '').strip()} "
                f"{str(r.get('Last Name') or '').strip()}"
            ).strip()
            by_division.setdefault(division_id, []).append({
                "division_id": division_id,
                "sport_type": sport_type,
                "sport_gender": sport_gender,
                "sport_format": str(r.get("sport_format") or "").strip(),
                "participant_ids": [pid] if pid else [],
                "participant_names": {pid: full_name or pid} if pid else {},
                "primary_sports": {
                    pid: self._normalize_primary_sport_name(
                        r.get("participant_primary_sport")
                    )
                } if pid else {},
                "player_name": full_name,
            })

        # Assign deterministic, reproducible entry IDs per division.
        for division_id, entries in by_division.items():
            entries.sort(key=lambda e: (tuple(e["participant_ids"]), e["player_name"]))
            for idx, entry in enumerate(entries, start=1):
                entry["entry_id"] = f"{division_id}-S{idx:02d}"
                entry["label"] = (
                    f"{entry['entry_id']} ({entry['player_name']})"
                    if entry["player_name"] else entry["entry_id"]
                )

        return by_division

    # Validation issue types that persist unresolved doubles registrations.
    _POD_PARTNER_ISSUE_TYPES = frozenset({
        "missing_doubles_partner",
        "doubles_partner_unmatched",
    })

    @staticmethod
    def _expected_partner_issue_type(reason: str) -> str:
        """Map a resolver unresolved reason to its persisted issue type."""
        return (
            "missing_doubles_partner"
            if str(reason or "").strip() == "MissingPartner"
            else "doubles_partner_unmatched"
        )

    @staticmethod
    def _partner_validation_key(
        participant_id: Any,
        sport_type: Any,
        sport_format: Any,
        sport_gender: Any = "",
    ) -> Tuple[str, str, str, str]:
        """Return a canonical participant + doubles-event reconciliation key."""
        raw_format = str(sport_format or "").strip()
        if raw_format in FORMAT_MAPPINGS:
            mapped_format, mapped_gender = FORMAT_MAPPINGS[raw_format]
        elif "double" in raw_format.casefold():
            mapped_format = SPORT_FORMAT["DOUBLES"]
            mapped_gender = str(sport_gender or "").strip()
            if not mapped_gender:
                format_cf = raw_format.casefold()
                mapped_gender = next(
                    (
                        gender for gender in ("Men", "Women", "Mixed")
                        if gender.casefold() in format_cf
                    ),
                    "",
                )
        else:
            mapped_format = raw_format
            mapped_gender = str(sport_gender or "").strip()
        return (
            str(participant_id or "").strip(),
            str(sport_type or "").strip().casefold(),
            str(mapped_format or "").strip().casefold(),
            str(mapped_gender or sport_gender or "").strip().casefold(),
        )

    @classmethod
    def _reconcile_pod_validation(
        cls,
        pod_unprotected_entries: List[Dict[str, Any]],
        confirmed_by_division: Dict[str, List[Dict[str, Any]]],
        validation_rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """Reconcile scheduler doubles diagnostics with persisted validation issues.

        Issue #160 acceptance criterion: every scheduler-unprotected doubles
        registration must have a corresponding OPEN partner validation issue
        (missing_doubles_partner / doubles_partner_unmatched) for the same
        participant + sport + format + gender, and confirmed pairs must not
        have one. This check
        runs on the same snapshot the scheduler consumes, so any divergence
        between WordPress validation state and roster-based resolution is
        surfaced instead of silently disagreeing.

        Side effect: annotates each unprotected entry in place with
        ``validation_issue_status`` (Matched / MismatchedIssueType /
        MissingValidationIssue / NoParticipantId) so the Conflict-Audit tab can
        show reconciliation state per entry.
        """
        # Full doubles-event key -> persisted issue details for that event.
        open_partner_issues: Dict[Tuple[str, str, str, str], Dict[str, Any]] = {}
        for v in validation_rows:
            issue_type = str(v.get("Issue Type") or "").strip()
            if issue_type not in cls._POD_PARTNER_ISSUE_TYPES:
                continue
            if str(v.get("Status") or "").strip().lower() != "open":
                continue
            pid = str(v.get("Participant ID (WP)") or "").strip()
            if not pid or pid == "0":
                continue
            sport = str(v.get("sport_type") or "").strip()
            sport_format = str(v.get("sport_format") or "").strip()
            key = cls._partner_validation_key(pid, sport, sport_format)
            bucket = open_partner_issues.setdefault(
                key,
                {
                    "issue_types": set(),
                    "sport_type": sport,
                    "sport_format": sport_format,
                },
            )
            bucket["issue_types"].add(issue_type)

        missing: List[Dict[str, Any]] = []
        mismatched: List[Dict[str, Any]] = []
        matched_count = 0
        scheduler_keys: set = set()

        for entry in pod_unprotected_entries:
            pid = str(entry.get("participant_id") or "").strip()
            sport = str(entry.get("sport_type") or "").strip()
            sport_format = str(entry.get("sport_format") or "").strip()
            sport_gender = str(entry.get("sport_gender") or "").strip()
            reason = str(entry.get("reason") or "").strip()
            record = {
                "participant_id": pid or None,
                "participant_name": entry.get("participant_name", ""),
                "church_code": entry.get("church_code", ""),
                "division_id": entry.get("division_id", ""),
                "sport_type": sport,
                "sport_format": sport_format,
                "sport_gender": sport_gender,
                "reason": reason,
                "expected_issue_type": cls._expected_partner_issue_type(reason),
            }
            if not pid:
                entry["validation_issue_status"] = "NoParticipantId"
                missing.append(record)
                continue

            key = cls._partner_validation_key(
                pid, sport, sport_format, sport_gender
            )
            scheduler_keys.add(key)
            issue_types = open_partner_issues.get(key, {}).get("issue_types", set())
            if record["expected_issue_type"] in issue_types:
                matched_count += 1
                entry["validation_issue_status"] = "Matched"
            elif issue_types:
                record["open_issue_types"] = sorted(issue_types)
                entry["validation_issue_status"] = "MismatchedIssueType"
                mismatched.append(record)
            else:
                entry["validation_issue_status"] = "MissingValidationIssue"
                missing.append(record)

        # Confirmed pairs with a still-open partner issue contradict validation.
        contradictory: List[Dict[str, Any]] = []
        for division_id, entries in sorted((confirmed_by_division or {}).items()):
            for confirmed in entries:
                sport = str(confirmed.get("sport_type") or "").strip()
                sport_format = str(confirmed.get("sport_format") or "").strip()
                sport_gender = str(confirmed.get("sport_gender") or "").strip()
                for pid in confirmed.get("participant_ids", []):
                    key = cls._partner_validation_key(
                        pid, sport, sport_format, sport_gender
                    )
                    scheduler_keys.add(key)
                    bucket = open_partner_issues.get(key)
                    if bucket:
                        contradictory.append({
                            "participant_id": str(pid),
                            "participant_name": confirmed.get(
                                "participant_names", {}
                            ).get(pid, ""),
                            "division_id": division_id,
                            "sport_type": sport,
                            "sport_format": sport_format,
                            "sport_gender": sport_gender,
                            "open_issue_types": sorted(bucket["issue_types"]),
                        })

        # Open partner issues with no scheduler-side record at all (for
        # example, the participant is absent from the exported roster).
        validation_only = [
            {
                "participant_id": pid,
                "sport_type": bucket["sport_type"],
                "sport_format": bucket["sport_format"],
                "open_issue_types": sorted(bucket["issue_types"]),
            }
            for key, bucket in sorted(open_partner_issues.items())
            for pid in (key[0],)
            if key not in scheduler_keys
        ]

        problem_count = (
            len(missing)
            + len(mismatched)
            + len(contradictory)
            + len(validation_only)
        )
        reconciliation = {
            "is_clean": problem_count == 0,
            "problem_count": problem_count,
            "matched_count": matched_count,
            "missing_validation_issues": missing,
            "mismatched_issue_types": mismatched,
            "contradictory_open_issues": contradictory,
            "validation_only_issues": validation_only,
        }

        if problem_count:
            logger.warning(
                "[VAY SM] Pod doubles validation reconciliation found "
                f"{len(missing)} unprotected entrie(s) without an open partner "
                f"validation issue, {len(mismatched)} with a mismatched issue "
                f"type, {len(contradictory)} confirmed pair member(s) with "
                f"a contradictory open issue, and {len(validation_only)} "
                "validation-only issue(s). Scheduler diagnostics and the "
                "Validation-Issues tab disagree — investigate before publishing."
            )
            for record in missing:
                logger.warning(
                    "[VAY SM] Missing partner validation issue: "
                    f"participant_id={record['participant_id']} "
                    f"church={record['church_code']} "
                    f"division={record['division_id']} reason={record['reason']} "
                    f"expected_issue_type={record['expected_issue_type']}"
                )
        else:
            logger.info(
                "[VAY SM] Pod doubles validation reconciliation clean: "
                f"{matched_count} unprotected entrie(s) all have matching open "
                f"partner validation issues; {len(validation_only)} open partner "
                "issue(s) reference participants outside the exported roster."
            )

        return reconciliation

    # ── Venue capacity ───────────────────────────────────────────────────────

    def _build_venue_capacity_rows(self, roster_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows = []

        # Team sports — count churches with a complete roster
        for event_name in COURT_ESTIMATE_EVENTS:
            min_team_size = self._get_min_team_size(event_name)
            counts = self._count_estimating_teams(roster_rows, event_name, min_team_size)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(event_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM)
            pool_plan = self._summarize_pool_policy(counts["n_estimating"], gpg)
            actual = len(self._make_pool_game_pairs("_", counts["n_estimating"], gpg))
            s = self._compute_court_slots(counts["n_estimating"], minutes_per_game=mpg,
                                          pool_games_per_team=gpg,
                                          actual_pool_games=actual,
                                          event_name=event_name)
            rows.append({
                "Event": event_name,
                "Potential Teams/Entries": counts["n_potential"],
                "Estimating Teams/Entries": counts["n_estimating"],
                "Teams": counts["team_codes"],
                "Target Pool Games/Team": pool_plan["target_pool_games_per_team"],
                "Actual Pool Games/Team": pool_plan["actual_pool_games_per_team"],
                "Pool Composition": pool_plan["pool_composition"],
                "BYE Slots": pool_plan["bye_slots"],
                "Minutes Per Game": s["minutes_per_game"],
                "Pool Slots": s["pool_slots"],
                "Playoff Teams": s["playoff_teams"],
                "Playoff Slots": s["playoff_slots"],
                "Third Place?": "Yes" if s["include_third_place"] else "No",
                "Third Place Slots": s["third_place_slots"],
                "Total Court Slots": s["total_slots"],
                "Estimated Court Hours": s["court_hours"],
            })

        # Bible Challenge — sequential single-classroom (Jeopardy, 3 teams/game)
        # Games never run concurrently; "court hours" = total room hours.
        bc_event = SPORT_TYPE["BIBLE_CHALLENGE"]
        bc_min = self._get_min_team_size(bc_event)
        bc_counts = self._count_estimating_teams(roster_rows, bc_event, bc_min)
        n_bc = bc_counts["n_estimating"]
        bc_can_run_rr = n_bc >= COURT_ESTIMATE_BC_TEAMS_PER_GAME
        bc_rr_games = (
            ceil(n_bc * COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM / COURT_ESTIMATE_BC_TEAMS_PER_GAME)
            if bc_can_run_rr else 0
        )
        bc_has_playoff = bc_can_run_rr and n_bc >= COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF
        bc_playoff_games = COURT_ESTIMATE_BC_PLAYOFF_GAMES if bc_has_playoff else 0
        bc_total = bc_rr_games + bc_playoff_games
        bc_hours = round(bc_total * COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE / 60, 2)
        bc_actual_gpg = COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM if bc_can_run_rr else 0
        bc_note = f"Sequential, 1 classroom — {COURT_ESTIMATE_BC_TEAMS_PER_GAME} teams/game"
        if not bc_can_run_rr and n_bc > 0:
            bc_note += (
                f" (waiting for at least {COURT_ESTIMATE_BC_TEAMS_PER_GAME} teams to open "
                "the round-robin queue)"
            )
        elif not bc_has_playoff:
            bc_note += f" (< {COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF} teams: no playoff)"
        rows.append({
            "Event": bc_event,
            "Potential Teams/Entries": bc_counts["n_potential"],
            "Estimating Teams/Entries": n_bc,
            "Teams": bc_counts["team_codes"],
            "Target Pool Games/Team": COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM,
            "Actual Pool Games/Team": bc_actual_gpg,
            "Pool Composition": bc_note,
            "BYE Slots": None,
            "Minutes Per Game": COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
            "Pool Slots": bc_rr_games,
            "Playoff Teams": COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF if bc_has_playoff else 0,
            "Playoff Slots": bc_playoff_games,
            "Third Place?": "No",
            "Third Place Slots": 0,
            "Total Court Slots": bc_total,
            "Estimated Court Hours": bc_hours,
        })

        # Racquet sports — count complete pairs + singles
        for sport_name in COURT_ESTIMATE_RACQUET_EVENTS:
            counts = self._count_racquet_entries(roster_rows, sport_name)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(sport_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
            s = self._compute_court_slots(counts["n_estimating"], minutes_per_game=mpg, event_name=sport_name)
            rows.append({
                "Event": sport_name,
                "Potential Teams/Entries": counts["n_potential"],
                "Estimating Teams/Entries": counts["n_estimating"],
                "Teams": counts["team_codes"],
                "Target Pool Games/Team": None,
                "Actual Pool Games/Team": None,
                "Pool Composition": "",
                "BYE Slots": None,
                "Minutes Per Game": s["minutes_per_game"],
                "Pool Slots": s["pool_slots"],
                "Playoff Teams": s["playoff_teams"],
                "Playoff Slots": s["playoff_slots"],
                "Third Place?": "Yes" if s["include_third_place"] else "No",
                "Third Place Slots": s["third_place_slots"],
                "Total Court Slots": s["total_slots"],
                "Estimated Court Hours": s["court_hours"],
            })

        return rows

    @staticmethod
    def _pool_assignments_sidecar_path(base_dir: Path) -> Path:
        """Return the default sidecar file used to persist editable pool seeds."""
        return Path(base_dir) / "pool_assignments.json"

    @staticmethod
    def _normalize_pool_seed(value: Any) -> Optional[int]:
        """Normalize blank/zero-like seed input to None, else return a positive int."""
        if value in (None, "", "0", 0):
            return None
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    @staticmethod
    def _positive_int_or_none(value: Any) -> Optional[int]:
        """Return a positive int when possible; otherwise None."""
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    @classmethod
    def _pool_assignment_event_prefix(cls, event_name: str) -> str:
        """Return the placeholder prefix used for one pool-assignment event."""
        for known_event, prefix in cls._POOL_ASSIGNMENT_EVENT_DEFS:
            if known_event == event_name:
                return prefix
        return event_name[:3].upper()

    @classmethod
    def _event_sort_index(cls, event_name: str) -> int:
        """Return a stable event ordering for the Pool-Assignment tab."""
        for idx, (known_event, _) in enumerate(cls._POOL_ASSIGNMENT_EVENT_DEFS):
            if known_event == event_name:
                return idx
        return len(cls._POOL_ASSIGNMENT_EVENT_DEFS)

    @classmethod
    def _load_pool_assignment_state(
        cls,
        sidecar_path: Optional[Path],
    ) -> Dict[Tuple[str, str], Dict[str, Any]]:
        """Load persisted seed/draw metadata keyed by (event, team_id)."""
        if not sidecar_path:
            return {}
        sidecar_path = Path(sidecar_path)
        if not sidecar_path.exists():
            return {}

        try:
            payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
        except Exception as exc:
            logger.warning(
                f"Could not parse pool-assignment sidecar '{sidecar_path}': {exc}. "
                "Ignoring persisted seeds for this build."
            )
            return {}

        rows = payload.get("rows", []) if isinstance(payload, dict) else []
        if not isinstance(rows, list):
            logger.warning(
                f"Pool-assignment sidecar '{sidecar_path}' has invalid rows content. "
                "Ignoring persisted seeds for this build."
            )
            return {}

        state: Dict[Tuple[str, str], Dict[str, Any]] = {}
        for row in rows:
            if not isinstance(row, dict):
                continue
            event_name = str(row.get("Event") or row.get("event") or "").strip()
            team_id = str(row.get("Team ID") or row.get("team_id") or "").strip()
            if not event_name or not team_id:
                continue
            state[(event_name, team_id)] = {
                "Seed": cls._normalize_pool_seed(row.get("Seed", row.get("seed"))),
                "Random Draw Order": cls._positive_int_or_none(
                    row.get("Random Draw Order", row.get("random_draw_order"))
                ),
                "Notes": str(row.get("Notes", row.get("notes")) or "").strip(),
            }
        return state

    @classmethod
    def _write_pool_assignment_state(
        cls,
        sidecar_path: Path,
        rows: List[Dict[str, Any]],
    ) -> None:
        """Persist editable Pool-Assignment state to a JSON sidecar."""
        payload_rows: List[Dict[str, Any]] = []
        for row in rows:
            event_name = str(row.get("Event") or "").strip()
            team_id = str(row.get("Team ID") or "").strip()
            if not event_name or not team_id:
                continue
            payload_rows.append({
                "event": event_name,
                "church_code": str(row.get("Church Team") or "").strip(),
                "team_order": str(row.get("Team Order") or "").strip(),
                "team_id": team_id,
                "seed": cls._normalize_pool_seed(row.get("Seed")),
                "random_draw_order": cls._positive_int_or_none(
                    row.get("Random Draw Order")
                ),
                "notes": str(row.get("Notes") or "").strip(),
            })

        sidecar_path = Path(sidecar_path)
        sidecar_path.parent.mkdir(parents=True, exist_ok=True)
        sidecar_path.write_text(
            json.dumps(
                {
                    "version": 1,
                    "updated_at": datetime.now().isoformat(timespec="seconds"),
                    "rows": payload_rows,
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    def _build_pool_assignment_base_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        persisted_state: Optional[Dict[Tuple[str, str], Dict[str, Any]]] = None,
    ) -> List[Dict[str, Any]]:
        """Build one Pool-Assignment row per eligible core gym team."""
        persisted_state = persisted_state or {}
        rows: List[Dict[str, Any]] = []

        for event_name, _prefix in self._POOL_ASSIGNMENT_EVENT_DEFS:
            min_team_size = self._get_min_team_size(event_name)
            target_type, target_gender, _target_format = self._decompose_event_name(event_name)

            counts_by_key: Dict[Tuple[str, str], int] = {}
            for roster_row in roster_rows:
                r_type = str(roster_row.get("sport_type") or "").strip()
                r_gender = str(roster_row.get("sport_gender") or "").strip()
                r_format = str(roster_row.get("sport_format") or "").strip()
                if (
                    r_type.casefold() != target_type.casefold()
                    and r_type.casefold() != event_name.casefold()
                ):
                    continue
                if target_gender and r_gender.casefold() != target_gender.casefold():
                    continue
                if r_format and r_format.casefold() != SPORT_FORMAT["TEAM"].casefold():
                    continue

                church_code = str(roster_row.get("Church Team") or "").strip().upper()
                if not church_code:
                    continue
                team_order = str(roster_row.get("team_order") or "").strip().upper()
                counts_by_key[(church_code, team_order)] = (
                    counts_by_key.get((church_code, team_order), 0) + 1
                )

            for (church_code, team_order), roster_count in sorted(counts_by_key.items()):
                if roster_count < min_team_size:
                    continue

                team_id = church_code if not team_order else f"{church_code}-{team_order}"
                persisted = persisted_state.get((event_name, team_id), {})
                rows.append({
                    "Event": event_name,
                    "Church Team": church_code,
                    "Team Order": team_order,
                    "Team ID": team_id,
                    "Team Label": team_id,
                    "Team Source": "ExplicitTeamOrder" if team_order else "ChurchLevel",
                    "Roster Count": roster_count,
                    "Min Team Size": min_team_size,
                    "Seed": persisted.get("Seed"),
                    "Random Draw Order": persisted.get("Random Draw Order"),
                    "Draw Position": None,
                    "Pool ID": "",
                    "Pool Slot": "",
                    "Assignment Basis": "",
                    "Notes": persisted.get("Notes", ""),
                })

        return rows

    @staticmethod
    def _default_random_draw_orders(
        event_name: str,
        team_ids: List[str],
    ) -> Dict[str, int]:
        """Return a stable pseudo-random ordering for unseeded teams."""
        ordered_ids = sorted(team_ids)
        rng = random.Random(f"vaysf-pool-draw|{event_name}|{'|'.join(ordered_ids)}")
        rng.shuffle(ordered_ids)
        return {team_id: idx for idx, team_id in enumerate(ordered_ids, start=1)}

    @staticmethod
    def _serpentine_pool_slots(pool_sizes: List[int]) -> List[Tuple[str, str]]:
        """Return pool slots in serpentine fill order."""
        slots: List[Tuple[str, str]] = []
        if not pool_sizes:
            return slots

        max_size = max(pool_sizes)
        for slot_idx in range(1, max_size + 1):
            eligible = [pool_idx for pool_idx, size in enumerate(pool_sizes, start=1) if size >= slot_idx]
            if slot_idx % 2 == 0:
                eligible = list(reversed(eligible))
            for pool_idx in eligible:
                slots.append((f"P{pool_idx}", f"T{slot_idx}"))
        return slots

    def _pool_sizes_for_assignment(
        self,
        event_name: str,
        n_teams: int,
    ) -> List[int]:
        """Return the pool sizes implied by the current placeholder-pool policy."""
        if n_teams < 2:
            return []

        prefix = self._pool_assignment_event_prefix(event_name)
        gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
            event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
        )
        pairs = self._make_pool_game_pairs(prefix, n_teams, gpg)
        slot_ids_by_pool: Dict[str, set] = {}
        for team_a_id, team_b_id, pool_id in pairs:
            for team_id in (team_a_id, team_b_id):
                match = re.search(r"-P\d+-T(\d+)$", str(team_id))
                if not match:
                    continue
                slot_ids_by_pool.setdefault(pool_id, set()).add(int(match.group(1)))

        if not slot_ids_by_pool:
            return [n_teams]

        def _pool_key(pool_id: str) -> int:
            try:
                return int(str(pool_id).replace("P", ""))
            except ValueError:
                return 0

        return [
            len(slot_ids_by_pool[pool_id])
            for pool_id in sorted(slot_ids_by_pool.keys(), key=_pool_key)
        ]

    def _apply_pool_assignments_to_rows(
        self,
        rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Compute draw order and pool placement for Pool-Assignment rows."""
        normalized_rows = [dict(row) for row in rows]
        grouped_rows: Dict[str, List[Dict[str, Any]]] = {}
        for row in normalized_rows:
            event_name = str(row.get("Event") or "").strip()
            team_id = str(row.get("Team ID") or "").strip()
            if not event_name or not team_id:
                continue
            row["Seed"] = self._normalize_pool_seed(row.get("Seed"))
            row["Random Draw Order"] = self._positive_int_or_none(row.get("Random Draw Order"))
            row["Notes"] = str(row.get("Notes") or "").strip()
            grouped_rows.setdefault(event_name, []).append(row)

        output_rows: List[Dict[str, Any]] = []
        for event_name, _prefix in self._POOL_ASSIGNMENT_EVENT_DEFS:
            event_rows = grouped_rows.get(event_name, [])
            if not event_rows:
                continue

            duplicate_seed_values = {
                seed
                for seed in {
                    row.get("Seed")
                    for row in event_rows
                    if row.get("Seed") is not None
                }
                if sum(1 for row in event_rows if row.get("Seed") == seed) > 1
            }
            if duplicate_seed_values:
                duplicates = ", ".join(
                    f"{seed}: "
                    + "/".join(
                        sorted(
                            str(row.get("Team ID") or "").strip()
                            for row in event_rows
                            if row.get("Seed") == seed
                        )
                    )
                    for seed in sorted(duplicate_seed_values)
                )
                logger.warning(
                    f"Pool-Assignment duplicate seeds detected for event '{event_name}': "
                    f"{duplicates}"
                )

            existing_draw_orders = {
                row["Team ID"]: row["Random Draw Order"]
                for row in event_rows
                if self._positive_int_or_none(row.get("Random Draw Order")) is not None
            }
            next_draw_order = max(existing_draw_orders.values(), default=0)
            missing_draw_ids = [
                str(row.get("Team ID") or "").strip()
                for row in event_rows
                if str(row.get("Team ID") or "").strip() not in existing_draw_orders
            ]
            if missing_draw_ids:
                for team_id in sorted(
                    missing_draw_ids,
                    key=lambda value: self._default_random_draw_orders(event_name, missing_draw_ids)[value],
                ):
                    next_draw_order += 1
                    existing_draw_orders[team_id] = next_draw_order

            for row in event_rows:
                row["Random Draw Order"] = existing_draw_orders.get(
                    str(row.get("Team ID") or "").strip()
                )

            seeded_rows = sorted(
                [row for row in event_rows if row.get("Seed") is not None],
                key=lambda row: (int(row["Seed"]), str(row.get("Team ID") or "")),
            )
            unseeded_rows = sorted(
                [row for row in event_rows if row.get("Seed") is None],
                key=lambda row: (
                    int(row.get("Random Draw Order") or 0),
                    str(row.get("Team ID") or ""),
                ),
            )
            ordered_rows = seeded_rows + unseeded_rows

            if len(ordered_rows) < 2:
                for draw_position, row in enumerate(ordered_rows, start=1):
                    row["Draw Position"] = draw_position
                    row["Pool ID"] = ""
                    row["Pool Slot"] = ""
                    row["Assignment Basis"] = "WaitingForMoreTeams"
                output_rows.extend(ordered_rows)
                continue

            pool_sizes = self._pool_sizes_for_assignment(event_name, len(ordered_rows))
            slots = self._serpentine_pool_slots(pool_sizes)
            if len(slots) != len(ordered_rows):
                logger.warning(
                    f"Pool-assignment slot count mismatch for event '{event_name}': "
                    f"{len(slots)} slots for {len(ordered_rows)} teams. Leaving pool cells blank."
                )
                for draw_position, row in enumerate(ordered_rows, start=1):
                    row["Draw Position"] = draw_position
                    row["Pool ID"] = ""
                    row["Pool Slot"] = ""
                    row["Assignment Basis"] = (
                        "SeededDuplicate"
                        if row.get("Seed") in duplicate_seed_values
                        else ("Seeded" if row.get("Seed") is not None else "RandomDraw")
                    )
                output_rows.extend(ordered_rows)
                continue

            for draw_position, (row, slot) in enumerate(zip(ordered_rows, slots), start=1):
                row["Draw Position"] = draw_position
                row["Pool ID"] = slot[0]
                row["Pool Slot"] = slot[1]
                row["Assignment Basis"] = (
                    "SeededDuplicate"
                    if row.get("Seed") in duplicate_seed_values
                    else ("Seeded" if row.get("Seed") is not None else "RandomDraw")
                )

            output_rows.extend(ordered_rows)

        return sorted(
            output_rows,
            key=lambda row: (
                self._event_sort_index(str(row.get("Event") or "")),
                int(row.get("Draw Position") or 0),
                str(row.get("Team ID") or ""),
            ),
        )

    def _build_pool_assignment_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        sidecar_path: Optional[Path],
    ) -> List[Dict[str, Any]]:
        """Build Pool-Assignment rows from roster data plus persisted seed state."""
        persisted_state = self._load_pool_assignment_state(sidecar_path)
        base_rows = self._build_pool_assignment_base_rows(roster_rows, persisted_state)
        return self._apply_pool_assignments_to_rows(base_rows)

    @staticmethod
    def _normalize_primary_sport_name(value: Any) -> str:
        """Normalize a declared primary sport value for conflict weighting."""
        return str(value or "").strip()

    @classmethod
    def _solver_team_id(cls, event_name: str, team_id: str) -> str:
        """Return an event-scoped internal team id safe for cross-sport solving."""
        return f"{cls._pool_assignment_event_prefix(event_name)}::{team_id}"

    @classmethod
    def _pool_assignment_placeholder_map(
        cls,
        pool_assignment_rows: List[Dict[str, Any]],
    ) -> Dict[str, Dict[str, Dict[str, Any]]]:
        """Return {event: {PREFIX-Px-Ty: team metadata}} from assigned pool rows."""
        grouped: Dict[str, Dict[str, Dict[str, Any]]] = {}
        for row in pool_assignment_rows:
            event_name = str(row.get("Event") or "").strip()
            pool_id = str(row.get("Pool ID") or "").strip()
            pool_slot = str(row.get("Pool Slot") or "").strip()
            team_id = str(row.get("Team ID") or "").strip()
            if not event_name or not pool_id or not pool_slot or not team_id:
                continue
            prefix = cls._pool_assignment_event_prefix(event_name)
            placeholder_id = f"{prefix}-{pool_id}-{pool_slot}"
            display_label = str(row.get("Team Label") or team_id).strip() or team_id
            grouped.setdefault(event_name, {})[placeholder_id] = {
                "solver_team_id": cls._solver_team_id(event_name, team_id),
                "display_label": display_label,
                "team_id": team_id,
                "pool_id": pool_id,
                "pool_slot": pool_slot,
            }
        return grouped

    @classmethod
    def _build_core_gym_team_lookup(
        cls,
        roster_rows: List[Dict[str, Any]],
    ) -> Dict[Tuple[str, str], Dict[str, Any]]:
        """Return team membership metadata keyed by (event_name, team_id)."""
        team_lookup: Dict[Tuple[str, str], Dict[str, Any]] = {}

        for event_name, _prefix in cls._POOL_ASSIGNMENT_EVENT_DEFS:
            min_team_size = cls._get_min_team_size(event_name)
            target_type, target_gender, _target_format = cls._decompose_event_name(event_name)
            provisional: Dict[Tuple[str, str], Dict[str, Any]] = {}

            for roster_row in roster_rows:
                r_type = str(roster_row.get("sport_type") or "").strip()
                r_gender = str(roster_row.get("sport_gender") or "").strip()
                r_format = str(roster_row.get("sport_format") or "").strip()
                if (
                    r_type.casefold() != target_type.casefold()
                    and r_type.casefold() != event_name.casefold()
                ):
                    continue
                if target_gender and r_gender.casefold() != target_gender.casefold():
                    continue
                if r_format and r_format.casefold() != SPORT_FORMAT["TEAM"].casefold():
                    continue

                church_code = str(roster_row.get("Church Team") or "").strip().upper()
                if not church_code:
                    continue
                team_order = str(roster_row.get("team_order") or "").strip().upper()
                team_id = church_code if not team_order else f"{church_code}-{team_order}"
                key = (event_name, team_id)
                team_state = provisional.setdefault(
                    key,
                    {
                        "event": event_name,
                        "team_id": team_id,
                        "solver_team_id": cls._solver_team_id(event_name, team_id),
                        "display_label": team_id,
                        "participant_ids": set(),
                        "participant_names": {},
                        "primary_sports": {},
                    },
                )
                participant_id = str(
                    roster_row.get("Participant ID (WP)")
                    or roster_row.get("ChMeetings ID")
                    or ""
                ).strip()
                if not participant_id:
                    continue

                team_state["participant_ids"].add(participant_id)
                full_name = (
                    f"{str(roster_row.get('First Name') or '').strip()} "
                    f"{str(roster_row.get('Last Name') or '').strip()}"
                ).strip()
                if full_name:
                    team_state["participant_names"][participant_id] = full_name
                team_state["primary_sports"][participant_id] = cls._normalize_primary_sport_name(
                    roster_row.get("participant_primary_sport")
                )

            for key, team_state in provisional.items():
                if len(team_state["participant_ids"]) < min_team_size:
                    continue
                team_lookup[key] = team_state

        return team_lookup

    @classmethod
    def _build_gym_team_conflicts(
        cls,
        roster_rows: List[Dict[str, Any]],
        pool_assignment_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Return cross-sport shared-athlete edges for the current Phase 1 team sports."""
        team_lookup = cls._build_core_gym_team_lookup(roster_rows)
        if not team_lookup:
            return []

        assigned_rows = {
            (str(row.get("Event") or "").strip(), str(row.get("Team ID") or "").strip())
            for row in pool_assignment_rows
            if str(row.get("Pool ID") or "").strip() and str(row.get("Pool Slot") or "").strip()
        }
        ordered_keys = sorted(
            [key for key in team_lookup.keys() if key in assigned_rows],
            key=lambda item: (cls._event_sort_index(item[0]), item[1]),
        )
        units = {key: cls._team_state_to_unit(team_lookup[key]) for key in ordered_keys}

        conflicts: List[Dict[str, Any]] = []
        for idx, key_a in enumerate(ordered_keys):
            unit_a = units[key_a]
            if not unit_a["participant_ids"]:
                continue
            for key_b in ordered_keys[idx + 1:]:
                if key_a[0] == key_b[0]:
                    continue
                edge = cls._make_shared_athlete_edge(unit_a, units[key_b])
                if edge:
                    conflicts.append(edge)

        return conflicts

    @staticmethod
    def _team_state_to_unit(team_state: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize a gym team-lookup entry into a shared-athlete conflict unit."""
        return {
            "unit_id": team_state["solver_team_id"],
            "label": team_state["display_label"],
            "event": team_state["event"],
            "participant_ids": team_state["participant_ids"],
            "participant_names": team_state["participant_names"],
            "primary_sports": team_state["primary_sports"],
        }

    @staticmethod
    def _make_shared_athlete_edge(
        unit_a: Dict[str, Any],
        unit_b: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """Return one shared-athlete conflict edge, or None when no athlete overlaps.

        A unit is any scheduling entity with a roster: a gym team (Basketball,
        Volleyball, BC, Soccer) or a racquet doubles pair.  The edge dict shape
        is identical for every conflict class so the solver and Conflict-Audit
        tab consume team↔team, team↔racquet, and racquet↔racquet edges the same
        way.  A shared athlete is counted as a *primary* overlap when their
        declared primary sport matches one of the two units' events; otherwise
        it is *secondary* (their primary sport is a third event).
        """
        shared_ids = sorted(set(unit_a["participant_ids"]) & set(unit_b["participant_ids"]))
        if not shared_ids:
            return None

        primary_overlap_count = 0
        shared_names: List[str] = []
        for participant_id in shared_ids:
            primary_sport = (
                unit_a["primary_sports"].get(participant_id)
                or unit_b["primary_sports"].get(participant_id)
                or ""
            )
            if primary_sport and primary_sport.casefold() in {
                str(unit_a["event"]).casefold(),
                str(unit_b["event"]).casefold(),
            }:
                primary_overlap_count += 1
            shared_names.append(
                unit_a["participant_names"].get(
                    participant_id,
                    unit_b["participant_names"].get(participant_id, participant_id),
                )
            )

        return {
            "team_a_id": unit_a["unit_id"],
            "team_a_label": unit_a["label"],
            "event_a": unit_a["event"],
            "team_b_id": unit_b["unit_id"],
            "team_b_label": unit_b["label"],
            "event_b": unit_b["event"],
            "shared_participant_ids": shared_ids,
            "shared_participant_names": shared_names,
            "shared_count": len(shared_ids),
            "primary_overlap_count": primary_overlap_count,
            "secondary_only_count": len(shared_ids) - primary_overlap_count,
        }

    def _build_cross_sport_conflicts(
        self,
        roster_rows: List[Dict[str, Any]],
        pool_assignment_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Return team↔racquet and racquet↔racquet shared-athlete edges (#158, #164).

        Covers the conflict classes the gym-only builder misses:
          (b) a participant in a team sport and a racquet event, and
          (c) a participant in two racquet events.
        Racquet units are confirmed doubles pairs (Issue #158) and all singles
        entries (Issue #164) — singles membership is always known, so every
        singles player is protected in their Round-1 game.  UnresolvedDoubles
        cannot be protected because their membership is unknown.  Edges
        reference the same stable racquet entry IDs that
        `_build_pod_game_objects` assigns to R1 games, so the solver's
        cross-pool avoidance can act on them.  Bye entries and post-R1 rounds
        remain unprotected: their game participation depends on match results.
        """
        # Gym units: team-sport teams actually assigned to a pool.
        team_lookup = self._build_core_gym_team_lookup(roster_rows)
        assigned_rows = {
            (str(row.get("Event") or "").strip(), str(row.get("Team ID") or "").strip())
            for row in pool_assignment_rows
            if str(row.get("Pool ID") or "").strip() and str(row.get("Pool Slot") or "").strip()
        }
        gym_units = [
            self._team_state_to_unit(team_lookup[key])
            for key in team_lookup
            if key in assigned_rows
        ]

        # Racquet units: confirmed doubles pairs with stable entry IDs.
        confirmed_by_div, _unprotected = self._resolve_pod_doubles(roster_rows, validation_rows)
        racquet_units: List[Dict[str, Any]] = []
        for entries in confirmed_by_div.values():
            for entry in entries:
                racquet_units.append({
                    "unit_id": entry["entry_id"],
                    "label": entry["label"],
                    "event": entry["sport_type"],
                    "participant_ids": set(entry["participant_ids"]),
                    "participant_names": entry["participant_names"],
                    "primary_sports": entry["primary_sports"],
                })

        # Racquet units: singles entries with stable entry IDs (Issue #164).
        # Entries with no participant ID can never overlap and are skipped.
        singles_by_div = self._resolve_pod_singles(roster_rows)
        for division_id in sorted(singles_by_div):
            for entry in singles_by_div[division_id]:
                if not entry["participant_ids"]:
                    continue
                racquet_units.append({
                    "unit_id": entry["entry_id"],
                    "label": entry["label"],
                    "event": entry["sport_type"],
                    "participant_ids": set(entry["participant_ids"]),
                    "participant_names": entry["participant_names"],
                    "primary_sports": entry["primary_sports"],
                })

        edges: List[Dict[str, Any]] = []
        # (b) team ↔ racquet
        for gym_unit in gym_units:
            for racquet_unit in racquet_units:
                edge = self._make_shared_athlete_edge(gym_unit, racquet_unit)
                if edge:
                    edges.append(edge)
        # (c) racquet ↔ racquet (distinct entries; same player double-entered in
        # one division produces no overlap and is naturally skipped).
        for idx, unit_a in enumerate(racquet_units):
            for unit_b in racquet_units[idx + 1:]:
                edge = self._make_shared_athlete_edge(unit_a, unit_b)
                if edge:
                    edges.append(edge)
        return edges

    @staticmethod
    def _pool_numeric_suffix(value: str, prefix: str) -> int:
        """Extract the trailing numeric suffix from pool or pool-slot labels."""
        text = str(value or "").strip()
        if not text:
            return 0
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", text)
        if match:
            return int(match.group(1))
        return 0

    @classmethod
    def _bc_no_repeat_triplets(
        cls,
        all_rows: List[Dict[str, Any]],
    ) -> List[Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]]:
        """Generate BC round-robin triplets across all teams in one global pool.

        Each team appears in exactly COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM triplets.
        No pair of teams is in more than one triplet (the "no same opponent twice"
        rule the user expects from traditional BC Jeopardy format).

        Seeded teams (rows with a non-empty Seed value) are never placed in the
        same triplet as another seeded team, preserving the convention that top
        seeds meet each other only in the playoffs.

        Uses backtracking with a most-constrained-first pivot.  For n ≥ 7 with
        3 games/team a valid schedule always exists; for smaller n the constraint
        cannot be satisfied and the method returns [] with a warning.
        """
        ordered = sorted(
            all_rows,
            key=lambda r: (
                cls._pool_numeric_suffix(str(r.get("Pool ID") or ""), "P"),
                cls._pool_numeric_suffix(str(r.get("Pool Slot") or ""), "T"),
                str(r.get("Team ID") or ""),
            ),
        )
        n = len(ordered)
        gpt = COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM
        if n < COURT_ESTIMATE_BC_TEAMS_PER_GAME:
            return []
        n_games = (n * gpt) // COURT_ESTIMATE_BC_TEAMS_PER_GAME

        # Track which indices correspond to seeded teams so the validity check
        # can reject any triplet that pairs two seeded teams together.
        seeded: set = {
            i for i, r in enumerate(ordered)
            if str(r.get("Seed") or "").strip() not in ("", "0")
        }

        from itertools import combinations as _combinations
        all_triples: List[Tuple[int, int, int]] = list(_combinations(range(n), 3))

        used_pairs: set = set()
        count = [0] * n
        chosen: List[Tuple[int, int, int]] = []

        def _pair(a: int, b: int) -> Tuple[int, int]:
            return (a, b) if a < b else (b, a)

        def _valid(a: int, b: int, c: int) -> bool:
            # Seeded teams must not share a triplet with another seeded team.
            seed_count = (a in seeded) + (b in seeded) + (c in seeded)
            return (
                seed_count <= 1
                and count[a] < gpt and count[b] < gpt and count[c] < gpt
                and _pair(a, b) not in used_pairs
                and _pair(b, c) not in used_pairs
                and _pair(a, c) not in used_pairs
            )

        def _apply(a: int, b: int, c: int) -> None:
            count[a] += 1; count[b] += 1; count[c] += 1
            used_pairs.add(_pair(a, b))
            used_pairs.add(_pair(b, c))
            used_pairs.add(_pair(a, c))
            chosen.append((a, b, c))

        def _undo(a: int, b: int, c: int) -> None:
            count[a] -= 1; count[b] -= 1; count[c] -= 1
            used_pairs.discard(_pair(a, b))
            used_pairs.discard(_pair(b, c))
            used_pairs.discard(_pair(a, c))
            chosen.pop()

        def _solve() -> bool:
            if len(chosen) == n_games:
                return True
            incomplete = [i for i in range(n) if count[i] < gpt]
            pivot = min(incomplete, key=lambda i: count[i])
            for a, b, c in all_triples:
                if pivot not in (a, b, c):
                    continue
                if _valid(a, b, c):
                    _apply(a, b, c)
                    if _solve():
                        return True
                    _undo(a, b, c)
            return False

        if not _solve():
            logger.warning(
                "BC: no-repeat triplet schedule is not possible for %d teams "
                "with %d games/team (need n ≥ 7 for 3 games/team). "
                "BC pool games will be omitted.",
                n, gpt,
            )
            return []

        return [tuple(ordered[i] for i in t) for t in chosen]  # type: ignore[return-value]

    @classmethod
    def _build_assigned_bc_game_objects(
        cls,
        pool_assignment_rows: List[Dict[str, Any]],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Return BC queue games and precedence using the assigned BC pool draw.

        All BC teams are treated as a single global pool.  Triplets are generated
        by _bc_no_repeat_triplets so no pair of teams meets more than once before
        the playoffs — equivalent to last year's manually-scheduled 16-team format.
        """
        event_name = SPORT_TYPE["BIBLE_CHALLENGE"]
        prefix = cls._pool_assignment_event_prefix(event_name)
        bc_rows = [
            row
            for row in pool_assignment_rows
            if str(row.get("Event") or "").strip() == event_name
            and str(row.get("Pool ID") or "").strip()
            and str(row.get("Pool Slot") or "").strip()
        ]
        if len(bc_rows) < COURT_ESTIMATE_BC_TEAMS_PER_GAME:
            return [], []

        triplets = cls._bc_no_repeat_triplets(bc_rows)
        if not triplets:
            return [], []

        games: List[Dict[str, Any]] = []
        precedence: List[Dict[str, Any]] = []
        for game_num, trio in enumerate(triplets, start=1):
            solver_team_ids = [
                cls._solver_team_id(event_name, str(row.get("Team ID") or "").strip())
                for row in trio
            ]
            labels = [
                str(row.get("Team Label") or row.get("Team ID") or "").strip()
                for row in trio
            ]
            games.append({
                "game_id": f"{prefix}-RR-{game_num}",
                "event": event_name,
                "stage": "Pool",
                "pool_id": "",
                "round": game_num,
                "team_a_id": solver_team_ids[0],
                "team_b_id": solver_team_ids[1],
                "team_c_id": solver_team_ids[2],
                "team_a_label": labels[0],
                "team_b_label": labels[1],
                "team_c_label": labels[2],
                "duration_minutes": COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
                "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
                "earliest_slot": None,
                "latest_slot": None,
            })

        if len(bc_rows) >= COURT_ESTIMATE_BC_MIN_TEAMS_FOR_PLAYOFF:
            pool_game_ids = [
                str(game.get("game_id") or "").strip()
                for game in games
                if str(game.get("stage") or "").strip() == "Pool"
            ]
            semi_ids: List[str] = []
            for semi_idx in range(1, 4):
                semi_id = f"{prefix}-Semi-{semi_idx}"
                semi_ids.append(semi_id)
                games.append({
                    "game_id": semi_id,
                    "event": event_name,
                    "stage": "Semi",
                    "pool_id": "",
                    "round": semi_idx,
                    "team_a_id": f"{prefix}-Semi-{semi_idx}-A",
                    "team_b_id": f"{prefix}-Semi-{semi_idx}-B",
                    "team_c_id": f"{prefix}-Semi-{semi_idx}-C",
                    "team_a_label": f"Semi {semi_idx} Qualifier A",
                    "team_b_label": f"Semi {semi_idx} Qualifier B",
                    "team_c_label": f"Semi {semi_idx} Qualifier C",
                    "duration_minutes": COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
                    "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
                    "earliest_slot": None,
                    "latest_slot": None,
                })

            precedence.extend(
                {
                    "before_game_id": pool_game_id,
                    "after_game_id": semi_id,
                    "min_gap_slots": 1,
                }
                for pool_game_id in pool_game_ids
                for semi_id in semi_ids
            )

            final_id = f"{prefix}-Final"
            games.append({
                "game_id": final_id,
                "event": event_name,
                "stage": "Final",
                "pool_id": "",
                "round": 1,
                "team_a_id": f"WIN-{semi_ids[0]}",
                "team_b_id": f"WIN-{semi_ids[1]}",
                "team_c_id": f"WIN-{semi_ids[2]}",
                "team_a_label": "Winner Semi 1",
                "team_b_label": "Winner Semi 2",
                "team_c_label": "Winner Semi 3",
                "duration_minutes": COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE,
                "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
                "earliest_slot": None,
                "latest_slot": None,
            })
            precedence.extend(
                {
                    "before_game_id": semi_id,
                    "after_game_id": final_id,
                    "min_gap_slots": 1,
                }
                for semi_id in semi_ids
            )

        return games, precedence

    @classmethod
    def _build_assigned_soccer_game_objects(
        cls,
        roster_rows: List[Dict[str, Any]],
        pool_assignment_rows: List[Dict[str, Any]],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Return Soccer field games plus playoff precedence using the assigned pool draw."""
        if not SOCCER_ENABLED:
            return [], []

        event_name = SPORT_TYPE["SOCCER"]
        prefix = cls._pool_assignment_event_prefix(event_name)
        min_team_size = cls._get_min_team_size(event_name)
        counts = cls._count_estimating_teams(roster_rows, event_name, min_team_size)
        slot_map = cls._pool_assignment_placeholder_map(pool_assignment_rows).get(event_name, {})
        if slot_map:
            n_teams = len(slot_map)
        elif counts["n_estimating"] >= 2:
            n_teams = counts["n_estimating"]
        else:
            return [], []

        gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
            event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
        )
        mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(
            event_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        )

        games: List[Dict[str, Any]] = []
        precedence: List[Dict[str, Any]] = []
        pool_pairs = cls._make_pool_game_pairs(prefix, n_teams, gpg)
        for pair_idx, (team_a_id, team_b_id, pool_id) in enumerate(pool_pairs, start=1):
            team_a_meta = slot_map.get(team_a_id)
            team_b_meta = slot_map.get(team_b_id)
            if slot_map and (team_a_meta is None or team_b_meta is None):
                logger.warning(
                    f"Pool-assignment map for event '{event_name}' is missing "
                    f"{team_a_id!r} or {team_b_id!r}; falling back to placeholders."
                )
            games.append({
                "game_id": f"{prefix}-{pair_idx:02d}",
                "event": event_name,
                "stage": "Pool",
                "pool_id": pool_id,
                "round": pair_idx,
                "team_a_id": (
                    team_a_meta["solver_team_id"] if team_a_meta is not None else team_a_id
                ),
                "team_b_id": (
                    team_b_meta["solver_team_id"] if team_b_meta is not None else team_b_id
                ),
                "team_a_label": (
                    team_a_meta["display_label"] if team_a_meta is not None else team_a_id
                ),
                "team_b_label": (
                    team_b_meta["display_label"] if team_b_meta is not None else team_b_id
                ),
                "duration_minutes": mpg,
                "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
                "earliest_slot": None,
                "latest_slot": None,
            })

        playoff_teams = cls._get_playoff_teams_for_event(event_name, n_teams)
        if playoff_teams >= 4:
            pool_game_ids = [str(game.get("game_id") or "").strip() for game in games]
            semi_ids = [f"{prefix}-Semi-1", f"{prefix}-Semi-2"]
            semi_seed_pairs = [("Seed 1", "Seed 4"), ("Seed 2", "Seed 3")]
            for semi_idx, (semi_id, labels) in enumerate(zip(semi_ids, semi_seed_pairs), start=1):
                games.append({
                    "game_id": semi_id,
                    "event": event_name,
                    "stage": "Semi",
                    "pool_id": "",
                    "round": semi_idx,
                    "team_a_id": f"{prefix}-Seed-{semi_idx}",
                    "team_b_id": f"{prefix}-Seed-{5 - semi_idx}",
                    "team_a_label": labels[0],
                    "team_b_label": labels[1],
                    "duration_minutes": mpg,
                    "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
                    "earliest_slot": None,
                    "latest_slot": None,
                })

            precedence.extend(
                {
                    "before_game_id": pool_game_id,
                    "after_game_id": semi_id,
                    "min_gap_slots": 1,
                }
                for pool_game_id in pool_game_ids
                for semi_id in semi_ids
            )

            final_id = f"{prefix}-Final"
            games.append({
                "game_id": final_id,
                "event": event_name,
                "stage": "Final",
                "pool_id": "",
                "round": 1,
                "team_a_id": f"WIN-{semi_ids[0]}",
                "team_b_id": f"WIN-{semi_ids[1]}",
                "team_a_label": "Winner Semi 1",
                "team_b_label": "Winner Semi 2",
                "duration_minutes": mpg,
                "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
                "earliest_slot": None,
                "latest_slot": None,
            })
            precedence.extend(
                {
                    "before_game_id": semi_id,
                    "after_game_id": final_id,
                    "min_gap_slots": 1,
                }
                for semi_id in semi_ids
            )
            if COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME:
                third_id = f"{prefix}-3rd"
                games.append({
                    "game_id": third_id,
                    "event": event_name,
                    "stage": "3rd",
                    "pool_id": "",
                    "round": 1,
                    "team_a_id": f"LOS-{semi_ids[0]}",
                    "team_b_id": f"LOS-{semi_ids[1]}",
                    "team_a_label": "Loser Semi 1",
                    "team_b_label": "Loser Semi 2",
                    "duration_minutes": mpg,
                    "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
                    "earliest_slot": None,
                    "latest_slot": None,
                })
                precedence.extend(
                    {
                        "before_game_id": semi_id,
                        "after_game_id": third_id,
                        "min_gap_slots": 1,
                    }
                    for semi_id in semi_ids
                )

        return games, precedence

    @staticmethod
    def _warn_if_resource_slot_minutes_differ_from_config(
        all_games: List[Dict[str, Any]],
        all_resources: List[Dict[str, Any]],
    ) -> None:
        """Log advisory warnings when venue slot sizes differ from config durations."""
        expected_minutes_by_resource_type: Dict[str, int] = {
            GYM_RESOURCE_TYPE_BASKETBALL: int(
                COURT_ESTIMATE_MINUTES_PER_GAME.get(
                    SPORT_TYPE["BASKETBALL"],
                    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                )
            ),
            GYM_RESOURCE_TYPE_VOLLEYBALL: int(
                COURT_ESTIMATE_MINUTES_PER_GAME.get(
                    SPORT_TYPE["VOLLEYBALL_MEN"],
                    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                )
            ),
            TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE: int(COURT_ESTIMATE_MINUTES_BIBLE_CHALLENGE),
            TEAM_RESOURCE_TYPE_SOCCER: int(
                COURT_ESTIMATE_MINUTES_PER_GAME.get(
                    SPORT_TYPE["SOCCER"],
                    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                )
            ),
        }

        pod_expected_minutes: Dict[str, set[int]] = defaultdict(set)
        for event_name, resource_type in POD_RESOURCE_EVENT_TYPE.items():
            pod_expected_minutes[resource_type].add(
                int(
                    COURT_ESTIMATE_MINUTES_PER_GAME.get(
                        event_name,
                        COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                    )
                )
            )
        for resource_type, minute_values in pod_expected_minutes.items():
            if len(minute_values) == 1:
                expected_minutes_by_resource_type[resource_type] = next(iter(minute_values))

        scheduled_resource_types = {
            str(game.get("resource_type") or "").strip()
            for game in all_games
            if str(game.get("resource_type") or "").strip()
        }
        slot_minutes_by_resource_type: Dict[str, set[int]] = defaultdict(set)
        for resource in all_resources:
            resource_type = str(resource.get("resource_type") or "").strip()
            if resource_type not in scheduled_resource_types:
                continue
            try:
                slot_minutes = int(resource.get("slot_minutes") or 0)
            except (TypeError, ValueError):
                continue
            if slot_minutes > 0:
                slot_minutes_by_resource_type[resource_type].add(slot_minutes)

        for resource_type in sorted(scheduled_resource_types):
            expected_minutes = expected_minutes_by_resource_type.get(resource_type)
            actual_slot_minutes = sorted(slot_minutes_by_resource_type.get(resource_type, set()))
            if expected_minutes is None or not actual_slot_minutes:
                continue
            if actual_slot_minutes == [expected_minutes]:
                continue

            actual_text = ", ".join(str(value) for value in actual_slot_minutes)
            logger.warning(
                f"Layer 2 duration mismatch for '{resource_type}': config.py game duration is "
                f"{expected_minutes}m but venue_input.xlsx uses slot_minutes [{actual_text}]. "
                "This is only a warning. The solver keeps the config game duration and uses "
                "venue_input slot sizes for capacity, so games may span multiple slots or "
                "consume padded time. If the venue_input values are an intentional real-world "
                "override, you can ignore this warning."
            )

    @classmethod
    def _build_single_elim_playoff(
        cls,
        event_name: str,
        prefix: str,
        playoff_teams: int,
        pool_game_ids: List[str],
        extra_fields: Dict[str, Any],
        include_third: bool,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Build single-elimination playoff games + precedence for a 4- or 8-team bracket.

        Seed pairings (highest seed plays lowest):
          4 teams: Semi-1 = Seed 1 vs Seed 4; Semi-2 = Seed 2 vs Seed 3
          8 teams: QF-1=S1vS8, QF-2=S4vS5, QF-3=S2vS7, QF-4=S3vS6;
                   Semi-1 = WIN-QF-1 vs WIN-QF-2; Semi-2 = WIN-QF-3 vs WIN-QF-4
        Final = WIN-Semi-1 vs WIN-Semi-2; 3rd = LOS-Semi-1 vs LOS-Semi-2 (when
        include_third).  Precedence wires Pool→QF (if any)→Semi→Final/3rd with a
        one-slot gap between rounds.  Returns ([], []) for unsupported sizes.

        ``extra_fields`` is merged into each generated game dict so callers can
        attach sport-specific fields (resource_type, duration_minutes, solver_pool).
        """
        if playoff_teams not in (4, 8):
            return [], []

        games: List[Dict[str, Any]] = []
        precedence: List[Dict[str, Any]] = []

        semi_ids = [f"{prefix}-Semi-1", f"{prefix}-Semi-2"]

        if playoff_teams == 8:
            qf_ids = [f"{prefix}-QF-{i}" for i in range(1, 5)]
            qf_seeds = [(1, 8), (4, 5), (2, 7), (3, 6)]
            for qf_idx, (qf_id, (seed_a, seed_b)) in enumerate(
                zip(qf_ids, qf_seeds), start=1
            ):
                qf_game = {
                    "game_id": qf_id,
                    "event": event_name,
                    "stage": "QF",
                    "pool_id": "",
                    "round": qf_idx,
                    "team_a_id": f"{prefix}-Seed-{seed_a}",
                    "team_b_id": f"{prefix}-Seed-{seed_b}",
                    "team_a_label": f"Seed {seed_a}",
                    "team_b_label": f"Seed {seed_b}",
                }
                qf_game.update(extra_fields)
                games.append(qf_game)
            precedence.extend(
                {"before_game_id": pid, "after_game_id": qid, "min_gap_slots": 1}
                for pid in pool_game_ids for qid in qf_ids
            )
            precedence.extend([
                {"before_game_id": qf_ids[0], "after_game_id": semi_ids[0], "min_gap_slots": 2},
                {"before_game_id": qf_ids[1], "after_game_id": semi_ids[0], "min_gap_slots": 2},
                {"before_game_id": qf_ids[2], "after_game_id": semi_ids[1], "min_gap_slots": 2},
                {"before_game_id": qf_ids[3], "after_game_id": semi_ids[1], "min_gap_slots": 2},
            ])
            semi_team_pairs = [
                (f"WIN-{qf_ids[0]}", f"WIN-{qf_ids[1]}", "Winner QF-1", "Winner QF-2"),
                (f"WIN-{qf_ids[2]}", f"WIN-{qf_ids[3]}", "Winner QF-3", "Winner QF-4"),
            ]
        else:
            semi_team_pairs = [
                (f"{prefix}-Seed-1", f"{prefix}-Seed-4", "Seed 1", "Seed 4"),
                (f"{prefix}-Seed-2", f"{prefix}-Seed-3", "Seed 2", "Seed 3"),
            ]
            precedence.extend(
                {"before_game_id": pid, "after_game_id": sid, "min_gap_slots": 1}
                for pid in pool_game_ids for sid in semi_ids
            )

        for semi_idx, (semi_id, (team_a, team_b, label_a, label_b)) in enumerate(
            zip(semi_ids, semi_team_pairs), start=1
        ):
            semi_game = {
                "game_id": semi_id,
                "event": event_name,
                "stage": "Semi",
                "pool_id": "",
                "round": semi_idx,
                "team_a_id": team_a,
                "team_b_id": team_b,
                "team_a_label": label_a,
                "team_b_label": label_b,
            }
            semi_game.update(extra_fields)
            games.append(semi_game)

        final_id = f"{prefix}-Final"
        final_game = {
            "game_id": final_id,
            "event": event_name,
            "stage": "Final",
            "pool_id": "",
            "round": 1,
            "team_a_id": f"WIN-{semi_ids[0]}",
            "team_b_id": f"WIN-{semi_ids[1]}",
            "team_a_label": "Winner Semi 1",
            "team_b_label": "Winner Semi 2",
        }
        final_game.update(extra_fields)
        games.append(final_game)
        precedence.extend(
            {"before_game_id": sid, "after_game_id": final_id, "min_gap_slots": 2}
            for sid in semi_ids
        )

        if include_third:
            third_id = f"{prefix}-3rd"
            third_game = {
                "game_id": third_id,
                "event": event_name,
                "stage": "3rd",
                "pool_id": "",
                "round": 1,
                "team_a_id": f"LOS-{semi_ids[0]}",
                "team_b_id": f"LOS-{semi_ids[1]}",
                "team_a_label": "Loser Semi 1",
                "team_b_label": "Loser Semi 2",
            }
            third_game.update(extra_fields)
            games.append(third_game)
            precedence.extend(
                {"before_game_id": sid, "after_game_id": third_id, "min_gap_slots": 2}
                for sid in semi_ids
            )

        return games, precedence

    @classmethod
    def _build_assigned_gym_game_objects(
        cls,
        roster_rows: List[Dict[str, Any]],
        pool_assignment_rows: List[Dict[str, Any]],
        allow_placeholder_fallback: bool = True,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Return gym games (pool + auto-generated playoffs) plus precedence.

        Pool games come from the seeded draw in ``pool_assignment_rows``.  Playoff
        games (QF/Semi/Final/3rd) are auto-generated based on the playoff-team count
        and wired into precedence so the solver can place them in any open slot.
        Operators who want a specific game pinned to a court/time still do that
        via the Playoff-Slots tab in venue_input.xlsx — those rows override the
        auto-generated assignments via merge_playoff_slot_assignments.
        """
        sport_defs = [
            (SPORT_TYPE["BASKETBALL"], "BBM", GYM_RESOURCE_TYPE_BASKETBALL),
            (SPORT_TYPE["VOLLEYBALL_MEN"], "VBM", GYM_RESOURCE_TYPE_VOLLEYBALL),
            (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW", GYM_RESOURCE_TYPE_VOLLEYBALL),
        ]
        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        games: List[Dict[str, Any]] = []
        precedence: List[Dict[str, Any]] = []
        placeholder_map_by_event = cls._pool_assignment_placeholder_map(pool_assignment_rows)

        for event_name, prefix, resource_type in sport_defs:
            min_team_size = cls._get_min_team_size(event_name)
            counts = cls._count_estimating_teams(roster_rows, event_name, min_team_size)
            slot_map = placeholder_map_by_event.get(event_name, {})
            if slot_map:
                n_teams = len(slot_map)
            elif counts["n_estimating"] >= 2:
                n_teams = counts["n_estimating"]
            elif allow_placeholder_fallback:
                n_teams = 8
            else:
                continue

            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
                event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
            )
            pool_pairs = cls._make_pool_game_pairs(prefix, n_teams, gpg)
            sport_pool_game_ids: List[str] = []
            for pair_idx, (team_a_id, team_b_id, pool_id) in enumerate(pool_pairs, start=1):
                team_a_meta = slot_map.get(team_a_id)
                team_b_meta = slot_map.get(team_b_id)
                if slot_map and (team_a_meta is None or team_b_meta is None):
                    logger.warning(
                        f"Pool-assignment map for event '{event_name}' is missing "
                        f"{team_a_id!r} or {team_b_id!r}; falling back to placeholders."
                    )
                pool_game_id = f"{prefix}-{pair_idx:02d}"
                sport_pool_game_ids.append(pool_game_id)
                games.append({
                    "game_id": pool_game_id,
                    "event": event_name,
                    "stage": "Pool",
                    "pool_id": pool_id,
                    "round": pair_idx,
                    "team_a_id": (
                        team_a_meta["solver_team_id"] if team_a_meta is not None else team_a_id
                    ),
                    "team_b_id": (
                        team_b_meta["solver_team_id"] if team_b_meta is not None else team_b_id
                    ),
                    "team_a_label": (
                        team_a_meta["display_label"] if team_a_meta is not None else team_a_id
                    ),
                    "team_b_label": (
                        team_b_meta["display_label"] if team_b_meta is not None else team_b_id
                    ),
                    "duration_minutes": mpg,
                    "resource_type": resource_type,
                    "solver_pool": cls._GYM_CORE_SOLVER_POOL,
                    "earliest_slot": None,
                    "latest_slot": None,
                })

            playoff_teams = cls._get_playoff_teams_for_event(event_name, n_teams)
            if playoff_teams >= 4:
                playoff_games, playoff_precedence = cls._build_single_elim_playoff(
                    event_name=event_name,
                    prefix=prefix,
                    playoff_teams=playoff_teams,
                    pool_game_ids=sport_pool_game_ids,
                    extra_fields={
                        "duration_minutes": mpg,
                        "resource_type": resource_type,
                        "solver_pool": cls._GYM_CORE_SOLVER_POOL,
                        "earliest_slot": None,
                        "latest_slot": None,
                    },
                    include_third=COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME,
                )
                games.extend(playoff_games)
                precedence.extend(playoff_precedence)

        return games, precedence

    # ── Schedule-Input JSON builders ─────────────────────────────────────────

    def _build_gym_game_objects(
        self,
        roster_rows: List[Dict[str, Any]],
        allow_placeholder_fallback: bool = True,
    ) -> List[Dict[str, Any]]:
        """Return pool-play game placeholder dicts for gym sports (Basketball, VB Men, VB Women).

        Pool games carry stable placeholder team IDs (e.g. BBM-P1-T1, BBM-P1-T2)
        and non-empty pool_id values so the solver can enforce team-overlap and
        min-rest constraints even before final church assignments are known.

        Playoff games (QF/Semi/Final/3rd) are pre-assigned via the Playoff-Slots
        tab in venue_input.xlsx and are not included here.

        When allow_placeholder_fallback is True, sports with fewer than two
        estimating teams fall back to the legacy 8-team planning scaffold so
        offline sketching still works without venue data. When False, those
        sports are omitted from the solver input entirely.
        """
        sport_defs = [
            (SPORT_TYPE["BASKETBALL"],       "BBM", GYM_RESOURCE_TYPE_BASKETBALL),
            (SPORT_TYPE["VOLLEYBALL_MEN"],   "VBM", GYM_RESOURCE_TYPE_VOLLEYBALL),
            (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW", GYM_RESOURCE_TYPE_VOLLEYBALL),
        ]
        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        games: List[Dict[str, Any]] = []

        for event_name, prefix, resource_type in sport_defs:
            min_sz = self._get_min_team_size(event_name)
            counts = self._count_estimating_teams(roster_rows, event_name, min_sz)
            if counts["n_estimating"] >= 2:
                n_teams = counts["n_estimating"]
            elif allow_placeholder_fallback:
                n_teams = 8
            else:
                continue
            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
                event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
            )

            # Pool games — stable team IDs and non-empty pool_id
            pool_pairs = self._make_pool_game_pairs(prefix, n_teams, gpg)
            for pair_idx, (team_a_id, team_b_id, pool_id) in enumerate(pool_pairs, start=1):
                games.append({
                    "game_id": f"{prefix}-{pair_idx:02d}",
                    "event": event_name,
                    "stage": "Pool",
                    "pool_id": pool_id,
                    "round": pair_idx,
                    "team_a_id": team_a_id,
                    "team_b_id": team_b_id,
                    "duration_minutes": mpg,
                    "resource_type": resource_type,
                    "earliest_slot": None,
                    "latest_slot":   None,
                })

        return games

    def _build_pod_game_objects(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Return (games, precedence) for single-elimination brackets for pod (racquet) sports.

        Uses planning_entries (confirmed + provisional) from Pod-Divisions.
        Elimination games per division = planning_entries − 1, plus an optional
        third-place game when two Semis are played.
        Divisions with fewer than 2 planning entries are skipped.

        Late rounds receive stage-aware IDs so operators can pin them in
        Playoff-Slots (Issue #130):
          - ``-QF-N``   quarter-final games (bracket size >= 8)
          - ``-Semi-N`` semi-final games
          - ``-Final``  championship game
          - ``-3rd``    third-place game when enabled and two Semis are played
        Early rounds keep sequential numeric IDs (``-01``, ``-02``, ...).

        Precedence edges enforce round ordering: every game in round R must
        complete (min_gap_slots=1) before any game in round R+1 can start.

        For doubles divisions, the first ``floor(c / 2)`` games (where ``c`` is
        the number of *confirmed* pairs) are real Round-1 matchups carrying the
        stable entry IDs from `_resolve_pod_doubles`, so the solver can keep a
        shared athlete's R1 game off another sport's slot (Issue #158,
        Decision 1).  Singles divisions get the same Round-1 protection using
        the ``-S{nn}`` entry IDs from `_resolve_pod_singles` (Issue #164).
        Later-round games, bye entries, and any games beyond the known
        matchups keep ``team_a_id``/``team_b_id`` of ``None`` because their
        participants are not knowable until earlier rounds are played.
        """
        div_rows = self._build_pod_divisions_rows(roster_rows, validation_rows)
        confirmed_by_div, _unprotected = self._resolve_pod_doubles(roster_rows, validation_rows)
        singles_by_div = self._resolve_pod_singles(roster_rows)
        all_games: List[Dict[str, Any]] = []
        all_precedence: List[Dict[str, Any]] = []

        for div in div_rows:
            if div["division_status"] in ("Empty", "AnomalyOnly"):
                continue
            fmt_class = self._pod_format_class(str(div.get("sport_format") or ""))
            if fmt_class == "singles":
                n_entries = len(singles_by_div.get(div["division_id"], []))
            else:
                n_entries = div["planning_entries"]
            if n_entries < 2:
                continue
            division_id = div["division_id"]
            sport_type = div["sport_type"]
            resource_type = div["resource_type"]
            mpg = div["minutes_per_game"]

            # Bracket size P = smallest power of 2 >= n_entries.
            P = 1
            while P < n_entries:
                P <<= 1
            n_rounds = P.bit_length() - 1  # log2(P)

            # Stage name for each bracket round (from the Final backward).
            def _round_stage(r: int, _nr: int = n_rounds) -> str:
                rounds_from_end = _nr - r
                if rounds_from_end == 0:
                    return "Final"
                if rounds_from_end == 1:
                    return "Semi"
                if rounds_from_end == 2:
                    return "QF"
                return f"R{r}"  # pre-QF rounds for very large brackets

            # Round-1 matchup assignment for divisions with known entries —
            # confirmed doubles pairs (Issue #158) and all singles players
            # (Issue #164).  Bracket math mirrors the outer P computation so
            # byes are placed correctly.
            r1_matchups: List[Tuple[Optional[str], Optional[str]]] = []
            if fmt_class == "doubles":
                entries = confirmed_by_div.get(division_id, [])
            elif fmt_class == "singles":
                entries = singles_by_div.get(division_id, [])
            else:
                entries = []
            if entries and n_entries >= 2:
                byes = P - n_entries
                r1_match_count = (n_entries - byes) // 2
                planned_entry_ids: List[Optional[str]] = [
                    entry["entry_id"] for entry in entries[:n_entries]
                ]
                planned_entry_ids.extend([None] * (n_entries - len(planned_entry_ids)))
                r1_starters = planned_entry_ids[n_entries - 2 * r1_match_count:]
                for k in range(0, 2 * r1_match_count, 2):
                    r1_matchups.append((r1_starters[k], r1_starters[k + 1]))

            # Generate games round by round.
            # Round 1 actual games: n_entries - P//2  (byes happen only in round 1)
            # Round r > 1: P // 2^r games  (no more byes)
            div_games: List[Dict[str, Any]] = []
            games_by_round: Dict[int, List[str]] = {}
            stage_counters: Dict[str, int] = {}
            early_seq = 0   # sequential counter for pre-QF numeric IDs
            game_idx = 0    # total games generated (for doubles matchup assignment)

            for r in range(1, n_rounds + 1):
                stage = _round_stage(r)
                n_games = (
                    max(0, n_entries - P // 2) if r == 1 else P // (2 ** r)
                )
                stage_counters.setdefault(stage, 0)
                games_by_round[r] = []

                for _ in range(n_games):
                    stage_counters[stage] += 1
                    cnt = stage_counters[stage]

                    if stage == "Final":
                        game_id = f"{division_id}-Final"
                    elif stage in ("Semi", "QF"):
                        game_id = f"{division_id}-{stage}-{cnt}"
                    else:
                        early_seq += 1
                        game_id = f"{division_id}-{early_seq:02d}"

                    games_by_round[r].append(game_id)

                    if game_idx < len(r1_matchups):
                        team_a_id, team_b_id = r1_matchups[game_idx]
                    else:
                        team_a_id, team_b_id = None, None
                    game_idx += 1

                    div_games.append({
                        "game_id":              game_id,
                        "division_id":          division_id,
                        "division_entry_count": n_entries,
                        "event":                sport_type,
                        "stage":                stage,
                        "pool_id":              "",
                        "round":                r,
                        "team_a_id":            team_a_id,
                        "team_b_id":            team_b_id,
                        "duration_minutes":     mpg,
                        "resource_type":        resource_type,
                        "earliest_slot":        None,
                        "latest_slot":          None,
                    })

            # Precedence: every game in round R must finish before round R+1 starts.
            for r in range(1, n_rounds):
                for before_id in games_by_round[r]:
                    for after_id in games_by_round[r + 1]:
                        all_precedence.append({
                            "before_game_id": before_id,
                            "after_game_id":  after_id,
                            "min_gap_slots":  1,
                        })

            semi_round = n_rounds - 1
            semi_ids = games_by_round.get(semi_round, [])
            if (
                COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME
                and len(semi_ids) == 2
            ):
                third_id = f"{division_id}-3rd"
                div_games.append({
                    "game_id":              third_id,
                    "division_id":          division_id,
                    "division_entry_count": n_entries,
                    "event":                sport_type,
                    "stage":                "3rd",
                    "pool_id":              "",
                    "round":                n_rounds,
                    "team_a_id":            None,
                    "team_b_id":            None,
                    "duration_minutes":     mpg,
                    "resource_type":        resource_type,
                    "earliest_slot":        None,
                    "latest_slot":          None,
                })
                all_precedence.extend(
                    {
                        "before_game_id": semi_id,
                        "after_game_id": third_id,
                        "min_gap_slots": 1,
                    }
                    for semi_id in semi_ids
                )

            all_games.extend(div_games)

        return all_games, all_precedence

    @staticmethod
    def _build_gym_resource_objects(
        n_basketball: int = 2,
        n_volleyball: int = 2,
    ) -> List[Dict[str, Any]]:
        """Fallback gym resource builder (no venue_input.xlsx).

        Generates n_basketball + n_volleyball resources per session across four
        sessions (Sat-1, Sun-1, Sat-2, Sun-2) using SCHEDULE_SKETCH_* time
        windows.  Basketball courts are numbered first within each session,
        volleyball courts second.
        """
        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        close_sat = f"{SCHEDULE_SKETCH_SATURDAY_LAST_GAME + mpg // 60:02d}:00"
        close_sun = f"{SCHEDULE_SKETCH_SUNDAY_LAST_GAME + mpg // 60:02d}:00"
        sessions = [
            ("Sat-1", f"{SCHEDULE_SKETCH_SATURDAY_START:02d}:00", close_sat),
            ("Sun-1", f"{SCHEDULE_SKETCH_SUNDAY_START:02d}:00",   close_sun),
            ("Sat-2", f"{SCHEDULE_SKETCH_SATURDAY_START:02d}:00", close_sat),
            ("Sun-2", f"{SCHEDULE_SKETCH_SUNDAY_START:02d}:00",   close_sun),
        ]
        type_blocks = [
            (GYM_RESOURCE_TYPE_BASKETBALL, n_basketball),
            (GYM_RESOURCE_TYPE_VOLLEYBALL, n_volleyball),
        ]
        resources: List[Dict[str, Any]] = []
        for day_label, open_time, close_time in sessions:
            c = 0
            for rtype, count in type_blocks:
                for local in range(1, count + 1):
                    c += 1
                    resources.append({
                        "resource_id":     f"GYM-{day_label}-{c}",
                        "resource_type":   rtype,
                        "label":           f"Court-{c}",
                        "day":             day_label,
                        "open_time":       open_time,
                        "close_time":      close_time,
                        "slot_minutes":    mpg,
                        "exclusive_group": "",
                    })
        return resources

    @staticmethod
    def _build_gym_resources_from_allocator(decisions) -> List[Dict[str, Any]]:
        """Convert Stage-A AllocationDecision objects into schedule_input resources.

        Courts within each decision are numbered per-day sequentially so that
        resource IDs remain stable across re-runs with the same venue configuration.
        """
        resources: List[Dict[str, Any]] = []
        court_counter: Dict[str, int] = {}  # day → running counter

        # Sort for ID stability: day order, then open_time, then gym name.
        from gym_allocator import _day_sort_key
        sorted_decisions = sorted(
            decisions,
            key=lambda d: (_day_sort_key(d.day), d.open_time, d.gym_name),
        )
        for decision in sorted_decisions:
            day = decision.day
            for local in range(1, decision.courts + 1):
                court_counter[day] = court_counter.get(day, 0) + 1
                n = court_counter[day]
                resources.append({
                    "resource_id":     f"GYM-{day}-{n}",
                    "resource_type":   decision.mode,
                    "label":           f"Court-{local}",
                    "day":             day,
                    "open_time":       decision.open_time,
                    "close_time":      decision.close_time,
                    "slot_minutes":    decision.slot_minutes,
                    "venue_name":      decision.gym_name,
                    "exclusive_group": decision.gym_name,
                })
        return resources

    def _resolve_venue_playoff_slots(
        self,
        playoff_slots: List[Dict[str, Any]],
        venue_rows: List[Dict[str, Any]],
        date_day_map: Dict[str, str],
        gym_modes: Dict[str, Dict[str, int]],
        allocator_active: bool,
        game_duration_by_id: Optional[Dict[str, int]] = None,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Any], set]:
        """Resolve venue-centric Playoff-Slots rows into concrete resources (Issue #127).

        A venue-centric row specifies gym_name + date + start_time instead of an
        internal resource_id + slot.  Resolution validates the venue against the
        Venue-Input rows and fills resource_id/slot in place:

        - **Allocator-managed gyms** (row's Exclusive Venue Group has a Gym-Modes
          entry and the Stage-A allocator is active): a playoff-pinned synthetic
          resource is created covering only the pinned window, and that window is
          returned as a reservation the allocator must skip.  Contiguous pins on
          the same gym/sport merge into one synthetic court track, mirroring the
          legacy same-resource_id merge behavior.
        - **Direct/standalone resources**: the row resolves to an existing
          expanded resource_id; validate_playoff_slots() reserves the exact
          (resource, slot) pair from pool play as it always has.

        Rows that already carry resource_id + slot pass through untouched (the
        explicit form remains valid as an override). Invalid venue-centric rows
        fail the build together so playoff intent is never silently omitted.

        Returns (resolved_playoff_slots, synthetic_resources, reserved_windows,
        synthetic_resource_ids).
        """
        from gym_allocator import EVENT_TO_MODE as _EVENT_TO_MODE, GymBlock

        event_to_resource_type = dict(_EVENT_TO_MODE)
        event_to_resource_type[SPORT_TYPE["BIBLE_CHALLENGE"]] = TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE
        event_to_resource_type[SPORT_TYPE["SOCCER"]] = TEAM_RESOURCE_TYPE_SOCCER

        def _minutes(text: str) -> int:
            hours, mins = text.split(":")
            return int(hours) * 60 + int(mins)

        def _clock(total: int) -> str:
            return f"{total // 60:02d}:{total % 60:02d}"

        def _overlaps(left: Tuple[int, int], right: Tuple[int, int]) -> bool:
            return left[0] < right[1] and right[0] < left[1]

        def _record_error(message: str) -> None:
            logger.error(message)
            errors.append(message)

        game_duration_by_id = game_duration_by_id or {}
        resolved: List[Dict[str, Any]] = []
        synthetic_resources: List[Dict[str, Any]] = []
        reserved_windows: List[Any] = []
        synthetic_ids: set = set()
        errors: List[str] = []
        # (gym exclusive_group, day, resource_type) -> pins for court tracks.
        managed_pending: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
        managed_by_gym_day: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
        used_intervals_by_resource: Dict[str, List[Tuple[int, int, str]]] = defaultdict(list)

        for entry in playoff_slots:
            if entry.get("resource_id") and entry.get("slot"):
                if entry.get("gym_name") or entry.get("date") or entry.get("start_time"):
                    logger.info(
                        f"Playoff slot {entry['game_id']!r}: explicit resource_id+slot "
                        "takes precedence over gym_name/date/start_time."
                    )
                game_id = str(entry.get("game_id") or "").strip()
                if game_id in game_duration_by_id:
                    entry.setdefault(
                        "duration_minutes", game_duration_by_id[game_id]
                    )
                resolved.append(entry)
                continue

            game_id = entry.get("game_id", "<unknown>")
            gym_name = str(entry.get("gym_name") or "").strip()
            date_text = str(entry.get("date") or "").strip()
            start_text = str(entry.get("start_time") or "").strip()

            if re.fullmatch(r"[A-Za-z]+-\d+", date_text):
                day = date_text
            else:
                day = date_day_map.get(date_text, "")
            if not day:
                _record_error(
                    f"Playoff slot {game_id!r}: date {date_text!r} matches no "
                    "Venue-Input date. Use a date that appears in "
                    "the Venue-Input tab (or a day label such as 'Sun-2')."
                )
                continue

            event = str(entry.get("event") or "").strip()
            resource_type = event_to_resource_type.get(event)
            if not resource_type:
                _record_error(
                    f"Playoff slot {game_id!r}: cannot infer a resource type from "
                    f"event {event!r}. The event must exactly match "
                    "a scheduled sport name."
                )
                continue

            gym_key = gym_name.casefold()
            candidates = [
                row for row in venue_rows
                if str(row.get("day") or "").strip() == day
                and str(row.get("resource_type") or "").strip() == resource_type
                and gym_key in (
                    str(row.get("venue_name") or "").strip().casefold(),
                    str(row.get("exclusive_group") or "").strip().casefold(),
                )
            ]
            if not candidates:
                _record_error(
                    f"Playoff slot {game_id!r}: no Venue-Input row found for gym "
                    f"{gym_name!r} on {day} with a {resource_type!r}. "
                    "Check the gym name, date, and that the venue offers this sport."
                )
                continue

            start_min = _minutes(start_text)
            window_rows: List[Tuple[Dict[str, Any], int, int, int, int]] = []
            for row in candidates:
                row_slot_min = int(row.get("slot_minutes") or 60)
                exclusive_group = str(row.get("exclusive_group") or "").strip()
                managed = bool(
                    allocator_active
                    and exclusive_group
                    and exclusive_group in gym_modes
                )
                grid_minutes = (
                    int(entry.get("slot_minutes") or row_slot_min)
                    if managed
                    else row_slot_min
                )
                if (
                    not managed
                    and entry.get("slot_minutes")
                    and int(entry["slot_minutes"]) != row_slot_min
                ):
                    continue
                duration = int(
                    entry.get("duration_minutes")
                    or game_duration_by_id.get(str(game_id))
                    or grid_minutes
                )
                occupied_minutes = (
                    (duration + grid_minutes - 1) // grid_minutes
                ) * grid_minutes
                open_min = _minutes(str(row.get("open_time")))
                close_min = _minutes(str(row.get("close_time")))
                if open_min <= start_min and start_min + occupied_minutes <= close_min:
                    window_rows.append(
                        (row, grid_minutes, duration, occupied_minutes, open_min)
                    )
            if not window_rows:
                windows = sorted({
                    f"{row.get('open_time')}-{row.get('close_time')}" for row in candidates
                })
                _record_error(
                    f"Playoff slot {game_id!r}: start_time {start_text!r} and its "
                    "game duration do not fit "
                    f"inside any {gym_name!r} window on {day} ({', '.join(windows)}) "
                    "on the venue's slot grid."
                )
                continue

            source_row, grid_minutes, duration, occupied_minutes, _open_min = window_rows[0]
            exclusive_group = str(source_row.get("exclusive_group") or "").strip()
            managed = bool(
                allocator_active and exclusive_group and exclusive_group in gym_modes
            )
            entry.setdefault("duration_minutes", duration)

            if managed:
                pending = {
                    "entry":       entry,
                    "start_min":   start_min,
                    "end_min":     start_min + occupied_minutes,
                    "duration":    duration,
                    "grid_minutes": grid_minutes,
                    "resource_type": resource_type,
                    "source_row":  source_row,
                }
                managed_pending[(exclusive_group, day, resource_type)].append(pending)
                managed_by_gym_day[(exclusive_group, day)].append(pending)
                resolved.append(entry)
                continue

            # Direct/standalone resource: resolve to an existing expanded
            # resource whose slot grid contains this start and whose full
            # occupied interval is still free.
            slot_label = f"{day}-{_clock(start_min)}"
            chosen = None
            chosen_interval: Optional[Tuple[int, int, str]] = None
            for row, row_grid_min, row_duration, _occupied, open_min in sorted(
                window_rows,
                key=lambda item: (
                    len(str(item[0].get("resource_id") or "")),
                    str(item[0].get("resource_id") or ""),
                ),
            ):
                if (start_min - open_min) % row_grid_min != 0:
                    continue
                rid = str(row.get("resource_id") or "").strip()
                end_min = start_min + (
                    (row_duration + row_grid_min - 1) // row_grid_min
                ) * row_grid_min
                interval = (start_min, end_min)
                if any(
                    _overlaps(interval, (used_start, used_end))
                    for used_start, used_end, _used_game
                    in used_intervals_by_resource[rid]
                ):
                    continue
                chosen = rid
                chosen_interval = (start_min, end_min, str(game_id))
                break
            if chosen is None:
                _record_error(
                    f"Playoff slot {game_id!r}: every {resource_type!r} at "
                    f"{gym_name!r} is already occupied during "
                    f"{_clock(start_min)}-{_clock(start_min + occupied_minutes)} "
                    f"or has a slot grid that does not start at {start_text!r}."
                )
                continue
            entry["resource_id"] = chosen
            entry["slot"] = slot_label
            if chosen_interval is not None:
                used_intervals_by_resource[chosen].append(chosen_interval)
            resolved.append(entry)

        # Gym-Modes describes mutually-exclusive physical configurations. At
        # every instant a managed gym may host only one mode, and that mode may
        # not exceed its configured court/table count.
        for (exclusive_group, day), pins in managed_by_gym_day.items():
            boundaries = sorted({
                boundary
                for pin in pins
                for boundary in (pin["start_min"], pin["end_min"])
            })
            for segment_start, segment_end in zip(boundaries, boundaries[1:]):
                active = [
                    pin for pin in pins
                    if pin["start_min"] < segment_end
                    and segment_start < pin["end_min"]
                ]
                if not active:
                    continue
                active_types = {
                    str(pin["resource_type"]) for pin in active
                }
                game_ids = sorted(
                    str(pin["entry"].get("game_id") or "<unknown>")
                    for pin in active
                )
                if len(active_types) > 1:
                    _record_error(
                        f"Playoff-Slots pins "
                        f"{', '.join(repr(game_id) for game_id in game_ids)} "
                        f"overlap at mutually-exclusive gym {exclusive_group!r} on "
                        f"{day} {_clock(segment_start)}-{_clock(segment_end)} using "
                        f"different modes {sorted(active_types)}."
                    )
                    continue
                resource_type = next(iter(active_types))
                capacity = int(
                    gym_modes.get(exclusive_group, {}).get(resource_type, 0)
                )
                if len(active) > capacity:
                    _record_error(
                        f"Playoff-Slots pins "
                        f"{', '.join(repr(game_id) for game_id in game_ids)} "
                        f"need {len(active)} concurrent {resource_type!r} resources "
                        f"at {exclusive_group!r} on {day} "
                        f"{_clock(segment_start)}-{_clock(segment_end)}, but "
                        f"Gym-Modes provides {capacity}."
                    )

        if errors:
            details = "\n".join(f"  - {message}" for message in errors)
            raise ValueError(
                "Invalid venue-centric Playoff-Slots configuration:\n" + details
            )

        # Track pass for allocator-managed gyms: contiguous pins share one
        # synthetic court; concurrent pins get separate courts.
        synthetic_counters: Dict[Tuple[str, str], int] = {}
        for (exclusive_group, day, resource_type), pins in managed_pending.items():
            pins.sort(key=lambda p: p["start_min"])
            tracks: List[Dict[str, Any]] = []
            prefix = self._resource_id_prefix(resource_type)
            for pin in pins:
                track = next(
                    (
                        t for t in tracks
                        if t["close_min"] == pin["start_min"]
                        and t["slot_minutes"] == pin["grid_minutes"]
                    ),
                    None,
                )
                if track is None:
                    counter_key = (prefix, day)
                    n = synthetic_counters.get(counter_key, 0) + 1
                    synthetic_counters[counter_key] = n
                    label_kind = "Table" if "table" in resource_type.lower() else "Court"
                    track = {
                        "resource_id": f"{prefix}-{day}-PF{n}",
                        "label":       f"{label_kind}-PF{n}",
                        "open_min":    pin["start_min"],
                        "close_min":   pin["end_min"],
                        "slot_minutes": pin["grid_minutes"],
                        "source_row":  pin["source_row"],
                    }
                    tracks.append(track)
                else:
                    track["close_min"] = pin["end_min"]
                entry = pin["entry"]
                entry["resource_id"] = track["resource_id"]
                entry["slot"] = f"{day}-{_clock(pin['start_min'])}"

            for track in tracks:
                source_row = track["source_row"]
                synthetic = {
                    "resource_id":     track["resource_id"],
                    "resource_type":   resource_type,
                    "label":           track["label"],
                    "day":             day,
                    "open_time":       _clock(track["open_min"]),
                    "close_time":      _clock(track["close_min"]),
                    "slot_minutes":    track["slot_minutes"],
                    "venue_name":      source_row.get("venue_name", ""),
                    "exclusive_group": exclusive_group,
                    "playoff_pinned":  True,
                }
                synthetic_resources.append(synthetic)
                synthetic_ids.add(track["resource_id"])
                reserved_windows.append(GymBlock(
                    gym_name=exclusive_group,
                    day=day,
                    open_time=synthetic["open_time"],
                    close_time=synthetic["close_time"],
                    slot_minutes=track["slot_minutes"],
                    resource_types=frozenset({resource_type}),
                ))
                logger.info(
                    f"Playoff venue pin: reserved {exclusive_group!r} "
                    f"{day} {synthetic['open_time']}–{synthetic['close_time']} as "
                    f"{track['resource_id']!r} ({resource_type}) — excluded from "
                    "pool-play allocation."
                )

        return resolved, synthetic_resources, reserved_windows, synthetic_ids

    def _build_schedule_input(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
        venue_input_path: Path,
        pool_assignment_path: Optional[Path] = None,
    ) -> Dict[str, Any]:
        """Assemble the full schedule_input package consumed by OR-Tools.

        Returns a dict with keys: generated_at, gym_court_scenario, game_count,
        resource_count, games, resources, playoff_slots, gym_modes, gym_allocation.

        When venue_input.xlsx is present with a Gym-Modes tab, the Layer-2
        Stage-A greedy allocator runs and produces real gym resources keyed to
        the booked venue.  Otherwise falls back to the SCHEDULE_SOLVER_GYM_COURTS
        constant split evenly between basketball and volleyball when no explicit
        venue rows exist. If venue rows exist but the allocator cannot run, the
        explicit Venue-Input rows are used directly.
        """
        from gym_allocator import (
            aggregate_demand_by_mode, extract_gym_blocks, allocate,
        )
        gym_modes = self._load_gym_modes(venue_input_path)
        venue_rows, day_order = self._load_venue_input_rows(venue_input_path)
        playoff_slots = self._load_playoff_slots(venue_input_path)
        gym_blocks = extract_gym_blocks(venue_rows)
        explicit_gym_resource_types = {
            GYM_RESOURCE_TYPE_BASKETBALL,
            GYM_RESOURCE_TYPE_VOLLEYBALL,
        }
        has_explicit_gym_rows = any(
            resource.get("resource_type") in explicit_gym_resource_types
            for resource in venue_rows
        )
        gym_resource_strategy = (
            "allocator"
            if gym_blocks and gym_modes
            else "direct_venue_input"
            if has_explicit_gym_rows
            else "fallback"
        )

        pool_assignment_rows = self._build_pool_assignment_rows(
            roster_rows,
            pool_assignment_path,
        )
        gym_games, gym_precedence = self._build_assigned_gym_game_objects(
            roster_rows,
            pool_assignment_rows,
            allow_placeholder_fallback=(gym_resource_strategy == "fallback"),
        )
        bc_games, precedence = self._build_assigned_bc_game_objects(pool_assignment_rows)
        soccer_games, soccer_precedence = self._build_assigned_soccer_game_objects(
            roster_rows,
            pool_assignment_rows,
        )
        pod_games, pod_precedence = self._build_pod_game_objects(roster_rows, validation_rows)
        all_games = gym_games + bc_games + soccer_games + pod_games
        team_conflicts = self._build_gym_team_conflicts(roster_rows, pool_assignment_rows)
        team_conflicts += self._build_cross_sport_conflicts(
            roster_rows, pool_assignment_rows, validation_rows
        )
        _confirmed_pods, pod_unprotected_entries = self._resolve_pod_doubles(
            roster_rows, validation_rows
        )
        pod_validation_reconciliation = self._reconcile_pod_validation(
            pod_unprotected_entries, _confirmed_pods, validation_rows
        )
        precedence.extend(gym_precedence)
        precedence.extend(soccer_precedence)
        precedence.extend(pod_precedence)

        # Resolve venue-centric Playoff-Slots after game generation so physical
        # reservations use each game's real duration, but still before Stage-A
        # allocation so those windows are removed from pool-play inventory.
        date_day_map = self._load_venue_date_day_map(venue_input_path)
        game_duration_by_id = {
            str(game.get("game_id") or "").strip(): int(
                game.get("duration_minutes") or 0
            )
            for game in all_games
            if str(game.get("game_id") or "").strip()
            and int(game.get("duration_minutes") or 0) > 0
        }
        (
            playoff_slots,
            playoff_synthetic_resources,
            playoff_reserved_windows,
            playoff_synthetic_ids,
        ) = self._resolve_venue_playoff_slots(
            playoff_slots,
            venue_rows,
            date_day_map,
            gym_modes,
            allocator_active=(gym_resource_strategy == "allocator"),
            game_duration_by_id=game_duration_by_id,
        )

        gym_allocation: Optional[Dict[str, Any]] = None
        if gym_resource_strategy == "allocator":
            venue_capacity_rows = self._build_venue_capacity_rows(roster_rows)
            demand = aggregate_demand_by_mode(venue_capacity_rows)
            # Days that carry pinned playoff slots are excluded from the
            # spreading pass in allocate() — those blocks are handled by the
            # playoff-slot promotion path below and must not be pre-empted.
            _playoff_days: set = set()
            for _ps in playoff_slots:
                _slot = str(_ps.get("slot") or "").strip()
                if _slot:
                    _parts = _slot.rsplit("-", 1)
                    if len(_parts) == 2:
                        _playoff_days.add(_parts[0])
            alloc_result = allocate(demand, gym_modes, gym_blocks,
                                    spreading_excluded_days=_playoff_days,
                                    reserved_windows=playoff_reserved_windows)
            gym_resources = self._build_gym_resources_from_allocator(alloc_result.decisions)
            # Rows with no exclusive_group are standalone resources — include directly.
            # Rows whose exclusive_group has no Gym-Modes entry were not seen by the
            # allocator; include them directly too, and warn so the operator knows
            # mutual exclusivity is not enforced for those venues.
            covered_groups = set(gym_modes.keys())
            uncovered_groups = {
                r["exclusive_group"] for r in venue_rows
                if r.get("exclusive_group") and r["exclusive_group"] not in covered_groups
            }
            if uncovered_groups:
                logger.warning(
                    f"Exclusive venue group(s) {sorted(uncovered_groups)} appear in Venue-Input "
                    "but have no entry in the Gym-Modes tab. Their rows are included as direct "
                    "resources without mode-exclusivity enforcement. Add them to Gym-Modes if "
                    "the courts in those venues cannot be used simultaneously."
                )
            direct_resources = [
                r for r in venue_rows
                if not r.get("exclusive_group") or r["exclusive_group"] not in covered_groups
            ]
            all_resources = gym_resources + direct_resources
            gym_allocation = {
                "source":        "allocator",
                "decisions":     [
                    {
                        "gym_name":     d.gym_name,
                        "day":          d.day,
                        "open_time":    d.open_time,
                        "close_time":   d.close_time,
                        "mode":         d.mode,
                        "courts":       d.courts,
                        "slot_minutes": d.slot_minutes,
                    }
                    for d in alloc_result.decisions
                ],
                "mode_supply":    alloc_result.mode_supply,
                "mode_demand":    alloc_result.mode_demand,
                "mode_shortfall": alloc_result.mode_shortfall,
                "switch_count":   alloc_result.switch_count,
            }
            logger.info(
                f"Gym allocation (Stage A): {len(alloc_result.decisions)} blocks assigned, "
                f"{alloc_result.switch_count} mode switches"
            )
        elif gym_resource_strategy == "direct_venue_input":
            all_resources = venue_rows
            if gym_blocks and not gym_modes:
                reason = "grouped_rows_without_gym_modes"
                logger.warning(
                    "Gym allocation skipped: Venue-Input contains Exclusive Venue Group rows "
                    "but no Gym-Modes tab. Using Venue-Input rows directly; mutual exclusivity "
                    "is not enforced in this mode."
                )
            elif gym_modes and not gym_blocks:
                reason = "gym_modes_without_grouped_rows"
                logger.info(
                    "Gym allocation skipped: Gym-Modes tab is present but no Exclusive Venue "
                    "Group rows were found. Using Venue-Input rows directly."
                )
            else:
                reason = "explicit_venue_rows_without_allocator"
            gym_allocation = {"source": "direct_venue_input", "reason": reason}
        else:
            n_bb = SCHEDULE_SOLVER_GYM_COURTS // 2
            n_vb = SCHEDULE_SOLVER_GYM_COURTS - n_bb
            gym_resources = self._build_gym_resource_objects(n_bb, n_vb)
            all_resources = gym_resources + venue_rows
            gym_allocation = {"source": "fallback", "gym_court_scenario": SCHEDULE_SOLVER_GYM_COURTS}
            logger.info(
                f"Gym allocation: fallback mode — {n_bb} basketball + {n_vb} volleyball courts "
                f"per session (SCHEDULE_SOLVER_GYM_COURTS={SCHEDULE_SOLVER_GYM_COURTS})"
            )

        # Synthetic playoff-pinned resources from venue-centric Playoff-Slots
        # rows (Issue #127).  Added as a new list so the direct_venue_input
        # branch's venue_rows alias is never mutated.
        if playoff_synthetic_resources:
            all_resources = all_resources + playoff_synthetic_resources

        # Promote any playoff-pinned resource that the allocator didn't emit.
        # Rather than adding the whole multi-slot venue row (which would expose
        # unused slots to pool play), we synthesise a one-slot resource covering
        # only the exact time window referenced in the playoff entry.  The
        # playoff_pinned flag keeps the resource out of capacity diagnostics;
        # gym-sport synthetics still join the Gym Core solver pool below so
        # precedence rules involving the pinned game stay enforceable.
        #
        # Venue rows for gym sports use BB-*/VB-* resource_ids (per
        # RESOURCE_ID_PREFIX_BY_TYPE), while allocator-generated resources use
        # GYM-* ids — so we cannot look up by resource_id directly.  Instead we
        # derive day from the slot label and resource_type from the game event,
        # then find a representative venue row for slot_minutes / venue metadata.
        from gym_allocator import EVENT_TO_MODE as _EVENT_TO_MODE

        grouped_rows = [row for row in venue_rows if row.get("exclusive_group")]
        if grouped_rows and playoff_slots:
            block_mode_rows: Dict[Tuple[Tuple[str, str, str, str], str], List[Dict[str, Any]]] = defaultdict(list)
            block_capacity: Dict[Tuple[str, str, str, str], int] = {}
            day_blocks: Dict[str, List[Tuple[str, str, str, str]]] = defaultdict(list)

            for venue_row in grouped_rows:
                block_key = (
                    str(venue_row.get("day") or "").strip(),
                    str(venue_row.get("exclusive_group") or "").strip(),
                    str(venue_row.get("open_time") or "").strip(),
                    str(venue_row.get("close_time") or "").strip(),
                )
                resource_type = str(venue_row.get("resource_type") or "").strip()
                block_mode_rows[(block_key, resource_type)].append(venue_row)

            for (block_key, _resource_type), rows in block_mode_rows.items():
                block_capacity[block_key] = max(block_capacity.get(block_key, 0), len(rows))

            for block_key in block_capacity:
                day_blocks[block_key[0]].append(block_key)

            for day_label, blocks in day_blocks.items():
                blocks.sort(key=lambda item: (
                    self._parse_hour(item[2]),
                    item[1],
                    self._parse_hour(item[3]),
                ))

            block_ranges: Dict[Tuple[str, str, str, str], Tuple[int, int]] = {}
            for day_label, blocks in day_blocks.items():
                ordinal = 0
                for block_key in blocks:
                    start_ordinal = ordinal + 1
                    ordinal += block_capacity[block_key]
                    block_ranges[block_key] = (start_ordinal, ordinal)

            resources_by_id: Dict[str, Dict[str, Any]] = {
                str(resource.get("resource_id") or "").strip(): resource
                for resource in all_resources
            }

            for playoff_slot in playoff_slots:
                game_id = str(playoff_slot.get("game_id") or "").strip() or "<unknown>"
                resource_id = str(playoff_slot.get("resource_id") or "").strip()
                slot_label = str(playoff_slot.get("slot") or "").strip()
                event = str(playoff_slot.get("event") or "").strip()
                if not resource_id or not slot_label:
                    continue

                # Venue-centric rows were already resolved (and their windows
                # reserved) by _resolve_venue_playoff_slots — the legacy GYM-*
                # ordinal promotion below does not apply to them.
                if resource_id in playoff_synthetic_ids:
                    continue

                existing_resource = resources_by_id.get(resource_id)
                if existing_resource is not None and not existing_resource.get("playoff_pinned"):
                    continue

                day, time_part = self._split_slot_label(slot_label)
                if not day or not time_part:
                    logger.warning(
                        f"Playoff slot {game_id!r}: cannot parse day/time from slot {slot_label!r} — skipped."
                    )
                    continue

                resource_type = _EVENT_TO_MODE.get(event)
                if not resource_type:
                    logger.warning(
                        f"Playoff slot {game_id!r}: cannot infer resource_type from "
                        f"event {event!r} — skipped. Ensure event matches a gym sport name."
                    )
                    continue

                expected_prefix = f"GYM-{day}-"
                if not resource_id.startswith(expected_prefix):
                    logger.warning(
                        f"Playoff slot {game_id!r}: resource_id {resource_id!r} does not match "
                        f"slot day {day!r}; expected prefix {expected_prefix!r}. Skipped."
                    )
                    continue
                ordinal_text = resource_id[len(expected_prefix):]
                if not ordinal_text.isdigit():
                    logger.warning(
                        f"Playoff slot {game_id!r}: resource_id {resource_id!r} does not end "
                        "with a numeric court ordinal; skipped."
                    )
                    continue
                requested_ordinal = int(ordinal_text)
                slot_hour = self._parse_hour(time_part)

                matched_rows: List[Dict[str, Any]] = []
                matched_local_index = -1
                for block_key in day_blocks.get(day, []):
                    block_open = self._parse_hour(block_key[2])
                    block_close = self._parse_hour(block_key[3])
                    if not (block_open <= slot_hour < block_close):
                        continue

                    start_ordinal, end_ordinal = block_ranges[block_key]
                    if not (start_ordinal <= requested_ordinal <= end_ordinal):
                        continue

                    candidate_rows = sorted(
                        block_mode_rows.get((block_key, resource_type), []),
                        key=lambda row: (
                            str(row.get("label") or "").strip(),
                            str(row.get("resource_id") or "").strip(),
                        ),
                    )
                    if not candidate_rows:
                        continue

                    local_index = requested_ordinal - start_ordinal
                    if local_index >= len(candidate_rows):
                        continue

                    matched_rows = candidate_rows
                    matched_local_index = local_index
                    break

                if matched_local_index < 0:
                    logger.warning(
                        f"Playoff slot {game_id!r}: resource_id {resource_id!r} is not a plausible "
                        f"{resource_type} court for {day} at {time_part}. Check the requested "
                        "court ordinal or add a direct venue row instead."
                    )
                    continue

                source_row = matched_rows[matched_local_index]
                slot_minutes = int(source_row.get("slot_minutes") or 60)
                try:
                    hour, minute = (int(x) for x in time_part.split(":"))
                    close_total = hour * 60 + minute + slot_minutes
                    close_time = f"{close_total // 60:02d}:{close_total % 60:02d}"
                except (ValueError, AttributeError):
                    close_time = str(source_row.get("close_time") or time_part)

                if existing_resource is not None:
                    if (
                        str(existing_resource.get("day") or "").strip() != day
                        or str(existing_resource.get("resource_type") or "").strip() != resource_type
                    ):
                        logger.warning(
                            f"Playoff slot {game_id!r}: resource_id {resource_id!r} was already "
                            "promoted for a different day/resource_type; skipped."
                        )
                        continue
                    existing_resource["open_time"] = min(
                        str(existing_resource.get("open_time") or time_part),
                        time_part,
                        key=self._parse_hour,
                    )
                    existing_resource["close_time"] = max(
                        str(existing_resource.get("close_time") or close_time),
                        close_time,
                        key=self._parse_hour,
                    )
                    continue

                synthetic = {
                    "resource_id":     resource_id,
                    "resource_type":   resource_type,
                    "label":           source_row.get("label", "Court-1"),
                    "day":             day,
                    "open_time":       time_part,
                    "close_time":      close_time,
                    "slot_minutes":    slot_minutes,
                    "venue_name":      source_row.get("venue_name", ""),
                    "exclusive_group": source_row.get("exclusive_group", ""),
                    "playoff_pinned":  True,
                }
                all_resources.append(synthetic)
                resources_by_id[resource_id] = synthetic
                logger.info(
                    f"Promoted playoff-pinned resource {resource_id!r} ({resource_type}, "
                    f"{day} {time_part}) — single-slot, excluded from pool play."
                )

        if bc_games and not any(
            str(resource.get("resource_type") or "").strip() == TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE
            for resource in all_resources
        ):
            logger.warning(
                "Bible Challenge games were generated but no 'BC Station' resources were found "
                "in venue_input.xlsx. Those games will be unscheduled until a BC Station row is added."
            )

        if soccer_games and not any(
            str(resource.get("resource_type") or "").strip() == TEAM_RESOURCE_TYPE_SOCCER
            for resource in all_resources
        ):
            logger.warning(
                "Soccer games were generated but no 'Soccer Field' resources were found "
                "in venue_input.xlsx. Those games will be unscheduled until a Soccer Field row is added."
            )

        self._warn_if_resource_slot_minutes_differ_from_config(all_games, all_resources)

        # Playoff-pinned BB/VB resources join the Gym Core pool too: a pinned
        # Semi/Final game takes its pool from its pinned resource, and the
        # auto-generated QF→Semi→Final precedence rules can only be enforced
        # when the pinned game shares a pool with its pool-play siblings.
        # Pool play still cannot use a pinned resource — its window covers
        # only the pinned slots, all reserved by validate_playoff_slots.
        for resource in all_resources:
            if resource.get("resource_type") in (
                GYM_RESOURCE_TYPE_BASKETBALL,
                GYM_RESOURCE_TYPE_VOLLEYBALL,
            ):
                resource["solver_pool"] = self._GYM_CORE_SOLVER_POOL

        # Constrain QF games to end no later than the last slot on the day
        # BEFORE the Finals pinned day for that sport.  When the user pins a
        # Final to e.g. Sun-2-14:00, they almost always intend QFs to run the
        # day before (Sat-2) — that's why Sat-2 capacity exists in venue_input.
        # Without this constraint, CP-SAT can FEASIBLY (but undesirably) push
        # a QF onto the Finals day next to its Semi/Final.
        finals_day_by_event: Dict[str, str] = {}
        for ps in playoff_slots:
            if str(ps.get("stage") or "").strip().lower() != "final":
                continue
            slot = str(ps.get("slot") or "").strip()
            event_label = str(ps.get("event") or "").strip()
            if not slot or not event_label:
                continue
            day, _ = self._split_slot_label(slot)
            if day:
                finals_day_by_event[event_label] = day

        if finals_day_by_event:
            for game in all_games:
                if str(game.get("stage") or "").strip() not in ("QF", "Semi"):
                    continue
                event_label = str(game.get("event") or "").strip()
                finals_day = finals_day_by_event.get(event_label)
                if not finals_day:
                    continue
                try:
                    finals_idx = day_order.index(finals_day)
                except ValueError:
                    continue
                if finals_idx <= 0:
                    continue
                day_before = day_order[finals_idx - 1]
                court_type = str(game.get("resource_type") or "").strip()
                if not court_type:
                    continue
                latest = ScheduleWorkbookBuilder._last_slot_label_on_day(
                    all_resources, court_type, day_before,
                )
                if latest:
                    game["latest_slot"] = latest

        return {
            "generated_at":       datetime.now().isoformat(timespec="seconds"),
            "gym_court_scenario": SCHEDULE_SOLVER_GYM_COURTS,
            "game_count":         len(all_games),
            "resource_count":     len(all_resources),
            "games":              all_games,
            "resources":          all_resources,
            "playoff_slots":      playoff_slots,
            "gym_modes":          gym_modes,
            "gym_allocation":     gym_allocation,
            "team_conflicts":     team_conflicts,
            "pod_unprotected_entries": pod_unprotected_entries,
            "pod_validation_reconciliation": pod_validation_reconciliation,
            "precedence":         precedence,
            "day_order":          day_order,
        }






    @classmethod
    def _write_pool_assignment_tab(
        cls,
        ws,
        pool_assignment_rows: List[Dict[str, Any]],
    ) -> None:
        """Write the editable Pool-Assignment planning tab."""
        from openpyxl.styles import Alignment, Font

        columns = cls._POOL_ASSIGNMENT_COLUMNS
        ws.cell(row=1, column=1, value=cls._POOL_ASSIGNMENT_NOTE)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        note_cell = ws.cell(row=1, column=1)
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        note_cell.font = Font(italic=True)
        ws.row_dimensions[1].height = 30
        ws.append(columns)
        for row in pool_assignment_rows:
            values = []
            for column in columns:
                value = row.get(column)
                if column == "Seed" and value in (None, 0):
                    value = ""
                values.append(value)
            ws.append(values)

        cls._annotate_pool_assignment_tab(ws, len(columns))

    def refresh_pool_assignments(
        self,
        workbook_path: Path,
        output_path: Optional[Path] = None,
        sidecar_path: Optional[Path] = None,
    ) -> List[Dict[str, Any]]:
        """Refresh Pool-Assignment rows from an edited workbook and persist them."""
        from openpyxl import load_workbook

        workbook_path = Path(workbook_path)
        output_path = Path(output_path) if output_path else workbook_path
        effective_sidecar_path = (
            Path(sidecar_path)
            if sidecar_path is not None
            else self._pool_assignments_sidecar_path(output_path.parent)
        )

        sheet_rows = self._read_pool_assignment_rows(workbook_path)
        if not sheet_rows:
            raise ValueError(
                f"Workbook '{workbook_path}' has no usable Pool-Assignment sheet."
            )

        normalized_rows: List[Dict[str, Any]] = []
        for row in sheet_rows:
            event_name = str(row.get("Event") or "").strip()
            team_id = str(row.get("Team ID") or "").strip()
            if not event_name or not team_id:
                continue
            normalized_rows.append({
                "Event": event_name,
                "Church Team": str(row.get("Church Team") or "").strip(),
                "Team Order": str(row.get("Team Order") or "").strip(),
                "Team ID": team_id,
                "Team Label": str(row.get("Team Label") or team_id).strip(),
                "Team Source": str(row.get("Team Source") or "").strip(),
                "Roster Count": self._positive_int_or_none(row.get("Roster Count")) or 0,
                "Min Team Size": self._positive_int_or_none(row.get("Min Team Size")) or 0,
                "Seed": self._normalize_pool_seed(row.get("Seed")),
                "Random Draw Order": self._positive_int_or_none(row.get("Random Draw Order")),
                "Draw Position": self._positive_int_or_none(row.get("Draw Position")),
                "Pool ID": str(row.get("Pool ID") or "").strip(),
                "Pool Slot": str(row.get("Pool Slot") or "").strip(),
                "Assignment Basis": str(row.get("Assignment Basis") or "").strip(),
                "Notes": str(row.get("Notes") or "").strip(),
            })

        refreshed_rows = self._apply_pool_assignments_to_rows(normalized_rows)
        self._write_pool_assignment_state(effective_sidecar_path, refreshed_rows)

        wb = load_workbook(workbook_path)
        if "Pool-Assignment" not in wb.sheetnames:
            raise ValueError(
                f"Workbook '{workbook_path}' does not contain a Pool-Assignment sheet."
            )

        sheet_index = wb.sheetnames.index("Pool-Assignment")
        wb.remove(wb["Pool-Assignment"])
        ws = wb.create_sheet(title="Pool-Assignment", index=sheet_index)
        self._write_pool_assignment_tab(ws, refreshed_rows)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(
            f"Pool-Assignment tab refreshed: {len(refreshed_rows)} rows -> {output_path}"
        )
        logger.info(f"Pool-assignment sidecar written to: {effective_sidecar_path}")
        return refreshed_rows



    # ── Pool planning ─────────────────────────────────────────────────────────


    @staticmethod
    def _two_game_pool_sizes(n_teams: int) -> List[int]:
        """Return deterministic pool sizes for the normalized 2-game/team policy."""
        if n_teams < 2:
            return []
        if n_teams in (2, 3, 4, 5):
            return [n_teams]

        for n_fours in range(n_teams // 4, -1, -1):
            remainder = n_teams - (4 * n_fours)
            if remainder >= 0 and remainder % 3 == 0:
                return ([4] * n_fours) + ([3] * (remainder // 3))

        raise ValueError(f"Unable to build normalized 2-game pools for n_teams={n_teams}")

    @staticmethod
    def _three_game_even_pool_sizes(n_teams: int) -> List[int]:
        """Return exact 3-game/team pool sizes for even team counts."""
        if n_teams < 4 or n_teams % 2:
            raise ValueError(
                f"Exact normalized 3-game pools require an even team count >= 4; got {n_teams}"
            )

        for n_sixes in range(n_teams // 6, -1, -1):
            remainder = n_teams - (6 * n_sixes)
            if remainder >= 0 and remainder % 4 == 0:
                return ([6] * n_sixes) + ([4] * (remainder // 4))

        raise ValueError(f"Unable to build normalized 3-game pools for n_teams={n_teams}")

    @staticmethod
    def _three_game_pool_sizes(n_teams: int) -> List[int]:
        """Return deterministic pool sizes for the supported 3-game/team policy."""
        if n_teams < 2:
            return []
        if n_teams in (2, 3, 4, 5, 7):
            return [n_teams]
        if n_teams % 2 == 0:
            return ScheduleWorkbookBuilder._three_game_even_pool_sizes(n_teams)
        if n_teams >= 9:
            return ScheduleWorkbookBuilder._three_game_even_pool_sizes(n_teams - 5) + [5]

        raise ValueError(f"Unable to build normalized 3-game pools for n_teams={n_teams}")

    @staticmethod
    def _format_pool_games_per_team(value: float) -> Any:
        """Return ints when whole, otherwise a rounded operator-facing float."""
        rounded = round(float(value), 2)
        return int(rounded) if float(rounded).is_integer() else rounded

    @staticmethod
    def _format_pool_composition(pool_sizes: List[int], gpg: int) -> str:
        """Return the operator-facing pool composition summary string."""
        if not pool_sizes:
            return ""

        base = " + ".join(str(size) for size in pool_sizes)
        notes: List[str] = []

        if gpg == 2:
            if 5 in pool_sizes:
                notes.append("5-team pools include one bye per round")
        elif gpg == 3:
            if any(size in (2, 3) for size in pool_sizes):
                notes.append("small pools cannot reach the full 3-game target")
            if 5 in pool_sizes:
                notes.append("5-team pools give T5 the extra 4th game")
            if 7 in pool_sizes:
                notes.append("7-team pools give T7 the extra 4th game")

        if not notes:
            return base
        return f"{base} ({'; '.join(notes)})"

    @staticmethod
    def _summarize_pool_policy(n_teams: int, gpg: int) -> Dict[str, Any]:
        """Return operator-facing metadata for the current pool-generation policy."""
        if n_teams < 2:
            return {
                "target_pool_games_per_team": gpg,
                "actual_pool_games_per_team": 0,
                "pool_composition": "",
                "bye_slots": 0,
                "actual_pool_games": 0,
            }

        if gpg == 2:
            pool_sizes = ScheduleWorkbookBuilder._two_game_pool_sizes(n_teams)
            games_by_pool_size = {2: 1, 3: 3, 4: 4, 5: 5}
            actual_pool_games = sum(games_by_pool_size[size] for size in pool_sizes)
            actual_pool_games_per_team = 1 if n_teams == 2 else 2
            bye_slots = 5 * pool_sizes.count(5)
        elif gpg == 3:
            pool_sizes = ScheduleWorkbookBuilder._three_game_pool_sizes(n_teams)
            games_by_pool_size = {2: 1, 3: 3, 4: 6, 5: 8, 6: 9, 7: 11}
            actual_pool_games = sum(games_by_pool_size[size] for size in pool_sizes)
            actual_pool_games_per_team = ScheduleWorkbookBuilder._format_pool_games_per_team(
                (2 * actual_pool_games) / n_teams
            )
            bye_slots = 0
        else:
            raise ValueError(
                f"Unsupported team-sport pool target {gpg}. Only 2 and 3 games/team are supported."
            )

        return {
            "target_pool_games_per_team": gpg,
            "actual_pool_games_per_team": actual_pool_games_per_team,
            "pool_composition": ScheduleWorkbookBuilder._format_pool_composition(pool_sizes, gpg),
            "bye_slots": bye_slots,
            "actual_pool_games": actual_pool_games,
        }

    @staticmethod
    def _make_legacy_pool_game_pairs(
        prefix: str, n_teams: int, gpg: int
    ) -> List[Tuple[str, str, str]]:
        """Legacy balanced round-robin fallback for non-default pool-game targets."""
        if n_teams < 2:
            return []

        target_pool_size = max(2, gpg + 1)
        n_pools = max(1, n_teams // target_pool_size)

        pools: List[List[int]] = [[] for _ in range(n_pools)]
        for i in range(n_teams):
            pools[i % n_pools].append(i + 1)

        team_id: Dict[int, str] = {}
        for p_idx, pool_teams in enumerate(pools, start=1):
            for t_idx, team_num in enumerate(pool_teams, start=1):
                team_id[team_num] = f"{prefix}-P{p_idx}-T{t_idx}"

        pairs: List[Tuple[str, str, str]] = []
        for p_idx, pool_teams in enumerate(pools, start=1):
            pool_id = f"P{p_idx}"
            for i in range(len(pool_teams)):
                for j in range(i + 1, len(pool_teams)):
                    pairs.append((
                        team_id[pool_teams[i]],
                        team_id[pool_teams[j]],
                        pool_id,
                    ))
        return pairs

    @staticmethod
    def _make_pool_game_pairs(
        prefix: str, n_teams: int, gpg: int
    ) -> List[Tuple[str, str, str]]:
        """Return (team_a_id, team_b_id, pool_id) tuples for pool-play games.

        Supported team-sport planning policies are:

        2-game mode:
        - 2 teams  -> one direct match
        - 3 teams  -> 3-team round robin
        - 4 teams  -> 4-match matrix (every team plays exactly twice)
        - 5 teams  -> 5-match cycle (every team plays exactly twice)
        - 6+ teams -> deterministic composition of 3-team and 4-team pools

        3-game mode:
        - 4 teams  -> 4-team round robin (every team plays exactly three)
        - 6 teams  -> 6-team 3-round matrix (every team plays exactly three)
        - odd team counts may require one 5-team or 7-team pool where the
          highest slot (T5 or T7) receives the extra 4th game

        Any target other than 2 or 3 is rejected loudly.

        Team IDs are stable planning placeholders: {prefix}-P{pool}-T{slot}.
        The same placeholder is reused across all games involving that team,
        allowing the solver to enforce team-overlap and min-rest constraints.
        """
        if n_teams < 2:
            return []

        if gpg == 2:
            template_pairs = {
                2: [(0, 1)],
                3: [(0, 1), (0, 2), (1, 2)],
                4: [(0, 1), (2, 3), (0, 2), (1, 3)],
                5: [(0, 1), (1, 2), (2, 3), (3, 4), (4, 0)],
            }
            pool_sizes = ScheduleWorkbookBuilder._two_game_pool_sizes(n_teams)
        elif gpg == 3:
            template_pairs = {
                2: [(0, 1)],
                3: [(0, 1), (0, 2), (1, 2)],
                4: [(0, 1), (0, 2), (0, 3), (1, 2), (1, 3), (2, 3)],
                5: [
                    (0, 1), (1, 2), (2, 3), (3, 4), (4, 0),
                    (4, 1), (4, 2), (0, 3),
                ],
                6: [
                    (0, 5), (1, 4), (2, 3),
                    (0, 4), (5, 3), (1, 2),
                    (0, 3), (4, 2), (5, 1),
                ],
                7: [
                    (0, 1), (1, 2), (2, 3), (3, 4), (4, 5), (5, 6), (6, 0),
                    (6, 1), (6, 3), (0, 4), (2, 5),
                ],
            }
            pool_sizes = ScheduleWorkbookBuilder._three_game_pool_sizes(n_teams)
        else:
            raise ValueError(
                f"Unsupported team-sport pool target {gpg}. Only 2 and 3 games/team are supported."
            )

        pairs: List[Tuple[str, str, str]] = []
        for p_idx, pool_size in enumerate(pool_sizes, start=1):
            pool_id = f"P{p_idx}"
            pool_team_ids = [
                f"{prefix}-P{p_idx}-T{slot_idx}"
                for slot_idx in range(1, pool_size + 1)
            ]
            for team_a_idx, team_b_idx in template_pairs[pool_size]:
                pairs.append((
                    pool_team_ids[team_a_idx],
                    pool_team_ids[team_b_idx],
                    pool_id,
                ))
        return pairs

    @staticmethod
    def _make_playoff_ids(
        prefix: str, playoff_teams: int, include_third: bool
    ) -> Tuple[List[str], List[str]]:
        """Return (early_ids, final_ids) split by which weekend they belong to.

        early_ids  — QF + Semi games, scheduled on 2nd Saturday.
        final_ids  — Final (+ optional 3rd-place), scheduled on 2nd Sunday.

        Bracket size is determined by playoff_teams (from COURT_ESTIMATE_PLAYOFF_RULES):
            0 teams  → no playoff games
            4 teams  → Semi-1, Semi-2 | Final [+ 3rd]
            8 teams  → QF-1…4, Semi-1, Semi-2 | Final [+ 3rd]

        To add a new bracket size (e.g. 16 teams with quarter-finals already
        called Round-of-16), extend the if/elif chain here and add matching
        rows to COURT_ESTIMATE_PLAYOFF_RULES in config.py.
        """
        early_ids: List[str] = []
        if playoff_teams >= 8:
            for i in range(1, 5):
                early_ids.append(f"{prefix}-QF-{i}")
            early_ids.extend([f"{prefix}-Semi-1", f"{prefix}-Semi-2"])
        elif playoff_teams >= 4:
            early_ids.extend([f"{prefix}-Semi-1", f"{prefix}-Semi-2"])

        final_ids: List[str] = []
        if playoff_teams >= 4:
            final_ids.append(f"{prefix}-Final")
            if include_third:
                final_ids.append(f"{prefix}-3rd")
        return early_ids, final_ids

    # ── Court-schedule sketch ────────────────────────────────────────────────


    # ── Pod-Resource-Estimate helpers ────────────────────────────────────────

    @staticmethod
    def _load_available_slots_from_schedule_input(
        schedule_input: Dict[str, Any],
    ) -> Dict[str, int]:
        """Summarize total available slots per resource_type from schedule_input."""
        totals: Dict[str, int] = {}
        for res in schedule_input.get("resources", []):
            resource_type = ScheduleWorkbookBuilder._clean_excel_text(
                res.get("resource_type")
            )
            if not resource_type:
                continue
            open_time = ScheduleWorkbookBuilder._clean_excel_text(res.get("open_time"))
            close_time = ScheduleWorkbookBuilder._clean_excel_text(res.get("close_time"))
            slot_min = int(res.get("slot_minutes", 0) or 0)
            if not open_time or not close_time or slot_min <= 0:
                continue
            start = ScheduleWorkbookBuilder._parse_hour(open_time)
            close = ScheduleWorkbookBuilder._parse_hour(close_time)
            if close < start:
                continue
            available = int(((close - start) * 60 / slot_min))
            totals[resource_type] = totals.get(resource_type, 0) + max(available, 0)
        logger.debug(f"Derived availability from schedule_input resources: {totals}")
        return totals



    # ── produce-schedule renderer ────────────────────────────────────────────

    # ── ALL-workbook readers (build-schedule-workbook input) ─────────────────

    @staticmethod
    def _read_pool_assignment_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
        """Read Pool-Assignment while tolerating the operator note row above headers."""
        try:
            df = pd.read_excel(
                xlsx_path,
                sheet_name="Pool-Assignment",
                header=None,
                engine="openpyxl",
            )
        except Exception as e:
            logger.warning(f"Could not read 'Pool-Assignment' tab from {xlsx_path}: {e}")
            return []

        df = df.astype(object).where(pd.notna(df), None)
        header_idx: Optional[int] = None
        for idx in range(min(len(df.index), 3)):
            first_cell = str(df.iat[idx, 0] or "").strip()
            if first_cell == "Event":
                header_idx = idx
                break
        if header_idx is None:
            logger.warning(
                f"Could not find Pool-Assignment header row in {xlsx_path}; expected a row starting with 'Event'."
            )
            return []

        header_values = [
            str(value).strip() if value is not None else ""
            for value in df.iloc[header_idx].tolist()
        ]
        data_df = df.iloc[header_idx + 1 :].copy()
        data_df.columns = header_values
        data_df = data_df.loc[:, [column for column in header_values if column]]
        rows = data_df.to_dict("records")
        logger.debug(
            f"Read {len(rows)} rows from 'Pool-Assignment' tab of {xlsx_path} "
            f"(header row {header_idx + 1})"
        )
        return rows

    @staticmethod
    def read_roster_validation_rows(
        xlsx_path: Optional[Path],
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """Parse the Roster and Validation-Issues tabs of an exported ALL workbook.

        Returns (roster_rows, validation_rows) in the list-of-dicts shape the
        scheduling builders expect.  A missing path or missing tabs degrade
        gracefully to empty lists with a WARNING — build-schedule-workbook still
        produces the Schedule-Input tab (echo of the JSON) without roster data.
        """
        if not xlsx_path or not Path(xlsx_path).exists():
            logger.warning(
                f"ALL workbook not found at {xlsx_path!r}; "
                "scheduling tabs that need roster data will be empty."
            )
            return [], []
        xlsx_path = Path(xlsx_path)
        roster_rows = ScheduleWorkbookBuilder._read_xlsx_sheet_rows(xlsx_path, "Roster")
        validation_rows = ScheduleWorkbookBuilder._read_xlsx_sheet_rows(
            xlsx_path, "Validation-Issues"
        )
        return roster_rows, validation_rows

    # ── Public entry points ──────────────────────────────────────────────────

    def write_schedule_input_json(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
        venue_input_path: Path,
        json_path: Path,
        pool_assignment_path: Optional[Path] = None,
    ) -> Dict[str, Any]:
        """Build schedule_input dict and write it as JSON. Returns the dict.
        Always called by export-church-teams, regardless of whether venue_input.xlsx
        exists (graceful degradation is handled inside _build_schedule_input).
        """
        schedule_input = self._build_schedule_input(
            roster_rows,
            validation_rows,
            venue_input_path,
            pool_assignment_path=pool_assignment_path,
        )
        json_path.write_text(json.dumps(schedule_input, indent=2, default=str), encoding="utf-8")
        logger.info(
            f"Schedule-Input: {schedule_input['game_count']} games, "
            f"{schedule_input['resource_count']} resources -> {json_path}"
        )
        return schedule_input

    def write_schedule_workbook(
        self,
        output_path: Path,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
        schedule_input: Dict[str, Any],
        venue_input_path: Optional[Path],
        pool_assignment_path: Optional[Path] = None,
    ) -> None:
        """Write the Schedule_Workbook xlsx with all scheduling tabs.
        Called by build-schedule-workbook command (Step 3).
        For the solver-rendered two-tab workbook, use write_schedule_output_workbook().
        When venue_input_path is None, derive resource availability from the
        schedule_input resources so offline builds stay self-consistent.
        """
        # Build workbook with pandas ExcelWriter for the DataFrame-based tabs,
        # then attach the openpyxl-native tabs using writer.book.
        venue_rows = self._build_venue_capacity_rows(roster_rows)
        venue_cols = [
            "Event", "Potential Teams/Entries", "Estimating Teams/Entries", "Teams",
            "Target Pool Games/Team", "Actual Pool Games/Team",
            "Pool Composition", "BYE Slots", "Minutes Per Game", "Pool Slots",
            "Playoff Teams", "Playoff Slots", "Third Place?",
            "Third Place Slots", "Total Court Slots", "Estimated Court Hours",
        ]

        pod_div_rows = self._build_pod_divisions_rows(roster_rows, validation_rows)
        pod_div_cols = [
            "division_id", "sport_type", "sport_gender", "sport_format",
            "resource_type", "minutes_per_game",
            "planning_entries", "confirmed_entries", "provisional_entries",
            "anomaly_count", "division_status", "notes",
        ]

        pod_entry_rows = self._build_pod_entries_review_rows(roster_rows, validation_rows)
        pod_entry_cols = [
            "entry_id", "division_id", "entry_type",
            "participant_1_name", "participant_2_name",
            "source_participant_ids", "church_team",
            "partner_status", "review_status", "notes",
        ]
        pool_assignment_rows = self._build_pool_assignment_rows(
            roster_rows,
            pool_assignment_path,
        )

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Venue-Estimator tab (pandas)
            df_venue = pd.DataFrame(venue_rows, columns=venue_cols)
            df_venue.to_excel(writer, sheet_name="Venue-Estimator", index=False, startrow=0)
            snapshot_note = (
                f"Roster snapshot as of {datetime.now().strftime('%Y-%m-%d')} — "
                "Estimating = complete entries; Potential = rule-aware ceiling from current "
                "registrations (including incomplete doubles pairings), capped by 2026 entry limits. "
                "Approval-agnostic. Updates with each export run."
            )
            venue_ws = writer.sheets["Venue-Estimator"]
            self._annotate_venue_estimator_tab(venue_ws, len(venue_cols))
            note_row = len(df_venue) + 3
            venue_ws.cell(row=note_row, column=1, value=snapshot_note)
            logger.debug(f"Venue-Estimator tab: {len(df_venue)} rows.")

            # Pool-Assignment tab (openpyxl native - editable seed/draw workspace)
            pool_ws = writer.book.create_sheet(title="Pool-Assignment", index=1)
            self._write_pool_assignment_tab(pool_ws, pool_assignment_rows)
            logger.debug(f"Pool-Assignment tab: {len(pool_assignment_rows)} rows.")

            # Pod-Divisions tab (pandas)
            df_pod_div = pd.DataFrame(pod_div_rows, columns=pod_div_cols)
            df_pod_div.to_excel(writer, sheet_name="Pod-Divisions", index=False)
            self._annotate_pod_divisions_tab(writer.sheets["Pod-Divisions"], len(pod_div_cols))
            logger.debug(f"Pod-Divisions tab: {len(df_pod_div)} rows.")

            # Pod-Entries-Review tab (pandas)
            df_pod_entries = pd.DataFrame(pod_entry_rows, columns=pod_entry_cols)
            df_pod_entries.to_excel(writer, sheet_name="Pod-Entries-Review", index=False)
            self._annotate_pod_entries_review_tab(
                writer.sheets["Pod-Entries-Review"], len(pod_entry_cols)
            )
            logger.debug(f"Pod-Entries-Review tab: {len(df_pod_entries)} rows.")

            # Court-Schedule-Sketch tab (openpyxl native)
            sketch_ws = writer.book.create_sheet(title="Court-Schedule-Sketch")
            self._write_court_schedule_sketch(sketch_ws, roster_rows)

            # Pod-Resource-Estimate tab (openpyxl native)
            if venue_input_path is None:
                available_by_resource = self._load_available_slots_from_schedule_input(
                    schedule_input
                )
                availability_source_label = "schedule_input.json resources"
            else:
                available_by_resource = self._load_venue_input(venue_input_path)
                availability_source_label = VENUE_INPUT_FILENAME
            pod_res_rows = self._build_pod_resource_rows(roster_rows, available_by_resource)
            pod_ws = writer.book.create_sheet(title="Pod-Resource-Estimate")
            self._write_pod_resource_estimate(
                pod_ws,
                pod_res_rows,
                available_by_resource,
                availability_source_label=availability_source_label,
            )

            # Schedule-Input tab (openpyxl native — echo of the JSON)
            si_ws = writer.book.create_sheet(title="Schedule-Input")
            self._write_schedule_input_tab(si_ws, schedule_input)

            # Gym-Allocation tab (openpyxl native — Stage-A allocator summary)
            gym_alloc_ws = writer.book.create_sheet(title="Gym-Allocation")
            self._write_gym_allocation_tab(gym_alloc_ws, schedule_input.get("gym_allocation"))

            # Summary tab (openpyxl native — operator guide / command cheat sheet)
            summary_ws = writer.book.create_sheet(title="Summary", index=0)
            self._write_summary_tab(summary_ws)
            self._stamp_known_tab_statuses(writer.book)

        logger.info(f"Schedule workbook written to: {output_path}")

    @staticmethod
    def write_schedule_output_workbook(
        output_path: Path,
        schedule_output: Dict[str, Any],
        schedule_input: Dict[str, Any],
    ) -> None:
        """Write the standalone Schedule-by-Time / Schedule-by-Sport workbook."""
        ScheduleWorkbookBuilder._write_schedule_output_report(
            Path(output_path), schedule_output, schedule_input
        )

    # ── Backward-compat class aliases — extracted to scheduling/ (Issue #152) ──
    _clean_excel_text = staticmethod(xlsx_utils._clean_excel_text)
    _float_from_excel = staticmethod(xlsx_utils._float_from_excel)
    _normalize_resource_type_name = staticmethod(xlsx_utils._normalize_resource_type_name)
    _resource_id_prefix = staticmethod(xlsx_utils._resource_id_prefix)
    _ordinal = staticmethod(xlsx_utils._ordinal)
    _day_sort_key = staticmethod(xlsx_utils._day_sort_key)
    _day_display_label = staticmethod(xlsx_utils._day_display_label)
    _coerce_excel_date = staticmethod(xlsx_utils._coerce_excel_date)
    _derive_day_labels_from_dates = staticmethod(xlsx_utils._derive_day_labels_from_dates)
    _set_excel_comment = staticmethod(xlsx_utils._set_excel_comment)
    _make_excel_note_shapes_visible = staticmethod(xlsx_utils._make_excel_note_shapes_visible)
    _stamp_tab_status_banner = staticmethod(xlsx_utils._stamp_tab_status_banner)
    _stamp_known_tab_statuses = staticmethod(xlsx_utils._stamp_known_tab_statuses)
    _annotate_header_row = staticmethod(xlsx_utils._annotate_header_row)
    _parse_hour = staticmethod(xlsx_utils._parse_hour)
    _read_xlsx_sheet_rows = staticmethod(xlsx_utils._read_xlsx_sheet_rows)
    _load_venue_input_rows = staticmethod(venue_loader._load_venue_input_rows)
    _load_playoff_slots = staticmethod(venue_loader._load_playoff_slots)
    _load_venue_date_day_map = staticmethod(venue_loader._load_venue_date_day_map)
    _split_slot_label = staticmethod(venue_loader._split_slot_label)
    _last_slot_label_on_day = staticmethod(venue_loader._last_slot_label_on_day)
    _load_gym_modes = staticmethod(venue_loader._load_gym_modes)
    _load_venue_input = staticmethod(venue_loader._load_venue_input)
    _warn_if_schedules_mismatched = staticmethod(output_report._warn_if_schedules_mismatched)
    _build_schedule_output_flat_rows = staticmethod(output_report._build_schedule_output_flat_rows)
    _write_schedule_diagnostics_tab = staticmethod(output_report._write_schedule_diagnostics_tab)
    _write_schedule_output_report = staticmethod(output_report._write_schedule_output_report)

    # ── Step 4: planning-tab renderers extracted to scheduling/planning_tabs.py (Issue #152) ──
    _write_summary_tab = staticmethod(planning_tabs._write_summary_tab)
    _build_scenario_schedule = staticmethod(planning_tabs._build_scenario_schedule)

    @classmethod
    def _annotate_venue_estimator_tab(cls, ws, n_cols: int) -> None:
        return planning_tabs._annotate_venue_estimator_tab(cls, ws, n_cols)

    @classmethod
    def _annotate_pod_divisions_tab(cls, ws, n_cols: int) -> None:
        return planning_tabs._annotate_pod_divisions_tab(cls, ws, n_cols)

    @classmethod
    def _annotate_pod_entries_review_tab(cls, ws, n_cols: int) -> None:
        return planning_tabs._annotate_pod_entries_review_tab(cls, ws, n_cols)

    @classmethod
    def _annotate_pool_assignment_tab(cls, ws, n_cols: int) -> None:
        return planning_tabs._annotate_pool_assignment_tab(cls, ws, n_cols)

    @classmethod
    def _write_schedule_input_tab(cls, ws, schedule_input: Dict[str, Any]) -> None:
        return planning_tabs._write_schedule_input_tab(cls, ws, schedule_input)

    @classmethod
    def _write_gym_allocation_tab(cls, ws, gym_allocation: Optional[Dict[str, Any]]) -> None:
        return planning_tabs._write_gym_allocation_tab(cls, ws, gym_allocation)

    def _write_court_schedule_sketch(self, ws, roster_rows: List[Dict[str, Any]]) -> None:
        return planning_tabs._write_court_schedule_sketch(self, ws, roster_rows)

    def _build_pod_resource_rows(self, roster_rows: List[Dict[str, Any]], available_by_resource: Dict[str, int]) -> List[Dict[str, Any]]:
        return planning_tabs._build_pod_resource_rows(self, roster_rows, available_by_resource)

    @classmethod
    def _write_pod_resource_estimate(self, ws, pod_rows: List[Dict[str, Any]], available_by_resource: Dict[str, int], availability_source_label: str=VENUE_INPUT_FILENAME) -> None:
        return planning_tabs._write_pod_resource_estimate(self, ws, pod_rows, available_by_resource, availability_source_label)
