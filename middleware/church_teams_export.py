# church_teams_export.py
# Version 1.3.0
import json
import pandas as pd
from pathlib import Path
from loguru import logger
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime
from collections import deque
import re
import unicodedata

from config import (
    Config,
    DATA_DIR,
    CHECK_BOXES,
    CHM_FIELDS,
    MEMBERSHIP_QUESTION,
    FORMAT_MAPPINGS,
    RULE_LEVEL,
    VALIDATION_SEVERITY,
    ATHLETE_FEE_STANDARD,
    ATHLETE_FEE_OTHER_EVENTS_ONLY,
    ATHLETE_FEE_LATE,
    REGISTRATION_DEADLINE,
    SPORT_TYPE,
    RACQUET_SPORTS,
    is_racquet_sport,
    COURT_ESTIMATE_EVENTS,
    COURT_ESTIMATE_RACQUET_EVENTS,
    COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
    COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME,
    COURT_ESTIMATE_MIN_TEAM_SIZE,
    COURT_ESTIMATE_MINUTES_PER_GAME,
    COURT_ESTIMATE_PLAYOFF_RULES,
    POD_SPORT_ABBREV,
    SCHEDULE_SKETCH_SATURDAY_START,
    SCHEDULE_SKETCH_SATURDAY_LAST_GAME,
    SCHEDULE_SKETCH_SUNDAY_START,
    SCHEDULE_SKETCH_SUNDAY_LAST_GAME,
    SCHEDULE_SKETCH_N_COURTS,
    SCHEDULE_SKETCH_COLOR_BASKETBALL,
    SCHEDULE_SKETCH_COLOR_VB_MEN,
    SCHEDULE_SKETCH_COLOR_VB_WOMEN,
    SCHEDULE_SKETCH_COLOR_SECTION,
    SCHEDULE_SKETCH_COLOR_HEADER,
    GYM_RESOURCE_TYPE,
    SCHEDULE_SOLVER_GYM_COURTS,
    VENUE_INPUT_FILENAME,
    POD_RESOURCE_EVENT_TYPE,
    POD_FIT_COLOR_GREEN,
    POD_FIT_COLOR_YELLOW,
    POD_FIT_COLOR_RED,
    POD_FIT_YELLOW_MAX,
)
from validation.name_matcher import normalized_name as _norm_name
from chmeetings.backend_connector import ChMeetingsConnector
from wordpress.frontend_connector import WordPressConnector
from validation.models import RulesManager
from math import ceil

class ChurchTeamsExporter: # MODIFIED CLASS NAME
    """
    Generates Excel reports for church team statuses, combining data from
    ChMeetings and WordPress.
    """

    def __init__(self):
        logger.info("Initializing ChurchTeamsExporter...") # MODIFIED LOGGER MESSAGE
        self.chm_connector = ChMeetingsConnector()
        self.wp_connector = WordPressConnector()
        self.sports_fest_date = datetime.strptime(Config.SPORTS_FEST_DATE, "%Y-%m-%d").date()
        self.latest_chm_update_by_church: Dict[str, str] = {} # Initialize here
        self.last_orphaned_memberships_by_church: Dict[str, int] = {}
        logger.info("ChurchTeamsExporter initialized.") # MODIFIED LOGGER MESSAGE

    def _calculate_age(self, birthdate_str: Optional[str]) -> Optional[int]:
        """Calculates age as of the Sports Fest date."""
        if not birthdate_str:
            return None
        try:
            birth_date_obj = datetime.strptime(birthdate_str, "%Y-%m-%d").date()
            age = self.sports_fest_date.year - birth_date_obj.year - \
                  ((self.sports_fest_date.month, self.sports_fest_date.day) < (birth_date_obj.month, birth_date_obj.day))
            return age
        except ValueError:
            logger.warning(f"Invalid birthdate format encountered: '{birthdate_str}' for age calculation.")
            return None

    def _parse_church_code_from_group_name(self, group_name: str) -> Optional[str]:
        """Extracts church code from a ChMeetings group name like 'Team ABC'."""
        if group_name.startswith(Config.TEAM_PREFIX + " "):
            return group_name[len(Config.TEAM_PREFIX + " "):].strip().upper()
        logger.warning(f"Could not parse church code from group name: '{group_name}'")
        return None

    def _get_completion_checklist_statuses(self, checklist_str: Optional[str]) -> Dict[str, str]:
        """Parses the 'Completion Check List' string and returns status for each box."""
        statuses = {f"Box {i+1}": "" for i in range(len(CHECK_BOXES))}
        if not checklist_str:
            return statuses

        for i, (key, text_value) in enumerate(CHECK_BOXES.items()):
            # Ensure we're checking for the full text of the checklist item
            if text_value in checklist_str:
                statuses[f"Box {i+1}"] = text_value # Store the constant text
        return statuses

    def _fetch_open_validation_issues(self, church_id: Optional[int]) -> List[Dict[str, Any]]:
        """Fetch all open validation issues for one church from WordPress."""
        if not church_id:
            return []

        issues: List[Dict[str, Any]] = []
        current_page = 1
        fetch_per_page = 200

        while True:
            page_issues = self.wp_connector.get_validation_issues({
                "church_id": church_id,
                "status": "open",
                "page": current_page,
                "per_page": fetch_per_page,
            })
            if not page_issues:
                break

            issues.extend(page_issues)
            if len(page_issues) < fetch_per_page:
                break

            current_page += 1
            if current_page > 50:
                logger.warning(
                    f"Reached validation-issue page limit while exporting church_id={church_id}. "
                    "Stopping after 50 pages."
                )
                break

        return issues

    @staticmethod
    def _normalized_sport_type(sport_type: Optional[str]) -> str:
        """Normalize sport labels so team issues can match roster rows."""
        return str(sport_type or "").split(" - ")[0].strip().casefold()

    @staticmethod
    def _issue_rule_level(issue: Dict[str, Any]) -> str:
        return str(issue.get("rule_level") or "").strip() or RULE_LEVEL["INDIVIDUAL"]

    @staticmethod
    def _issue_severity(issue: Dict[str, Any]) -> str:
        return str(issue.get("severity") or "").strip() or VALIDATION_SEVERITY["ERROR"]

    @staticmethod
    def _issue_format_type(issue: Dict[str, Any]) -> str:
        issue_format = str(issue.get("sport_format") or "").strip()
        if issue_format in FORMAT_MAPPINGS:
            return FORMAT_MAPPINGS[issue_format][0]
        return issue_format

    @staticmethod
    def _issue_gender(issue: Dict[str, Any]) -> str:
        issue_format = str(issue.get("sport_format") or "").strip()
        if issue_format in FORMAT_MAPPINGS:
            return str(FORMAT_MAPPINGS[issue_format][1])

        sport_type = str(issue.get("sport_type") or "")
        for gender in ("Women", "Men", "Mixed"):
            if gender.casefold() in sport_type.casefold():
                return gender
        return ""

    def _team_issue_matches_roster(self, issue: Dict[str, Any], roster_entry: Dict[str, Any]) -> bool:
        """Return True when a TEAM issue applies to a roster row."""
        if self._issue_rule_level(issue) != RULE_LEVEL["TEAM"]:
            return False

        if self._normalized_sport_type(issue.get("sport_type")) != self._normalized_sport_type(roster_entry.get("sport_type")):
            return False

        issue_gender = self._issue_gender(issue)
        roster_gender = str(roster_entry.get("sport_gender") or "").strip()
        if issue_gender and roster_gender and issue_gender.casefold() != roster_gender.casefold():
            return False

        issue_format_type = self._issue_format_type(issue)
        roster_format = str(roster_entry.get("sport_format") or "").strip()
        if issue_format_type and roster_format and issue_format_type.casefold() != roster_format.casefold():
            return False

        return True

    @staticmethod
    def _team_issue_scope_key(issue: Dict[str, Any]) -> Tuple[str, str, str]:
        return (
            ChurchTeamsExporter._normalized_sport_type(issue.get("sport_type")),
            ChurchTeamsExporter._issue_gender(issue).casefold(),
            ChurchTeamsExporter._issue_format_type(issue).casefold(),
        )

    @staticmethod
    def _normalized_gender(gender: Optional[str]) -> str:
        return str(gender or "").strip().casefold()

    @staticmethod
    def _normalized_name(name: Optional[str]) -> str:
        normalized = unicodedata.normalize("NFKD", str(name or "").strip())
        without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
        return " ".join(without_marks.casefold().split())

    @staticmethod
    def _token_matches(query_token: str, candidate_token: str) -> bool:
        if query_token == candidate_token:
            return True
        return len(query_token) >= 2 and candidate_token.startswith(query_token)

    @classmethod
    def _is_likely_name_match(cls, query_name_key: str, candidate_name_key: str) -> bool:
        query_tokens = [token for token in str(query_name_key or "").split() if token]
        candidate_tokens = [token for token in str(candidate_name_key or "").split() if token]
        if not query_tokens or not candidate_tokens:
            return False

        return all(
            any(cls._token_matches(query_token, candidate_token) for candidate_token in candidate_tokens)
            for query_token in query_tokens
        )

    @classmethod
    def _reverse_partner_suggestion_key(
        cls,
        participant_id: Any,
        sport_type: Optional[str],
        sport_gender: Optional[str],
        sport_format: Optional[str],
    ) -> Tuple[str, str, str, str]:
        return (
            str(participant_id or ""),
            cls._normalized_sport_type(sport_type),
            cls._normalized_gender(sport_gender),
            str(sport_format or "").strip().casefold(),
        )

    def _build_reverse_partner_suggestion_lookup(
        self,
        roster_rows: List[Dict[str, Any]],
    ) -> Dict[Tuple[str, str, str, str], List[str]]:
        """Infer reverse partner-name suggestions from same-event roster rows."""
        grouped_rows: Dict[Tuple[str, str, str, str], List[Dict[str, Any]]] = {}

        for row in roster_rows:
            participant_id = row.get("Participant ID (WP)")
            sport_type = row.get("sport_type")
            sport_gender = row.get("sport_gender")
            sport_format = row.get("sport_format")
            if participant_id in (None, "", 0, "0") or not sport_type or not sport_format:
                continue

            full_name = " ".join(
                part
                for part in (
                    str(row.get("First Name") or "").strip(),
                    str(row.get("Last Name") or "").strip(),
                )
                if part
            ).strip()
            if not full_name:
                continue

            event_key = (
                str(row.get("Church Team") or "").strip().upper(),
                self._normalized_sport_type(sport_type),
                self._normalized_gender(sport_gender),
                str(sport_format).strip().casefold(),
            )
            grouped_rows.setdefault(event_key, []).append({
                "participant_id": str(participant_id),
                "participant_name": full_name,
                "participant_name_key": self._normalized_name(full_name),
                "partner_name": str(row.get("partner_name") or "").strip(),
            })

        suggestions_by_key: Dict[Tuple[str, str, str, str], set[str]] = {}
        for (_, sport_type_key, sport_gender_key, sport_format_key), event_rows in grouped_rows.items():
            for claimant in event_rows:
                partner_name = claimant["partner_name"]
                if not partner_name:
                    continue

                partner_name_key = self._normalized_name(partner_name)
                matching_targets = [
                    candidate
                    for candidate in event_rows
                    if candidate["participant_id"] != claimant["participant_id"]
                    and self._is_likely_name_match(
                        partner_name_key,
                        candidate["participant_name_key"],
                    )
                ]

                unique_targets = {
                    candidate["participant_id"]: candidate
                    for candidate in matching_targets
                }
                if len(unique_targets) != 1:
                    continue

                target = next(iter(unique_targets.values()))
                suggestion_key = self._reverse_partner_suggestion_key(
                    target["participant_id"],
                    sport_type_key,
                    sport_gender_key,
                    sport_format_key,
                )
                suggestions_by_key.setdefault(suggestion_key, set()).add(claimant["participant_name"])

        return {
            key: sorted(names)
            for key, names in suggestions_by_key.items()
        }

    @staticmethod
    def _participant_name_from_lookup(
        participant_id: Any,
        participants_by_wp_id: Dict[str, Dict[str, Any]],
        fallback_first_name: Optional[str] = None,
        fallback_last_name: Optional[str] = None,
    ) -> str:
        participant_key = str(participant_id) if participant_id not in (None, "", 0, "0") else ""
        participant_info = participants_by_wp_id.get(participant_key, {})
        first_name = str(
            participant_info.get("First Name")
            or fallback_first_name
            or ""
        ).strip()
        last_name = str(
            participant_info.get("Last Name")
            or fallback_last_name
            or ""
        ).strip()
        return " ".join(part for part in (first_name, last_name) if part).strip()

    @staticmethod
    def _parse_partner_issue_names(issue_description: str) -> Tuple[str, str]:
        match = re.match(
            r"^(?P<claimant>.+?) listed (?P<partner>.+?) as their partner for ",
            str(issue_description or "").strip(),
        )
        if not match:
            return "", ""
        return match.group("claimant").strip(), match.group("partner").strip()

    def _build_issue_based_reverse_partner_suggestion_lookup(
        self,
        issues: List[Dict[str, Any]],
        participants_by_wp_id: Dict[str, Dict[str, Any]],
    ) -> Dict[Tuple[str, str, str, str], List[str]]:
        """Infer reverse partner suggestions from TEAM unmatched-partner warnings."""
        missing_targets_by_event: Dict[Tuple[str, str, str], List[Dict[str, str]]] = {}

        for issue in issues:
            if issue.get("issue_type") != "missing_doubles_partner":
                continue

            participant_id = issue.get("participant_id")
            if participant_id in (None, "", 0, "0"):
                continue

            participant_name = self._participant_name_from_lookup(
                participant_id,
                participants_by_wp_id,
                issue.get("first_name"),
                issue.get("last_name"),
            )
            if not participant_name:
                continue

            event_key = (
                self._normalized_sport_type(issue.get("sport_type")),
                self._normalized_gender(self._issue_gender(issue)),
                self._issue_format_type(issue).casefold(),
            )
            missing_targets_by_event.setdefault(event_key, []).append({
                "participant_id": str(participant_id),
                "participant_name": participant_name,
                "participant_name_key": self._normalized_name(participant_name),
            })

        suggestions_by_key: Dict[Tuple[str, str, str, str], set[str]] = {}
        for issue in issues:
            if issue.get("issue_type") != "doubles_partner_unmatched":
                continue

            event_key = (
                self._normalized_sport_type(issue.get("sport_type")),
                self._normalized_gender(self._issue_gender(issue)),
                self._issue_format_type(issue).casefold(),
            )
            event_targets = missing_targets_by_event.get(event_key, [])
            if not event_targets:
                continue

            parsed_claimant_name, parsed_partner_name = self._parse_partner_issue_names(
                str(issue.get("issue_description") or "")
            )
            if not parsed_partner_name:
                continue

            claimant_name = self._participant_name_from_lookup(
                issue.get("participant_id"),
                participants_by_wp_id,
            ) or parsed_claimant_name
            if not claimant_name:
                continue

            partner_name_key = self._normalized_name(parsed_partner_name)
            matching_targets = [
                target
                for target in event_targets
                if self._is_likely_name_match(partner_name_key, target["participant_name_key"])
            ]
            unique_targets = {
                target["participant_id"]: target
                for target in matching_targets
            }
            if len(unique_targets) != 1:
                continue

            target = next(iter(unique_targets.values()))
            suggestion_key = self._reverse_partner_suggestion_key(
                target["participant_id"],
                issue.get("sport_type"),
                self._issue_gender(issue),
                self._issue_format_type(issue),
            )
            suggestions_by_key.setdefault(suggestion_key, set()).add(claimant_name)

        return {
            key: sorted(names)
            for key, names in suggestions_by_key.items()
        }

    @staticmethod
    def _merge_reverse_partner_suggestions(
        *lookups: Dict[Tuple[str, str, str, str], List[str]],
    ) -> Dict[Tuple[str, str, str, str], List[str]]:
        merged: Dict[Tuple[str, str, str, str], set[str]] = {}
        for lookup in lookups:
            for key, names in lookup.items():
                merged.setdefault(key, set()).update(names)

        return {
            key: sorted(names)
            for key, names in merged.items()
        }

    def _build_validation_issue_rows(
        self,
        church_code: str,
        issues: List[Dict[str, Any]],
        participants_by_wp_id: Dict[str, Dict[str, Any]],
        reverse_partner_suggestions: Optional[Dict[Tuple[str, str, str, str], List[str]]] = None,
    ) -> List[Dict[str, Any]]:
        """Build export rows for the Validation-Issues tab."""
        rows: List[Dict[str, Any]] = []
        reverse_partner_suggestions = reverse_partner_suggestions or {}
        for issue in issues:
            participant_id = issue.get("participant_id")
            participant_key = str(participant_id) if participant_id not in (None, "", 0, "0") else ""
            participant_info = participants_by_wp_id.get(participant_key, {})
            first_name = str(issue.get("first_name") or participant_info.get("First Name") or "").strip()
            last_name = str(issue.get("last_name") or participant_info.get("Last Name") or "").strip()
            participant_name = " ".join(part for part in (first_name, last_name) if part).strip()
            issue_description = issue.get("issue_description", "")

            if (
                issue.get("issue_type") == "missing_doubles_partner"
                and participant_key
            ):
                suggestion_key = self._reverse_partner_suggestion_key(
                    participant_key,
                    issue.get("sport_type"),
                    self._issue_gender(issue),
                    self._issue_format_type(issue),
                )
                reverse_claimants = reverse_partner_suggestions.get(suggestion_key, [])
                if len(reverse_claimants) == 1:
                    issue_description = (
                        f"{issue_description}; perhaps {reverse_claimants[0]} listed you as partner."
                    )
                elif len(reverse_claimants) > 1:
                    suggestion_list = ", ".join(reverse_claimants)
                    issue_description = (
                        f"{issue_description}; ambiguous reverse partner claims: "
                        f"{suggestion_list}. Use full name."
                    )

            rows.append({
                "Church Team": church_code,
                "Rule Level": self._issue_rule_level(issue),
                "Severity": self._issue_severity(issue),
                "Status": issue.get("status", "open"),
                "Issue Type": issue.get("issue_type", ""),
                "Rule Code": issue.get("rule_code", ""),
                "Participant ID (WP)": participant_id or "",
                "ChMeetings ID": participant_info.get("ChMeetings ID", ""),
                "Participant Name": participant_name,
                "Approval_Status (WP)": participant_info.get("Approval_Status (WP)", ""),
                "sport_type": issue.get("sport_type"),
                "sport_format": issue.get("sport_format"),
                "Issue Description": issue_description,
            })

        return rows

    def _filter_reportable_validation_issues(
        self,
        church_code: str,
        issues: List[Dict[str, Any]],
        participants_by_wp_id: Dict[str, Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Keep only validation issues that belong in the current church snapshot.

        The export is driven by current ChMeetings Team-group membership. If a person
        was deleted/re-registered in ChMeetings, WordPress can still hold open
        individual issues tied to an older participant_id. Those stale issues should
        not appear in the current church workbook.
        """
        current_participant_ids = {
            str(participant_id)
            for participant_id in participants_by_wp_id
            if str(participant_id).strip()
        }
        filtered_issues: List[Dict[str, Any]] = []
        skipped_stale_individual_issues = 0

        for issue in issues:
            if self._issue_rule_level(issue) != RULE_LEVEL["INDIVIDUAL"]:
                filtered_issues.append(issue)
                continue

            participant_id = issue.get("participant_id")
            participant_key = (
                str(participant_id).strip()
                if participant_id not in (None, "", 0, "0")
                else ""
            )
            if not participant_key or participant_key in current_participant_ids:
                filtered_issues.append(issue)
                continue

            skipped_stale_individual_issues += 1

        if skipped_stale_individual_issues:
            logger.warning(
                f"Team {church_code}: filtered out {skipped_stale_individual_issues} "
                "stale INDIVIDUAL validation issue(s) that no longer map to a "
                "current participant in the ChMeetings Team-group snapshot."
            )

        return filtered_issues

    def _fetch_chm_church_team_data(self, target_church_code: Optional[str] = None) -> Dict[str, List[Dict[str, Any]]]:
        """
        Fetches data for all individuals in ChMeetings groups starting with 'Team '.
        Organizes data by church code.
        If target_church_code is specified, only fetches for that church.
        """
        logger.info(f"Fetching ChMeetings data. Target church: {target_church_code or 'ALL'}")
        self.last_orphaned_memberships_by_church = {}
        all_chm_groups = self.chm_connector.get_groups()
        if not all_chm_groups:
            logger.warning("No groups found in ChMeetings.")
            return {}

        chm_data_by_church: Dict[str, List[Dict[str, Any]]] = {}
        # Renamed from latest_chm_update_by_church to avoid conflict with instance variable if used directly
        _latest_chm_update_by_church_dt: Dict[str, Optional[datetime]] = {}
        orphaned_ids_by_church: Dict[str, List[str]] = {}


        team_groups = [g for g in all_chm_groups if g.get("name", "").startswith(Config.TEAM_PREFIX + " ")]
        logger.info(f"Found {len(team_groups)} ChMeetings groups with prefix '{Config.TEAM_PREFIX} '.")

        for group in team_groups:
            group_name = group.get("name", "")
            group_id = str(group.get("id"))
            church_code = self._parse_church_code_from_group_name(group_name)

            if not church_code:
                continue

            if target_church_code and church_code != target_church_code.upper():
                continue

            logger.debug(f"Processing ChMeetings group: '{group_name}' (ID: {group_id}, Church: {church_code})")
            chm_data_by_church.setdefault(church_code, [])
            _latest_chm_update_by_church_dt.setdefault(church_code, None)


            group_people_summaries = self.chm_connector.get_group_people(group_id)
            if not group_people_summaries:
                logger.info(f"No people found in ChMeetings group '{group_name}'.")
                continue

            for person_summary in group_people_summaries:
                person_id_str = str(person_summary.get("person_id")) # API uses person_id in group people
                if not person_id_str:
                    logger.warning(f"Missing person_id in summary from group '{group_name}': {person_summary}")
                    continue

                person_details_response = self.chm_connector.get_person(person_id_str)
                if not person_details_response:
                    if getattr(self.chm_connector, "last_get_person_status", None) == "not_found":
                        orphaned_ids_by_church.setdefault(church_code, []).append(person_id_str)
                    else:
                        logger.warning(
                            f"Could not fetch details for ChM Person ID: {person_id_str} "
                            f"from group '{group_name}'."
                        )
                    continue
                
                person_data = person_details_response if isinstance(person_details_response, dict) and "data" not in person_details_response else person_details_response.get("data", {})
                if not person_data:
                    logger.warning(f"Empty person data for ChM Person ID: {person_id_str} from group '{group_name}'.")
                    continue

                chm_id = str(person_data.get("id")) 

                additional_fields = {f.get("field_name"): f.get("value") for f in person_data.get("additional_fields", [])}
                
                updated_on_str = person_data.get("updated_on", "1970-01-01T00:00:00+00:00")
                try:
                    if 'Z' in updated_on_str:
                         updated_on_dt = datetime.fromisoformat(updated_on_str.replace("Z", "+00:00"))
                    elif 'T' in updated_on_str and ('+' in updated_on_str or '-' in updated_on_str[updated_on_str.rfind('T'):]): # Check for timezone info after T
                        updated_on_dt = datetime.fromisoformat(updated_on_str)
                    else: 
                        updated_on_dt = datetime.strptime(updated_on_str, "%Y-%m-%d %H:%M:%S")
                except ValueError as ve:
                    logger.warning(f"Could not parse ChM updated_on '{updated_on_str}' for person {chm_id}: {ve}. Using epoch.")
                    updated_on_dt = datetime.fromisoformat("1970-01-01T00:00:00+00:00")

                current_latest_dt = _latest_chm_update_by_church_dt[church_code]
                if current_latest_dt is None or updated_on_dt > current_latest_dt:
                    _latest_chm_update_by_church_dt[church_code] = updated_on_dt


                mapped_person = {
                    "Church Team": church_code,
                    "ChMeetings ID": chm_id,
                    "First Name": person_data.get("first_name", ""),
                    "Last Name": person_data.get("last_name", ""),
                    "Gender": person_data.get("gender", ""),
                    "Birthdate": person_data.get("birth_date", ""),
                    "Mobile Phone": person_data.get("mobile", ""),
                    "Email": person_data.get("email", "").strip(),
                    "Is_Member_ChM": additional_fields.get(MEMBERSHIP_QUESTION, "No") == "Yes",
                    "ChM_Roles": additional_fields.get(CHM_FIELDS["ROLES"], ""),
                    "ChM_Completion_Checklist": additional_fields.get(CHM_FIELDS["COMPLETION_CHECKLIST"], ""),
                    "ChM_Primary_Sport": additional_fields.get(CHM_FIELDS["PRIMARY_SPORT"], ""),
                    "ChM_Secondary_Sport": additional_fields.get(CHM_FIELDS["SECONDARY_SPORT"], ""),
                    "ChM_Other_Events": additional_fields.get(CHM_FIELDS["OTHER_EVENTS"], ""),
                    "Update_on_ChM": updated_on_dt.strftime("%Y-%m-%d %H:%M:%S") 
                }
                chm_data_by_church[church_code].append(mapped_person)
        
        self.latest_chm_update_by_church = { # Store on instance
            code: dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "N/A"
            for code, dt in _latest_chm_update_by_church_dt.items()
        }
        self.last_orphaned_memberships_by_church = {
            code: len(ids) for code, ids in orphaned_ids_by_church.items()
        }

        total_orphaned_memberships = sum(len(ids) for ids in orphaned_ids_by_church.values())
        if total_orphaned_memberships:
            for church_code, ids in sorted(orphaned_ids_by_church.items()):
                logger.warning(
                    f"Team {church_code}: skipped {len(ids)} orphaned member ID(s) — "
                    f"[{', '.join(ids)}]"
                )
            logger.warning(
                f"Skipped {total_orphaned_memberships} orphaned Team-group membership(s) "
                f"across {len(orphaned_ids_by_church)} church(es). "
                "Run 'python main.py audit-team-groups' to clean up."
            )
        
        if target_church_code and not chm_data_by_church.get(target_church_code.upper()):
             logger.warning(f"No ChMeetings group found or no members in group for target church: {target_church_code}")

        logger.info(f"Fetched ChMeetings data for {len(chm_data_by_church)} churches.")
        return chm_data_by_church

    def generate_reports(self, target_church_code: Optional[str], output_dir: Path,
                        force_resend_pending: bool = False, force_resend_validated1: bool = False, 
                        force_resend_validated2: bool = False, dry_run: bool = False,
                        target_resend_chm_id: Optional[str] = None) -> bool:
        """
        Generates Excel status reports for church teams.
        If target_church_code is provided, generates a single report for that church.
        Otherwise, generates a report for each church and one consolidated "ALL" report.
        Furthermore, the flags force_resend_pending, force_resend_validated1, and force_resend_validated2
        control whether to resend pastoral approval for participants with approval pending, validated1 (under church rep's review),
        and validated2 (no review yet).
        If target_resend_chm_id is provided, resend actions are limited to that participant.
        Dry run mode does not send the email yet but note the actions would be taken.
        """
        if not self.chm_connector.authenticate():
            logger.error("ChMeetings authentication failed. Cannot generate reports.")
            return False

        deadline_date = datetime.strptime(REGISTRATION_DEADLINE, "%Y-%m-%d").date()
        today = datetime.now().date()
        if today >= deadline_date:
            fee_tier = f"LATE (${ATHLETE_FEE_LATE}, past deadline {REGISTRATION_DEADLINE})"
        else:
            fee_tier = f"STANDARD (${ATHLETE_FEE_STANDARD} early-bird, deadline {REGISTRATION_DEADLINE})"
        logger.info(f"Athlete fee tier: {fee_tier}")

        logger.info(f"Starting report generation. Target Church: {target_church_code or 'ALL'}. Output Dir: {output_dir}")

        chm_data_by_church = self._fetch_chm_church_team_data(target_church_code)

        if not chm_data_by_church:
            logger.warning("No data fetched from ChMeetings. No reports will be generated.")
            return True 

        all_contacts_data: List[Dict[str, Any]] = []
        all_rosters_data: List[Dict[str, Any]] = []
        all_validation_data: List[Dict[str, Any]] = []
        summary_data_list: List[Dict[str, Any]] = []

        churches_to_process_codes = [target_church_code.upper()] if target_church_code else sorted(list(chm_data_by_church.keys()))

        for church_code_iter in churches_to_process_codes:
            if church_code_iter not in chm_data_by_church:
                logger.warning(f"Skipping report for {church_code_iter} as no ChM data was found (e.g., no 'Team {church_code_iter}' group).")
                continue
            
            logger.info(f"Processing data for church: {church_code_iter}")
            church_contacts_rows: List[Dict[str, Any]] = []
            church_rosters_rows: List[Dict[str, Any]] = []
            church_validation_rows: List[Dict[str, Any]] = []
            participants_by_wp_id: Dict[str, Dict[str, Any]] = {}

            total_members_chm = len(chm_data_by_church[church_code_iter])
            total_participants_wp = 0
            total_approved_wp = 0
            total_denied_wp = 0
            total_pending_participants_wp = 0
            total_with_open_errors_wp = 0
            total_athlete_fees = 0
            church_wp = self.wp_connector.get_church_by_code(church_code_iter)
            church_wp_id = church_wp.get("church_id") if church_wp else None
            open_validation_issues = self._fetch_open_validation_issues(church_wp_id)
            participant_error_lookup: Dict[str, List[Dict[str, Any]]] = {}
            team_validation_issues = [
                issue for issue in open_validation_issues
                if self._issue_rule_level(issue) == RULE_LEVEL["TEAM"]
            ]

            for issue in open_validation_issues:
                participant_id = issue.get("participant_id")
                if participant_id in (None, "", 0, "0"):
                    continue
                if self._issue_severity(issue) != VALIDATION_SEVERITY["ERROR"]:
                    continue
                participant_error_lookup.setdefault(str(participant_id), []).append(issue)

            for chm_person in chm_data_by_church[church_code_iter]:
                chm_id = chm_person["ChMeetings ID"]

                roles_str = chm_person.get("ChM_Roles", "")
                is_participant_chm = any(role.strip().lower() in ["athlete", "participant", "athlete/participant"] for role in roles_str.split(","))

                wp_participant_id_val = 0
                approval_status_val = "N/A"
                total_open_errors_val = 0
                first_open_error_desc_val = ""
                photo_url_val = "N/A"
                wp_created_at_str = ""

                if is_participant_chm:
                    wp_participants = self.wp_connector.get_participants({"chmeetings_id": chm_id})
                    if wp_participants:
                        wp_participant = wp_participants[0]
                        wp_participant_id_val = wp_participant.get("participant_id", 0)
                        approval_status_val = wp_participant.get("approval_status", "pending")
                        photo_url_val = wp_participant.get("photo_url", "N/A")
                        wp_created_at_str = wp_participant.get("created_at", "")
                        participant_issue_list = participant_error_lookup.get(str(wp_participant_id_val), [])
                        total_participants_wp += 1

                        if approval_status_val == "approved":
                            total_approved_wp += 1
                        if approval_status_val == "denied": # ADD THIS BLOCK
                            total_denied_wp += 1
                        if approval_status_val in ["pending", "validated", "pending_approval"]:
                            total_pending_participants_wp +=1

                        if wp_participant_id_val:
                            total_open_errors_val = len(participant_issue_list)
                            if participant_issue_list:
                                first_open_error_desc_val = participant_issue_list[0].get("issue_description", "")

                            participants_by_wp_id[str(wp_participant_id_val)] = {
                                "ChMeetings ID": chm_id,
                                "First Name": chm_person["First Name"],
                                "Last Name": chm_person["Last Name"],
                                "Approval_Status (WP)": approval_status_val,
                            }
                            wp_rosters = self.wp_connector.get_rosters({"participant_id": wp_participant_id_val})
                            for roster_entry in wp_rosters:
                                matching_team_issues = [
                                    issue for issue in team_validation_issues
                                    if self._team_issue_matches_roster(issue, roster_entry)
                                ]
                                church_rosters_rows.append({
                                    "Church Team": church_code_iter,
                                    "ChMeetings ID": chm_id,
                                    "Participant ID (WP)": wp_participant_id_val,
                                    "Approval_Status (WP)": approval_status_val,
                                    "Is_Member_ChM": chm_person.get("Is_Member_ChM", False),  # ADD THIS LINE
                                    "Photo": f'=IMAGE("{photo_url_val}")' if photo_url_val != "N/A" and photo_url_val.startswith(("http://", "https://")) else "",  # ADD THIS LINE
#NOTE: The above line assumes the Excel engine supports IMAGE formula, which is not standard in pandas and will insert "@" after "="
                                    "First Name": chm_person["First Name"], 
                                    "Last Name": chm_person["Last Name"],
                                    "Gender": chm_person["Gender"],
                                    "Age (at Event)": self._calculate_age(chm_person["Birthdate"]),
                                    "Mobile Phone": chm_person["Mobile Phone"],
                                    "Email": chm_person["Email"],
                                    "sport_type": roster_entry.get("sport_type"),
                                    "sport_gender": roster_entry.get("sport_gender"),
                                    "sport_format": roster_entry.get("sport_format"),
                                    "team_order": roster_entry.get("team_order"),
                                    "partner_name": roster_entry.get("partner_name"),
                                    "Open_TEAM_Issue_Count (WP)": len(matching_team_issues),
                                    "Open_TEAM_Issue_Desc (WP)": " | ".join(
                                        str(issue.get("issue_description", "")).strip()
                                        for issue in matching_team_issues
                                        if str(issue.get("issue_description", "")).strip()
                                    ),
                                })
                    else: 
                        approval_status_val = "Not in WordPress"
                
                checklist_statuses = self._get_completion_checklist_statuses(chm_person.get("ChM_Completion_Checklist"))

                registration_date_str = ""
                if is_participant_chm:
                    _primary = chm_person.get("ChM_Primary_Sport", "")
                    _secondary = chm_person.get("ChM_Secondary_Sport", "")
                    _other = chm_person.get("ChM_Other_Events", "")

                    if wp_created_at_str:
                        try:
                            created_date = datetime.strptime(wp_created_at_str.split()[0], "%Y-%m-%d").date()
                            registration_date_str = created_date.strftime("%Y-%m-%d")
                            deadline_date = datetime.strptime(REGISTRATION_DEADLINE, "%Y-%m-%d").date()

                            if not _primary and not _secondary and _other:
                                athlete_fee = ATHLETE_FEE_OTHER_EVENTS_ONLY
                            elif created_date >= deadline_date:
                                athlete_fee = ATHLETE_FEE_LATE
                            else:
                                athlete_fee = ATHLETE_FEE_STANDARD
                        except (ValueError, AttributeError):
                            athlete_fee = (
                                ATHLETE_FEE_OTHER_EVENTS_ONLY
                                if (not _primary and not _secondary and _other)
                                else ATHLETE_FEE_STANDARD
                            )
                    else:
                        athlete_fee = (
                            ATHLETE_FEE_OTHER_EVENTS_ONLY
                            if (not _primary and not _secondary and _other)
                            else ATHLETE_FEE_STANDARD
                        )
                    total_athlete_fees += athlete_fee
                else:
                    athlete_fee = ""

                contact_row = {
                    "Church Team": church_code_iter,
                    "ChMeetings ID": chm_id,
                    "First Name": chm_person["First Name"],
                    "Last Name": chm_person["Last Name"],
                    "Is_Participant": "Yes" if is_participant_chm else "No",
                    "Is_Member_ChM": "Yes" if chm_person["Is_Member_ChM"] else "No",
                    "Participant ID (WP)": wp_participant_id_val,
                    "Approval_Status (WP)": approval_status_val,
                    "Total_Open_ERRORs (WP)": total_open_errors_val,
                    "Gender": chm_person["Gender"],
                    "Birthdate": chm_person["Birthdate"],
                    "Age (at Event)": self._calculate_age(chm_person["Birthdate"]),
                    "Mobile Phone": chm_person["Mobile Phone"],
                    "Email": chm_person["Email"],
                    "Registration Date (WP)": registration_date_str,
                    "Athlete Fee": athlete_fee,
                    "First_Open_ERROR_Desc (WP)": first_open_error_desc_val,
                    **checklist_statuses,
                    "Photo URL (WP)": photo_url_val,
                    "Update_on_ChM": chm_person["Update_on_ChM"]
                }
                church_contacts_rows.append(contact_row)

            reverse_partner_suggestions = self._merge_reverse_partner_suggestions(
                self._build_reverse_partner_suggestion_lookup(church_rosters_rows),
                self._build_issue_based_reverse_partner_suggestion_lookup(
                    open_validation_issues,
                    participants_by_wp_id,
                ),
            )

            reportable_validation_issues = self._filter_reportable_validation_issues(
                church_code_iter,
                open_validation_issues,
                participants_by_wp_id,
            )
            reportable_team_validation_issues = [
                issue for issue in reportable_validation_issues
                if self._issue_rule_level(issue) == RULE_LEVEL["TEAM"]
            ]

            church_validation_rows = self._build_validation_issue_rows(
                church_code_iter,
                reportable_validation_issues,
                participants_by_wp_id,
                reverse_partner_suggestions,
            )

            individual_open_errors = [
                issue for issue in reportable_validation_issues
                if self._issue_rule_level(issue) == RULE_LEVEL["INDIVIDUAL"]
                and self._issue_severity(issue) == VALIDATION_SEVERITY["ERROR"]
            ]
            team_open_errors = [
                issue for issue in reportable_team_validation_issues
                if self._issue_severity(issue) == VALIDATION_SEVERITY["ERROR"]
            ]
            open_warnings = [
                issue for issue in reportable_validation_issues
                if self._issue_severity(issue) == VALIDATION_SEVERITY["WARNING"]
            ]
            participant_ids_with_errors = {
                str(issue.get("participant_id"))
                for issue in individual_open_errors
                if issue.get("participant_id") not in (None, "", 0, "0")
            }
            total_with_open_errors_wp = len(participant_ids_with_errors)
            total_sports_with_team_issues = len({
                self._team_issue_scope_key(issue)
                for issue in team_validation_issues
            })

            summary_data_list.append({
                "Church Code": church_code_iter,
                "Total Members (ChM Team Group)": total_members_chm,
                "Total Participants (in WP)": total_participants_wp,
                "Total Approved (WP)": total_approved_wp,
                "Total Denied (WP)": total_denied_wp,
                "Total Pending Approval (WP)": total_pending_participants_wp,
                "Total Participants w/ Open ERRORs (WP)": total_with_open_errors_wp,
                "Total Open Individual ERRORs (WP)": len(individual_open_errors),
                "Total Open TEAM ERRORs (WP)": len(team_open_errors),
                "Total Open WARNINGs (WP)": len(open_warnings),
                "Total Sports w/ Open TEAM Issues (WP)": total_sports_with_team_issues,
                "Total Athlete Fees": total_athlete_fees,
                "Latest ChM Record Update for Team": self.latest_chm_update_by_church.get(church_code_iter, "N/A")
            })
            
            all_contacts_data.extend(church_contacts_rows)
            all_rosters_data.extend(church_rosters_rows)
            all_validation_data.extend(church_validation_rows)

            if target_church_code: 
                safe_code = "".join(c if c.isalnum() else "_" for c in church_code_iter)
                filename = f"Church_Team_Status_{safe_code}.xlsx"
                self._write_excel_report(output_dir / filename,
                                         [summary_data_list[-1]], 
                                         church_contacts_rows,
                                         church_rosters_rows,
                                         church_validation_rows)
                
        if not target_church_code: 
            for church_code_iter in churches_to_process_codes: # Iterate using the original list of codes
                if church_code_iter not in chm_data_by_church: continue 

                current_church_contacts = [row for row in all_contacts_data if row["Church Team"] == church_code_iter]
                current_church_rosters = [row for row in all_rosters_data if row["Church Team"] == church_code_iter]
                current_church_validation = [row for row in all_validation_data if row["Church Team"] == church_code_iter]
                current_church_summary = [s_data for s_data in summary_data_list if s_data["Church Code"] == church_code_iter]

                safe_code = "".join(c if c.isalnum() else "_" for c in church_code_iter)
                filename = f"Church_Team_Status_{safe_code}.xlsx"
                self._write_excel_report(output_dir / filename,
                                         current_church_summary,
                                         current_church_contacts,
                                         current_church_rosters,
                                         current_church_validation)
            
            all_filename = f"Church_Team_Status_ALL_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            self._write_excel_report(output_dir / all_filename,
                                     summary_data_list,
                                     all_contacts_data,
                                     all_rosters_data,
                                     all_validation_data,
                                     include_venue_capacity=True)

        # Handle force resend options
        if force_resend_pending or force_resend_validated1 or force_resend_validated2:
            resend_count = self._handle_force_resend(
                all_contacts_data, force_resend_pending, force_resend_validated1, 
                force_resend_validated2, dry_run, target_resend_chm_id=target_resend_chm_id
            )
            logger.info(f"Force resend completed. Total emails {'would be sent' if dry_run else 'sent'}: {resend_count}")
        
        logger.info("Report generation process finished.")
        return True

    @staticmethod
    def _decompose_event_name(event_name: str) -> Tuple[str, str, str]:
        """Split a canonical event name like 'Basketball - Men Team' into
        (sport_type, sport_gender, sport_format) using the same casing the
        roster rows store after sync."""
        parts = event_name.split(" - ", 1)
        sport_type = parts[0].strip()
        if len(parts) < 2:
            return sport_type, "", "Team"
        suffix = parts[1].upper()
        if "WOMEN" in suffix:
            gender = "Women"
        elif "MEN" in suffix:
            gender = "Men"
        elif "MIXED" in suffix or "COED" in suffix:
            gender = "Mixed"
        else:
            gender = ""
        sport_format = "Singles" if "SINGLES" in suffix else "Team"
        return sport_type, gender, sport_format

    def _get_min_team_size(self, event_name: str) -> int:
        """Look up minimum team size from the validation ruleset; fall back
        to COURT_ESTIMATE_MIN_TEAM_SIZE if the JSON rule is absent."""
        if not hasattr(self, "_rules_manager_cache"):
            try:
                self._rules_manager_cache = RulesManager(collection="SUMMER_2026")
            except Exception as e:
                logger.warning(f"Could not load validation rules for venue estimate: {e}")
                self._rules_manager_cache = None
        if self._rules_manager_cache is not None:
            for rule in self._rules_manager_cache.get_rules_for_sport(event_name):
                if rule.get("rule_type") == "team_size" and rule.get("category") == "min":
                    try:
                        return int(rule.get("value"))
                    except (TypeError, ValueError):
                        pass
        return int(COURT_ESTIMATE_MIN_TEAM_SIZE.get(event_name, 0))

    def _count_estimating_teams(self, roster_rows: List[Dict[str, Any]],
                                 event_name: str, min_team_size: int) -> Dict[str, Any]:
        """Return estimating/potential team counts and the qualifying church codes.

        Approval-agnostic — every roster entry counts.

        Returns a dict with:
          n_estimating  – churches with >= min_team_size entries (ready to compete)
          n_potential   – churches with >= 1 but < min_team_size entries (still forming)
          team_codes    – sorted, comma-separated list of estimating church codes
        """
        if min_team_size <= 0:
            return {"n_estimating": 0, "n_potential": 0, "team_codes": ""}
        target_type, target_gender, _ = self._decompose_event_name(event_name)
        counts_by_church: Dict[str, int] = {}
        for r in roster_rows:
            r_type = str(r.get("sport_type") or "").strip()
            r_gender = str(r.get("sport_gender") or "").strip()
            # Primary/secondary sports are stored as the base name (e.g. "Basketball");
            # Other-Events sports are stored as the full SPORT_TYPE value verbatim.
            # Accept either so both paths match.
            if (r_type.casefold() != target_type.casefold() and
                    r_type.casefold() != event_name.casefold()):
                continue
            if target_gender and r_gender.casefold() != target_gender.casefold():
                continue
            church = str(r.get("Church Team") or "").strip()
            if not church:
                continue
            counts_by_church[church] = counts_by_church.get(church, 0) + 1
        estimating = sorted(c for c, n in counts_by_church.items() if n >= min_team_size)
        partial = [c for c, n in counts_by_church.items() if 0 < n < min_team_size]
        return {
            "n_estimating": len(estimating),
            "n_potential": len(estimating) + len(partial),  # all churches with >= 1 entry
            "team_codes": ", ".join(estimating),
        }

    @staticmethod
    def _get_playoff_teams(n_teams: int) -> int:
        for rule in COURT_ESTIMATE_PLAYOFF_RULES:
            if rule["min_teams"] <= n_teams <= rule["max_teams"]:
                return int(rule["playoff_teams"])
        return 0

    def _compute_court_slots(self, n_teams: int,
                              minutes_per_game: int = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
                              pool_games_per_team: int = COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM) -> Dict[str, Any]:
        include_third = COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME

        pool_slots = ceil((n_teams * pool_games_per_team) / 2) if n_teams > 0 else 0
        playoff_teams = self._get_playoff_teams(n_teams)
        playoff_slots = max(playoff_teams - 1, 0)
        third_place_slots = 1 if include_third and playoff_teams >= 4 else 0
        total_slots = pool_slots + playoff_slots + third_place_slots
        return {
            "pool_games_per_team": pool_games_per_team,
            "minutes_per_game": minutes_per_game,
            "pool_slots": pool_slots,
            "playoff_teams": playoff_teams,
            "playoff_slots": playoff_slots,
            "include_third_place": include_third,
            "third_place_slots": third_place_slots,
            "total_slots": total_slots,
            "court_hours": round(total_slots * minutes_per_game / 60, 2),
        }

    def _count_racquet_entries(self, roster_rows: List[Dict[str, Any]],
                               sport_name: str) -> Dict[str, Any]:
        """Count racquet sport entries for the venue estimator.

        Estimating Entries = complete pairs floor(n_doubles / 2) + n_singles.
        Potential Entries  = total individual registrations (one person may be
                             waiting for a partner to sign up).
        """
        n_singles = 0
        n_doubles = 0
        for r in roster_rows:
            if str(r.get("sport_type") or "").strip().casefold() != sport_name.casefold():
                continue
            fmt = str(r.get("sport_format") or "").strip().casefold()
            if "singles" in fmt:
                n_singles += 1
            else:
                n_doubles += 1
        n_estimating = n_singles + (n_doubles // 2)
        n_potential = n_singles + n_doubles
        return {
            "n_estimating": n_estimating,
            "n_potential": n_potential,
            "team_codes": "",
        }

    # ── Pod-Divisions / Pod-Entries-Review helpers (Issue #88) ──────────────

    @staticmethod
    def _pod_format_class(sport_format: str) -> str:
        """Classify a sport_format string into 'singles', 'doubles', or 'anomaly'."""
        fmt = (sport_format or "").strip().casefold()
        if "single" in fmt:
            return "singles"
        if "double" in fmt:
            return "doubles"
        return "anomaly"

    @staticmethod
    def _make_division_id(sport_type: str, sport_gender: str, format_class: str) -> str:
        """Return a canonical division_id string, e.g. 'BAD-Men-Singles'."""
        abbrev = POD_SPORT_ABBREV.get(sport_type, sport_type[:3].upper())
        gender_part = sport_gender or "Unknown"
        format_part = format_class.title() if format_class != "anomaly" else "Anomaly"
        return f"{abbrev}-{gender_part}-{format_part}"

    @staticmethod
    def _build_pod_error_lookup(
        validation_rows: List[Dict[str, Any]],
    ) -> Dict[str, set]:
        """Return {str(participant_id) → set(sport_types)} for open ERRORs."""
        lookup: Dict[str, set] = {}
        for v in validation_rows:
            if str(v.get("Severity", "")).upper() != "ERROR":
                continue
            if str(v.get("Status", "")).lower() != "open":
                continue
            pid = str(v.get("Participant ID (WP)") or "").strip()
            if not pid or pid in ("0",):
                continue
            sport = str(v.get("sport_type") or "").strip()
            lookup.setdefault(pid, set()).add(sport)
        return lookup

    def _build_pod_divisions_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Build one summary row per pod division for the Pod-Divisions tab."""
        error_lookup = self._build_pod_error_lookup(validation_rows)

        # Accumulate per-division counts.
        # key: (sport_type, sport_gender, format_class)
        divs: Dict[tuple, Dict[str, Any]] = {}

        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            sport_gender = str(r.get("sport_gender") or "").strip()
            sport_format = str(r.get("sport_format") or "").strip()
            fmt_class = self._pod_format_class(sport_format)

            key = (sport_type, sport_gender, fmt_class)
            if key not in divs:
                divs[key] = {
                    "sport_type": sport_type,
                    "sport_gender": sport_gender,
                    "sport_format": sport_format,
                    "n_total": 0,
                    "n_confirmed": 0,
                    "n_anomaly": 0,
                }

            pid = str(r.get("Participant ID (WP)") or "").strip()
            has_error = bool(pid and pid not in ("0",) and sport_type in error_lookup.get(pid, set()))

            if fmt_class == "anomaly":
                divs[key]["n_anomaly"] += 1
            else:
                divs[key]["n_total"] += 1
                if not has_error:
                    divs[key]["n_confirmed"] += 1

        rows: List[Dict[str, Any]] = []
        for (sport_type, sport_gender, fmt_class), div in sorted(divs.items()):
            division_id = self._make_division_id(sport_type, sport_gender, fmt_class)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(sport_type, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)

            if fmt_class == "doubles":
                planning = div["n_total"] // 2
                confirmed = div["n_confirmed"] // 2
            else:
                planning = div["n_total"]
                confirmed = div["n_confirmed"]

            provisional = planning - confirmed
            anomaly_count = div["n_anomaly"]

            if planning == 0 and anomaly_count == 0:
                status = "Empty"
            elif planning == 0:
                status = "AnomalyOnly"
            elif anomaly_count > 0 or provisional > 0:
                status = "Partial"
            else:
                status = "Ready"

            rows.append({
                "division_id": division_id,
                "sport_type": sport_type,
                "sport_gender": sport_gender,
                "sport_format": div["sport_format"],
                # Keep pod division rows aligned with venue_input.xlsx and the solver's
                # exact C4 resource-type matching. Generic "Court" breaks all pod games.
                "resource_type": POD_RESOURCE_EVENT_TYPE.get(sport_type, "Court"),
                "minutes_per_game": mpg,
                "planning_entries": planning,
                "confirmed_entries": confirmed,
                "provisional_entries": provisional,
                "anomaly_count": anomaly_count,
                "division_status": status,
                "notes": "",
            })

        return rows

    def _build_pod_entries_review_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Build one review row per singles player, doubles pair, or anomaly.

        Doubles are matched by reciprocal partner_name declarations using
        normalized name comparison. Unmatched doubles surface as UnresolvedDoubles.
        Confirmed/provisional status is determined by open ERROR presence.
        """
        error_lookup = self._build_pod_error_lookup(validation_rows)

        # Separate racquet entries by format class.
        singles_rows: List[Dict[str, Any]] = []
        doubles_rows: List[Dict[str, Any]] = []
        anomaly_rows: List[Dict[str, Any]] = []

        for r in roster_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            if sport_type not in RACQUET_SPORTS:
                continue
            fmt_class = self._pod_format_class(str(r.get("sport_format") or ""))
            if fmt_class == "singles":
                singles_rows.append(r)
            elif fmt_class == "doubles":
                doubles_rows.append(r)
            else:
                anomaly_rows.append(r)

        entry_rows: List[Dict[str, Any]] = []
        entry_counter = 0

        def _full_name(r: Dict[str, Any]) -> str:
            return f"{r.get('First Name', '')} {r.get('Last Name', '')}".strip()

        def _pid(r: Dict[str, Any]) -> str:
            return str(r.get("Participant ID (WP)") or r.get("ChMeetings ID") or "").strip()

        def _has_error(r: Dict[str, Any]) -> bool:
            pid = _pid(r)
            sport = str(r.get("sport_type") or "").strip()
            return bool(pid and pid not in ("0",) and sport in error_lookup.get(pid, set()))

        # Singles — one entry per participant.
        for r in singles_rows:
            entry_counter += 1
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            division_id = self._make_division_id(sport_type, sport_gender, "singles")
            entry_rows.append({
                "entry_id": entry_counter,
                "division_id": division_id,
                "entry_type": "Singles",
                "participant_1_name": _full_name(r),
                "participant_2_name": "",
                "source_participant_ids": _pid(r),
                "church_team": str(r.get("Church Team") or ""),
                "partner_status": "N/A",
                "review_status": "NeedsReview" if _has_error(r) else "OK",
                "notes": "",
            })

        # Doubles — attempt reciprocal pairing within each division.
        # Group by (sport_type, sport_gender) across all churches.
        doubles_by_div: Dict[tuple, List[Dict[str, Any]]] = {}
        for r in doubles_rows:
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            doubles_by_div.setdefault((sport_type, sport_gender), []).append(r)

        for (sport_type, sport_gender), div_rows in sorted(doubles_by_div.items()):
            division_id = self._make_division_id(sport_type, sport_gender, "doubles")
            name_to_rows: Dict[str, List[Dict[str, Any]]] = {}
            for r in div_rows:
                key = _norm_name(_full_name(r))
                if key:
                    name_to_rows.setdefault(key, []).append(r)

            paired_pids: set = set()

            for r in div_rows:
                pid_a = _pid(r)
                if pid_a and pid_a in paired_pids:
                    continue

                name_a = _full_name(r)
                partner_decl = str(r.get("partner_name") or "").strip()

                if not partner_decl:
                    entry_counter += 1
                    entry_rows.append({
                        "entry_id": entry_counter,
                        "division_id": division_id,
                        "entry_type": "UnresolvedDoubles",
                        "participant_1_name": name_a,
                        "participant_2_name": "",
                        "source_participant_ids": pid_a,
                        "church_team": str(r.get("Church Team") or ""),
                        "partner_status": "MissingPartner",
                        "review_status": "NeedsReview",
                        "notes": "No partner declared",
                    })
                    if pid_a:
                        paired_pids.add(pid_a)
                    continue

                partner_key = _norm_name(partner_decl)
                candidates = name_to_rows.get(partner_key, [])

                # Look for a reciprocal candidate not yet paired.
                name_a_key = _norm_name(name_a)
                reciprocal = next(
                    (
                        c for c in candidates
                        if _pid(c) not in paired_pids
                        and _norm_name(str(c.get("partner_name") or "")) == name_a_key
                    ),
                    None,
                )

                if reciprocal:
                    pid_b = _pid(reciprocal)
                    entry_counter += 1
                    both_ok = not _has_error(r) and not _has_error(reciprocal)
                    churches = ", ".join(sorted({
                        str(r.get("Church Team") or ""),
                        str(reciprocal.get("Church Team") or ""),
                    }))
                    entry_rows.append({
                        "entry_id": entry_counter,
                        "division_id": division_id,
                        "entry_type": "DoublesPair",
                        "participant_1_name": name_a,
                        "participant_2_name": _full_name(reciprocal),
                        "source_participant_ids": ", ".join(filter(None, [pid_a, pid_b])),
                        "church_team": churches,
                        "partner_status": "Confirmed",
                        "review_status": "OK" if both_ok else "NeedsReview",
                        "notes": "",
                    })
                    if pid_a:
                        paired_pids.add(pid_a)
                    if pid_b:
                        paired_pids.add(pid_b)
                else:
                    # Partner not found or non-reciprocal.
                    if candidates:
                        reason = "NonReciprocal"
                        note = f"Partner '{partner_decl}' found but does not list this player back"
                    else:
                        reason = "PartnerNotFound"
                        note = f"Partner '{partner_decl}' not in same-division roster"
                    entry_counter += 1
                    entry_rows.append({
                        "entry_id": entry_counter,
                        "division_id": division_id,
                        "entry_type": "UnresolvedDoubles",
                        "participant_1_name": name_a,
                        "participant_2_name": "",
                        "source_participant_ids": pid_a,
                        "church_team": str(r.get("Church Team") or ""),
                        "partner_status": reason,
                        "review_status": "NeedsReview",
                        "notes": note,
                    })
                    if pid_a:
                        paired_pids.add(pid_a)

        # Anomalies — non-standard format rows for racquet sports.
        for r in anomaly_rows:
            entry_counter += 1
            sport_type = str(r.get("sport_type") or "").strip()
            sport_gender = str(r.get("sport_gender") or "").strip()
            division_id = self._make_division_id(sport_type, sport_gender, "anomaly")
            entry_rows.append({
                "entry_id": entry_counter,
                "division_id": division_id,
                "entry_type": "Anomaly",
                "participant_1_name": _full_name(r),
                "participant_2_name": "",
                "source_participant_ids": _pid(r),
                "church_team": str(r.get("Church Team") or ""),
                "partner_status": "N/A",
                "review_status": "NeedsReview",
                "notes": f"Unexpected format '{r.get('sport_format', '')}' for racquet sport",
            })

        return entry_rows

    # ── End pod helpers ──────────────────────────────────────────────────────

    def _build_venue_capacity_rows(self, roster_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows = []

        # Team sports — count churches with a complete roster
        for event_name in COURT_ESTIMATE_EVENTS:
            min_team_size = self._get_min_team_size(event_name)
            counts = self._count_estimating_teams(roster_rows, event_name, min_team_size)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(event_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM)
            s = self._compute_court_slots(counts["n_estimating"], minutes_per_game=mpg, pool_games_per_team=gpg)
            rows.append({
                "Event": event_name,
                "Potential Teams/Entries": counts["n_potential"],
                "Estimating Teams/Entries": counts["n_estimating"],
                "Teams": counts["team_codes"],
                "Pool Games Per Team": s["pool_games_per_team"],
                "Minutes Per Game": s["minutes_per_game"],
                "Pool Slots": s["pool_slots"],
                "Playoff Teams": s["playoff_teams"],
                "Playoff Slots": s["playoff_slots"],
                "Third Place?": "Yes" if s["include_third_place"] else "No",
                "Third Place Slots": s["third_place_slots"],
                "Total Court Slots": s["total_slots"],
                "Estimated Court Hours": s["court_hours"],
            })

        # Racquet sports — count complete pairs + singles
        for sport_name in COURT_ESTIMATE_RACQUET_EVENTS:
            counts = self._count_racquet_entries(roster_rows, sport_name)
            mpg = COURT_ESTIMATE_MINUTES_PER_GAME.get(sport_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
            s = self._compute_court_slots(counts["n_estimating"], minutes_per_game=mpg)
            rows.append({
                "Event": sport_name,
                "Potential Teams/Entries": counts["n_potential"],
                "Estimating Teams/Entries": counts["n_estimating"],
                "Teams": counts["team_codes"],
                "Pool Games Per Team": s["pool_games_per_team"],
                "Minutes Per Game": s["minutes_per_game"],
                "Pool Slots": s["pool_slots"],
                "Playoff Teams": s["playoff_teams"],
                "Playoff Slots": s["playoff_slots"],
                "Third Place?": "Yes" if s["include_third_place"] else "No",
                "Third Place Slots": s["third_place_slots"],
                "Total Court Slots": s["total_slots"],
                "Estimated Court Hours": s["court_hours"],
            })

        return rows

    # ── Schedule-Input helpers (Issue #87) ──────────────────────────────────

    def _build_gym_game_objects(
        self, roster_rows: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Return pool-play game placeholder dicts for gym sports (Basketball, VB Men, VB Women).

        Pool games carry stable placeholder team IDs (e.g. BBM-P1-T1, BBM-P1-T2)
        and non-empty pool_id values so the solver can enforce team-overlap and
        min-rest constraints even before final church assignments are known.

        Playoff games (QF/Semi/Final/3rd) are pre-assigned via the Playoff-Slots
        tab in venue_input.xlsx and are not included here.
        """
        sport_defs = [
            (SPORT_TYPE["BASKETBALL"],       "BBM"),
            (SPORT_TYPE["VOLLEYBALL_MEN"],   "VBM"),
            (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW"),
        ]
        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        games: List[Dict[str, Any]] = []

        for event_name, prefix in sport_defs:
            min_sz = self._get_min_team_size(event_name)
            counts = self._count_estimating_teams(roster_rows, event_name, min_sz)
            n_teams = counts["n_estimating"] if counts["n_estimating"] >= 2 else 8
            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
                event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
            )

            # Pool games — stable team IDs and non-empty pool_id
            pool_pairs = self._make_pool_game_pairs(prefix, n_teams, gpg)
            for pair_idx, (team_a_id, team_b_id, pool_id) in enumerate(pool_pairs, start=1):
                games.append({
                    "game_id": f"{prefix}-{pair_idx:02d}",
                    "event": event_name,
                    "stage": "Pool",
                    "pool_id": pool_id,
                    "round": pair_idx,
                    "team_a_id": team_a_id,
                    "team_b_id": team_b_id,
                    "duration_minutes": mpg,
                    "resource_type": GYM_RESOURCE_TYPE,
                    "earliest_slot": None,
                    "latest_slot":   None,
                })

        return games

    def _build_pod_game_objects(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Return single-elimination game placeholder dicts for pod (racquet) sports.

        Uses planning_entries (confirmed + provisional) from Pod-Divisions.
        Number of games per division = planning_entries − 1 (single elimination).
        Divisions with fewer than 2 planning entries are skipped.
        """
        div_rows = self._build_pod_divisions_rows(roster_rows, validation_rows)
        games: List[Dict[str, Any]] = []
        for div in div_rows:
            if div["division_status"] in ("Empty", "AnomalyOnly"):
                continue
            n_entries = div["planning_entries"]
            if n_entries < 2:
                continue
            division_id = div["division_id"]
            sport_type = div["sport_type"]
            resource_type = div["resource_type"]
            mpg = div["minutes_per_game"]
            for i in range(1, n_entries):  # n_entries - 1 games
                games.append({
                    "game_id": f"{division_id}-{i:02d}",
                    "event": sport_type,
                    "stage": "R1",
                    "pool_id": "",
                    "round": i,
                    "team_a_id": None,
                    "team_b_id": None,
                    "duration_minutes": mpg,
                    "resource_type": resource_type,
                    "earliest_slot": None,
                    "latest_slot": None,
                })
        return games

    @staticmethod
    def _build_gym_resource_objects(n_courts: int = 4) -> List[Dict[str, Any]]:
        """Return one resource object per (session × court) for gym sports.

        Four sessions: 1st Saturday, 1st Sunday, 2nd Saturday, 2nd Sunday.
        Time windows are taken from SCHEDULE_SKETCH_* config constants.
        close_time = last game start + slot_minutes (one slot after last start).
        """
        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        close_sat = f"{SCHEDULE_SKETCH_SATURDAY_LAST_GAME + mpg // 60:02d}:00"
        close_sun = f"{SCHEDULE_SKETCH_SUNDAY_LAST_GAME + mpg // 60:02d}:00"
        sessions = [
            ("Sat-1", f"{SCHEDULE_SKETCH_SATURDAY_START:02d}:00", close_sat),
            ("Sun-1", f"{SCHEDULE_SKETCH_SUNDAY_START:02d}:00",   close_sun),
            ("Sat-2", f"{SCHEDULE_SKETCH_SATURDAY_START:02d}:00", close_sat),
            ("Sun-2", f"{SCHEDULE_SKETCH_SUNDAY_START:02d}:00",   close_sun),
        ]
        resources: List[Dict[str, Any]] = []
        for day_label, open_time, close_time in sessions:
            for c in range(1, n_courts + 1):
                resources.append({
                    "resource_id":   f"GYM-{day_label}-{c}",
                    "resource_type": GYM_RESOURCE_TYPE,
                    "label":         f"Court-{c}",
                    "day":           day_label,
                    "open_time":     open_time,
                    "close_time":    close_time,
                    "slot_minutes":  mpg,
                    "exclusive_group": "",
                })
        return resources

    @staticmethod
    def _clean_excel_text(val) -> str:
        """Normalize spreadsheet cells so blanks/NaN become an empty string."""
        if pd.isna(val):
            return ""
        return str(val).strip()

    @staticmethod
    def _float_from_excel(val, default: float) -> float:
        """Convert spreadsheet cells to float while treating blanks/NaN as a default."""
        if pd.isna(val) or val in (None, ""):
            return default
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _load_venue_input_rows(venue_input_path: Path) -> List[Dict[str, Any]]:
        """Expand venue_input.xlsx into per-resource objects for schedule_input.json.

        Each row with Quantity=N emits N resource objects labelled Court-1…N or
        Table-1…N.  Returns an empty list if the file does not exist.
        """
        if not venue_input_path.exists():
            return []
        try:
            df = pd.read_excel(
                venue_input_path, sheet_name="Venue-Input", engine="openpyxl"
            )
        except Exception as e:
            logger.warning(f"Could not read venue input rows from {venue_input_path}: {e}")
            return []

        rows: List[Dict[str, Any]] = []
        resource_counts: Dict[str, int] = {}

        for _, row in df.iterrows():
            resource_type = ChurchTeamsExporter._clean_excel_text(row.get("Resource Type"))
            if not resource_type:
                continue
            # Exclusive Venue Group: rows sharing a group value compete for the
            # same physical gym (only one mode active per time block). Optional
            # column — blank means the resource stands alone.
            exclusive_group = ChurchTeamsExporter._clean_excel_text(
                row.get("Exclusive Venue Group")
            )
            qty = max(1, int(ChurchTeamsExporter._float_from_excel(row.get("Quantity"), 1)))
            slot_min = max(1, int(ChurchTeamsExporter._float_from_excel(row.get("Slot Minutes"), 60)))
            start = ChurchTeamsExporter._parse_hour(row.get("Start Time"))
            last_start = ChurchTeamsExporter._parse_hour(row.get("Last Start Time"))
            open_time = f"{int(start):02d}:{int(round((start % 1) * 60)):02d}"
            close_decimal = last_start + slot_min / 60.0
            close_time = f"{int(close_decimal):02d}:{int(round((close_decimal % 1) * 60)):02d}"

            abbrev = resource_type.split()[0][:3].upper()
            rc = resource_counts.get(resource_type, 0)

            for i in range(1, qty + 1):
                rc += 1
                label = (
                    f"Table-{i}" if "table" in resource_type.lower() else f"Court-{i}"
                )
                rows.append({
                    "resource_id":     f"{abbrev}-{rc}",
                    "resource_type":   resource_type,
                    "label":           label,
                    "day":             "Day-1",
                    "open_time":       open_time,
                    "close_time":      close_time,
                    "slot_minutes":    slot_min,
                    "exclusive_group": exclusive_group,
                })
            resource_counts[resource_type] = rc

        logger.debug(f"Loaded {len(rows)} venue resource rows from {venue_input_path}")
        return rows

    @staticmethod
    def _load_playoff_slots(venue_input_path: Path) -> List[Dict[str, Any]]:
        """Load pre-assigned playoff game slots from the Playoff-Slots tab in venue_input.xlsx.

        Returns an empty list (with a WARNING) if the file or tab is absent.
        Expected columns: game_id, event, stage, resource_id, slot
        Optional columns: team_a_id, team_b_id, duration_minutes
        """
        if not venue_input_path.exists():
            return []
        try:
            df = pd.read_excel(venue_input_path, sheet_name="Playoff-Slots", engine="openpyxl")
        except Exception:
            logger.warning(
                "venue_input.xlsx is present but has no 'Playoff-Slots' tab — "
                "playoff games will not appear in the schedule. "
                "Add a 'Playoff-Slots' tab to include them."
            )
            return []

        required = {"game_id", "event", "stage", "resource_id", "slot"}
        cols = {str(c).strip() for c in df.columns}
        missing = required - cols
        if missing:
            logger.warning(
                f"Playoff-Slots tab is missing required columns {sorted(missing)}; "
                "playoff games will not appear in the schedule."
            )
            return []

        slots: List[Dict[str, Any]] = []
        for _, row in df.iterrows():
            game_id = ChurchTeamsExporter._clean_excel_text(row.get("game_id"))
            if not game_id:
                continue
            entry: Dict[str, Any] = {
                "game_id":     game_id,
                "event":       ChurchTeamsExporter._clean_excel_text(row.get("event", "")),
                "stage":       ChurchTeamsExporter._clean_excel_text(row.get("stage", "")),
                "resource_id": ChurchTeamsExporter._clean_excel_text(row.get("resource_id", "")),
                "slot":        ChurchTeamsExporter._clean_excel_text(row.get("slot", "")),
            }
            for optional in ("team_a_id", "team_b_id", "duration_minutes"):
                val = row.get(optional)
                if val is not None and str(val).strip() not in ("", "nan"):
                    entry[optional] = ChurchTeamsExporter._clean_excel_text(str(val)) if optional != "duration_minutes" else int(val)
            if entry["resource_id"] and entry["slot"]:
                slots.append(entry)
            else:
                logger.warning(f"Playoff-Slots row for {game_id!r} missing resource_id or slot; skipped.")
        return slots

    @staticmethod
    def _load_gym_modes(venue_input_path: Path) -> Dict[str, Dict[str, int]]:
        """Load per-gym mode capacities from the Gym-Modes tab in venue_input.xlsx.

        A gym that can be configured as either-or (e.g. 1 Basketball Court OR
        2 Volleyball Courts per time block) records both options on one row.
        Returns {gym_name: {resource_type: courts_per_block}}; 0 means that
        mode is not available in that gym.

        Returns an empty dict (with a WARNING) if the file or tab is absent —
        the schedule is still produced; the gym-mode capacity estimator simply
        has no mode data to work with.
        """
        # Maps a Gym-Modes column header to the resource_type it represents.
        mode_column_map = {
            "Basketball Courts": "Basketball Court",
            "Volleyball Courts": "Volleyball Court",
            "Badminton Courts":  "Badminton Court",
            "Pickleball Courts": "Pickleball Court",
            "Soccer Fields":     "Soccer Field",
        }
        if not venue_input_path.exists():
            return {}
        try:
            df = pd.read_excel(venue_input_path, sheet_name="Gym-Modes", engine="openpyxl")
        except Exception:
            logger.warning(
                "venue_input.xlsx is present but has no 'Gym-Modes' tab — "
                "gym-mode capacity estimation will be skipped. "
                "Add a 'Gym-Modes' tab to enable it."
            )
            return {}

        df = df.rename(columns=lambda c: str(c).strip())
        cols = set(df.columns)
        if "Gym Name" not in cols:
            logger.warning(
                "Gym-Modes tab is missing the 'Gym Name' column — "
                "gym-mode capacity estimation will be skipped."
            )
            return {}

        active_modes = {col: rt for col, rt in mode_column_map.items() if col in cols}
        if not active_modes:
            logger.warning(
                "Gym-Modes tab has no recognized mode columns "
                f"({sorted(mode_column_map)}); gym-mode capacity estimation "
                "will be skipped."
            )
            return {}

        gym_modes: Dict[str, Dict[str, int]] = {}
        for _, row in df.iterrows():
            gym_name = ChurchTeamsExporter._clean_excel_text(row.get("Gym Name"))
            if not gym_name:
                continue
            capacities = {
                rt: int(ChurchTeamsExporter._float_from_excel(row.get(col), 0))
                for col, rt in active_modes.items()
            }
            # Skip note/blank rows — a "gym" with zero capacity in every mode
            # is the documentation footer, not a real venue.
            if not any(capacities.values()):
                continue
            gym_modes[gym_name] = capacities

        logger.debug(f"Loaded {len(gym_modes)} gym mode rows from {venue_input_path}")
        return gym_modes

    def _build_schedule_input(
        self,
        roster_rows: List[Dict[str, Any]],
        validation_rows: List[Dict[str, Any]],
        venue_input_path: Path,
    ) -> Dict[str, Any]:
        """Assemble the full schedule_input package consumed by OR-Tools.

        Returns a dict with keys: generated_at, gym_court_scenario, game_count,
        resource_count, games, resources, playoff_slots, gym_modes.

        Gym resources are built from the explicit SCHEDULE_SOLVER_GYM_COURTS
        constant (config.py) so the solver knows exactly which court scenario
        was chosen for this run.
        """
        gym_games = self._build_gym_game_objects(roster_rows)
        pod_games = self._build_pod_game_objects(roster_rows, validation_rows)
        all_games = gym_games + pod_games

        gym_resources = self._build_gym_resource_objects(SCHEDULE_SOLVER_GYM_COURTS)
        pod_resources = self._load_venue_input_rows(venue_input_path)
        all_resources = gym_resources + pod_resources

        playoff_slots = self._load_playoff_slots(venue_input_path)
        gym_modes = self._load_gym_modes(venue_input_path)

        return {
            "generated_at":       datetime.now().isoformat(timespec="seconds"),
            "gym_court_scenario": SCHEDULE_SOLVER_GYM_COURTS,
            "game_count":         len(all_games),
            "resource_count":     len(all_resources),
            "games":              all_games,
            "resources":          all_resources,
            "playoff_slots":      playoff_slots,
            "gym_modes":          gym_modes,
        }

    @staticmethod
    def _write_schedule_input_tab(ws, schedule_input: Dict[str, Any]) -> None:
        """Write Schedule-Input tab with Games, Resources, and Playoff-Slots sections."""
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        hdr_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER, fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        sec_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
        sec_font = Font(bold=True)

        game_cols = [
            "game_id", "event", "stage", "pool_id", "round",
            "team_a_id", "team_b_id", "duration_minutes",
            "resource_type", "earliest_slot", "latest_slot",
        ]
        resource_cols = [
            "resource_id", "resource_type", "label", "day",
            "open_time", "close_time", "slot_minutes", "exclusive_group",
        ]
        playoff_slot_cols = ["game_id", "event", "stage", "resource_id", "slot"]

        current_row = 1

        # Meta row
        ws.cell(row=current_row, column=1, value="generated_at").font = sec_font
        ws.cell(row=current_row, column=2, value=schedule_input["generated_at"])
        ws.cell(
            row=current_row, column=3,
            value=f"Games: {schedule_input['game_count']}  Resources: {schedule_input['resource_count']}",
        )
        current_row += 2

        def _write_section(title: str, cols: List[str], rows: List[Dict]) -> None:
            nonlocal current_row
            sec_cell = ws.cell(row=current_row, column=1, value=title)
            sec_cell.fill = sec_fill
            sec_cell.font = sec_font
            current_row += 1
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=current_row, column=c_idx, value=col)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            current_row += 1
            for data_row in rows:
                for c_idx, col in enumerate(cols, start=1):
                    ws.cell(row=current_row, column=c_idx, value=data_row.get(col))
                current_row += 1
            current_row += 1  # blank separator

        playoff_slots = schedule_input.get("playoff_slots", [])
        playoff_note_rows = (
            playoff_slots if playoff_slots
            else [{"game_id": "No playoff slots loaded — add Playoff-Slots tab to venue_input.xlsx"}]
        )

        gym_modes = schedule_input.get("gym_modes", {})
        gym_mode_rtypes = sorted({rt for caps in gym_modes.values() for rt in caps})
        gym_mode_cols = ["gym_name", *gym_mode_rtypes]
        gym_mode_rows = (
            [{"gym_name": name, **caps} for name, caps in sorted(gym_modes.items())]
            if gym_modes
            else [{"gym_name": "No Gym-Modes tab loaded — add Gym-Modes tab to venue_input.xlsx"}]
        )

        _write_section("GAMES",          game_cols,          schedule_input["games"])
        _write_section("RESOURCES",      resource_cols,      schedule_input["resources"])
        _write_section("PLAYOFF-SLOTS",  playoff_slot_cols,  playoff_note_rows)
        _write_section("GYM-MODES",      gym_mode_cols,      gym_mode_rows)

        # Column widths
        col_widths = [20, 30, 10, 10, 8, 16, 16, 18, 22, 14, 12]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── End Schedule-Input helpers ───────────────────────────────────────────

    @staticmethod
    def _build_scenario_schedule(
        n_courts: int,
        pool_queues: List[List[str]],
        early_playoff_queues: List[List[str]],
        final_queues: List[List[str]],
        n_sat: int,
        n_sun: int,
    ) -> List[List[List[str]]]:
        """
        Build a 4-session court schedule for a given number of courts.

        Returns a list of 4 session grids:
            grids[0] = 1st Saturday  (n_sat time slots)
            grids[1] = 1st Sunday    (n_sun time slots)
            grids[2] = 2nd Saturday  (n_sat time slots)
            grids[3] = 2nd Sunday    (n_sun time slots)
        Each grid is grids[session][time_slot][court_index] = game_id or "".

        ── Court allocation ──────────────────────────────────────────────────
        Courts are divided into contiguous "primary blocks", one per sport,
        allocated proportionally.  Remainder courts go to the first sport(s)
        (i.e., Basketball gets an extra court before Volleyball does).

        Example with 5 courts and 3 sports:
            base = 5 // 3 = 1, extras = 5 % 3 = 2
            BBM → courts [0, 1]   (base 1 + 1 extra)
            VBM → courts [2, 3]   (base 1 + 1 extra)
            VBW → courts [4]      (base 1, no extra)

        Rationale: keeps each court dedicated to one sport type, so no
        net-height adjustment or equipment swap is needed mid-court.

        ── Phase 1 — Pool fill (sat1 → sun1 → sat2) ─────────────────────────
        For every time slot, courts are visited left-to-right (court 0, 1, …):

        • If the primary sport for that court still has pool games left,
          place the next game there (primary-first rule).
        • If the primary sport has finished its pool games, the court is
          idle.  The idle court is given to whichever sport currently has
          the most remaining pool games (greedy-most-needy rule).

        Effect in 5-court scenario (equal teams, 12 pool games each):
            Slots 0–5   : BBM fills courts 0-1, VBM fills courts 2-3,
                          VBW fills court 4  (all 3 sports running in parallel)
            Slot 6+     : BBM and VBM are done; their 4 courts become idle.
                          VBW still has 6 games → claims all 4 idle courts
                          plus its own, running 5 VBW games simultaneously.
                          VBW finishes at slot 7 (≈15:00) instead of slot 11
                          (≈19:00) — the whole church leaves ~4 hours earlier.

        Pool games never spill into sun2; that session is reserved for finals.

        ── Phase 2 — Early playoffs (QF + Semis) on sat2 ───────────────────
        After pool fill, each sport's empty cells in sat2 are collected in
        (time_slot, court) order.  Early-round playoff games (QF-1…4 if 8
        playoff teams, Semi-1 and Semi-2 otherwise) are placed there.

        Playoffs are placed on the sport's primary courts only — no court
        sharing — so the same nets and equipment remain in place.

        ── Phase 3 — Finals on sun2 ─────────────────────────────────────────
        Final and 3rd-place games are placed on each sport's empty cells in
        sun2, again primary courts only.  This guarantees that championship
        games always fall on the last day of the festival, regardless of how
        pool play distributes across the earlier sessions.

        ── Changing the algorithm ───────────────────────────────────────────
        • Court count scenarios: edit SCHEDULE_SKETCH_N_COURTS in config.py.
        • Session hours: edit SCHEDULE_SKETCH_SATURDAY_START / LAST_GAME and
          SCHEDULE_SKETCH_SUNDAY_START / LAST_GAME in config.py.
        • Court allocation order: the sport order in sport_defs inside
          _write_court_schedule_sketch controls which sport gets extra courts
          (earlier in the list = higher priority for extras).
        • Pool overflow policy: replace the greedy-most-needy rule (the
          `max(range(n_sports), key=lambda i: len(pool_remaining[i]))` line)
          with any other priority function — e.g., fixed sport priority,
          round-robin, or "same-sport block only" to revert to strict blocks.
        • Playoff session assignment: swap early_playoff_queues and
          final_queues arguments, or add a third category (e.g. Semis on
          sun1) by adding a new fill phase following the same pattern.
        """
        n_sports = len(pool_queues)
        n_slots = [n_sat, n_sun, n_sat, n_sun]
        grids: List[List[List[str]]] = [
            [[""] * n_courts for _ in range(n)] for n in n_slots
        ]

        # Court block allocation
        base = n_courts // n_sports
        extras = n_courts % n_sports
        court_blocks: List[List[int]] = []
        cur = 0
        for i in range(n_sports):
            k = base + (1 if i < extras else 0)
            court_blocks.append(list(range(cur, cur + k)))
            cur += k

        court_to_primary = {c: i for i, courts in enumerate(court_blocks) for c in courts}

        # Phase 1: pool fill — primary-first, then greedy-most-needy for idle courts
        pool_remaining = [deque(q) for q in pool_queues]
        for sess_idx in range(3):  # sat1, sun1, sat2
            for t in range(n_slots[sess_idx]):
                for c in range(n_courts):
                    primary = court_to_primary[c]
                    if pool_remaining[primary]:
                        grids[sess_idx][t][c] = pool_remaining[primary].popleft()
                    else:
                        most_needy = max(range(n_sports), key=lambda i: len(pool_remaining[i]))
                        if pool_remaining[most_needy]:
                            grids[sess_idx][t][c] = pool_remaining[most_needy].popleft()

        # Phase 2: early playoffs (QF + Semi) on primary courts in sat2
        for early_q, courts in zip(early_playoff_queues, court_blocks):
            cells = [
                (2, t, c)
                for t in range(n_sat)
                for c in courts
                if not grids[2][t][c]
            ]
            for i, game_id in enumerate(early_q):
                if i < len(cells):
                    s, t, c = cells[i]
                    grids[s][t][c] = game_id

        # Phase 3: finals (Final + 3rd) on primary courts in sun2
        for final_q, courts in zip(final_queues, court_blocks):
            cells = [
                (3, t, c)
                for t in range(n_sun)
                for c in courts
                if not grids[3][t][c]
            ]
            for i, game_id in enumerate(final_q):
                if i < len(cells):
                    s, t, c = cells[i]
                    grids[s][t][c] = game_id

        return grids

    @staticmethod
    @staticmethod
    def _make_pool_game_pairs(
        prefix: str, n_teams: int, gpg: int
    ) -> List[Tuple[str, str, str]]:
        """Return (team_a_id, team_b_id, pool_id) tuples for pool-play games.

        Teams are split into balanced pools of size (gpg+1).  A full round-robin
        within a pool of that size gives each team exactly gpg games.  Edge pools
        (when n_teams is not divisible by gpg+1) produce fewer games for those teams.

        Team IDs are stable planning placeholders: {prefix}-P{pool}-T{slot}.
        The same placeholder is reused across all games involving that team,
        allowing the solver to enforce team-overlap and min-rest constraints.
        """
        if n_teams < 2:
            return []
        target_pool_size = max(2, gpg + 1)
        n_pools = max(1, -(-n_teams // target_pool_size))  # ceil division

        pools: List[List[int]] = [[] for _ in range(n_pools)]
        for i in range(n_teams):
            pools[i % n_pools].append(i + 1)  # 1-indexed team number

        team_id: Dict[int, str] = {}
        for p_idx, pool_teams in enumerate(pools, start=1):
            for t_idx, team_num in enumerate(pool_teams, start=1):
                team_id[team_num] = f"{prefix}-P{p_idx}-T{t_idx}"

        pairs: List[Tuple[str, str, str]] = []
        for p_idx, pool_teams in enumerate(pools, start=1):
            pool_id = f"P{p_idx}"
            for i in range(len(pool_teams)):
                for j in range(i + 1, len(pool_teams)):
                    pairs.append((
                        team_id[pool_teams[i]],
                        team_id[pool_teams[j]],
                        pool_id,
                    ))
        return pairs

    @staticmethod
    def _make_playoff_ids(
        prefix: str, playoff_teams: int, include_third: bool
    ) -> Tuple[List[str], List[str]]:
        """Return (early_ids, final_ids) split by which weekend they belong to.

        early_ids  — QF + Semi games, scheduled on 2nd Saturday.
        final_ids  — Final (+ optional 3rd-place), scheduled on 2nd Sunday.

        Bracket size is determined by playoff_teams (from COURT_ESTIMATE_PLAYOFF_RULES):
            0 teams  → no playoff games
            4 teams  → Semi-1, Semi-2 | Final [+ 3rd]
            8 teams  → QF-1…4, Semi-1, Semi-2 | Final [+ 3rd]

        To add a new bracket size (e.g. 16 teams with quarter-finals already
        called Round-of-16), extend the if/elif chain here and add matching
        rows to COURT_ESTIMATE_PLAYOFF_RULES in config.py.
        """
        early_ids: List[str] = []
        if playoff_teams >= 8:
            for i in range(1, 5):
                early_ids.append(f"{prefix}-QF-{i}")
            early_ids.extend([f"{prefix}-Semi-1", f"{prefix}-Semi-2"])
        elif playoff_teams >= 4:
            early_ids.extend([f"{prefix}-Semi-1", f"{prefix}-Semi-2"])

        final_ids: List[str] = []
        if playoff_teams >= 4:
            final_ids.append(f"{prefix}-Final")
            if include_third:
                final_ids.append(f"{prefix}-3rd")
        return early_ids, final_ids

    def _write_court_schedule_sketch(
        self, ws, roster_rows: List[Dict[str, Any]]
    ) -> None:
        """
        Write the Court-Schedule-Sketch tab.

        Three scenarios (3, 4, 5 courts) are rendered side-by-side on one
        worksheet, separated by an empty column.  Game IDs are sequential
        placeholders (BBM01…, VBM01…, VBW01…); no actual team assignments
        or conflict enforcement is performed here.  This is an Excel-only
        planning artifact — no data is written to WordPress sf_schedules.
        """
        from openpyxl.styles import PatternFill, Font, Alignment

        mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
        include_third = COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME

        # Sports covered by this sketch (shared court type: basketball / volleyball)
        sport_defs = [
            (SPORT_TYPE["BASKETBALL"],       "BBM", SCHEDULE_SKETCH_COLOR_BASKETBALL),
            (SPORT_TYPE["VOLLEYBALL_MEN"],   "VBM", SCHEDULE_SKETCH_COLOR_VB_MEN),
            (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW", SCHEDULE_SKETCH_COLOR_VB_WOMEN),
        ]

        # --- Compute game IDs per sport (per-sport pool games per team) ---
        sport_meta: Dict[str, Dict] = {}
        for event_name, prefix, color in sport_defs:
            min_sz = self._get_min_team_size(event_name)
            counts = self._count_estimating_teams(roster_rows, event_name, min_sz)
            n_teams = counts["n_estimating"] if counts["n_estimating"] >= 2 else 8
            gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM)
            s = self._compute_court_slots(n_teams, mpg, pool_games_per_team=gpg)
            early_ids, final_ids = self._make_playoff_ids(
                prefix, s["playoff_teams"], include_third
            )
            sport_meta[event_name] = {
                "prefix": prefix,
                "color": color,
                "n_teams": n_teams,
                "pool_gpg": gpg,
                "pool_ids":   [f"{prefix}-{i:02d}" for i in range(1, s["pool_slots"] + 1)],
                "early_ids":  early_ids,   # QF + Semi → 2nd Saturday
                "final_ids":  final_ids,   # Final + 3rd → 2nd Sunday
            }

        # --- Per-sport game queues (pool overflow + dedicated playoff courts) ---
        pool_queues_by_sport          = [sport_meta[ev]["pool_ids"]  for ev, _, _ in sport_defs]
        early_playoff_queues_by_sport = [sport_meta[ev]["early_ids"] for ev, _, _ in sport_defs]
        final_queues_by_sport         = [sport_meta[ev]["final_ids"] for ev, _, _ in sport_defs]

        # --- Time slot helpers ---
        n_sat = SCHEDULE_SKETCH_SATURDAY_LAST_GAME - SCHEDULE_SKETCH_SATURDAY_START + 1
        n_sun = SCHEDULE_SKETCH_SUNDAY_LAST_GAME - SCHEDULE_SKETCH_SUNDAY_START + 1
        sat_times = [
            f"{h:02d}:00"
            for h in range(SCHEDULE_SKETCH_SATURDAY_START, SCHEDULE_SKETCH_SATURDAY_LAST_GAME + 1)
        ]
        sun_times = [
            f"{h:02d}:00"
            for h in range(SCHEDULE_SKETCH_SUNDAY_START, SCHEDULE_SKETCH_SUNDAY_LAST_GAME + 1)
        ]
        sessions = [
            ("1st Saturday", sat_times),
            ("1st Sunday",   sun_times),
            ("2nd Saturday", sat_times),
            ("2nd Sunday",   sun_times),
        ]

        # --- Styles ---
        section_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
        hdr_fill     = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER,  fill_type="solid")
        hdr_font     = Font(bold=True, color="FFFFFF")
        bold_font    = Font(bold=True)
        center       = Alignment(horizontal="center", vertical="center")
        prefix_fill  = {
            "BBM": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_BASKETBALL, fill_type="solid"),
            "VBM": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_VB_MEN,     fill_type="solid"),
            "VBW": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_VB_WOMEN,   fill_type="solid"),
        }

        # --- Column layout ---
        # Scenario A (3 cts): col 1=Time, 2-4=Courts → 4 cols; gap col 5
        # Scenario B (4 cts): col 6=Time, 7-10=Courts → 5 cols; gap col 11
        # Scenario C (5 cts): col 12=Time, 13-17=Courts → 6 cols
        n_courts_list = SCHEDULE_SKETCH_N_COURTS
        scenario_starts: List[int] = []
        cur_col = 1
        for n in n_courts_list:
            scenario_starts.append(cur_col)
            cur_col += (1 + n) + 1  # time col + court cols + gap

        INPUTS_ROW      = 1
        SCENARIO_HDR_ROW = 3
        COL_HDR_ROW     = 4
        DATA_START_ROW  = 5

        # --- Row 1: inputs summary (per-sport pool games per team) ---
        ws.cell(row=INPUTS_ROW, column=1, value="Inputs:").font = bold_font
        col = 2
        for ev, prefix, _ in sport_defs:
            ws.cell(row=INPUTS_ROW, column=col,
                    value=f"{prefix} pool games/team: {sport_meta[ev]['pool_gpg']}")
            col += 3
        ws.cell(row=INPUTS_ROW, column=col,     value=f"Minutes/game: {mpg}")
        ws.cell(row=INPUTS_ROW, column=col + 3, value=f"3rd place: {'Yes' if include_third else 'No'}")

        # --- Row 2: per-sport game counts ---
        ws.cell(row=2, column=1, value="Game totals:").font = bold_font
        col_offset = 2
        for ev, prefix, _ in sport_defs:
            meta = sport_meta[ev]
            total = len(meta["pool_ids"]) + len(meta["early_ids"]) + len(meta["final_ids"])
            label = f"{prefix}: {meta['n_teams']} teams, {total} games ({len(meta['pool_ids'])} pool)"
            ws.cell(row=2, column=col_offset, value=label)
            col_offset += 5

        # --- Pre-compute per-scenario schedules ---
        scenario_grids: Dict[int, List[List[List[str]]]] = {}
        for n_courts in n_courts_list:
            scenario_grids[n_courts] = self._build_scenario_schedule(
                n_courts,
                pool_queues_by_sport,
                early_playoff_queues_by_sport,
                final_queues_by_sport,
                n_sat, n_sun,
            )

        # --- Render scenario headers and column headers ---
        for n_courts, start_col in zip(n_courts_list, scenario_starts):
            end_col = start_col + n_courts  # time col + n court cols (inclusive)
            # Scenario header
            sc_cell = ws.cell(row=SCENARIO_HDR_ROW, column=start_col, value=f"Scenario: {n_courts} Courts")
            sc_cell.font = hdr_font
            sc_cell.fill = hdr_fill
            sc_cell.alignment = center
            ws.merge_cells(
                start_row=SCENARIO_HDR_ROW, start_column=start_col,
                end_row=SCENARIO_HDR_ROW,   end_column=end_col,
            )
            # Column sub-headers
            t_cell = ws.cell(row=COL_HDR_ROW, column=start_col, value="Time")
            t_cell.font = bold_font
            for c in range(n_courts):
                ct_cell = ws.cell(row=COL_HDR_ROW, column=start_col + 1 + c, value=f"Court {c + 1}")
                ct_cell.font = bold_font
                ct_cell.alignment = center

        # --- Render session sections and time-slot rows ---
        current_row = DATA_START_ROW
        for sess_idx, (sess_label, times) in enumerate(sessions):
            # Section header
            for n_courts, start_col in zip(n_courts_list, scenario_starts):
                end_col = start_col + n_courts
                sh_cell = ws.cell(row=current_row, column=start_col, value=sess_label)
                sh_cell.fill = section_fill
                sh_cell.font = bold_font
                sh_cell.alignment = center
                ws.merge_cells(
                    start_row=current_row, start_column=start_col,
                    end_row=current_row,   end_column=end_col,
                )
            current_row += 1

            # Time slot rows
            for t, time_str in enumerate(times):
                for n_courts, start_col in zip(n_courts_list, scenario_starts):
                    ws.cell(row=current_row, column=start_col, value=time_str)
                    grid = scenario_grids[n_courts]
                    for c in range(n_courts):
                        game_id = grid[sess_idx][t][c]
                        cell = ws.cell(row=current_row, column=start_col + 1 + c, value=game_id)
                        if game_id:
                            fill = prefix_fill.get(game_id.split("-")[0])
                            if fill:
                                cell.fill = fill
                current_row += 1

        # --- Column widths ---
        from openpyxl.utils import get_column_letter
        for n_courts, start_col in zip(n_courts_list, scenario_starts):
            ws.column_dimensions[get_column_letter(start_col)].width = 10      # Time
            for c in range(n_courts):
                ws.column_dimensions[get_column_letter(start_col + 1 + c)].width = 12  # Courts

        total_pool  = sum(len(q) for q in pool_queues_by_sport)
        total_early = sum(len(q) for q in early_playoff_queues_by_sport)
        total_final = sum(len(q) for q in final_queues_by_sport)
        logger.debug(
            f"Court-Schedule-Sketch tab: {total_pool} pool + {total_early} early-playoff "
            f"+ {total_final} finals across {len(n_courts_list)} scenarios."
        )

    # ── Pod-Resource-Estimate helpers (Issue #86) ──────────────────────────────

    @staticmethod
    def _parse_hour(val) -> float:
        """Convert a cell value to a decimal hour (e.g. datetime.time(13,0) → 13.0)."""
        import datetime as _dt
        if pd.isna(val):
            return 0.0
        if isinstance(val, _dt.time):
            return val.hour + val.minute / 60.0
        try:
            return float(val)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _load_venue_input(venue_input_path: Path) -> Dict[str, int]:
        """Read venue_input.xlsx and return {resource_type: total_available_slots}.

        Each row may have Available Slots pre-computed (formula or number).
        If Available Slots is missing/zero, falls back to computing from
        Quantity, Start Time, Last Start Time, and Slot Minutes.
        Returns an empty dict if the file does not exist.
        """
        if not venue_input_path.exists():
            return {}
        try:
            df = pd.read_excel(venue_input_path, sheet_name="Venue-Input", engine="openpyxl")
        except Exception as e:
            logger.warning(f"Could not read venue input file {venue_input_path}: {e}")
            return {}

        totals: Dict[str, int] = {}
        for _, row in df.iterrows():
            resource_type = ChurchTeamsExporter._clean_excel_text(row.get("Resource Type"))
            if not resource_type:
                continue
            avail = row.get("Available Slots")
            if pd.isna(avail) or not avail:
                # Formula wasn't cached — compute from component columns.
                qty       = ChurchTeamsExporter._float_from_excel(row.get("Quantity"), 0)
                start     = ChurchTeamsExporter._parse_hour(row.get("Start Time"))
                last_start = ChurchTeamsExporter._parse_hour(row.get("Last Start Time"))
                slot_min  = ChurchTeamsExporter._float_from_excel(row.get("Slot Minutes"), 1)
                if slot_min > 0 and qty > 0 and last_start >= start:
                    avail = qty * ((last_start - start) * 60 / slot_min + 1)
                else:
                    avail = 0
            totals[resource_type] = totals.get(resource_type, 0) + int(
                ChurchTeamsExporter._float_from_excel(avail, 0)
            )
        logger.debug(f"Loaded venue input: {totals}")
        return totals

    def _build_pod_resource_rows(
        self,
        roster_rows: List[Dict[str, Any]],
        available_by_resource: Dict[str, int],
    ) -> List[Dict[str, Any]]:
        """Build Pod-Resource-Estimate output rows.

        Required slots use single-elimination: entries - 1
        (doubles counted as complete pairs, same as _count_racquet_entries).
        """
        rows = []
        for sport_name in COURT_ESTIMATE_RACQUET_EVENTS:
            counts = self._count_racquet_entries(roster_rows, sport_name)
            n = counts["n_estimating"]
            resource_type = POD_RESOURCE_EVENT_TYPE.get(sport_name, "")
            required = max(0, n - 1)
            available = available_by_resource.get(resource_type, 0)
            surplus = available - required
            if not available_by_resource:
                fit_status = "No venue data"
            elif surplus >= 0:
                fit_status = "Green"
            elif surplus >= -POD_FIT_YELLOW_MAX:
                fit_status = "Yellow"
            else:
                fit_status = "Red"
            rows.append({
                "Event":              sport_name,
                "Resource Type":      resource_type,
                "Entries / Teams":    n,
                "Required Slots":     required,
                "Available Slots":    available,
                "Surplus / Shortage": surplus,
                "Fit Status":         fit_status,
            })
        return rows

    def _write_pod_resource_estimate(
        self,
        ws,
        pod_rows: List[Dict[str, Any]],
        available_by_resource: Dict[str, int],
    ) -> None:
        """Write Pod-Resource-Estimate tab content with colour-coded Fit Status."""
        from openpyxl.styles import PatternFill, Font, Alignment

        cols = ["Event", "Resource Type", "Entries / Teams",
                "Required Slots", "Available Slots", "Surplus / Shortage", "Fit Status"]

        header_fill = PatternFill("solid", fgColor=SCHEDULE_SKETCH_COLOR_HEADER)
        header_font = Font(color="FFFFFF", bold=True)

        # Header row
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        fit_colors = {
            "Green":  POD_FIT_COLOR_GREEN,
            "Yellow": POD_FIT_COLOR_YELLOW,
            "Red":    POD_FIT_COLOR_RED,
        }

        if not available_by_resource:
            notice = (
                "No venue input loaded — "
                f"create {VENUE_INPUT_FILENAME} from the template and re-run the export"
            )
            ws.cell(row=2, column=1, value=notice)
            for c_idx, col in enumerate(cols, start=1):
                row_cell = ws.cell(row=2, column=c_idx)
                if c_idx == 1:
                    row_cell.value = notice
                else:
                    row_cell.value = None

        for r_idx, row in enumerate(pod_rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
                if col in ("Entries / Teams", "Required Slots", "Available Slots",
                           "Surplus / Shortage"):
                    cell.alignment = Alignment(horizontal="right")
                if col == "Fit Status":
                    color = fit_colors.get(row["Fit Status"])
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
                    cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22
        for letter in ["C", "D", "E", "F", "G"]:
            ws.column_dimensions[letter].width = 16

        # Snapshot note
        note_row = len(pod_rows) + 3
        ws.cell(
            row=note_row, column=1,
            value=(
                f"Venue data loaded from {VENUE_INPUT_FILENAME}. "
                "Required = entries − 1 (single elimination). "
                f"Green ≥ 0 | Yellow short 1–{POD_FIT_YELLOW_MAX} | Red short {POD_FIT_YELLOW_MAX + 1}+."
            ),
        )
        logger.debug(f"Pod-Resource-Estimate tab: {len(pod_rows)} rows.")

    def _write_excel_report(self, filepath: Path,
                            summary_rows: List[Dict[str, Any]],
                            contacts_rows: List[Dict[str, Any]],
                            roster_rows: List[Dict[str, Any]],
                            validation_rows: List[Dict[str, Any]],
                            include_venue_capacity: bool = False):
        """Writes the collected data to an Excel file with specified tabs and formatting."""
        logger.info(f"Writing Excel report to: {filepath}")
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Summary Tab
                df_summary = pd.DataFrame(summary_rows)
                if not df_summary.empty:
                    summary_cols = [
                        "Church Code", "Total Members (ChM Team Group)", "Total Participants (in WP)",
                        "Total Approved (WP)", "Total Pending Approval (WP)", "Total Denied (WP)",
                        "Total Participants w/ Open ERRORs (WP)",
                        "Total Open Individual ERRORs (WP)", "Total Open TEAM ERRORs (WP)",
                        "Total Open WARNINGs (WP)", "Total Sports w/ Open TEAM Issues (WP)",
                        "Total Athlete Fees",
                        "Latest ChM Record Update for Team"
                    ]
                    # Ensure all summary columns exist
                    for col in summary_cols:
                        if col not in df_summary.columns:
                            df_summary[col] = None 
                    df_summary = df_summary.reindex(columns=summary_cols).sort_values(by="Church Code")
                df_summary.to_excel(writer, sheet_name="Summary", index=False)
                
                logger.debug(f"Summary tab: {len(df_summary)} rows.")

                # Contacts-Status Tab
                # Build sports-registered lookup keyed by Participant ID (WP) and ChMeetings ID
                sports_by_wp_id: Dict[str, list] = {}
                sports_by_chm_id: Dict[str, list] = {}
                for rrow in roster_rows:
                    label_parts = [
                        str(rrow.get("sport_type") or "").strip(),
                        str(rrow.get("sport_gender") or "").strip(),
                        str(rrow.get("sport_format") or "").strip(),
                    ]
                    label = " ".join(p for p in label_parts if p)
                    if not label:
                        continue
                    wp_pid = str(rrow.get("Participant ID (WP)") or "").strip()
                    chm_pid = str(rrow.get("ChMeetings ID") or "").strip()
                    if wp_pid and wp_pid not in ("0", ""):
                        sports_by_wp_id.setdefault(wp_pid, [])
                        if label not in sports_by_wp_id[wp_pid]:
                            sports_by_wp_id[wp_pid].append(label)
                    if chm_pid and chm_pid not in ("0", ""):
                        sports_by_chm_id.setdefault(chm_pid, [])
                        if label not in sports_by_chm_id[chm_pid]:
                            sports_by_chm_id[chm_pid].append(label)

                for crow in contacts_rows:
                    wp_pid = str(crow.get("Participant ID (WP)") or "").strip()
                    chm_pid = str(crow.get("ChMeetings ID") or "").strip()
                    sports = (
                        sports_by_wp_id.get(wp_pid)
                        or sports_by_chm_id.get(chm_pid)
                        or []
                    )
                    crow["Sports Registered"] = ", ".join(sorted(sports)) if sports else ""

                df_contacts = pd.DataFrame(contacts_rows)
                if not df_contacts.empty:

                    photo_url_col_name = "Photo URL (WP)"
                    if photo_url_col_name in df_contacts.columns:
                        # Create the hyperlink formula if the URL is not "N/A" and looks like a URL
                        df_contacts[photo_url_col_name] = df_contacts[photo_url_col_name].apply(
                            lambda url: f'=HYPERLINK("{url}", "{url}")'
                            if isinstance(url, str) and url != "N/A" and (url.startswith("http://") or url.startswith("https://"))
                            else url # Keep "N/A" or other non-URL values as is
                        )

                    contacts_cols = [
                        "Church Team", "ChMeetings ID", "First Name", "Last Name", "Is_Participant",
                        "Is_Member_ChM", "Participant ID (WP)", "Approval_Status (WP)",
                        "Total_Open_ERRORs (WP)", "Gender", "Birthdate", "Age (at Event)",
                        "Mobile Phone", "Email", "Registration Date (WP)", "Sports Registered", "Athlete Fee",
                        "First_Open_ERROR_Desc (WP)",
                        "Box 1", "Box 2", "Box 3", "Box 4", "Box 5", "Box 6",
                        photo_url_col_name, "Update_on_ChM"
                    ]
                    for col in contacts_cols: # Ensure all contact columns exist
                        if col not in df_contacts.columns:
                            df_contacts[col] = None
                    df_contacts = df_contacts.reindex(columns=contacts_cols).sort_values(
                        by=["Church Team", "Total_Open_ERRORs (WP)", "Is_Participant", "Last Name", "First Name"],
                        ascending=[True, False, False, True, True]
                    )
                df_contacts.to_excel(writer, sheet_name="Contacts-Status", index=False)
                logger.debug(f"Contacts-Status tab: {len(df_contacts)} rows.")

                # Roster Tab
                df_roster = pd.DataFrame(roster_rows)
                if not df_roster.empty:
                    roster_cols = [
                        "Church Team", "ChMeetings ID", "Participant ID (WP)", "Approval_Status (WP)",
                        "Is_Member_ChM", "Photo",          # ADD THIS LINE
                        "First Name", "Last Name", "Gender", "Age (at Event)", "Mobile Phone", "Email",
                        "sport_type", "sport_gender", "sport_format", "team_order", "partner_name",
                        "Open_TEAM_Issue_Count (WP)", "Open_TEAM_Issue_Desc (WP)"
                    ]
                    for col in roster_cols: # Ensure all roster columns exist
                        if col not in df_roster.columns:
                            df_roster[col] = None
                    df_roster = df_roster.reindex(columns=roster_cols).sort_values(
                        by=["Church Team", "sport_type", "sport_gender", "Last Name", "First Name", 
                            "Approval_Status (WP)", "sport_format"] # Removed team_order and partner_name from sort if they are often None
                    )
                df_roster.to_excel(writer, sheet_name="Roster", index=False)
                logger.debug(f"Roster tab: {len(df_roster)} rows.")

                # Validation-Issues Tab
                df_validation = pd.DataFrame(validation_rows)
                if not df_validation.empty:
                    validation_cols = [
                        "Church Team", "Rule Level", "Severity", "Status",
                        "Issue Type", "Rule Code",
                        "Participant ID (WP)", "ChMeetings ID", "Participant Name",
                        "Approval_Status (WP)", "sport_type", "sport_format",
                        "Issue Description"
                    ]
                    for col in validation_cols:
                        if col not in df_validation.columns:
                            df_validation[col] = None
                    df_validation = df_validation.reindex(columns=validation_cols).sort_values(
                        by=["Church Team", "Rule Level", "Severity", "Participant Name", "Issue Type"],
                        ascending=[True, True, True, True, True]
                    )
                df_validation.to_excel(writer, sheet_name="Validation-Issues", index=False)
                logger.debug(f"Validation-Issues tab: {len(df_validation)} rows.")

                # Venue-Estimator Tab (only on the consolidated ALL export — see Issue #83)
                if include_venue_capacity:
                    venue_rows = self._build_venue_capacity_rows(roster_rows)
                    venue_cols = [
                        "Event", "Potential Teams/Entries", "Estimating Teams/Entries", "Teams",
                        "Pool Games Per Team", "Minutes Per Game", "Pool Slots",
                        "Playoff Teams", "Playoff Slots", "Third Place?",
                        "Third Place Slots", "Total Court Slots", "Estimated Court Hours",
                    ]
                    df_venue = pd.DataFrame(venue_rows, columns=venue_cols)
                    snapshot_note = (
                        f"Roster snapshot as of {datetime.now().strftime('%Y-%m-%d')} — "
                        "Estimating = complete entries; Potential = all registrations including partial. "
                        "Approval-agnostic. Updates with each export run."
                    )
                    # Data first, then a blank row, then the snapshot disclaimer at the bottom.
                    df_venue.to_excel(writer, sheet_name="Venue-Estimator", index=False, startrow=0)
                    venue_ws = writer.sheets["Venue-Estimator"]
                    note_row = len(df_venue) + 3  # header + data rows + blank row
                    venue_ws.cell(row=note_row, column=1, value=snapshot_note)
                    logger.debug(f"Venue-Estimator tab: {len(df_venue)} rows.")

                    # Pod-Divisions Tab (Issue #88) — division-level pod planning summary.
                    pod_div_rows = self._build_pod_divisions_rows(roster_rows, validation_rows)
                    pod_div_cols = [
                        "division_id", "sport_type", "sport_gender", "sport_format",
                        "resource_type", "minutes_per_game",
                        "planning_entries", "confirmed_entries", "provisional_entries",
                        "anomaly_count", "division_status", "notes",
                    ]
                    df_pod_div = pd.DataFrame(pod_div_rows, columns=pod_div_cols)
                    df_pod_div.to_excel(writer, sheet_name="Pod-Divisions", index=False)
                    logger.debug(f"Pod-Divisions tab: {len(df_pod_div)} rows.")

                    # Pod-Entries-Review Tab (Issue #88) — coordinator review of each entry.
                    pod_entry_rows = self._build_pod_entries_review_rows(roster_rows, validation_rows)
                    pod_entry_cols = [
                        "entry_id", "division_id", "entry_type",
                        "participant_1_name", "participant_2_name",
                        "source_participant_ids", "church_team",
                        "partner_status", "review_status", "notes",
                    ]
                    df_pod_entries = pd.DataFrame(pod_entry_rows, columns=pod_entry_cols)
                    df_pod_entries.to_excel(writer, sheet_name="Pod-Entries-Review", index=False)
                    logger.debug(f"Pod-Entries-Review tab: {len(df_pod_entries)} rows.")
                    # Court-Schedule-Sketch Tab (Excel-only planning — no WordPress writes)
                    sketch_ws = writer.book.create_sheet(title="Court-Schedule-Sketch")
                    self._write_court_schedule_sketch(sketch_ws, roster_rows)

                    # Pod-Resource-Estimate Tab (Excel-only planning — no WordPress writes)
                    venue_input_path = DATA_DIR / VENUE_INPUT_FILENAME
                    available_by_resource = self._load_venue_input(venue_input_path)
                    pod_rows = self._build_pod_resource_rows(roster_rows, available_by_resource)
                    pod_ws = writer.book.create_sheet(title="Pod-Resource-Estimate")
                    self._write_pod_resource_estimate(pod_ws, pod_rows, available_by_resource)

                    # Schedule-Input Tab + JSON (Issue #87) — OR-Tools-ready planning artifact
                    schedule_input = self._build_schedule_input(
                        roster_rows, validation_rows, venue_input_path
                    )
                    si_ws = writer.book.create_sheet(title="Schedule-Input")
                    self._write_schedule_input_tab(si_ws, schedule_input)
                    json_path = filepath.parent / "schedule_input.json"
                    json_path.write_text(
                        json.dumps(schedule_input, indent=2, default=str),
                        encoding="utf-8",
                    )
                    logger.info(
                        f"Schedule-Input: {schedule_input['game_count']} games, "
                        f"{schedule_input['resource_count']} resources → {json_path}"
                    )

                # Add yellow note to Photo column in Roster sheet
                if not df_roster.empty:
                    from openpyxl.comments import Comment
                    from openpyxl.styles import PatternFill
                    
                    roster_ws = writer.sheets["Roster"]
                    
                    # Add comment to Photo column (column F)
                    photo_comment = Comment(
                        "In Office365 edition 2023 and later, you can remove the @ from the formula to display the image",
                        "Bumble",200,300 # Width and height in pixels
                    )
                    roster_ws["F1"].comment = photo_comment
                    
                    # Add yellow background
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    roster_ws["F1"].fill = yellow_fill

                # Sport-Specific Tabs
                if not df_roster.empty and "sport_type" in df_roster.columns:
                    # Group roster data by sport combinations
                    sport_groups = {}
                    
                    for _, row in df_roster.iterrows():
                        sport_type = row.get("sport_type", "")
                        sport_gender = row.get("sport_gender", "")
                        
                        if not sport_type:
                            continue
                            
                        # Special handling for Volleyball - separate by gender
                        if sport_type.upper() == "VOLLEYBALL" or sport_type.upper().startswith("VB"):
                            if sport_gender.upper() == "MEN" or sport_gender.upper() == "MALE":
                                tab_name = "VB Men"
                            elif sport_gender.upper() == "WOMEN" or sport_gender.upper() == "FEMALE":
                                tab_name = "VB Women"
                            else:
                                tab_name = "Volleyball"  # Fallback if gender unclear
                        else:
                            # For all other sports, use sport_type only
                            tab_name = sport_type
                        
                        # Create group if it doesn't exist
                        if tab_name not in sport_groups:
                            sport_groups[tab_name] = []
                        
                        sport_groups[tab_name].append(row.to_dict())
                    
                    # Create tabs for each sport group
                    for sport_name, sport_data in sport_groups.items():
                        if not sport_data:  # Skip empty sports
                            continue
                            
                        df_sport = pd.DataFrame(sport_data)
                        
                        # Modify columns: replace Last Name and First Name with Full Name
                        sport_cols = [
                            "Church Team", "ChMeetings ID", "Participant ID (WP)", "Approval_Status (WP)",
                            "Is_Member_ChM", "Photo", "Full Name", "Gender", "Age (at Event)", 
                            "Mobile Phone", "Email", "sport_type", "sport_gender", "sport_format", 
                            "team_order", "partner_name",
                            "Open_TEAM_Issue_Count (WP)", "Open_TEAM_Issue_Desc (WP)"
                        ]
                        
                        # Create Full Name column
                        df_sport["Full Name"] = df_sport["Last Name"].astype(str) + " " + df_sport["First Name"].astype(str)
                        
                        # Ensure all sport columns exist
                        for col in sport_cols:
                            if col not in df_sport.columns:
                                df_sport[col] = None
                        
                        # Reindex and sort
                        df_sport = df_sport.reindex(columns=sport_cols).sort_values(
                            by=["Church Team", "sport_type", "sport_gender", "Full Name", "Approval_Status (WP)"],
                            ascending=[True, True, True, True, True]
                        )
                        
                        # Write to Excel with sport name as tab name
                        # Clean tab name for Excel compatibility (max 31 characters, no special chars)
                        clean_tab_name = "".join(c for c in sport_name if c.isalnum() or c in " -_")[:31]
                        df_sport.to_excel(writer, sheet_name=clean_tab_name, index=False)
                        logger.debug(f"{clean_tab_name} tab: {len(df_sport)} rows.")

                # SINGLE Excel Formatting Block - Add right before closing writer
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Add auto filter to all data (if worksheet has data)
                    if worksheet.max_row > 1:  # More than just headers
                        worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))/2 # Divide by 2 for better fit
                            except:
                                pass
                        
                        # Set width with some padding
                        adjusted_width = min(max_length + 2, 50)  # Max width of 50
                        worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"Successfully wrote Excel report: {filepath}")
        except Exception as e:
            logger.error(f"Failed to write Excel file {filepath}: {e}", exc_info=True)

    def _handle_force_resend(self, contacts_data: List[Dict[str, Any]], 
                            force_pending: bool, force_validated1: bool, force_validated2: bool,
                            dry_run: bool, target_resend_chm_id: Optional[str] = None) -> int:
        """Handle force resend based on participant categories."""
        
        from sync.manager import SyncManager  # Import here to avoid circular imports
        
        participants_to_resend = []
        
        for contact in contacts_data:
            approval_status = contact.get("Approval_Status (WP)", "").lower()
            
            # Category 1: Pending participants
            if force_pending and approval_status in ["pending", "pending_approval"]:
                participants_to_resend.append(contact)
                logger.debug(f"Added pending participant: {contact.get('First Name')} {contact.get('Last Name')} ({contact.get('Church Team')})")
                
            # Category 2 & 3: Validated participants  
            elif approval_status == "validated":
                # Check Box 1-6 data to categorize
                has_box_data = any(contact.get(f"Box {i}", "") for i in range(1, 7))
                
                if force_validated1 and has_box_data:
                    # Under review by church rep
                    participants_to_resend.append(contact)
                    logger.debug(f"Added validated-under-review participant: {contact.get('First Name')} {contact.get('Last Name')} ({contact.get('Church Team')})")
                elif force_validated2 and not has_box_data:
                    # Not reviewed yet by church rep
                    participants_to_resend.append(contact)
                    logger.debug(f"Added validated-not-reviewed participant: {contact.get('First Name')} {contact.get('Last Name')} ({contact.get('Church Team')})")

        if target_resend_chm_id:
            normalized_target = str(target_resend_chm_id).strip()
            participants_to_resend = [
                contact
                for contact in participants_to_resend
                if str(contact.get("ChMeetings ID", "")).strip() == normalized_target
            ]
            logger.info(
                f"Applied resend filter for ChMeetings ID {normalized_target}. "
                f"Matching participants: {len(participants_to_resend)}"
            )
        
        logger.info(f"Found {len(participants_to_resend)} participants matching criteria")
        
        if dry_run:
            logger.info("DRY RUN - Would resend emails to:")
            
            # Load churches cache to get email addresses for dry run preview
            temp_sync_manager = SyncManager()
            with temp_sync_manager:
                # Load churches cache
                wp_churches = temp_sync_manager.wordpress_connector.get_churches()
                if wp_churches:
                    churches_cache = {c["church_code"]: c for c in wp_churches}
                else:
                    logger.error("Failed to load churches for dry-run preview")
                    return len(participants_to_resend)
                
                for p in participants_to_resend:
                    participant_name = f"{p.get('First Name')} {p.get('Last Name')}"
                    church_code = p.get('Church Team')
                    participant_email = p.get('Email', 'N/A')
                    status = p.get('Approval_Status (WP)')
                    
                    # Get church details for email preview
                    if church_code in churches_cache:
                        church = churches_cache[church_code]
                        pastor_email = church.get('pastor_email', 'N/A')
                        church_rep_email = church.get('church_rep_email', 'N/A')
                        
                        logger.info(f"  {participant_name} ({church_code}) - Status: {status}")
                        logger.info(f"    → Pastor approval email TO: {pastor_email}")
                        logger.info(f"    → Participant notification TO: {participant_email}")
                        if church_rep_email != 'N/A':
                            logger.info(f"    → Church rep CC: {church_rep_email}")
                        logger.info("")  # Empty line for readability
                    else:
                        logger.info(f"  {participant_name} ({church_code}) - Status: {status}")
                        logger.error(f"    → ERROR: Church {church_code} not found in cache!")
                        logger.info("")
            
            return len(participants_to_resend)
        
        # Actually send the emails
        sync_manager = SyncManager()
        with sync_manager:
            success_count = 0
            for participant_contact in participants_to_resend:
                wp_participant_id = participant_contact.get("Participant ID (WP)")
                if wp_participant_id and self._resend_approval_for_participant(sync_manager, participant_contact):
                    success_count += 1
                    logger.info(f"Resent approval email for {participant_contact.get('First Name')} {participant_contact.get('Last Name')}")
                else:
                    logger.error(f"Failed to resend for {participant_contact.get('First Name')} {participant_contact.get('Last Name')}")
        
        return success_count


    # In church_teams_export.py, update both methods:
    def _resend_approval_for_participant(self, sync_manager, participant_contact: Dict[str, Any]) -> bool:
        """Resend approval email for a specific participant using contact data."""
        try:
            # Extract data from contact
            wp_participant_id = participant_contact.get("Participant ID (WP)")
            first_name = participant_contact.get("First Name", "")
            last_name = participant_contact.get("Last Name", "")
            church_team = participant_contact.get("Church Team", "")
            chm_id = participant_contact.get("ChMeetings ID", "")
            email = participant_contact.get("Email", "")
            
            if not wp_participant_id:
                logger.error(f"No WordPress participant ID found for {first_name} {last_name}")
                return False
            
            # Ensure churches cache is loaded
            if not sync_manager.churches_cache:
                logger.info("Loading churches cache for resend operation...")
                wp_churches = sync_manager.wordpress_connector.get_churches()
                if wp_churches:
                    sync_manager.churches_cache = {c["church_code"]: c for c in wp_churches}
                else:
                    logger.error("Failed to load churches cache")
                    return False
            
            church_code = church_team
            if not church_code or church_code not in sync_manager.churches_cache:
                logger.error(f"Church {church_code} not found for participant {first_name} {last_name}")
                return False
            
            church = sync_manager.churches_cache[church_code]
            pastor_email = church.get("pastor_email")
            if not pastor_email:
                logger.error(f"No pastor email for church {church_code}")
                return False
            
            # Generate new token and expiry
            from uuid import uuid4
            import datetime
            from config import Config
            
            token = str(uuid4())
            expiry_date = datetime.datetime.now() + datetime.timedelta(days=Config.TOKEN_EXPIRY_DAYS)
            
            # Prepare approval data for CREATE endpoint (which now handles UNIQUE constraint properly)
            church_wp_id = church["church_id"]
            approval_data = {
                "participant_id": wp_participant_id,
                "church_id": church_wp_id,
                "approval_token": token,
                "token_expiry": expiry_date.strftime("%Y-%m-%d %H:%M:%S"),
                "pastor_email": pastor_email,
                "approval_status": "pending",
                "synced_to_chmeetings": False
            }
            
            # Use CREATE endpoint (which now handles UNIQUE constraint with INSERT ... ON DUPLICATE KEY UPDATE)
            logger.info(f"Creating/updating approval record for participant {wp_participant_id} via CREATE endpoint")
            
            result = sync_manager.wordpress_connector.create_approval(approval_data)
            
            if not result:
                logger.error(f"Failed to create/update approval record for participant {wp_participant_id}")
                return False
            
            logger.info(f"Successfully created/updated approval record for participant {wp_participant_id}")
            
            # Build participant data for email
            participant_data = {
                "participant_id": wp_participant_id,
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "church_code": church_code,
                "chmeetings_id": chm_id,
                "is_church_member": participant_contact.get("Is_Member_ChM", "No") == "Yes",
                "photo_url": participant_contact.get("Photo URL (WP)", "")
            }
            
            # Send approval email
            participant_name = f"{first_name} {last_name}"
            success = sync_manager.send_pastor_approval_email(
                pastor_email, participant_name, token, participant_data, expiry_date
            )
            
            if success:
                logger.info(f"Successfully resent approval email for {participant_name} (ID: {wp_participant_id})")
                return True
            else:
                logger.error(f"Failed to send approval email for {participant_name} (ID: {wp_participant_id})")
                return False
            
        except Exception as e:
            logger.error(f"Error resending approval for participant {participant_contact.get('First Name')} {participant_contact.get('Last Name')}: {e}", exc_info=True)
            return False
                    
    @staticmethod
    def _build_schedule_output_flat_rows(
        schedule_output: Dict[str, Any],
        schedule_input: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """Build sorted flat-list rows for the Schedule-by-Sport tab.

        Each row joins one assignment from schedule_output with game metadata
        from schedule_input.  Rows are sorted by event → stage order → round → slot.
        """
        game_meta = {g["game_id"]: g for g in schedule_input.get("games", [])}
        res_meta  = {r["resource_id"]: r for r in schedule_input.get("resources", [])}
        _STAGE_ORDER = {"Pool": 0, "R1": 1, "QF": 2, "Semi": 3, "Final": 4, "3rd": 5}
        _DAY_DISPLAY = {
            "Sat-1": "1st Sat", "Sun-1": "1st Sun",
            "Sat-2": "2nd Sat", "Sun-2": "2nd Sun",
        }
        rows: List[Dict[str, Any]] = []
        for a in schedule_output.get("assignments", []):
            gid  = a["game_id"]
            rid  = a["resource_id"]
            slot = a["slot"]
            # Fall back to the assignment dict itself for playoff games whose
            # game_id is not in schedule_input games (they carry event/stage fields).
            game = game_meta.get(gid, a)
            res  = res_meta.get(rid, {})
            time_part = slot.rsplit("-", maxsplit=1)[-1] if "-" in slot else slot
            rows.append({
                "game_id":          gid,
                "event":            game.get("event", ""),
                "stage":            game.get("stage", ""),
                "round":            game.get("round", ""),
                "team_a_id":        game.get("team_a_id", ""),
                "team_b_id":        game.get("team_b_id", ""),
                "resource_label":   res.get("label", rid),
                "day":              _DAY_DISPLAY.get(res.get("day", ""), res.get("day", "")),
                "slot":             time_part,
                "duration_minutes": game.get("duration_minutes", ""),
            })
        rows.sort(key=lambda r: (
            r["event"],
            _STAGE_ORDER.get(str(r["stage"]), 99),
            int(r["round"]) if isinstance(r["round"], int) else 0,
            r["slot"],
        ))
        return rows

    @staticmethod
    def _write_schedule_output_report(
        filepath: Path,
        schedule_output: Dict[str, Any],
        schedule_input: Dict[str, Any],
    ) -> None:
        """Write Schedule-by-Time and Schedule-by-Sport Excel tabs from solver output.

        Tab 1 — Schedule-by-Time: grid (rows = time slots, columns = courts),
          colour-coded by sport, with session sections separated by grey rows.
        Tab 2 — Schedule-by-Sport: flat list sorted by event → stage → round,
          with auto-filter and an unscheduled section when applicable.
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        game_meta: Dict[str, Dict[str, Any]] = {
            g["game_id"]: g for g in schedule_input.get("games", [])
        }
        res_meta: Dict[str, Dict[str, Any]] = {
            r["resource_id"]: r for r in schedule_input.get("resources", [])
        }
        assign_map: Dict[Tuple[str, str], Dict[str, Any]] = {
            (a["resource_id"], a["slot"]): game_meta.get(a["game_id"], {"game_id": a["game_id"]})
            for a in schedule_output.get("assignments", [])
        }

        solved_at     = schedule_output.get("solved_at", "")
        status        = schedule_output.get("status", "")
        n_assigned    = len(schedule_output.get("assignments", []))
        n_unscheduled = len(schedule_output.get("unscheduled", []))
        snapshot      = (
            f"Generated: {solved_at}  |  Status: {status}  |  "
            f"Scheduled: {n_assigned}  |  Unscheduled: {n_unscheduled}"
        )

        sec_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
        hdr_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER,  fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)
        center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        red_fill = PatternFill(fgColor="FFC7CE", fill_type="solid")

        _SPORT_COLORS: Dict[str, str] = {
            SPORT_TYPE["BASKETBALL"]:       SCHEDULE_SKETCH_COLOR_BASKETBALL,
            SPORT_TYPE["VOLLEYBALL_MEN"]:   SCHEDULE_SKETCH_COLOR_VB_MEN,
            SPORT_TYPE["VOLLEYBALL_WOMEN"]: SCHEDULE_SKETCH_COLOR_VB_WOMEN,
        }
        _DAY_ORDER   = {"Sat-1": 0, "Sun-1": 1, "Sat-2": 2, "Sun-2": 3}
        _DAY_DISPLAY = {
            "Sat-1": "1st Saturday", "Sun-1": "1st Sunday",
            "Sat-2": "2nd Saturday", "Sun-2": "2nd Sunday",
        }

        def _sport_fill(event: str) -> PatternFill:
            return PatternFill(
                fgColor=_SPORT_COLORS.get(event, "EBF1DE"), fill_type="solid"
            )

        def _slot_times(res: Dict[str, Any]) -> List[str]:
            o_h, o_m = map(int, res["open_time"].split(":"))
            c_h, c_m = map(int, res["close_time"].split(":"))
            sm        = res["slot_minutes"]
            open_min  = o_h * 60 + o_m
            close_min = c_h * 60 + c_m
            times: List[str] = []
            t = open_min
            while t + sm <= close_min:
                times.append(f"{t // 60:02d}:{t % 60:02d}")
                t += sm
            return times

        def _cell_text(game: Dict[str, Any]) -> str:
            gid = game.get("game_id", "")
            a   = str(game.get("team_a_id") or "")
            b   = str(game.get("team_b_id") or "")
            # Show teams only when they look like real church codes (≤5 chars, no hyphens)
            if a and b and len(a) <= 5 and "-" not in a:
                return f"{gid}\n{a} vs {b}"
            return gid

        def _time_sort_key(hhmm: str) -> int:
            h, m = map(int, hhmm.split(":"))
            return h * 60 + m

        def _resource_group_key(res: Dict[str, Any]) -> Tuple[str, str, str, str, int]:
            return (
                str(res.get("day", "")),
                str(res.get("resource_type", "")),
                str(res.get("open_time", "")),
                str(res.get("close_time", "")),
                int(res.get("slot_minutes", 0) or 0),
            )

        # Group resources by uniform day/window/resource pool so pod schedules with
        # mixed slot lengths do not get collapsed into one broken "Day-1" grid.
        resource_groups: Dict[Tuple[str, str, str, str, int], List[Dict[str, Any]]] = {}
        for res in schedule_input.get("resources", []):
            resource_groups.setdefault(_resource_group_key(res), []).append(res)

        group_counts_by_day: Dict[str, int] = {}
        for day, _, _, _, _ in resource_groups.keys():
            group_counts_by_day[day] = group_counts_by_day.get(day, 0) + 1

        sorted_group_keys = sorted(
            resource_groups.keys(),
            key=lambda key: (
                _DAY_ORDER.get(key[0], 99),
                key[0],
                _time_sort_key(key[2]) if key[2] else 0,
                _time_sort_key(key[3]) if key[3] else 0,
                key[4],
                key[1],
            ),
        )
        max_resources = max((len(v) for v in resource_groups.values()), default=4)
        n_cols        = 1 + max_resources

        def _section_label(group_key: Tuple[str, str, str, str, int]) -> str:
            day, resource_type, open_time, close_time, slot_minutes = group_key
            day_label = _DAY_DISPLAY.get(day, day)
            if (
                day in _DAY_DISPLAY
                and resource_type == GYM_RESOURCE_TYPE
                and group_counts_by_day.get(day, 0) == 1
            ):
                return day_label
            return (
                f"{day_label} — {resource_type} "
                f"({open_time}-{close_time}, {slot_minutes}m)"
            )

        wb = Workbook()

        # ── Tab 1: Schedule-by-Time ──────────────────────────────────────────
        ws1       = wb.active
        ws1.title = "Schedule-by-Time"

        # Row 1 — report title (merged)
        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        c = ws1.cell(row=1, column=1, value="VAY Sports Fest — Schedule by Time")
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, center

        ws1.freeze_panes = "A2"

        cur_row = 3
        for group_key in sorted_group_keys:
            day_res = sorted(resource_groups[group_key], key=lambda r: r["resource_id"])
            if not day_res:
                continue

            # Section header (grey, merged)
            ws1.merge_cells(
                start_row=cur_row, start_column=1,
                end_row=cur_row, end_column=n_cols,
            )
            c = ws1.cell(row=cur_row, column=1, value=_section_label(group_key))
            c.fill, c.font, c.alignment = sec_fill, bold_font, center
            cur_row += 1

            # Column headers for this group
            ws1.cell(row=cur_row, column=1, value="Time").font = bold_font
            ws1.cell(row=cur_row, column=1).fill = sec_fill
            ws1.cell(row=cur_row, column=1).alignment = center
            for ci, res in enumerate(day_res, start=2):
                c = ws1.cell(row=cur_row, column=ci, value=res["label"])
                c.fill, c.font, c.alignment = sec_fill, bold_font, center
            cur_row += 1

            day = group_key[0]
            # Data rows — one per time slot in this uniform resource group
            for t_str in _slot_times(day_res[0]):
                slot_label = f"{day}-{t_str}"
                ws1.cell(row=cur_row, column=1, value=t_str).alignment = center
                for ci, res in enumerate(day_res, start=2):
                    game = assign_map.get((res["resource_id"], slot_label))
                    cell = ws1.cell(row=cur_row, column=ci)
                    if game:
                        cell.value = _cell_text(game)
                        cell.fill  = _sport_fill(game.get("event", ""))
                    cell.alignment = center
                cur_row += 1

            cur_row += 1  # blank row between sessions

        ws1.cell(row=cur_row + 1, column=1, value=snapshot)
        ws1.column_dimensions["A"].width = 7
        for ci in range(2, n_cols + 1):
            ws1.column_dimensions[get_column_letter(ci)].width = 18

        # ── Tab 2: Schedule-by-Sport ─────────────────────────────────────────
        ws2       = wb.create_sheet("Schedule-by-Sport")
        flat_rows = ChurchTeamsExporter._build_schedule_output_flat_rows(
            schedule_output, schedule_input
        )
        col_defs = [
            ("game_id",          14),
            ("event",            28),
            ("stage",             8),
            ("round",             6),
            ("team_a_id",        20),
            ("team_b_id",        20),
            ("resource_label",   14),
            ("day",              10),
            ("slot",              8),
            ("duration_minutes", 16),
        ]
        cols = [col for col, _ in col_defs]

        for ci, (col, _) in enumerate(col_defs, start=1):
            cell = ws2.cell(row=1, column=ci, value=col)
            cell.fill, cell.font = hdr_fill, hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}1"

        for ri, row in enumerate(flat_rows, start=2):
            fill = _sport_fill(row.get("event", ""))
            for ci, col in enumerate(cols, start=1):
                cell = ws2.cell(row=ri, column=ci, value=row.get(col, ""))
                cell.fill, cell.alignment = fill, left

        # Unscheduled section at bottom
        unscheduled = schedule_output.get("unscheduled", [])
        ri = len(flat_rows) + 3
        if unscheduled:
            ws2.merge_cells(
                start_row=ri, start_column=1, end_row=ri, end_column=len(col_defs)
            )
            c = ws2.cell(
                row=ri, column=1,
                value=f"Unscheduled Games ({len(unscheduled)})",
            )
            c.fill, c.font = red_fill, Font(bold=True)
            for gid in unscheduled:
                ri += 1
                ws2.cell(row=ri, column=1, value=gid).fill = red_fill
            ri += 2

        # Pool results summary — shown when pool_results present
        pool_results = schedule_output.get("pool_results", [])
        if pool_results:
            ws2.merge_cells(
                start_row=ri, start_column=1, end_row=ri, end_column=len(col_defs)
            )
            c = ws2.cell(row=ri, column=1, value="Pool Results")
            c.fill, c.font = sec_fill, Font(bold=True)
            ri += 1
            for pr in pool_results:
                pr_status = pr.get("status", "")
                pr_fill   = red_fill if pr_status not in ("OPTIMAL", "FEASIBLE") else PatternFill(
                    fgColor="C6EFCE", fill_type="solid"
                )
                ws2.cell(row=ri, column=1, value=pr.get("resource_type", "")).fill = pr_fill
                ws2.cell(row=ri, column=2, value=pr_status).fill                   = pr_fill
                ws2.cell(row=ri, column=3, value=f"Assigned: {len(pr.get('assignments', []))}").fill  = pr_fill
                ws2.cell(row=ri, column=4, value=f"Unscheduled: {len(pr.get('unscheduled', []))}").fill = pr_fill
                ri += 1
                for diag in pr.get("diagnostics", []):
                    for line in (diag.get("missing_resource_events") or []):
                        ws2.cell(row=ri, column=2,
                                 value=f"  No resources: {line.get('event','')} ({line.get('game_count',0)} games)"
                                 ).fill = red_fill
                        ri += 1
                    if diag.get("shortage_slots", 0) > 0:
                        ws2.cell(row=ri, column=2,
                                 value=f"  Short {diag['shortage_slots']} slot(s): "
                                       f"need {diag['required_slots']}, have {diag['available_slots']}"
                                 ).fill = red_fill
                        ri += 1
            ri += 1

        ws2.cell(row=ri, column=1, value=snapshot)
        for ci, (_, width) in enumerate(col_defs, start=1):
            ws2.column_dimensions[get_column_letter(ci)].width = width

        wb.save(filepath)
        logger.info(f"Schedule output report written to: {filepath}")

    def close(self):
        """Closes any open connections (if necessary)."""
        logger.info("Closing ChurchTeamsExporter resources.") # MODIFIED LOGGER MESSAGE
        if self.chm_connector:
            self.chm_connector.close()
        logger.info("ChurchTeamsExporter resources closed.") # MODIFIED LOGGER MESSAGE

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

# end of the ChurchTeamsExporter class
# end of the middleware/church_teams_export.py file
