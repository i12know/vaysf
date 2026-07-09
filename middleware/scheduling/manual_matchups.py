"""manual_matchups - 2026 fixed team-sport matchup import.

The 2026 team-sport pool games are approved in a meeting workbook. This module
turns that workbook into a small JSON sidecar that schedule_input generation
can consume without silently depending on raw Excel every run.
"""
from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from loguru import logger
from openpyxl import load_workbook

from config import (
    COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM,
    COURT_ESTIMATE_BC_TEAMS_PER_GAME,
    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
    COURT_ESTIMATE_MINUTES_PER_GAME,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
    SPORT_FORMAT,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
    TEAM_RESOURCE_TYPE_SOCCER,
)

MANUAL_TEAM_MATCHUP_WORKBOOK_FILENAME = (
    "2026-VAY-Lottery-Drawing_Team-Assignment(ALL-TEAM-SPORTS)_template.xlsx"
)
MANUAL_TEAM_MATCHUP_SIDECAR_FILENAME = "manual_team_matchups.json"

_GYM_CORE_SOLVER_POOL = "Gym Core"

_DEFAULT_SHEET_CONFIGS: List[Dict[str, Any]] = [
    {
        "sheet_name": "BB_round2",
        "event": SPORT_TYPE["BASKETBALL"],
        "prefix": "BBM",
        "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
        "expected_games_per_team": 3,
        "team_count": 2,
        "solver_pool": _GYM_CORE_SOLVER_POOL,
        "active": True,
    },
    {
        "sheet_name": "MVB",
        "event": SPORT_TYPE["VOLLEYBALL_MEN"],
        "prefix": "VBM",
        "resource_type": GYM_RESOURCE_TYPE_VOLLEYBALL,
        "expected_games_per_team": 3,
        "team_count": 2,
        "solver_pool": _GYM_CORE_SOLVER_POOL,
        "active": True,
    },
    {
        "sheet_name": "WVB",
        "event": SPORT_TYPE["VOLLEYBALL_WOMEN"],
        "prefix": "VBW",
        "resource_type": GYM_RESOURCE_TYPE_VOLLEYBALL,
        "expected_games_per_team": 4,
        "team_count": 2,
        "solver_pool": _GYM_CORE_SOLVER_POOL,
        "active": True,
    },
    {
        "sheet_name": "SOC",
        "event": SPORT_TYPE["SOCCER"],
        "prefix": "SOC",
        "resource_type": TEAM_RESOURCE_TYPE_SOCCER,
        "expected_games_per_team": COURT_ESTIMATE_POOL_GAMES_PER_TEAM[
            SPORT_TYPE["SOCCER"]
        ],
        "team_count": 2,
        "active": True,
    },
    {
        "sheet_name": "BC",
        "event": SPORT_TYPE["BIBLE_CHALLENGE"],
        "prefix": "BC",
        "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
        "expected_games_per_team": COURT_ESTIMATE_BC_RR_GAMES_PER_TEAM,
        "team_count": COURT_ESTIMATE_BC_TEAMS_PER_GAME,
        "game_id_style": "round_robin",
        "active": True,
    },
    {
        "sheet_name": "BB (2)",
        "event": SPORT_TYPE["BASKETBALL"],
        "prefix": "BBM",
        "resource_type": GYM_RESOURCE_TYPE_BASKETBALL,
        "expected_games_per_team": 3,
        "team_count": 2,
        "solver_pool": _GYM_CORE_SOLVER_POOL,
        "active": False,
    },
]

_TEAM_SIDES = ("a", "b", "c")


def default_sidecar_path(base_dir: Path) -> Path:
    """Return the default sidecar path for imported manual matchups."""
    return Path(base_dir) / MANUAL_TEAM_MATCHUP_SIDECAR_FILENAME


def default_workbook_path(data_dir: Path) -> Path:
    """Return the default 2026 manual matchup workbook path."""
    return Path(data_dir) / MANUAL_TEAM_MATCHUP_WORKBOOK_FILENAME


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _clean_team_code(value: Any) -> str:
    return _clean_text(value).upper()


def _is_bye(value: Any) -> bool:
    return _clean_text(value).casefold() == "bye"


def _int_or_none(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _clean_text(value)
    if text.isdigit():
        return int(text)
    return None


def _pool_id_from_label(label: str) -> str:
    text = _clean_text(label).upper()
    if text.startswith("POOL "):
        suffix = text.split("POOL ", 1)[1].strip()
    else:
        suffix = text.strip()
    if len(suffix) == 1 and "A" <= suffix <= "Z":
        return f"P{ord(suffix) - ord('A') + 1}"
    return text.replace(" ", "-") if text else ""


def _event_target(event_name: str) -> Tuple[str, Sequence[str]]:
    """Return (sport_type, accepted genders) expected in roster rows for an event."""
    if event_name == SPORT_TYPE["BASKETBALL"]:
        return "Basketball", ("Men",)
    if event_name == SPORT_TYPE["VOLLEYBALL_MEN"]:
        return "Volleyball", ("Men",)
    if event_name == SPORT_TYPE["VOLLEYBALL_WOMEN"]:
        return "Volleyball", ("Women",)
    if event_name == SPORT_TYPE["SOCCER"]:
        return "Soccer", ("Coed", "Mixed")
    if event_name == SPORT_TYPE["BIBLE_CHALLENGE"]:
        return "Bible Challenge", ("Mixed",)
    return event_name, ()


def _extract_roster_team_codes(
    roster_rows: Optional[List[Dict[str, Any]]],
) -> Dict[str, set[str]]:
    """Return roster church/team codes by canonical event name."""
    result: Dict[str, set[str]] = defaultdict(set)
    if not roster_rows:
        return result

    event_names = []
    seen_events: set[str] = set()
    for config in _DEFAULT_SHEET_CONFIGS:
        event_name = str(config["event"])
        if event_name not in seen_events:
            event_names.append(event_name)
            seen_events.add(event_name)

    for event_name in event_names:
        target_type, target_genders = _event_target(event_name)
        accepted_genders = {gender.casefold() for gender in target_genders if gender}
        for row in roster_rows:
            r_type = _clean_text(row.get("sport_type"))
            r_gender = _clean_text(row.get("sport_gender"))
            r_format = _clean_text(row.get("sport_format"))
            if (
                r_type.casefold() != target_type.casefold()
                and r_type.casefold() != event_name.casefold()
            ):
                continue
            if accepted_genders and r_gender.casefold() not in accepted_genders:
                continue
            if r_format and r_format.casefold() != SPORT_FORMAT["TEAM"].casefold():
                continue
            code = _clean_team_code(row.get("Church Team"))
            if code:
                team_order = _clean_team_code(row.get("team_order"))
                result[event_name].add(code if not team_order else f"{code}-{team_order}")
    return result


def _active_sheet_configs(
    active_sheets: Optional[Iterable[str]] = None,
) -> List[Dict[str, Any]]:
    allowed = {name.strip() for name in active_sheets or [] if name and name.strip()}
    configs = []
    for config in _DEFAULT_SHEET_CONFIGS:
        if allowed:
            if config["sheet_name"] in allowed:
                configs.append(dict(config, active=True))
        elif config.get("active"):
            configs.append(dict(config))
    return configs


def _extract_pool_members(ws) -> Tuple[Dict[int, str], Dict[str, Dict[str, str]]]:
    """Return slot->pool_id and pool_id->{slot: team_code} from the sheet."""
    slot_to_pool: Dict[int, str] = {}
    pool_members: Dict[str, Dict[str, str]] = defaultdict(dict)
    max_row = ws.max_row

    heading_cells: List[Tuple[int, int, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            text = _clean_text(cell.value).upper()
            # Real pool headings live in the left/middle table.  The WVB sheet
            # repeats labels far right as visual markers; ignore those.
            if text.startswith("POOL ") and cell.column <= 10:
                heading_cells.append((cell.row, cell.column, text))

    heading_cells.sort()
    for idx, (heading_row, heading_col, label) in enumerate(heading_cells):
        next_heading_row = (
            heading_cells[idx + 1][0] if idx + 1 < len(heading_cells) else max_row + 1
        )
        pool_id = _pool_id_from_label(label)
        for row_idx in range(heading_row + 1, next_heading_row):
            team = _clean_team_code(ws.cell(row=row_idx, column=heading_col).value)
            if not team or team.startswith("POOL ") or team.startswith("UPDATE:"):
                continue
            slot = (
                _int_or_none(ws.cell(row=row_idx, column=heading_col + 1).value)
                or _int_or_none(ws.cell(row=row_idx, column=heading_col - 1).value)
            )
            if slot is None:
                continue
            slot_to_pool[slot] = pool_id
            pool_members[pool_id][str(slot)] = team

    return slot_to_pool, dict(pool_members)


def _matchup_values_from_row(
    values: List[Any],
    marker_idx: int,
    team_count: int,
) -> Tuple[List[Any], List[str]]:
    """Read slot/team values around a '>>' marker."""
    raw_slots = [
        values[marker_idx - (2 * (team_count - idx) - 1)]
        if marker_idx >= (2 * (team_count - idx) - 1)
        else None
        for idx in range(team_count)
    ]
    teams = [
        _clean_team_code(
            values[marker_idx + (2 * idx + 1)]
            if marker_idx + (2 * idx + 1) < len(values)
            else None
        )
        for idx in range(team_count)
    ]
    return raw_slots, teams


def _is_matchup_header(raw_slots: List[Any], teams: List[str]) -> bool:
    """Return True for visible worksheet header rows around a '>>' marker."""
    slot_texts = [_clean_text(value).casefold() for value in raw_slots]
    team_texts = [_clean_text(value).casefold() for value in teams]
    return (
        any(text.startswith("slot ") for text in slot_texts)
        and any(text.startswith("team ") for text in team_texts)
    )


def _pool_game_id(prefix: str, game_number: int, config: Dict[str, Any]) -> str:
    """Return the imported pool-game id matching the generated event style."""
    if str(config.get("game_id_style") or "") == "round_robin":
        return f"{prefix}-RR-{game_number}"
    return f"{prefix}-{game_number:02d}"


def _side_key(idx: int) -> str:
    return _TEAM_SIDES[idx]


def _parse_sheet(ws, config: Dict[str, Any]) -> Dict[str, Any]:
    event_name = str(config["event"])
    prefix = str(config["prefix"])
    resource_type = str(config["resource_type"])
    team_count = int(config.get("team_count") or 2)
    if team_count < 2 or team_count > len(_TEAM_SIDES):
        raise ValueError(f"Unsupported manual matchup team_count={team_count} for {ws.title}")
    solver_pool = _clean_text(config.get("solver_pool"))
    requires_pool = bool(config.get("requires_pool", False))
    duration = int(
        COURT_ESTIMATE_MINUTES_PER_GAME.get(event_name, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
    )
    expected_games = int(config["expected_games_per_team"])
    slot_to_pool, pool_members = _extract_pool_members(ws)

    games: List[Dict[str, Any]] = []
    skipped_byes: List[Dict[str, Any]] = []
    malformed_rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    matchup_seen: set[Tuple[str, ...]] = set()
    duplicate_matchups: List[Dict[str, Any]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row)
        for marker_idx, value in enumerate(values):
            if _clean_text(value) != ">>":
                continue
            raw_slots, teams = _matchup_values_from_row(values, marker_idx, team_count)
            if _is_matchup_header(raw_slots, teams):
                continue

            slots = [_int_or_none(raw_slot) for raw_slot in raw_slots]
            if any(slot is None for slot in slots) or any(not team for team in teams):
                malformed_rows.append({
                    "sheet": ws.title,
                    "row": row_idx,
                    "slots": raw_slots,
                    "teams": teams,
                    "reason": "missing slot/team around matchup marker",
                })
                continue

            if any(_is_bye(team) for team in teams):
                skipped_byes.append({
                    "sheet": ws.title,
                    "row": row_idx,
                    "slots": slots,
                    "teams": teams,
                })
                continue

            matchup_key = tuple(sorted(teams))
            if matchup_key in matchup_seen:
                duplicate_matchups.append({
                    "sheet": ws.title,
                    "row": row_idx,
                    "teams": teams,
                })
            matchup_seen.add(matchup_key)

            inferred_pools = [slot_to_pool.get(slot or -1, "") for slot in slots]
            nonempty_pools = [pool_id for pool_id in inferred_pools if pool_id]
            if requires_pool and len(set(nonempty_pools)) > 1:
                slot_text = "/".join(str(slot) for slot in slots)
                warnings.append(
                    f"{ws.title} row {row_idx}: slots {slot_text} map to "
                    f"{'/'.join(inferred_pools)}; using {nonempty_pools[0]}."
                )
            pool_id = nonempty_pools[0] if nonempty_pools else ""
            if requires_pool and not pool_id:
                warnings.append(
                    f"{ws.title} row {row_idx}: no pool could be inferred for "
                    f"slots {'/'.join(str(slot) for slot in slots)}."
                )

            game_id = _pool_game_id(prefix, len(games) + 1, config)
            game = {
                "game_id": game_id,
                "event": event_name,
                "stage": "Pool",
                "pool_id": pool_id,
                "round": len(games) + 1,
                "duration_minutes": duration,
                "resource_type": resource_type,
                "earliest_slot": None,
                "latest_slot": None,
                "x_manual_source_sheet": ws.title,
                "x_manual_source_row": row_idx,
                "x_manual_team_count": team_count,
            }
            if solver_pool:
                game["solver_pool"] = solver_pool
            for idx, (team, slot) in enumerate(zip(teams, slots)):
                side = _side_key(idx)
                game[f"team_{side}_id"] = f"{prefix}::{team}"
                game[f"team_{side}_label"] = team
                game[f"x_manual_slot_{side}"] = slot
            games.append(game)

    game_counts = Counter()
    imported_codes: set[str] = set()
    for game in games:
        for key in ("team_a_label", "team_b_label", "team_c_label"):
            code = _clean_team_code(game.get(key))
            if code:
                game_counts[code] += 1
                imported_codes.add(code)

    count_warnings = []
    for team_code in sorted(imported_codes):
        count = game_counts[team_code]
        if count != expected_games:
            count_warnings.append({
                "team": team_code,
                "actual_games": count,
                "expected_games": expected_games,
            })
            warnings.append(
                f"{ws.title}: {team_code} has {count} real game(s); expected "
                f"{expected_games} for {event_name}."
            )

    return {
        "sheet_name": ws.title,
        "event": event_name,
        "prefix": prefix,
        "resource_type": resource_type,
        "expected_games_per_team": expected_games,
        "team_count": team_count,
        "requires_pool": requires_pool,
        "games": games,
        "pool_members": pool_members,
        "team_game_counts": dict(sorted(game_counts.items())),
        "teams": sorted(imported_codes),
        "skipped_byes": skipped_byes,
        "malformed_rows": malformed_rows,
        "duplicate_matchups": duplicate_matchups,
        "count_warnings": count_warnings,
        "warnings": warnings,
    }


def build_manual_matchup_payload(
    workbook_path: Path,
    *,
    roster_rows: Optional[List[Dict[str, Any]]] = None,
    active_sheets: Optional[Iterable[str]] = None,
) -> Dict[str, Any]:
    """Parse and validate the manual team-sport matchup workbook."""
    workbook_path = Path(workbook_path)
    errors: List[str] = []
    warnings: List[str] = []
    sheet_configs = _active_sheet_configs(active_sheets)
    if not workbook_path.exists():
        errors.append(f"manual matchup workbook not found: {workbook_path}")
        return {
            "version": 1,
            "source_workbook": str(workbook_path),
            "imported_at": datetime.now().isoformat(timespec="seconds"),
            "events": [],
            "games": [],
            "validation": {"errors": errors, "warnings": warnings},
        }

    wb = load_workbook(workbook_path, data_only=True)
    events: List[Dict[str, Any]] = []
    all_games: List[Dict[str, Any]] = []

    for config in sheet_configs:
        sheet_name = str(config["sheet_name"])
        if sheet_name not in wb.sheetnames:
            errors.append(f"active manual matchup sheet {sheet_name!r} not found")
            continue
        parsed = _parse_sheet(wb[sheet_name], config)
        if parsed["malformed_rows"]:
            errors.append(
                f"{sheet_name}: {len(parsed['malformed_rows'])} malformed matchup row(s)"
            )
        if parsed["duplicate_matchups"]:
            errors.append(
                f"{sheet_name}: {len(parsed['duplicate_matchups'])} duplicate matchup(s)"
            )
        warnings.extend(parsed["warnings"])
        events.append(parsed)
        all_games.extend(parsed["games"])

    roster_codes = _extract_roster_team_codes(roster_rows)
    if roster_rows is None:
        warnings.append(
            "No roster context supplied; skipped roster-vs-matchup team-code validation."
        )
    elif not roster_rows:
        warnings.append(
            "Roster context was empty; skipped roster-vs-matchup team-code validation."
        )
    else:
        for event in events:
            event_name = event["event"]
            imported = set(event["teams"])
            roster = roster_codes.get(event_name, set())
            missing_in_roster = sorted(imported - roster)
            missing_in_matchups = sorted(roster - imported)
            event["missing_in_roster"] = missing_in_roster
            event["missing_in_matchups"] = missing_in_matchups
            if missing_in_roster:
                warnings.append(
                    f"{event['sheet_name']}: imported team code(s) missing from roster for "
                    f"{event_name}: {', '.join(missing_in_roster)}"
                )
            if missing_in_matchups:
                warnings.append(
                    f"{event['sheet_name']}: roster team code(s) absent from manual "
                    f"matchups for {event_name}: {', '.join(missing_in_matchups)}"
                )

    return {
        "version": 1,
        "source_workbook": str(workbook_path),
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "active_sheets": [config["sheet_name"] for config in sheet_configs],
        "events": events,
        "games": all_games,
        "validation": {
            "errors": errors,
            "warnings": warnings,
            "game_count": len(all_games),
            "event_count": len(events),
            "bye_count": sum(len(event["skipped_byes"]) for event in events),
        },
    }


def write_manual_matchup_sidecar(payload: Dict[str, Any], output_path: Path) -> None:
    """Write an imported manual matchup payload to JSON."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_manual_matchup_sidecar(sidecar_path: Optional[Path]) -> Optional[Dict[str, Any]]:
    """Load an imported manual matchup sidecar, returning None when absent."""
    if not sidecar_path:
        return None
    path = Path(sidecar_path)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(
            f"Could not parse manual team matchup sidecar '{path}': {exc}. "
            "Generated team-sport pairings will be used."
        )
        return None
    games = payload.get("games")
    if not isinstance(games, list):
        logger.warning(
            f"Manual team matchup sidecar '{path}' has no games list. "
            "Generated team-sport pairings will be used."
        )
        return None
    return payload


def summarize_payload_for_log(payload: Dict[str, Any]) -> List[str]:
    """Return concise human-readable summary lines."""
    lines: List[str] = []
    validation = payload.get("validation", {}) or {}
    lines.append(
        f"manual matchups: {validation.get('game_count', 0)} imported game(s), "
        f"{validation.get('bye_count', 0)} bye row(s) skipped"
    )
    for event in payload.get("events", []) or []:
        counts = event.get("team_game_counts", {}) or {}
        count_text = ", ".join(f"{team}={count}" for team, count in sorted(counts.items()))
        lines.append(
            f"{event.get('sheet_name')}: {event.get('event')} -> "
            f"{len(event.get('games', []) or [])} game(s); {count_text}"
        )
    return lines
