"""Import approved preliminary game workbooks for WordPress publication.

The approved workbooks are operator-authored source-of-truth sheets for event
day score entry.  This importer normalizes exact preliminary games into stable
game keys, writes an auditable sidecar, and emits a minimal schedule_input /
schedule_output pair that can be fed to publish-schedule.  It does not write to
WordPress directly.
"""
from __future__ import annotations

import copy
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import (
    POD_RESOURCE_TYPE_BADMINTON,
    POD_RESOURCE_TYPE_TABLE_TENNIS,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
    TEAM_RESOURCE_TYPE_SOCCER,
)
from scheduling import master_schedule
from scheduling import match_schedule_overrides

APPROVED_GAMES_SIDECAR_FILENAME = "approved_schedule_games.json"
APPROVED_GAMES_AUDIT_FILENAME = "approved_schedule_games.audit.json"
APPROVED_SCHEDULE_INPUT_FILENAME = "approved_schedule_input.json"
APPROVED_SCHEDULE_OUTPUT_FILENAME = "approved_schedule_output.json"

DEFAULT_MAIN_SCHEDULE_FILENAME = "2026 Main Schedule draft 11.xlsx"
DEFAULT_BADMINTON_FILENAME = "2026 VAY Badminton Schedule_draft_v1_10Jul2026.xlsx"
DEFAULT_SOCCER_FILENAME = "COED SOCCER SCHEDULE.xlsx"
DEFAULT_TABLE_TENNIS_FILENAME = "Schedule_Roster - Table Tennis (PingPong) 2026.xlsx"

_TEAM_CODE_RE = re.compile(r"^[A-Z0-9]{2,5}(?:-\d+)?$")
_DATE_TO_DAY = {
    "7/18": "Sat-1",
    "7/19": "Sun-1",
    "7/24": "Fri-1",
    "7/25": "Sat-2",
    "7/26": "Sun-2",
}
_EVENT_PREFIX = {
    SPORT_TYPE["BASKETBALL"]: "BBM",
    SPORT_TYPE["VOLLEYBALL_MEN"]: "VBM",
    SPORT_TYPE["VOLLEYBALL_WOMEN"]: "VBW",
    SPORT_TYPE["BIBLE_CHALLENGE"]: "BC",
    SPORT_TYPE["SOCCER"]: "SOC",
    SPORT_TYPE["BADMINTON"]: "BAD",
    SPORT_TYPE["TABLE_TENNIS"]: "TT",
    SPORT_TYPE["TABLE_TENNIS_35"]: "TT35",
}
_BADMINTON_CATEGORY_TO_PREFIX = {
    "Men's Doubles": "BAD-MD",
    "Women's Doubles": "BAD-WD",
    "Mixed Doubles": "BAD-XD",
}


def default_main_schedule_path(data_dir: Path) -> Path:
    return Path(data_dir) / DEFAULT_MAIN_SCHEDULE_FILENAME


def default_badminton_path(data_dir: Path) -> Path:
    return Path(data_dir) / DEFAULT_BADMINTON_FILENAME


def default_soccer_path(data_dir: Path) -> Path:
    return Path(data_dir) / DEFAULT_SOCCER_FILENAME


def default_table_tennis_path(data_dir: Path) -> Path:
    return Path(data_dir) / DEFAULT_TABLE_TENNIS_FILENAME


def default_sidecar_path(base_dir: Path) -> Path:
    return Path(base_dir) / APPROVED_GAMES_SIDECAR_FILENAME


def default_audit_path(base_dir: Path) -> Path:
    return Path(base_dir) / APPROVED_GAMES_AUDIT_FILENAME


def default_publish_input_path(base_dir: Path) -> Path:
    return Path(base_dir) / APPROVED_SCHEDULE_INPUT_FILENAME


def default_publish_output_path(base_dir: Path) -> Path:
    return Path(base_dir) / APPROVED_SCHEDULE_OUTPUT_FILENAME


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _team_code(value: Any) -> str:
    return _clean_text(value).upper()


def _is_bye(value: Any) -> bool:
    return _clean_text(value).casefold() == "bye"


def _cell_ref(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def _cell_range(sheet: str, row: int, start_col: int, end_col: int) -> str:
    return f"{sheet}!{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"


def _time_label(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, dt_time):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = _clean_text(value)
    match = re.fullmatch(r"(\d{1,2}):(\d{2})\s*([AP]M)?", text, re.I)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    meridiem = match.group(3)
    if meridiem:
        hour %= 12
        if meridiem.upper() == "PM":
            hour += 12
    return f"{hour:02d}:{minute:02d}"


def _date_to_day(value: Any) -> Optional[str]:
    text = _clean_text(value).upper()
    match = re.search(r"(\d{1,2}/\d{1,2})", text)
    if not match:
        return None
    return _DATE_TO_DAY.get(match.group(1))


def _slot(day: str, start_time: str) -> str:
    return f"{day}-{start_time}"


def _parse_minutes(value: str) -> int:
    hours, minutes = str(value).split(":")
    return int(hours) * 60 + int(minutes)


def _slot_fits_resource(resource: Mapping[str, Any], slot: str) -> bool:
    try:
        day, start_time = slot.rsplit("-", 1)
        open_min = _parse_minutes(str(resource.get("open_time")))
        close_min = _parse_minutes(str(resource.get("close_time")))
        start_min = _parse_minutes(start_time)
    except (ValueError, AttributeError):
        return False
    return str(resource.get("day") or "") == day and open_min <= start_min < close_min


def _lane_number(label: Any, fallback: str) -> int:
    text = _clean_text(label) or fallback
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else 999


def _resolve_lane_resource(
    resources: Sequence[Mapping[str, Any]],
    *,
    resource_type: str,
    slot: str,
    lane: Optional[int] = None,
) -> tuple[Optional[str], Optional[str]]:
    candidates = [
        r for r in resources
        if str(r.get("resource_type") or "") == resource_type
        and _slot_fits_resource(r, slot)
    ]
    candidates.sort(key=lambda r: (_lane_number(r.get("label"), str(r.get("resource_id"))), str(r.get("resource_id"))))
    if not candidates:
        return None, f"no {resource_type} resource has slot {slot}"
    if lane is not None:
        if 1 <= lane <= len(candidates):
            return str(candidates[lane - 1].get("resource_id")), None
        return None, f"visual lane {lane} exceeds {len(candidates)} {resource_type} resource(s) at {slot}"
    return str(candidates[0].get("resource_id")), None


def _source_hash(record: Mapping[str, Any]) -> str:
    subset = {
        key: record.get(key)
        for key in (
            "game_key", "event", "stage", "pool_id", "round_number",
            "sub_event", "team_a_key", "team_a_label", "team_b_key",
            "team_b_label", "team_c_key", "team_c_label", "team_ids_json",
            "resource_id", "scheduled_slot", "source_workbook",
            "source_sheet", "source_cell", "raw_source_text",
        )
    }
    canonical = json.dumps(subset, sort_keys=True, separators=(",", ":"), ensure_ascii=False)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _team_ids_json(*team_keys: Optional[str]) -> str:
    return json.dumps([team for team in team_keys if team], separators=(",", ":"))


def _existing_pair_lookup(games: Sequence[Mapping[str, Any]]) -> dict[tuple[str, frozenset[str]], Mapping[str, Any]]:
    lookup: dict[tuple[str, frozenset[str]], Mapping[str, Any]] = {}
    for game in games:
        event = _clean_text(game.get("event"))
        team_a = _team_code(game.get("team_a_label"))
        team_b = _team_code(game.get("team_b_label"))
        if event and team_a and team_b:
            lookup[(event, frozenset((team_a, team_b)))] = game
    return lookup


def _existing_triple_lookup(games: Sequence[Mapping[str, Any]]) -> dict[tuple[str, frozenset[str]], Mapping[str, Any]]:
    lookup: dict[tuple[str, frozenset[str]], Mapping[str, Any]] = {}
    for game in games:
        event = _clean_text(game.get("event"))
        teams = [
            _team_code(game.get("team_a_label")),
            _team_code(game.get("team_b_label")),
            _team_code(game.get("team_c_label")),
        ]
        if event and all(teams):
            lookup[(event, frozenset(teams))] = game
    return lookup


def _approved_record(
    *,
    game_key: str,
    event: str,
    stage: str,
    resource_type: str,
    duration_minutes: int,
    scheduled_slot: Optional[str],
    resource_id: Optional[str],
    source_workbook: Path,
    source_sheet: str,
    source_cell: str,
    raw_source_text: str,
    pool_id: str = "",
    round_number: Optional[int] = None,
    sub_event: str = "",
    team_a_key: Optional[str] = None,
    team_a_label: Optional[str] = None,
    team_b_key: Optional[str] = None,
    team_b_label: Optional[str] = None,
    team_c_key: Optional[str] = None,
    team_c_label: Optional[str] = None,
    x_extra: Optional[Mapping[str, Any]] = None,
) -> dict[str, Any]:
    record: dict[str, Any] = {
        "game_key": game_key,
        "event": event,
        "stage": stage,
        "pool_id": pool_id,
        "round_number": round_number,
        "sub_event": sub_event,
        "team_a_key": team_a_key,
        "team_a_label": team_a_label,
        "team_b_key": team_b_key,
        "team_b_label": team_b_label,
        "team_c_key": team_c_key,
        "team_c_label": team_c_label,
        "team_ids_json": _team_ids_json(team_a_key, team_b_key, team_c_key),
        "resource_type": resource_type,
        "duration_minutes": duration_minutes,
        "resource_id": resource_id,
        "scheduled_slot": scheduled_slot,
        "game_status": "scheduled",
        "source_workbook": str(source_workbook),
        "source_sheet": source_sheet,
        "source_cell": source_cell,
        "raw_source_text": raw_source_text,
    }
    if x_extra:
        record.update({f"x_{key}": value for key, value in x_extra.items()})
    record["source_hash"] = _source_hash(record)
    return record


def _record_from_existing_game(
    game: Mapping[str, Any],
    *,
    resource_id: Optional[str],
    scheduled_slot: str,
    source_workbook: Path,
    source_sheet: str,
    source_cell: str,
    raw_source_text: str,
) -> dict[str, Any]:
    event = _clean_text(game.get("event"))
    prefix = _EVENT_PREFIX.get(event, event.replace(" ", "-"))
    return _approved_record(
        game_key=_clean_text(game.get("game_id")),
        event=event,
        stage=_clean_text(game.get("stage")) or "Pool",
        pool_id=_clean_text(game.get("pool_id")),
        round_number=game.get("round") if isinstance(game.get("round"), int) else None,
        sub_event=_clean_text(game.get("division_id")),
        team_a_key=_clean_text(game.get("team_a_id")) or f"{prefix}::{_team_code(game.get('team_a_label'))}",
        team_a_label=_clean_text(game.get("team_a_label")) or None,
        team_b_key=_clean_text(game.get("team_b_id")) or f"{prefix}::{_team_code(game.get('team_b_label'))}",
        team_b_label=_clean_text(game.get("team_b_label")) or None,
        team_c_key=_clean_text(game.get("team_c_id")) or (f"{prefix}::{_team_code(game.get('team_c_label'))}" if game.get("team_c_label") else None),
        team_c_label=_clean_text(game.get("team_c_label")) or None,
        resource_type=_clean_text(game.get("resource_type")),
        duration_minutes=int(game.get("duration_minutes") or 60),
        resource_id=resource_id,
        scheduled_slot=scheduled_slot,
        source_workbook=source_workbook,
        source_sheet=source_sheet,
        source_cell=source_cell,
        raw_source_text=raw_source_text,
    )


def _parse_main_schedule(
    path: Path,
    *,
    games: Sequence[Mapping[str, Any]],
    resources: Sequence[Mapping[str, Any]],
    warnings: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    parsed = match_schedule_overrides.parse_match_schedule_overrides_workbook(path)
    pair_lookup = _existing_pair_lookup(games)
    triple_lookup = _existing_triple_lookup(games)
    records: list[dict[str, Any]] = []
    placeholders: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()

    for row in parsed.get("rows", []) or []:
        kind = row.get("kind")
        if kind not in {"two_team_game", "three_team_game", "block"}:
            continue
        if kind == "block":
            placeholders.append({
                "source_workbook": str(path),
                "source_sheet": row.get("source_sheet"),
                "source_cell": row.get("source_cell"),
                "raw_source_text": row.get("raw_value"),
                "classification": "placeholder_or_block",
            })
            continue

        resource_row = {
            "resource_type": row.get("resource_type"),
            "day": row.get("day"),
            "slot": row.get("slot"),
            "sport": row.get("sport"),
            "visual_venue": row.get("visual_venue"),
        }
        resource_id, reason = master_schedule._resolve_resource_id(resource_row, resources)
        if not resource_id:
            warnings.append(f"{row.get('source_cell')}: {reason or 'could not resolve resource'}")

        if kind == "two_team_game":
            event = row["event"]
            team_a, team_b = row["team_a"], row["team_b"]
            game = pair_lookup.get((event, frozenset((team_a, team_b))))
            if game:
                records.append(_record_from_existing_game(
                    game,
                    resource_id=resource_id,
                    scheduled_slot=row["slot"],
                    source_workbook=path,
                    source_sheet=row.get("source_sheet", ""),
                    source_cell=row.get("source_cell", ""),
                    raw_source_text=row.get("raw_value", ""),
                ))
                continue

            prefix = _EVENT_PREFIX.get(event, "GAME")
            counters[event] += 1
            warnings.append(
                f"{row.get('source_cell')}: {team_a} vs {team_b} has no generated game; created {prefix}-APP-{counters[event]:02d}"
            )
            records.append(_approved_record(
                game_key=f"{prefix}-APP-{counters[event]:02d}",
                event=event,
                stage="Pool",
                resource_type=row["resource_type"],
                duration_minutes=60,
                resource_id=resource_id,
                scheduled_slot=row["slot"],
                team_a_key=f"{prefix}::{team_a}",
                team_a_label=team_a,
                team_b_key=f"{prefix}::{team_b}",
                team_b_label=team_b,
                source_workbook=path,
                source_sheet=row.get("source_sheet", ""),
                source_cell=row.get("source_cell", ""),
                raw_source_text=row.get("raw_value", ""),
            ))
            continue

        event = row["event"]
        teams = list(row.get("teams") or [])
        game = triple_lookup.get((event, frozenset(teams)))
        if game:
            records.append(_record_from_existing_game(
                game,
                resource_id=resource_id,
                scheduled_slot=row["slot"],
                source_workbook=path,
                source_sheet=row.get("source_sheet", ""),
                source_cell=row.get("source_cell", ""),
                raw_source_text=row.get("raw_value", ""),
            ))
            continue

        counters[event] += 1
        game_key = f"BC-RR-APP-{counters[event]:02d}"
        warnings.append(f"{row.get('source_cell')}: BC teams {'/'.join(teams)} have no generated game; created {game_key}")
        records.append(_approved_record(
            game_key=game_key,
            event=event,
            stage="Pool",
            resource_type=TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
            duration_minutes=60,
            resource_id=resource_id,
            scheduled_slot=row["slot"],
            team_a_key=f"BC::{teams[0]}",
            team_a_label=teams[0],
            team_b_key=f"BC::{teams[1]}",
            team_b_label=teams[1],
            team_c_key=f"BC::{teams[2]}",
            team_c_label=teams[2],
            source_workbook=path,
            source_sheet=row.get("source_sheet", ""),
            source_cell=row.get("source_cell", ""),
            raw_source_text=row.get("raw_value", ""),
        ))

    return records, placeholders


def _badminton_team_category_map(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    if len(wb.worksheets) < 2:
        return {}
    ws = wb.worksheets[1]
    mapping: dict[str, str] = {}
    blocks = [
        (1, "Men's Doubles"),
        (6, "Women's Doubles"),
        (11, "Mixed Doubles"),
    ]
    for start_col, category in blocks:
        team_col = start_col + 3
        for row in range(3, ws.max_row + 1):
            code = _team_code(ws.cell(row=row, column=team_col).value)
            if code:
                mapping[code] = category
    return mapping


def _parse_badminton(
    path: Path,
    *,
    resources: Sequence[Mapping[str, Any]],
    warnings: list[str],
) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    team_categories = _badminton_team_category_map(path)
    records: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        start_time = _time_label(ws.cell(row=row, column=1).value)
        if not start_time:
            continue
        slot = _slot("Fri-1", start_time)
        for lane, start_col in enumerate((2, 6, 10), start=1):
            match_no = ws.cell(row=row, column=start_col).value
            team_a = _team_code(ws.cell(row=row, column=start_col + 1).value)
            middle = _clean_text(ws.cell(row=row, column=start_col + 2).value).casefold()
            team_b = _team_code(ws.cell(row=row, column=start_col + 3).value)
            if not match_no or middle != "v" or not team_a or not team_b:
                continue
            if _is_bye(team_a) or _is_bye(team_b):
                continue
            category = team_categories.get(team_a) or team_categories.get(team_b)
            if not category:
                warnings.append(f"{_cell_range(ws.title, row, start_col, start_col + 3)}: no Sheet2 category found for {team_a} vs {team_b}")
                category = "Unknown Doubles"
            prefix = _BADMINTON_CATEGORY_TO_PREFIX.get(category, "BAD-UNK")
            try:
                match_int = int(match_no)
            except (TypeError, ValueError):
                warnings.append(f"{_cell_ref(ws.title, row, start_col)}: badminton match number {match_no!r} is not numeric")
                continue
            resource_id, reason = _resolve_lane_resource(
                resources,
                resource_type=POD_RESOURCE_TYPE_BADMINTON,
                slot=slot,
                lane=lane,
            )
            if not resource_id:
                warnings.append(f"{_cell_range(ws.title, row, start_col, start_col + 3)}: {reason}")
            records.append(_approved_record(
                game_key=f"{prefix}-{match_int:02d}",
                event=SPORT_TYPE["BADMINTON"],
                stage="Pool",
                sub_event=category,
                round_number=match_int,
                resource_type=POD_RESOURCE_TYPE_BADMINTON,
                duration_minutes=20,
                resource_id=resource_id,
                scheduled_slot=slot,
                team_a_key=f"{prefix}::{team_a}",
                team_a_label=team_a,
                team_b_key=f"{prefix}::{team_b}",
                team_b_label=team_b,
                source_workbook=path,
                source_sheet=ws.title,
                source_cell=_cell_range(ws.title, row, start_col, start_col + 3),
                raw_source_text=f"{match_no}: {team_a} v {team_b}",
                x_extra={"court": lane},
            ))
    return records


def _parse_soccer(
    path: Path,
    *,
    resources: Sequence[Mapping[str, Any]],
    warnings: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    records: list[dict[str, Any]] = []
    placeholders: list[dict[str, Any]] = []
    for row in range(11, ws.max_row + 1):
        game_no = _clean_text(ws.cell(row=row, column=3).value)
        if not game_no:
            continue
        team_a = _clean_text(ws.cell(row=row, column=4).value)
        middle = _clean_text(ws.cell(row=row, column=5).value).casefold()
        team_b = _clean_text(ws.cell(row=row, column=6).value)
        raw = " ".join(part for part in (game_no, team_a, "v", team_b) if part)
        source_cell = _cell_range(ws.title, row, 1, 7)
        game_key = f"SOC-{game_no.split()[0].upper()}"
        if _is_bye(team_a) or _is_bye(team_b):
            placeholders.append({
                "game_key": game_key,
                "source_workbook": str(path),
                "source_sheet": ws.title,
                "source_cell": source_cell,
                "raw_source_text": raw,
                "classification": "bye",
            })
            continue
        day = _date_to_day(ws.cell(row=row, column=1).value)
        start_time = _time_label(ws.cell(row=row, column=2).value)
        exact_teams = bool(_TEAM_CODE_RE.match(team_a.upper()) and _TEAM_CODE_RE.match(team_b.upper()))
        if not day or not start_time or middle != "v" or not exact_teams:
            placeholders.append({
                "game_key": game_key,
                "source_workbook": str(path),
                "source_sheet": ws.title,
                "source_cell": source_cell,
                "raw_source_text": raw,
                "classification": "unresolved_soccer_placeholder",
            })
            continue
        slot = _slot(day, start_time)
        resource_id, reason = _resolve_lane_resource(
            resources,
            resource_type=TEAM_RESOURCE_TYPE_SOCCER,
            slot=slot,
            lane=1,
        )
        if not resource_id:
            warnings.append(f"{source_cell}: {reason}")
        records.append(_approved_record(
            game_key=game_key,
            event=SPORT_TYPE["SOCCER"],
            stage="Pool",
            round_number=int(re.sub(r"\D", "", game_no) or "0") or None,
            resource_type=TEAM_RESOURCE_TYPE_SOCCER,
            duration_minutes=60,
            resource_id=resource_id,
            scheduled_slot=slot,
            team_a_key=f"SOC::{team_a.upper()}",
            team_a_label=team_a.upper(),
            team_b_key=f"SOC::{team_b.upper()}",
            team_b_label=team_b.upper(),
            source_workbook=path,
            source_sheet=ws.title,
            source_cell=source_cell,
            raw_source_text=raw,
            x_extra={"referee": _clean_text(ws.cell(row=row, column=7).value)},
        ))
    return records, placeholders


def _side_key(prefix: str, label: str) -> str:
    compact = re.sub(r"[^A-Z0-9]+", "-", label.upper()).strip("-")
    return f"{prefix}::{compact}"


def _table_tennis_category(cell_text: str, column: int) -> tuple[str, str, str]:
    text = cell_text.upper()
    if "(35+)" in text:
        return SPORT_TYPE["TABLE_TENNIS_35"], "35+ Doubles", "TT-35P-D"
    if "(U35)" in text:
        return SPORT_TYPE["TABLE_TENNIS"], "U35 Doubles", "TT-U35-D"
    if column == 2:
        return SPORT_TYPE["TABLE_TENNIS"], "Women's Singles", "TT-W-S"
    return SPORT_TYPE["TABLE_TENNIS"], "Men's Singles", "TT-M-S"


def _strip_table_tennis_marker(text: str) -> str:
    return re.sub(r"^\((?:U35|35\+)\)\s*", "", text.strip(), flags=re.I)


def _parse_table_tennis(
    path: Path,
    *,
    resources: Sequence[Mapping[str, Any]],
    warnings: list[str],
    errors: list[str],
    waive_discrepancy: bool,
) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    records: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()
    roster_u35_codes = {
        _team_code(value).replace("(U35) ", "")
        for value in (ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1))
        if _clean_text(value).upper().startswith("(U35)")
    }
    schedule_u35_codes: set[str] = set()

    for row in range(3, ws.max_row + 1):
        start_time = _time_label(ws.cell(row=row, column=1).value)
        if not start_time:
            continue
        slot = _slot("Fri-1", start_time)
        for column in range(2, 6):
            raw = _clean_text(ws.cell(row=row, column=column).value)
            if not raw or " - " not in raw:
                continue
            if re.search(r"\b(?:SF|FINAL|3RD)\b", raw, re.I):
                continue
            left, right = [part.strip() for part in raw.split(" - ", 1)]
            team_a_label = _strip_table_tennis_marker(left)
            team_b_label = _strip_table_tennis_marker(right)
            if _is_bye(team_a_label) or _is_bye(team_b_label):
                continue
            event, sub_event, prefix = _table_tennis_category(raw, column)
            if prefix == "TT-U35-D":
                schedule_u35_codes.update({
                    _team_code(team_a_label),
                    _team_code(team_b_label),
                })
            counters[prefix] += 1
            resource_id, reason = _resolve_lane_resource(
                resources,
                resource_type=POD_RESOURCE_TYPE_TABLE_TENNIS,
                slot=slot,
                lane=column - 1,
            )
            if not resource_id:
                warnings.append(f"{_cell_ref(ws.title, row, column)}: {reason}")
            records.append(_approved_record(
                game_key=f"{prefix}-{counters[prefix]:02d}",
                event=event,
                stage="Pool",
                sub_event=sub_event,
                round_number=counters[prefix],
                resource_type=POD_RESOURCE_TYPE_TABLE_TENNIS,
                duration_minutes=20,
                resource_id=resource_id,
                scheduled_slot=slot,
                team_a_key=_side_key(prefix, team_a_label),
                team_a_label=team_a_label,
                team_b_key=_side_key(prefix, team_b_label),
                team_b_label=team_b_label,
                source_workbook=path,
                source_sheet=ws.title,
                source_cell=_cell_ref(ws.title, row, column),
                raw_source_text=raw,
                x_extra={"table": column - 1},
            ))

    missing_from_schedule = sorted(code for code in roster_u35_codes if code and code not in schedule_u35_codes)
    if "SBC" in missing_from_schedule and not waive_discrepancy:
        errors.append(
            "Table Tennis U35 roster includes SBC, but the schedule does not; "
            "rerun with --waive-table-tennis-discrepancy only after operator approval."
        )
    elif "SBC" in missing_from_schedule:
        warnings.append("Table Tennis U35 SBC roster/schedule discrepancy was explicitly waived.")
    return records


def _validate_records(records: Sequence[Mapping[str, Any]], errors: list[str], warnings: list[str]) -> None:
    seen_keys: dict[str, str] = {}
    seen_slots: dict[tuple[str, str], str] = {}
    for record in records:
        game_key = _clean_text(record.get("game_key"))
        source = _clean_text(record.get("source_cell"))
        if game_key in seen_keys:
            errors.append(f"{source}: duplicate game_key {game_key!r}; first seen at {seen_keys[game_key]}")
        else:
            seen_keys[game_key] = source

        teams = [record.get("team_a_key"), record.get("team_b_key"), record.get("team_c_key")]
        team_count = len([team for team in teams if team])
        event = _clean_text(record.get("event"))
        if event == SPORT_TYPE["BIBLE_CHALLENGE"] and team_count != 3:
            errors.append(f"{source}: Bible Challenge game {game_key} must have exactly three teams")
        elif event != SPORT_TYPE["BIBLE_CHALLENGE"] and team_count != 2:
            errors.append(f"{source}: game {game_key} must have exactly two teams")

        resource_id = _clean_text(record.get("resource_id"))
        slot = _clean_text(record.get("scheduled_slot"))
        if not resource_id or not slot:
            errors.append(f"{source}: game {game_key} has no resolved resource/slot")
            continue
        slot_key = (resource_id, slot)
        if slot_key in seen_slots:
            errors.append(
                f"{source}: {resource_id} at {slot} is double-booked by {game_key} "
                f"and {seen_slots[slot_key]}"
            )
        else:
            seen_slots[slot_key] = game_key

    by_event = Counter(str(record.get("event") or "") for record in records)
    if by_event.get(SPORT_TYPE["BADMINTON"], 0) != 39:
        warnings.append(
            f"Badminton imported {by_event.get(SPORT_TYPE['BADMINTON'], 0)} preliminary game(s); expected 39."
        )
    if by_event.get(SPORT_TYPE["SOCCER"], 0) != 6:
        warnings.append(
            f"Soccer imported {by_event.get(SPORT_TYPE['SOCCER'], 0)} exact preliminary game(s); expected 6."
        )


def _game_contract_row(record: Mapping[str, Any]) -> dict[str, Any]:
    game = {
        "game_id": record["game_key"],
        "event": record["event"],
        "stage": record.get("stage") or "Pool",
        "pool_id": record.get("pool_id") or "",
        "round": record.get("round_number"),
        "team_a_id": record.get("team_a_key"),
        "team_b_id": record.get("team_b_key"),
        "team_c_id": record.get("team_c_key"),
        "team_ids": [
            value for value in (
                record.get("team_a_key"),
                record.get("team_b_key"),
                record.get("team_c_key"),
            ) if value
        ],
        "team_a_label": record.get("team_a_label"),
        "team_b_label": record.get("team_b_label"),
        "team_c_label": record.get("team_c_label"),
        "duration_minutes": float(record.get("duration_minutes") or 60),
        "resource_type": record.get("resource_type"),
        "division_id": record.get("sub_event") or None,
        "x_approved_source": record.get("source_workbook"),
        "x_approved_sheet": record.get("source_sheet"),
        "x_approved_cell": record.get("source_cell"),
        "x_approved_hash": record.get("source_hash"),
    }
    return {key: value for key, value in game.items() if value is not None}


def _assignment_contract_row(record: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "game_id": record["game_key"],
        "event": record["event"],
        "stage": record.get("stage") or "Pool",
        "resource_type": record.get("resource_type"),
        "resource_id": record["resource_id"],
        "slot": record["scheduled_slot"],
        "duration_minutes": int(record.get("duration_minutes") or 60),
        "team_a_id": record.get("team_a_key"),
        "team_b_id": record.get("team_b_key"),
        "x_approved_source": record.get("source_workbook"),
        "x_approved_sheet": record.get("source_sheet"),
        "x_approved_cell": record.get("source_cell"),
        "x_approved_hash": record.get("source_hash"),
    }


def _build_publish_artifacts(
    records: Sequence[Mapping[str, Any]],
    source_schedule_input: Mapping[str, Any],
    imported_at: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    schedule_input = {
        "generated_at": imported_at,
        "games": [_game_contract_row(record) for record in records],
        "resources": copy.deepcopy(list(source_schedule_input.get("resources", []) or [])),
        "playoff_slots": [],
        "gym_modes": copy.deepcopy(source_schedule_input.get("gym_modes", {}) or {}),
        "approved_games": {
            "source": APPROVED_GAMES_SIDECAR_FILENAME,
            "game_count": len(records),
        },
    }
    schedule_output = {
        "solved_at": imported_at,
        "status": "FEASIBLE",
        "solver_wall_seconds": 0.0,
        "assignments": [_assignment_contract_row(record) for record in records],
        "unscheduled": [],
        "approved_games": {
            "source": APPROVED_GAMES_SIDECAR_FILENAME,
            "game_count": len(records),
        },
    }
    return schedule_input, schedule_output


def build_approved_games_payload(
    *,
    main_schedule_path: Path,
    badminton_path: Path,
    soccer_path: Path,
    table_tennis_path: Path,
    schedule_input: Mapping[str, Any],
    venue_input_path: Optional[Path] = None,
    waive_table_tennis_discrepancy: bool = False,
) -> dict[str, Any]:
    imported_at = datetime.now().isoformat(timespec="seconds")
    games = list(schedule_input.get("games", []) or [])
    resources = list(schedule_input.get("resources", []) or [])
    warnings: list[str] = []
    errors: list[str] = []
    placeholders: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []

    for path in (main_schedule_path, badminton_path, soccer_path, table_tennis_path):
        if not Path(path).exists():
            errors.append(f"source workbook not found: {path}")

    if not errors:
        main_records, main_placeholders = _parse_main_schedule(
            main_schedule_path, games=games, resources=resources, warnings=warnings,
        )
        records.extend(main_records)
        placeholders.extend(main_placeholders)
        records.extend(_parse_badminton(badminton_path, resources=resources, warnings=warnings))
        soccer_records, soccer_placeholders = _parse_soccer(
            soccer_path, resources=resources, warnings=warnings,
        )
        records.extend(soccer_records)
        placeholders.extend(soccer_placeholders)
        records.extend(_parse_table_tennis(
            table_tennis_path,
            resources=resources,
            warnings=warnings,
            errors=errors,
            waive_discrepancy=waive_table_tennis_discrepancy,
        ))
        _validate_records(records, errors, warnings)

    by_event = Counter(str(record.get("event") or "") for record in records)
    by_sub_event = Counter(
        f"{record.get('event')}::{record.get('sub_event') or record.get('stage')}"
        for record in records
    )
    schedule_input_artifact, schedule_output_artifact = _build_publish_artifacts(
        records, schedule_input, imported_at
    )
    return {
        "version": 1,
        "imported_at": imported_at,
        "sources": {
            "main_schedule": str(main_schedule_path),
            "badminton": str(badminton_path),
            "soccer": str(soccer_path),
            "table_tennis": str(table_tennis_path),
            "venue_input": str(venue_input_path) if venue_input_path else None,
        },
        "games": records,
        "placeholders": placeholders,
        "validation": {
            "errors": errors,
            "warnings": warnings,
            "counts_by_event": dict(sorted(by_event.items())),
            "counts_by_sub_event": dict(sorted(by_sub_event.items())),
            "placeholder_count": len(placeholders),
        },
        "publish_artifacts": {
            "schedule_input": schedule_input_artifact,
            "schedule_output": schedule_output_artifact,
        },
    }


def write_json(payload: Mapping[str, Any], path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def summarize_payload_for_log(payload: Mapping[str, Any]) -> list[str]:
    validation = payload.get("validation", {}) or {}
    counts = validation.get("counts_by_event", {}) or {}
    parts = [f"{event}={count}" for event, count in counts.items()]
    lines = [
        "approved games: "
        f"{len(payload.get('games', []) or [])} parsed exact game(s), "
        f"{validation.get('placeholder_count', 0)} placeholder/block row(s); "
        + (", ".join(parts) if parts else "no games"),
    ]
    lines.append(
        "approved games validation: "
        f"{len(validation.get('errors', []) or [])} error(s), "
        f"{len(validation.get('warnings', []) or [])} warning(s)"
    )
    return lines
