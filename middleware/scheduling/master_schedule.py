"""master_schedule - import the 2026 visual master schedule workbook.

The master schedule workbook is an operator-authored allocation sheet.  It is
not a roster or matchup source; it records when/where already-known games should
land.  This module parses that visual worksheet into a JSON sidecar and can
resolve confident rows into fixed scheduler assignments.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook

from config import (
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
    SPORT_TYPE,
    TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
)

MASTER_SCHEDULE_WORKBOOK_FILENAME = "VAY2026_Main_Schedule_draft_4.xlsx"
MASTER_SCHEDULE_SIDECAR_FILENAME = "manual_schedule_overrides.json"

_SHEET_NAME = "2026_Draft (4)"

_DATE_TO_DAY = {
    "7/18": "Sat-1",
    "7/19": "Sun-1",
    "7/24": "Fri-1",
    "7/25": "Sat-2",
    "7/26": "Sun-2",
}

_DATE_HEADER_RE = re.compile(r"^(?:SAT|SUN|FRI)\s+(\d{1,2}/\d{1,2})$", re.I)

_LEGEND_CELLS = {
    "B55": "Badminton",
    "F55": "Basketball",
    "J55": "MVB",
    "N55": "WVB",
    "R55": "BC",
    "V55": "Table Tennis",
}

_EVENT_BY_SPORT = {
    "Basketball": SPORT_TYPE["BASKETBALL"],
    "MVB": SPORT_TYPE["VOLLEYBALL_MEN"],
    "WVB": SPORT_TYPE["VOLLEYBALL_WOMEN"],
    "BC": SPORT_TYPE["BIBLE_CHALLENGE"],
}

_PREFIX_BY_SPORT = {
    "Basketball": "BBM",
    "MVB": "VBM",
    "WVB": "VBW",
}

_RESOURCE_TYPE_BY_SPORT = {
    "Basketball": GYM_RESOURCE_TYPE_BASKETBALL,
    "MVB": GYM_RESOURCE_TYPE_VOLLEYBALL,
    "WVB": GYM_RESOURCE_TYPE_VOLLEYBALL,
    "BC": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
}

_CONTROL_WORDS = {
    "CLOSED",
    "SET UP",
    "DINNER",
    "BREAK",
    "LUNCH / COMMUNITY GAMES",
    "TRACK & FIELD / TUG OF WAR",
}

_STAGE_TEXT_MAP = {
    "BB QF": ("Basketball", "BBM-QF"),
    "BB SF": ("Basketball", "BBM-Semi"),
    "BASKETBALL 3RD": ("Basketball", "BBM-3rd-Place"),
    "BASKETBALL FINAL": ("Basketball", "BBM-Final"),
    "MVB QF": ("MVB", "VBM-QF"),
    "MVB SF": ("MVB", "VBM-Semi"),
    "MVB 3RD": ("MVB", "VBM-3rd-Place"),
    "MVB FINAL": ("MVB", "VBM-Final"),
    "WVB QF": ("WVB", "VBW-QF"),
    "WVB SF": ("WVB", "VBW-Semi"),
    "WVB 3RD": ("WVB", "VBW-3rd-Place"),
    "WVB FINAL": ("WVB", "VBW-Final"),
    "BC SF 1": ("BC", "BC-Semi-1"),
    "BC SF 2": ("BC", "BC-Semi-2"),
    "BC SF 3": ("BC", "BC-Semi-3"),
}


def default_workbook_path(data_dir: Path) -> Path:
    return Path(data_dir) / MASTER_SCHEDULE_WORKBOOK_FILENAME


def default_sidecar_path(base_dir: Path) -> Path:
    return Path(base_dir) / MASTER_SCHEDULE_SIDECAR_FILENAME


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _team_code(value: Any) -> str:
    return _clean_text(value).upper()


def _fill_signature(cell) -> Tuple[str, Any, float]:
    color = cell.fill.fgColor
    key = color.rgb if color.type == "rgb" else color.theme
    return (str(color.type), key, round(float(color.tint or 0), 6))


def _has_visible_fill(cell) -> bool:
    return bool(cell.fill and cell.fill.fill_type)


def _time_label(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        minutes = int(
            round(
                value.hour * 60
                + value.minute
                + (value.second / 60.0)
                + (value.microsecond / 60_000_000.0)
            )
        )
    else:
        text = _clean_text(value)
        if not text or text.casefold() == "legend":
            return None
        parts = text.split(":")
        if len(parts) < 2:
            return None
        try:
            hour = int(parts[0])
            minute_float = float(parts[1])
            second_float = float(parts[2]) if len(parts) > 2 else 0.0
        except ValueError:
            return None
        minutes = int(round(hour * 60 + minute_float + (second_float / 60.0)))
    hour = (minutes // 60) % 24
    minute = minutes % 60
    return f"{hour:02d}:{minute:02d}"


def _date_header_day(value: Any) -> Optional[Tuple[str, str]]:
    text = _clean_text(value).upper()
    match = _DATE_HEADER_RE.match(text)
    if not match:
        return None
    date_key = match.group(1)
    day = _DATE_TO_DAY.get(date_key)
    if not day:
        return None
    return text, day


def _nearest_anchor_column(column: int, anchors: Sequence[int]) -> Optional[int]:
    eligible = [anchor for anchor in anchors if anchor <= column]
    return max(eligible) if eligible else None


def _column_lane(column: int) -> int:
    """Return the visual lane number within the six 4-column blocks."""
    if column < 2:
        return 0
    return ((column - 2) // 4) + 1


def _slot_label(day: str, start_time: str) -> str:
    return f"{day}-{start_time}"


def _normal_stage_text(value: Any) -> str:
    return " ".join(_clean_text(value).upper().split())


def _is_control_text(value: Any) -> bool:
    text = _normal_stage_text(value)
    if not text:
        return True
    if text in _CONTROL_WORDS:
        return True
    return any(marker in text for marker in ("CEREMONY", "SERVICE", "SCRIPTURE"))


def _stage_game_id_for_text(
    text: str,
    stage_counters: Counter[str],
) -> Optional[Tuple[str, str]]:
    mapped = _STAGE_TEXT_MAP.get(text)
    if not mapped:
        return None
    sport, base = mapped
    if base.endswith("-QF") or base.endswith("-Semi"):
        stage_counters[base] += 1
        return sport, f"{base}-{stage_counters[base]}"
    return sport, base


def _legend_map(ws) -> Dict[Tuple[str, Any, float], str]:
    mapping: Dict[Tuple[str, Any, float], str] = {}
    for coord, sport in _LEGEND_CELLS.items():
        mapping[_fill_signature(ws[coord])] = sport
    # First-weekend MVB cells use a theme fill while the legend cell uses RGB.
    if ws.max_row >= 5 and ws.max_column >= 11 and _has_visible_fill(ws["K5"]):
        mapping[_fill_signature(ws["K5"])] = "MVB"
    return mapping


def _extract_bc_row(
    ws,
    row_idx: int,
    day: str,
    start_time: str,
    visual_venue: str,
    legend: Mapping[Tuple[str, Any, float], str],
) -> Optional[Dict[str, Any]]:
    teams = [_team_code(ws.cell(row=row_idx, column=col).value) for col in (22, 23, 24)]
    if not all(re.fullmatch(r"[A-Z0-9]{3}", team or "") for team in teams):
        return None
    if not any(
        legend.get(_fill_signature(ws.cell(row=row_idx, column=col))) == "BC"
        for col in (22, 23, 24)
    ):
        return None
    return {
        "kind": "bc_three_team",
        "source_sheet": ws.title,
        "source_cell": f"V{row_idx}:X{row_idx}",
        "day": day,
        "start_time": start_time,
        "slot": _slot_label(day, start_time),
        "visual_venue": visual_venue or "BC [LIBRARY]",
        "visual_lane": 6,
        "sport": "BC",
        "event": SPORT_TYPE["BIBLE_CHALLENGE"],
        "resource_type": TEAM_RESOURCE_TYPE_BIBLE_CHALLENGE,
        "teams": teams,
        "raw_value": " / ".join(teams),
    }


def parse_master_schedule_workbook(workbook_path: Path) -> Dict[str, Any]:
    """Parse the visual master schedule workbook into raw override intent rows."""
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active
    legend = _legend_map(ws)

    rows: List[Dict[str, Any]] = []
    diagnostics: Dict[str, Any] = {
        "warnings": [],
        "errors": [],
        "unmapped_cells": [],
    }
    current_date = ""
    current_day = ""
    venue_by_anchor: Dict[int, str] = {}
    stage_counters: Counter[str] = Counter()
    seen_bc_rows: set[int] = set()

    for row_idx in range(1, ws.max_row + 1):
        header = _date_header_day(ws.cell(row=row_idx, column=1).value)
        if header:
            current_date, current_day = header
            venue_by_anchor = {
                col: _clean_text(ws.cell(row=row_idx, column=col).value)
                for col in (2, 6, 10, 14, 18, 22)
                if _clean_text(ws.cell(row=row_idx, column=col).value)
            }
            continue
        start_time = _time_label(ws.cell(row=row_idx, column=1).value)
        if not current_day or not start_time:
            continue

        bc_row = _extract_bc_row(
            ws,
            row_idx,
            current_day,
            start_time,
            venue_by_anchor.get(22, "BC [LIBRARY]"),
            legend,
        )
        if bc_row:
            rows.append(bc_row)
            seen_bc_rows.add(row_idx)

        for cell in ws[row_idx]:
            if cell.column < 2 or cell.column > 24:
                continue
            if row_idx in seen_bc_rows and cell.column in (22, 23, 24):
                continue
            raw_value = cell.value
            if raw_value in (None, ""):
                continue
            sport = legend.get(_fill_signature(cell))
            if not sport:
                continue
            text = _normal_stage_text(raw_value)
            if _is_control_text(raw_value):
                continue
            if row_idx == 55:
                continue

            anchor = _nearest_anchor_column(cell.column, sorted(venue_by_anchor))
            visual_venue = venue_by_anchor.get(anchor or 0, "")
            base_row = {
                "source_sheet": ws.title,
                "source_cell": cell.coordinate,
                "day": current_day,
                "start_time": start_time,
                "slot": _slot_label(current_day, start_time),
                "visual_venue": visual_venue,
                "visual_lane": _column_lane(cell.column),
                "sport": sport,
                "raw_value": raw_value,
            }

            if sport in _PREFIX_BY_SPORT and isinstance(raw_value, (int, float)) and not isinstance(raw_value, bool):
                number = int(raw_value)
                prefix = _PREFIX_BY_SPORT[sport]
                rows.append({
                    **base_row,
                    "kind": "numbered_game",
                    "event": _EVENT_BY_SPORT[sport],
                    "resource_type": _RESOURCE_TYPE_BY_SPORT[sport],
                    "game_id": f"{prefix}-{number:02d}",
                    "game_number": number,
                })
                continue

            stage_game = _stage_game_id_for_text(text, stage_counters)
            if stage_game:
                mapped_sport, game_id = stage_game
                rows.append({
                    **base_row,
                    "kind": "stage_game",
                    "sport": mapped_sport,
                    "event": _EVENT_BY_SPORT[mapped_sport],
                    "resource_type": _RESOURCE_TYPE_BY_SPORT[mapped_sport],
                    "game_id": game_id,
                })
                continue

            rows.append({
                **base_row,
                "kind": "block",
                "status": "unmapped_block",
            })
            diagnostics["unmapped_cells"].append(
                f"{cell.coordinate}: {raw_value!r} is a broad/manual block, not a game id"
            )

    diagnostics["row_count"] = len(rows)
    diagnostics["candidate_count"] = sum(1 for row in rows if row.get("kind") != "block")
    diagnostics["block_count"] = sum(1 for row in rows if row.get("kind") == "block")

    return {
        "version": 1,
        "source_workbook": str(workbook_path),
        "source_sheet": ws.title,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "diagnostics": diagnostics,
    }


def _game_lookup(games: Iterable[Mapping[str, Any]]) -> Tuple[Dict[str, Mapping[str, Any]], Dict[frozenset[str], str]]:
    by_id: Dict[str, Mapping[str, Any]] = {}
    bc_by_teams: Dict[frozenset[str], str] = {}
    for game in games:
        game_id = _clean_text(game.get("game_id"))
        if game_id:
            by_id[game_id] = game
        if game.get("event") == SPORT_TYPE["BIBLE_CHALLENGE"] and game.get("stage") == "Pool":
            labels = [
                _team_code(game.get("team_a_label") or str(game.get("team_a_id", "")).split("::")[-1]),
                _team_code(game.get("team_b_label") or str(game.get("team_b_id", "")).split("::")[-1]),
                _team_code(game.get("team_c_label") or str(game.get("team_c_id", "")).split("::")[-1]),
            ]
            if all(labels):
                bc_by_teams[frozenset(labels)] = game_id
    return by_id, bc_by_teams


def _resource_slots(resource: Mapping[str, Any]) -> List[str]:
    day = _clean_text(resource.get("day"))
    open_time = _clean_text(resource.get("open_time"))
    close_time = _clean_text(resource.get("close_time"))
    slot_minutes = int(resource.get("slot_minutes") or 60)
    if not day or not open_time or not close_time:
        return []
    start = _to_minutes(open_time)
    end = _to_minutes(close_time)
    if end <= start:
        return []
    return [f"{day}-{_from_minutes(minute)}" for minute in range(start, end, slot_minutes)]


def _to_minutes(value: str) -> int:
    hour, minute = value.split(":", maxsplit=1)
    return int(hour) * 60 + int(minute)


def _from_minutes(value: int) -> str:
    return f"{(value // 60) % 24:02d}:{value % 60:02d}"


def _resource_sort_key(resource: Mapping[str, Any]) -> Tuple[str, str]:
    return (_clean_text(resource.get("label")), _clean_text(resource.get("resource_id")))


def _visual_venue_lane(row: Mapping[str, Any]) -> Optional[int]:
    text = _clean_text(row.get("visual_venue")).upper()
    if not text:
        return None
    if row.get("sport") == "Basketball":
        match = re.search(r"\bBB\s*(\d+)\b", text)
    elif row.get("sport") in {"MVB", "WVB"}:
        match = re.search(r"\bVB\s*(\d+)\b", text)
    else:
        match = None
    if match:
        return int(match.group(1))
    return None


def _desired_lane(row: Mapping[str, Any]) -> Optional[int]:
    sport = row.get("sport")
    lane = int(row.get("visual_lane") or 0)
    venue_lane = _visual_venue_lane(row)
    if venue_lane is not None:
        return venue_lane
    if sport == "BC":
        return 1
    if row.get("day") == "Sun-2" and sport in {"MVB", "WVB"}:
        return 2 if lane == 1 else 1 if lane == 3 else None
    if sport == "Basketball" and lane in {1, 2}:
        return lane
    if sport in {"MVB", "WVB"}:
        volleyball_lane = lane - 2
        if 1 <= volleyball_lane <= 3:
            return volleyball_lane
    return None


def _resolve_resource_id(
    row: Mapping[str, Any],
    resources: Sequence[Mapping[str, Any]],
) -> Tuple[Optional[str], Optional[str]]:
    resource_type = _clean_text(row.get("resource_type"))
    day = _clean_text(row.get("day"))
    slot = _clean_text(row.get("slot"))
    candidates = sorted(
        [
            resource for resource in resources
            if _clean_text(resource.get("resource_type")) == resource_type
            and _clean_text(resource.get("day")) == day
        ],
        key=_resource_sort_key,
    )
    if not candidates:
        return None, f"no resources for {resource_type} on {day}"

    valid = [resource for resource in candidates if slot in _resource_slots(resource)]
    if not valid:
        return None, f"no {resource_type} resource has slot {slot}"

    desired = _desired_lane(row)
    if desired is not None and 1 <= desired <= len(candidates):
        resource = candidates[desired - 1]
        if slot in _resource_slots(resource):
            return _clean_text(resource.get("resource_id")), None
        return (
            None,
            f"visual lane {desired} resource "
            f"{_clean_text(resource.get('resource_id')) or '<unknown>'} lacks slot {slot}",
        )
    if desired is not None:
        return None, f"visual lane {desired} exceeds {len(candidates)} {resource_type} resource(s)"

    return _clean_text(valid[0].get("resource_id")), None


def resolve_master_schedule_payload(
    payload: Mapping[str, Any],
    games: Sequence[Mapping[str, Any]],
    resources: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Resolve parsed master schedule rows into fixed assignment rows."""
    game_by_id, bc_by_teams = _game_lookup(games)
    fixed: List[Dict[str, Any]] = []
    warnings: List[str] = []
    errors: List[str] = []
    unresolved: List[Dict[str, Any]] = []

    for row in payload.get("rows", []) or []:
        if not isinstance(row, dict) or row.get("kind") == "block":
            continue
        game_id = _clean_text(row.get("game_id"))
        if row.get("kind") == "bc_three_team":
            game_id = bc_by_teams.get(frozenset(_team_code(t) for t in row.get("teams", []) or []), "")
        if not game_id:
            unresolved.append(dict(row, reason="could not map row to a game_id"))
            continue
        game = game_by_id.get(game_id)
        if not game:
            unresolved.append(dict(row, game_id=game_id, reason="game_id not present in schedule_input games"))
            continue

        resource_id, reason = _resolve_resource_id(row, resources)
        if not resource_id:
            unresolved.append(dict(row, game_id=game_id, reason=reason or "could not resolve resource"))
            continue

        fixed.append({
            "game_id": game_id,
            "event": game.get("event", row.get("event", "")),
            "stage": game.get("stage", ""),
            "resource_id": resource_id,
            "slot": row.get("slot"),
            "duration_minutes": int(game.get("duration_minutes") or 60),
            "x_master_schedule_source": payload.get("source_workbook", ""),
            "x_master_schedule_sheet": row.get("source_sheet", payload.get("source_sheet", "")),
            "x_master_schedule_cell": row.get("source_cell", ""),
            "x_master_schedule_raw": _clean_text(row.get("raw_value")),
        })

    seen_games: Dict[str, str] = {}
    seen_slots: Dict[Tuple[str, str], str] = {}
    deduped: List[Dict[str, Any]] = []
    for row in fixed:
        game_id = _clean_text(row.get("game_id"))
        source = _clean_text(row.get("x_master_schedule_cell"))
        if game_id in seen_games:
            errors.append(f"{source}: duplicate master schedule pin for game {game_id}; first seen at {seen_games[game_id]}")
            continue
        key = (_clean_text(row.get("resource_id")), _clean_text(row.get("slot")))
        if key in seen_slots:
            errors.append(f"{source}: duplicate master schedule slot {key[0]} {key[1]}; first game {seen_slots[key]}")
            continue
        seen_games[game_id] = source
        seen_slots[key] = game_id
        deduped.append(row)

    for row in unresolved:
        game_hint = _clean_text(row.get("game_id"))
        game_text = f" game_id={game_hint}" if game_hint else ""
        warnings.append(
            f"{row.get('source_cell', '<unknown>')}: {row.get('reason', 'unresolved')}"
            f"{game_text} "
            f"(raw={row.get('raw_value')!r})"
        )

    diagnostics = {
        "fixed_count": len(deduped),
        "unresolved_count": len(unresolved),
        "warnings": warnings,
        "errors": errors,
        "unresolved": unresolved,
    }
    return deduped, diagnostics


def build_master_schedule_payload(
    workbook_path: Path,
    *,
    games: Optional[Sequence[Mapping[str, Any]]] = None,
    resources: Optional[Sequence[Mapping[str, Any]]] = None,
) -> Dict[str, Any]:
    payload = parse_master_schedule_workbook(workbook_path)
    if games is not None and resources is not None:
        fixed, diagnostics = resolve_master_schedule_payload(payload, games, resources)
        payload["fixed_assignments"] = fixed
        payload["validation"] = diagnostics
    return payload


def load_master_schedule_sidecar(path: Optional[Path]) -> Dict[str, Any]:
    if path is None:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def write_master_schedule_sidecar(payload: Mapping[str, Any], path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def summarize_payload_for_log(payload: Mapping[str, Any]) -> List[str]:
    rows = payload.get("rows", []) or []
    by_kind = Counter(str(row.get("kind")) for row in rows if isinstance(row, dict))
    lines = [
        "master schedule: "
        f"{len(rows)} parsed row(s); "
        + ", ".join(f"{kind}={count}" for kind, count in sorted(by_kind.items()))
    ]
    validation = payload.get("validation") or {}
    if validation:
        lines.append(
            "master schedule validation: "
            f"{validation.get('fixed_count', 0)} fixed assignment(s), "
            f"{validation.get('unresolved_count', 0)} unresolved row(s)"
        )
    return lines


def merge_master_schedule_into_playoff_slots(
    playoff_slots: Sequence[Mapping[str, Any]],
    payload: Mapping[str, Any],
    games: Sequence[Mapping[str, Any]],
    resources: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    if not payload:
        return [dict(row) for row in playoff_slots], {}
    fixed, diagnostics = resolve_master_schedule_payload(payload, games, resources)
    merged_by_game: Dict[str, Dict[str, Any]] = {
        _clean_text(row.get("game_id")): dict(row)
        for row in playoff_slots
        if _clean_text(row.get("game_id"))
    }
    for row in fixed:
        merged_by_game[_clean_text(row.get("game_id"))] = row
    return list(merged_by_game.values()), {
        "source_workbook": payload.get("source_workbook", ""),
        "source_sheet": payload.get("source_sheet", ""),
        "imported_at": payload.get("imported_at", ""),
        "parsed_row_count": len(payload.get("rows", []) or []),
        **diagnostics,
    }
