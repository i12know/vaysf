"""pool_assignment — Pool-Assignment sidecar state, seeding/draw logic, and
tab refresh extracted from ScheduleWorkbookBuilder.

Step 5 of the Issue #152 decomposition. Pure extraction, no behavior changes.

Functions that need builder state (class-level column/event definitions or
other helpers) take the ``ScheduleWorkbookBuilder`` ``builder`` (the original
``self``/``cls``) as their first parameter and reach everything through it.
"""
import json
import random
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
from loguru import logger
from config import (
    SPORT_FORMAT,
    COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
)


def _pool_assignments_sidecar_path(base_dir: Path) -> Path:
    """Return the default sidecar file used to persist editable pool seeds."""
    return Path(base_dir) / "pool_assignments.json"

def _normalize_pool_seed(value: Any) -> Optional[int]:
    """Normalize blank/zero-like seed input to None, else return a positive int."""
    if value in (None, "", "0", 0):
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None

def _positive_int_or_none(value: Any) -> Optional[int]:
    """Return a positive int when possible; otherwise None."""
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed > 0 else None

def _pool_assignment_event_prefix(builder, event_name: str) -> str:
    """Return the placeholder prefix used for one pool-assignment event."""
    for known_event, prefix in builder._POOL_ASSIGNMENT_EVENT_DEFS:
        if known_event == event_name:
            return prefix
    return event_name[:3].upper()

def _event_sort_index(builder, event_name: str) -> int:
    """Return a stable event ordering for the Pool-Assignment tab."""
    for idx, (known_event, _) in enumerate(builder._POOL_ASSIGNMENT_EVENT_DEFS):
        if known_event == event_name:
            return idx
    return len(builder._POOL_ASSIGNMENT_EVENT_DEFS)

def _load_pool_assignment_state(
    builder,
    sidecar_path: Optional[Path],
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """Load persisted seed/draw metadata keyed by (event, team_id)."""
    if not sidecar_path:
        return {}
    sidecar_path = Path(sidecar_path)
    if not sidecar_path.exists():
        return {}

    try:
        payload = json.loads(sidecar_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(
            f"Could not parse pool-assignment sidecar '{sidecar_path}': {exc}. "
            "Ignoring persisted seeds for this build."
        )
        return {}

    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        logger.warning(
            f"Pool-assignment sidecar '{sidecar_path}' has invalid rows content. "
            "Ignoring persisted seeds for this build."
        )
        return {}

    state: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        event_name = str(row.get("Event") or row.get("event") or "").strip()
        team_id = str(row.get("Team ID") or row.get("team_id") or "").strip()
        if not event_name or not team_id:
            continue
        state[(event_name, team_id)] = {
            "Seed": builder._normalize_pool_seed(row.get("Seed", row.get("seed"))),
            "Random Draw Order": builder._positive_int_or_none(
                row.get("Random Draw Order", row.get("random_draw_order"))
            ),
            "Notes": str(row.get("Notes", row.get("notes")) or "").strip(),
        }
    return state

def _write_pool_assignment_state(
    builder,
    sidecar_path: Path,
    rows: List[Dict[str, Any]],
) -> None:
    """Persist editable Pool-Assignment state to a JSON sidecar."""
    payload_rows: List[Dict[str, Any]] = []
    for row in rows:
        event_name = str(row.get("Event") or "").strip()
        team_id = str(row.get("Team ID") or "").strip()
        if not event_name or not team_id:
            continue
        payload_rows.append({
            "event": event_name,
            "church_code": str(row.get("Church Team") or "").strip(),
            "team_order": str(row.get("Team Order") or "").strip(),
            "team_id": team_id,
            "seed": builder._normalize_pool_seed(row.get("Seed")),
            "random_draw_order": builder._positive_int_or_none(
                row.get("Random Draw Order")
            ),
            "notes": str(row.get("Notes") or "").strip(),
        })

    sidecar_path = Path(sidecar_path)
    sidecar_path.parent.mkdir(parents=True, exist_ok=True)
    sidecar_path.write_text(
        json.dumps(
            {
                "version": 1,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
                "rows": payload_rows,
            },
            indent=2,
        ),
        encoding="utf-8",
    )

def _build_pool_assignment_base_rows(
    builder,
    roster_rows: List[Dict[str, Any]],
    persisted_state: Optional[Dict[Tuple[str, str], Dict[str, Any]]] = None,
) -> List[Dict[str, Any]]:
    """Build one Pool-Assignment row per eligible core gym team."""
    persisted_state = persisted_state or {}
    rows: List[Dict[str, Any]] = []

    for event_name, _prefix in builder._POOL_ASSIGNMENT_EVENT_DEFS:
        min_team_size = builder._get_min_team_size(event_name)
        target_type, target_gender, _target_format = builder._decompose_event_name(event_name)

        counts_by_key: Dict[Tuple[str, str], int] = {}
        for roster_row in roster_rows:
            r_type = str(roster_row.get("sport_type") or "").strip()
            r_gender = str(roster_row.get("sport_gender") or "").strip()
            r_format = str(roster_row.get("sport_format") or "").strip()
            if (
                r_type.casefold() != target_type.casefold()
                and r_type.casefold() != event_name.casefold()
            ):
                continue
            if target_gender and r_gender.casefold() != target_gender.casefold():
                continue
            if r_format and r_format.casefold() != SPORT_FORMAT["TEAM"].casefold():
                continue

            church_code = str(roster_row.get("Church Team") or "").strip().upper()
            if not church_code:
                continue
            team_order = str(roster_row.get("team_order") or "").strip().upper()
            counts_by_key[(church_code, team_order)] = (
                counts_by_key.get((church_code, team_order), 0) + 1
            )

        for (church_code, team_order), roster_count in sorted(counts_by_key.items()):
            if roster_count < min_team_size:
                continue

            team_id = church_code if not team_order else f"{church_code}-{team_order}"
            persisted = persisted_state.get((event_name, team_id), {})
            rows.append({
                "Event": event_name,
                "Church Team": church_code,
                "Team Order": team_order,
                "Team ID": team_id,
                "Team Label": team_id,
                "Team Source": "ExplicitTeamOrder" if team_order else "ChurchLevel",
                "Roster Count": roster_count,
                "Min Team Size": min_team_size,
                "Seed": persisted.get("Seed"),
                "Random Draw Order": persisted.get("Random Draw Order"),
                "Draw Position": None,
                "Pool ID": "",
                "Pool Slot": "",
                "Assignment Basis": "",
                "Notes": persisted.get("Notes", ""),
            })

    return rows

def _default_random_draw_orders(
    event_name: str,
    team_ids: List[str],
) -> Dict[str, int]:
    """Return a stable pseudo-random ordering for unseeded teams."""
    ordered_ids = sorted(team_ids)
    rng = random.Random(f"vaysf-pool-draw|{event_name}|{'|'.join(ordered_ids)}")
    rng.shuffle(ordered_ids)
    return {team_id: idx for idx, team_id in enumerate(ordered_ids, start=1)}

def _serpentine_pool_slots(pool_sizes: List[int]) -> List[Tuple[str, str]]:
    """Return pool slots in serpentine fill order."""
    slots: List[Tuple[str, str]] = []
    if not pool_sizes:
        return slots

    max_size = max(pool_sizes)
    for slot_idx in range(1, max_size + 1):
        eligible = [pool_idx for pool_idx, size in enumerate(pool_sizes, start=1) if size >= slot_idx]
        if slot_idx % 2 == 0:
            eligible = list(reversed(eligible))
        for pool_idx in eligible:
            slots.append((f"P{pool_idx}", f"T{slot_idx}"))
    return slots

def _pool_sizes_for_assignment(
    builder,
    event_name: str,
    n_teams: int,
) -> List[int]:
    """Return the pool sizes implied by the current placeholder-pool policy."""
    if n_teams < 2:
        return []

    prefix = builder._pool_assignment_event_prefix(event_name)
    gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(
        event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM
    )
    pairs = builder._make_pool_game_pairs(prefix, n_teams, gpg)
    slot_ids_by_pool: Dict[str, set] = {}
    for team_a_id, team_b_id, pool_id in pairs:
        for team_id in (team_a_id, team_b_id):
            match = re.search(r"-P\d+-T(\d+)$", str(team_id))
            if not match:
                continue
            slot_ids_by_pool.setdefault(pool_id, set()).add(int(match.group(1)))

    if not slot_ids_by_pool:
        return [n_teams]

    def _pool_key(pool_id: str) -> int:
        try:
            return int(str(pool_id).replace("P", ""))
        except ValueError:
            return 0

    return [
        len(slot_ids_by_pool[pool_id])
        for pool_id in sorted(slot_ids_by_pool.keys(), key=_pool_key)
    ]

def _apply_pool_assignments_to_rows(
    builder,
    rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Compute draw order and pool placement for Pool-Assignment rows."""
    normalized_rows = [dict(row) for row in rows]
    grouped_rows: Dict[str, List[Dict[str, Any]]] = {}
    for row in normalized_rows:
        event_name = str(row.get("Event") or "").strip()
        team_id = str(row.get("Team ID") or "").strip()
        if not event_name or not team_id:
            continue
        row["Seed"] = builder._normalize_pool_seed(row.get("Seed"))
        row["Random Draw Order"] = builder._positive_int_or_none(row.get("Random Draw Order"))
        row["Notes"] = str(row.get("Notes") or "").strip()
        grouped_rows.setdefault(event_name, []).append(row)

    output_rows: List[Dict[str, Any]] = []
    for event_name, _prefix in builder._POOL_ASSIGNMENT_EVENT_DEFS:
        event_rows = grouped_rows.get(event_name, [])
        if not event_rows:
            continue

        duplicate_seed_values = {
            seed
            for seed in {
                row.get("Seed")
                for row in event_rows
                if row.get("Seed") is not None
            }
            if sum(1 for row in event_rows if row.get("Seed") == seed) > 1
        }
        if duplicate_seed_values:
            duplicates = ", ".join(
                f"{seed}: "
                + "/".join(
                    sorted(
                        str(row.get("Team ID") or "").strip()
                        for row in event_rows
                        if row.get("Seed") == seed
                    )
                )
                for seed in sorted(duplicate_seed_values)
            )
            logger.warning(
                f"Pool-Assignment duplicate seeds detected for event '{event_name}': "
                f"{duplicates}"
            )

        existing_draw_orders = {
            row["Team ID"]: row["Random Draw Order"]
            for row in event_rows
            if builder._positive_int_or_none(row.get("Random Draw Order")) is not None
        }
        next_draw_order = max(existing_draw_orders.values(), default=0)
        missing_draw_ids = [
            str(row.get("Team ID") or "").strip()
            for row in event_rows
            if str(row.get("Team ID") or "").strip() not in existing_draw_orders
        ]
        if missing_draw_ids:
            for team_id in sorted(
                missing_draw_ids,
                key=lambda value: builder._default_random_draw_orders(event_name, missing_draw_ids)[value],
            ):
                next_draw_order += 1
                existing_draw_orders[team_id] = next_draw_order

        for row in event_rows:
            row["Random Draw Order"] = existing_draw_orders.get(
                str(row.get("Team ID") or "").strip()
            )

        seeded_rows = sorted(
            [row for row in event_rows if row.get("Seed") is not None],
            key=lambda row: (int(row["Seed"]), str(row.get("Team ID") or "")),
        )
        unseeded_rows = sorted(
            [row for row in event_rows if row.get("Seed") is None],
            key=lambda row: (
                int(row.get("Random Draw Order") or 0),
                str(row.get("Team ID") or ""),
            ),
        )
        ordered_rows = seeded_rows + unseeded_rows

        if len(ordered_rows) < 2:
            for draw_position, row in enumerate(ordered_rows, start=1):
                row["Draw Position"] = draw_position
                row["Pool ID"] = ""
                row["Pool Slot"] = ""
                row["Assignment Basis"] = "WaitingForMoreTeams"
            output_rows.extend(ordered_rows)
            continue

        pool_sizes = builder._pool_sizes_for_assignment(event_name, len(ordered_rows))
        slots = builder._serpentine_pool_slots(pool_sizes)
        if len(slots) != len(ordered_rows):
            logger.warning(
                f"Pool-assignment slot count mismatch for event '{event_name}': "
                f"{len(slots)} slots for {len(ordered_rows)} teams. Leaving pool cells blank."
            )
            for draw_position, row in enumerate(ordered_rows, start=1):
                row["Draw Position"] = draw_position
                row["Pool ID"] = ""
                row["Pool Slot"] = ""
                row["Assignment Basis"] = (
                    "SeededDuplicate"
                    if row.get("Seed") in duplicate_seed_values
                    else ("Seeded" if row.get("Seed") is not None else "RandomDraw")
                )
            output_rows.extend(ordered_rows)
            continue

        for draw_position, (row, slot) in enumerate(zip(ordered_rows, slots), start=1):
            row["Draw Position"] = draw_position
            row["Pool ID"] = slot[0]
            row["Pool Slot"] = slot[1]
            row["Assignment Basis"] = (
                "SeededDuplicate"
                if row.get("Seed") in duplicate_seed_values
                else ("Seeded" if row.get("Seed") is not None else "RandomDraw")
            )

        output_rows.extend(ordered_rows)

    return sorted(
        output_rows,
        key=lambda row: (
            builder._event_sort_index(str(row.get("Event") or "")),
            int(row.get("Draw Position") or 0),
            str(row.get("Team ID") or ""),
        ),
    )

def _build_pool_assignment_rows(
    builder,
    roster_rows: List[Dict[str, Any]],
    sidecar_path: Optional[Path],
) -> List[Dict[str, Any]]:
    """Build Pool-Assignment rows from roster data plus persisted seed state."""
    persisted_state = builder._load_pool_assignment_state(sidecar_path)
    base_rows = builder._build_pool_assignment_base_rows(roster_rows, persisted_state)
    return builder._apply_pool_assignments_to_rows(base_rows)

def _normalize_primary_sport_name(value: Any) -> str:
    """Normalize a declared primary sport value for conflict weighting."""
    return str(value or "").strip()

def _solver_team_id(builder, event_name: str, team_id: str) -> str:
    """Return an event-scoped internal team id safe for cross-sport solving."""
    return f"{builder._pool_assignment_event_prefix(event_name)}::{team_id}"

def _pool_assignment_placeholder_map(
    builder,
    pool_assignment_rows: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """Return {event: {PREFIX-Px-Ty: team metadata}} from assigned pool rows."""
    grouped: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for row in pool_assignment_rows:
        event_name = str(row.get("Event") or "").strip()
        pool_id = str(row.get("Pool ID") or "").strip()
        pool_slot = str(row.get("Pool Slot") or "").strip()
        team_id = str(row.get("Team ID") or "").strip()
        if not event_name or not pool_id or not pool_slot or not team_id:
            continue
        prefix = builder._pool_assignment_event_prefix(event_name)
        placeholder_id = f"{prefix}-{pool_id}-{pool_slot}"
        display_label = str(row.get("Team Label") or team_id).strip() or team_id
        grouped.setdefault(event_name, {})[placeholder_id] = {
            "solver_team_id": builder._solver_team_id(event_name, team_id),
            "display_label": display_label,
            "team_id": team_id,
            "pool_id": pool_id,
            "pool_slot": pool_slot,
        }
    return grouped

def _write_pool_assignment_tab(
    builder,
    ws,
    pool_assignment_rows: List[Dict[str, Any]],
) -> None:
    """Write the editable Pool-Assignment planning tab."""
    from openpyxl.styles import Alignment, Font

    columns = builder._POOL_ASSIGNMENT_COLUMNS
    ws.cell(row=1, column=1, value=builder._POOL_ASSIGNMENT_NOTE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    note_cell = ws.cell(row=1, column=1)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.font = Font(italic=True)
    ws.row_dimensions[1].height = 30
    ws.append(columns)
    for row in pool_assignment_rows:
        values = []
        for column in columns:
            value = row.get(column)
            if column == "Seed" and value in (None, 0):
                value = ""
            values.append(value)
        ws.append(values)

    builder._annotate_pool_assignment_tab(ws, len(columns))

def refresh_pool_assignments(
    builder,
    workbook_path: Path,
    output_path: Optional[Path] = None,
    sidecar_path: Optional[Path] = None,
) -> List[Dict[str, Any]]:
    """Refresh Pool-Assignment rows from an edited workbook and persist them."""
    from openpyxl import load_workbook

    workbook_path = Path(workbook_path)
    output_path = Path(output_path) if output_path else workbook_path
    effective_sidecar_path = (
        Path(sidecar_path)
        if sidecar_path is not None
        else builder._pool_assignments_sidecar_path(output_path.parent)
    )

    sheet_rows = builder._read_pool_assignment_rows(workbook_path)
    if not sheet_rows:
        raise ValueError(
            f"Workbook '{workbook_path}' has no usable Pool-Assignment sheet."
        )

    normalized_rows: List[Dict[str, Any]] = []
    for row in sheet_rows:
        event_name = str(row.get("Event") or "").strip()
        team_id = str(row.get("Team ID") or "").strip()
        if not event_name or not team_id:
            continue
        normalized_rows.append({
            "Event": event_name,
            "Church Team": str(row.get("Church Team") or "").strip(),
            "Team Order": str(row.get("Team Order") or "").strip(),
            "Team ID": team_id,
            "Team Label": str(row.get("Team Label") or team_id).strip(),
            "Team Source": str(row.get("Team Source") or "").strip(),
            "Roster Count": builder._positive_int_or_none(row.get("Roster Count")) or 0,
            "Min Team Size": builder._positive_int_or_none(row.get("Min Team Size")) or 0,
            "Seed": builder._normalize_pool_seed(row.get("Seed")),
            "Random Draw Order": builder._positive_int_or_none(row.get("Random Draw Order")),
            "Draw Position": builder._positive_int_or_none(row.get("Draw Position")),
            "Pool ID": str(row.get("Pool ID") or "").strip(),
            "Pool Slot": str(row.get("Pool Slot") or "").strip(),
            "Assignment Basis": str(row.get("Assignment Basis") or "").strip(),
            "Notes": str(row.get("Notes") or "").strip(),
        })

    refreshed_rows = builder._apply_pool_assignments_to_rows(normalized_rows)
    builder._write_pool_assignment_state(effective_sidecar_path, refreshed_rows)

    wb = load_workbook(workbook_path)
    if "Pool-Assignment" not in wb.sheetnames:
        raise ValueError(
            f"Workbook '{workbook_path}' does not contain a Pool-Assignment sheet."
        )

    sheet_index = wb.sheetnames.index("Pool-Assignment")
    wb.remove(wb["Pool-Assignment"])
    ws = wb.create_sheet(title="Pool-Assignment", index=sheet_index)
    builder._write_pool_assignment_tab(ws, refreshed_rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        f"Pool-Assignment tab refreshed: {len(refreshed_rows)} rows -> {output_path}"
    )
    logger.info(f"Pool-assignment sidecar written to: {effective_sidecar_path}")
    return refreshed_rows

def _read_pool_assignment_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Read Pool-Assignment while tolerating the operator note row above headers."""
    try:
        df = pd.read_excel(
            xlsx_path,
            sheet_name="Pool-Assignment",
            header=None,
            engine="openpyxl",
        )
    except Exception as e:
        logger.warning(f"Could not read 'Pool-Assignment' tab from {xlsx_path}: {e}")
        return []

    df = df.astype(object).where(pd.notna(df), None)
    header_idx: Optional[int] = None
    for idx in range(min(len(df.index), 3)):
        first_cell = str(df.iat[idx, 0] or "").strip()
        if first_cell == "Event":
            header_idx = idx
            break
    if header_idx is None:
        logger.warning(
            f"Could not find Pool-Assignment header row in {xlsx_path}; expected a row starting with 'Event'."
        )
        return []

    header_values = [
        str(value).strip() if value is not None else ""
        for value in df.iloc[header_idx].tolist()
    ]
    data_df = df.iloc[header_idx + 1 :].copy()
    data_df.columns = header_values
    data_df = data_df.loc[:, [column for column in header_values if column]]
    rows = data_df.to_dict("records")
    logger.debug(
        f"Read {len(rows)} rows from 'Pool-Assignment' tab of {xlsx_path} "
        f"(header row {header_idx + 1})"
    )
    return rows
