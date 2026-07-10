"""planning_tabs — offline planning-workbook tab renderers extracted from ScheduleWorkbookBuilder.

Step 4 of the Issue #152 decomposition. Pure extraction, no behavior changes.

Functions that need builder state (class methods, class-level header-note dicts,
or instance methods such as ``_compute_court_slots``) take the
``ScheduleWorkbookBuilder`` ``builder`` (the original ``self``/``cls``) as their
first parameter and reach everything through it. ``_write_summary_tab`` and
``_build_scenario_schedule`` have no builder dependency and stay free functions.
"""
from collections import deque
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from loguru import logger

from config import (
    COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME,
    COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME,
    COURT_ESTIMATE_POOL_GAMES_PER_TEAM,
    COURT_ESTIMATE_RACQUET_EVENTS,
    POD_FIT_COLOR_GREEN,
    POD_FIT_COLOR_RED,
    POD_FIT_COLOR_YELLOW,
    POD_FIT_YELLOW_MAX,
    POD_RESOURCE_EVENT_TYPE,
    SCHEDULE_SKETCH_COLOR_BASKETBALL,
    SCHEDULE_SKETCH_COLOR_HEADER,
    SCHEDULE_SKETCH_COLOR_SECTION,
    SCHEDULE_SKETCH_COLOR_VB_MEN,
    SCHEDULE_SKETCH_COLOR_VB_WOMEN,
    SCHEDULE_SKETCH_N_COURTS,
    SCHEDULE_SKETCH_SATURDAY_LAST_GAME,
    SCHEDULE_SKETCH_SATURDAY_START,
    SCHEDULE_SKETCH_SUNDAY_LAST_GAME,
    SCHEDULE_SKETCH_SUNDAY_START,
    SOCCER_ENABLED,
    SPORT_TYPE,
    VENUE_INPUT_FILENAME,
)

def _write_summary_tab(ws) -> None:
    """Write an operator-facing guide for using the planning workbook."""
    from openpyxl.styles import PatternFill, Font, Alignment

    title_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER, fill_type="solid")
    section_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    section_font = Font(bold=True)
    body_font = Font(size=11)
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    rows: List[Tuple[str, str]] = [
        (
            "What This Workbook Is",
            "This workbook is the Layer 1 planning aid for Sports Fest scheduling. "
            "Use it to estimate court demand, review pod/racquet entries, inspect "
            "resource IDs, and iterate on venue_input.xlsx before producing the final "
            "Layer 2 floor schedule.",
        ),
        (
            "What This Workbook Is Not",
            "This is not the final game timetable for coordinators on event day. "
            "The final Layer 2 output is VAYSF_Schedule_YYYY-MM-DD.xlsx, produced after "
            "running the solver.",
        ),
        (
            "Where To Start",
            "Work from the middleware folder. Start by refreshing the live exports with "
            "run-me.bat or, if you only need the scheduling artifacts, "
            "python main.py export-church-teams.",
        ),
        (
            "Required Roster Input",
            "build-schedule-workbook should read the consolidated ALL-church export "
            "(Church_Team_Status_ALL_YYYY-MM-DD.xlsx) as its roster and validation "
            "context. Do not point it at a single-church workbook if you want full "
            "Venue-Estimator / pod planning results.",
        ),
        (
            "Layer 1 Loop",
            "1. Edit middleware/data/venue_input.xlsx.\n"
            "2. If you need to change team-sport pool games per team, edit the "
            "COURT_ESTIMATE_POOL_GAMES_* constants in middleware/config.py. "
            "Supported live values are 2 or 3; Venue-Estimator is read-only.\n"
            "3. Run: python main.py export-church-teams\n"
            "4. Run: python main.py build-schedule-workbook "
            "--input-xlsx \"...\\Church_Team_Status_ALL_YYYY-MM-DD.xlsx\"\n"
            "   If omitted, the command tries to auto-detect the newest ALL workbook "
            "beside schedule_input.json or in EXPORT_DIR.\n"
            "5. Review the planning tabs in this workbook.\n"
            "6. Edit the Pool-Assignment tab if you want to seed BB/VBM/VBW/BC"
            f"{'/SOC' if SOCCER_ENABLED else ''} teams, then run:\n"
            "   python main.py assign-pools --workbook \"...\\Schedule_Workbook_YYYY-MM-DD.xlsx\"\n"
            "7. Repeat until venue capacity, seeding, pod planning, and resource IDs look right.",
        ),
        (
            "Layer 2 Commands",
            "When Layer 1 looks good, run Layer 2 from the middleware folder:\n"
            "run-schedule.bat\n"
            "Or run the two commands separately:\n"
            "python main.py solve-schedule\n"
            "python main.py produce-schedule",
        ),
        (
            "Tabs In This Workbook",
            "Summary: operator guide and command cheat sheet.\n"
            "Venue-Estimator: rough demand estimate for team/racquet sports.\n"
            "Pool-Assignment: editable BB/VBM/VBW/BC"
            f"{'/SOC' if SOCCER_ENABLED else ''} seed and pool-draw workspace.\n"
            "Pod-Divisions: planned pod divisions for racquet/pod events.\n"
            "Pod-Entries-Review: detailed entry review for pod sports.\n"
            "Court-Schedule-Sketch: quick planning sketch using Layer 1 assumptions.\n"
            "Pod-Resource-Estimate: compare pod demand against available venue resources.\n"
            "Schedule-Input: readable echo of schedule_input.json, including resource IDs.\n"
            "Gym-Allocation: Stage-A Layer 2 gym-mode allocation summary.",
        ),
        (
            "Most Important Checks",
            "Use Venue-Estimator and Pod-Resource-Estimate to see whether the booked venue "
            "is large enough. Use Schedule-Input to copy exact resource_id values into the "
            "Playoff-Slots tab of venue_input.xlsx. Use Gym-Allocation to confirm how gym "
            "time blocks are being assigned across basketball and volleyball modes. If you "
            "change team-sport Target Pool Games/Team from 2 to 3 in config.py, rebuild "
            "the workbook and rerun assign-pools because pool sizes and slot meanings can change.",
        ),
        (
            "Where To Read More",
            "For the full operator walkthrough, open docs/SCHEDULE-HOW-TO.md. "
            "For the deeper technical reference, open docs/SCHEDULING.md.",
        ),
    ]

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "VAY Sports Fest — Schedule Workbook Guide"
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = wrap_left
    ws.merge_cells("A2:B2")

    current_row = 4
    for heading, body in rows:
        head_cell = ws.cell(row=current_row, column=1, value=heading)
        head_cell.fill = section_fill
        head_cell.font = section_font
        head_cell.alignment = wrap_left

        body_cell = ws.cell(row=current_row, column=2, value=body)
        body_cell.font = body_font
        body_cell.alignment = wrap_left

        line_count = max(2, body.count("\n") + 1)
        ws.row_dimensions[current_row].height = max(24, 18 * line_count)
        current_row += 1


def _annotate_venue_estimator_tab(builder, ws, n_cols: int) -> None:
    """Add operator-facing comments to the Venue-Estimator header row."""
    builder._annotate_header_row(
        ws,
        1,
        n_cols,
        builder._VENUE_ESTIMATOR_HEADER_NOTES,
        width_map={
            "A": 26,
            "B": 16,
            "C": 18,
            "D": 20,
            "E": 18,
            "F": 18,
            "G": 18,
            "H": 12,
            "I": 16,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 12,
            "N": 16,
            "O": 16,
            "P": 20,
        },
        freeze_panes="A2",
        autofilter=True,
    )


def _annotate_pod_divisions_tab(builder, ws, n_cols: int) -> None:
    """Add operator-facing Excel comments and light usability affordances."""
    builder._annotate_header_row(
        ws,
        1,
        n_cols,
        builder._POD_DIVISION_HEADER_NOTES,
        width_map={
            "A": 20,
            "B": 16,
            "C": 14,
            "D": 14,
            "E": 18,
            "F": 16,
            "G": 16,
            "H": 17,
            "I": 18,
            "J": 14,
            "K": 16,
            "L": 24,
        },
        freeze_panes="A2",
        autofilter=True,
    )


def _annotate_pod_entries_review_tab(builder, ws, n_cols: int) -> None:
    """Add operator-facing comments to the Pod-Entries-Review header row."""
    builder._annotate_header_row(
        ws,
        1,
        n_cols,
        builder._POD_ENTRY_HEADER_NOTES,
        width_map={
            "A": 10,
            "B": 20,
            "C": 20,
            "D": 22,
            "E": 22,
            "F": 18,
            "G": 12,
            "H": 18,
            "I": 16,
            "J": 36,
        },
        freeze_panes="A2",
        autofilter=True,
    )


def _annotate_pool_assignment_tab(builder, ws, n_cols: int) -> None:
    """Add operator-facing comments to the Pool-Assignment header row."""
    builder._annotate_header_row(
        ws,
        2,
        n_cols,
        builder._POOL_ASSIGNMENT_HEADER_NOTES,
        width_map={
            "A": 24,
            "B": 14,
            "C": 12,
            "D": 14,
            "E": 16,
            "F": 18,
            "G": 12,
            "H": 12,
            "I": 10,
            "J": 18,
            "K": 12,
            "L": 10,
            "M": 10,
            "N": 18,
            "O": 28,
        },
        freeze_panes="A3",
        autofilter=True,
    )


def _write_schedule_input_tab(builder, ws, schedule_input: Dict[str, Any]) -> None:
    """Write Schedule-Input tab with Games, Resources, and Playoff-Slots sections."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER, fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    sec_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    sec_font = Font(bold=True)

    game_cols = [
        "game_id", "event", "stage", "pool_id", "round",
        "team_a_id", "team_b_id", "team_c_id", "duration_minutes",
        "resource_type", "earliest_slot", "latest_slot",
    ]
    resource_cols = [
        "resource_id", "resource_type", "label", "day",
        "open_time", "close_time", "slot_minutes", "exclusive_group",
    ]
    playoff_slot_cols = [
        "game_id", "event", "stage", "resource_id", "slot",
        "x_master_schedule_cell", "x_master_schedule_raw",
    ]
    precedence_cols = ["before_game_id", "after_game_id", "min_gap_slots"]

    current_row = 1

    # Meta row
    ws.cell(row=current_row, column=1, value="generated_at").font = sec_font
    ws.cell(row=current_row, column=2, value=schedule_input["generated_at"])
    ws.cell(
        row=current_row, column=3,
        value=f"Games: {schedule_input['game_count']}  Resources: {schedule_input['resource_count']}",
    )
    current_row += 2

    def _write_section(
        title: str,
        cols: List[str],
        rows: List[Dict],
        header_notes: Dict[str, str],
        section_note: str,
    ) -> None:
        nonlocal current_row
        sec_cell = ws.cell(row=current_row, column=1, value=title)
        sec_cell.fill = sec_fill
        sec_cell.font = sec_font
        builder._set_excel_comment(sec_cell, section_note)
        current_row += 1
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=current_row, column=c_idx, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            note = header_notes.get(col)
            if title == "GYM-MODES" and not note and col != "gym_name":
                note = (
                    f"Maximum concurrent {col} resources this gym can provide when allocated "
                    "to that mode."
                )
            builder._set_excel_comment(cell, note)
        ws.row_dimensions[current_row].height = max(ws.row_dimensions[current_row].height or 0, 30)
        current_row += 1
        for data_row in rows:
            for c_idx, col in enumerate(cols, start=1):
                ws.cell(row=current_row, column=c_idx, value=data_row.get(col))
            current_row += 1
        current_row += 1  # blank separator

    playoff_slots = schedule_input.get("playoff_slots", [])
    playoff_note_rows = (
        playoff_slots if playoff_slots
        else [{"game_id": "No playoff slots loaded — add Playoff-Slots tab to venue_input.xlsx"}]
    )

    gym_modes = schedule_input.get("gym_modes", {})
    gym_mode_rtypes = sorted({rt for caps in gym_modes.values() for rt in caps})
    gym_mode_cols = ["gym_name", *gym_mode_rtypes]
    gym_mode_rows = (
        [{"gym_name": name, **caps} for name, caps in sorted(gym_modes.items())]
        if gym_modes
        else [{"gym_name": "No Gym-Modes tab loaded — add Gym-Modes tab to venue_input.xlsx"}]
    )

    builder._set_excel_comment(
        ws.cell(row=1, column=1),
        "Timestamp when schedule_input.json was generated."
    )
    builder._set_excel_comment(
        ws.cell(row=1, column=3),
        "Quick counts of total games and resources in this schedule-input snapshot."
    )

    _write_section(
        "GAMES",
        game_cols,
        schedule_input["games"],
        builder._SCHEDULE_INPUT_GAME_HEADER_NOTES,
        "Game rows the Layer 2 solver must place into resource slots.",
    )
    _write_section(
        "RESOURCES",
        resource_cols,
        schedule_input["resources"],
        builder._SCHEDULE_INPUT_RESOURCE_HEADER_NOTES,
        "Resource rows available to the Layer 2 solver.",
    )
    _write_section(
        "PLAYOFF-SLOTS",
        playoff_slot_cols,
        playoff_note_rows,
        builder._SCHEDULE_INPUT_PLAYOFF_HEADER_NOTES,
        "Optional fixed-slot playoff constraints loaded from the Playoff-Slots tab in venue_input.xlsx.",
    )
    _write_section(
        "PRECEDENCE",
        precedence_cols,
        schedule_input.get("precedence", []),
        builder._SCHEDULE_INPUT_PRECEDENCE_HEADER_NOTES,
        "Optional ordering constraints between generated games. The after_game_id "
        "must start at least min_gap_slots after before_game_id.",
    )
    _write_section(
        "GYM-MODES",
        gym_mode_cols,
        gym_mode_rows,
        {"gym_name": "Venue block / gym name from the Gym-Modes sheet."},
        "Stage-A gym capability matrix showing which sport modes each grouped gym can host.",
    )

    # Column widths
    col_widths = [20, 30, 10, 10, 8, 16, 16, 16, 18, 22, 14, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def _write_gym_allocation_tab(builder, ws, gym_allocation: Optional[Dict[str, Any]]) -> None:
    """Write the Gym-Allocation tab summarising the Stage-A allocator output."""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER, fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    sec_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    sec_font = Font(bold=True)
    cur_row = [1]  # mutable so nested fn can advance it

    def _hrow(label: str) -> None:
        cell = ws.cell(row=cur_row[0], column=1, value=label)
        cell.fill = sec_fill
        cell.font = sec_font
        cur_row[0] += 1

    def _header_row(cols: List[str]) -> None:
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=cur_row[0], column=c_idx, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        cur_row[0] += 1

    def _data_row(cols: List[str], data: Dict) -> None:
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=cur_row[0], column=c_idx, value=data.get(col))
        cur_row[0] += 1

    source = gym_allocation.get("source") if gym_allocation else None
    if not gym_allocation or source in ("fallback", "direct_venue_input"):
        if not gym_allocation:
            message = "Gym allocation data not available."
        elif source == "fallback":
            message = (
                "Gym allocation not run — no Gym-Modes tab or no venue blocks with "
                "Exclusive Venue Group found in venue_input.xlsx.  "
                f"Fallback: {gym_allocation.get('gym_court_scenario', '?')} courts per session "
                "(SCHEDULE_SOLVER_GYM_COURTS)."
            )
        elif gym_allocation.get("reason") == "grouped_rows_without_gym_modes":
            message = (
                "Gym allocation not run — Venue-Input contains Exclusive Venue Group rows "
                "but no Gym-Modes tab. Using Venue-Input rows directly, so mutual exclusivity "
                "is not enforced."
            )
        elif gym_allocation.get("reason") == "gym_modes_without_grouped_rows":
            message = (
                "Gym allocation not run — Gym-Modes tab is present but no Exclusive Venue "
                "Group rows were found. Using Venue-Input rows directly."
            )
        else:
            message = "Gym allocation not run — using Venue-Input rows directly."
        ws.cell(
            row=1, column=1,
            value=message,
        )
        builder._set_excel_comment(
            ws.cell(row=1, column=1),
            "This tab only shows detailed allocation tables when grouped gym rows and Gym-Modes data are available."
        )
        ws.column_dimensions["A"].width = 80
        return

    source = gym_allocation.get("source", "unknown")
    ws.cell(row=cur_row[0], column=1, value=f"Source: {source}").font = sec_font
    ws.cell(row=cur_row[0], column=2, value=f"Mode switches: {gym_allocation.get('switch_count', '?')}")
    builder._set_excel_comment(
        ws.cell(row=cur_row[0], column=1),
        "Where this allocation summary came from. allocator means Stage-A Gym-Modes allocation was used."
    )
    builder._set_excel_comment(
        ws.cell(row=cur_row[0], column=2),
        "How many times the chosen gym mode switches between adjacent grouped venue blocks."
    )
    cur_row[0] += 2

    # Decisions
    _hrow("ALLOCATION DECISIONS")
    dec_cols = ["gym_name", "day", "open_time", "close_time", "mode", "courts", "slot_minutes"]
    _header_row(dec_cols)
    for idx, col in enumerate(dec_cols, start=1):
        builder._set_excel_comment(
            ws.cell(row=cur_row[0] - 1, column=idx),
            builder._GYM_ALLOCATION_DECISION_HEADER_NOTES.get(col),
        )
    ws.row_dimensions[cur_row[0] - 1].height = 30
    for dec in gym_allocation.get("decisions", []):
        _data_row(dec_cols, dec)
    cur_row[0] += 1

    # Demand vs supply
    _hrow("MODE DEMAND vs SUPPLY")
    ds_cols = ["mode", "demand", "supply", "shortfall"]
    _header_row(ds_cols)
    for idx, col in enumerate(ds_cols, start=1):
        builder._set_excel_comment(
            ws.cell(row=cur_row[0] - 1, column=idx),
            builder._GYM_ALLOCATION_SUPPLY_HEADER_NOTES.get(col),
        )
    ws.row_dimensions[cur_row[0] - 1].height = 30
    demand = gym_allocation.get("mode_demand", {})
    supply = gym_allocation.get("mode_supply", {})
    shortfall = gym_allocation.get("mode_shortfall", {})
    for mode in sorted(demand):
        _data_row(ds_cols, {
            "mode":      mode,
            "demand":    demand.get(mode, 0),
            "supply":    supply.get(mode, 0),
            "shortfall": shortfall.get(mode, 0),
        })

    col_widths = [22, 8, 10, 10, 22, 8, 14, 10, 10, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def _build_scenario_schedule(
    n_courts: int,
    pool_queues: List[List[str]],
    early_playoff_queues: List[List[str]],
    final_queues: List[List[str]],
    n_sat: int,
    n_sun: int,
) -> List[List[List[str]]]:
    """
    Build a 4-session court schedule for a given number of courts.

    Returns a list of 4 session grids:
        grids[0] = 1st Saturday  (n_sat time slots)
        grids[1] = 1st Sunday    (n_sun time slots)
        grids[2] = 2nd Saturday  (n_sat time slots)
        grids[3] = 2nd Sunday    (n_sun time slots)
    Each grid is grids[session][time_slot][court_index] = game_id or "".

    ── Court allocation ──────────────────────────────────────────────────
    Courts are divided into contiguous "primary blocks", one per sport,
    allocated proportionally.  Remainder courts go to the first sport(s)
    (i.e., Basketball gets an extra court before Volleyball does).

    Example with 5 courts and 3 sports:
        base = 5 // 3 = 1, extras = 5 % 3 = 2
        BBM → courts [0, 1]   (base 1 + 1 extra)
        VBM → courts [2, 3]   (base 1 + 1 extra)
        VBW → courts [4]      (base 1, no extra)

    Rationale: keeps each court dedicated to one sport type, so no
    net-height adjustment or equipment swap is needed mid-court.

    ── Phase 1 — Pool fill (sat1 → sun1 → sat2) ─────────────────────────
    For every time slot, courts are visited left-to-right (court 0, 1, …):

    • If the primary sport for that court still has pool games left,
      place the next game there (primary-first rule).
    • If the primary sport has finished its pool games, the court is
      idle.  The idle court is given to whichever sport currently has
      the most remaining pool games (greedy-most-needy rule).

    Effect in 5-court scenario (equal teams, 12 pool games each):
        Slots 0–5   : BBM fills courts 0-1, VBM fills courts 2-3,
                      VBW fills court 4  (all 3 sports running in parallel)
        Slot 6+     : BBM and VBM are done; their 4 courts become idle.
                      VBW still has 6 games → claims all 4 idle courts
                      plus its own, running 5 VBW games simultaneously.
                      VBW finishes at slot 7 (≈15:00) instead of slot 11
                      (≈19:00) — the whole church leaves ~4 hours earlier.

    Pool games never spill into sun2; that session is reserved for finals.

    ── Phase 2 — Early playoffs (QF + Semis) on sat2 ───────────────────
    After pool fill, each sport's empty cells in sat2 are collected in
    (time_slot, court) order.  Early-round playoff games (QF-1…4 if 8
    playoff teams, Semi-1 and Semi-2 otherwise) are placed there.

    Playoffs are placed on the sport's primary courts only — no court
    sharing — so the same nets and equipment remain in place.

    ── Phase 3 — Finals on sun2 ─────────────────────────────────────────
    Final and 3rd-place games are placed on each sport's empty cells in
    sun2, again primary courts only.  This guarantees that championship
    games always fall on the last day of the festival, regardless of how
    pool play distributes across the earlier sessions.

    ── Changing the algorithm ───────────────────────────────────────────
    • Court count scenarios: edit SCHEDULE_SKETCH_N_COURTS in config.py.
    • Session hours: edit SCHEDULE_SKETCH_SATURDAY_START / LAST_GAME and
      SCHEDULE_SKETCH_SUNDAY_START / LAST_GAME in config.py.
    • Court allocation order: the sport order in sport_defs inside
      _write_court_schedule_sketch controls which sport gets extra courts
      (earlier in the list = higher priority for extras).
    • Pool overflow policy: replace the greedy-most-needy rule (the
      `max(range(n_sports), key=lambda i: len(pool_remaining[i]))` line)
      with any other priority function — e.g., fixed sport priority,
      round-robin, or "same-sport block only" to revert to strict blocks.
    • Playoff session assignment: swap early_playoff_queues and
      final_queues arguments, or add a third category (e.g. Semis on
      sun1) by adding a new fill phase following the same pattern.
    """
    n_sports = len(pool_queues)
    n_slots = [n_sat, n_sun, n_sat, n_sun]
    grids: List[List[List[str]]] = [
        [[""] * n_courts for _ in range(n)] for n in n_slots
    ]

    # Court block allocation
    base = n_courts // n_sports
    extras = n_courts % n_sports
    court_blocks: List[List[int]] = []
    cur = 0
    for i in range(n_sports):
        k = base + (1 if i < extras else 0)
        court_blocks.append(list(range(cur, cur + k)))
        cur += k

    court_to_primary = {c: i for i, courts in enumerate(court_blocks) for c in courts}

    # Phase 1: pool fill — primary-first, then greedy-most-needy for idle courts
    pool_remaining = [deque(q) for q in pool_queues]
    for sess_idx in range(3):  # sat1, sun1, sat2
        for t in range(n_slots[sess_idx]):
            for c in range(n_courts):
                primary = court_to_primary[c]
                if pool_remaining[primary]:
                    grids[sess_idx][t][c] = pool_remaining[primary].popleft()
                else:
                    most_needy = max(range(n_sports), key=lambda i: len(pool_remaining[i]))
                    if pool_remaining[most_needy]:
                        grids[sess_idx][t][c] = pool_remaining[most_needy].popleft()

    # Phase 2: early playoffs (QF + Semi) on primary courts in sat2
    for early_q, courts in zip(early_playoff_queues, court_blocks):
        cells = [
            (2, t, c)
            for t in range(n_sat)
            for c in courts
            if not grids[2][t][c]
        ]
        for i, game_id in enumerate(early_q):
            if i < len(cells):
                s, t, c = cells[i]
                grids[s][t][c] = game_id

    # Phase 3: finals (Final + 3rd) on primary courts in sun2
    for final_q, courts in zip(final_queues, court_blocks):
        cells = [
            (3, t, c)
            for t in range(n_sun)
            for c in courts
            if not grids[3][t][c]
        ]
        for i, game_id in enumerate(final_q):
            if i < len(cells):
                s, t, c = cells[i]
                grids[s][t][c] = game_id

    return grids


def _write_court_schedule_sketch(
    builder, ws, roster_rows: List[Dict[str, Any]]
) -> None:
    """
    Write the Court-Schedule-Sketch tab.

    Three scenarios (3, 4, 5 courts) are rendered side-by-side on one
    worksheet, separated by an empty column.  Game IDs are sequential
    placeholders (BBM01…, VBM01…, VBW01…); no actual team assignments
    or conflict enforcement is performed here.  This is an Excel-only
    planning artifact — no data is written to WordPress sf_schedules.
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    mpg = COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME
    include_third = COURT_ESTIMATE_INCLUDE_THIRD_PLACE_GAME

    # Sports covered by this sketch (shared court type: basketball / volleyball)
    sport_defs = [
        (SPORT_TYPE["BASKETBALL"],       "BBM", SCHEDULE_SKETCH_COLOR_BASKETBALL),
        (SPORT_TYPE["VOLLEYBALL_MEN"],   "VBM", SCHEDULE_SKETCH_COLOR_VB_MEN),
        (SPORT_TYPE["VOLLEYBALL_WOMEN"], "VBW", SCHEDULE_SKETCH_COLOR_VB_WOMEN),
    ]

    # --- Compute game IDs per sport (per-sport pool games per team) ---
    sport_meta: Dict[str, Dict] = {}
    for event_name, prefix, color in sport_defs:
        min_sz = builder._get_min_team_size(event_name)
        counts = builder._count_estimating_teams(roster_rows, event_name, min_sz)
        n_teams = counts["n_estimating"] if counts["n_estimating"] >= 2 else 8
        gpg = COURT_ESTIMATE_POOL_GAMES_PER_TEAM.get(event_name, COURT_ESTIMATE_DEFAULT_POOL_GAMES_PER_TEAM)
        pool_plan = builder._summarize_pool_policy(n_teams, gpg)
        actual = len(builder._make_pool_game_pairs("_", n_teams, gpg))
        s = builder._compute_court_slots(
            n_teams,
            mpg,
            pool_games_per_team=gpg,
            actual_pool_games=actual,
            event_name=event_name,
        )
        early_ids, final_ids = builder._make_playoff_ids(
            prefix, s["playoff_teams"], include_third
        )
        sport_meta[event_name] = {
            "prefix": prefix,
            "color": color,
            "n_teams": n_teams,
            "target_pool_gpg": pool_plan["target_pool_games_per_team"],
            "actual_pool_gpg": pool_plan["actual_pool_games_per_team"],
            "pool_composition": pool_plan["pool_composition"],
            "pool_ids":   [f"{prefix}-{i:02d}" for i in range(1, s["pool_slots"] + 1)],
            "early_ids":  early_ids,   # QF + Semi → 2nd Saturday
            "final_ids":  final_ids,   # Final + 3rd → 2nd Sunday
        }

    # --- Per-sport game queues (pool overflow + dedicated playoff courts) ---
    pool_queues_by_sport          = [sport_meta[ev]["pool_ids"]  for ev, _, _ in sport_defs]
    early_playoff_queues_by_sport = [sport_meta[ev]["early_ids"] for ev, _, _ in sport_defs]
    final_queues_by_sport         = [sport_meta[ev]["final_ids"] for ev, _, _ in sport_defs]

    # --- Time slot helpers ---
    n_sat = SCHEDULE_SKETCH_SATURDAY_LAST_GAME - SCHEDULE_SKETCH_SATURDAY_START + 1
    n_sun = SCHEDULE_SKETCH_SUNDAY_LAST_GAME - SCHEDULE_SKETCH_SUNDAY_START + 1
    sat_times = [
        f"{h:02d}:00"
        for h in range(SCHEDULE_SKETCH_SATURDAY_START, SCHEDULE_SKETCH_SATURDAY_LAST_GAME + 1)
    ]
    sun_times = [
        f"{h:02d}:00"
        for h in range(SCHEDULE_SKETCH_SUNDAY_START, SCHEDULE_SKETCH_SUNDAY_LAST_GAME + 1)
    ]
    sessions = [
        ("1st Saturday", sat_times),
        ("1st Sunday",   sun_times),
        ("2nd Saturday", sat_times),
        ("2nd Sunday",   sun_times),
    ]

    # --- Styles ---
    section_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    hdr_fill     = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER,  fill_type="solid")
    hdr_font     = Font(bold=True, color="FFFFFF")
    bold_font    = Font(bold=True)
    center       = Alignment(horizontal="center", vertical="center")
    prefix_fill  = {
        "BBM": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_BASKETBALL, fill_type="solid"),
        "VBM": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_VB_MEN,     fill_type="solid"),
        "VBW": PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_VB_WOMEN,   fill_type="solid"),
    }

    # --- Column layout ---
    # Scenario A (3 cts): col 1=Time, 2-4=Courts → 4 cols; gap col 5
    # Scenario B (4 cts): col 6=Time, 7-10=Courts → 5 cols; gap col 11
    # Scenario C (5 cts): col 12=Time, 13-17=Courts → 6 cols
    n_courts_list = SCHEDULE_SKETCH_N_COURTS
    scenario_starts: List[int] = []
    cur_col = 1
    for n in n_courts_list:
        scenario_starts.append(cur_col)
        cur_col += (1 + n) + 1  # time col + court cols + gap

    INPUTS_ROW      = 1
    SCENARIO_HDR_ROW = 3
    COL_HDR_ROW     = 4
    DATA_START_ROW  = 5

    # --- Row 1: inputs summary (target/actual pool games per team) ---
    ws.cell(row=INPUTS_ROW, column=1, value="Inputs:").font = bold_font
    builder._set_excel_comment(
        ws.cell(row=INPUTS_ROW, column=1),
        "Planning assumptions used to generate this Layer 1 sketch, including pool-game targets and minutes per game."
    )
    col = 2
    for ev, prefix, _ in sport_defs:
        meta = sport_meta[ev]
        ws.cell(
            row=INPUTS_ROW,
            column=col,
            value=(
                f"{prefix} pool target/actual: "
                f"{meta['target_pool_gpg']}/{meta['actual_pool_gpg']}"
            ),
        )
        col += 3
    ws.cell(row=INPUTS_ROW, column=col,     value=f"Minutes/game: {mpg}")
    ws.cell(row=INPUTS_ROW, column=col + 3, value=f"3rd place: {'Yes' if include_third else 'No'}")

    # --- Row 2: per-sport game counts ---
    ws.cell(row=2, column=1, value="Game totals:").font = bold_font
    builder._set_excel_comment(
        ws.cell(row=2, column=1),
        "High-level game totals used in this what-if sketch. These are planning placeholders, not final assignments."
    )
    col_offset = 2
    for ev, prefix, _ in sport_defs:
        meta = sport_meta[ev]
        total = len(meta["pool_ids"]) + len(meta["early_ids"]) + len(meta["final_ids"])
        label = (
            f"{prefix}: {meta['n_teams']} teams, {total} games "
            f"({len(meta['pool_ids'])} pool, pools {meta['pool_composition']})"
        )
        ws.cell(row=2, column=col_offset, value=label)
        col_offset += 5

    # --- Pre-compute per-scenario schedules ---
    scenario_grids: Dict[int, List[List[List[str]]]] = {}
    for n_courts in n_courts_list:
        scenario_grids[n_courts] = builder._build_scenario_schedule(
            n_courts,
            pool_queues_by_sport,
            early_playoff_queues_by_sport,
            final_queues_by_sport,
            n_sat, n_sun,
        )

    # --- Render scenario headers and column headers ---
    for n_courts, start_col in zip(n_courts_list, scenario_starts):
        end_col = start_col + n_courts  # time col + n court cols (inclusive)
        # Scenario header
        sc_cell = ws.cell(row=SCENARIO_HDR_ROW, column=start_col, value=f"Scenario: {n_courts} Courts")
        sc_cell.font = hdr_font
        sc_cell.fill = hdr_fill
        sc_cell.alignment = center
        builder._set_excel_comment(
            sc_cell,
            f"What-if sketch assuming {n_courts} simultaneous shared gym courts across Basketball, Volleyball Men, and Volleyball Women."
        )
        ws.merge_cells(
            start_row=SCENARIO_HDR_ROW, start_column=start_col,
            end_row=SCENARIO_HDR_ROW,   end_column=end_col,
        )
        # Column sub-headers
        t_cell = ws.cell(row=COL_HDR_ROW, column=start_col, value="Time")
        t_cell.font = bold_font
        builder._set_excel_comment(
            t_cell,
            "Start time for the slot within the session block."
        )
        for c in range(n_courts):
            ct_cell = ws.cell(row=COL_HDR_ROW, column=start_col + 1 + c, value=f"Court {c + 1}")
            ct_cell.font = bold_font
            ct_cell.alignment = center
            builder._set_excel_comment(
                ct_cell,
                "Placeholder court lane in this scenario. Colored BBM/VBM/VBW IDs are Layer 1 planning placeholders, not final team assignments."
            )

    # --- Render session sections and time-slot rows ---
    current_row = DATA_START_ROW
    for sess_idx, (sess_label, times) in enumerate(sessions):
        # Section header
        for n_courts, start_col in zip(n_courts_list, scenario_starts):
            end_col = start_col + n_courts
            sh_cell = ws.cell(row=current_row, column=start_col, value=sess_label)
            sh_cell.fill = section_fill
            sh_cell.font = bold_font
            sh_cell.alignment = center
            ws.merge_cells(
                start_row=current_row, start_column=start_col,
                end_row=current_row,   end_column=end_col,
            )
        current_row += 1

        # Time slot rows
        for t, time_str in enumerate(times):
            for n_courts, start_col in zip(n_courts_list, scenario_starts):
                ws.cell(row=current_row, column=start_col, value=time_str)
                grid = scenario_grids[n_courts]
                for c in range(n_courts):
                    game_id = grid[sess_idx][t][c]
                    cell = ws.cell(row=current_row, column=start_col + 1 + c, value=game_id)
                    if game_id:
                        fill = prefix_fill.get(game_id.split("-")[0])
                        if fill:
                            cell.fill = fill
            current_row += 1

    # --- Column widths ---
    from openpyxl.utils import get_column_letter
    for n_courts, start_col in zip(n_courts_list, scenario_starts):
        ws.column_dimensions[get_column_letter(start_col)].width = 10      # Time
        for c in range(n_courts):
            ws.column_dimensions[get_column_letter(start_col + 1 + c)].width = 12  # Courts
    ws.freeze_panes = "A5"

    total_pool  = sum(len(q) for q in pool_queues_by_sport)
    total_early = sum(len(q) for q in early_playoff_queues_by_sport)
    total_final = sum(len(q) for q in final_queues_by_sport)
    logger.debug(
        f"Court-Schedule-Sketch tab: {total_pool} pool + {total_early} early-playoff "
        f"+ {total_final} finals across {len(n_courts_list)} scenarios."
    )


def _build_pod_resource_rows(
    builder,
    roster_rows: List[Dict[str, Any]],
    available_by_resource: Dict[str, int],
) -> List[Dict[str, Any]]:
    """Build Pod-Resource-Estimate output rows.

    Required slots use single-elimination: entries - 1
    (doubles counted as complete pairs, same as _count_racquet_entries).
    """
    rows = []
    for sport_name in COURT_ESTIMATE_RACQUET_EVENTS:
        counts = builder._count_racquet_entries(roster_rows, sport_name)
        n = counts["n_estimating"]
        resource_type = POD_RESOURCE_EVENT_TYPE.get(sport_name, "")
        required = max(0, n - 1)
        available = available_by_resource.get(resource_type, 0)
        surplus = available - required
        if not available_by_resource:
            fit_status = "No venue data"
        elif surplus >= 0:
            fit_status = "Green"
        elif surplus >= -POD_FIT_YELLOW_MAX:
            fit_status = "Yellow"
        else:
            fit_status = "Red"
        rows.append({
            "Event":              sport_name,
            "Resource Type":      resource_type,
            "Entries / Teams":    n,
            "Required Slots":     required,
            "Available Slots":    available,
            "Surplus / Shortage": surplus,
            "Fit Status":         fit_status,
        })
    return rows


def _write_pod_resource_estimate(
    builder,
    ws,
    pod_rows: List[Dict[str, Any]],
    available_by_resource: Dict[str, int],
    availability_source_label: str = VENUE_INPUT_FILENAME,
) -> None:
    """Write Pod-Resource-Estimate tab content with colour-coded Fit Status."""
    from openpyxl.styles import PatternFill, Font, Alignment

    cols = ["Event", "Resource Type", "Entries / Teams",
            "Required Slots", "Available Slots", "Surplus / Shortage", "Fit Status"]

    header_fill = PatternFill("solid", fgColor=SCHEDULE_SKETCH_COLOR_HEADER)
    header_font = Font(color="FFFFFF", bold=True)

    # Header row
    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        builder._set_excel_comment(cell, builder._POD_RESOURCE_HEADER_NOTES.get(col))
    ws.row_dimensions[1].height = 30

    fit_colors = {
        "Green":  POD_FIT_COLOR_GREEN,
        "Yellow": POD_FIT_COLOR_YELLOW,
        "Red":    POD_FIT_COLOR_RED,
    }

    if not available_by_resource:
        notice = (
            "No venue input loaded — "
            f"create {VENUE_INPUT_FILENAME} from the template and re-run the export"
        )
        if availability_source_label != VENUE_INPUT_FILENAME:
            notice = f"No availability data loaded from {availability_source_label}."
        ws.cell(row=2, column=1, value=notice)
        for c_idx, col in enumerate(cols, start=1):
            row_cell = ws.cell(row=2, column=c_idx)
            if c_idx == 1:
                row_cell.value = notice
            else:
                row_cell.value = None

    for r_idx, row in enumerate(pod_rows, start=2):
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            if col in ("Entries / Teams", "Required Slots", "Available Slots",
                       "Surplus / Shortage"):
                cell.alignment = Alignment(horizontal="right")
            if col == "Fit Status":
                color = fit_colors.get(row["Fit Status"])
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    for letter in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[letter].width = 16
    ws.freeze_panes = "A2"

    # Snapshot note
    note_row = len(pod_rows) + 3
    ws.cell(
        row=note_row, column=1,
        value=(
            f"Available slots loaded from {availability_source_label}. "
            "Required = entries − 1 (single elimination). "
            f"Green ≥ 0 | Yellow short 1–{POD_FIT_YELLOW_MAX} | Red short {POD_FIT_YELLOW_MAX + 1}+."
        ),
    )
    if availability_source_label != VENUE_INPUT_FILENAME:
        ws.cell(
            row=note_row + 1,
            column=1,
            value=(
                f"Offline build: available slots derived from {availability_source_label}."
            ),
        )
    logger.debug(f"Pod-Resource-Estimate tab: {len(pod_rows)} rows.")

