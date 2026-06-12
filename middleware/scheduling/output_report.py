"""output_report — schedule output Excel report builders extracted from ScheduleWorkbookBuilder.

Pure functions; no class state.  Extracted as part of Issue #152.
"""
from pathlib import Path
from typing import Any, Dict, List, Tuple

from loguru import logger

from config import (
    SCHEDULE_SKETCH_COLOR_SECTION,
    SCHEDULE_SKETCH_COLOR_HEADER,
    GYM_RESOURCE_TYPE,
    GYM_RESOURCE_TYPE_BASKETBALL,
    GYM_RESOURCE_TYPE_VOLLEYBALL,
)
from schedule_styles import (
    SPORT_STYLES,
    CATEGORY_STYLES,
    style_for_game,
    sport_style,
    format_compact_label,
    category_style,
)
from scheduling.xlsx_utils import (
    _day_display_label,
    _day_sort_key,
    _make_excel_note_shapes_visible,
    _stamp_known_tab_statuses,
)

_GYM_CORE_SOLVER_POOL = "Gym Core"


def _warn_if_schedules_mismatched(
    schedule_output: Dict[str, Any],
    schedule_input: Dict[str, Any],
) -> bool:
    """Warn if schedule_output assignments reference game IDs absent from schedule_input.

    Returns True when the files are consistent, False when orphaned game IDs are found.
    Orphaned IDs typically mean --input and --constraint came from different runs, which
    causes produce-schedule to silently render rows with blank event/stage fields (B5).
    """
    known_ids = (
        {g["game_id"] for g in schedule_input.get("games", [])}
        | {ps["game_id"] for ps in schedule_input.get("playoff_slots", [])}
    )
    assignment_ids = {
        a["game_id"] for a in schedule_output.get("assignments", [])
    }
    orphaned = assignment_ids - known_ids
    if orphaned:
        logger.warning(
            f"{len(orphaned)} assignment game_id(s) not found in schedule_input — "
            "--input and --constraint may be from different runs. "
            "Affected rows will render with blank event/stage. "
            f"Orphaned IDs: {sorted(orphaned)}"
        )
        return False
    return True


def _build_schedule_output_flat_rows(
    schedule_output: Dict[str, Any],
    schedule_input: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Build sorted flat-list rows for the Schedule-by-Sport tab.

    Each row joins one assignment from schedule_output with game metadata
    from schedule_input.  Rows are sorted by event → stage order → round → slot.
    """
    import re
    _warn_if_schedules_mismatched(schedule_output, schedule_input)
    game_meta = {g["game_id"]: g for g in schedule_input.get("games", [])}
    res_meta  = {r["resource_id"]: r for r in schedule_input.get("resources", [])}
    def _stage_order(stage: Any) -> Tuple[int, int]:
        text = str(stage or "").strip()
        early_round = re.fullmatch(r"R(\d+)", text)
        if early_round:
            return (1, int(early_round.group(1)))
        return {
            "Pool": (0, 0),
            "QF": (2, 0),
            "Semi": (3, 0),
            "Final": (4, 0),
            "3rd": (5, 0),
        }.get(text, (99, 0))
    rows: List[Dict[str, Any]] = []
    for a in schedule_output.get("assignments", []):
        gid  = a["game_id"]
        rid  = a["resource_id"]
        slot = a["slot"]
        # Fall back to the assignment dict itself for playoff games whose
        # game_id is not in schedule_input games (they carry event/stage fields).
        game = game_meta.get(gid, a)
        res  = res_meta.get(rid, {})
        time_part = slot.rsplit("-", maxsplit=1)[-1] if "-" in slot else slot
        _, _, cat_code = style_for_game(game)
        rows.append({
            "game_id":          gid,
            "category":         cat_code,
            "event":            game.get("event", ""),
            "stage":            game.get("stage", ""),
            "round":            game.get("round", ""),
            "team_a_id":        game.get("team_a_label", game.get("team_a_id", "")),
            "team_b_id":        game.get("team_b_label", game.get("team_b_id", "")),
            "team_c_id":        game.get("team_c_label", game.get("team_c_id", "")),
            "resource_label":   res.get("label", rid),
            "day":              _day_display_label(
                str(res.get("day", "")),
                short=True,
            ),
            "slot":             time_part,
            "duration_minutes": game.get("duration_minutes", ""),
        })
    rows.sort(key=lambda r: (
        r["event"],
        _stage_order(r["stage"]),
        int(r["round"]) if isinstance(r["round"], int) else 0,
        r["slot"],
    ))
    return rows


def _write_schedule_diagnostics_tab(
    ws,
    schedule_output: Dict[str, Any],
    schedule_input: Dict[str, Any],
) -> None:
    """Render the operator-facing diagnose-schedule summary into Excel."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from schedule_diagnostics import build_schedule_diagnostics

    diagnostics = build_schedule_diagnostics(schedule_input, schedule_output)
    summary = diagnostics.get("summary", {}) or {}
    audit = diagnostics.get("audit", {}) or {}
    supply = diagnostics.get("supply", {}) or {}
    control = diagnostics.get("control", {}) or {}
    resource_contract = diagnostics.get("resource_contract", {}) or {}

    header_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER, fill_type="solid")
    section_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    good_fill = PatternFill(fgColor="D9EAD3", fill_type="solid")
    warn_fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
    high_fill = PatternFill(fgColor="F4CCCC", fill_type="solid")
    info_fill = PatternFill(fgColor="D9EAF7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _severity_fill(severity: str) -> PatternFill:
        if severity == "high":
            return high_fill
        if severity == "medium":
            return warn_fill
        return info_fill

    def _section(row: int, title: str, width: int = 5) -> int:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
        cell = ws.cell(row=row, column=1, value=title)
        cell.fill = section_fill
        cell.font = bold_font
        cell.alignment = center
        return row + 1

    def _headers(row: int, columns: List[str]) -> int:
        for col_idx, value in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        return row + 1

    ws.title = "Schedule-Diagnostics"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell = ws.cell(row=1, column=1, value="VAY Sports Fest - Schedule Diagnostics")
    title_cell.fill = header_fill
    title_cell.font = header_font
    title_cell.alignment = center
    ws.freeze_panes = "A8"

    row = 3
    row = _section(row, "Summary")
    row = _headers(row, ["Metric", "Value"])
    summary_rows = [
        ("Generated at", summary.get("generated_at") or ""),
        ("Schedule output", "yes" if summary.get("has_schedule_output") else "no"),
        ("Games", summary.get("game_count", 0)),
        ("Resources", summary.get("resource_count", 0)),
        ("Solver status", audit.get("status") or ""),
        ("Assigned", audit.get("assigned_count", 0)),
        ("Unscheduled", audit.get("unscheduled_count", 0)),
    ]
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    if resource_contract:
        row += 1
        row = _section(row, "Resource Contract")
        row = _headers(row, ["Status", "Source", "Exclusive groups", "Gym-Modes", "Issue count"])
        contract_status = str(resource_contract.get("status") or "")
        issue_count = len(resource_contract.get("issues", []) or [])
        fill = high_fill if contract_status == "error" else warn_fill if contract_status == "warn" else good_fill
        values = [
            contract_status,
            str(resource_contract.get("allocation_source") or ""),
            resource_contract.get("exclusive_group_count", 0),
            resource_contract.get("gym_modes_count", 0),
            issue_count,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = left
        row += 1
        if issue_count:
            row = _headers(row, ["Severity", "Code", "Message"])
            for issue in resource_contract.get("issues", []) or []:
                severity = str(issue.get("severity") or "")
                fill = _severity_fill(severity)
                values = [
                    severity,
                    str(issue.get("code") or ""),
                    str(issue.get("message") or ""),
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.fill = fill
                    cell.alignment = left
                row += 1

    row += 1
    row = _section(row, "Next Actions")
    row = _headers(row, ["Severity", "Vector", "Message"])
    actions = diagnostics.get("next_actions", []) or []
    for action in actions:
        severity = str(action.get("severity") or "")
        fill = _severity_fill(severity)
        values = [
            severity,
            str(action.get("vector") or ""),
            str(action.get("message") or ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = left
        row += 1
    if not actions:
        ws.cell(row=row, column=1, value="No suggested next actions.")
        row += 1

    quality_warnings = diagnostics.get("quality_warnings", []) or []
    if quality_warnings:
        row += 1
        row = _section(row, "Quality Warnings")
        row = _headers(row, ["Severity", "Check", "Event", "Day", "Message"])
        for warning in quality_warnings:
            severity = str(warning.get("severity") or "info")
            fill = _severity_fill(severity)
            values = [
                severity,
                str(warning.get("check") or ""),
                str(warning.get("event") or ""),
                str(warning.get("day") or ""),
                str(warning.get("message") or ""),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = left
            row += 1

    overlaps = supply.get("exclusive_group_overlaps", []) or []
    if overlaps:
        row += 1
        row = _section(row, "Physical Venue Overlaps")
        row = _headers(
            row,
            ["Venue", "Day", "First window", "Second window", "Example resources"],
        )
        for overlap in overlaps:
            values = [
                overlap.get("exclusive_group", ""),
                overlap.get("day", ""),
                (
                    f"{overlap.get('first_resource_type', '')} "
                    f"{overlap.get('first_open_time', '')}-{overlap.get('first_close_time', '')}"
                ),
                (
                    f"{overlap.get('second_resource_type', '')} "
                    f"{overlap.get('second_open_time', '')}-{overlap.get('second_close_time', '')}"
                ),
                "; ".join(
                    " + ".join(pair)
                    for pair in overlap.get("example_resource_ids", []) or []
                ),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = high_fill
                cell.alignment = left
            row += 1

    gym_shortfall = (
        control.get("gym_allocation", {}).get("mode_shortfall", {}) or {}
    )
    if gym_shortfall:
        row += 1
        row = _section(row, "Gym Mode Capacity Notes")
        row = _headers(row, ["Mode", "Shortfall slots"])
        for mode, shortfall in sorted(gym_shortfall.items()):
            ws.cell(row=row, column=1, value=mode)
            ws.cell(row=row, column=2, value=shortfall)
            row += 1

    capacity_rows = diagnostics.get("capacity_pressure", []) or []
    if capacity_rows:
        row += 1
        row = _section(row, "Capacity Pressure")
        row = _headers(
            row,
            [
                "Resource type",
                "Required slots",
                "Available slots",
                "Shortage slots",
                "Missing events",
            ],
        )
        for capacity in capacity_rows:
            missing = "; ".join(
                (
                    f"{item.get('event', '')} "
                    f"({item.get('game_count', 0)} game(s))"
                )
                for item in capacity.get("missing_resource_events", []) or []
            )
            shortage = capacity.get("shortage_slots", 0)
            fill = high_fill if shortage else good_fill
            values = [
                capacity.get("resource_type", ""),
                capacity.get("required_slots", 0),
                capacity.get("available_slots", 0),
                shortage,
                missing,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = left
            row += 1

    widths = [24, 18, 78, 34, 42]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 24


def _write_schedule_output_report(
    filepath: Path,
    schedule_output: Dict[str, Any],
    schedule_input: Dict[str, Any],
) -> None:
    """Write Schedule-by-Time, Schedule-by-Sport, and Conflict-Audit tabs.

    Tab 1 — Schedule-by-Time: grid (rows = time slots, columns = courts),
      colour-coded by sport, with session sections separated by grey rows.
    Tab 2 — Schedule-by-Sport: flat list sorted by event → stage → round,
      with auto-filter and an unscheduled section when applicable.
    Tab 3 — Conflict-Audit: cross-sport shared-athlete audit rows when available.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    game_meta: Dict[str, Dict[str, Any]] = {
        g["game_id"]: g for g in schedule_input.get("games", [])
    }
    res_meta: Dict[str, Dict[str, Any]] = {
        r["resource_id"]: r for r in schedule_input.get("resources", [])
    }
    assign_map: Dict[Tuple[str, str], Dict[str, Any]] = {
        (a["resource_id"], a["slot"]): game_meta.get(a["game_id"], {"game_id": a["game_id"]})
        for a in schedule_output.get("assignments", [])
    }

    solved_at     = schedule_output.get("solved_at", "")
    status        = schedule_output.get("status", "")
    n_assigned    = len(schedule_output.get("assignments", []))
    n_unscheduled = len(schedule_output.get("unscheduled", []))
    snapshot      = (
        f"Generated: {solved_at}  |  Status: {status}  |  "
        f"Scheduled: {n_assigned}  |  Unscheduled: {n_unscheduled}"
    )

    sec_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_SECTION, fill_type="solid")
    hdr_fill = PatternFill(fgColor=SCHEDULE_SKETCH_COLOR_HEADER,  fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    red_fill      = PatternFill(fgColor="FFC7CE", fill_type="solid")
    conflict_fill = PatternFill(fgColor="FFCC00", fill_type="solid")

    def _sport_fill(event: str) -> PatternFill:
        return PatternFill(fgColor=sport_style(event).fill_color, fill_type="solid")

    def _category_font(game: Dict[str, Any], bold: bool = False) -> Font:
        _, cat_style, _ = style_for_game(game)
        return Font(color=cat_style.text_color, bold=bold)

    def _slot_times(res: Dict[str, Any]) -> List[str]:
        o_h, o_m = map(int, res["open_time"].split(":"))
        c_h, c_m = map(int, res["close_time"].split(":"))
        sm        = res["slot_minutes"]
        open_min  = o_h * 60 + o_m
        close_min = c_h * 60 + c_m
        times: List[str] = []
        t = open_min
        while t + sm <= close_min:
            times.append(f"{t // 60:02d}:{t % 60:02d}")
            t += sm
        return times

    def _cell_text(game: Dict[str, Any]) -> str:
        gid    = game.get("game_id", "")
        event  = str(game.get("event") or "")
        badge  = format_compact_label(event,
                                      str(game.get("sport_format") or ""),
                                      game.get("category"))
        a   = str(game.get("team_a_label") or game.get("team_a_id") or "")
        b   = str(game.get("team_b_label") or game.get("team_b_id") or "")
        c   = str(game.get("team_c_label") or game.get("team_c_id") or "")
        # Two-line layout: badge + gid on line 1, matchup on line 2 if compact.
        header = f"{badge}  {gid}" if gid else badge
        if a and b and c and len(a) <= 12 and len(b) <= 12 and len(c) <= 12:
            return f"{header}\n{a} / {b} / {c}"
        if a and b and len(a) <= 12 and len(b) <= 12:
            return f"{header}\n{a} vs {b}"
        return header

    def _time_sort_key(hhmm: str) -> int:
        h, m = map(int, hhmm.split(":"))
        return h * 60 + m

    def _resource_group_key(res: Dict[str, Any]) -> Tuple[str, str, str, str, int]:
        solver_pool = str(res.get("solver_pool") or "").strip()
        day = str(res.get("day", ""))
        resource_type = str(res.get("resource_type", ""))
        slot_minutes = int(res.get("slot_minutes", 0) or 0)
        if solver_pool == _GYM_CORE_SOLVER_POOL:
            # Render one continuous operator-facing section per Day/resource_type
            # for the shared gym solver pool, even when the allocator produced
            # multiple overlapping time windows for the same sport.
            return (day, resource_type, "", "", slot_minutes)
        return (
            day,
            resource_type,
            str(res.get("open_time", "")),
            str(res.get("close_time", "")),
            slot_minutes,
        )

    def _group_open_close(day_res: List[Dict[str, Any]]) -> Tuple[str, str]:
        open_times = [
            str(res.get("open_time", "")).strip()
            for res in day_res
            if str(res.get("open_time", "")).strip()
        ]
        close_times = [
            str(res.get("close_time", "")).strip()
            for res in day_res
            if str(res.get("close_time", "")).strip()
        ]
        merged_open = min(open_times, key=_time_sort_key) if open_times else ""
        merged_close = max(close_times, key=_time_sort_key) if close_times else ""
        return merged_open, merged_close

    def _group_slot_times(day_res: List[Dict[str, Any]]) -> List[str]:
        return sorted(
            {
                t_str
                for res in day_res
                for t_str in _slot_times(res)
            },
            key=_time_sort_key,
        )

    def _resource_header_labels(day_res: List[Dict[str, Any]]) -> Dict[str, str]:
        labels_by_resource: Dict[str, str] = {}
        base_labels: Dict[str, str] = {}
        for res in day_res:
            resource_id = str(res.get("resource_id", "")).strip()
            base_label = str(res.get("label") or resource_id).strip() or resource_id
            solver_pool = str(res.get("solver_pool") or "").strip()
            venue_name = (
                str(res.get("exclusive_group") or "").strip()
                or str(res.get("venue_name") or "").strip()
            )
            base_labels[resource_id] = base_label
            if (
                venue_name
                and solver_pool == _GYM_CORE_SOLVER_POOL
            ):
                labels_by_resource[resource_id] = f"{venue_name} {base_label}"
            else:
                labels_by_resource[resource_id] = base_label

        def _counts() -> Dict[str, int]:
            counts: Dict[str, int] = {}
            for resource_id, label in labels_by_resource.items():
                counts[label] = counts.get(label, 0) + 1
            return counts

        duplicate_labels = {
            label for label, count in _counts().items() if count > 1
        }
        if duplicate_labels:
            for res in day_res:
                resource_id = str(res.get("resource_id", "")).strip()
                if labels_by_resource.get(resource_id) not in duplicate_labels:
                    continue
                venue_name = str(res.get("venue_name") or "").strip()
                if venue_name:
                    labels_by_resource[resource_id] = (
                        f"{venue_name} {base_labels[resource_id]}"
                    )

        duplicate_labels = {
            label for label, count in _counts().items() if count > 1
        }
        if duplicate_labels:
            for res in day_res:
                resource_id = str(res.get("resource_id", "")).strip()
                if labels_by_resource.get(resource_id) not in duplicate_labels:
                    continue
                open_time = str(res.get("open_time") or "").strip()
                close_time = str(res.get("close_time") or "").strip()
                window = f"{open_time}-{close_time}" if open_time and close_time else resource_id
                labels_by_resource[resource_id] = (
                    f"{labels_by_resource[resource_id]} [{window}]"
                )

        duplicate_labels = {
            label for label, count in _counts().items() if count > 1
        }
        if duplicate_labels:
            for res in day_res:
                resource_id = str(res.get("resource_id", "")).strip()
                if labels_by_resource.get(resource_id) not in duplicate_labels:
                    continue
                labels_by_resource[resource_id] = (
                    f"{labels_by_resource[resource_id]} ({resource_id})"
                )

        return labels_by_resource

    # Group resources by uniform day/window/resource pool so pod schedules with
    # mixed slot lengths do not get collapsed into one broken "Day-1" grid.
    resource_groups: Dict[Tuple[str, str, str, str, int], List[Dict[str, Any]]] = {}
    for res in schedule_input.get("resources", []):
        resource_groups.setdefault(_resource_group_key(res), []).append(res)

    group_counts_by_day: Dict[str, int] = {}
    for day, _, _, _, _ in resource_groups.keys():
        group_counts_by_day[day] = group_counts_by_day.get(day, 0) + 1

    sorted_group_keys = sorted(
        resource_groups.keys(),
        key=lambda key: (
            _day_sort_key(key[0]),
            _time_sort_key(key[2]) if key[2] else 0,
            _time_sort_key(key[3]) if key[3] else 0,
            key[4],
            key[1],
        ),
    )
    max_resources = max((len(v) for v in resource_groups.values()), default=4)
    n_cols        = 1 + max_resources

    def _section_label(
        group_key: Tuple[str, str, str, str, int],
        day_res: List[Dict[str, Any]],
    ) -> str:
        day, resource_type, open_time, close_time, slot_minutes = group_key
        day_label = _day_display_label(day)
        if not open_time or not close_time:
            open_time, close_time = _group_open_close(day_res)
        if (
            day_label != day
            and resource_type in (GYM_RESOURCE_TYPE, GYM_RESOURCE_TYPE_BASKETBALL, GYM_RESOURCE_TYPE_VOLLEYBALL)
            and group_counts_by_day.get(day, 0) == 1
        ):
            return day_label
        return (
            f"{day_label} — {resource_type} "
            f"({open_time}-{close_time}, {slot_minutes}m)"
        )

    wb = Workbook()

    # ── Tab 1: Schedule-by-Time ──────────────────────────────────────────
    ws1       = wb.active
    ws1.title = "Schedule-by-Time"

    # Row 1 — report title (merged)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws1.cell(row=1, column=1, value="VAY Sports Fest — Schedule by Time")
    c.fill, c.font, c.alignment = hdr_fill, hdr_font, center

    # Legend row: one chip per sport (fill color + abbrev) followed by
    # one chip per category code (text color + prefix).  Keeps the
    # schedule self-explanatory even when printed in black-and-white.
    legend_row = 2
    legend_col = 1
    for event_name, style in SPORT_STYLES.items():
        cell = ws1.cell(row=legend_row, column=legend_col, value=style.abbrev)
        cell.fill = PatternFill(fgColor=style.fill_color, fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = center
        legend_col += 1
        if legend_col > n_cols:
            break
    if legend_col <= n_cols:
        for cat in CATEGORY_STYLES.values():
            cell = ws1.cell(row=legend_row, column=legend_col, value=cat.prefix)
            cell.font = Font(color=cat.text_color, bold=True)
            cell.alignment = center
            legend_col += 1
            if legend_col > n_cols:
                break

    ws1.freeze_panes = "A3"

    cur_row = 4
    for group_key in sorted_group_keys:
        day_res = sorted(
            resource_groups[group_key],
            key=lambda r: (
                _time_sort_key(str(r.get("open_time") or "00:00")),
                str(r.get("exclusive_group") or ""),
                str(r.get("label") or ""),
                r["resource_id"],
            ),
        )
        if not day_res:
            continue
        header_labels = _resource_header_labels(day_res)

        # Section header (grey, merged)
        ws1.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=n_cols,
        )
        c = ws1.cell(row=cur_row, column=1, value=_section_label(group_key, day_res))
        c.fill, c.font, c.alignment = sec_fill, bold_font, center
        cur_row += 1

        # Column headers for this group
        ws1.cell(row=cur_row, column=1, value="Time").font = bold_font
        ws1.cell(row=cur_row, column=1).fill = sec_fill
        ws1.cell(row=cur_row, column=1).alignment = center
        for ci, res in enumerate(day_res, start=2):
            c = ws1.cell(
                row=cur_row,
                column=ci,
                value=header_labels.get(
                    str(res.get("resource_id") or "").strip(),
                    res.get("label"),
                ),
            )
            c.fill, c.font, c.alignment = sec_fill, bold_font, center
        cur_row += 1

        day = group_key[0]
        # Data rows — one per unioned time slot in this resource group.
        for t_str in _group_slot_times(day_res):
            slot_label = f"{day}-{t_str}"
            ws1.cell(row=cur_row, column=1, value=t_str).alignment = center
            for ci, res in enumerate(day_res, start=2):
                game = assign_map.get((res["resource_id"], slot_label))
                cell = ws1.cell(row=cur_row, column=ci)
                if game:
                    cell.value = _cell_text(game)
                    cell.fill  = _sport_fill(game.get("event", ""))
                    cell.font  = _category_font(game, bold=True)
                cell.alignment = center
            cur_row += 1

        cur_row += 1  # blank row between sessions

    ws1.cell(row=cur_row + 1, column=1, value=snapshot)
    ws1.column_dimensions["A"].width = 7
    for ci in range(2, n_cols + 1):
        ws1.column_dimensions[get_column_letter(ci)].width = 18

    # ── Tab 2: Schedule-by-Sport ─────────────────────────────────────────
    ws2       = wb.create_sheet("Schedule-by-Sport")
    flat_rows = _build_schedule_output_flat_rows(
        schedule_output, schedule_input
    )
    col_defs = [
        ("game_id",          14),
        ("category",          8),
        ("event",            28),
        ("stage",             8),
        ("round",             6),
        ("team_a_id",        20),
        ("team_b_id",        20),
        ("team_c_id",        20),
        ("resource_label",   14),
        ("day",              10),
        ("slot",              8),
        ("duration_minutes", 16),
    ]
    cols = [col for col, _ in col_defs]

    for ci, (col, _) in enumerate(col_defs, start=1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill, cell.font = hdr_fill, hdr_font
        cell.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(col_defs))}1"

    for ri, row in enumerate(flat_rows, start=2):
        fill = _sport_fill(row.get("event", ""))
        row_font = _category_font(row)
        for ci, col in enumerate(cols, start=1):
            cell = ws2.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.fill, cell.alignment, cell.font = fill, left, row_font

    # Unscheduled section at bottom
    unscheduled = schedule_output.get("unscheduled", [])
    ri = len(flat_rows) + 3
    if unscheduled:
        ws2.merge_cells(
            start_row=ri, start_column=1, end_row=ri, end_column=len(col_defs)
        )
        c = ws2.cell(
            row=ri, column=1,
            value=f"Unscheduled Games ({len(unscheduled)})",
        )
        c.fill, c.font = red_fill, Font(bold=True)
        for gid in unscheduled:
            ri += 1
            ws2.cell(row=ri, column=1, value=gid).fill = red_fill
        ri += 2

    # Pool results summary — shown when pool_results present
    pool_results = schedule_output.get("pool_results", [])
    if pool_results:
        ws2.merge_cells(
            start_row=ri, start_column=1, end_row=ri, end_column=len(col_defs)
        )
        c = ws2.cell(row=ri, column=1, value="Pool Results")
        c.fill, c.font = sec_fill, Font(bold=True)
        ri += 1
        for pr in pool_results:
            pr_status = pr.get("status", "")
            pr_fill   = red_fill if pr_status not in ("OPTIMAL", "FEASIBLE") else PatternFill(
                fgColor="C6EFCE", fill_type="solid"
            )
            ws2.cell(row=ri, column=1, value=pr.get("resource_type", "")).fill = pr_fill
            ws2.cell(row=ri, column=2, value=pr_status).fill                   = pr_fill
            ws2.cell(row=ri, column=3, value=f"Assigned: {len(pr.get('assignments', []))}").fill  = pr_fill
            ws2.cell(row=ri, column=4, value=f"Unscheduled: {len(pr.get('unscheduled', []))}").fill = pr_fill
            ri += 1
            for diag in pr.get("diagnostics", []):
                for line in (diag.get("missing_resource_events") or []):
                    ws2.cell(row=ri, column=2,
                             value=f"  No resources: {line.get('event','')} ({line.get('game_count',0)} games)"
                             ).fill = red_fill
                    ri += 1
                if diag.get("shortage_slots", 0) > 0:
                    ws2.cell(row=ri, column=2,
                             value=f"  Short {diag['shortage_slots']} slot(s): "
                                   f"need {diag['required_slots']}, have {diag['available_slots']}"
                             ).fill = red_fill
                    ri += 1
        ri += 1

    ws2.cell(row=ri, column=1, value=snapshot)
    for ci, (_, width) in enumerate(col_defs, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = width

    # ── Tab 3: Conflict-Audit ────────────────────────────────────────────
    ws3 = wb.create_sheet("Conflict-Audit")
    conflict_summary = schedule_output.get("conflict_audit_summary", {}) or {}
    conflict_rows = schedule_output.get("conflict_audit", []) or []
    ws3.cell(row=1, column=1, value="Cross-Sport Conflict Audit").fill = hdr_fill
    ws3.cell(row=1, column=1).font = hdr_font

    summary_lines = [
        (
            "Summary",
            (
                f"Edges: {conflict_summary.get('total_edges', 0)}  |  "
                f"Separated: {conflict_summary.get('separated_edges', 0)}  |  "
                f"Remaining: {conflict_summary.get('overlapping_edges', 0)}  |  "
                f"Planning-only: {conflict_summary.get('planning_only_edges', 0)}  |  "
                f"Incomplete: {conflict_summary.get('incomplete_edges', 0)}"
            ),
        ),
        (
            "Remaining Penalties",
            (
                f"Primary: {conflict_summary.get('remaining_primary_overlap_penalty', 0)}  |  "
                f"Secondary-only: {conflict_summary.get('remaining_secondary_overlap_penalty', 0)}"
            ),
        ),
    ]
    for row_idx, (label, value) in enumerate(summary_lines, start=3):
        ws3.cell(row=row_idx, column=1, value=label).font = bold_font
        ws3.cell(row=row_idx, column=2, value=value)

    audit_headers = [
        ("team_a_label", 16),
        ("event_a", 24),
        ("team_b_label", 16),
        ("event_b", 24),
        ("shared_count", 12),
        ("primary_overlap_count", 18),
        ("secondary_only_count", 18),
        ("status", 20),
        ("overlap_count", 14),
        ("scheduled_team_a_games", 20),
        ("scheduled_team_b_games", 20),
        ("shared_participant_names", 40),
        ("overlap_game_pairs", 48),
    ]
    header_row = 6
    for ci, (col, width) in enumerate(audit_headers, start=1):
        cell = ws3.cell(row=header_row, column=ci, value=col)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, center
        ws3.column_dimensions[get_column_letter(ci)].width = width
    ws3.freeze_panes = "A7"
    ws3.auto_filter.ref = f"A{header_row}:{get_column_letter(len(audit_headers))}{header_row}"

    if conflict_rows:
        for ri3, row in enumerate(conflict_rows, start=header_row + 1):
            row_fill = red_fill if row.get("status") == "ConflictRemains" else PatternFill(
                fgColor="C6EFCE", fill_type="solid"
            )
            if row.get("status") == "IncompleteSchedule":
                row_fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
            elif row.get("status") == "PlanningOnly":
                row_fill = PatternFill(fgColor="DDEBF7", fill_type="solid")
            for ci, (col, _width) in enumerate(audit_headers, start=1):
                cell = ws3.cell(row=ri3, column=ci, value=row.get(col, ""))
                cell.fill = row_fill
                cell.alignment = left
    else:
        ws3.cell(
            row=header_row + 1,
            column=1,
            value="No cross-sport conflict audit rows were produced for this schedule.",
        )

    # Unprotected racquet doubles entries (Issue #158): UnresolvedDoubles
    # have unknown membership, so they cannot be conflict-protected.  List
    # them so operators can chase down the missing/non-reciprocal partners.
    unprotected = schedule_output.get("pod_unprotected_entries", []) or []
    if unprotected:
        section_row = header_row + 1 + max(len(conflict_rows), 1) + 2
        note_cell = ws3.cell(
            row=section_row,
            column=1,
            value="Unprotected Racquet Doubles (UnresolvedDoubles — not conflict-protected)",
        )
        note_cell.font = bold_font
        note_cell.fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
        unprotected_headers = [
            "division_id", "participant_name", "reason",
            "validation_issue_status", "notes",
        ]
        head_r = section_row + 1
        for ci, col in enumerate(unprotected_headers, start=1):
            cell = ws3.cell(row=head_r, column=ci, value=col)
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, center
        warn_fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
        for ri, entry in enumerate(unprotected, start=head_r + 1):
            for ci, col in enumerate(unprotected_headers, start=1):
                cell = ws3.cell(row=ri, column=ci, value=entry.get(col, ""))
                cell.fill = warn_fill
                cell.alignment = left

    # ── Tab 4: Master-Schedule ───────────────────────────────────────────
    # Diagnostic summary tab: same signal as diagnose-schedule, embedded in
    # the generated workbook so operators do not need to inspect JSON first.
    ws_diag = wb.create_sheet("Schedule-Diagnostics")
    _write_schedule_diagnostics_tab(
        ws_diag,
        schedule_output,
        schedule_input,
    )

    # Tab 5: Master-Schedule
    # Grid: rows = time slots grouped by day; columns = physical courts/
    # tables grouped by venue.  One column per (venue_group, label) pair
    # so that multiple resource_ids on the same physical court (created by
    # the gym allocator for different sports) collapse into one column.
    ws4 = wb.create_sheet("Master-Schedule")

    all_resources: List[Dict[str, Any]] = list(schedule_input.get("resources", []))

    # Chronological day order — calendar-date-derived when available.
    ms_day_order: List[str] = list(schedule_input.get("day_order") or [])
    if not ms_day_order:
        import re as _re
        ms_day_order = sorted(
            {str(r.get("day", "")) for r in all_resources if r.get("day")},
            key=_day_sort_key,
        )

    def _ms_day_rank(day: str) -> int:
        try:
            return ms_day_order.index(day)
        except ValueError:
            return len(ms_day_order)

    # Physical venue group label for a resource.
    def _venue_group(res: Dict[str, Any]) -> str:
        solver_pool = str(res.get("solver_pool") or "").strip()
        if solver_pool == _GYM_CORE_SOLVER_POOL:
            return (
                str(res.get("exclusive_group") or "").strip()
                or str(res.get("venue_name") or "").strip()
                or "Other"
            )
        return str(res.get("venue_name") or "").strip() or "Other"

    def _res_label(res: Dict[str, Any]) -> str:
        rid = str(res.get("resource_id") or "").strip()
        return str(res.get("label") or rid).strip()

    # Sort venue groups by the priority of their dominant resource type so
    # the column layout follows the sport order (Soccer → BC → Basketball →
    # Volleyball → Tennis → Pickleball → Badminton → Table Tennis) rather
    # than alphabetical venue names.  When a venue hosts multiple resource
    # types (e.g. 4 Tennis + 4 Pickleball at EHS Tennis Court), it sorts by
    # the lowest-numbered (highest priority) type it contains.
    _resource_type_rank = {
        "Soccer Field":        0,
        "BC Station":          1,
        "Basketball Court":    2,
        "Volleyball Court":    3,
        "Tennis Court":        4,
        "Pickleball Court":    5,
        "Badminton Court":     6,
        "Table Tennis Table":  7,
    }
    _UNRANKED = 99

    vg_to_min_rank: Dict[str, int] = {}
    for res in all_resources:
        vg = _venue_group(res)
        rt = str(res.get("resource_type") or "").strip()
        rank = _resource_type_rank.get(rt, _UNRANKED)
        if vg not in vg_to_min_rank or rank < vg_to_min_rank[vg]:
            vg_to_min_rank[vg] = rank

    # "Other" (fallback when venue_name is blank) sorts after all named
    # venues regardless of its resource types, so it never splits two
    # sport-related venue groups that should be adjacent (e.g. Tennis and
    # Pickleball).  Named venues still sort by sport-type priority.
    def _vg_sort_key(vg: str) -> Tuple[int, str]:
        if vg == "Other":
            return (_UNRANKED, "Other")
        return (vg_to_min_rank.get(vg, _UNRANKED), vg)

    def _label_sort_key(label: str) -> Tuple[str, int, str]:
        import re as _re2
        # Natural sort: split on the last run of digits so "Court-2" < "Court-10".
        m = _re2.search(r"(\d+)(\D*)$", label)
        if m:
            prefix = label[: m.start()]
            return (prefix, int(m.group(1)), m.group(2))
        return (label, 0, "")

    sorted_resources = sorted(
        all_resources,
        key=lambda r: (_vg_sort_key(_venue_group(r)), _label_sort_key(_res_label(r)), str(r.get("resource_id") or "")),
    )

    # Build one column per physical court = (venue_group, label) pair.
    # Multiple resource_ids that share the same (venue_group, label) —
    # e.g. the gym allocator creates separate IDs per sport on the same
    # court — all map to the same column so no duplicate Court-N columns.
    col_keys: List[Tuple[str, str]] = []       # ordered (vg, label) list
    col_key_set: set = set()
    rid_to_col_key: Dict[str, Tuple[str, str]] = {}  # resource_id → (vg, label)
    for res in sorted_resources:
        rid = str(res.get("resource_id") or "").strip()
        key = (_venue_group(res), _res_label(res))
        rid_to_col_key[rid] = key
        if key not in col_key_set:
            col_keys.append(key)
            col_key_set.add(key)

    # Drop columns that have no assignments — avoids empty Table-N columns
    # for venues the solver left unused (e.g. EHS Practice Gym TT workaround).
    active_col_keys: set = {
        rid_to_col_key[rid]
        for rid in (a["resource_id"] for a in schedule_output.get("assignments", []))
        if rid in rid_to_col_key
    }
    col_keys = [k for k in col_keys if k in active_col_keys]

    venue_groups_ordered: List[str] = list(dict.fromkeys(vg for vg, _ in col_keys))

    # Columns start at 3 (col 1 = Day, col 2 = Time).
    col_key_idx: Dict[Tuple[str, str], int] = {}
    col_idx = 3
    for key in col_keys:
        col_key_idx[key] = col_idx
        col_idx += 1

    total_cols = max(col_idx - 1, 3)

    # Compact cell text for Master-Schedule: game_id + matchup only.
    # The sport badge from _cell_text() is redundant here because cells
    # are already colour-coded by sport.
    def _master_cell_text(game: Dict[str, Any]) -> str:
        gid = str(game.get("game_id") or "")
        a = str(game.get("team_a_label") or game.get("team_a_id") or "")
        b = str(game.get("team_b_label") or game.get("team_b_id") or "")
        c = str(game.get("team_c_label") or game.get("team_c_id") or "")
        if a and b and c and len(a) <= 12 and len(b) <= 12 and len(c) <= 12:
            return f"{gid} {a} / {b} / {c}"
        if a and b and len(a) <= 12 and len(b) <= 12:
            return f"{gid} {a} v {b}"
        return gid

    # ── Header rows ─────────────────────────────────────────────────────
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws4.cell(row=1, column=1, value="VAY Sports Fest — Master Schedule")
    c.fill, c.font, c.alignment = hdr_fill, hdr_font, center

    ws4.cell(row=2, column=1, value="Day").fill   = hdr_fill
    ws4.cell(row=2, column=1).font                = hdr_font
    ws4.cell(row=2, column=1).alignment           = center
    ws4.cell(row=2, column=2, value="Time").fill  = hdr_fill
    ws4.cell(row=2, column=2).font                = hdr_font
    ws4.cell(row=2, column=2).alignment           = center

    for vg in venue_groups_ordered:
        vg_cols = [col_key_idx[k] for k in col_keys if k[0] == vg]
        first_col, last_col = min(vg_cols), max(vg_cols)
        if first_col < last_col:
            ws4.merge_cells(
                start_row=2, start_column=first_col,
                end_row=2, end_column=last_col,
            )
        c = ws4.cell(row=2, column=first_col, value=vg)
        c.fill, c.font, c.alignment = hdr_fill, hdr_font, center

    ws4.cell(row=3, column=1, value="").fill = sec_fill
    ws4.cell(row=3, column=2, value="").fill = sec_fill
    for (vg, label), col in col_key_idx.items():
        c = ws4.cell(row=3, column=col, value=label)
        c.fill, c.font, c.alignment = sec_fill, bold_font, center

    ws4.freeze_panes = "C4"

    # ── Build unified time-slot grid ─────────────────────────────────────
    slot_index: Dict[Tuple[int, int], Tuple[str, str]] = {}
    for res in all_resources:
        day = str(res.get("day") or "").strip()
        if not day:
            continue
        day_rank = _ms_day_rank(day)
        for t_str in _slot_times(res):
            h, m = map(int, t_str.split(":"))
            slot_index[(day_rank, h * 60 + m)] = (day, t_str)

    sorted_slot_keys = sorted(slot_index.keys())

    # Pre-compute which slot labels have at least one assignment so we can
    # skip entirely empty rows (e.g. 12:30 half-hour gaps between games).
    occupied_slots: set = {a["slot"] for a in schedule_output.get("assignments", [])}

    # ── Data rows ────────────────────────────────────────────────────────
    cur_row4 = 4
    prev_day: str = ""
    prev_day_display: str = ""
    # Track whether we've written the day header yet; defer it until the
    # first non-empty row so a day with all-empty slots is also suppressed.
    pending_day_header: Any = None  # (day_display, day) or None

    for day_rank, time_min in sorted_slot_keys:
        day, t_str = slot_index[(day_rank, time_min)]
        slot_label = f"{day}-{t_str}"

        if slot_label not in occupied_slots:
            # No game anywhere at this time — skip the row entirely.
            if day != prev_day:
                day_display = _day_display_label(day)
                pending_day_header = (day_display, day)
            continue

        # Emit deferred day-section header now that we know a row follows.
        if day != prev_day or pending_day_header is not None:
            if pending_day_header is not None:
                day_display, _ = pending_day_header
                pending_day_header = None
            else:
                day_display = _day_display_label(day)
            ws4.merge_cells(
                start_row=cur_row4, start_column=1,
                end_row=cur_row4, end_column=total_cols,
            )
            c = ws4.cell(row=cur_row4, column=1, value=day_display)
            c.fill, c.font, c.alignment = sec_fill, bold_font, center
            cur_row4 += 1
            prev_day = day
            prev_day_display = day_display

        day_cell = ws4.cell(row=cur_row4, column=1, value=prev_day_display)
        day_cell.alignment = center
        day_cell.font = Font(color="808080")

        time_cell = ws4.cell(row=cur_row4, column=2, value=t_str)
        time_cell.alignment = center

        for res in all_resources:
            rid = str(res.get("resource_id") or "").strip()
            col_key = rid_to_col_key.get(rid)
            if col_key is None or col_key not in col_key_idx:
                continue
            col = col_key_idx[col_key]

            res_day = str(res.get("day") or "").strip()
            if res_day != day:
                continue

            o_h, o_m = map(int, res["open_time"].split(":"))
            c_h, c_m = map(int, res["close_time"].split(":"))
            sm = int(res.get("slot_minutes") or 20)
            open_min  = o_h * 60 + o_m
            close_min = c_h * 60 + c_m
            if not (open_min <= time_min and time_min + sm <= close_min):
                continue

            game = assign_map.get((rid, slot_label))
            if not game:
                continue

            cell = ws4.cell(row=cur_row4, column=col)
            if cell.value is not None:
                # Two resources share the same physical column at this slot
                # (exclusive_group double-booking across solver pools).
                # First write wins; yellow fill + red text + always-visible Note.
                from openpyxl.comments import Comment
                from openpyxl.styles import Font as _Font
                conflict_text = _master_cell_text(game)
                note_text = (
                    f"SCHEDULING CONFLICT\n"
                    f"This court is double-booked at {slot_label}.\n"
                    f"Showing: {cell.value}\n"
                    f"Also assigned: {conflict_text}\n\n"
                    f"Fix: adjust Venue_Input.xlsx so only one sport\n"
                    f"uses this court at this time."
                )
                cell.fill = conflict_fill
                cell.font = _Font(color="FF2400", bold=True)
                if cell.comment is None:
                    c = Comment(note_text, "VAYSF Scheduler")
                    c.visible = True
                    cell.comment = c
                continue

            cell.value     = _master_cell_text(game)
            cell.fill      = _sport_fill(game.get("event", ""))
            cell.font      = _category_font(game, bold=True)
            cell.alignment = center

        cur_row4 += 1

    ws4.cell(row=cur_row4 + 1, column=1, value=snapshot)

    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 7
    for col in col_key_idx.values():
        ws4.column_dimensions[get_column_letter(col)].width = 14

    for row_num in range(4, cur_row4):
        ws4.row_dimensions[row_num].height = 36

    _stamp_known_tab_statuses(wb)
    wb.save(filepath)
    _make_excel_note_shapes_visible(filepath)
    logger.info(f"Schedule output report written to: {filepath}")
