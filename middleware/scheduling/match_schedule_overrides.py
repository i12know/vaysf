"""match_schedule_overrides - import Loc's authoritative visual match schedule.

Loc, the human Sports Fest scheduler, fills in the real team codes for every
Basketball / Men's Volleyball / Women's Volleyball pool game directly on a
visual day-by-day grid workbook (one 3-column [team, "v", team] block per
court, one row per time slot; Bible Challenge uses a 3-column all-team-code
block instead). This module parses that grid, validates the team codes
against the roster, cross-references each pairing against the pool games
already generated in schedule_input.json (see import-team-matchups /
assign-pools), and pins the matching game's time/court slot.

This is a distinct workbook family from master_schedule.py's numbered-game
grid: it has no numeric game ids, and its own LEGEND row's swatch colors do
NOT reliably match the actual game-cell fill colors (verified against the
real "2026 Main Schedule draft 11.xlsx" export - e.g. the legend's Basketball
swatch is F4B084 while real BB cells are F8CBAD, and the legend's Men's/
Women's volleyball swatches are each closer to the *other* sport's real cell
color than to their own). Sport is instead resolved by nearest-color match
against schedule_styles.SPORT_STYLES - the same canonical per-sport palette
already used to render every other schedule export in this codebase - which
matches every real color in the draft with a wide margin.
"""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime
from datetime import time as _dt_time
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from openpyxl import load_workbook

from config import COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME, COURT_ESTIMATE_MINUTES_PER_GAME
from schedule_styles import SPORT_STYLES
from scheduling import master_schedule
from scheduling.manual_matchups import _extract_roster_team_codes, _is_bye

MATCH_SCHEDULE_OVERRIDES_WORKBOOK_FILENAME = "2026 Main Schedule draft 11.xlsx"
MATCH_SCHEDULE_OVERRIDES_SIDECAR_FILENAME = "match_schedule_overrides.json"

DEFAULT_EVENT_CODES = ("BB", "MVB", "WVB")

# Translate the CLI-facing short codes (--events BB,MVB,WVB) into the
# "sport" vocabulary master_schedule.py already uses ("Basketball", "MVB",
# "WVB", "BC") so resource resolution (_resolve_resource_id) and the
# event/prefix/resource-type lookup tables can be reused as-is.
_SPORT_BY_EVENT_CODE = {"BB": "Basketball", "MVB": "MVB", "WVB": "WVB", "BC": "BC"}
_EVENT_CODE_BY_SPORT = {sport: code for code, sport in _SPORT_BY_EVENT_CODE.items()}

_CANONICAL_RGB_BY_SPORT = {
    sport: SPORT_STYLES[event_name].fill_color
    for sport, event_name in master_schedule._EVENT_BY_SPORT.items()
    if event_name in SPORT_STYLES
}
_COLOR_MATCH_DISTANCE_THRESHOLD = 15000

_CONTROL_WORDS = master_schedule._CONTROL_WORDS | {
    "SET UP - TAPE LINES",
    "SET UP - AUDIO / VISUAL",
    "SCRIPTURE",
    "MEMORIZATION",
    "CONTEST",
    "[CLASS RM]",
    "PLAYOFF / 3RD / FINAL",
    "BASKETBALL 3RD",
    "BASKETBALL FINAL",
}

_TEAM_CODE_RE = re.compile(r"^[A-Z0-9]{2,5}$")


def default_workbook_path(data_dir: Path) -> Path:
    return Path(data_dir) / MATCH_SCHEDULE_OVERRIDES_WORKBOOK_FILENAME


def default_sidecar_path(base_dir: Path) -> Path:
    return Path(base_dir) / MATCH_SCHEDULE_OVERRIDES_SIDECAR_FILENAME


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _team_code(value: Any) -> str:
    return _clean_text(value).upper()


def _time_label(value: Any) -> Optional[str]:
    """Parse a time cell into 24-hour "HH:MM".

    Draft 11 stores times as plain "12:00 PM"-style text rather than Excel
    time values or master_schedule.py's assumed 24-hour "H:MM[:SS]" text, so
    this needs its own AM/PM-aware parser rather than reusing
    master_schedule._time_label.
    """
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, _dt_time):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = _clean_text(value)
    if not text or text.casefold() == "legend":
        return None
    match = re.fullmatch(r"(\d{1,2}):(\d{2})\s*([AP]M)?", text, re.I)
    if not match:
        return None
    hour, minute, meridiem = int(match.group(1)), int(match.group(2)), match.group(3)
    if meridiem:
        hour %= 12
        if meridiem.upper() == "PM":
            hour += 12
    return f"{hour % 24:02d}:{minute:02d}"


def _is_control_text(value: Any) -> bool:
    text = _clean_text(value).upper()
    if not text:
        return True
    if text in _CONTROL_WORDS:
        return True
    if re.fullmatch(r"(BB|MVB|WVB|BC)\s+(QF|SF|FINAL|3RD)", text):
        return True
    if re.search(r"CEREMONY|SERVICE|SET\s*UP", text):
        return True
    return False


def _rgb_of(cell) -> Optional[str]:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color.type != "rgb" or not isinstance(color.rgb, str) or len(color.rgb) < 6:
        return None
    return color.rgb[-6:].upper()


def _rgb_distance(a: str, b: str) -> int:
    return sum(
        (int(a[i : i + 2], 16) - int(b[i : i + 2], 16)) ** 2 for i in (0, 2, 4)
    )


def _nearest_sport_by_color(rgb: Optional[str]) -> Optional[str]:
    if not rgb:
        return None
    best_sport, best_distance = None, None
    for sport, canonical_rgb in _CANONICAL_RGB_BY_SPORT.items():
        distance = _rgb_distance(rgb, canonical_rgb)
        if best_distance is None or distance < best_distance:
            best_sport, best_distance = sport, distance
    if best_distance is not None and best_distance <= _COLOR_MATCH_DISTANCE_THRESHOLD:
        return best_sport
    return None


def _venue_blocks_for_header_row(ws, row_idx: int) -> List[Tuple[int, int, str]]:
    """Return (start_col, end_col, venue_name) for each venue on a header row.

    Column widths vary by date section (weekend-1 pool play is a fixed 3-col
    block per court; weekend-2 playoff headers are narrower/wider), so widths
    are read from the header row's own merged-cell ranges rather than assumed.
    """
    blocks: List[Tuple[int, int, str]] = []
    covered: set[int] = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row != row_idx or merged_range.max_row != row_idx:
            continue
        if merged_range.min_col < 2:
            continue
        name = _clean_text(ws.cell(row=row_idx, column=merged_range.min_col).value)
        if name:
            blocks.append((merged_range.min_col, merged_range.max_col, name))
            covered.update(range(merged_range.min_col, merged_range.max_col + 1))
    for col in range(2, ws.max_column + 1):
        if col in covered:
            continue
        name = _clean_text(ws.cell(row=row_idx, column=col).value)
        if name:
            blocks.append((col, col, name))
    return sorted(blocks)


def _classify_block(
    cells: Sequence[Any],
    day: str,
    start_time: str,
    slot: str,
    venue_name: str,
    sheet_name: str,
) -> Optional[Dict[str, Any]]:
    values = [_clean_text(cell.value) for cell in cells]
    nonempty = [value for value in values if value]
    if not nonempty:
        return None
    if len(nonempty) == 1 and _is_control_text(nonempty[0]):
        return None

    cell_range = (
        f"{cells[0].coordinate}:{cells[-1].coordinate}" if len(cells) > 1 else cells[0].coordinate
    )
    raw_value = " / ".join(nonempty)
    base = {
        "source_sheet": sheet_name,
        "source_cell": cell_range,
        "day": day,
        "start_time": start_time,
        "slot": slot,
        "visual_venue": venue_name,
        "raw_value": raw_value,
    }

    rgb = next((_rgb_of(cell) for cell, value in zip(cells, values) if value), None)
    sport = _nearest_sport_by_color(rgb)
    looks_team_coded = any(_TEAM_CODE_RE.match(value.upper()) for value in nonempty)

    if len(values) == 3 and values[1].strip().casefold() == "v" and values[0] and values[2]:
        team_a, team_b = _team_code(values[0]), _team_code(values[2])
        if sport not in ("Basketball", "MVB", "WVB"):
            return {
                **base,
                "kind": "block",
                "severity": "error" if looks_team_coded else "warning",
                "note": f"{cell_range}: {raw_value!r} has a 'v' pairing but its fill color "
                "doesn't resolve to Basketball/MVB/WVB",
            }
        if _is_bye(team_a) or _is_bye(team_b):
            return {
                **base,
                "kind": "bye",
                "sport": sport,
                "event": master_schedule._EVENT_BY_SPORT[sport],
                "team_a": None if _is_bye(team_a) else team_a,
                "team_b": None if _is_bye(team_b) else team_b,
            }
        return {
            **base,
            "kind": "two_team_game",
            "sport": sport,
            "event": master_schedule._EVENT_BY_SPORT[sport],
            "resource_type": master_schedule._RESOURCE_TYPE_BY_SPORT[sport],
            "team_a": team_a,
            "team_b": team_b,
        }

    if len(values) == 3 and len(nonempty) == 3 and all(
        _TEAM_CODE_RE.match(value.upper()) for value in values
    ):
        if sport != "BC":
            return {
                **base,
                "kind": "block",
                "severity": "warning",
                "note": f"{cell_range}: {raw_value!r} looks like a 3-team row but its fill "
                "color doesn't resolve to Bible Challenge",
            }
        return {
            **base,
            "kind": "three_team_game",
            "sport": "BC",
            "event": master_schedule._EVENT_BY_SPORT["BC"],
            "resource_type": master_schedule._RESOURCE_TYPE_BY_SPORT["BC"],
            "teams": [value.upper() for value in values],
        }

    return {
        **base,
        "kind": "block",
        "severity": "error" if (sport and looks_team_coded) else "warning",
        "note": f"{cell_range}: {raw_value!r} is not a recognized team-code pattern",
    }


def parse_match_schedule_overrides_workbook(workbook_path: Path) -> Dict[str, Any]:
    """Parse the visual match schedule workbook into raw override intent rows."""
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb.active

    rows: List[Dict[str, Any]] = []
    diagnostics: Dict[str, Any] = {"warnings": [], "errors": [], "unmapped_cells": []}
    current_day = ""
    venue_blocks: List[Tuple[int, int, str]] = []

    for row_idx in range(1, ws.max_row + 1):
        header = master_schedule._date_header_day(ws.cell(row=row_idx, column=1).value)
        if header:
            _current_date, current_day = header
            venue_blocks = _venue_blocks_for_header_row(ws, row_idx)
            continue
        start_time = _time_label(ws.cell(row=row_idx, column=1).value)
        if not current_day or not start_time or not venue_blocks:
            continue
        slot = master_schedule._slot_label(current_day, start_time)

        for start_col, end_col, venue_name in venue_blocks:
            cells = [ws.cell(row=row_idx, column=col) for col in range(start_col, end_col + 1)]
            result = _classify_block(cells, current_day, start_time, slot, venue_name, ws.title)
            if result is None:
                continue
            rows.append(result)
            if result.get("kind") == "block":
                note = result["note"]
                if result.get("severity") == "error":
                    diagnostics["errors"].append(note)
                else:
                    diagnostics["unmapped_cells"].append(note)

    diagnostics["row_count"] = len(rows)
    diagnostics["candidate_count"] = sum(
        1 for row in rows if row.get("kind") in ("two_team_game", "three_team_game")
    )
    diagnostics["bye_count"] = sum(1 for row in rows if row.get("kind") == "bye")
    diagnostics["block_count"] = sum(1 for row in rows if row.get("kind") == "block")

    return {
        "version": 1,
        "source_workbook": str(workbook_path),
        "source_sheet": ws.title,
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "diagnostics": diagnostics,
    }


def _pair_lookup(games: Sequence[Mapping[str, Any]]) -> Dict[Tuple[str, frozenset], str]:
    lookup: Dict[Tuple[str, frozenset], str] = {}
    for game in games:
        event = _clean_text(game.get("event"))
        team_a = _team_code(game.get("team_a_label"))
        team_b = _team_code(game.get("team_b_label"))
        game_id = _clean_text(game.get("game_id"))
        if event and team_a and team_b and game_id:
            lookup[(event, frozenset((team_a, team_b)))] = game_id
    return lookup


def resolve_match_schedule_overrides(
    payload: Mapping[str, Any],
    *,
    events: Sequence[str] = DEFAULT_EVENT_CODES,
    roster_rows: Optional[List[Dict[str, Any]]] = None,
    games: Optional[Sequence[Mapping[str, Any]]] = None,
    resources: Optional[Sequence[Mapping[str, Any]]] = None,
) -> Dict[str, Any]:
    """Validate parsed rows against the roster and resolve/pin against schedule_input.

    Rows for events not in ``events`` are ignored entirely (default scope is
    BB/MVB/WVB only, per Issue #214 - Bible Challenge already has its own
    team-code mechanism via master_schedule.py's three-team rows).
    """
    wanted_sports = {_SPORT_BY_EVENT_CODE[code] for code in events if code in _SPORT_BY_EVENT_CODE}
    roster_codes = _extract_roster_team_codes(roster_rows) if roster_rows else {}
    games = games or []
    resources = resources or []
    pair_lookup = _pair_lookup(games)

    errors: List[str] = []
    warnings: List[str] = []
    resolved_games: List[Dict[str, Any]] = []
    resolved_slots: List[Dict[str, Any]] = []
    seen_team_at_slot: Dict[Tuple[str, str], str] = {}
    next_round_by_event: Counter = Counter()

    if roster_rows is None:
        warnings.append("No roster context supplied; team-code validation was skipped.")

    for row in payload.get("rows", []) or []:
        if row.get("kind") != "two_team_game" or row.get("sport") not in wanted_sports:
            continue
        event = row["event"]
        cell = row.get("source_cell", "<unknown>")
        team_a, team_b = row["team_a"], row["team_b"]

        codes_ok = True
        if roster_rows:
            valid_codes = roster_codes.get(event, set())
            for team in (team_a, team_b):
                if team in valid_codes:
                    continue
                elsewhere = sorted(
                    other_event
                    for other_event, codes in roster_codes.items()
                    if other_event != event and team in codes
                )
                hint = f" (registered for {', '.join(elsewhere)} instead)" if elsewhere else ""
                errors.append(f"{cell}: team code {team!r} not found in {event} roster{hint}")
                codes_ok = False

        for team in (team_a, team_b):
            # Scoped per event: a church code fielding both an MVB and a WVB
            # team is two different rosters of people, so both legitimately
            # playing at the same slot is not a conflict.
            key = (row["slot"], event, team)
            if key in seen_team_at_slot:
                errors.append(
                    f"{cell}: team {team!r} is double-booked in {event} at {row['slot']} "
                    f"(also at {seen_team_at_slot[key]})"
                )
                codes_ok = False
            else:
                seen_team_at_slot[key] = cell

        if not codes_ok:
            continue

        game_id = pair_lookup.get((event, frozenset((team_a, team_b))))
        if game_id is None:
            prefix = master_schedule._PREFIX_BY_SPORT[row["sport"]]
            next_round_by_event[event] += 1
            game_id = f"{prefix}-LOC-{next_round_by_event[event]:02d}"
            duration = int(
                COURT_ESTIMATE_MINUTES_PER_GAME.get(event, COURT_ESTIMATE_DEFAULT_MINUTES_PER_GAME)
            )
            resolved_games.append({
                "game_id": game_id,
                "event": event,
                "stage": "Pool",
                "pool_id": "",
                "team_a_id": f"{prefix}::{team_a}",
                "team_b_id": f"{prefix}::{team_b}",
                "team_a_label": team_a,
                "team_b_label": team_b,
                "resource_type": row["resource_type"],
                "duration_minutes": duration,
                "x_match_schedule_source": payload.get("source_workbook", ""),
                "x_match_schedule_sheet": row.get("source_sheet", ""),
                "x_match_schedule_cell": cell,
            })
            warnings.append(
                f"{cell}: {team_a} vs {team_b} ({event}) has no matching generated pool game; "
                f"created {game_id} from the visual schedule. Run assign-pools / "
                "import-team-matchups first if this pairing should already exist."
            )

        resource_row = {
            "resource_type": row["resource_type"],
            "day": row["day"],
            "slot": row["slot"],
            "sport": row["sport"],
            "visual_venue": row["visual_venue"],
        }
        resource_id, reason = master_schedule._resolve_resource_id(resource_row, resources)
        if not resource_id:
            warnings.append(f"{cell}: {reason or 'could not resolve a court/resource'}")
            continue

        resolved_slots.append({
            "game_id": game_id,
            "event": event,
            "resource_id": resource_id,
            "slot": row["slot"],
            "x_match_schedule_team_a_label": team_a,
            "x_match_schedule_team_b_label": team_b,
            "x_match_schedule_source": payload.get("source_workbook", ""),
            "x_match_schedule_sheet": row.get("source_sheet", ""),
            "x_match_schedule_cell": cell,
            "x_match_schedule_raw": row.get("raw_value", ""),
        })

    seen_slot_pins: Dict[str, str] = {}
    deduped_slots: List[Dict[str, Any]] = []
    for slot_row in resolved_slots:
        key = f"{slot_row['resource_id']}|{slot_row['slot']}"
        if key in seen_slot_pins:
            errors.append(
                f"{slot_row['x_match_schedule_cell']}: duplicate pin for "
                f"{slot_row['resource_id']} {slot_row['slot']}; first seen at "
                f"{seen_slot_pins[key]}"
            )
            continue
        seen_slot_pins[key] = slot_row["x_match_schedule_cell"]
        deduped_slots.append(slot_row)

    return {
        "resolved_games": resolved_games,
        "resolved_slots": deduped_slots,
        "fixed_count": len(deduped_slots),
        "created_game_count": len(resolved_games),
        "errors": errors,
        "warnings": warnings,
    }


def build_match_schedule_overrides_payload(
    workbook_path: Path,
    *,
    events: Sequence[str] = DEFAULT_EVENT_CODES,
    roster_rows: Optional[List[Dict[str, Any]]] = None,
    games: Optional[Sequence[Mapping[str, Any]]] = None,
    resources: Optional[Sequence[Mapping[str, Any]]] = None,
) -> Dict[str, Any]:
    payload = parse_match_schedule_overrides_workbook(workbook_path)
    payload["events"] = list(events)
    payload["validation"] = resolve_match_schedule_overrides(
        payload, events=events, roster_rows=roster_rows, games=games, resources=resources,
    )
    return payload


def load_match_schedule_overrides_sidecar(path: Optional[Path]) -> Dict[str, Any]:
    if path is None:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def write_match_schedule_overrides_sidecar(payload: Mapping[str, Any], path: Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def summarize_payload_for_log(payload: Mapping[str, Any]) -> List[str]:
    diagnostics = payload.get("diagnostics", {}) or {}
    lines = [
        "match schedule overrides: "
        f"{diagnostics.get('candidate_count', 0)} candidate game(s), "
        f"{diagnostics.get('bye_count', 0)} bye row(s), "
        f"{diagnostics.get('block_count', 0)} unmapped block(s)"
    ]
    validation = payload.get("validation") or {}
    if validation:
        lines.append(
            "match schedule overrides validation: "
            f"{validation.get('fixed_count', 0)} pinned assignment(s), "
            f"{validation.get('created_game_count', 0)} newly-created game(s)"
        )
    return lines


def _fixed_slot_key(slot: Mapping[str, Any]) -> Tuple[str, str]:
    return (_clean_text(slot.get("resource_id")), _clean_text(slot.get("slot")))


def _is_supersedable_master_pool_pin(slot: Mapping[str, Any]) -> bool:
    """Return True for older visual master-schedule pool pins this import replaces.

    The numbered master schedule and the team-code visual schedule describe the
    same BB/MVB/WVB pool cells.  When both sidecars exist, the team-code sidecar
    is the more authoritative source because it names the actual pairing in the
    cell; keep hard conflicts for every other fixed-slot source.
    """
    event = _clean_text(slot.get("event"))
    stage = _clean_text(slot.get("stage"))
    has_master_source = any(
        _clean_text(slot.get(key))
        for key in ("x_master_schedule_source", "x_master_schedule_cell", "x_master_schedule_raw")
    )
    return (
        has_master_source
        and stage.casefold() == "pool"
        and event
        in {
            master_schedule._EVENT_BY_SPORT["Basketball"],
            master_schedule._EVENT_BY_SPORT["MVB"],
            master_schedule._EVENT_BY_SPORT["WVB"],
        }
    )


def merge_match_schedule_overrides_into_schedule_input(
    games: Sequence[Mapping[str, Any]],
    playoff_slots: Sequence[Mapping[str, Any]],
    payload: Mapping[str, Any],
    resources: Sequence[Mapping[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    """Merge a resolved match_schedule_overrides payload into schedule_input.

    Returns (games, playoff_slots, summary). New games from the visual
    schedule are appended (they never replace an existing game_id); pinned
    slots are merged into playoff_slots keyed by game_id, same as
    master_schedule.merge_master_schedule_into_playoff_slots.
    """
    if not payload:
        return [dict(game) for game in games], [dict(slot) for slot in playoff_slots], {}

    validation = resolve_match_schedule_overrides(
        payload,
        events=payload.get("events", DEFAULT_EVENT_CODES),
        games=games,
        resources=resources,
    )

    existing_ids = {_clean_text(game.get("game_id")) for game in games}
    merged_games = [dict(game) for game in games]
    for new_game in validation["resolved_games"]:
        if new_game["game_id"] not in existing_ids:
            merged_games.append(new_game)
            existing_ids.add(new_game["game_id"])

    merged_by_game: Dict[str, Dict[str, Any]] = {
        _clean_text(slot.get("game_id")): dict(slot)
        for slot in playoff_slots
        if _clean_text(slot.get("game_id"))
    }
    occupied_slots: Dict[Tuple[str, str], str] = {
        _fixed_slot_key(slot): _clean_text(slot.get("game_id"))
        for slot in playoff_slots
        if _clean_text(slot.get("resource_id"))
        and _clean_text(slot.get("slot"))
        and _clean_text(slot.get("game_id"))
    }
    for slot_row in validation["resolved_slots"]:
        game_id = _clean_text(slot_row.get("game_id"))
        previous_game_pin = merged_by_game.get(game_id)
        if previous_game_pin:
            previous_slot_key = _fixed_slot_key(previous_game_pin)
            if occupied_slots.get(previous_slot_key) == game_id:
                occupied_slots.pop(previous_slot_key, None)

        slot_key = _fixed_slot_key(slot_row)
        existing_game_id = occupied_slots.get(slot_key)
        if existing_game_id and existing_game_id != game_id:
            existing_slot = merged_by_game.get(existing_game_id, {})
            if _is_supersedable_master_pool_pin(existing_slot):
                merged_by_game.pop(existing_game_id, None)
                occupied_slots.pop(slot_key, None)
                validation["warnings"].append(
                    f"{slot_row.get('x_match_schedule_cell', '<unknown>')}: "
                    f"team-code match schedule override for game {game_id} superseded "
                    f"numbered master-schedule pool pin {existing_game_id} at "
                    f"{slot_key[0]} {slot_key[1]}"
                )
            else:
                validation["errors"].append(
                    f"{slot_row.get('x_match_schedule_cell', '<unknown>')}: "
                    f"match schedule override for game {game_id} conflicts with existing fixed "
                    f"assignment {existing_game_id} at {slot_key[0]} {slot_key[1]}"
                )
                continue
        if occupied_slots.get(slot_key) and occupied_slots[slot_key] != game_id:
            validation["errors"].append(
                f"{slot_row.get('x_match_schedule_cell', '<unknown>')}: "
                f"match schedule override for game {game_id} conflicts with existing fixed "
                f"assignment {occupied_slots[slot_key]} at {slot_key[0]} {slot_key[1]}"
            )
            continue
        merged_by_game[game_id] = slot_row
        if slot_key[0] and slot_key[1]:
            occupied_slots[slot_key] = game_id

    return (
        merged_games,
        list(merged_by_game.values()),
        {
            "source_workbook": payload.get("source_workbook", ""),
            "source_sheet": payload.get("source_sheet", ""),
            "imported_at": payload.get("imported_at", ""),
            **{key: value for key, value in validation.items() if key not in ("resolved_games", "resolved_slots")},
        },
    )
