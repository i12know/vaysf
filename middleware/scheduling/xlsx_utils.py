"""xlsx_utils — low-level Excel / spreadsheet helpers extracted from ScheduleWorkbookBuilder.

Pure functions; no class state.  Extracted as part of Issue #152.
"""
import os
import re
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import RESOURCE_TYPE_ALIASES, RESOURCE_ID_PREFIX_BY_TYPE


def _clean_excel_text(val) -> str:
    """Normalize spreadsheet cells so blanks/NaN become an empty string."""
    if pd.isna(val):
        return ""
    return str(val).strip()


def _float_from_excel(val, default: float) -> float:
    """Convert spreadsheet cells to float while treating blanks/NaN as a default."""
    if pd.isna(val) or val in (None, ""):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _normalize_resource_type_name(val) -> str:
    """Normalize venue resource names to the canonical scheduler vocabulary."""
    cleaned = _clean_excel_text(val)
    if not cleaned:
        return ""
    key = re.sub(r"[\s_-]+", " ", cleaned).strip().casefold()
    return RESOURCE_TYPE_ALIASES.get(key, cleaned)


def _resource_id_prefix(resource_type: str) -> str:
    """Return the canonical resource-id prefix for one resource type."""
    return RESOURCE_ID_PREFIX_BY_TYPE.get(
        resource_type,
        resource_type.split()[0][:3].upper(),
    )


def _ordinal(n: int) -> str:
    """Return 1st / 2nd / 3rd / 4th style ordinals."""
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _day_sort_key(day_label: str) -> Tuple[int, int, str]:
    """Sort logical day labels like Fri-1, Sat-1, Sun-2 chronologically."""
    cleaned = str(day_label or "").strip()
    if not cleaned:
        return (99, 99, "")
    match = re.fullmatch(r"([A-Za-z]+)-(\d+)", cleaned)
    if not match:
        return (99, 99, cleaned)
    prefix = match.group(1)
    cycle = int(match.group(2))
    weekday_order = {
        "Mon": 0, "Tue": 1, "Wed": 2, "Thu": 3,
        "Fri": 4, "Sat": 5, "Sun": 6, "Day": 7,
    }
    return (cycle, weekday_order.get(prefix, 99), cleaned)


def _day_display_label(day_label: str, short: bool = False) -> str:
    """Return a human-friendly label for one logical day key."""
    cleaned = str(day_label or "").strip()
    match = re.fullmatch(r"([A-Za-z]+)-(\d+)", cleaned)
    if not match:
        return cleaned
    prefix = match.group(1)
    cycle = int(match.group(2))
    names = {
        "Mon": ("Monday", "Mon"),
        "Tue": ("Tuesday", "Tue"),
        "Wed": ("Wednesday", "Wed"),
        "Thu": ("Thursday", "Thu"),
        "Fri": ("Friday", "Fri"),
        "Sat": ("Saturday", "Sat"),
        "Sun": ("Sunday", "Sun"),
        "Day": ("Day", "Day"),
    }
    long_name, short_name = names.get(prefix, (prefix, prefix))
    if prefix == "Day":
        return cleaned
    ordinal = _ordinal(cycle)
    return f"{ordinal} {short_name if short else long_name}"


def _coerce_excel_date(val) -> Optional[datetime]:
    """Convert an Excel date-like cell to datetime, or None when unavailable."""
    if pd.isna(val) or val in (None, ""):
        return None
    if isinstance(val, datetime):
        return val
    try:
        parsed = pd.to_datetime(val)
    except Exception:
        return None
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _derive_day_labels_from_dates(values: List[Any]) -> Dict[str, str]:
    """Map unique venue dates to logical labels such as Fri-1 / Sat-1 / Sun-2."""
    unique_dates: List[datetime] = []
    seen_keys: set[str] = set()
    for value in values:
        parsed = _coerce_excel_date(value)
        if not parsed:
            continue
        key = parsed.date().isoformat()
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_dates.append(parsed)

    unique_dates.sort()
    labels: Dict[str, str] = {}
    weekday_counts: Dict[str, int] = defaultdict(int)
    weekday_prefix = {
        0: "Mon",
        1: "Tue",
        2: "Wed",
        3: "Thu",
        4: "Fri",
        5: "Sat",
        6: "Sun",
    }
    for dt_value in unique_dates:
        prefix = weekday_prefix.get(dt_value.weekday(), "Day")
        weekday_counts[prefix] += 1
        label = f"{prefix}-{weekday_counts[prefix]}"
        labels[dt_value.date().isoformat()] = label
    return labels


def _set_excel_comment(cell, note: Optional[str]) -> None:
    """Attach a standard Excel note/comment to a cell when note text exists."""
    if not note:
        return
    from openpyxl.comments import Comment

    cell.comment = Comment(note, "VAYSF")


def _make_excel_note_shapes_visible(xlsx_path: Path) -> bool:
    """Patch openpyxl's hidden VML note shapes so Excel opens them by default.

    openpyxl writes comments as legacy VML notes and currently ignores
    Comment.visible. The workbook XML is otherwise valid, so we patch only
    the generated VML note shapes after save, preserving openpyxl's existing
    Excel namespace prefix instead of injecting a conflicting namespace.
    """
    xlsx_path = Path(xlsx_path)
    excel_ns = "urn:schemas-microsoft-com:office:excel"
    patched_any = False
    tmp_path = xlsx_path.with_suffix(f"{xlsx_path.suffix}.tmp")

    try:
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            tmp_path, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (
                    item.filename.startswith("xl/drawings/")
                    and item.filename.endswith(".vml")
                    and b'ObjectType="Note"' in data
                ):
                    text = data.decode("utf-8")
                    prefix_match = re.search(
                        rf'xmlns:([A-Za-z_][\w.-]*)="{re.escape(excel_ns)}"',
                        text,
                    )
                    if prefix_match:
                        excel_prefix = prefix_match.group(1)
                        client_data_re = re.compile(
                            rf"(<{re.escape(excel_prefix)}:ClientData\b"
                            rf"(?=[^>]*\bObjectType=\"Note\")[^>]*>)"
                            rf"(.*?)"
                            rf"(</{re.escape(excel_prefix)}:ClientData>)",
                            re.DOTALL,
                        )

                        def _show_note(match: re.Match) -> str:
                            nonlocal patched_any
                            body = match.group(2)
                            if f"<{excel_prefix}:Visible" not in body:
                                patched_any = True
                                body = f"{body}<{excel_prefix}:Visible />"
                            return f"{match.group(1)}{body}{match.group(3)}"

                        text = client_data_re.sub(_show_note, text)

                    visible_text = re.sub(
                        r"visibility\s*:\s*hidden",
                        "visibility:visible",
                        text,
                    )
                    if visible_text != text:
                        patched_any = True
                    data = visible_text.encode("utf-8")

                zout.writestr(item, data)

        if patched_any:
            os.replace(tmp_path, xlsx_path)
        else:
            tmp_path.unlink(missing_ok=True)
        return patched_any
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        logger.warning(f"Could not patch Excel note visibility in {xlsx_path}: {exc}")
        return False


def _stamp_tab_status_banner(
    ws,
    status: str,
    guidance: str,
    *,
    fill_color: str,
) -> None:
    """Add a non-disruptive tab role banner outside the data table."""
    start_col = max(ws.max_column or 1, 1) + 2
    end_col = start_col + 3
    ws.merge_cells(
        start_row=1,
        start_column=start_col,
        end_row=1,
        end_column=end_col,
    )
    ws.merge_cells(
        start_row=2,
        start_column=start_col,
        end_row=2,
        end_column=end_col,
    )

    title_cell = ws.cell(row=1, column=start_col, value=f"STATUS: {status}")
    body_cell = ws.cell(row=2, column=start_col, value=guidance)
    fill = PatternFill("solid", fgColor=fill_color)
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    for row_idx in (1, 2):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    title_cell.font = Font(bold=True, color="1F1F1F")
    body_cell.font = Font(italic=True, color="1F1F1F")
    ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 22)
    ws.row_dimensions[2].height = max(ws.row_dimensions[2].height or 0, 36)
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18
    ws.sheet_properties.tabColor = fill_color


def _annotate_header_row(
    ws,
    row_idx: int,
    n_cols: int,
    header_notes: Dict[str, str],
    *,
    width_map: Optional[Dict[str, float]] = None,
    freeze_panes: Optional[str] = None,
    autofilter: bool = False,
) -> None:
    """Add consistent header comments and simple usability affordances."""
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if autofilter and n_cols > 0:
        ws.auto_filter.ref = f"A{row_idx}:{get_column_letter(n_cols)}{row_idx}"
    if width_map:
        for col_letter, width in width_map.items():
            ws.column_dimensions[col_letter].width = width

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        header = str(cell.value or "").strip()
        _set_excel_comment(cell, header_notes.get(header))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if row_idx not in ws.row_dimensions:
        ws.row_dimensions[row_idx].height = 30
    else:
        ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, 30)


_TAB_STATUS_GUIDE: Dict[str, Tuple[str, str, str]] = {
    "Venue-Input": (
        "EDITABLE INPUT",
        "Edit booked venue resource blocks here, then rerun export-church-teams.",
        "D9EAD3",
    ),
    "Gym-Modes": (
        "EDITABLE INPUT",
        "Edit physical gym mode capacities here when a gym can switch sports.",
        "D9EAD3",
    ),
    "Playoff-Slots": (
        "EDITABLE OVERRIDE INPUT",
        "Optional pinned playoff overrides. Copy exact resource_id and slot values from Schedule-Input.",
        "FCE5CD",
    ),
    "Summary": (
        "READ-ONLY OUTPUT",
        "Generated guide or status summary. Do not edit; rerun the source command instead.",
        "D9EAF7",
    ),
    "Contacts-Status": (
        "READ-ONLY OUTPUT",
        "Generated church contact and approval snapshot. Do not edit this tab.",
        "D9EAF7",
    ),
    "Roster": (
        "GENERATED DATA SOURCE",
        "Generated roster source for scheduling. Fix source data, then rerun export-church-teams.",
        "D9EAF7",
    ),
    "Validation-Issues": (
        "GENERATED DATA SOURCE",
        "Generated validation source for scheduling. Resolve issues upstream, then rerun export-church-teams.",
        "D9EAF7",
    ),
    "Venue-Estimator": (
        "READ-ONLY PLANNING OUTPUT",
        "Use for capacity planning only. Change venue capacity in venue_input.xlsx.",
        "D9EAF7",
    ),
    "Court-Schedule-Sketch": (
        "READ-ONLY PLANNING OUTPUT",
        "Layer-1 planning sketch. Rerun build-schedule-workbook after changing inputs.",
        "D9EAF7",
    ),
    "Pod-Divisions": (
        "READ-ONLY OUTPUT",
        "Generated pod division planning view. Do not edit this tab.",
        "D9EAF7",
    ),
    "Pod-Entries-Review": (
        "READ-ONLY REVIEW OUTPUT",
        "Review pod entries here, but fix roster data upstream instead of editing cells.",
        "D9EAF7",
    ),
    "Pod-Resource-Estimate": (
        "READ-ONLY CAPACITY CHECK",
        "Use to compare pod demand against venue capacity. Edit venue_input.xlsx for changes.",
        "D9EAF7",
    ),
    "Schedule-Input": (
        "GENERATED LOOKUP / MACHINE CONTRACT VIEW",
        "Copy exact resource_id, slot, and game_id values from here; do not hand-edit.",
        "D9EAF7",
    ),
    "Pool-Assignment": (
        "TEMPORARY EDIT SURFACE",
        "Edit seeds or pool placements here, then rerun assign-pools --workbook.",
        "FFF2CC",
    ),
    "Gym-Allocation": (
        "READ-ONLY ALLOCATION AUDIT",
        "Generated gym-mode allocation audit. Change venue inputs, then rerun the workflow.",
        "D9EAF7",
    ),
    "Schedule-by-Time": (
        "FINAL OUTPUT",
        "Final schedule view by time. Do not edit; rerun the scheduler if inputs change.",
        "D9EAD3",
    ),
    "Schedule-by-Sport": (
        "FINAL OUTPUT",
        "Final schedule view by sport. Do not edit; rerun the scheduler if inputs change.",
        "D9EAD3",
    ),
    "Master-Schedule": (
        "FINAL OUTPUT",
        "Final master grid. Do not edit; rerun the scheduler if inputs change.",
        "D9EAD3",
    ),
    "Conflict-Audit": (
        "AUDIT OUTPUT",
        "Audit conflicts and warnings here before publishing the final schedule.",
        "F4CCCC",
    ),
    "Schedule-Diagnostics": (
        "DIAGNOSTIC OUTPUT",
        "Review suggested next actions before changing venue inputs or rerunning the scheduler.",
        "FCE5CD",
    ),
}


def _stamp_known_tab_statuses(wb, *, default_unknown: Optional[Tuple[str, str, str]] = None) -> None:
    """Stamp known workbook sheets with operator-facing role guidance."""
    for ws in wb.worksheets:
        guide = _TAB_STATUS_GUIDE.get(ws.title)
        if guide is None:
            guide = default_unknown
        if guide is None:
            continue
        status, guidance, fill_color = guide
        _stamp_tab_status_banner(ws, status, guidance, fill_color=fill_color)


def _parse_hour(val) -> float:
    """Convert a cell value to a decimal hour (e.g. datetime.time(13,0) → 13.0)."""
    import datetime as _dt
    if pd.isna(val):
        return 0.0
    if isinstance(val, _dt.time):
        return val.hour + val.minute / 60.0
    if isinstance(val, str) and ":" in val:
        try:
            hour_str, minute_str = val.split(":", 1)
            return int(hour_str) + int(minute_str) / 60.0
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _read_xlsx_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    """Read one sheet of an exported workbook into a list of row dicts.

    NaN cells are normalized to None so the scheduling builders' common
    `str(row.get(col) or "")` idiom collapses blanks to empty strings
    (a bare NaN float is truthy and would otherwise stringify to 'nan').
    Returns an empty list with a WARNING when the sheet is absent.
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, engine="openpyxl")
    except Exception as e:
        logger.warning(f"Could not read '{sheet_name}' tab from {xlsx_path}: {e}")
        return []
    # astype(object) first: assigning None to a float64 column silently
    # reverts to NaN, so the column must be object-typed before the mask.
    df = df.astype(object).where(pd.notna(df), None)
    rows = df.to_dict("records")
    logger.debug(f"Read {len(rows)} rows from '{sheet_name}' tab of {xlsx_path}")
    return rows
