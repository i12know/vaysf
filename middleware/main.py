# main.py

import os
import sys
import argparse
import time
import datetime
import json
from typing import Optional
from loguru import logger
import schedule
from tenacity import retry, stop_after_attempt, wait_exponential
from config import Config, DATA_DIR, EXPORT_DIR
from pathlib import Path

# Add parent directory to import path
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)

from sync.manager import SyncManager
from chmeetings.backend_connector import ChMeetingsConnector  # Import for export command
from wordpress.frontend_connector import WordPressConnector   # Import for export command
from church_teams_export import ChurchTeamsExporter           # Import for export command
from season_reset import SeasonResetter                       # Import for reset-season command

def parse_args() -> argparse.Namespace:
    """Parse command-line arguments for the VAYSF middleware."""
    parser = argparse.ArgumentParser(description="Sports Fest 2026 ChMeetings Integration")
    subparsers = parser.add_subparsers(dest="command", help="Command to run", required=True)

    # Sync command
    sync_parser = subparsers.add_parser("sync", help="Sync data between systems")
    sync_parser.add_argument("--type", choices=["churches", "participants", "approvals", "validation", "full"],
                             default="full", help="Type of data to sync")
    sync_parser.add_argument("--chm-id", type=str, default=None,
                             help="Optional ChMeetings ID to target when --type is 'participants' or 'approvals'")
    sync_parser.add_argument("--excel-fallback", action="store_true",
                             help="Use Excel export instead of API for syncing approvals to ChMeetings")

    # Sync-churches command
    sync_churches_parser = subparsers.add_parser("sync-churches", help="Sync churches from Excel file")
    sync_churches_parser.add_argument("--file", default=os.path.join("data", "Church Application Form.xlsx"),
                                      help="Path to the church Excel file")

    # Group assignment command
    group_assignment_parser = subparsers.add_parser(
        "assign-groups",
        help="Assign ChMeetings people to church team groups via direct API calls",
    )
    group_assignment_parser.add_argument("--output", help="Output directory path",
                                         default=DATA_DIR)
    group_assignment_parser.add_argument(
        "--file",
        default=None,
        help="Optional path to the current-season Individual Application export "
             "used to limit assignments to this year's registrants",
    )
    group_assignment_parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview only - show who would be assigned without making API calls",
    )

    # Team-group clearing command
    clear_team_groups_parser = subparsers.add_parser(
        "clear-team-groups",
        help="Clear memberships from ChMeetings Team XXX groups via direct API calls",
    )
    clear_team_groups_parser.add_argument(
        "--church-code",
        help="Limit the run to a single team group such as Team RPC",
    )
    clear_mode = clear_team_groups_parser.add_mutually_exclusive_group(required=True)
    clear_mode.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview only - show who would be removed without making API calls",
    )
    clear_mode.add_argument(
        "--execute",
        action="store_true",
        help="Actually remove current members from the target Team XXX groups",
    )

    audit_team_groups_parser = subparsers.add_parser(
        "audit-team-groups",
        help="Audit ChMeetings Team XXX groups for orphaned member IDs",
    )
    audit_team_groups_parser.add_argument(
        "--church-code",
        help="Limit the audit to a single team group such as Team GAC",
    )
    audit_team_groups_parser.add_argument(
        "--remove-orphans",
        action="store_true",
        help="Remove orphaned memberships from ChMeetings after identifying them (irreversible)",
    )

    # Export command
    export_parser = subparsers.add_parser("export-church-teams", help="Export church team status reports")
    export_parser.add_argument("--church-code", help="Export for specific church code (if omitted, exports for all churches)")
    export_parser.add_argument("--output", help="Output directory path", default=EXPORT_DIR)
    export_parser.add_argument("--force-resend-pending", action="store_true",
                            help="Resend approval emails for participants with pending/pending_approval status")
    export_parser.add_argument("--force-resend-validated1", action="store_true", 
                            help="Resend approval emails for 'validated' participants WITH data in Box 1-6 (under review)")
    export_parser.add_argument("--force-resend-validated2", action="store_true",
                            help="Resend approval emails for 'validated' participants with NO data in Box 1-6 (not reviewed yet)")
    export_parser.add_argument("--dry-run", action="store_true",
                            help="Show what would be resent without actually sending emails")
    export_parser.add_argument("--chm-id",
                            help="Limit force-resend operations to one ChMeetings ID")

    # Config command
    config_parser = subparsers.add_parser("config", help="Configure system settings")
    config_parser.add_argument("--validate", action="store_true", help="Validate current configuration")

    # Schedule command
    schedule_parser = subparsers.add_parser("schedule", help="Run scheduled sync jobs")
    schedule_parser.add_argument("--interval", type=int, default=Config.SYNC_INTERVAL_MINUTES,
                                 help="Interval in minutes between sync jobs")
    schedule_parser.add_argument("--daemon", action="store_true", help="Run as daemon process")

    # Test command
    test_parser = subparsers.add_parser("test", help="Test connectivity and functionality of systems")
    test_parser.add_argument("--system", choices=["chmeetings", "wordpress", "all"],
                             default="all", help="System to test")
    test_parser.add_argument("--test-type", choices=["connectivity", "churches", "email", "api-inspect", "all"],
                             default="connectivity", help="Type of test to run (connectivity, churches, email, api-inspect, or all)")
    test_parser.add_argument("--test-email", default=os.getenv("TEST_EMAIL", "PastorBumble@gmail.com"),
                             help="Email address for email test")

    # Inspect-person command
    inspect_parser = subparsers.add_parser(
        "inspect-person",
        help="Inspect a single ChMeetings person ID and any matching WordPress records",
    )
    inspect_parser.add_argument("--chm-id", type=str, required=True,
                                help="ChMeetings person ID to inspect")

    # Reset-season command
    reset_parser = subparsers.add_parser(
        "reset-season",
        help="Archive and clear Sports Fest custom fields for all VAY-SM members"
    )
    reset_parser.add_argument("--year", type=int, required=True,
                              help="Season year to archive (e.g. 2025)")
    reset_parser.add_argument("--dry-run", action="store_true",
                              help="Preview what would be archived/reset without making any changes")
    reset_parser.add_argument("--archive-only", action="store_true",
                              help="Write archive notes only; do not reset custom fields")
    reset_parser.add_argument("--reset-only", action="store_true",
                              help="Reset custom fields only; skip writing archive notes")
    reset_parser.add_argument("--person-id", type=str, default=None,
                              help="Process a single ChMeetings person ID instead of the whole group (for testing)")
    reset_parser.add_argument("--probe", action="store_true",
                              help="Diagnostic: test what the PUT endpoint accepts for a single person (requires --person-id)")

    # Solve-schedule command
    solve_schedule_parser = subparsers.add_parser(
        "solve-schedule",
        help="Run CP-SAT scheduler: reads schedule_input.json, writes schedule_output.json",
    )
    solve_schedule_parser.add_argument(
        "--input",
        default=None,
        help="Path to schedule_input.json (default: DATA_DIR/schedule_input.json)",
    )
    solve_schedule_parser.add_argument(
        "--output",
        default=None,
        help="Path for schedule_output.json (default: DATA_DIR/schedule_output.json)",
    )

    # Produce-schedule command
    produce_schedule_parser = subparsers.add_parser(
        "produce-schedule",
        help="Render schedule_output.json as a human-readable Excel timetable",
    )
    produce_schedule_parser.add_argument(
        "--input",
        default=None,
        dest="schedule_output",
        help="Path to schedule_output.json (default: DATA_DIR/schedule_output.json)",
    )
    produce_schedule_parser.add_argument(
        "--constraint",
        default=None,
        dest="schedule_input",
        help="Path to schedule_input.json (default: DATA_DIR/schedule_input.json)",
    )
    produce_schedule_parser.add_argument(
        "--output",
        default=None,
        help="Output path for xlsx (default: EXPORT_DIR/VAYSF_Schedule_YYYY-MM-DD.xlsx)",
    )

    # Generate-venue-template command
    venue_template_parser = subparsers.add_parser(
        "generate-venue-template",
        help="Create (or regenerate) the blank venue input template for Pod-Resource-Estimate",
    )
    venue_template_parser.add_argument(
        "--output",
        default=None,
        help="Output path for the template xlsx (default: data/SportsFest_2026_Venue_Input_Template.xlsx)",
    )

    # Check-consent command
    check_consent_parser = subparsers.add_parser(
        "check-consent",
        help="Match consent-form export rows to participants and auto-check the consent checklist box",
    )
    check_consent_parser.add_argument(
        "--file",
        required=True,
        help="Path to the consent-form export xlsx file",
    )
    check_consent_parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview matches and write the audit xlsx without updating ChMeetings",
    )
    check_consent_parser.add_argument(
        "--church-code",
        help="Limit the run to a single church code such as RPC",
    )

    return parser.parse_args()

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def run_sync(manager: SyncManager, sync_type: str = "full", chm_id: Optional[str] = None,
             excel_fallback: bool = False) -> bool:
    """Run synchronization process with retry logic.

    Args:
        manager: SyncManager instance to use.
        sync_type: Type of sync to perform (churches, participants, approvals, validation, full).
        chm_id: Optional ChMeetings ID of a single participant to sync.
        excel_fallback: If True, use Excel export for approval sync instead of API.

    Returns:
        bool: True if successful, False otherwise.
    """
    logger.info(f"Running {sync_type} synchronization")
    if chm_id and sync_type == "participants":
        logger.info(f"Targeting single participant with ChMeetings ID: {chm_id}")

    try:
        if not manager.authenticate():
            logger.error("Authentication failed")
            return False
        if sync_type == "churches":
            excel_path = os.path.join("data", "Church Application Form.xlsx")
            if not os.path.exists(excel_path):
                logger.error(f"Excel file not found at {excel_path}")
                return False
            return manager.sync_churches_from_excel(excel_path)
        elif sync_type == "participants":
            # Pass chm_id to the manager's sync_participants method
            return manager.sync_participants(chm_id=chm_id)
        elif sync_type == "approvals":
            success1 = manager.generate_approvals(chm_id_to_target=chm_id)
            success2 = manager.sync_approvals_to_chmeetings(
                chm_id_to_target=chm_id,
                use_excel_fallback=excel_fallback,
            )
            return success1 and success2
        elif sync_type == "validation":
            return manager.validate_data()
        elif sync_type == "full":
            # For a 'full' sync, we generally don't sync a single participant by ID,
            # so chm_id is typically None here. If a chm_id were passed for 'full',
            # the current SyncManager.run_full_sync doesn't use it.
            # If you need 'full' sync to also be capable of targeting a specific CHM ID
            # for its participant sync portion, SyncManager.run_full_sync would need adjustment.
            # For now, assuming chm_id is primarily for direct 'participants' sync type.
            if chm_id:
                logger.warning("Warning: --chm-id is provided with --type=full. The participant sync portion of the full sync will currently run for all, not the specific ID.")
            stats = manager.run_full_sync() # run_full_sync internally calls manager.sync_participants without an ID.
            logger.info(f"Full sync completed with stats: {stats}")
            return True
        else:
            logger.error(f"Invalid sync type: {sync_type}")
            return False
    except Exception as e:
        logger.exception(f"Sync failed: {e}")
        return False
# END --- Modified run_sync function in main.py ---

def validate_config() -> bool:
    """Validate system configuration."""
    logger.info("Validating configuration")
    valid = Config.validate()
    if valid:
        logger.info("Configuration is valid")
    else:
        logger.error("Configuration validation failed")
    return valid

def run_scheduled_sync(interval: int, daemon: bool = False) -> None:
    """Run sync jobs at scheduled intervals."""
    logger.info(f"Starting scheduled sync with {interval}-minute intervals")
    with SyncManager() as manager:
        schedule.every(interval).minutes.do(run_sync, manager=manager, sync_type="full")
        if daemon:
            logger.info("Running as daemon process")
            run_sync(manager, "full")  # Initial run
            while True:
                try:
                    schedule.run_pending()
                    time.sleep(min(60, interval * 60))
                except KeyboardInterrupt:
                    logger.info("Scheduled sync interrupted by user")
                    break
                except Exception as e:
                    logger.exception(f"Scheduled sync error: {e}")
        else:
            logger.info("Running scheduled tasks once")
            schedule.run_all()

def test_connectivity(system: str = "all", test_type: str = "connectivity", test_email: str = "PastorBumble@gmail.com") -> bool:
    """Test connectivity and functionality to ChMeetings and/or WordPress."""
    logger.info(f"Testing {test_type} for {system}")
    success = True

    if system in ["chmeetings", "all"]:
        from chmeetings.backend_connector import ChMeetingsConnector
        logger.info("Testing ChMeetings functionality")
        try:
            with ChMeetingsConnector() as connector:
                if test_type in ["connectivity", "all"]:
                    if connector.authenticate():
                        logger.info("Successfully connected to ChMeetings")
                    else:
                        logger.error("Failed to connect to ChMeetings")
                        success = False

                if test_type == "api-inspect":
                    import json
                    if not connector.authenticate():
                        logger.error("Cannot inspect API - authentication failed")
                        success = False
                    else:
                        # 1. Dump field definitions
                        logger.info("=" * 60)
                        logger.info("FIELD DEFINITIONS (GET /api/v1/people/fields)")
                        logger.info("=" * 60)
                        fields = connector.get_fields()
                        if fields:
                            sections = fields.get("sections", []) if isinstance(fields, dict) else fields
                            for section in (sections if isinstance(sections, list) else []):
                                logger.info(f"  Section: {section.get('title', '(untitled)')} (id={section.get('section_id')})")
                                for field in section.get("fields", []):
                                    opts = field.get("options", [])
                                    opts_str = f" options={[(o.get('id'), o.get('name')) for o in opts]}" if opts else ""
                                    logger.info(f"    field_id={field.get('field_id')} | name={field.get('field_name')!r} | type={field.get('field_type')}{opts_str}")
                            # 1b. Cross-reference CHM_FIELDS constants against live API
                            from config import CHM_FIELDS
                            all_api_field_names = set()
                            for section in (sections if isinstance(sections, list) else []):
                                for field in section.get("fields", []):
                                    fname = field.get("field_name")
                                    if fname:
                                        all_api_field_names.add(fname)

                            logger.info("=" * 60)
                            logger.info("FIELD MAPPING VALIDATION (CHM_FIELDS vs live API)")
                            logger.info("=" * 60)
                            all_matched = True
                            for key, expected_name in CHM_FIELDS.items():
                                if expected_name in all_api_field_names:
                                    logger.info(f"  OK  CHM_FIELDS[{key!r}] = {expected_name!r}")
                                else:
                                    logger.warning(f"  MISSING  CHM_FIELDS[{key!r}] = {expected_name!r} - NOT found in API fields!")
                                    all_matched = False
                            if all_matched:
                                logger.info("  All CHM_FIELDS matched successfully.")
                            else:
                                logger.warning("  Some CHM_FIELDS did not match. Update config.py CHM_FIELDS if field names changed.")
                        else:
                            logger.warning("No field definitions returned")

                        # 2. List groups
                        logger.info("=" * 60)
                        logger.info("GROUPS (GET /api/v1/groups)")
                        logger.info("=" * 60)
                        groups = connector.get_groups()
                        for g in groups:
                            logger.info(f"  id={g.get('id')} | name={g.get('name')}")
                        logger.info(f"  Total: {len(groups)} groups")

                        # 3. Fetch 2 sample people with additional_fields (single page, no full pagination)
                        logger.info("=" * 60)
                        logger.info("SAMPLE PEOPLE (GET /api/v1/people?include_additional_fields=true&page_size=2&page=1)")
                        logger.info("=" * 60)
                        _resp = connector._api_request("GET", "api/v1/people", params={
                            "include_additional_fields": True,
                            "include_family_members": False,
                            "include_organizations": False,
                            "page_size": 2, "page": 1,
                        })
                        _resp.raise_for_status()
                        sample_people = connector._extract_data(_resp.json())
                        for person in sample_people[:2]:
                            logger.info(f"--- Person: {person.get('first_name', '?')} {person.get('last_name', '?')} (id={person.get('id', person.get('person_id', '?'))}) ---")
                            logger.info(f"  Keys: {list(person.keys())}")
                            af = person.get("additional_fields", [])
                            if af:
                                logger.info(f"  additional_fields ({len(af)} fields):")
                                for f in af[:5]:
                                    logger.info(f"    {json.dumps(f)}")
                                if len(af) > 5:
                                    logger.info(f"    ... and {len(af) - 5} more fields")
                            else:
                                logger.info("  additional_fields: (empty or not present)")

                        # 4. Fetch one person by ID if we have any
                        if sample_people:
                            pid = sample_people[0].get("id", sample_people[0].get("person_id"))
                            if pid:
                                logger.info("=" * 60)
                                logger.info(f"SINGLE PERSON (GET /api/v1/people/{pid})")
                                logger.info("=" * 60)
                                single = connector.get_person(str(pid))
                                if single:
                                    logger.info(f"  Keys: {list(single.keys())}")
                                    af = single.get("additional_fields", [])
                                    logger.info(f"  additional_fields: {len(af) if isinstance(af, list) else 'N/A'} fields")
                                    if af and isinstance(af, list):
                                        for f in af[:3]:
                                            logger.info(f"    {json.dumps(f)}")
                                else:
                                    logger.warning(f"  Could not fetch person {pid}")
        except Exception as e:
            logger.error(f"ChMeetings test failed: {e}")
            success = False

    if system in ["wordpress", "all"]:
        from wordpress.frontend_connector import WordPressConnector
        logger.info("Testing WordPress functionality")
        try:
            with WordPressConnector() as connector:
                if test_type == "connectivity":
                    churches = connector.get_churches()
                    if churches is not None:
                        logger.info(f"Successfully connected to WordPress: Retrieved {len(churches)} churches")
                    else:
                        logger.error("Failed to connect to WordPress")
                        success = False
                elif test_type == "churches":
                    test_church = {
                        "church_name": "Test Church API",
                        "church_code": "API",
                        "pastor_name": "Pastor Test",
                        "pastor_email": "pastor@testchurch.org",
                        "pastor_phone": "555-123-4567",
                        "church_rep_name": "Rep Test",
                        "church_rep_email": "rep@testchurch.org",
                        "church_rep_phone": "555-987-6543",
                        "sports_ministry_level": 2
                    }
                    new_church = connector.create_church(test_church)
                    if new_church:
                        logger.info(f"Created church: {new_church.get('church_id')}")
                    else:
                        logger.error("Failed to create church")
                        success = False
                    church_by_code = connector.get_church_by_code("API")
                    if church_by_code:
                        logger.info(f"Retrieved church by code: {church_by_code['church_name']}")
                    else:
                        logger.error("Failed to get church by code")
                        success = False
                    update_data = {"church_name": "Updated Test Church API", "sports_ministry_level": 3}
                    updated_church = connector.update_church_by_code("API", update_data)
                    if updated_church:
                        logger.info(f"Updated church: {updated_church['church_name']}")
                    else:
                        logger.error("Failed to update church by code")
                        success = False
                elif test_type == "email":
                    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    email_data = {
                        "to": test_email,
                        "subject": "Test Email from Sports Fest API",
                        "message": f"""
                        <h2>Sports Fest Email Test</h2>
                        <p>This is a test email sent via the middleware.</p>
                        <p>Time of test: {current_time}</p>
                        """,
                        "from_email": "SportsFest Staff <info@sportsfest.vayhub.us>"
                    }
                    result = connector.send_email(**email_data)
                    if result.get("success", False):
                        logger.info(f"Email sent successfully to {test_email}")
                    else:
                        logger.error(f"Failed to send email: {result.get('message', 'Unknown error')}")
                        success = False
                elif test_type == "all":
                    churches = connector.get_churches()
                    if churches is not None:
                        logger.info(f"Successfully connected to WordPress: Retrieved {len(churches)} churches")
                    else:
                        logger.error("Failed to connect to WordPress")
                        success = False
                    test_church = {
                        "church_name": "Test Church API",
                        "church_code": "API",
                        "pastor_name": "Pastor Test",
                        "pastor_email": "pastor@testchurch.org",
                        "pastor_phone": "555-123-4567",
                        "church_rep_name": "Rep Test",
                        "church_rep_email": "rep@testchurch.org",
                        "church_rep_phone": "555-987-6543",
                        "sports_ministry_level": 2
                    }
                    new_church = connector.create_church(test_church)
                    if new_church:
                        logger.info(f"Created church: {new_church.get('church_id')}")
                    else:
                        logger.error("Failed to create church")
                        success = False
                    church_by_code = connector.get_church_by_code("API")
                    if church_by_code:
                        logger.info(f"Retrieved church by code: {church_by_code['church_name']}")
                    else:
                        logger.error("Failed to get church by code")
                        success = False
                    update_data = {"church_name": "Updated Test Church API", "sports_ministry_level": 3}
                    updated_church = connector.update_church_by_code("API", update_data)
                    if updated_church:
                        logger.info(f"Updated church: {updated_church['church_name']}")
                    else:
                        logger.error("Failed to update church by code")
                        success = False
                    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    email_data = {
                        "to": test_email,
                        "subject": "Test Email from Sports Fest API",
                        "message": f"""
                        <h2>Sports Fest Email Test</h2>
                        <p>This is a test email sent via the middleware.</p>
                        <p>Time of test: {current_time}</p>
                        """,
                        "from_email": "SportsFest Staff <info@sportsfest.vayhub.us>"
                    }
                    result = connector.send_email(**email_data)
                    if result.get("success", False):
                        logger.info(f"Email sent successfully to {test_email}")
                    else:
                        logger.error(f"Failed to send email: {result.get('message', 'Unknown error')}")
                        success = False
        except Exception as e:
            logger.error(f"WordPress test failed: {e}")
            success = False

    return success


def generate_venue_template(output_path: Optional[Path] = None) -> bool:
    """Create (or regenerate) the blank venue input template xlsx.

    Writes a Venue-Input sheet with column headers and example rows pre-filled
    with the Available Slots formula so staff can adjust times and see the cell
    update automatically.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from config import (
        DATA_DIR, VENUE_TEMPLATE_FILENAME,
        POD_RESOURCE_TYPE_TENNIS, POD_RESOURCE_TYPE_PICKLEBALL,
        POD_RESOURCE_TYPE_TABLE_TENNIS, POD_RESOURCE_TYPE_BADMINTON,
        SCHEDULE_SKETCH_COLOR_HEADER,
    )

    if output_path is None:
        output_path = DATA_DIR / VENUE_TEMPLATE_FILENAME

    wb = Workbook()
    ws = wb.active
    ws.title = "Venue-Input"

    headers = [
        "Pod Name", "Venue Name", "Resource Type", "Quantity",
        "Date", "Start Time", "Last Start Time", "Slot Minutes",
        "Available Slots", "Contact", "Cost", "Notes",
    ]

    header_fill = PatternFill("solid", fgColor=SCHEDULE_SKETCH_COLOR_HEADER)
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Example rows: (pod, venue, resource_type, qty, date, start_hr, last_hr, slot_min, contact, cost, notes)
    examples = [
        ("North OC Tennis Pod",  "City Park",        POD_RESOURCE_TYPE_TENNIS,       4, "2026-07-19", 13, 18, 60,  "TBD", "TBD", "Staff verified usable"),
        ("Pickleball Pod A",     "Community Center", POD_RESOURCE_TYPE_PICKLEBALL,   6, "2026-07-19", 13, 18, 45,  "TBD", "TBD", "Includes 35+"),
        ("Indoor Table Pod",     "Church Hall",      POD_RESOURCE_TYPE_TABLE_TENNIS, 6, "2026-07-20", 18, 21, 30,  "TBD", "TBD", "Includes 35+"),
        ("Badminton Pod",        "School Gym",       POD_RESOURCE_TYPE_BADMINTON,    4, "2026-07-20", 18, 21, 45,  "TBD", "TBD", "Staff verified usable"),
    ]

    for row_idx, ex in enumerate(examples, start=2):
        pod, venue, rtype, qty, date, start_hr, last_hr, slot_min, contact, cost, notes = ex
        row_vals = [pod, venue, rtype, qty, date, start_hr, last_hr, slot_min, None, contact, cost, notes]
        for col_idx, val in enumerate(row_vals, start=1):
            if col_idx == 9:  # Available Slots — formula; col letters D=4 F=6 G=7 H=8
                # Formula: Qty * ((LastStart - Start) * 60 / SlotMin + 1)
                # where Start/LastStart are stored as decimal hours (integers like 13, 18)
                cell = ws.cell(
                    row=row_idx, column=col_idx,
                    value=f"=D{row_idx}*(((G{row_idx}-F{row_idx})*60/H{row_idx})+1)",
                )
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Column widths
    col_widths = [22, 22, 22, 10, 12, 12, 16, 14, 16, 16, 10, 28]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Help note below examples
    note_row = len(examples) + 3
    ws.cell(
        row=note_row, column=1,
        value=(
            "Available Slots formula: =D*(((G-F)*60/H)+1) "
            "where D=Quantity, F=Start Time (decimal hour), G=Last Start Time (decimal hour), H=Slot Minutes. "
            "Add one row per pod per date. Staff-entered resources are assumed valid."
        ),
    )

    try:
        wb.save(output_path)
        logger.info(f"Venue input template written to: {output_path}")
        return True
    except OSError as e:
        logger.error(f"Failed to write venue template to {output_path}: {e}")
        return False


def _log_json_block(title: str, payload) -> None:
    """Log structured JSON payloads in a readable block."""
    logger.info("=" * 60)
    logger.info(title)
    logger.info("=" * 60)
    logger.info(json.dumps(payload, indent=2, ensure_ascii=False, default=str))


def inspect_person(chm_id: str) -> bool:
    """Inspect one ChMeetings person ID and related WordPress records.

    This is intentionally read-only and operator-friendly. If the ChMeetings
    record is gone, it still looks for any matching WordPress participant so we
    can identify who the stale ID used to belong to.
    """
    logger.info(f"Inspecting ChMeetings person ID: {chm_id}")
    found_any = False

    with ChMeetingsConnector() as chm_connector:
        if not chm_connector.authenticate():
            logger.error("Failed to authenticate with ChMeetings for inspect-person")
        else:
            person = chm_connector.get_person(chm_id)
            if person:
                found_any = True
                _log_json_block(f"ChMeetings person {chm_id}", person)
            elif chm_connector.last_get_person_status == "not_found":
                logger.warning(f"ChMeetings person {chm_id} returned 404 Not Found.")
            else:
                logger.warning(f"Could not retrieve ChMeetings person {chm_id}. Check logs above for details.")

    with WordPressConnector() as wp_connector:
        participants = wp_connector.get_participants(
            params={"chmeetings_id": chm_id, "per_page": 100}
        )

        if not participants:
            logger.info(f"No WordPress participants found with chmeetings_id={chm_id}.")
            return found_any

        found_any = True
        _log_json_block(f"WordPress participants matching chmeetings_id={chm_id}", participants)

        for participant in participants:
            wp_participant_id = participant.get("participant_id")
            if not wp_participant_id:
                continue

            rosters = wp_connector.get_rosters(params={"participant_id": wp_participant_id})
            approvals = wp_connector.get_approvals(params={"participant_id": wp_participant_id})
            validation_issues = wp_connector.get_validation_issues(params={"participant_id": wp_participant_id})

            _log_json_block(f"WordPress rosters for participant_id={wp_participant_id}", rosters)
            _log_json_block(f"WordPress approvals for participant_id={wp_participant_id}", approvals)
            _log_json_block(
                f"WordPress validation issues for participant_id={wp_participant_id}",
                validation_issues,
            )

    return found_any
    
def main() -> None:
    """Main entry point for the VAYSF middleware."""
    args = parse_args()
    logger.info(f"Executing command: {args.command}")

# START --- Modified main() function's sync block in main.py ---
    if args.command == "sync":
        # Retrieve chm_id from args. It will be None if not provided.
        participant_chm_id = args.chm_id if hasattr(args, 'chm_id') else None

        # Optional: Add a check or warning if --chm-id is used with a sync type other than 'participants'  or 'approvals'
        if participant_chm_id and args.type not in ["participants", "approvals"]: # Ensure "approvals" is in this list
            logger.warning(f"--chm-id '{participant_chm_id}' was provided with sync type '{args.type}'. "
                           "The --chm-id argument is only used when --type is 'participants' or 'approvals'. "
                           "The specified ID will be ignored for this operation.")
            # Reset to None if not applicable to ensure run_sync behaves as expected for other types or let run_sync handle the warning as implemented above.
            # For clarity here, if it's not for 'participants', it shouldn't be passed as a specific ID.
            # However, run_sync already has a conditional warning for 'full' type.
            # Let's pass it and let run_sync decide.

        excel_fallback = args.excel_fallback if hasattr(args, 'excel_fallback') else False
        manager = SyncManager()
        with manager:
            # Pass the participant_chm_id and excel_fallback to run_sync
            success = run_sync(manager, args.type, chm_id=participant_chm_id, excel_fallback=excel_fallback)
# END --- Modified main() function's sync block in main.py ---
    elif args.command == "sync-churches":
        if not os.path.exists(args.file):
            logger.error(f"Excel file not found at {args.file}")
            success = False
        else:
            with SyncManager() as manager:
                success = manager.sync_churches_from_excel(args.file)  # Fixed
    elif args.command == "assign-groups":
        from group_assignment import assign_people_to_church_team_groups
        dry_run = getattr(args, "dry_run", False)
        success = assign_people_to_church_team_groups(
            dry_run=dry_run,
            source_file=getattr(args, "file", None),
        )
        if success:
            if dry_run:
                logger.info("Dry-run complete. Check data/church_team_assignments.xlsx for the preview.")
            else:
                logger.info("Group assignment complete. Check data/church_team_assignments.xlsx for the audit log.")
    elif args.command == "clear-team-groups":
        from group_assignment import clear_team_groups
        success = clear_team_groups(
            church_code=args.church_code,
            dry_run=args.dry_run,
            execute=args.execute,
        )
        if success:
            if args.dry_run:
                logger.info("Dry-run complete. Check data/team_group_clearing_audit.xlsx for the preview.")
            else:
                logger.info("Team-group clearing complete. Check data/team_group_clearing_audit.xlsx for the audit log.")
    elif args.command == "audit-team-groups":
        from group_assignment import audit_team_groups
        success = audit_team_groups(church_code=args.church_code,
                                    remove_orphans=args.remove_orphans)
        if success:
            logger.info("Team-group audit complete. Check data/team_group_orphan_audit.xlsx for the audit log.")
    elif args.command == "export-church-teams":
        output_path = Path(args.output) 
        try:
            output_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"Report output directory set to: {output_path.resolve()}")
        except OSError as e:
            logger.error(f"Failed to create report output directory {output_path}: {e}")
            success = False
        else:
            try:
                # ChurchTeamsExporter is a context manager
                with ChurchTeamsExporter() as exporter: 
                    success = exporter.generate_reports( 
                        target_church_code=args.church_code,
                        output_dir=output_path,
                        force_resend_pending=args.force_resend_pending,
                        force_resend_validated1=args.force_resend_validated1,
                        force_resend_validated2=args.force_resend_validated2,
                        dry_run=args.dry_run,
                        target_resend_chm_id=args.chm_id,
                    )
                if success: 
                    logger.info(f"Church team reports generated successfully in {output_path.resolve()}.")
                else:
                    logger.error("Failed to generate church team reports (exporter returned False).")
            except Exception as e:
                logger.error(f"An exception occurred during report export: {e}", exc_info=True)
                success = False
    elif args.command == "solve-schedule":
        from scheduler import run_solve_schedule
        input_path = Path(args.input) if args.input else DATA_DIR / "schedule_input.json"
        output_path = Path(args.output) if args.output else DATA_DIR / "schedule_output.json"
        exit_code = run_solve_schedule(input_path, output_path)
        sys.exit(exit_code)
    elif args.command == "produce-schedule":
        from schedule_workbook import ScheduleWorkbookBuilder
        so_path = Path(args.schedule_output) if args.schedule_output else DATA_DIR / "schedule_output.json"
        si_path = Path(args.schedule_input)  if args.schedule_input  else DATA_DIR / "schedule_input.json"
        if args.output:
            out_path = Path(args.output)
        else:
            today = datetime.date.today().strftime("%Y-%m-%d")
            out_path = Path(EXPORT_DIR) / f"VAYSF_Schedule_{today}.xlsx"
        try:
            so_data = json.loads(so_path.read_text(encoding="utf-8"))
            si_data = json.loads(si_path.read_text(encoding="utf-8"))
        except FileNotFoundError as exc:
            logger.error(f"export-schedule: required file not found — {exc.filename}")
            success = False
        else:
            out_path.parent.mkdir(parents=True, exist_ok=True)
            ScheduleWorkbookBuilder.write_schedule_output_workbook(
                out_path, so_data, si_data
            )
            logger.info(f"Schedule Excel written to: {out_path.resolve()}")
            success = True
    elif args.command == "generate-venue-template":
        out = Path(args.output) if args.output else None
        success = generate_venue_template(out)
        if success:
            logger.info("Venue input template created. Copy it to data/venue_input.xlsx, fill in your pod details, then re-run export-church-teams.")
    elif args.command == "config":
        success = validate_config()
    elif args.command == "schedule":
        run_scheduled_sync(args.interval, args.daemon)
        success = True  # No exit until interrupted
    elif args.command == "test":
        success = test_connectivity(args.system, args.test_type, args.test_email)
    elif args.command == "inspect-person":
        success = inspect_person(args.chm_id)
    elif args.command == "reset-season":
        if args.archive_only and args.reset_only:
            logger.error("--archive-only and --reset-only are mutually exclusive.")
            success = False
        elif args.probe:
            if not args.person_id:
                logger.error("--probe requires --person-id.")
                success = False
            else:
                with ChMeetingsConnector() as chm_conn, WordPressConnector() as wp_conn:
                    resetter = SeasonResetter(chm_conn, wp_conn)
                    success = resetter.probe_put_endpoint(args.person_id)
        else:
            with ChMeetingsConnector() as chm_conn, WordPressConnector() as wp_conn:
                resetter = SeasonResetter(chm_conn, wp_conn)
                success = resetter.run(
                    args.year,
                    dry_run=args.dry_run,
                    archive_only=args.archive_only,
                    reset_only=args.reset_only,
                    person_id=args.person_id,
                )
    elif args.command == "check-consent":
        if not os.path.exists(args.file):
            logger.error(f"Consent export file not found at {args.file}")
            success = False
        else:
            from sync.consent_checker import ConsentChecker

            with ChMeetingsConnector() as chm_conn, WordPressConnector() as wp_conn:
                checker = ConsentChecker(chm_conn, wp_conn)
                summary = checker.run(
                    args.file,
                    dry_run=args.dry_run,
                    church_code=args.church_code,
                )
                success = summary["api_error"] == 0
    else:
        logger.error(f"Unknown command: {args.command}")
        success = False

    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
