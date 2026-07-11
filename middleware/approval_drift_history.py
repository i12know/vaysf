"""Build audit workbooks and operator actions for approval drift.

The history command is read-only. The acceptance command is an explicit
operator action for final-week cases where a church confirms sport/event
changes were legitimate and the prior approval should be restored.
"""

from __future__ import annotations

import datetime as dt
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


STATUS_WORKBOOK_GLOB = "Church_Team_Status_ALL_*.xlsx"
LOG_GLOB = "sportsfest_*.log"
APPROVED_STATUS = "approved"
DEFAULT_REAPPROVAL_STATUS = "reapproval_required"
REAPPROVAL_REASON_ISSUE_TYPES = (
    "approval_identity_drift",
    "approval_registration_drift",
    "reapproval_required_reason_missing",
)
OPEN_STATUS = "open"
RESOLVED_STATUS = "resolved"

_HARD_DRIFT_RE = re.compile(
    r"^(?P<detected_at>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})"
    r".*APPROVAL IDENTITY DRIFT for chm_id=(?P<chm_id>\d+) "
    r"\(WP participant_id=(?P<wp_participant_id>\d+)\): "
    r"(?P<summary>.*)\. Prior '(?P<prior_status>[^']*)' invalidated"
)

_SOFT_BIRTHDATE_RE = re.compile(
    r"^(?P<detected_at>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})"
    r".*Birthdate correction for chm_id=(?P<chm_id>\d+): "
    r"'(?P<old_value>[^']*)'.*'(?P<new_value>[^']*)'"
)

_DELTA_RE = re.compile(
    r"^(?P<field>[^:]+): '(?P<old_value>.*?)'.*?'(?P<new_value>.*?)'$"
)


@dataclass(frozen=True)
class CurrentParticipant:
    church: str
    chmeetings_id: str
    wp_participant_id: str
    name: str
    is_member: str
    approval_status: str
    sports_registered: str
    latest_chm_update: str
    first_open_error: str


@dataclass(frozen=True)
class DriftEvent:
    detected_at: str
    chmeetings_id: str
    wp_participant_id: str
    source_log: str
    event_type: str
    prior_status: str
    raw_summary: str
    field: str
    old_value: str
    new_value: str


def find_latest_status_workbook(*roots: Path) -> Optional[Path]:
    """Return the newest Church_Team_Status_ALL workbook from the given roots."""
    candidates: list[Path] = []
    for root in roots:
        if root and root.exists():
            candidates.extend(root.glob(STATUS_WORKBOOK_GLOB))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_current_participants(
    workbook_path: Path,
    *,
    status: str = "reapproval_required",
    church_code: Optional[str] = None,
) -> dict[str, CurrentParticipant]:
    """Load current participants matching the requested approval status."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if "Contacts-Status" not in wb.sheetnames:
        raise ValueError(f"{workbook_path} does not contain a Contacts-Status tab")

    ws = wb["Contacts-Status"]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    idx = {header: i for i, header in enumerate(headers) if header}

    required = [
        "Church Team",
        "ChMeetings ID",
        "First Name",
        "Last Name",
        "Participant ID (WP)",
        "Approval_Status (WP)",
        "Sports Registered",
    ]
    missing = [name for name in required if name not in idx]
    if missing:
        raise ValueError(
            f"{workbook_path} Contacts-Status is missing required columns: {', '.join(missing)}"
        )

    participants: dict[str, CurrentParticipant] = {}
    for row in rows:
        row_status = str(row[idx["Approval_Status (WP)"]] or "").strip()
        if status and row_status != status:
            continue

        church = str(row[idx["Church Team"]] or "").strip()
        if church_code and church.upper() != church_code.upper():
            continue

        chmeetings_id = str(row[idx["ChMeetings ID"]] or "").strip()
        if not chmeetings_id:
            continue

        first_name = str(row[idx["First Name"]] or "").strip()
        last_name = str(row[idx["Last Name"]] or "").strip()
        is_member = _membership_answer(row, idx)
        participants[chmeetings_id] = CurrentParticipant(
            church=church,
            chmeetings_id=chmeetings_id,
            wp_participant_id=str(row[idx["Participant ID (WP)"]] or "").strip(),
            name=f"{first_name} {last_name}".strip(),
            is_member=is_member,
            approval_status=row_status,
            sports_registered=str(row[idx["Sports Registered"]] or "").strip(),
            latest_chm_update=str(row[idx.get("Update_on_ChM", -1)] or "").strip()
            if "Update_on_ChM" in idx
            else "",
            first_open_error=str(row[idx.get("First_Open_ERROR_Desc (WP)", -1)] or "").strip()
            if "First_Open_ERROR_Desc (WP)" in idx
            else "",
        )
    return participants


def _membership_answer(row: tuple[object, ...], idx: dict[str, int]) -> str:
    """Return the ChMeetings membership-question answer from Contacts-Status."""
    for column_name in (
        "Is_Member_ChM",
        "Is Member",
        "Is Church Member",
        "Would the team's Senior Pastor say that you belong to his church?",
    ):
        if column_name in idx:
            value = row[idx[column_name]]
            if isinstance(value, bool):
                return "Yes" if value else "No"
            text = str(value or "").strip()
            if text in {"1", "TRUE", "True", "true"}:
                return "Yes"
            if text in {"0", "FALSE", "False", "false"}:
                return "No"
            return text
    return ""


def _split_hard_drift_summary(summary: str) -> list[tuple[str, str, str]]:
    """Split a log drift summary into individual field old/new deltas."""
    deltas: list[tuple[str, str, str]] = []
    for part in summary.split("; "):
        match = _DELTA_RE.match(part.strip())
        if match:
            deltas.append(
                (
                    match.group("field").strip(),
                    match.group("old_value"),
                    match.group("new_value"),
                )
            )
        else:
            deltas.append((part.strip(), "", ""))
    return deltas


def parse_drift_events(log_paths: Iterable[Path]) -> list[DriftEvent]:
    """Parse approval drift and soft birthdate correction entries from logs."""
    events: list[DriftEvent] = []
    for log_path in sorted(log_paths):
        if not log_path.exists() or not log_path.is_file():
            continue
        for line in log_path.read_text(encoding="utf-8", errors="replace").splitlines():
            hard = _HARD_DRIFT_RE.search(line)
            if hard:
                summary = hard.group("summary")
                for field, old_value, new_value in _split_hard_drift_summary(summary):
                    events.append(
                        DriftEvent(
                            detected_at=hard.group("detected_at"),
                            chmeetings_id=hard.group("chm_id"),
                            wp_participant_id=hard.group("wp_participant_id"),
                            source_log=log_path.name,
                            event_type="approval_identity_drift",
                            prior_status=hard.group("prior_status"),
                            raw_summary=summary,
                            field=field,
                            old_value=old_value,
                            new_value=new_value,
                        )
                    )
                continue

            soft = _SOFT_BIRTHDATE_RE.search(line)
            if soft:
                events.append(
                    DriftEvent(
                        detected_at=soft.group("detected_at"),
                        chmeetings_id=soft.group("chm_id"),
                        wp_participant_id="",
                        source_log=log_path.name,
                        event_type="approval_birthdate_correction",
                        prior_status="approval_preserved",
                        raw_summary="Birthdate correction with age unchanged",
                        field="Birthdate (age unchanged)",
                        old_value=soft.group("old_value"),
                        new_value=soft.group("new_value"),
                    )
                )
    return events


def build_history_rows(
    current: dict[str, CurrentParticipant],
    events: Iterable[DriftEvent],
) -> list[dict[str, str]]:
    """Return audit rows for current participants and their matching log events."""
    events_by_chm: dict[str, list[DriftEvent]] = {}
    for event in events:
        if event.chmeetings_id in current:
            events_by_chm.setdefault(event.chmeetings_id, []).append(event)

    rows: list[dict[str, str]] = []
    for participant in sorted(current.values(), key=lambda p: (p.church, p.name)):
        participant_events = sorted(
            events_by_chm.get(participant.chmeetings_id, []),
            key=lambda e: (e.detected_at, e.source_log, e.field),
        )
        if participant_events:
            for grouped_events in _collapse_repeated_events(participant_events):
                rows.append(_row(participant, grouped_events))
        else:
            rows.append(
                _row(
                    participant,
                    [
                        DriftEvent(
                            detected_at="",
                            chmeetings_id=participant.chmeetings_id,
                            wp_participant_id=participant.wp_participant_id,
                            source_log="",
                            event_type="no_local_drift_log_found",
                            prior_status="",
                            raw_summary=(
                                "No APPROVAL IDENTITY DRIFT or birthdate correction log entry "
                                "was found for this current participant."
                            ),
                            field="",
                            old_value="",
                            new_value="",
                        )
                    ],
                )
            )
    return rows


def _collapse_repeated_events(events: list[DriftEvent]) -> list[list[DriftEvent]]:
    """Collapse repeated sync-log sightings of the same underlying change."""
    groups: dict[tuple[str, str, str, str, str], list[DriftEvent]] = {}
    order: list[tuple[str, str, str, str, str]] = []
    for event in events:
        key = (
            event.event_type,
            event.field,
            event.old_value,
            event.new_value,
            event.prior_status,
        )
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(event)
    return [groups[key] for key in order]


def _row(participant: CurrentParticipant, events: list[DriftEvent]) -> dict[str, str]:
    event = events[0]
    detected_values = sorted(e.detected_at for e in events if e.detected_at)
    source_logs = sorted({e.source_log for e in events if e.source_log})
    return {
        "Church": participant.church,
        "Name": participant.name,
        "Is Member": participant.is_member,
        "ChMeetings ID": participant.chmeetings_id,
        "WP Participant ID": participant.wp_participant_id or event.wp_participant_id,
        "Current Approval Status": participant.approval_status,
        "Sports Registered Now": participant.sports_registered,
        "Latest ChMeetings Update": participant.latest_chm_update,
        "First Open Error": participant.first_open_error,
        "First Detected At": detected_values[0] if detected_values else "",
        "Last Seen At": detected_values[-1] if detected_values else "",
        "Times Seen": str(len(events)) if event.event_type != "no_local_drift_log_found" else "0",
        "Source Logs": ", ".join(source_logs),
        "Event Type": event.event_type,
        "Prior Status": event.prior_status,
        "Changed Field": event.field,
        "Old Value": event.old_value,
        "New Value": event.new_value,
        "Raw Change Summary": event.raw_summary,
    }


def write_history_workbook(rows: list[dict[str, str]], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approval-Drift-History"

    headers = list(rows[0].keys()) if rows else [
        "Church",
        "Name",
        "Is Member",
        "ChMeetings ID",
        "WP Participant ID",
        "Current Approval Status",
        "Sports Registered Now",
        "Latest ChMeetings Update",
        "First Open Error",
        "First Detected At",
        "Last Seen At",
        "Times Seen",
        "Source Logs",
        "Event Type",
        "Prior Status",
        "Changed Field",
        "Old Value",
        "New Value",
        "Raw Change Summary",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 10,
        "B": 24,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 22,
        "G": 42,
        "H": 22,
        "I": 40,
        "J": 20,
        "K": 20,
        "L": 12,
        "M": 34,
        "N": 28,
        "O": 18,
        "P": 26,
        "Q": 38,
        "R": 38,
        "S": 80,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    participant_ids = {row["ChMeetings ID"] for row in rows}
    participant_with_history = {
        row["ChMeetings ID"]
        for row in rows
        if row["Event Type"] != "no_local_drift_log_found"
    }
    event_counts = Counter(row["Event Type"] for row in rows)
    summary_rows = [
        ("Participants in current status scope", len(participant_ids)),
        ("Participants with local drift history", len(participant_with_history)),
        ("Audit rows", len(rows)),
    ]
    summary_rows.extend((f"Rows: {event_type}", count) for event_type, count in sorted(event_counts.items()))
    for item in summary_rows:
        summary.append(list(item))
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 16

    wb.save(output_path)
    return output_path


def accept_reviewed_drift(
    *,
    wordpress_connector: Any,
    workbook_path: Path,
    output_path: Path,
    logs_dir: Optional[Path] = None,
    church_code: Optional[str] = None,
    chm_id: Optional[str] = None,
    status: str = DEFAULT_REAPPROVAL_STATUS,
    reason: Optional[str] = None,
    execute: bool = False,
    force_approved: bool = False,
) -> dict[str, object]:
    """Reset reviewed approval drift targets to the prior approval state.

    Exactly one target scope is required: ``church_code`` for a whole church team
    or ``chm_id`` for one participant. Targets are selected from the current
    Church_Team_Status_ALL workbook and must still be in ``status``. By default,
    the target status must be unambiguous in the local approval drift logs. Pass
    ``force_approved`` only when the operator intentionally wants to override
    the prior-state check and mark selected targets approved.
    """
    if bool(church_code) == bool(chm_id):
        raise ValueError("Pass exactly one of church_code or chm_id.")

    current = load_current_participants(
        workbook_path,
        status=status,
        church_code=church_code,
    )
    if chm_id:
        current = {chm_id: current[chm_id]} if chm_id in current else {}

    prior_statuses = _load_prior_statuses(logs_dir) if logs_dir else {}
    accepted_at = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    operator_reason = (reason or "Operator accepted reviewed approval drift.").strip()
    rows: list[dict[str, str]] = []

    for participant in sorted(current.values(), key=lambda p: (p.church, p.name)):
        target = _resolve_target_status(
            participant,
            prior_statuses=prior_statuses,
            force_approved=force_approved,
        )
        rows.append(
            _accept_one_participant(
                wordpress_connector=wordpress_connector,
                participant=participant,
                expected_status=status,
                target_status=target["target_status"],
                prior_status_candidates=target["prior_status_candidates"],
                prior_status_resolution=target["prior_status_resolution"],
                reason=operator_reason,
                accepted_at=accepted_at,
                execute=execute,
                force_approved=force_approved,
            )
        )

    written = write_acceptance_workbook(rows, output_path)
    actions = Counter(row["Action"] for row in rows)
    return {
        "workbook": str(workbook_path),
        "output": str(written),
        "execute": execute,
        "force_approved": force_approved,
        "targets": len(current),
        "accepted": actions.get("accepted", 0),
        "would_accept": actions.get("would_accept", 0),
        "skipped": sum(
            count
            for action, count in actions.items()
            if action.startswith("skipped_") or action.startswith("blocked_")
        ),
        "errors": actions.get("partial_error", 0) + actions.get("error", 0),
        "actions": dict(actions),
    }


def _load_prior_statuses(logs_dir: Optional[Path]) -> dict[str, list[str]]:
    if not logs_dir or not logs_dir.exists():
        return {}
    statuses: dict[str, list[str]] = {}
    for event in parse_drift_events(sorted(logs_dir.glob(LOG_GLOB))):
        prior = event.prior_status.strip()
        if not prior or prior == "approval_preserved":
            continue
        chm_statuses = statuses.setdefault(event.chmeetings_id, [])
        if prior not in chm_statuses:
            chm_statuses.append(prior)
    return statuses


def _resolve_target_status(
    participant: CurrentParticipant,
    *,
    prior_statuses: dict[str, list[str]],
    force_approved: bool,
) -> dict[str, str]:
    candidates = prior_statuses.get(participant.chmeetings_id, [])
    ignored_statuses = {participant.approval_status, DEFAULT_REAPPROVAL_STATUS}
    actionable_candidates = [
        status for status in candidates if status and status not in ignored_statuses
    ]
    if force_approved:
        return {
            "target_status": APPROVED_STATUS,
            "prior_status_candidates": ", ".join(candidates),
            "prior_status_resolution": "forced_approved",
        }
    if len(actionable_candidates) == 1:
        resolution = "single_prior_status"
        if len(candidates) > len(actionable_candidates):
            resolution = "single_prior_status_after_ignoring_reapproval_required"
        return {
            "target_status": actionable_candidates[0],
            "prior_status_candidates": ", ".join(candidates),
            "prior_status_resolution": resolution,
        }
    if not actionable_candidates:
        return {
            "target_status": "",
            "prior_status_candidates": ", ".join(candidates),
            "prior_status_resolution": "blocked_missing_prior_status",
        }
    return {
        "target_status": "",
        "prior_status_candidates": ", ".join(candidates),
        "prior_status_resolution": "blocked_ambiguous_prior_status",
    }


def _accept_one_participant(
    *,
    wordpress_connector: Any,
    participant: CurrentParticipant,
    expected_status: str,
    target_status: str,
    prior_status_candidates: str,
    prior_status_resolution: str,
    reason: str,
    accepted_at: str,
    execute: bool,
    force_approved: bool,
) -> dict[str, str]:
    row = _acceptance_base_row(
        participant,
        accepted_at,
        reason,
        execute,
        force_approved=force_approved,
        target_status=target_status,
        prior_status_candidates=prior_status_candidates,
        prior_status_resolution=prior_status_resolution,
    )
    if prior_status_resolution.startswith("blocked_"):
        row.update(
            {
                "Action": prior_status_resolution,
                "Message": (
                    "Could not safely determine exactly one prior approval state from local "
                    "approval drift logs. Pass --force-approved only if this participant "
                    "should be marked approved anyway."
                ),
            }
        )
        return row

    wp_participant = _find_wp_participant(wordpress_connector, participant)
    if not wp_participant:
        row.update(
            {
                "Action": "blocked_missing_wordpress_participant",
                "Message": "No WordPress participant was found for this ChMeetings ID.",
            }
        )
        return row

    wp_participant_id = str(wp_participant.get("participant_id") or "")
    wp_status = str(wp_participant.get("approval_status") or "").strip()
    row["WP Participant ID"] = wp_participant_id
    row["WP Current Approval Status"] = wp_status

    if wp_status != expected_status:
        row.update(
            {
                "Action": "skipped_current_status_changed",
                "Message": (
                    f"WordPress participant is '{wp_status}', not '{expected_status}'. "
                    "No approval reset needed."
                ),
            }
        )
        return row

    approvals = wordpress_connector.get_approvals(
        params={"participant_id": int(wp_participant_id)}
    )
    row["Approvals Found"] = str(len(approvals))
    if not approvals:
        row.update(
            {
                "Action": "blocked_missing_approval_record",
                "Message": "No approval record found; participant status was not changed.",
            }
        )
        return row

    drift_issues = []
    for issue_type in REAPPROVAL_REASON_ISSUE_TYPES:
        drift_issues.extend(
            wordpress_connector.get_validation_issues(
                params={
                    "participant_id": int(wp_participant_id),
                    "issue_type": issue_type,
                    "status": OPEN_STATUS,
                    "per_page": 100,
                }
            )
        )
    row["Open Drift Issues Found"] = str(len(drift_issues))

    if not execute:
        row.update(
            {
                "Action": "would_accept",
                "Message": (
                    f"Dry run only; participant, approval, and validation issue rows would be "
                    f"restored to '{target_status}'."
                ),
            }
        )
        return row

    participant_updated = wordpress_connector.update_participant(
        int(wp_participant_id),
        {"approval_status": target_status},
    )
    row["Participant Updated"] = "yes" if participant_updated else "no"

    approval_note = _approval_note(
        approvals[0].get("approval_notes"),
        accepted_at=accepted_at,
        reason=reason,
        target_status=target_status,
    )
    approval_successes = 0
    for approval in approvals:
        approval_id = int(approval["approval_id"])
        result = wordpress_connector.update_approval(
            approval_id,
            {
                "approval_status": target_status,
                "approval_notes": approval_note,
                "synced_to_chmeetings": False,
            },
        )
        if result:
            approval_successes += 1
    row["Approvals Updated"] = str(approval_successes)

    issue_successes = 0
    for issue in drift_issues:
        issue_id = int(issue["issue_id"])
        result = wordpress_connector.update_validation_issue(
            issue_id,
            {"status": RESOLVED_STATUS},
        )
        if result:
            issue_successes += 1
    row["Drift Issues Resolved"] = str(issue_successes)

    if participant_updated and approval_successes == len(approvals) and issue_successes == len(drift_issues):
        row.update(
            {
                "Action": "accepted",
                "Message": f"Reviewed drift accepted; WordPress approval restored to '{target_status}'.",
            }
        )
    else:
        row.update(
            {
                "Action": "partial_error",
                "Message": "One or more WordPress updates failed; review this row before continuing.",
            }
        )
    return row


def _find_wp_participant(
    wordpress_connector: Any,
    participant: CurrentParticipant,
) -> Optional[dict[str, Any]]:
    matches = wordpress_connector.get_participants(
        {
            "chmeetings_id": participant.chmeetings_id,
            "per_page": 100,
        }
    )
    if not matches:
        return None
    if participant.wp_participant_id:
        for match in matches:
            if str(match.get("participant_id") or "") == participant.wp_participant_id:
                return match
    return matches[0]


def _approval_note(
    existing_note: Optional[str],
    *,
    accepted_at: str,
    reason: str,
    target_status: str,
) -> str:
    appended = (
        f"{accepted_at} - Approval restored to '{target_status}' after reviewed approval drift. "
        f"{reason}"
    )
    if existing_note and str(existing_note).strip():
        return f"{str(existing_note).strip()}\n\n{appended}"
    return appended


def _acceptance_base_row(
    participant: CurrentParticipant,
    accepted_at: str,
    reason: str,
    execute: bool,
    *,
    force_approved: bool,
    target_status: str,
    prior_status_candidates: str,
    prior_status_resolution: str,
) -> dict[str, str]:
    return {
        "Mode": "execute" if execute else "dry-run",
        "Force Approved": "yes" if force_approved else "no",
        "Accepted At": accepted_at,
        "Church": participant.church,
        "Name": participant.name,
        "Is Member": participant.is_member,
        "ChMeetings ID": participant.chmeetings_id,
        "WP Participant ID": participant.wp_participant_id,
        "Workbook Approval Status": participant.approval_status,
        "WP Current Approval Status": "",
        "Target Approval Status": target_status,
        "Prior Status Candidates": prior_status_candidates,
        "Prior Status Resolution": prior_status_resolution,
        "Sports Registered Now": participant.sports_registered,
        "Reason": reason,
        "Approvals Found": "0",
        "Open Drift Issues Found": "0",
        "Participant Updated": "no",
        "Approvals Updated": "0",
        "Drift Issues Resolved": "0",
        "Action": "",
        "Message": "",
    }


def write_acceptance_workbook(rows: list[dict[str, str]], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approval-Drift-Acceptance"

    headers = list(rows[0].keys()) if rows else [
        "Mode",
        "Force Approved",
        "Accepted At",
        "Church",
        "Name",
        "Is Member",
        "ChMeetings ID",
        "WP Participant ID",
        "Workbook Approval Status",
        "WP Current Approval Status",
        "Target Approval Status",
        "Prior Status Candidates",
        "Prior Status Resolution",
        "Sports Registered Now",
        "Reason",
        "Approvals Found",
        "Open Drift Issues Found",
        "Participant Updated",
        "Approvals Updated",
        "Drift Issues Resolved",
        "Action",
        "Message",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="7030A0")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 10,
        "B": 14,
        "C": 20,
        "D": 10,
        "E": 24,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 24,
        "J": 24,
        "K": 22,
        "L": 28,
        "M": 30,
        "N": 42,
        "O": 52,
        "P": 14,
        "Q": 20,
        "R": 16,
        "S": 16,
        "T": 20,
        "U": 32,
        "V": 60,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    action_counts = Counter(row.get("Action", "") for row in rows)
    summary_rows = [
        ("Target rows", len(rows)),
    ]
    summary_rows.extend((f"Rows: {action}", count) for action, count in sorted(action_counts.items()))
    for item in summary_rows:
        summary.append(list(item))
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 16

    wb.save(output_path)
    return output_path


def run(
    *,
    workbook_path: Path,
    logs_dir: Path,
    output_path: Path,
    status: str = "reapproval_required",
    church_code: Optional[str] = None,
    since: Optional[str] = None,
) -> dict[str, object]:
    current = load_current_participants(
        workbook_path,
        status=status,
        church_code=church_code,
    )
    log_paths = sorted(logs_dir.glob(LOG_GLOB))
    if since:
        log_paths = [
            path
            for path in log_paths
            if _log_date_from_name(path.name) is None or _log_date_from_name(path.name) >= since
        ]
    events = parse_drift_events(log_paths)
    rows = build_history_rows(current, events)
    written = write_history_workbook(rows, output_path)
    with_history = {
        row["ChMeetings ID"]
        for row in rows
        if row["Event Type"] != "no_local_drift_log_found"
    }
    return {
        "workbook": str(workbook_path),
        "logs_dir": str(logs_dir),
        "output": str(written),
        "participants": len(current),
        "participants_with_history": len(with_history),
        "rows": len(rows),
    }


def _log_date_from_name(name: str) -> Optional[str]:
    match = re.search(r"sportsfest_(\d{8})", name)
    if not match:
        return None
    raw = match.group(1)
    return f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
